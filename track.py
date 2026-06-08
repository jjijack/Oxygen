import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.path import Path as MplPath
import geopandas as gpd
import inspect
from netCDF4 import Dataset
import re
import os
from pathlib import Path
import pickle
from matplotlib.cm import ScalarMappable
from matplotlib.colors import Normalize
import glob
from scipy.interpolate import RegularGridInterpolator
from scipy.ndimage import gaussian_filter
from scipy.stats import pearsonr, spearmanr
import copy
import gsw
from collections import defaultdict
import multiprocessing
import h5py
import time as tm
import shutil
from dataclasses import dataclass, field
from matplotlib.lines import Line2D
from matplotlib.patches import Ellipse
import dask.dataframe as dd
from dask.distributed import Client, LocalCluster, as_completed
from dask import delayed, compute
from dask.diagnostics import ProgressBar
from tqdm.auto import tqdm
import yaml
import traceback
import json, pyarrow as pa, pyarrow.parquet as pq
import math
import zarr
import cartopy.crs as ccrs
import cartopy.feature as cfeature
from adjustText import adjust_text
# -------------------- 区域配置加载 --------------------
def _load_region_config(config_path: str | Path = 'config/regions.yml', region: str | None = None):
    """加载区域配置文件并返回指定区域字典。

    若未安装 PyYAML 或文件缺失，则回退到黑潮延伸体默认范围。
    """
    fallback = {
        'lon_min': 140 - 2.5,
        'lon_max': 180 + 2.5,
        'lat_min': 28 - 2.5,
        'lat_max': 40 + 2.5,
        'crosses_dateline': True,
        '_fallback': True
    }
    cfg_path = Path(config_path)
    if yaml is None or not cfg_path.exists():
        return fallback
    try:
        with open(cfg_path, 'r') as f:
            cfg = yaml.safe_load(f)
        if region is None:
            region = cfg.get('default_region')
        region_dict = cfg['regions'][region]
        # 统一键名，缺失时使用 fallback
        for k, v in fallback.items():
            region_dict.setdefault(k, v)
        region_dict['_fallback'] = False
        region_dict['_region_key'] = str(region)
        return region_dict
    except Exception:
        return fallback

_REGION_CFG = _load_region_config()
lonmin, lonmax = _REGION_CFG['lon_min'], _REGION_CFG['lon_max']
latmin, latmax = _REGION_CFG['lat_min'], _REGION_CFG['lat_max']

# -------------------- 数据路径与处理参数配置加载（Paths & Processing Config） --------------------
def _load_yaml(path: str | Path) -> dict:
    if yaml is None:
        return {}
    p = Path(path)
    if not p.exists():
        return {}
    try:
        with open(p, 'r') as f:
            return yaml.safe_load(f) or {}
    except Exception:
        return {}

_PATHS_CFG = _load_yaml('config/paths.yml')
_PROC_CFG = _load_yaml('config/processing.yml')

argo_origin_path = Path(_PATHS_CFG.get('paths', {}).get('argo_origin', './Argo_origin'))
tmp_parquet_path = Path(_PATHS_CFG.get('paths', {}).get('argo_intermediate', './Argo_data_tmp'))
argo_path = Path(_PATHS_CFG.get('paths', {}).get('argo_parquet', './Argo_data'))
argo_mat_input_path = Path(_PATHS_CFG.get('paths', {}).get('argo_mat_input', './Argo_addFloat'))
Glorys_path = _PATHS_CFG.get('paths', {}).get('glorys_root', '../copernicus/GLORYS')
glorys_processed_root = Path(_PATHS_CFG.get('paths', {}).get('glorys_processed', './GLORYS_processed'))
META_root_path = Path(_PATHS_CFG.get('paths', {}).get('meta_root', '../META3.2_DT_allsat'))
plots_output_root = Path(_PATHS_CFG.get('paths', {}).get('plots_output', './plot_outputs'))

# 用下划线隐藏内部配置值，提供 getter 避免随处写死名称
circle_enlargement_factor = float(
    _PROC_CFG.get('processing', {}).get('circle_enlargement_factor', 1.2)
)
# 已弃用的平均米/度常量删除，统一使用 approximate_degree_length 计算局地尺度
_default_delta_do_threshold = float(
    _PROC_CFG.get('processing', {}).get('default_delta_do_threshold', 50.0)
)
_default_salinity_threshold = float(
    _PROC_CFG.get('processing', {}).get('default_salinity_threshold', 0.0)
)
_default_temperature_threshold = float(
    _PROC_CFG.get('processing', {}).get('default_temperature_threshold', 0.0)
)
_default_depth_interval = float(
    _PROC_CFG.get('processing', {}).get('depth_interval', 100.0)
)
_default_depth_merge_tolerance = float(
    _PROC_CFG.get('processing', {}).get('depth_merge_tolerance', 10.0)
)
_default_duplicate_depth_strategy = _PROC_CFG.get('processing', {}).get('duplicate_depth_strategy', 'best_qc')
_cfg_anomaly_min_depth = float(
    _PROC_CFG.get('processing', {}).get('anomaly_min_depth', 300.0)
)
_cfg_anomaly_max_depth = float(
    _PROC_CFG.get('processing', {}).get('anomaly_max_depth', 1500.0)
)
_cfg_do_near_zero_threshold = float(
    _PROC_CFG.get('processing', {}).get('do_near_zero_threshold', 1.0)
)
_cfg_do_near_zero_max_count = int(
    _PROC_CFG.get('processing', {}).get('do_near_zero_max_count', 7)
)
_default_subduction_detection_method = str(
    _PROC_CFG.get('processing', {}).get('subduction_detection_method', 'do')
).strip().lower()
if _default_subduction_detection_method not in {'do', 'aou', 'trim'}:
    _default_subduction_detection_method = 'do'
_default_aou_threshold = float(
    _PROC_CFG.get('processing', {}).get('aou_threshold', -10.0)
)
_default_pi_threshold = float(
    _PROC_CFG.get('processing', {}).get('pi_threshold', 0.05)
)
_default_aou_pi_depth_tolerance = float(
    _PROC_CFG.get('processing', {}).get('aou_pi_depth_tolerance', 30.0)
)
_default_trim_cutoff = float(
    _PROC_CFG.get('processing', {}).get('trim_cutoff', 1.96)
)
_default_trim_window = float(
    _PROC_CFG.get('processing', {}).get('trim_window', 60.0)
)
_default_trim_bin_width_outlier = float(
    _PROC_CFG.get('processing', {}).get('trim_bin_width_outlier', 40.0)
)
_default_trim_bin_width_check = float(
    _PROC_CFG.get('processing', {}).get('trim_bin_width_check', 20.0)
)
_default_trim_depth_min = float(
    _PROC_CFG.get('processing', {}).get('trim_depth_min', 200.0)
)
_default_trim_depth_max = float(
    _PROC_CFG.get('processing', {}).get('trim_depth_max', 1000.0)
)
_default_adaptive_lat_threshold = float(
    _PROC_CFG.get('processing', {}).get('adaptive_lat_threshold', 70.0)
)
_default_adaptive_distance_threshold_km = float(
    _PROC_CFG.get('processing', {}).get('adaptive_distance_threshold_km', 300.0)
)
_default_spice_percentile_threshold = float(
    _PROC_CFG.get('processing', {}).get('spice', {}).get('percentile_threshold', 10.0)
)
_default_vertical_profile_spacing_km = float(
    _PROC_CFG.get('processing', {}).get('vertical_profile_spacing_km', 2.0)
)
_default_vertical_profile_depth_spacing_m = float(
    _PROC_CFG.get('processing', {}).get('vertical_profile_depth_spacing_m', 5.0)
)

# Heave + Ventilation 等密面出露诊断参数（从 processing.yml:processing.heave 读取）
_HEAVE_CFG = _PROC_CFG.get('processing', {}).get('heave', {})
_heave_search_range = float(
    _HEAVE_CFG.get('search_range', 0.5)
)
_heave_depth_threshold = float(
    _HEAVE_CFG.get('depth_threshold', 150.0)
)
_heave_magnitude_threshold = float(
    _HEAVE_CFG.get('magnitude_threshold', 100.0)
)
_heave_x_window_km = float(
    _HEAVE_CFG.get('x_window_km', 200.0)
)
_heave_local_x_window_km = float(
    _HEAVE_CFG.get('local_x_window_km', 50.0)
)
_heave_z_search_m = float(
    _HEAVE_CFG.get('z_search_m', 200.0)
)

# Argo 3D 高斯核断面重建参数（从 processing.yml:processing.argo_reconstruction 读取）
_ARGO_RECON_CFG = _PROC_CFG.get('processing', {}).get('argo_reconstruction', {})
_argo_recon_h_bw_km = float(
    _ARGO_RECON_CFG.get('h_bw_km', 60.0)
)
_argo_recon_min_weight = float(
    _ARGO_RECON_CFG.get('min_weight', 3.0)
)
_argo_recon_depth_bw_m = float(
    _ARGO_RECON_CFG.get('depth_bw_m', 25.0)
)
_argo_recon_h_spacing_deg = float(
    _ARGO_RECON_CFG.get('h_spacing_deg', 0.1)
)
_argo_recon_z_max_m = float(
    _ARGO_RECON_CFG.get('z_max_m', 1500.0)
)
_argo_recon_z_spacing_m = float(
    _ARGO_RECON_CFG.get('z_spacing_m', 10.0)
)
_argo_recon_x_spacing_km = float(
    _ARGO_RECON_CFG.get('x_spacing_km', 5.0)
)
_argo_recon_radius_km = float(
    _ARGO_RECON_CFG.get('radius_km', 400.0)
)
_argo_recon_day_window = int(
    _ARGO_RECON_CFG.get('day_window_days', 15)
)
_argo_recon_min_profiles = int(
    _ARGO_RECON_CFG.get('min_profiles', 20)
)
_argo_recon_min_coverage_top = float(
    _ARGO_RECON_CFG.get('min_coverage_top1000', 0.5)
)
_argo_recon_coverage_probe_spacing_m = float(
    _ARGO_RECON_CFG.get('coverage_probe_spacing_m', 100.0)
)

# GLORYS 逐点残差 / 垂向细结构丢失统计参数（从 processing.yml:processing.glorys_residual 读取）
_GLORYS_RESID_CFG = _PROC_CFG.get('processing', {}).get('glorys_residual', {})
_glorys_resid_z_max_m = float(
    _GLORYS_RESID_CFG.get('z_max_m', 1000.0)
)
_glorys_resid_z_spacing_m = float(
    _GLORYS_RESID_CFG.get('z_spacing_m', 5.0)
)
_glorys_resid_fine_window = int(
    _GLORYS_RESID_CFG.get('fine_struct_window', 15)
)
_glorys_resid_sameloc_radius_km = float(
    _GLORYS_RESID_CFG.get('sameloc_radius_km', 40.0)
)
_glorys_resid_match_radius_km = float(
    _GLORYS_RESID_CFG.get('match_radius_km', 140.0)
)
_glorys_resid_match_window_deg = float(
    _GLORYS_RESID_CFG.get('match_window_deg', 1.0)
)
_glorys_resid_match_window_days = int(
    _GLORYS_RESID_CFG.get('match_window_days', 1)
)
_glorys_resid_match_min_cov = float(
    _GLORYS_RESID_CFG.get('match_min_depth_coverage', 0.8)
)
_glorys_resid_so_lat = float(
    _GLORYS_RESID_CFG.get('so_lat_threshold', -40.0)
)
# 残差 worker 进程内的 Argo 年缓存及其上限（防跨十几年累积把内存吃爆；任务按年份排序后命中率仍高）
_RESID_ARGO_CACHE: dict = {}
_RESID_ARGO_CACHE_MAX = 3

_DETECTION_METHODS = {'do', 'aou', 'trim'}
_cfg_cbar_defaults = _PROC_CFG.get('processing', {}).get('cbar_defaults', {})
_CBAR_FALLBACK = {
    'do': (50.0, 150.0),
    'aou': (-60.0, -10.0),
    'trim': (2.0, 6.0),
}

def _format_detection_value(value) -> str:
    """文件名安全的短数字串。"""
    try:
        if value is None or pd.isna(value):
            return 'NA'
    except Exception:
        if value is None:
            return 'NA'
    try:
        return f"{float(value):g}".replace('.', 'p').replace('-', 'n')
    except Exception:
        return str(value).replace('.', 'p').replace('-', 'n').replace(' ', '')

def _normalize_detection_method(method: str | None) -> str:
    method_norm = str(method or _default_subduction_detection_method).strip().lower()
    if method_norm not in _DETECTION_METHODS:
        raise ValueError(
            f"method must be one of {sorted(_DETECTION_METHODS)}, got {method!r}."
        )
    return method_norm

def _method_cbar_defaults(method: str) -> tuple[float, float]:
    """按识别方法返回默认色标上下限。"""
    method_norm = _normalize_detection_method(method)
    defaults = _cfg_cbar_defaults.get(method_norm, {}) if isinstance(_cfg_cbar_defaults, dict) else {}
    fallback_lo, fallback_hi = _CBAR_FALLBACK.get(method_norm, (0.0, 100.0))
    if isinstance(defaults, dict):
        lo = defaults.get('min', fallback_lo)
        hi = defaults.get('max', fallback_hi)
    else:
        lo, hi = fallback_lo, fallback_hi
    return float(lo), float(hi)

@dataclass
class DetectionConfig:
    """单一异常识别配置对象，贯穿 calculate_delta_do 及其下游流程。"""

    method: str = field(default_factory=lambda: _default_subduction_detection_method)

    # DO 模式
    do_threshold: float = field(default_factory=lambda: _default_delta_do_threshold)

    # AOU 模式
    aou_threshold: float = field(default_factory=lambda: _default_aou_threshold)
    pi_threshold: float = field(default_factory=lambda: _default_pi_threshold)
    aou_pi_depth_tolerance: float = field(default_factory=lambda: _default_aou_pi_depth_tolerance)

    # TRIM 模式
    trim_cutoff: float = field(default_factory=lambda: _default_trim_cutoff)
    trim_window: float = field(default_factory=lambda: _default_trim_window)
    trim_bin_width_outlier: float = field(default_factory=lambda: _default_trim_bin_width_outlier)
    trim_bin_width_check: float = field(default_factory=lambda: _default_trim_bin_width_check)
    trim_depth_min: float = field(default_factory=lambda: _default_trim_depth_min)
    trim_depth_max: float = field(default_factory=lambda: _default_trim_depth_max)

    # Spiciness T-S 偏离阈值（百分位）
    spice_percentile_threshold: float = field(default_factory=lambda: _default_spice_percentile_threshold)

    # 通用预处理与过滤
    depth_interval: float = field(default_factory=lambda: _default_depth_interval)
    salinity_threshold: float = field(default_factory=lambda: _default_salinity_threshold)
    temperature_threshold: float = field(default_factory=lambda: _default_temperature_threshold)
    anomaly_min_depth: float = field(default_factory=lambda: _cfg_anomaly_min_depth)
    anomaly_max_depth: float = field(default_factory=lambda: _cfg_anomaly_max_depth)
    depth_merge_tolerance: float = field(default_factory=lambda: _default_depth_merge_tolerance)
    duplicate_depth_strategy: str = field(default_factory=lambda: _default_duplicate_depth_strategy)
    do_near_zero_threshold: float = field(default_factory=lambda: _cfg_do_near_zero_threshold)
    do_near_zero_max_count: int = field(default_factory=lambda: _cfg_do_near_zero_max_count)

    # 绘图显示
    cbar_min: float | None = None
    cbar_max: float | None = None
    cbar_ticks: list | None = None

    def __post_init__(self):
        self.method = _normalize_detection_method(self.method)

    def resolved_cbar(self) -> tuple[float, float]:
        lo_default, hi_default = _method_cbar_defaults(self.method)
        lo = self.cbar_min if self.cbar_min is not None else lo_default
        hi = self.cbar_max if self.cbar_max is not None else hi_default
        return float(lo), float(hi)

    def score_col(self) -> str:
        return 'anomaly_score'

    def color_col(self) -> str:
        if self.method == 'aou':
            return 'delta_aou'
        if self.method == 'trim':
            return 'trim_score'
        return 'delta_do'

    def timeseries_variable(self) -> str:
        if self.method in {'aou', 'trim'}:
            return 'AOU'
        return 'DO'

    def color_label(self) -> str:
        if self.method == 'aou':
            return 'ΔAOU / μmol·kg⁻¹'
        if self.method == 'trim':
            return 'Trim residual score / σ'
        return 'ΔDO / μmol·kg⁻¹'

    def cmap(self) -> str:
        if self.method == 'aou':
            return 'Blues_r'
        if self.method == 'trim':
            return 'magma'
        return 'Reds'

    def threshold_label(self) -> str:
        if self.method == 'aou':
            return (
                f"ΔAOU ≤ {self.aou_threshold:g} μmol kg⁻¹, "
                f"|ΔPI| ≥ {self.pi_threshold:g}"
            )
        if self.method == 'trim':
            return f"trim residual > {self.trim_cutoff:g}σ"
        return f"ΔDO ≥ {self.do_threshold:g} μmol kg⁻¹"

    def file_stem(self) -> str:
        if self.method == 'aou':
            core = (
                f"aou{_format_detection_value(self.aou_threshold)}"
                f"_pi{_format_detection_value(self.pi_threshold)}"
            )
        elif self.method == 'trim':
            core = f"trim{_format_detection_value(self.trim_cutoff)}"
        else:
            core = f"do{_format_detection_value(self.do_threshold)}"

        depth = ''
        if self.anomaly_min_depth is not None and float(self.anomaly_min_depth) > 0:
            depth = f"_depth{_format_detection_value(self.anomaly_min_depth)}m"
        return f"{core}{depth}"

    def output_dir(self, fn_name: str, region_slug: str | None = None, plots_root: str | Path | None = None) -> Path:
        slug = region_slug or _current_region_key()
        base = Path(plots_root) if plots_root is not None else Path(plots_output_root)
        return base / self.method / slug / fn_name

def _shared_output_dir(
    fn_name: str,
    region_slug: str | None = None,
    plots_root: str | Path | None = None,
) -> Path:
    """返回与异常判定方法无关的共享输出目录。"""
    slug = region_slug or _current_region_key()
    base = Path(plots_root) if plots_root is not None else Path(plots_output_root)
    return base / 'shared' / str(slug) / fn_name

def _detection_output_dir_from_meta(
    fn_name: str,
    meta: dict | None = None,
    *,
    region_slug: str | None = None,
    plots_root: str | Path | None = None,
) -> Path:
    """根据 summary 元信息返回带异常判定方法的输出目录。"""
    meta = meta or {}
    method = _normalize_detection_method(meta.get('detection_method') or 'do')
    slug = region_slug or meta.get('region_key') or _current_region_key()
    base = Path(plots_root) if plots_root is not None else Path(plots_output_root)
    return base / method / str(slug) / fn_name

def make_detection_config(
    method: str | DetectionConfig | None = None,
    **overrides,
) -> DetectionConfig:
    """从全局默认配置构建 DetectionConfig，并应用临时覆盖。

    传入 DetectionConfig 实例时按其字段克隆再覆盖；传入方法名字符串（或 None）时先规范化方法名，
    再用 processing.yml 默认值构建。覆盖项通过 ``**overrides`` 透传给 DetectionConfig。

    参数:
        - method (str | DetectionConfig | None): 检测方法名（'do'/'aou'/'trim'），或一个已有的 DetectionConfig（克隆后覆盖），或 None（取默认方法）。
        - **overrides: 透传给 DetectionConfig 的字段覆盖（如 do_threshold、anomaly_min_depth 等）。

    返回:
        - DetectionConfig: 构建好的检测配置实例。
    """
    if isinstance(method, DetectionConfig):
        values = method.__dict__.copy()
        values.update(overrides)
        return DetectionConfig(**values)
    method_norm = _normalize_detection_method(method)
    return DetectionConfig(method=method_norm, **overrides)

def _resolve_detection_config(
    detection_config: DetectionConfig | None = None,
    **overrides,
) -> DetectionConfig:
    """公共入口的 DetectionConfig 解析器；None 时使用 processing.yml 默认值。"""
    if detection_config is not None and not isinstance(detection_config, DetectionConfig):
        raise TypeError("detection_config 必须是 DetectionConfig；请先用 make_detection_config(...) 构建。")
    return make_detection_config(
        detection_config,
        **{k: v for k, v in overrides.items() if v is not None},
    )

def _keep_best_anomaly_per_profile(anomalies: pd.DataFrame, detection_config: DetectionConfig | None = None) -> pd.DataFrame:
    """每个 Profile_number 只保留统一 anomaly_score 最大的一行。"""
    if anomalies is None or anomalies.empty or 'Profile_number' not in anomalies.columns:
        return pd.DataFrame() if anomalies is None else anomalies
    score_col = (detection_config.score_col() if detection_config is not None else 'anomaly_score')
    if score_col not in anomalies.columns:
        fallback_col = 'delta_do' if 'delta_do' in anomalies.columns else None
        if fallback_col is None:
            return anomalies.drop_duplicates(subset='Profile_number', keep='first')
        score_col = fallback_col
    return (
        anomalies.sort_values(score_col, ascending=False)
        .drop_duplicates(subset='Profile_number', keep='first')
    )

def _color_values_for_anomalies(anomalies: pd.DataFrame, detection_config: DetectionConfig) -> tuple[pd.Series | None, str, str, str]:
    """返回 anomalies 的着色列、列名、色标标签和 cmap。"""
    col = detection_config.color_col()
    if col not in anomalies.columns and detection_config.score_col() in anomalies.columns:
        col = detection_config.score_col()
    if col not in anomalies.columns:
        return None, col, detection_config.color_label(), detection_config.cmap()
    return pd.to_numeric(anomalies[col], errors='coerce'), col, detection_config.color_label(), detection_config.cmap()

def _apply_detection_colorbar_ticks(cbar, detection_config: DetectionConfig, cbar_min: float, cbar_max: float):
    if detection_config.cbar_ticks is not None:
        cbar.set_ticks(detection_config.cbar_ticks)
        return
    rng = float(cbar_max) - float(cbar_min)
    if rng > 30:
        mid = (float(cbar_min) + float(cbar_max)) / 2.0
        cbar.set_ticks([cbar_min, mid, cbar_max])
    else:
        cbar.set_ticks([cbar_min, cbar_max])

def _num_from_record(record, key: str) -> float:
    try:
        val = pd.to_numeric(pd.Series([record.get(key)]), errors='coerce').iloc[0]
        return float(val) if np.isfinite(val) else np.nan
    except Exception:
        return np.nan

def _annotation_text_from_anomaly_record(record, detection_config: DetectionConfig) -> tuple[str, str]:
    """从异常记录生成简短的标题指标文本与深度文本。"""
    parts = []
    if detection_config.method == 'aou':
        delta_aou = _num_from_record(record, 'delta_aou')
        delta_pi = _num_from_record(record, 'delta_pi')
        if np.isfinite(delta_aou):
            parts.append(f"ΔAOU={delta_aou:.2f}")
        if np.isfinite(delta_pi):
            parts.append(f"ΔPI={delta_pi:.3f}")
    elif detection_config.method == 'trim':
        trim_score = _num_from_record(record, 'trim_score')
        score_aou = _num_from_record(record, 'trim_scale_res_rob_aou')
        score_sal = _num_from_record(record, 'trim_scale_res_rob_abs_sal')
        if np.isfinite(trim_score):
            parts.append(f"trim score={trim_score:.2f}σ")
        if np.isfinite(score_aou):
            parts.append(f"AOU res={score_aou:.2f}σ")
        if np.isfinite(score_sal):
            parts.append(f"SA res={score_sal:.2f}σ")
    else:
        delta_do = _num_from_record(record, 'delta_do')
        if np.isfinite(delta_do):
            parts.append(f"ΔDO={delta_do:.2f}")

    delta_temp = _num_from_record(record, 'delta_temperature')
    delta_sal = _num_from_record(record, 'delta_salinity')
    if np.isfinite(delta_temp):
        parts.append(f"ΔTemperature={delta_temp:.2f}")
    if np.isfinite(delta_sal):
        parts.append(f"ΔSalinity={delta_sal:.2f}")

    depth_val = _num_from_record(record, 'depth')
    depth_text = f" @{depth_val:.1f}m" if np.isfinite(depth_val) else ""
    metric_text = ", " + ", ".join(parts) if parts else ""
    return metric_text, depth_text

def _load_basemap_colors() -> dict:
    """加载底图颜色配置。

    未配置的键使用默认值；缺失 plot.basemap_colors 则全部取默认。
    """
    defaults = {
        'ocean': '#f8fafc',
        'land': '#d9d9d9',
        'coastline': '#787d85',
        'border': '#9aa0a8',
        'grid': '#b6bdc6',
    }
    if not isinstance(_PROC_CFG, dict):
        return defaults
    bm = _PROC_CFG.get('plot', {}).get('basemap_colors', {})
    merged = defaults.copy()
    if isinstance(bm, dict):
        merged.update({k: str(v) for k, v in bm.items() if k in defaults})
    return merged

_BASEMAP_COLORS = _load_basemap_colors()

def _load_hotspot_plot_colors() -> tuple[dict, dict, dict]:
    """加载 hotspot_type / spice_type / cross 三组拆图配色。"""
    plot_cfg = _PROC_CFG.get('plot', {}) if isinstance(_PROC_CFG, dict) else {}
    type_defaults = {'type_1': 'black', 'type_2': 'dimgray', 'type_3': 'red'}
    spice_defaults = {'cold_fresh': '#1f77b4', 'background': '#7f7f7f', 'warm_salty': '#2ca02c'}
    cross_defaults = {
        'T1_cold_fresh': '#1f77b4', 'T1_background': '#aec7e8',
        'T2_cold_fresh': '#d62728', 'T2_background': '#ff9896', 'T3_OMZ': '#ff7f0e',
    }
    def _merge(defaults, key):
        raw = plot_cfg.get(key, {})
        merged = defaults.copy()
        if isinstance(raw, dict):
            merged.update({k: str(v) for k, v in raw.items() if k in defaults})
        return merged
    return _merge(type_defaults, 'hotspot_type_colors'), _merge(spice_defaults, 'hotspot_spice_colors'), _merge(cross_defaults, 'hotspot_cross_colors')

_HOTSPOT_TYPE_COLORS, _HOTSPOT_SPICE_COLORS, _HOTSPOT_CROSS_COLORS = _load_hotspot_plot_colors()

# -------------------- 自带底图加载（使用本地 Natural Earth, 简洁版） --------------------
def _load_world_geodataframe():
    """从配置读取底图路径，默认 external/natural_earth/ne_110m_admin_0_countries.shp。"""
    cfg_paths = _PATHS_CFG.get('paths', {}) if isinstance(_PATHS_CFG, dict) else {}
    shp_path = Path(cfg_paths.get('gpd_world_shp', './external/natural_earth/ne_110m_admin_0_countries.shp'))
    return gpd.read_file(str(shp_path))

# -------------------------------------------------------------------------------

def print_current_processing_defaults():
        """打印当前生效的处理参数全局默认值（从 processing.yml 读取/回退），用于调试与运行时确认配置，无返回值。

        涵盖 circle_enlargement_factor、distance_deg_per_meter、subduction_detection_method 与 do/aou/trim
        识别参数、salinity_threshold/temperature_threshold、depth_interval/depth_merge_tolerance/
        duplicate_depth_strategy 等。
        """
        print("[Processing Defaults]")
        print(f"  circle_enlargement_factor : {circle_enlargement_factor}")
        print(f"  delta_do_threshold        : {_default_delta_do_threshold}")
        print(f"  salinity_threshold        : {_default_salinity_threshold}")
        print(f"  temperature_threshold     : {_default_temperature_threshold}")
        print(f"  depth_interval            : {_default_depth_interval}")
        print(f"  depth_merge_tolerance     : {_default_depth_merge_tolerance}")
        print(f"  duplicate_depth_strategy  : {_default_duplicate_depth_strategy}")
        print(f"  anomaly_min_depth         : {_cfg_anomaly_min_depth}")
        print(f"  anomaly_max_depth         : {_cfg_anomaly_max_depth}")
        print(f"  do_near_zero_threshold    : {_cfg_do_near_zero_threshold}")
        print(f"  do_near_zero_max_count    : {_cfg_do_near_zero_max_count}")
        print(f"  subduction_detection_method: {_default_subduction_detection_method}")
        print(f"  aou_threshold             : {_default_aou_threshold}")
        print(f"  pi_threshold              : {_default_pi_threshold}")
        print(f"  aou_pi_depth_tolerance    : {_default_aou_pi_depth_tolerance}")
        print(f"  trim_cutoff               : {_default_trim_cutoff}")
        print(f"  trim_window               : {_default_trim_window}")
        print(f"  trim_bin_width_outlier    : {_default_trim_bin_width_outlier}")
        print(f"  trim_bin_width_check      : {_default_trim_bin_width_check}")
        print(f"  trim_depth_min            : {_default_trim_depth_min}")
        print(f"  trim_depth_max            : {_default_trim_depth_max}")
        print(f"  vertical_profile_spacing_km: {_default_vertical_profile_spacing_km}")
        print(f"  vertical_profile_depth_spacing_m: {_default_vertical_profile_depth_spacing_m}")
        print(f"  --- Heave / OI diagnostics ---")
        print(f"  heave.search_range          : {_heave_search_range}")
        print(f"  heave.depth_threshold       : {_heave_depth_threshold}")
        print(f"  heave.magnitude_threshold   : {_heave_magnitude_threshold}")
        print(f"  heave.x_window_km           : {_heave_x_window_km}")
        print(f"  heave.local_x_window_km     : {_heave_local_x_window_km}")
        print(f"  heave.z_search_m            : {_heave_z_search_m}")

def approximate_degree_length(lat: float | np.ndarray, lon: float | np.ndarray | None = None) -> dict:
    """计算指定纬度（可选经度）处经纬度与距离的近似换算关系。

    采用 WGS84 椭球常用的近似级数公式（单位：米/度），适用于绝大多数海洋学分析的精度需求；相比旧的
    单一平均值，提供随纬度变化的更精细米/度估计。

    参数:
        - lat (float | np.ndarray): 纬度（度），可为标量或 numpy 数组。
        - lon (float | np.ndarray | None): 经度（度），对当前计算无影响，仅为接口对称保留（可传与 lat 同形状数组，将被忽略），便于未来扩展（如地形加权）。

    返回:
        - dict: 含四个键 —— meters_per_degree_lat / meters_per_degree_lon（该纬度上一度纬差/经差对应的米数）、degrees_per_meter_lat / degrees_per_meter_lon（前两者的倒数，度/米）；输入为数组时各字段为同形状 numpy 数组。

    说明:
        - 近似级数（展开到 cos(5φ)/cos(6φ)）：meters_per_degree_lat ≈ 111132.92 − 559.82·cos2φ + 1.175·cos4φ − 0.0023·cos6φ；meters_per_degree_lon ≈ 111412.84·cosφ − 93.5·cos3φ + 0.118·cos5φ。
        - 将“米单位半径”换算到“角度半径”时推荐：radius_deg_lat = radius_m / meters_per_degree_lat。
    """
    # 转换为 numpy 数组以统一处理
    lat_arr = np.asarray(lat, dtype=float)
    phi = np.deg2rad(lat_arr)

    # 分别计算纬度与经度方向一度的米数（WGS84 近似）
    meters_per_degree_lat = (
        111132.92
        - 559.82 * np.cos(2 * phi)
        + 1.175 * np.cos(4 * phi)
        - 0.0023 * np.cos(6 * phi)
    )
    meters_per_degree_lon = (
        111412.84 * np.cos(phi)
        - 93.5 * np.cos(3 * phi)
        + 0.118 * np.cos(5 * phi)
    )

    # 倒数（度/米）。对 0 做防护：若 cos(phi) ~ 0（极点），经向一度长度 → 0，避免除零设为 np.inf
    with np.errstate(divide='ignore', invalid='ignore'):
        degrees_per_meter_lat = 1.0 / meters_per_degree_lat
        degrees_per_meter_lon = np.where(meters_per_degree_lon != 0, 1.0 / meters_per_degree_lon, np.inf)

    result = {
        'meters_per_degree_lat': meters_per_degree_lat,
        'meters_per_degree_lon': meters_per_degree_lon,
        'degrees_per_meter_lat': degrees_per_meter_lat,
        'degrees_per_meter_lon': degrees_per_meter_lon,
    }

    # 若输入是标量（非数组或0维），将结果中对应项转换为原生 float，避免下游出现 0-d ndarray 序列化/打印差异
    if np.asarray(lat).shape == ():
        for k, v in result.items():
            # v 可能是 ndarray 标量或 Python float
            if isinstance(v, np.ndarray) and v.shape == ():
                result[k] = v.item()
    return result

def _minimal_lon_diff_deg(lon: float | np.ndarray, lon0: float) -> np.ndarray:
    """计算经度差并映射到 (-180, 180] 区间，支持标量或数组。

    用于跨日界线区域，保证 179.8° 与 -179.7° 的差为 0.5° 而不是 -359.5°。
    """
    d = np.asarray(lon) - lon0
    return (d + 180.0) % 360.0 - 180.0

def _normalize_lon_array(val: np.ndarray | float) -> np.ndarray:
    """将任意经度值映射到 (-180, 180] 区间，返回 np.ndarray。"""
    arr = np.asarray(val, dtype=float)
    return (arr + 180.0) % 360.0 - 180.0

def _region_lon_mask(lon_vals: np.ndarray, lon_min_cfg: float, lon_max_cfg: float) -> np.ndarray:
    """根据区域经度范围（允许跨日界线）生成布尔掩码。"""
    lon_arr = np.asarray(lon_vals, dtype=float)
    if lon_arr.size == 0:
        return np.zeros(0, dtype=bool)

    lon_min_norm = float(_normalize_lon_array(lon_min_cfg))
    lon_max_norm = float(_normalize_lon_array(lon_max_cfg))

    raw_span = abs(float(lon_max_cfg) - float(lon_min_cfg))
    eff_span = (lon_max_norm - lon_min_norm) % 360.0
    is_global_lon = (
        (raw_span >= 359.5)
        or (eff_span >= 359.5)
        or np.isclose(eff_span, 0.0, atol=1e-6)
    )
    if is_global_lon:
        return np.ones(lon_arr.shape, dtype=bool)

    lon_norm = _normalize_lon_array(lon_arr)
    if lon_min_norm <= lon_max_norm:
        return (lon_norm >= lon_min_norm) & (lon_norm <= lon_max_norm)
    return (lon_norm >= lon_min_norm) | (lon_norm <= lon_max_norm)

def local_xy_distance_m(lon: float | np.ndarray, lat: float | np.ndarray,
                        lon0: float, lat0: float, wrap_dateline: bool = True) -> np.ndarray:
    """计算点 (lon, lat) 到参考点 (lon0, lat0) 的局地平面近似距离（米）。

    使用纬度依赖的经/纬一度长度（WGS84 近似）。在中低纬、距离 <~500 km 时平面近似已足够；距离很大
    或靠近极区时误差增大，可改用 great_circle_distance_m。

    参数:
        - lon (float | np.ndarray): 目标点经度，可为标量或数组（广播到与 lat 相同形状）。
        - lat (float | np.ndarray): 目标点纬度，可为标量或数组。
        - lon0 (float): 参考中心经度（标量）。
        - lat0 (float): 参考中心纬度（标量）。
        - wrap_dateline (bool): 是否对经度差做跨日界线（±180°）最短差处理，默认 True；为 True 时能正确处理如 179.9° 与 -179.9° 仅相差 0.2° 的情况。

    返回:
        - np.ndarray: 与广播后输入同形状的距离（米）。
    """
    scale = approximate_degree_length(lat0)
    m_per_deg_lat = scale['meters_per_degree_lat']
    m_per_deg_lon = scale['meters_per_degree_lon']
    dlon = _minimal_lon_diff_deg(lon, lon0) if wrap_dateline else (np.asarray(lon) - lon0)
    dlat = np.asarray(lat) - lat0
    dx_m = dlon * m_per_deg_lon
    dy_m = dlat * m_per_deg_lat
    return np.hypot(dx_m, dy_m)

def great_circle_distance_m(lon: float | np.ndarray, lat: float | np.ndarray,
                            lon0: float, lat0: float, wrap_dateline: bool = True,
                            radius_earth_m: float = 6371000.0) -> np.ndarray:
    """计算球面大圆距离（Haversine 公式），单位：米。

    设计目标是简单稳定、无自动切换逻辑，与 local_xy_distance_m 并存，供需要更精确、大尺度或高纬场景
    手动调用。采用 Haversine 公式（a = sin²(Δφ/2) + cosφ₁·cosφ₂·sin²(Δλ/2)，c = 2·asin√a，d = R·c），
    中短距离下与球面真值非常接近，对极区与大尺度优于局地平面近似。

    参数:
        - lon (float | np.ndarray): 目标点经度（标量或一维数组）。
        - lat (float | np.ndarray): 目标点纬度（标量或一维数组）。
        - lon0 (float): 中心点经度（标量）。
        - lat0 (float): 中心点纬度（标量）。
        - wrap_dateline (bool): 为 True 时对经度差做跨日界线（±180°）最短差归一，默认 True。
        - radius_earth_m (float): 地球半径（米），默认 6371000.0，可按需改为更精确的椭球平均半径。

    返回:
        - np.ndarray | float: 与输入 (lon, lat) 形状一致的距离（米）；标量输入返回标量。
    """
    lon_arr = np.asarray(lon, dtype=float)
    lat_arr = np.asarray(lat, dtype=float)
    dlon_deg = _minimal_lon_diff_deg(lon_arr, lon0) if wrap_dateline else (lon_arr - lon0)
    dlat_deg = lat_arr - lat0
    dlon = np.deg2rad(dlon_deg)
    dlat = np.deg2rad(dlat_deg)
    lat1 = np.deg2rad(lat0)
    lat2 = np.deg2rad(lat_arr)
    sin_dlat = np.sin(dlat / 2.0)
    sin_dlon = np.sin(dlon / 2.0)
    a = sin_dlat**2 + np.cos(lat1) * np.cos(lat2) * sin_dlon**2
    # 数值稳定保护：浮点误差可能导致 a 略超 1
    c = 2.0 * np.arcsin(np.minimum(1.0, np.sqrt(a)))
    dist = radius_earth_m * c
    if np.ndim(dist) == 0:
        return float(dist)
    return dist

def adaptive_distance_m(
    lon: float | np.ndarray,
    lat: float | np.ndarray,
    lon0: float | np.ndarray,
    lat0: float | np.ndarray,
    wrap_dateline: bool = True,
    gc_lat_threshold: float | None = None,
    gc_distance_threshold_km: float | None = None,
    force_great_circle: bool = False,
    radius_earth_m: float = 6371000.0
) -> np.ndarray | tuple[np.ndarray, np.ndarray]:
    """自适应距离（米）。在保持平面估算速度的前提下，自动在高纬或大尺度条件下改用大圆距离。

    参数:
        - lon (float | np.ndarray): 目标点经度（标量或数组）。
        - lat (float | np.ndarray): 目标点纬度（标量或数组）。
        - lon0 (float | np.ndarray): 中心点经度（标量或数组，可与目标点一一对应广播）。
        - lat0 (float | np.ndarray): 中心点纬度（标量或数组）。
        - wrap_dateline (bool): 为 True 时对经度差做跨日界线（±180°）最短差归一，默认 True。
        - gc_lat_threshold (float | None): 高纬触发大圆的绝对纬度阈值；None 时取配置 adaptive_lat_threshold。
        - gc_distance_threshold_km (float | None): 平面最大距离超过该值（km）触发大圆；None 时取配置 adaptive_distance_threshold_km。
        - force_great_circle (bool): 强制全程使用大圆，默认 False。
        - radius_earth_m (float): 地球半径（米），默认 6371000.0。

    返回:
        - np.ndarray | float: 与 (lon, lat) 形状一致的距离（米）；标量输入返回 float。

    说明:
        策略:

            - force_great_circle=True 时直接使用大圆（适合全球/大范围批处理）。
            - 若 |lat0| >= gc_lat_threshold，直接大圆，跳过平面计算。
            - 否则先算一次平面近似；若最大平面距离 > gc_distance_threshold_km 则改用大圆，否则保留平面结果。

        使用建议:

            - 全球或大量远距点：直接 force_great_circle=True，减少双重计算。
            - 中低纬涡旋附近（多数点在半径数倍内）：保持默认，兼顾平面速度与精度。
            - 极区或大半径且对精度敏感：force_great_circle=True。
    """
    if gc_lat_threshold is None:
        gc_lat_threshold = _default_adaptive_lat_threshold
    if gc_distance_threshold_km is None:
        gc_distance_threshold_km = _default_adaptive_distance_threshold_km

    lon_arr = np.asarray(lon, dtype=float)
    lat_arr = np.asarray(lat, dtype=float)
    lon0_arr = np.asarray(lon0, dtype=float)
    lat0_arr = np.asarray(lat0, dtype=float)

    # 广播到统一形状，便于处理逐点中心（例如逐日涡旋中心）
    try:
        broadcast_shape = np.broadcast(lon_arr, lat_arr, lon0_arr, lat0_arr).shape
    except ValueError as exc:
        raise ValueError("adaptive_distance_m: inputs cannot be broadcast to a common shape") from exc
    lon_arr = np.broadcast_to(lon_arr, broadcast_shape)
    lat_arr = np.broadcast_to(lat_arr, broadcast_shape)
    lon0_arr = np.broadcast_to(lon0_arr, broadcast_shape)
    lat0_arr = np.broadcast_to(lat0_arr, broadcast_shape)

    # (1) 强制模式：直接大圆
    if force_great_circle:
        gc = great_circle_distance_m(lon_arr, lat_arr, lon0_arr, lat0_arr, wrap_dateline=wrap_dateline, radius_earth_m=radius_earth_m)
        return gc

    # (2) 高纬提前判定：无需先算平面
    if gc_lat_threshold is not None and np.any(np.abs(lat0_arr) >= gc_lat_threshold):
        gc = great_circle_distance_m(lon_arr, lat_arr, lon0_arr, lat0_arr, wrap_dateline=wrap_dateline, radius_earth_m=radius_earth_m)
        return gc

    # (3) 需要距离阈值判定，才计算平面距离
    scale = approximate_degree_length(lat0_arr)
    dlon = _minimal_lon_diff_deg(lon_arr, lon0_arr) if wrap_dateline else (lon_arr - lon0_arr)
    dlat = lat_arr - lat0_arr
    planar = np.hypot(
        dlon * scale['meters_per_degree_lon'],
        dlat * scale['meters_per_degree_lat']
    )

    if gc_distance_threshold_km is not None and np.max(planar) > gc_distance_threshold_km * 1000.0:
        gc = great_circle_distance_m(lon_arr, lat_arr, lon0_arr, lat0_arr, wrap_dateline=wrap_dateline, radius_earth_m=radius_earth_m)
        return gc

    if np.ndim(planar) == 0:
        return float(planar)
    return planar

def inside_radius_m_adaptive(
    lon: float | np.ndarray,
    lat: float | np.ndarray,
    lon0: float,
    lat0: float,
    radius_m: float | np.ndarray,
    enlarge: float = 1.0,
    wrap_dateline: bool = True,
    gc_lat_threshold: float | None = None,
    gc_distance_threshold_km: float | None = None,
    force_great_circle: bool = False
) -> np.ndarray | bool:
    """自适应包含判定：点是否落在以 (lon0, lat0) 为心、radius_m 为半径（可放大 enlarge 倍）的范围内。

    距离计算与 adaptive_distance_m 策略一致（高纬或大尺度自动切换大圆）。

    参数:
        - lon (float | np.ndarray): 目标点经度（标量或数组）。
        - lat (float | np.ndarray): 目标点纬度（标量或数组）。
        - lon0 (float): 中心点经度。
        - lat0 (float): 中心点纬度。
        - radius_m (float | np.ndarray): 判定半径（米），可为标量或与点广播一致的数组。
        - enlarge (float): 半径放大系数，默认 1.0。
        - wrap_dateline (bool): 为 True 时对经度差做跨日界线最短差归一，默认 True。
        - gc_lat_threshold (float | None): 高纬触发大圆的绝对纬度阈值；None 时取配置默认。
        - gc_distance_threshold_km (float | None): 平面最大距离超过该值（km）触发大圆；None 时取配置默认。
        - force_great_circle (bool): 强制全程使用大圆，默认 False。

    返回:
        - np.ndarray | bool: 与点同形状的布尔包含掩码；标量输入返回 bool。
    """
    dist = adaptive_distance_m(
        lon, lat, lon0, lat0,
        wrap_dateline=wrap_dateline,
        gc_lat_threshold=gc_lat_threshold,
        gc_distance_threshold_km=gc_distance_threshold_km,
        force_great_circle=force_great_circle,
    )
    r_arr = np.asarray(radius_m, dtype=float)
    if r_arr.shape not in ((), np.shape(dist)):
        r_arr = np.broadcast_to(r_arr, np.shape(dist))
    inside = dist <= r_arr * enlarge
    if np.shape(inside) == ():
        return bool(inside)
    return inside

def ellipse_patch_for_eddy(lon0: float, lat0: float, radius_m: float, enlarge: float = 1.0,
                           **kwargs):
    """构造在经纬度坐标下表示真实米尺度涡旋半径的椭圆补丁。

    在 lon-lat 图上同一米半径在经向与纬向的角度跨度不同（经向按 cos(lat) 收缩），故绘制为椭圆
    （width=Δlon、height=Δlat）。

    参数:
        - lon0 (float): 涡旋中心经度。
        - lat0 (float): 涡旋中心纬度。
        - radius_m (float): 涡旋半径（米）。
        - enlarge (float): 半径放大系数，默认 1.0。
        - **kwargs: 透传给 matplotlib.patches.Ellipse 的样式参数（默认 edgecolor='gray'、facecolor='none'、linestyle='--'、linewidth=1.0、zorder=3）。

    返回:
        - matplotlib.patches.Ellipse: 可直接 add_patch 到坐标轴的椭圆补丁。
    """
    scale = approximate_degree_length(lat0)
    m_per_deg_lat = scale['meters_per_degree_lat']
    m_per_deg_lon = scale['meters_per_degree_lon']
    radius_eff = radius_m * enlarge
    dlat = radius_eff / m_per_deg_lat
    dlon = radius_eff / m_per_deg_lon
    default_kwargs = dict(edgecolor='gray', facecolor='none', linestyle='--', linewidth=1.0, zorder=3)
    default_kwargs.update(kwargs)
    return Ellipse((lon0, lat0), width=2*dlon, height=2*dlat, **default_kwargs)

def switch_region(
    region_name: str,
    config_path: str | Path = 'config/regions.yml',
    verbose: bool = True,
):
    """在运行时切换默认区域（无需改 YAML），并刷新全局 lonmin/lonmax/latmin/latmax 与 _REGION_CFG。

    推荐先 ``import track`` 再 ``track.switch_region(...)``，之后通过 ``track.lonmin`` 等访问最新值；
    不要用 ``from track import lonmin`` 绑定数值副本，否则后续切换不会更新它。

    参数:
        - region_name (str): 在 regions.yml 中定义的区域 key。
        - config_path (str | Path): 配置文件路径，默认 'config/regions.yml'。
        - verbose (bool): 是否打印区域切换提示，默认 True；并行 worker 中可传 False 避免刷屏。

    说明:
        - 抛出 KeyError：区域未找到或配置加载失败（进入 fallback）。
    """
    global _REGION_CFG, lonmin, lonmax, latmin, latmax
    new_cfg = _load_region_config(config_path=config_path, region=region_name)
    if new_cfg.get('_fallback'):
        raise KeyError(f"Region '{region_name}' not found or config load failed; fallback config in use.")
    _REGION_CFG = new_cfg
    lonmin, lonmax = _REGION_CFG['lon_min'], _REGION_CFG['lon_max']
    latmin, latmax = _REGION_CFG['lat_min'], _REGION_CFG['lat_max']
    if verbose:
        if _REGION_CFG.get('crosses_dateline') and (lonmax < lonmin):
            print(f"[RegionConfig] Region '{region_name}' crosses dateline; implement split-range filtering if needed.")
        else:
            print(f"[RegionConfig] Switched to region '{region_name}': lon[{lonmin}, {lonmax}], lat[{latmin}, {latmax}]")

def load_meta_data(path: str | os.PathLike | None = None, version: float = 3.2):
    '''
    加载 META 涡旋数据，返回 (ACS, ACL, CS, CL)。

    参数:
        - path (str | os.PathLike | None): META 数据根目录；None 时取配置 paths.meta_root（默认 '../META3.2_DT_allsat'）。
        - version (float): META 版本，3.1 或 3.2，默认 3.2。

    返回:
        - tuple: (ACS, ACL, CS, CL) 四个已打开的 netCDF4.Dataset，依次为反气旋短/长生命期、气旋短/长生命期。

    说明:
        - 目录需包含对应命名的 NetCDF 文件，如 version=3.2 时 `META3.2_DT_allsat_{Anticyclonic,Cyclonic}_{short,long}_*.nc`。
    '''
    if path is None:
        path = META_root_path
    else:
        path = Path(path)
    if version == 3.2:
        ACS= Dataset(path / 'META3.2_DT_allsat_Anticyclonic_short_19930101_20220209.nc')
        ACL= Dataset(path / 'META3.2_DT_allsat_Anticyclonic_long_19930101_20220209.nc')
        CS= Dataset(path / 'META3.2_DT_allsat_Cyclonic_short_19930101_20220209.nc')
        CL= Dataset(path / 'META3.2_DT_allsat_Cyclonic_long_19930101_20220209.nc')
    elif version == 3.1:
        ACS=Dataset(path / 'META3.1exp_DT_allsat_Anticyclonic_short_19930101_20200307.nc')
        ACL=Dataset(path / 'META3.1exp_DT_allsat_Anticyclonic_long_19930101_20200307.nc')
        CS=Dataset(path / 'META3.1exp_DT_allsat_Cyclonic_short_19930101_20200307.nc')
        CL=Dataset(path / 'META3.1exp_DT_allsat_Cyclonic_long_19930101_20200307.nc')
    else:
        raise ValueError("Unsupported version. Please use 3.1 or 3.2.")

    return ACS, ACL, CS, CL


def area_limit(DS, latmin, latmax, lonmin, lonmax):
    '''
    将 META 涡旋数据集按经纬度范围裁剪到指定区域，并按轨迹分组返回逐字段记录。

    参数:
        - DS (netCDF4.Dataset): 已打开的 META 涡旋数据集。
        - latmin (float): 纬度下界。
        - latmax (float): 纬度上界。
        - lonmin (float): 经度下界（经度掩码经 _region_lon_mask 做跨日界线安全处理）。
        - lonmax (float): 经度上界。

    返回:
        - list: 按轨迹分组的嵌套列表；每条轨迹含若干日项，每个日项依次为 [涡旋序号, 时间, 中心点经度, 中心点纬度, 最值点经度, 最值点纬度, 边界经度, 边界纬度, 半径, 速度边界经度, 速度边界纬度]。
    '''
    time = DS.variables['time'][:].data
    if np.issubdtype(time.dtype, np.floating):
        # 如果时间是浮点数，转换为整数
        time = np.round(time).astype(np.uint32)
    # time = convert_date(time)
    center_lon = DS.variables['longitude'][:].data
    center_lat = DS.variables['latitude'][:].data
    max_lon = DS.variables['longitude_max'][:].data
    max_lat = DS.variables['latitude_max'][:].data
    contour_lon = DS.variables['effective_contour_longitude'][:].data
    contour_lat = DS.variables['effective_contour_latitude'][:].data
    radius = DS.variables['effective_radius'][:].data
    track = DS.variables['track'][:].data
    speed_contour_lon = DS.variables['speed_contour_longitude'][:].data
    speed_contour_lat = DS.variables['speed_contour_latitude'][:].data

    lon_mask = _region_lon_mask(center_lon, lonmin, lonmax)
    mask = (center_lat >= latmin) & (center_lat <= latmax) & lon_mask
    indices = np.nonzero(mask)[0]
    ds = []
    current_track = None
    current_list = []
    for i in indices:
       if current_track is None:
            current_track = track[i]
       # 当track变化时，保存之前的列表，并新建一个列表
       if track[i] != current_track:
            ds.append(current_list)
            current_list = []
            current_track = track[i]
       current_list.append([i,time[i], center_lon[i], center_lat[i], 
              max_lon[i], max_lat[i], contour_lon[i], contour_lat[i], 
              radius[i], speed_contour_lon[i], speed_contour_lat[i]])
    if current_list:
              ds.append(current_list)
    return ds

def _ensure_meta_tracks_root(root: str | Path | None = None) -> Path:
    """获取/创建 META_tracks 根目录。优先读取配置 `paths.META_tracks_root`，兼容 `paths.meta_tracks_root`；默认 `./META_tracks`。"""
    default_root = Path('./META_tracks')
    paths_cfg = _PATHS_CFG.get('paths', {})
    cfg_val = paths_cfg.get('META_tracks_root', default_root)
    cfg_root = Path(cfg_val)
    if root is not None:
        cfg_root = Path(root)
    cfg_root.mkdir(parents=True, exist_ok=True)
    return cfg_root

def _current_region_key() -> str:
    # regions.yml 中的 key 并未直接保存在 _REGION_CFG；此处基于 fallback 标志无法反推 key。
    # 方案：若 _REGION_CFG 包含 name 且不是 fallback，取 name 的 slug；否则 generic。
    name = _REGION_CFG.get('name') or 'region'
    slug = name.lower().replace(' ', '_')
    return slug


def _current_region_config_key() -> str | None:
    """返回当前区域在 regions.yml 中的配置 key（如 global、kuroshio_extension）。"""
    key = _REGION_CFG.get('_region_key')
    if key is None:
        return None
    return str(key)

def export_meta_tracks(ds: Dataset,
                       kind: str,
                       region_key: str | None = None,
                       output_root: str | Path | None = None,
                       chunk_size: int = 100_000,
                       write_contours: bool = False,
                       build_track_summary: bool = True,
                       keep_legacy_pickle: bool = False,
                       use_dask: bool = False,
                       dask_num_workers: int | None = None,
                       compact_after: bool = True) -> dict:
    """流式导出 META 涡旋数据为标准化拆表格式（Parquet + Zarr）。

    参数:
        - ds (netCDF4.Dataset): 已打开的 META 文件。
        - kind (str): 标识（如 'acs'/'acl'/'cs'/'cl'），用于文件前缀。
        - region_key (str | None): 区域 key / slug；默认基于当前 _REGION_CFG['name'] 生成。
        - output_root (str | Path | None): 根输出目录；默认 ./META_tracks 或配置 paths.META_tracks_root。
        - chunk_size (int): indices 分块大小（按满足区域筛选的记录索引数量），默认 100000。
        - write_contours (bool): 是否同时写 contours 的 Parquet 拆表（顶点逐行），Zarr 始终生成，默认 False。
        - build_track_summary (bool): 是否生成轨迹汇总表，默认 True。
        - keep_legacy_pickle (bool): 是否额外写出旧嵌套结构 pickle（临时兼容调试），默认 False。
        - use_dask (bool): 是否用 Dask 并行按块处理，默认 False。
        - dask_num_workers (int | None): Dask 并行进程数。
        - compact_after (bool): 完成后是否将目录形式的 Parquet 压实为单文件并删除目录，默认 True。

    返回:
        - dict: 写出的关键文件路径（如 daily_file/daily_dir、tracks_path、contours_zarr、metadata 等）。

    输出:
        - `<kind>_daily.parquet`：每个涡旋每日一行（标量 & 计数）。
        - `<kind>_contours.zarr`：轮廓顶点按 1D 扁平数组存储，含 prefix index（N+1）。
        - `<kind>_contours.parquet`：（可选）展开轮廓顶点逐行记录（用于 SQL/分析）。
        - `<kind>_tracks.parquet`：（可选）汇总轨迹层（聚合统计）。
        - `<kind>_metadata.json`：基本元信息（区域、生成时间、列说明、Zarr 路径）。
    """
    # 记录总长度用于分块；避免一次性加载变量
    try:
        total_len = ds.dimensions['time'].size
    except Exception:
        total_len = ds.variables['time'].shape[0]

    region_slug = region_key or _current_region_key()
    root = _ensure_meta_tracks_root(output_root)
    region_dir = root / region_slug
    region_dir.mkdir(parents=True, exist_ok=True)

    # 输出路径：分片写入使用临时目录，压实后生成同名 .parquet 单文件，避免目录/文件冲突
    daily_stage = region_dir / f"{kind}_daily_tmp"
    contours_stage = region_dir / f"{kind}_contours_tmp"
    contours_zarr_path = region_dir / f"{kind}_contours.zarr"
    contours_parts_dir = region_dir / f"{kind}_contours_zarr_parts"
    tracks_path = region_dir / f"{kind}_tracks.parquet"
    meta_path = region_dir / f"{kind}_metadata.json"
    legacy_pickle_path = region_dir / f"{kind}_legacy_nested.pkl"

    # 清理旧的分片临时目录（若存在）
    if daily_stage.exists():
        if daily_stage.is_dir():
            shutil.rmtree(daily_stage)
        else:
            daily_stage.unlink()
    if write_contours and contours_stage.exists():
        if contours_stage.is_dir():
            shutil.rmtree(contours_stage)
        else:
            contours_stage.unlink()
    # 清理旧 Zarr 目标和分片目录
    if contours_zarr_path.exists():
        shutil.rmtree(contours_zarr_path)
    if contours_parts_dir.exists():
        shutil.rmtree(contours_parts_dir)

    daily_stage.mkdir(parents=True, exist_ok=True)
    if write_contours:
        contours_stage.mkdir(parents=True, exist_ok=True)
    contours_parts_dir.mkdir(parents=True, exist_ok=True)

    daily_schema = pa.schema([
        pa.field('track_id', pa.int64()),
        pa.field('time', pa.timestamp('ns')),
        pa.field('center_lon', pa.float64()),
        pa.field('center_lat', pa.float64()),
        pa.field('max_lon', pa.float64()),
        pa.field('max_lat', pa.float64()),
        pa.field('radius', pa.float64()),
        pa.field('contour_vertex_count', pa.int32()),
        pa.field('speed_contour_vertex_count', pa.int32()),
        pa.field('orig_index', pa.int64()),
    ])

    contour_schema = pa.schema([
        pa.field('track_id', pa.int64()),
        pa.field('time', pa.timestamp('ns')),
        pa.field('kind', pa.string()),  # 'effective' or 'speed'
        pa.field('vertex_index', pa.int32()),
        pa.field('lon', pa.float64()),
        pa.field('lat', pa.float64()),
    ]) if write_contours else None

    # 构造 chunk 写入
    def write_daily_chunk(tbl: pa.Table, idx: int):
        pq.write_table(tbl, daily_stage / f"part_{idx:05d}.parquet")

    def write_contour_chunk(tbl: pa.Table, idx: int, suffix: str):
        pq.write_table(tbl, contours_stage / f"part_{suffix}_{idx:05d}.parquet")

    # 旧结构临时兼容：按轨迹分组的嵌套列表（外层按 track，内层按日）
    # 为减少内存占用，使用 dict 累积，每个 track_id -> list[day_item]
    # day_item 结构与 area_limit 输出一致：[orig_i, time_YYYYMMDD, center_lon, center_lat, max_lon, max_lat, contour_lon[], contour_lat[], radius, speed_contour_lon[], speed_contour_lat[]]
    legacy_nested = {} if (keep_legacy_pickle and not use_dask) else None

    # 自动决定是否跳过区域筛选：当配置等同于全球范围时跳过
    is_global_lon = (lonmin <= -179.999) and (lonmax >= 179.999)
    is_global_lat = (latmin <= -89.999) and (latmax >= 89.999)
    auto_skip = bool(is_global_lon and is_global_lat)

    # 单块处理函数：可被串行调用，或用 dask.delayed 并发调用
    def _process_chunk(file_path: str, part_idx: int, start: int, end: int) -> int:
        ds_local = Dataset(file_path)
        try:
            # 读取该块变量
            time_var = ds_local.variables['time'][start:end]
            center_lon = ds_local.variables['longitude'][start:end]
            center_lat = ds_local.variables['latitude'][start:end]
            n_chunk = len(center_lon)
            # 区域过滤（单块）；若跳过则全选
            if auto_skip:
                sel = np.arange(n_chunk, dtype=int)
            else:
                lon_mask = _region_lon_mask(center_lon, lonmin, lonmax)
                mask = (center_lat >= latmin) & (center_lat <= latmax) & lon_mask
                if not np.any(mask):
                    return 0
                sel = np.nonzero(mask)[0]
            # 仅在需要时读取其它变量
            max_lon = ds_local.variables['longitude_max'][start:end]
            max_lat = ds_local.variables['latitude_max'][start:end]
            radius = ds_local.variables['effective_radius'][start:end]
            track_id = ds_local.variables['track'][start:end]
            eff_lon = ds_local.variables['effective_contour_longitude'][start:end]
            eff_lat = ds_local.variables['effective_contour_latitude'][start:end]
            spd_lon = ds_local.variables['speed_contour_longitude'][start:end]
            spd_lat = ds_local.variables['speed_contour_latitude'][start:end]

            # 时间转换到 datetime64[ns]
            time_dt = convert_date(pd.Series(time_var)).to_numpy().astype('datetime64[ns]')

            # 组装并写出
            rows = []
            eff_rows = []
            spd_rows = []
            # Zarr 分片缓存
            e_counts_list = []
            s_counts_list = []
            eff_concat_lon = []
            eff_concat_lat = []
            spd_concat_lon = []
            spd_concat_lat = []
            for off in sel:
                orig_i = start + int(off)
                v_track = int(track_id[off])
                v_time = time_dt[off]
                c_lon = float(center_lon[off])
                c_lat = float(center_lat[off])
                mx_lon = float(max_lon[off])
                mx_lat = float(max_lat[off])
                rad = float(radius[off])
                e_lon = eff_lon[off]
                e_lat = eff_lat[off]
                s_lon = spd_lon[off]
                s_lat = spd_lat[off]
                e_cnt = len(e_lon) if hasattr(e_lon, '__len__') else 0
                s_cnt = len(s_lon) if hasattr(s_lon, '__len__') else 0
                rows.append((v_track, v_time, c_lon, c_lat, mx_lon, mx_lat, rad, e_cnt, s_cnt, int(orig_i)))

                if write_contours and e_cnt:
                    for vi in range(e_cnt):
                        eff_rows.append((v_track, v_time, 'effective', vi, float(e_lon[vi]), float(e_lat[vi])))
                if write_contours and s_cnt:
                    for vi in range(s_cnt):
                        spd_rows.append((v_track, v_time, 'speed', vi, float(s_lon[vi]), float(s_lat[vi])))

                # 记录 Zarr 分片内容（每日计数与扁平顶点）
                e_counts_list.append(int(e_cnt))
                s_counts_list.append(int(s_cnt))
                if e_cnt:
                    eff_concat_lon.append(np.asarray(e_lon, dtype='f8'))
                    eff_concat_lat.append(np.asarray(e_lat, dtype='f8'))
                if s_cnt:
                    spd_concat_lon.append(np.asarray(s_lon, dtype='f8'))
                    spd_concat_lat.append(np.asarray(s_lat, dtype='f8'))

                if legacy_nested is not None:
                    time_ymd = int(str(v_time)[:10].replace('-', ''))
                    item = [
                        int(orig_i), time_ymd, c_lon, c_lat,
                        mx_lon, mx_lat, e_lon, e_lat, rad, s_lon, s_lat
                    ]
                    legacy_nested.setdefault(v_track, []).append(item)

            if rows:
                daily_tbl = pa.Table.from_pylist([
                    {
                        'track_id': r[0], 'time': r[1], 'center_lon': r[2], 'center_lat': r[3],
                        'max_lon': r[4], 'max_lat': r[5], 'radius': r[6],
                        'contour_vertex_count': r[7], 'speed_contour_vertex_count': r[8], 'orig_index': r[9]
                    } for r in rows
                ], schema=daily_schema)
                write_daily_chunk(daily_tbl, part_idx)

            if write_contours:
                if eff_rows:
                    eff_tbl = pa.Table.from_pylist([
                        {
                            'track_id': r[0], 'time': r[1], 'kind': r[2], 'vertex_index': r[3], 'lon': r[4], 'lat': r[5]
                        } for r in eff_rows
                    ], schema=contour_schema)
                    write_contour_chunk(eff_tbl, part_idx, 'eff')
                if spd_rows:
                    spd_tbl = pa.Table.from_pylist([
                        {
                            'track_id': r[0], 'time': r[1], 'kind': r[2], 'vertex_index': r[3], 'lon': r[4], 'lat': r[5]
                        } for r in spd_rows
                    ], schema=contour_schema)
                    write_contour_chunk(spd_tbl, part_idx, 'spd')

            # 将本分块的 Zarr 数据写为 NPZ 分片，供后续合并
            if rows:
                e_counts = np.asarray(e_counts_list, dtype='i8')
                s_counts = np.asarray(s_counts_list, dtype='i8')
                eff_lon_concat = np.concatenate(eff_concat_lon) if eff_concat_lon else np.empty((0,), dtype='f8')
                eff_lat_concat = np.concatenate(eff_concat_lat) if eff_concat_lat else np.empty((0,), dtype='f8')
                spd_lon_concat = np.concatenate(spd_concat_lon) if spd_concat_lon else np.empty((0,), dtype='f8')
                spd_lat_concat = np.concatenate(spd_concat_lat) if spd_concat_lat else np.empty((0,), dtype='f8')
                np.savez_compressed(
                    contours_parts_dir / f"part_{part_idx:05d}.npz",
                    eff_lon=eff_lon_concat,
                    eff_lat=eff_lat_concat,
                    eff_counts=e_counts,
                    spd_lon=spd_lon_concat,
                    spd_lat=spd_lat_concat,
                    spd_counts=s_counts,
                )
            return len(rows)
        finally:
            try:
                ds_local.close()
            except Exception:
                pass

    # 生成块边界
    file_path = ds.filepath() if hasattr(ds, 'filepath') else None
    if file_path is None:
        # netCDF4.Dataset 在某些场景没有 filepath 属性，尝试从 repr 中提取失败则报错
        raise RuntimeError("Dataset filepath unavailable; Dask/streaming requires file path.")

    # 在执行前，根据数据量与目标并发自动调整 chunk_size，避免只有 1 个分块导致只能单核
    expected_workers = 1
    if use_dask:
        expected_workers = int(dask_num_workers) if dask_num_workers else max(1, (os.cpu_count() or 1))
    parts = max(1, (total_len + chunk_size - 1) // chunk_size)
    if use_dask and parts < expected_workers:
        target_parts = max(expected_workers * 3, 2)
        new_chunk = max(25_000, math.ceil(total_len / target_parts))
        if new_chunk != chunk_size:
            print(f"[export_meta_tracks] Auto-tune chunk_size: {chunk_size} -> {new_chunk} (total={total_len}, workers={expected_workers}, target_parts≈{target_parts})")
            chunk_size = new_chunk
            parts = max(1, (total_len + chunk_size - 1) // chunk_size)

    print(f"[export_meta_tracks] Plan kind={kind}: total={total_len}, chunk_size={chunk_size}, parts={parts}, parallel={use_dask}, workers={expected_workers}")

    # dask 并行或串行执行
    wrote_any = False
    if use_dask:
        if keep_legacy_pickle:
            print("[export_meta_tracks] keep_legacy_pickle=True 在 Dask 模式下会占用大量内存，已强制关闭。")
            legacy_nested = None

    if use_dask:
        tasks = []
        part_idx = 0
        for start in range(0, total_len, chunk_size):
            end = min(total_len, start + chunk_size)
            task = delayed(_process_chunk)(file_path, part_idx, start, end)
            tasks.append(task)
            part_idx += 1
        scheduler_kwargs = {}
        if dask_num_workers is not None:
            scheduler_kwargs['num_workers'] = dask_num_workers
        with ProgressBar():
            results = compute(*tasks, scheduler='processes', **scheduler_kwargs)
        total_rows = int(np.sum(results))
        wrote_any = total_rows > 0
        print(f"[export_meta_tracks] kind={kind} dask-complete parts={len(tasks)} rows={total_rows}")
    else:
        parts = (total_len + chunk_size - 1) // chunk_size
        for part_idx, start in enumerate(tqdm(range(0, total_len, chunk_size), total=parts, desc=f"export {kind} serial", unit="part")):
            end = min(total_len, start + chunk_size)
            nrows = _process_chunk(file_path, part_idx, start, end)
            wrote_any = wrote_any or (nrows > 0)
            # 串行时不需要额外计数，最终通过分片合并生成 index

    # 轨迹汇总
    written = {
        'daily_dir': str(daily_stage),
        'contours_dir': str(contours_stage) if write_contours else None,
    }

    if build_track_summary and wrote_any:
        # 读回 daily parquet dataset 聚合（数据量相比 contours 小得多）
        daily_dataset = pq.ParquetDataset(daily_stage)
        daily_tbl = daily_dataset.read()
        df = daily_tbl.to_pandas()
        grp = df.groupby('track_id', sort=False)
        summary = grp.agg(
            n_points=('time', 'count'),
            time_start=('time', 'min'),
            time_end=('time', 'max'),
            center_lon_mean=('center_lon', 'mean'),
            center_lat_mean=('center_lat', 'mean'),
            radius_mean=('radius', 'mean'),
            radius_max=('radius', 'max'),
        ).reset_index()
        pq.write_table(pa.Table.from_pandas(summary), tracks_path)
        written['tracks_path'] = str(tracks_path)

    # 合并 NPZ 分片为单一 Zarr 存储（支持串行和 Dask）
    def _merge_zarr_parts(parts_dir: Path, out_path: Path, meta_info: dict):
        part_files = sorted(parts_dir.glob('part_*.npz'))
        if not part_files:
            return None
        # 第一遍：统计总长度
        eff_rows_total = 0
        spd_rows_total = 0
        eff_vert_total = 0
        spd_vert_total = 0
        for pf in part_files:
            with np.load(pf) as npz:
                ec = npz['eff_counts']
                sc = npz['spd_counts']
                eff_rows_total += int(ec.size)
                spd_rows_total += int(sc.size)
                eff_vert_total += int(npz['eff_lon'].size)
                spd_vert_total += int(npz['spd_lon'].size)
        # 创建 Zarr 存储
        if out_path.exists():
            shutil.rmtree(out_path)
        root = zarr.open_group(out_path, mode='w')
        z_eff = root.create_group('effective')
        z_spd = root.create_group('speed')
        z_meta = root.create_group('meta')
        eff_lon_arr = z_eff.create_array('lon', shape=(eff_vert_total,), chunks=(262144,), dtype='f8')
        eff_lat_arr = z_eff.create_array('lat', shape=(eff_vert_total,), chunks=(262144,), dtype='f8')
        spd_lon_arr = z_spd.create_array('lon', shape=(spd_vert_total,), chunks=(262144,), dtype='f8')
        spd_lat_arr = z_spd.create_array('lat', shape=(spd_vert_total,), chunks=(262144,), dtype='f8')
        eff_idx_arr = z_eff.create_array('index', shape=(eff_rows_total + 1,), chunks=(262144,), dtype='i8')
        spd_idx_arr = z_spd.create_array('index', shape=(spd_rows_total + 1,), chunks=(262144,), dtype='i8')
        eff_idx_arr[0] = 0
        spd_idx_arr[0] = 0
        z_meta.attrs.update(meta_info)
        # 第二遍：按顺序写入
        e_vert_pos = 0
        s_vert_pos = 0
        e_row_pos = 0
        s_row_pos = 0
        for pf in part_files:
            with np.load(pf) as npz:
                ev_lon = npz['eff_lon']; ev_lat = npz['eff_lat']; ec = npz['eff_counts']
                sv_lon = npz['spd_lon']; sv_lat = npz['spd_lat']; sc = npz['spd_counts']
                if ev_lon.size:
                    eff_lon_arr[e_vert_pos:e_vert_pos+ev_lon.size] = ev_lon
                    eff_lat_arr[e_vert_pos:e_vert_pos+ev_lat.size] = ev_lat
                    e_vert_pos += int(ev_lon.size)
                if sv_lon.size:
                    spd_lon_arr[s_vert_pos:s_vert_pos+sv_lon.size] = sv_lon
                    spd_lat_arr[s_vert_pos:s_vert_pos+sv_lat.size] = sv_lat
                    s_vert_pos += int(sv_lon.size)
                if ec.size:
                    csum = np.cumsum(ec, dtype=np.int64)
                    eff_idx_arr[e_row_pos+1:e_row_pos+1+csum.size] = eff_idx_arr[e_row_pos] + csum
                    e_row_pos += int(ec.size)
                if sc.size:
                    csum = np.cumsum(sc, dtype=np.int64)
                    spd_idx_arr[s_row_pos+1:s_row_pos+1+csum.size] = spd_idx_arr[s_row_pos] + csum
                    s_row_pos += int(sc.size)
        return str(out_path)

    contours_zarr_written = None
    if True:
        meta_info = {
            'kind': kind,
            'region': region_slug,
            'lon_min': float(lonmin), 'lon_max': float(lonmax),
            'lat_min': float(latmin), 'lat_max': float(latmax),
        }
        try:
            contours_zarr_written = _merge_zarr_parts(contours_parts_dir, contours_zarr_path, meta_info)
        finally:
            if contours_parts_dir.exists():
                try:
                    shutil.rmtree(contours_parts_dir)
                except Exception:
                    pass

    # 元信息写出
    meta = {
        'kind': kind,
        'region': region_slug,
        'lon_min': float(lonmin), 'lon_max': float(lonmax),
        'lat_min': float(latmin), 'lat_max': float(latmax),
        'generated_at': str(np.datetime64('now')),
        'columns_daily': [f.name for f in daily_schema],
        # 兼容旧键：Parquet 顶点拆表
        'has_contours': write_contours,
        'contour_schema': [f.name for f in contour_schema] if contour_schema else None,
        # 新增：Zarr 轮廓信息
        'has_contours_zarr': bool(contours_zarr_written),
        'contours_format': 'zarr' if contours_zarr_written else None,
        'contours_zarr': contours_zarr_written,
        'track_summary': bool(build_track_summary),
        'legacy_nested_pickle': bool(keep_legacy_pickle),
    }
    with open(meta_path, 'w') as f:
        json.dump(meta, f, ensure_ascii=False, indent=2)
    written['metadata'] = str(meta_path)

    if keep_legacy_pickle and legacy_nested is not None:
        # 把 dict 转为 list[list]，保持“首次出现”的轨迹顺序，以及每轨迹内的原始追加顺序
        nested_list = list(legacy_nested.values())
        with open(legacy_pickle_path, 'wb') as f:
            pickle.dump(nested_list, f)
        written['legacy_pickle'] = str(legacy_pickle_path)
        print(f"[export_meta_tracks] Legacy nested pickle written: {legacy_pickle_path}")

    # 压实：将目录 Dataset 合并为单文件，并删除目录
    if wrote_any and compact_after:
        try:
            daily_file = compact_parquet_dataset(kind, 'daily', region_slug, root, delete_source=True)
            written['daily_file'] = daily_file
            written['daily_dir'] = None
        except Exception as e:
            print(f"[export_meta_tracks] compact daily failed: {e}")
        if write_contours:
            try:
                contours_file = compact_parquet_dataset(kind, 'contours', region_slug, root, delete_source=True)
                written['contours_file'] = contours_file
                written['contours_dir'] = None
            except Exception as e:
                print(f"[export_meta_tracks] compact contours failed: {e}")

    # 简单 Zarr 校验
    if contours_zarr_written:
        try:
            zroot = zarr.open_group(contours_zarr_path, mode='r')
            eff_idx_len = int(zroot['effective']['index'].shape[0])
            spd_idx_len = int(zroot['speed']['index'].shape[0])
            if eff_idx_len != spd_idx_len:
                print(f"[export_meta_tracks] Warning: effective/speed index length mismatch: {eff_idx_len} vs {spd_idx_len}")
        except Exception as e:
            print(f"[export_meta_tracks] Zarr open failed for validation: {e}")

    if not wrote_any:
        print(f"[export_meta_tracks] No records inside region bounds for kind={kind}.")
    out_hint = written.get('daily_file', str(daily_stage))
    if contours_zarr_written:
        written['contours_zarr'] = contours_zarr_written
    print(f"[export_meta_tracks] Completed kind={kind}. Daily: {out_hint}; Contours(zarr): {contours_zarr_written}")
    return written

def convert_mat_to_parquet(year: int, input_dir: str | Path = None, output_dir: str | Path = None):
    """将某年份的逐月 Argo .mat 原始文件合并为单一 Parquet。

    参数:
        - year (int): 年份（如 2008）。
        - input_dir (str | Path | None): 每月 .mat 文件所在目录；None 时取配置 paths.argo_mat_input。
        - output_dir (str | Path | None): 输出目录；None 时取配置 paths.argo_parquet。

    说明:
        - 当前仅示例性地提取 'do' 数据集（若存在）并按固定列顺序写出；如需更多变量可在此扩展。
    """
    year_str = str(year)
    if input_dir is None:
        input_dir = argo_mat_input_path
    else:
        input_dir = Path(input_dir)
    if output_dir is None:
        output_dir = argo_path
    else:
        output_dir = Path(output_dir)

    columns = [
        "Year", "Month", "Day", "Longitude", "Latitude", "Depth_m", "DO_mol_kg", "Salinity_psu",
        "Temperature_degC", "Oxygen_flag", "Oxygen_flag2", "Profile_number", "Datasets_number",
        "Platform_number", "Cycle_number", "Float_serial_no"
    ]

    paths = [input_dir / f'Argo{year_str}_{m}.mat' for m in range(1, 13)]
    print(f"[convert_mat_to_parquet] Year {year_str}: expecting {len(paths)} monthly files under {input_dir}")

    all_monthly = []
    for p in paths:
        if not p.exists():
            print(f"  - Missing file: {p}")
            continue
        try:
            with h5py.File(p, 'r') as f:
                if 'do' not in f:
                    print(f"  - 'do' dataset absent in {p}, skipping")
                    continue
                do_data = f['do'][:].T  # 原数据转置为 (n_rows, n_cols)
                df = pd.DataFrame(do_data, columns=columns[:do_data.shape[1]])  # 防御性截断
                all_monthly.append(df)
                print(f"  + Loaded {p}")
        except Exception as e:
            print(f"  ! Error reading {p}: {e}")

    if not all_monthly:
        print(f"[convert_mat_to_parquet] No monthly data loaded for {year_str}; abort.")
        return

    final_df = pd.concat(all_monthly, ignore_index=True)
    output_dir.mkdir(parents=True, exist_ok=True)
    out_path = output_dir / f'Argo{year_str}.parquet'
    try:
        final_df.to_parquet(out_path, index=False)
        print(f"[convert_mat_to_parquet] Saved merged parquet: {out_path}")
    except Exception as e:
        print(f"[convert_mat_to_parquet] Failed to save {out_path}: {e}")

def compact_parquet_dataset(kind: str,
                            which: str,
                            region_key: str | None = None,
                            output_root: str | Path | None = None,
                            delete_source: bool = True) -> str:
    """将目录形式的 Parquet Dataset (many part_*.parquet) 压实为单一文件，并可选删除源目录。

    参数:
        - kind (str): 数据类型前缀，如 'acs'/'acl'/'cs'/'cl'。
        - which (str): 'daily' 或 'contours'。
        - region_key (str | None): 区域 slug；None 时基于当前区域生成。
        - output_root (str | Path | None): 根输出目录；None 时读取配置或用默认 './META_tracks'。
        - delete_source (bool): 压实完成后是否删除源目录，默认 True。

    返回:
        - str: 目标单一 Parquet 文件路径。
    """
    region_slug = region_key or _current_region_key()
    root = _ensure_meta_tracks_root(output_root)
    region_dir = root / region_slug
    if which not in { 'daily', 'contours' }:
        raise ValueError("which 必须是 'daily' 或 'contours'")

    # 优先使用 *_tmp 作为分片目录；若不存在，则兼容旧的 *_<which>.parquet 目录
    candidate_dirs = [
        region_dir / f"{kind}_{which}_tmp",
        region_dir / f"{kind}_{which}.parquet",
    ]
    dir_path = None
    for cand in candidate_dirs:
        if cand.exists() and cand.is_dir():
            dir_path = cand
            break
    if dir_path is None:
        raise FileNotFoundError(f"未找到分片目录: {candidate_dirs[0]} 或 {candidate_dirs[1]}")

    # 目标文件路径（最终与目录同名，但先写 tmp，再替换）
    tmp_path = region_dir / f"{kind}_{which}.parquet.tmp"
    final_file = region_dir / f"{kind}_{which}.parquet"

    # 收集分片文件
    part_files = sorted([p for p in dir_path.glob('*.parquet') if p.is_file()])
    if not part_files:
        raise RuntimeError(f"未发现任何分片文件于 {dir_path}")

    # 流式合并
    first_tbl = pq.read_table(part_files[0])
    schema = first_tbl.schema
    # 若存在旧 tmp，先移除
    if tmp_path.exists():
        tmp_path.unlink()
    rows_total = 0
    with pq.ParquetWriter(tmp_path, schema) as writer:
        for pf in part_files:
            tbl = pq.read_table(pf)
            writer.write_table(tbl)
            rows_total += tbl.num_rows

    # 删除源目录，重命名 tmp 为最终文件（若已存在则先删除）
    if delete_source:
        shutil.rmtree(dir_path)
    if final_file.exists():
        final_file.unlink()
    tmp_path.rename(final_file)
    print(f"[compact] {which} -> {final_file} (rows≈{rows_total})")
    return str(final_file)

def parse_argo_txt_file(file_path: Path) -> pd.DataFrame | None:
    """
    解析单个扁平的、由制表符分隔的 Argo 表格文件。

    假设输入文件第一行是列标题、后续行是数据、字段以制表符（Tab）分隔，并自动将占位符（如 -999）
    转换为标准 NaN。

    参数:
        - file_path (Path): 单个 Argo txt 文件的路径。

    返回:
        - pd.DataFrame | None: 解析成功且内容非空时返回 DataFrame；否则返回 None。
    """
    try:
        # 定义需要被视作NaN的占位符列表
        missing_values = [-999, -999.0, -999.00, -999.000, -999.0000, -999.000000]
        
        # 在读取时使用 na_values 参数直接完成替换
        df = pd.read_csv(
            file_path,
            sep='\t',
            na_values=missing_values
        )
        
        # 如果文件只有表头或完全为空，df.empty会是True
        if df.empty:
            return None
        return df
    except Exception as e:
        # print(f"  [Warning] 读取文件 {file_path.name} 失败: {e}")
        return None

def worker_process_file(task_args: tuple[Path, Path]):
    """
    Dask 工作单元：处理单个文件。

    接收含输入/输出路径的元组，调用 parse_argo_txt_file 解析，并将结果保存为中间 Parquet 文件。

    参数:
        - task_args (tuple[Path, Path]): (原始 txt 文件路径 input_path, 中间 Parquet 输出路径 output_path)。
    """
    input_path, output_path = task_args
    try:
        df = parse_argo_txt_file(input_path)
        if df is not None:
            output_path.parent.mkdir(parents=True, exist_ok=True)
            df.to_parquet(output_path, index=False)
    except Exception as e:
        print(f"\n[Error] 在工作进程中处理文件 {input_path.name} 时出错: {e}")


def process_argo_txt_to_yearly_parquet_dask(
    origin_dir: Path | str | None = None,
    temp_dir: Path | str | None = None,
    final_dir: Path | str | None = None,
    cleanup_temp_dir: bool = True
):
    """
    使用 Dask 统一并行处理 Argo txt 文件，高效完成大规模数据的 ETL。

    完整的 Dask 驱动 ETL：先并行将所有 .txt 转为临时 .parquet，再并行合并、清洗、排序并输出为按年份
    分区的临时目录，最后串行合并为单个年度文件；利用 Dask 的 HPC 自适应能力避免内存瓶颈，并按参数选择
    是否清理临时文件。

    参数:
        - origin_dir (Path | str | None): 存放原始 Argo .txt 文件的目录；None 时取配置 paths.argo_txt_input，缺省则回退 './Argo_origin'。
        - temp_dir (Path | str | None): 中间产物（初始 Parquet、映射表、分区数据）的临时目录；None 时取配置 paths.tmp_parquet_path，缺省则回退 final_dir/'_tmp_txt2parquet_dask'。
        - final_dir (Path | str | None): 保存最终年度 Parquet（ArgoYYYY.parquet）的目录；None 时取配置 paths.argo_parquet。
        - cleanup_temp_dir (bool): 是否在任务结束后删除临时目录，默认 True。
    """
    # --- 0. 解析/回退目录参数（与 META 一致的配置驱动） ---
    start_total_time = tm.time()
    origin_dir = Path(origin_dir) if origin_dir is not None else Path(_PATHS_CFG.get('paths', {}).get('argo_txt_input', './Argo_origin'))
    final_dir = Path(final_dir) if final_dir is not None else Path(argo_path)
    if temp_dir is None:
        # 使用已在模块顶层解析好的全局 tmp_parquet_path（paths.yml: argo_intermediate），
        # 若用户未配置则回退到 final_dir/_tmp_txt2parquet_dask
        temp_dir = Path(tmp_parquet_path) if tmp_parquet_path else (final_dir / '_tmp_txt2parquet_dask')
    else:
        temp_dir = Path(temp_dir)
    print("[*] Using directories for Argo TXT → Parquet:")
    print(f"    - origin_dir = {origin_dir}")
    print(f"    - temp_dir   = {temp_dir}")
    print(f"    - final_dir  = {final_dir}")
    
    # --- 准备工作：初始化Dask客户端并创建目录 ---
    client = Client()
    print(f"[*] Dask客户端已启动，仪表盘链接: {client.dashboard_link}")
    
    scheduler_info = client.scheduler_info()
    workers_count = len(scheduler_info['workers'])
    total_threads = sum(worker['nthreads'] for worker in scheduler_info['workers'].values())
    print(f"[*] Dask自动检测到 {workers_count} 个 worker，共 {total_threads} 个工作核心。")
    
    print("[*] 准备工作环境...")
    if temp_dir.exists():
        print(f"  - 发现旧的临时目录，正在清理: {temp_dir}")
        shutil.rmtree(temp_dir)
    temp_dir.mkdir(parents=True, exist_ok=True)
    final_dir.mkdir(parents=True, exist_ok=True)
    
    # --- 阶段一: 使用Dask并行解析 -> 中间文件 ---
    print("\n--- 阶段一: 开始并行解析原始 .txt 文件 ---")
    txt_files = list(origin_dir.glob('*.txt'))
    if not txt_files:
        print("[Error] 输入目录中未找到任何 .txt 文件。程序终止。")
        client.close()
        return
        
    tasks = [(p, temp_dir / f"{p.stem}.parquet") for p in txt_files]
    print(f"[*] 找到 {len(txt_files)} 个文件，将任务分发给Dask处理...")
    
    start_map_time = tm.time()
    futures = client.map(worker_process_file, tasks)
    client.gather(futures)
    print(f"--- 阶段一完成 --- (耗时: {tm.time() - start_map_time:.2f} 秒)")
    
    # --- 阶段二: Dask并行ETL并输出分区文件 ---
    print("\n--- 阶段二: Dask开始并行处理中间文件 ---")
    start_reduce_time = tm.time()

    print(f"[*] Dask正在逻辑读取 {temp_dir} 中的Parquet文件...")
    ddf = dd.read_parquet(temp_dir / '*.parquet')
    record_count = len(ddf)
    print(f"[*] 数据逻辑加载完成，总计 {record_count:,} 条记录。")

    print("[*] 正在重命名列...")
    rename_map = {'WMO': 'Platform_number', 'Cycle': 'Profile_number', 'Lon': 'Longitude', 'Lat': 'Latitude'}
    ddf = ddf.rename(columns=rename_map)

    print("[*] 正在并行生成全局唯一的 Profile_number...")
    unique_profiles = ddf[['Platform_number', 'Profile_number']].drop_duplicates().compute()
    unique_profiles_sorted = unique_profiles.sort_values(by=['Platform_number', 'Profile_number'])
    unique_profiles_sorted['Global_Profile_Number'] = range(1, len(unique_profiles_sorted) + 1)
    mapping_file_path = temp_dir / 'profile_mapping.parquet'
    unique_profiles_sorted.to_parquet(mapping_file_path, index=False)
    mapping_ddf = dd.read_parquet(mapping_file_path)
    ddf = ddf.merge(mapping_ddf, on=['Platform_number', 'Profile_number'], how='left')
    ddf = ddf.drop(columns=['Profile_number']).rename(columns={'Global_Profile_Number': 'Profile_number'})

    print("[*] 正在优化并转换数据类型...")
    float64_cols = list(ddf.select_dtypes(include='float64').columns)
    if float64_cols:
        for col in float64_cols:
            ddf[col] = ddf[col].astype('float32')
    int_conversion_map = {'Year': 'Int16', 'Month': 'Int8', 'Day': 'Int8', 'Platform_number': 'Int64', 'Profile_number': 'Int64'}
    for col, dtype in int_conversion_map.items():
        if col in ddf.columns:
            ddf[col] = ddf[col].astype(dtype)
    
    print("[*] Dask开始执行所有计算并按年份分区写入临时目录...")
    ddf = ddf.dropna(subset=['Year'])
    partitioned_output_dir = temp_dir / 'dask_partitioned_output'
    ddf.to_parquet(partitioned_output_dir, write_index=False, partition_on=['Year'])
    
    end_reduce_time = tm.time()
    print(f"--- 阶段二完成 --- (耗时: {end_reduce_time - start_reduce_time:.2f} 秒)")
    
    # --- 阶段三: 将分区文件合并为最终产物 ---
    print("\n--- 阶段三: 开始合并分区文件并重排定序 ---")
    start_consolidation_time = tm.time()
    
    year_dirs = sorted([d for d in partitioned_output_dir.iterdir() if d.is_dir() and d.name.startswith('Year=')])
    for year_dir in year_dirs:
        year_str = year_dir.name.split('=')[1]
        year_int = int(year_str)
        final_single_file_path = final_dir / f"Argo{year_int}.parquet"
        print(f"  -> 合并年份 {year_int} -> {final_single_file_path}")
        
        year_df = pd.read_parquet(year_dir)
        year_df['Year'] = year_int
        year_df.sort_values(by=['Year', 'Month', 'Day', 'Platform_number'], inplace=True)
        
        start_cols = ['Year', 'Month', 'Day']
        end_cols = ['Profile_number', 'Platform_number']
        middle_cols = [col for col in year_df.columns if col not in start_cols + end_cols]
        final_column_order = start_cols + middle_cols + end_cols
        year_df = year_df[final_column_order]
        
        year_df.to_parquet(final_single_file_path, index=False)
        
    end_consolidation_time = tm.time()
    print(f"--- 阶段三完成 --- (耗时: {end_consolidation_time - start_consolidation_time:.2f} 秒)")

    # --- 清理工作 ---
    if cleanup_temp_dir:
        print("\n--- 清理阶段: 删除临时文件 ---")
        try:
            shutil.rmtree(temp_dir)
            print(f"[*] 临时目录 {temp_dir} 已成功删除。")
        except Exception as e:
            print(f"[Error] 删除临时目录失败: {e}")
    else:
        print("\n--- 清理阶段: 已跳过 ---")
        print(f"[*] 临时文件已保留在: {temp_dir}")

    client.close()
    end_total_time = tm.time()
    print("\n==================================================")
    print(f"[Success] 所有任务完成！总耗时: {(end_total_time - start_total_time)/60:.2f} 分钟。")
    print("==================================================")

def load_argo_data(year: int, data_dir: str | Path = None,
                   variable_selection: dict | None = None,
                   verbose: bool = False) -> pd.DataFrame:
    """
    加载指定年份的 Argo Parquet 数据文件，并进行列名规范化和变量选择。

    自动将旧版列名（如 'Depth_m'）转换为新版标准名；并通过 variable_selection 灵活指定 Temperature/DO/
    Salinity 三个标准变量分别来源于文件中的哪一列。

    参数:
        - year (int): 需要加载的数据年份（如 2014）。
        - data_dir (str | Path | None): Argo Parquet 所在目录；None 时取配置 paths.argo_parquet。
        - variable_selection (dict | None): 覆盖默认变量来源映射，如 {'Salinity': 'PSAL_WOA'}；默认 {'Temperature': 'Temp_Adjusted', 'DO': 'DOXY_Adjusted', 'Salinity': 'PSAL_Adjusted'}。
        - verbose (bool): 是否输出详细日志，默认 False。

    返回:
        - pd.DataFrame: 处理后的 Argo 数据，列名与数据源均已按参数标准化。
    """
    if data_dir is None:
        data_dir = argo_path
    else:
        data_dir = Path(data_dir)

    # --- 1. 定义并合并变量选择 ---
    # 定义默认选择
    default_selection = {
        'Temperature': 'Temp_Adjusted',
        'Temperature_Flag': 'Temp_Adjusted_Flag',
        'DO': 'DOXY_Adjusted',
        'DO_Flag': 'DOXY_Adjusted_Flag',
        'Salinity': 'PSAL_Adjusted',
        'Salinity_Flag': 'PSAL_Adjusted_Flag'
    }
    # 如果用户提供了自定义选择，则用它来更新（覆盖）默认值
    if variable_selection:
        default_selection.update(variable_selection)
    
    # 最终生效的选择
    final_selection = default_selection
    
    # --- 2. 构建路径并加载文件 ---
    file_path = Path(data_dir) / f'Argo{year}.parquet'
    if verbose:
        print(f"Attempting to load Argo data from: {file_path}")
    if not file_path.exists():
        raise FileNotFoundError(f"Error: The file '{file_path}' was not found.")
    try:
        argo_df = pd.read_parquet(file_path)
    except Exception as e:
        raise RuntimeError(f"Failed to read Argo parquet file: {file_path}") from e

    # --- 3. 列名规范化：将所有已知的旧列名统一为新版中对应的名称 ---
    normalization_map = {
        'Depth_m': 'Depth',
        'Temperature_degC': 'Temp_Adjusted',
        'DO_mol_kg': 'DOXY_Adjusted',
        'Salinity_psu': 'PSAL_Adjusted'
    }
    argo_df.rename(columns=normalization_map, inplace=True)

    # --- 4. 变量选择与最终DataFrame构建 ---
    base_columns = [
        'Year', 'Month', 'Day', 'Longitude', 'Latitude', 'Depth',
        'Profile_number', 'Platform_number'
    ]
    existing_base_columns = [col for col in base_columns if col in argo_df.columns]
    final_df = argo_df[existing_base_columns].copy()

    for standard_name, source_col in final_selection.items():
        if source_col in argo_df.columns:
            final_df[standard_name] = argo_df[source_col]
        else:
            if verbose:
                print(f"Warning: Column '{source_col}' not found in {file_path}. Creating empty column '{standard_name}'.")
            final_df[standard_name] = pd.NA

    if verbose:
        print("Argo data loaded and processed successfully.")
    return final_df

def find_track(
    DS_or_kind: list | str,
    num: int | list | tuple | set | np.ndarray,
    *,
    region: str | None = None,
    output_root: str | Path | None = None,
    include_contours: bool = True,
    return_list: bool = False
) -> pd.DataFrame | list | dict:
    """查找一个或多个涡旋编号的整条轨迹（兼容老/新两种数据源），并统一输出约定。

    DS_or_kind 可为旧的嵌套列表（如从 pickle 读取的 legacy acl/acs/... 列表），或字符串 kind
    （'acs'|'acl'|'cs'|'cl'，从 `META_tracks/<region>` 读取 `<kind>_daily.parquet` 与可选的
    `<kind>_contours.zarr`）。

    参数:
        - DS_or_kind (list | str): 旧结构列表或新结构 kind 字符串。
        - num (int | list | tuple | set | np.ndarray): 单个 track_id 或可迭代的多个 track_id。
        - region (str | None): 区域 slug，None 时使用当前默认区域。
        - output_root (str | Path | None): META_tracks 根目录，None 时读取配置或用默认 './META_tracks'。
        - include_contours (bool): True 时加载等值线（新结构路径），False 时 'contour_*' 为空数组，默认 True。
        - return_list (bool): True 时返回旧版 list/dict 结构而非 DataFrame，默认 False。

    返回:
        - pd.DataFrame | list | dict: 形态随 return_list 与单/多 ID 而定，详见“说明”。

    说明:
        返回契约（return_list=False，默认）:

            - 单 ID：返回 DataFrame。新结构首列为 'track_id'，列为 ['track_id','time','center_lon','center_lat','max_lon','max_lat','contour_lon','contour_lat','radius','speed_contour_lon','speed_contour_lat','date']（'date' 由 'time' 规范化得到，不附加末尾重复的 'track_id'）；旧结构首列为 'index_org'，末尾追加 'track_id'。
            - 多 ID：返回合并后的 DataFrame（含 'track_id' 列），按 ['track_id','date'] 排序。

        返回契约（return_list=True）:

            - 单 ID：返回旧版 list[list]，每个日项首元素为真实 track_id：[track_id, YYYYMMDD, center_lon, center_lat, max_lon, max_lat, eff_contour_lon[], eff_contour_lat[], radius, speed_contour_lon[], speed_contour_lat[]]。
            - 多 ID：返回 {track_id: list[list]} 字典。

        异常:

            - TypeError：DS_or_kind 类型非法。
            - ValueError：指定 track_id 未找到。
    """
    # 规范化 num 输入，判断是否批量
    def _is_scalar_id(x) -> bool:
        return isinstance(x, (int, np.integer))

    if _is_scalar_id(num):
        id_set = {int(num)}
        multi = False
    else:
        try:
            id_set = set(int(x) for x in list(num))
        except Exception:
            id_set = {int(num)}
            multi = False
        else:
            multi = len(id_set) > 1
    # 1) 旧数据结构兼容：DS_or_kind 为 list 时沿用旧逻辑
    if isinstance(DS_or_kind, list):
        DS = DS_or_kind
        id_to_track = {}
        for track in DS:
            if not track:
                continue
            try:
                tid = int(track[0][0])
            except Exception:
                continue
            if tid in id_set:
                id_to_track[tid] = track

        missing = sorted(list(id_set - set(id_to_track.keys())))
        if missing:
            raise ValueError(f"Track not found for id(s): {missing}")

        ACS, ACL, CS, CL = load_meta_data()
        meta_map = {'ACS': ACS, 'ACL': ACL, 'CS': CS, 'CL': CL}

        if not multi:
            tid = next(iter(id_set))
            track = id_to_track[tid]
            if return_list:
                return track
            df = pd.DataFrame(track, columns=[
                'index_org','time','center_lon','center_lat','max_lon','max_lat',
                'contour_lon','contour_lat','radius','speed_contour_lon','speed_contour_lat'
            ])
            try:
                df['date'] = convert_date(df['time'])
            except Exception:
                df['date'] = pd.NaT

            matched_ds_name = 'UNKNOWN'
            try:
                caller_locals = inspect.currentframe().f_back.f_locals
                for var_name, var_val in caller_locals.items():
                    if var_val is DS_or_kind:
                        matched_ds_name = var_name.upper()
                        break
            except Exception:
                matched_ds_name = 'UNKNOWN'

            matched_meta = meta_map.get(matched_ds_name)

            df['track_id'] = matched_meta['track'][df['index_org']] if matched_meta is not None else int(tid)
            return df
        else:
            if return_list:
                return {tid: id_to_track[tid] for tid in sorted(id_set)}
            frames = []
            for tid, track in id_to_track.items():
                df = pd.DataFrame(track, columns=[
                    'index_org','time','center_lon','center_lat','max_lon','max_lat',
                    'contour_lon','contour_lat','radius','speed_contour_lon','speed_contour_lat'
                ])
                try:
                    df['date'] = convert_date(df['time'])
                except Exception:
                    df['date'] = pd.NaT

                matched_ds_name = 'UNKNOWN'
                try:
                    caller_locals = inspect.currentframe().f_back.f_locals
                    for var_name, var_val in caller_locals.items():
                        if var_val is DS_or_kind:
                            matched_ds_name = var_name.upper()
                            break
                except Exception:
                    matched_ds_name = 'UNKNOWN'

                matched_meta = meta_map.get(matched_ds_name)

                df['track_id'] = matched_meta['track'][df['index_org']] if matched_meta is not None else int(tid)

                frames.append(df)
            out = pd.concat(frames, ignore_index=True)
            out.sort_values(['track_id', 'date'], inplace=True)
            out.reset_index(drop=True, inplace=True)
            return out

    # 2) 新数据结构：按 kind 从 Parquet + Zarr 中提取
    if not isinstance(DS_or_kind, str):
        raise TypeError("find_track 新用法：请传入 kind 字符串（'acs'|'acl'|'cs'|'cl'）作为第一个参数，或传旧的 DS 列表。")
    kind = DS_or_kind.lower()
    if kind not in {'acs', 'acl', 'cs', 'cl'}:
        raise ValueError(f"未知 kind='{DS_or_kind}', 期望 'acs'|'acl'|'cs'|'cl'.")

    region_slug = region or _current_region_key()
    root = _ensure_meta_tracks_root(output_root)
    region_dir = Path(root) / region_slug
    if not region_dir.exists():
        raise FileNotFoundError(f"区域目录不存在：{region_dir}")

    # 2.1 定位 daily 源：优先单文件，其次目录（_tmp 或 .parquet 目录）
    daily_file = region_dir / f"{kind}_daily.parquet"
    daily_tmp_dir = region_dir / f"{kind}_daily_tmp"
    daily_dir = region_dir / f"{kind}_daily.parquet"  # 目录形式
    daily_source = None
    source_type = None  # 'file' | 'dir'
    if daily_file.exists() and daily_file.is_file():
        daily_source = daily_file
        source_type = 'file'
    elif daily_tmp_dir.exists() and daily_tmp_dir.is_dir():
        daily_source = daily_tmp_dir
        source_type = 'dir'
    elif daily_dir.exists() and daily_dir.is_dir():
        daily_source = daily_dir
        source_type = 'dir'
    else:
        raise FileNotFoundError(f"未找到 daily 数据：{daily_file} 或 {daily_tmp_dir} / {daily_dir}")

    # 2.2 读取元信息以定位 Zarr（若需要）
    zarr_path = region_dir / f"{kind}_contours.zarr"
    meta_path = region_dir / f"{kind}_metadata.json"
    if include_contours and meta_path.exists():
        try:
            with open(meta_path, 'r') as f:
                meta = json.load(f)
            if meta.get('contours_zarr'):
                zarr_path = Path(meta['contours_zarr'])
        except Exception:
            pass

    # 2.3 流式扫描 Parquet（按行组/分片顺序）收集该 track 的全部行及其“全局行号”
    def _scan_daily_rows_file(parquet_file: Path, target_ids: set[int]):
        pf = pq.ParquetFile(parquet_file)
        # 找到 track_id 列的索引（用于读取统计信息）
        try:
            track_idx = [i for i, n in enumerate(pf.schema.names) if n == 'track_id'][0]
        except Exception:
            track_idx = None
        cum = 0
        min_tid = min(target_ids)
        max_tid = max(target_ids)
        for rg in range(pf.num_row_groups):
            # 先用统计信息快速判断是否可能包含 num
            if track_idx is not None and pf.metadata is not None:
                try:
                    col_meta = pf.metadata.row_group(rg).column(track_idx)
                    stats = col_meta.statistics
                    if stats is not None and stats.has_min_max:
                        min_v, max_v = stats.min, stats.max
                        # 某些版本可能返回 bytes，需要安全转换
                        try:
                            min_v = int(min_v); max_v = int(max_v)
                        except Exception:
                            pass
                        if isinstance(min_v, (int, np.integer)) and isinstance(max_v, (int, np.integer)):
                            if (max_tid < min_v) or (min_tid > max_v):
                                # 不可能命中，跳过读取该行组
                                cum += pf.metadata.row_group(rg).num_rows
                                continue
                except Exception:
                    pass

            tbl = pf.read_row_group(rg, columns=[
                'track_id', 'time', 'center_lon', 'center_lat', 'max_lon', 'max_lat', 'radius', 'orig_index'
            ])
            t_id = tbl.column('track_id').to_numpy()
            mask = np.isin(t_id, list(target_ids))
            idxs = np.nonzero(mask)[0]
            if idxs.size:
                df = tbl.to_pandas()
                for j in idxs.tolist():
                    yield cum + j, df.iloc[j]
            cum += tbl.num_rows

    def _scan_daily_rows_dir(parquet_dir: Path, target_ids: set[int]):
        parts = sorted([p for p in parquet_dir.glob('*.parquet') if p.is_file()])
        cum = 0
        min_tid = min(target_ids)
        max_tid = max(target_ids)
        for p in parts:
            pf = pq.ParquetFile(p)
            # 找到 track_id 列的索引
            try:
                track_idx = [i for i, n in enumerate(pf.schema.names) if n == 'track_id'][0]
            except Exception:
                track_idx = None
            for rg in range(pf.num_row_groups):
                # 统计信息快速跳过
                if track_idx is not None and pf.metadata is not None:
                    try:
                        col_meta = pf.metadata.row_group(rg).column(track_idx)
                        stats = col_meta.statistics
                        if stats is not None and stats.has_min_max:
                            min_v, max_v = stats.min, stats.max
                            try:
                                min_v = int(min_v); max_v = int(max_v)
                            except Exception:
                                pass
                            if isinstance(min_v, (int, np.integer)) and isinstance(max_v, (int, np.integer)):
                                if (max_tid < min_v) or (min_tid > max_v):
                                    cum += pf.metadata.row_group(rg).num_rows
                                    continue
                    except Exception:
                        pass

                tbl = pf.read_row_group(rg, columns=[
                    'track_id', 'time', 'center_lon', 'center_lat', 'max_lon', 'max_lat', 'radius', 'orig_index'
                ])
                t_id = tbl.column('track_id').to_numpy()
                mask = np.isin(t_id, list(target_ids))
                idxs = np.nonzero(mask)[0]
                if idxs.size:
                    df = tbl.to_pandas()
                    for j in idxs.tolist():
                        yield cum + j, df.iloc[j]
                cum += tbl.num_rows

    scanner = _scan_daily_rows_file if source_type == 'file' else _scan_daily_rows_dir
    rows = list(scanner(daily_source, id_set))
    if not rows:
        raise ValueError(f"Track id(s) {sorted(list(id_set))} 未在 {daily_source} 中找到")

    # 2.4 可选：打开 Zarr 并准备索引数组
    eff_lon_arr = eff_lat_arr = spd_lon_arr = spd_lat_arr = None
    eff_idx = spd_idx = None
    if include_contours:
        if not zarr_path.exists():
            print(f"[find_track] 警告：未找到 Zarr 存储 {zarr_path}，将仅返回无轮廓的轨迹基础信息。")
            include_contours = False
        else:
            zroot = zarr.open_group(zarr_path, mode='r')
            eff_grp = zroot['effective']
            spd_grp = zroot['speed']
            # 注意：保留 Zarr 数组对象本身，按需切片，避免整数组载入内存
            eff_lon_arr = eff_grp['lon']; eff_lat_arr = eff_grp['lat']
            spd_lon_arr = spd_grp['lon']; spd_lat_arr = spd_grp['lat']
            eff_idx = eff_grp['index'];  spd_idx = spd_grp['index']

    # 2.5 组装为结果（按时间升序）
    # rows: list of (global_row_idx, pandas_series)
    def _ymd_i8(ts: pd.Timestamp) -> int:
        ts = pd.Timestamp(ts)
        return ts.year * 10000 + ts.month * 100 + ts.day

    # 排序（多数情况下已是有序，这里稳妥起见按 time 排）
    rows_sorted = sorted(rows, key=lambda kv: pd.Timestamp(kv[1]['time']))

    # 将结果分配到各个 track_id 下
    rows_per_id: dict[int, list] = {}
    for global_idx, r in rows_sorted:
        tid = int(r['track_id']) if 'track_id' in r else None
        if tid is None:
            continue
        # 基础字段
        t_ymd = _ymd_i8(r['time'])
        center_lon = float(r['center_lon'])
        center_lat = float(r['center_lat'])
        max_lon = float(r['max_lon']) if 'max_lon' in r else float('nan')
        max_lat = float(r['max_lat']) if 'max_lat' in r else float('nan')
        radius = float(r['radius']) if 'radius' in r else float('nan')

        # 轮廓
        eff_lon = np.array([], dtype='f8')
        eff_lat = np.array([], dtype='f8')
        spd_lon = np.array([], dtype='f8')
        spd_lat = np.array([], dtype='f8')
        if include_contours and eff_idx is not None:
            if 0 <= global_idx < (eff_idx.shape[0] - 1):
                s_e = eff_idx[global_idx:global_idx+2]
                s = int(s_e[0]); e = int(s_e[1])
                if e > s:
                    eff_lon = np.asarray(eff_lon_arr[s:e])
                    eff_lat = np.asarray(eff_lat_arr[s:e])
            if 0 <= global_idx < (spd_idx.shape[0] - 1):
                s_e = spd_idx[global_idx:global_idx+2]
                s = int(s_e[0]); e = int(s_e[1])
                if e > s:
                    spd_lon = np.asarray(spd_lon_arr[s:e])
                    spd_lat = np.asarray(spd_lat_arr[s:e])

        # 注意：为简化后续标注/检索，首字段使用 track_id（tid）而非原始行索引
        rows_per_id.setdefault(tid, []).append([
            int(tid), t_ymd, center_lon, center_lat, max_lon, max_lat,
            eff_lon, eff_lat, radius, spd_lon, spd_lat
        ])

    # 检查是否全部命中
    hit_ids = set(rows_per_id.keys())
    missing = sorted(list(id_set - hit_ids))
    if missing:
        raise ValueError(f"Track not found for id(s): {missing}")

    if not multi:
        tid = next(iter(id_set))
        rows_list = rows_per_id[tid]
        if return_list:
            return rows_list
        df = pd.DataFrame(rows_list, columns=[
            'track_id','time','center_lon','center_lat','max_lon','max_lat',
            'contour_lon','contour_lat','radius','speed_contour_lon','speed_contour_lat'
        ])
        try:
            df['date'] = convert_date(df['time'])
        except Exception:
            df['date'] = pd.NaT
        return df
    else:
        if return_list:
            return {tid: rows_per_id[tid] for tid in sorted(rows_per_id.keys())}
        frames = []
        for tid, rows_list in rows_per_id.items():
            df = pd.DataFrame(rows_list, columns=[
                'track_id','time','center_lon','center_lat','max_lon','max_lat',
                'contour_lon','contour_lat','radius','speed_contour_lon','speed_contour_lat'
            ])
            try:
                df['date'] = convert_date(df['time'])
            except Exception:
                df['date'] = pd.NaT
            frames.append(df)
        out = pd.concat(frames, ignore_index=True)
        out.sort_values(['track_id', 'date'], inplace=True)
        out.reset_index(drop=True, inplace=True)
        return out

def is_point_in_contour(row: pd.Series) -> bool:
    """
    检查一个数据行中的点坐标是否在其涡旋轮廓坐标内部。

    辅助函数，通常与 DataFrame 的 .apply() 配合使用，从一行数据中提取 'Longitude'/'Latitude'/
    'contour_lon'/'contour_lat' 并判断点是否在多边形内部。

    参数:
        - row (pd.Series): 一个 Pandas DataFrame 的数据行，必须包含经纬度和轮廓坐标。

    返回:
        - bool: 点在轮廓内部返回 True，否则 False。
    """
    try:
        contour_lon = np.asarray(row['contour_lon'], dtype=float)
        contour_lat = np.asarray(row['contour_lat'], dtype=float)
        if contour_lon.size < 3 or contour_lat.size < 3:
            return False
        if contour_lon.shape != contour_lat.shape:
            return False

        center_lon = row.get('center_lon', np.nan)
        if pd.isna(center_lon):
            if contour_lon.size:
                center_lon = float(contour_lon[0])
            else:
                center_lon = float(row['Longitude'])

        contour_lon_norm = center_lon + _minimal_lon_diff_deg(contour_lon, center_lon)
        point_lon_norm = center_lon + _minimal_lon_diff_deg(row['Longitude'], center_lon)

        contour_coords = list(zip(contour_lon_norm, contour_lat))
        if len(contour_coords) < 3:
            return False

        path = MplPath(contour_coords)
        return path.contains_point((point_lon_norm, row['Latitude']))
    except Exception:
        return False

def filtered_float_data(
    DS: list | str,
    no: int,
    argo_data_dir: str | Path = None,
    circle_enlargement_factor: float | None = None,
    use_adaptive_circle: bool = True,
    adaptive_lat_threshold: float = 70.0,
    adaptive_distance_threshold_km: float = 300.0,
    force_great_circle_circle: bool = False,
    track: list | pd.DataFrame | None = None
) -> pd.DataFrame:
    """
    根据涡旋轨迹，动态加载并筛选匹配的 Argo 浮标剖面数据。

    分析涡旋轨迹覆盖的年份并自动加载对应年份的 Argo 数据，再以向量化方式高效匹配同一天出现在涡旋区域
    内的浮标；筛选标准为浮标位置处于涡旋有效轮廓内，或处于扩大一定倍数后的有效半径内。

    参数:
        - DS (list | str): 旧结构传 legacy 轨迹列表（如从 pickle 读取的 acl/acs/... 数据）；新结构传字符串 kind（'acs'|'acl'|'cs'|'cl'），内部调用 find_track 从 META_tracks 区域目录动态装载。
        - no (int): 需要筛选的涡旋唯一编号。
        - argo_data_dir (str | Path | None): 存放 Argo Parquet 文件的目录；None 时取配置 paths.argo_parquet。
        - circle_enlargement_factor (float | None): 有效半径放大系数；None 时回退配置值。
        - use_adaptive_circle (bool): 为 True 时半径匹配距离用 adaptive_distance_m（高纬或大距离自动切换大圆），否则用局地平面近似，默认 True。
        - adaptive_lat_threshold (float): |lat| 高于该阈值触发自适应大圆距离计算，默认 70.0。
        - adaptive_distance_threshold_km (float): 平面近似距离超过该阈值（km）触发大圆距离计算，默认 300.0。
        - force_great_circle_circle (bool): 强制半径距离全部使用大圆（忽略阈值条件），默认 False。
        - track (list | pd.DataFrame | None): 预先加载好的涡旋轨迹；传入可避免函数内部再次调用 find_track 触发重复磁盘 I/O，默认 None。

    返回:
        - pd.DataFrame: 匹配的 Argo 剖面完整数据（所有深度层级）；无匹配时返回空 DataFrame。
    """
    # 参数回退
    if argo_data_dir is None:
        argo_data_dir = argo_path
    if circle_enlargement_factor is None:
        circle_enlargement_factor = globals().get('circle_enlargement_factor', 1.2)

    # --- 1. 准备涡旋数据 ---
    # print(f"[*] Preparing track data for eddy ID {no}...")
    if track is None:
        wanted_track = find_track(DS, no)
        # 兼容 find_track 返回 DataFrame 的新默认行为
        if wanted_track is None or (isinstance(wanted_track, pd.DataFrame) and wanted_track.empty):
            print(f"  - Track for eddy {no} not found, returning empty result.")
            return pd.DataFrame()
        if isinstance(wanted_track, pd.DataFrame):
            track_df = wanted_track.copy()
        else:
            # 将涡旋轨迹列表转换为DataFrame，方便后续处理（旧结构）
            track_df = pd.DataFrame(
                wanted_track,
                columns=['track_id', 'time', 'center_lon', 'center_lat', 'max_lon', 'max_lat', 
                         'contour_lon', 'contour_lat', 'radius', 'speed_contour_lon', 'speed_contour_lat']
            )
    elif isinstance(track, pd.DataFrame):
        track_df = track.copy()
        # 尝试确保必要列存在
        needed_cols = {'track_id','time','center_lon','center_lat','max_lon','max_lat','contour_lon','contour_lat','radius','speed_contour_lon','speed_contour_lat'}
        missing = needed_cols - set(track_df.columns)
        if missing:
            raise ValueError(f"track DataFrame is missing required columns: {sorted(list(missing))}")
    else:
        # 视为 list[list | tuple]
        track_df = pd.DataFrame(
            track,
            columns=['track_id', 'time', 'center_lon', 'center_lat', 'max_lon', 'max_lat', 
                     'contour_lon', 'contour_lat', 'radius', 'speed_contour_lon', 'speed_contour_lat']
        )
    # 使用convert_date函数转换日期（如果没有或类型不对）
    if 'date' not in track_df.columns:
        track_df['date'] = convert_date(track_df['time'])

    # --- 2. 动态加载所需的Argo数据 ---
    min_year, max_year = track_df['date'].min().year, track_df['date'].max().year
    print(f"[*] Eddy track spans years: {min_year} - {max_year}. Loading corresponding Argo data...")
    
    all_argo_data = []
    for year in range(min_year, max_year + 1):
        try:
            yearly_argo_data = load_argo_data(year, data_dir=argo_data_dir)
            if not yearly_argo_data.empty:
                all_argo_data.append(yearly_argo_data)
        except FileNotFoundError:
            print(f"  - Warning: Argo data file for year {year} not found, skipping.")

    if not all_argo_data:
        print(f"  - No Argo data was loaded, returning empty result.")
        return pd.DataFrame()
        
    argo_data = pd.concat(all_argo_data, ignore_index=True)
    # print(f"[*] All relevant yearly Argo data loaded, total records: {len(argo_data):,}.")

    # --- 3. 准备Argo数据用于合并 ---
    # 预处理：删除没有有效坐标的记录，并创建日期列
    argo_data.dropna(subset=['Longitude', 'Latitude'], inplace=True)
    argo_data['date'] = pd.to_datetime(argo_data[['Year', 'Month', 'Day']])
    
    # 每个剖面只需要一个位置点来做匹配，因此我们按Profile_number去重，只保留第一个出现的点
    argo_positions = argo_data.drop_duplicates(subset='Profile_number', keep='first').copy()
    
    # --- 4. 核心步骤：将Argo与涡旋数据按日期合并 ---
    # pd.merge会高效地找出在同一天既有涡旋轨迹、又有Argo浮标的记录
    merged_df = pd.merge(argo_positions, track_df, on='date')
    if merged_df.empty:
        print(f"  - No Argo floats found on the same day as the eddy track.")
        return pd.DataFrame()
    # print(f"[*] Found {len(merged_df)} potential 'float-eddy' pairs based on date matching.")

    # --- 5. 向量化地理筛选 ---
    # print(f"[*] Performing geographic location matching...")
    # 5.1 检查是否在多边形内部
    inside_poly_mask = merged_df.apply(is_point_in_contour, axis=1)

    # 5.2 圆内判定
    if use_adaptive_circle:
        dist_m = adaptive_distance_m(
            merged_df['Longitude'].values,
            merged_df['Latitude'].values,
            merged_df['center_lon'].values,
            merged_df['center_lat'].values,
            wrap_dateline=True,
            gc_lat_threshold=adaptive_lat_threshold,
            gc_distance_threshold_km=adaptive_distance_threshold_km,
            force_great_circle=force_great_circle_circle
        )
    else:
        scale_all = approximate_degree_length(merged_df['center_lat'].values)
        scale_lat = scale_all['meters_per_degree_lat']
        scale_lon = scale_all['meters_per_degree_lon']
        dlon_deg = _minimal_lon_diff_deg(
            merged_df['Longitude'].values,
            merged_df['center_lon'].values
        )
        dx_m = dlon_deg * scale_lon
        dy_m = (merged_df['Latitude'].values - merged_df['center_lat'].values) * scale_lat
        dist_m = np.hypot(dx_m, dy_m)
    inside_circle_mask = dist_m <= (merged_df['radius'].values * circle_enlargement_factor)
    
    # 5.3 合并两种筛选条件
    final_mask = inside_poly_mask | inside_circle_mask
    
    # 筛选出真正匹配的记录
    matched_profiles = merged_df[final_mask]
    
    if matched_profiles.empty:
        print(f"  - After geographic filtering, no matching Argo floats were found.")
        return pd.DataFrame()

    # --- 6. 根据筛选结果获取完整的剖面数据 ---
    # 从原始的、包含所有深度数据的argo_data中，筛选出所有匹配上的Profile_number
    matching_profile_numbers = matched_profiles['Profile_number'].unique()
    final_argo_data = argo_data[argo_data['Profile_number'].isin(matching_profile_numbers)].copy()
    
    # 在返回结果前，删除临时的'date'辅助列
    if 'date' in final_argo_data.columns:
        final_argo_data = final_argo_data.drop(columns=['date'])
    
    print(f"[*] Filtering complete! Found {len(matching_profile_numbers)} matching Argo profiles, with a total of {len(final_argo_data):,} data records.")
    
    return final_argo_data

def get_argo_profile_numbers_on_day(
    DS: list | str | tuple | dict,
    track_id: int,
    date: str | int | float | pd.Timestamp,
    *,
    sort_result: bool = True,
) -> list[int]:
    """返回指定涡旋在指定日期命中的全部 Argo Profile_number（唯一值）。

    复用 filtered_float_data 的匹配口径（同日 + 轮廓内或扩圈半径内），但只返回唯一 Profile 编号列表，
    不返回完整剖面数据。

    参数:
        - DS (list | str | tuple | dict): legacy 列表、kind 字符串、kind 序列或数据集字典。
        - track_id (int): 涡旋编号。
        - date (str | int | float | pd.Timestamp): 目标日期；支持 'YYYY-MM-DD'、YYYYMMDD、days-since-1950、Timestamp。
        - sort_result (bool): 是否按升序返回，默认 True。

    返回:
        - list[int]: 该日命中的唯一 Profile_number 列表；无命中返回空列表。
    """
    try:
        track_df, _ds_name, ds_source_for_filter = _resolve_track_context(DS, int(track_id), include_contours=True)
    except Exception:
        return []

    try:
        if isinstance(date, (str, bytes, np.str_)):
            parsed = pd.to_datetime(date, errors='coerce')
            if pd.isna(parsed):
                target_date = convert_date(date)
                if isinstance(target_date, pd.Series):
                    if target_date.empty or target_date.isna().all():
                        return []
                    target_ts = pd.Timestamp(target_date.iloc[0]).normalize()
                else:
                    if pd.isna(target_date):
                        return []
                    target_ts = pd.Timestamp(target_date).normalize()
            else:
                target_ts = pd.Timestamp(parsed).normalize()
        else:
            target_date = convert_date(date)
            if isinstance(target_date, pd.Series):
                if target_date.empty or target_date.isna().all():
                    return []
                target_ts = pd.Timestamp(target_date.iloc[0]).normalize()
            else:
                if pd.isna(target_date):
                    return []
                target_ts = pd.Timestamp(target_date).normalize()
    except Exception:
        return []

    argo_rows = filtered_float_data(ds_source_for_filter, int(track_id), track=track_df)
    if argo_rows.empty or 'Profile_number' not in argo_rows.columns:
        return []

    date_col = pd.to_datetime(argo_rows[['Year', 'Month', 'Day']], errors='coerce').dt.normalize()
    day_rows = argo_rows.loc[date_col == target_ts]
    if day_rows.empty:
        return []

    pnums = pd.to_numeric(day_rows['Profile_number'], errors='coerce').dropna().astype(int).unique().tolist()
    if sort_result:
        pnums.sort()
    return pnums

def _resolve_track_context(
    DS_input: list | str | tuple | dict,
    track_id: int,
    *,
    include_contours: bool = True,
) -> tuple[pd.DataFrame, str, list | str | tuple | dict]:
    """将多种轨迹数据输入统一解析为轨迹 DataFrame 与元信息。"""

    def _infer_dataset_name_from_caller(obj) -> str:
        frame = inspect.currentframe()
        plot_frame = None
        caller_frame = None
        try:
            plot_frame = frame.f_back
            caller_frame = plot_frame.f_back if plot_frame else None
            search_frame = caller_frame or plot_frame
            while search_frame:
                for var_name, var_val in search_frame.f_locals.items():
                    if var_val is obj:
                        return var_name.upper()
                search_frame = search_frame.f_back if search_frame is caller_frame else None
        except Exception:
            pass
        finally:
            del frame
            if plot_frame:
                del plot_frame
            if caller_frame:
                del caller_frame
        return "UNKNOWN"

    def _is_kind_sequence(value) -> bool:
        return isinstance(value, (list, tuple)) and len(value) > 0 and all(isinstance(item, str) for item in value)

    def _build_track_df(track_payload):
        required_cols = [
            'track_id', 'time', 'center_lon', 'center_lat', 'max_lon', 'max_lat',
            'contour_lon', 'contour_lat', 'radius', 'speed_contour_lon', 'speed_contour_lat'
        ]
        if isinstance(track_payload, pd.DataFrame):
            missing_cols = [col for col in required_cols if col not in track_payload.columns]
            if missing_cols:
                raise ValueError(f"Track data missing required columns: {missing_cols}")
            df = track_payload[required_cols].copy()
        else:
            df = pd.DataFrame(track_payload, columns=required_cols)
        if 'date' not in df.columns:
            df['date'] = convert_date(df['time'])
        return df

    ds_name = "UNKNOWN"
    ds_source_for_filter: list | str | tuple | dict = DS_input
    wanted_track = None
    candidate_kind_names: list[str] = []

    if isinstance(DS_input, str):
        candidate_kind_names = [DS_input]
        ds_name = DS_input.upper()
    elif _is_kind_sequence(DS_input):
        candidate_kind_names = [str(k) for k in DS_input]
    elif isinstance(DS_input, dict):
        for key, value in DS_input.items():
            if isinstance(value, list) and not _is_kind_sequence(value):
                try:
                    wanted_track = find_track(value, track_id)
                except ValueError:
                    continue
                ds_name = str(key).upper()
                ds_source_for_filter = value
                break
        if wanted_track is None:
            candidate_kind_names = [str(k) for k in DS_input.keys()]
    elif isinstance(DS_input, list):
        wanted_track = find_track(DS_input, track_id)
        ds_name = _infer_dataset_name_from_caller(DS_input)
    elif isinstance(DS_input, tuple):
        ds_list = list(DS_input)
        wanted_track = find_track(ds_list, track_id)
        ds_source_for_filter = ds_list
        ds_name = _infer_dataset_name_from_caller(ds_list)
    else:
        raise TypeError("Unsupported DS type. Expected legacy list, kind string, tuple, or dataset dict.")

    if wanted_track is None and candidate_kind_names:
        last_error: Exception | None = None
        for kind in candidate_kind_names:
            kind_str = str(kind).strip()
            if not kind_str:
                continue
            try:
                wanted_track = find_track(kind_str.lower(), track_id, include_contours=include_contours)
                ds_source_for_filter = kind_str.lower()
                ds_name = kind_str.upper()
                break
            except Exception as exc:
                last_error = exc
        if wanted_track is None:
            if last_error is not None:
                raise ValueError(f"Track for eddy {track_id} not found in datasets {candidate_kind_names}: {last_error}")
            raise ValueError(f"Track for eddy {track_id} not found in datasets {candidate_kind_names}.")

    if wanted_track is None:
        raise ValueError(f"Track for eddy {track_id} not found.")

    try:
        track_df = _build_track_df(wanted_track)
    except Exception as exc:
        raise ValueError(f"Failed to normalize track data for eddy {track_id}: {exc}") from exc

    if track_df.empty:
        raise ValueError(f"Track for eddy {track_id} is empty.")

    return track_df, ds_name, ds_source_for_filter

def plot_track(
    DS: list | str | tuple | dict, 
    no: int,
    save_fig: bool = False,
    show_fig: bool = True,
    plot_radius: bool = False,
    connection_threshold_days: int = 5,
    detection_config: DetectionConfig | None = None,
    plot_unrelated_argo: bool = True,
    fix_colorbar: bool = True,
    cbar_min: float | None = None,
    cbar_max: float | None = None,
    cbar_ticks: list | None = None,
    min_anomaly_count: int = 0
):
    """
    绘制指定编号涡旋的详细轨迹，并高亮显示与 Argo 异常剖面的交互情况。

    自动加载并筛选与该涡旋匹配的 Argo 数据，将涡旋完整轨迹绘为虚线；对存在 Argo 浮标的时期智能高亮
    （相邻存在日间隔小于 connection_threshold_days 时连成实线，孤立单日标为点），用 calculate_delta_do
    按 detection_config 识别异常（每个剖面保留 anomaly_score 最强一条）并以当前方法主变量着色，可选叠加
    涡旋在交互日的有效半径与轮廓。

    参数:
        - DS (list | str | tuple | dict): legacy 模式传入已加载的数据列表（如 ACL/ACS）；新模式可传字符串 kind（'acs' 等）或字符串列表/元组，自动从本地 META_tracks 检索对应轨迹；亦支持 {"ACS": acs, ...} 字典以兼容旧流程。
        - no (int): 需要绘制的涡旋唯一编号。
        - save_fig (bool): 是否将图像保存到文件，默认 False。
        - show_fig (bool): 是否在交互式环境中显示图像，默认 True。
        - plot_radius (bool): 是否以圆形式绘制涡旋在交互日的有效半径，默认 False。
        - connection_threshold_days (int): 连接 Argo 交互点的最大天数阈值，默认 5。
        - detection_config (DetectionConfig | None): 异常识别配置；None 时使用 processing.yml 默认值。
        - plot_unrelated_argo (bool): 是否绘制未被判定为异常的匹配 Argo 剖面基准位置（空心灰圈），默认 True。
        - fix_colorbar (bool): 是否固定异常主变量色标，默认 True。
        - cbar_min (float | None): 色标下限；None 时自动。
        - cbar_max (float | None): 色标上限；None 时自动。
        - cbar_ticks (list | None): 色标刻度；None 时自动。
        - min_anomaly_count (int): >0 时要求异常数量 ≥ 该值才绘图，=0 不做数量阈值过滤，默认 0。
    输出:
        - 图像（save_fig=True 时）：`plot_outputs/<method>/<region>/plot_track/Track_Analysis_{数据集}{编号}_{stem}.png`
    """
    # --- 1. 准备涡旋和Argo数据 ---
    print(f"[*] Preparing data for eddy ID {no}...")

    cfg = _resolve_detection_config(
        detection_config,
        cbar_min=cbar_min,
        cbar_max=cbar_max,
        cbar_ticks=cbar_ticks,
    )
    
    try:
        track_df, ds_name, ds_source_for_filter = _resolve_track_context(DS, no, include_contours=True)
    except Exception as exc:
        print(f"  - Error: {exc}")
        return
    num = track_df['track_id'].iloc[0]

    # 调用筛选函数，获取所有匹配的 Argo 数据（包含所有深度）
    argo_data_filtered = filtered_float_data(ds_source_for_filter, no, track=track_df)

    # 预筛选：若匹配到的剖面数不足 min_anomaly_count，直接跳过
    if argo_data_filtered.empty:
        print(f"  - Skip plotting {ds_name}{no}: no matched Argo profiles.")
        return
    matched_profile_count = argo_data_filtered['Profile_number'].nunique()
    if min_anomaly_count > 0 and matched_profile_count < min_anomaly_count:
        print(f"\033[31m  - Skip plotting {ds_name}{no}: matched profiles ({matched_profile_count}) < min_anomaly_count ({min_anomaly_count}).\033[0m")
        return

    # 使用 ΔDO 异常检测
    anomalies = pd.DataFrame()
    base_argo_positions = pd.DataFrame()
    if not argo_data_filtered.empty:
        # 基准剖面位置（去除深度重复）
        base_argo_positions = (
            argo_data_filtered.sort_values(['Profile_number','Depth'])
            .groupby('Profile_number', as_index=False)
            .first()[['Profile_number','Longitude','Latitude','Year','Month','Day']]
        )
        anomalies = calculate_delta_do(
            argo_data_filtered,
            detection_config=cfg,
            remove_outliers=True,
            verbose=False
        )
        if not anomalies.empty:
            anomalies = _keep_best_anomaly_per_profile(anomalies, cfg)

    # 若异常数量不足阈值，直接跳过
    anomaly_count = 0 if anomalies.empty else len(anomalies)
    if min_anomaly_count > 0 and anomaly_count < min_anomaly_count:
        print(f"\033[31m  - Skip plotting {ds_name}{no}: anomalies ({anomaly_count}) < min_anomaly_count ({min_anomaly_count}).\033[0m")
        return

    # --- 2. 准备绘图 ---
    prop_colors = plt.rcParams['axes.prop_cycle'].by_key()['color']
    eddy_color = prop_colors[1] if 'AC' in ds_name else prop_colors[0]
    
    # 加载地理底图
    world = _load_world_geodataframe()

    # 初始化画布与坐标轴
    fig, ax = plt.subplots(figsize=(30, 20))
    ax.set_title(f'Track Analysis for Eddy {ds_name}{num}', fontsize=20)
    ax.set_xlabel('Longitude', fontsize=20); ax.set_ylabel('Latitude', fontsize=20)
    world.plot(color='lightgrey', edgecolor='white', ax=ax)

    # --- 3. 绘制涡旋轨迹 (背景和高亮) ---
    # 绘制完整的背景轨迹 (虚线)
    ax.plot(track_df['center_lon'], track_df['center_lat'], color=eddy_color, linestyle='--', alpha=0.5, label='Full Track Path')
    
    # 找出有Argo交互的日期
    interaction_track_df = pd.DataFrame(columns=track_df.columns)
    if not anomalies.empty:
        interaction_dates = pd.to_datetime(anomalies[['Year', 'Month', 'Day']]).unique()
        interaction_track_df = track_df[track_df['date'].isin(interaction_dates)].copy()
        
        # 识别并分别绘制不连续的实线段
        if not interaction_track_df.empty:
            date_diffs = interaction_track_df['date'].diff()
            break_points = date_diffs > pd.Timedelta(f'{connection_threshold_days} days')
            segment_ids = break_points.cumsum()
            
            labeled_interaction_path = False
            labeled_interaction_point = False
            # 按段落ID分组并分别绘图
            for _, segment_df in interaction_track_df.groupby(segment_ids):
                start_segment_date = segment_df['date'].min()
                end_segment_date = segment_df['date'].max()
                full_path_segment = track_df[
                    (track_df['date'] >= start_segment_date) & 
                    (track_df['date'] <= end_segment_date)
                ]

                # 如果段内只有一个交互点，画成散点
                if len(segment_df) == 1:
                    label = 'Interaction Point' if not labeled_interaction_point else None
                    ax.scatter(full_path_segment['center_lon'], full_path_segment['center_lat'], 
                               facecolors='none', edgecolors=eddy_color, 
                                linewidths=1.5, s=100, zorder=5, label=label)
                    labeled_interaction_point = True
                # 否则，将这段时间的完整轨迹画成实线
                else:
                    label = 'Interaction Path' if not labeled_interaction_path else None
                    ax.plot(full_path_segment['center_lon'], full_path_segment['center_lat'], 
                            color=eddy_color, linestyle='-', linewidth=2.5, label=label)
                    labeled_interaction_path = True
    
    # 标记轨迹的起点和终点
    ax.plot(track_df['center_lon'].iloc[0], track_df['center_lat'].iloc[0], marker='o', color=eddy_color, markersize=10, label='Start')
    ax.plot(track_df['center_lon'].iloc[-1], track_df['center_lat'].iloc[-1], marker='x', color=eddy_color, markersize=12, mew=2.5, label='End')

    # --- 4. 绘制交互日的轮廓、半径和日期 ---
    labeled_contour = False
    labeled_radius = False
    for _, eddy_day in interaction_track_df.iterrows():
        # 绘制轮廓线
        if not labeled_contour:
            ax.plot(eddy_day['contour_lon'], eddy_day['contour_lat'], color=eddy_color, linewidth=1, alpha=0.6, label='Effective Contour')
            labeled_contour = True
        else:
            ax.plot(eddy_day['contour_lon'], eddy_day['contour_lat'], color=eddy_color, linewidth=1, alpha=0.6)
        
        # 绘制有效半径
        if plot_radius:
            radius_color = 'r' if 'AC' in ds_name else 'purple'
            circle_label = 'Effective Radius' if not labeled_radius else None
            scale = approximate_degree_length(eddy_day['center_lat'])
            deg_height = (eddy_day['radius'] * circle_enlargement_factor) / scale['meters_per_degree_lat']
            deg_width = (eddy_day['radius'] * circle_enlargement_factor) / scale['meters_per_degree_lon']
            ell = Ellipse((eddy_day['center_lon'], eddy_day['center_lat']), width=2*deg_width, height=2*deg_height,
                          edgecolor=radius_color, facecolor='none', linestyle='--', alpha=0.4, linewidth=1.5, label=circle_label)
            ax.add_patch(ell)
            labeled_radius = True
            
        # 标记交互的起始和结束日期
        if eddy_day['date'] == interaction_track_df['date'].min() or eddy_day['date'] == interaction_track_df['date'].max():
             ax.text(eddy_day['center_lon'], eddy_day['center_lat'] + 0.1, eddy_day['date'].strftime('%Y-%m-%d'), 
                     fontsize=16, color='black', ha='center', zorder=11)

    # --- 5. 绘制 ΔDO 异常与基准剖面 ---
    if plot_unrelated_argo and not base_argo_positions.empty:
        ax.scatter(
            base_argo_positions['Longitude'], base_argo_positions['Latitude'],
            facecolors='none', edgecolors='gray', linewidths=0.8, s=50,
            label='All Matched Argo Profiles', zorder=5
        )

    if not anomalies.empty:
        scatter_kwargs = {}
        if fix_colorbar:
            cbar_lo, cbar_hi = cfg.resolved_cbar()
            scatter_kwargs.update(dict(vmin=cbar_lo, vmax=cbar_hi))
        depth_label = (
            f' @ depth ≥ {cfg.anomaly_min_depth} m'
            if cfg.anomaly_min_depth is not None and cfg.anomaly_min_depth > 0
            else ''
        )
        color_values, _, color_label, cmap_name = _color_values_for_anomalies(anomalies, cfg)
        if color_values is None:
            color_values = pd.Series(np.arange(len(anomalies)), index=anomalies.index)
        sc = ax.scatter(
            anomalies['Longitude'], anomalies['Latitude'],
            c=color_values, cmap=cmap_name, s=90,
            edgecolors='black', linewidths=0.6,
            label=f'{cfg.threshold_label()}{depth_label}',
            zorder=10,
            **scatter_kwargs
        )
        cbar = plt.colorbar(sc, ax=ax, orientation='horizontal', fraction=0.046, pad=0.08)
        cbar.set_label(color_label, fontsize=18)
        cbar.ax.tick_params(labelsize=14)
        if fix_colorbar:
            _apply_detection_colorbar_ticks(cbar, cfg, cbar_lo, cbar_hi)

    # --- 6. 最终化绘图设置 ---
    # 设定边界时排除META中错误的contour数据
    valid_contours_lon = [lon for day_lons in track_df['contour_lon'] for lon in day_lons if lon != 180.0]
    valid_contours_lat = [lat for day_lats in track_df['contour_lat'] for lat in day_lats if lat != 0.0]
    if valid_contours_lon and valid_contours_lat:
        ax.set_xlim(min(valid_contours_lon) - 0.5, max(valid_contours_lon) + 0.5)
        ax.set_ylim(min(valid_contours_lat) - 0.5, max(valid_contours_lat) + 0.5)
        
    ax.set_aspect('equal')
    ax.legend(fontsize=18)
    ax.tick_params(axis='both', which='major', labelsize=16)

    # --- 7. 输出控制 ---
    if save_fig:
        region_slug = _current_region_key()
        run_tag = cfg.file_stem()
        output_dir = cfg.output_dir("plot_track", region_slug)
        output_dir.mkdir(exist_ok=True, parents=True)
        base_filename = f"Track_Analysis_{ds_name}{num}_{run_tag}.png"
        save_path = output_dir / base_filename
        plt.savefig(save_path, dpi=300, bbox_inches='tight')
        print(f"Figure saved to: {save_path}")
    
    if show_fig:
        plt.show()
    
    plt.close(fig)

def plot_track_variable_timeseries(
    DS: list | str | tuple | dict,
    no: int,
    variable: str | None = None,
    only_anomaly_profiles: bool = True,
    depth_col: str = 'Depth',
    max_depth: float | None = 1000.0,
    depth_bin_size: float = 25.0,
    cmap: str = 'RdYlBu_r',
    show_profile_hist: bool = True,
    start_date: str | int | float | pd.Timestamp | None = None,
    end_date: str | int | float | pd.Timestamp | None = None,
    platform_number: int | str | list | tuple | set | None = None,
    detection_config: DetectionConfig | None = None,
    remove_outliers: bool = True,
    save_fig: bool = False,
    show_fig: bool = True,
):
    """绘制位于涡旋内的 Argo 剖面变量在时间-深度平面上的连续等值分布。
    

    参数:
        - DS (list | str | tuple | dict): legacy 轨迹列表、kind 字符串、字符串序列或数据集字典。
        - no (int): 涡旋编号。
        - variable (str | None): 需要统计的变量列名；None 时按 detection_config 自动选择（do→'DO'，aou/trim→'AOU'）。
        - only_anomaly_profiles (bool): True 时先按 detection_config 筛出异常剖面再绘制其变量场，False 时使用全部匹配剖面，默认 True。
        - depth_col (str): 深度列名，默认 'Depth'。
        - max_depth (float | None): 最大深度（单位与 depth_col 一致），默认 1000.0；None 表示使用观测最大值。
        - depth_bin_size (float): 深度分箱大小（dbar），默认 25.0。
        - cmap (str): 色标名称，默认 'RdYlBu_r'。
        - show_profile_hist (bool): 是否在底部显示每日剖面数柱状条，默认 True。
        - start_date (str | int | float | pd.Timestamp | None): 横轴开始日期，可传 pandas.Timestamp、YYYYMMDD 整数、days-since-1950 或 ISO 字符串。
        - end_date (str | int | float | pd.Timestamp | None): 横轴结束日期，编码同 start_date。
        - platform_number (int | str | list | tuple | set | None): 仅绘制指定 Argo 浮标（单个或列表/集合），先于异常筛选过滤。
        - detection_config (DetectionConfig | None): 异常识别配置；None 时使用 processing.yml 默认值。
        - remove_outliers (bool): 是否执行基础 QC（Flag 过滤 + DO<=1 过滤），默认 True。
        - save_fig (bool): 是否保存图像，默认 False。
        - show_fig (bool): 是否显示图像，默认 True。
    输出:
        - 图像（save_fig=True 时）：`plot_outputs/<method>/<region>/plot_track_timeseries/{数据集}{编号}_{变量}_timeseries*_{stem}.png`
    """
    cfg = _resolve_detection_config(detection_config)
    if variable is None:
        variable = cfg.timeseries_variable()

    print(f"[*] Building {variable} time series for eddy {no}...")

    display_variable = variable
    db_variable_name = _map_plot_variable_name(display_variable)

    def _normalize_date_input(value, label):
        if value is None:
            return None
        if isinstance(value, pd.Timestamp):
            return value.normalize()
        try:
            converted = convert_date(value)
        except Exception as exc:
            raise ValueError(f"Invalid {label}: {exc}")
        if isinstance(converted, pd.Series):
            converted = converted.iloc[0]
        if pd.isna(converted):
            raise ValueError(f"Invalid {label}: could not parse {value!r}")
        return pd.to_datetime(converted).normalize()

    try:
        start_dt = _normalize_date_input(start_date, 'start_date')
        end_dt = _normalize_date_input(end_date, 'end_date')
    except ValueError as exc:
        print(f"  - {exc}")
        return

    if start_dt is not None and end_dt is not None and start_dt > end_dt:
        print("  - start_date must be earlier than or equal to end_date.")
        return

    try:
        track_df, ds_name, ds_source_for_filter = _resolve_track_context(DS, no, include_contours=False)
    except Exception as exc:
        print(f"  - Error: {exc}")
        return

    argo_data = filtered_float_data(ds_source_for_filter, no, track=track_df)
    if argo_data.empty:
        print(f"  - Skip plotting {ds_name}{no}: no matched Argo profiles.")
        return

    if db_variable_name == 'AOU':
        if not _has_plottable_profile_variable(argo_data, 'AOU'):
            print("  - Column 'AOU' not found and cannot be derived from DO/Temperature/Salinity.")
            return
        argo_data = argo_data.copy()
        argo_data['AOU'] = pd.to_numeric(_compute_profile_aou(argo_data), errors='coerce')
    elif db_variable_name not in argo_data.columns:
        if db_variable_name == display_variable:
            print(f"  - Column '{db_variable_name}' not found in matched Argo data.")
        else:
            print(f"  - Column '{display_variable}' (mapped to '{db_variable_name}') not found in matched Argo data.")
        return

    if depth_col not in argo_data.columns:
        print(f"  - Column '{depth_col}' not found in matched Argo data.")
        return

    argo_data = argo_data.copy()
    if platform_number is not None:
        if 'Platform_number' not in argo_data.columns:
            print("  - Column 'Platform_number' not available; cannot filter by platform_number.")
            return
        pn_list = platform_number
        if not isinstance(pn_list, (list, tuple, set)):
            pn_list = [pn_list]
        argo_data = argo_data[argo_data['Platform_number'].isin(pn_list)]
        if argo_data.empty:
            print("  - No data after filtering by platform_number.")
            return
    argo_data['date'] = pd.to_datetime(argo_data[['Year', 'Month', 'Day']], errors='coerce')
    if argo_data['date'].isna().all():
        print("  - No valid timestamps/depths available after parsing Year/Month/Day.")
        return
    if start_dt is not None:
        argo_data = argo_data[argo_data['date'] >= start_dt]
    if end_dt is not None:
        argo_data = argo_data[argo_data['date'] <= end_dt]
    if argo_data.empty:
        print("  - No data within the requested date window.")
        return

    if 'Profile_number' not in argo_data.columns:
        print("  - Column 'Profile_number' is required to align with anomaly filtering.")
        return

    anomaly_filtered = False
    if only_anomaly_profiles:
        anomalies = calculate_delta_do(
            argo_data,
            detection_config=cfg,
            depth_col=depth_col,
            remove_outliers=remove_outliers,
            verbose=False,
        )
        if anomalies.empty or 'Profile_number' not in anomalies.columns:
            print(f"  - No profiles satisfy the anomaly criteria ({cfg.threshold_label()}).")
            return
        anomalies_unique = _keep_best_anomaly_per_profile(anomalies, cfg)
        qualifying_profiles = anomalies_unique['Profile_number'].dropna().unique()
        if qualifying_profiles.size == 0:
            print(f"  - No profiles satisfy the anomaly criteria ({cfg.threshold_label()}).")
            return
        argo_data = argo_data[argo_data['Profile_number'].isin(qualifying_profiles)].copy()
        if argo_data.empty:
            print("  - No profiles remain after applying anomaly filter.")
            return
        anomaly_filtered = True

    if remove_outliers:
        if db_variable_name == 'AOU':
            argo_data['AOU'], _ = _compute_aou_for_plot(argo_data, remove_outliers=True)
        else:
            argo_data = _apply_basic_argo_qc(argo_data, db_variable_name)

    argo_data['_var'] = pd.to_numeric(argo_data[db_variable_name], errors='coerce')
    argo_data['_depth'] = pd.to_numeric(argo_data[depth_col], errors='coerce')
    argo_data.dropna(subset=['date', '_depth'], inplace=True)
    if argo_data.empty:
        print("  - No valid timestamps/depths available after parsing Year/Month/Day.")
        return

    mask = ~argo_data['_var'].isna()
    if max_depth is not None:
        mask &= argo_data['_depth'] <= max_depth
    selected = argo_data.loc[mask]
    if selected.empty:
        print("  - No profiles satisfy the selected variable/depth conditions.")
        return

    if depth_bin_size <= 0:
        raise ValueError("depth_bin_size must be positive.")

    if max_depth is None:
        max_depth_val = float(selected['_depth'].max())
    else:
        max_depth_val = float(max_depth)
    depth_min_val = float(max(selected['_depth'].min(), 0.0))
    depth_bins = np.arange(depth_min_val, max_depth_val + depth_bin_size, depth_bin_size)
    if depth_bins.size < 2:
        depth_bins = np.array([depth_min_val, depth_min_val + depth_bin_size])
    selected = selected.copy()
    selected['depth_bin'] = pd.cut(selected['_depth'], bins=depth_bins, include_lowest=True, right=False)
    selected.dropna(subset=['depth_bin'], inplace=True)
    selected['depth_mid'] = selected['depth_bin'].apply(lambda iv: iv.left + depth_bin_size / 2 if pd.notna(iv) else np.nan)
    selected.dropna(subset=['depth_mid'], inplace=True)

    grouped = (
        selected.groupby(['depth_mid', 'date'], observed=False)['_var']
        .mean()
        .reset_index()
    )
    if grouped.empty:
        print("  - No averages could be computed for the selected bins.")
        return

    pivot = grouped.pivot(index='depth_mid', columns='date', values='_var')
    pivot.sort_index(inplace=True)
    pivot = pivot.sort_index(axis=1)
    if pivot.isna().all().all():
        print("  - All grid cells are NaN; cannot plot.")
        return

    depth_vals = pivot.index.to_numpy()
    date_vals = pivot.columns.to_pydatetime()
    if len(date_vals) == 0:
        print("  - No valid dates after grouping.")
        return

    Z = np.ma.masked_invalid(pivot.to_numpy())
    if np.ma.count_masked(Z) == Z.size:
        print("  - No valid data points remain after masking.")
        return

    vmin = np.nanmin(pivot.values)
    vmax = np.nanmax(pivot.values)
    if not np.isfinite(vmin) or not np.isfinite(vmax):
        print("  - Unable to determine color scale (all values NaN/inf).")
        return
    if np.isclose(vmin, vmax):
        levels = 10
    else:
        levels = 20

    date_nums = mdates.date2num(date_vals)
    if show_profile_hist:
        hist_counts = selected.groupby('date').size().reindex(pivot.columns, fill_value=0)

    height_ratios = [5, 1] if show_profile_hist else [1]
    nrows = 2 if show_profile_hist else 1
    fig, axes = plt.subplots(
        nrows=nrows,
        ncols=1,
        sharex=True,
        figsize=(20, 10 if show_profile_hist else 8),
        gridspec_kw={'height_ratios': height_ratios},
        constrained_layout=True
    )
    if show_profile_hist:
        ax, ax_hist = axes
    else:
        ax = axes

    cf = ax.contourf(date_vals, depth_vals, Z, levels=levels, cmap=cmap)
    ax.invert_yaxis()
    ax.set_ylabel('Depth (dbar)', fontsize=14)
    title_suffix = ' (anomaly profiles)' if only_anomaly_profiles else ' (all profiles)'
    ax.set_title(f"{ds_name}{no} {display_variable} mean inside eddy{title_suffix}", fontsize=16)
    ax.xaxis_date()
    ax.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d'))

    if start_dt is not None or end_dt is not None:
        xmin = start_dt if start_dt is not None else date_vals[0]
        xmax = end_dt if end_dt is not None else date_vals[-1]
        ax.set_xlim(xmin, xmax)

    cbar_axes = [ax, ax_hist] if show_profile_hist else ax
    cbar = fig.colorbar(
        cf,
        ax=cbar_axes,
        orientation='vertical',
        pad=0.12,
        fraction=0.035,
        location='right'
    )
    cbar.set_label(f'{display_variable} mean', fontsize=14)

    if show_profile_hist:
        ax_hist.bar(hist_counts.index, hist_counts.values, width=0.8, color='gray')
        ax_hist.set_ylabel('Profiles', fontsize=12)
        ax_hist.set_xlabel('Date', fontsize=14)
        ax_hist.set_ylim(0, max(hist_counts.max() * 1.2, 1))
        ax_hist.xaxis_date()
        ax_hist.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d'))
        if start_dt is not None or end_dt is not None:
            xmin = start_dt if start_dt is not None else hist_counts.index.min()
            xmax = end_dt if end_dt is not None else hist_counts.index.max()
            ax_hist.set_xlim(xmin, xmax)
    else:
        ax.set_xlabel('Date', fontsize=14)

    fig.autofmt_xdate()

    if save_fig:
        region_slug = _current_region_key()
        run_tag = cfg.file_stem()
        output_dir = cfg.output_dir("plot_track_timeseries", region_slug)
        output_dir.mkdir(parents=True, exist_ok=True)
        profile_suffix = "_anomaly_profiles" if only_anomaly_profiles else "_all_profiles"
        depth_suffix = f"_depth{str(max_depth).replace('.', 'p')}" if max_depth is not None else ''
        base_filename = f"{ds_name}{no}_{variable}_timeseries{profile_suffix}{depth_suffix}_{run_tag}.png"
        save_path = output_dir / base_filename
        plt.savefig(save_path, dpi=300, bbox_inches='tight')
        print(f"Figure saved to: {save_path}")

    if show_fig:
        plt.show()
    plt.close(fig)
    
def convert_date(values: pd.Series | np.ndarray | list | str | int | float) -> pd.Series | pd.Timestamp:
    """统一将整数/字符串编码日期转为 pandas datetime（日精度）。

    支持两种编码：自 1950-01-01 起的天数（CF 常见 time 轴），或 YYYYMMDD 8 位整数（如 20220131）。

    参数:
        - values (pd.Series | np.ndarray | list | str | int | float): 待转换的日期，标量或序列。

    返回:
        - pd.Series | pd.Timestamp: 标量输入返回单个 pd.Timestamp；序列输入返回 pd.Series（datetime64[ns]），未能解析的元素为 NaT。
    """
    is_scalar_input = isinstance(values, (str, bytes, int, float, np.integer, np.floating, np.str_))
    if isinstance(values, pd.Series):
        ser = values
    elif is_scalar_input:
        ser = pd.Series([values])
    else:
        # 注意：若 values 是字符串且不想被逐字符拆分，需封装为列表
        if isinstance(values, (str, bytes, np.str_)):
            ser = pd.Series([values])
        else:
            ser = pd.Series(values)
    # 结果容器，避免在原 float Series 上就地赋值导致 dtype 警告
    result = pd.Series(pd.NaT, index=ser.index, dtype='datetime64[ns]')

    non_na = ser.dropna()
    if non_na.empty:
        return result.iloc[0] if is_scalar_input else result

    # 若输入已是 datetime64 类型，直接标准化返回
    if pd.api.types.is_datetime64_any_dtype(non_na):
        dt = pd.to_datetime(non_na).dt.normalize()
        result.loc[dt.index] = dt.values
        return result.iloc[0] if is_scalar_input else result

    # 判定是否全为 8 位 YYYYMMDD
    try:
        strs = non_na.astype('int64').astype(str)
    except Exception:
        strs = pd.Series([str(x) for x in non_na], index=non_na.index)
    is_all_yyyymmdd = strs.map(len).eq(8).all()

    if is_all_yyyymmdd:
        parsed = []
        for s in strs:
            try:
                y = int(s[0:4]); m = int(s[4:6]); d = int(s[6:8])
                parsed.append(pd.Timestamp(year=y, month=m, day=d))
            except Exception as e:
                raise ValueError(f"Invalid YYYYMMDD integer {s}: {e}")
        dt = pd.Series(parsed, index=strs.index)
        result.loc[dt.index] = dt.values
        return result.iloc[0] if is_scalar_input else result

    # 默认按 days since 1950-01-01 解析（对浮点会取整）
    t0 = pd.Timestamp('1950-01-01')
    try:
        deltas = pd.to_timedelta(non_na.astype('int64'), unit='D')
        dt = t0 + deltas
        result.loc[dt.index] = dt.values
        return result.iloc[0] if is_scalar_input else result
    except Exception as e:
        raise ValueError(f"Values neither valid YYYYMMDD nor days-since-1950: {e}")

def convert_eddy_number(
    kind: str,
    value: str | int | float | list | tuple | np.ndarray,
    order: str,
    *,
    meta_root: str | os.PathLike | None = None,
    version: float = 3.2
) -> int | list[int]:
    """
    在 legacy orig_index 与标准 track_id 之间转换涡旋编号。

    参数:
        - kind (str): 涡旋类型 'acs'|'acl'|'cs'|'cl'，必须指定以保证索引映射到正确数据集。
        - value (str | int | float | list | tuple | np.ndarray): 单个编号或编号可迭代；可为旧编号（orig_index）或新编号（track_id）。
        - order (str): 转换方向，支持别名：

            - legacy→new：'old_to_new'、'orig_to_track'、'legacy_to_track'、'to_track'。
            - new→legacy：'new_to_old'、'track_to_orig'、'track_to_legacy'、'to_orig'。
        - meta_root (str | os.PathLike | None): META 原始 NetCDF 目录；None 时取配置 paths.meta_root。
        - version (float): META 版本（3.1 或 3.2），默认 3.2。

    返回:
        - int | list[int]: 按输入形状返回 —— 标量输入返回 int，可迭代输入返回 list[int]。

    说明:
        - 旧编号对应 NetCDF 中的全局行索引（orig_index）。
        - 新编号为轨迹唯一的 track_id。
        - new->old 方向返回该轨迹首次出现的行索引（最早时间点）。
    """

    direction = order.lower().strip()
    if direction in {'old_to_new', 'orig_to_track', 'legacy_to_track', 'to_track', 'orig_to_new'}:
        mode = 'orig_to_track'
    elif direction in {'new_to_old', 'track_to_orig', 'track_to_legacy', 'to_orig'}:
        mode = 'track_to_orig'
    else:
        raise ValueError("order must be one of: old_to_new/orig_to_track/legacy_to_track/to_track or new_to_old/track_to_orig/track_to_legacy/to_orig")

    kind_norm = kind.lower().strip()
    if kind_norm not in {'acs', 'acl', 'cs', 'cl'}:
        raise ValueError("kind must be one of 'acs', 'acl', 'cs', 'cl'")

    ACS, ACL, CS, CL = load_meta_data(path=meta_root, version=version)
    ds_map = {'acs': ACS, 'acl': ACL, 'cs': CS, 'cl': CL}
    ds = ds_map[kind_norm]

    # 将 track 数组载入内存，便于快速索引或反查
    track_arr = np.asarray(ds.variables['track'][:], dtype=np.int64)

    def _convert_one(val) -> int:
        try:
            n = int(val)
        except Exception as e:
            raise ValueError(f"Cannot convert value '{val}' to int") from e

        if mode == 'orig_to_track':
            if n < 0 or n >= track_arr.shape[0]:
                raise ValueError(f"orig_index {n} out of range for {kind_norm}")
            return int(track_arr[n])

        # track_id -> 最早出现的 orig_index
        hits = np.nonzero(track_arr == n)[0]
        if hits.size == 0:
            raise ValueError(f"track_id {n} not found in {kind_norm}")
        return int(hits[0])

    is_scalar = np.isscalar(value) or isinstance(value, (str, bytes, np.str_))
    if is_scalar:
        return _convert_one(value)

    try:
        values = list(value)
    except Exception as e:
        raise ValueError("value must be scalar or iterable of scalars") from e

    return [_convert_one(v) for v in values]


def _map_plot_variable_name(var_name: str) -> str:
    """将绘图变量名映射到 Argo 标准列名。"""
    return 'Temperature' if str(var_name) == 'Temp' else str(var_name)


def _prepare_vertical_plot_variables(
    variables: list,
    *,
    combine_do_aou: bool = True,
) -> tuple[list[str], bool]:
    """根据 variables 生成实际子图变量列表，并决定 DO 轴是否叠加 AOU。

    规则:
        - combine_do_aou=True 且同时包含 DO 和 AOU: 合并为同一子图（DO 为主轴，AOU 顶部叠加）。
        - combine_do_aou=False 且同时包含 DO 和 AOU: 分成两个子图（DO 与 AOU 分别绘制）。
        - 仅包含 DO 或仅包含 AOU: 第一张图绘制对应变量。
        - 其它变量保持原顺序，去重按标准列名进行。
    """
    if variables is None:
        variables = []

    normalized: list[str] = []
    for item in variables:
        name = str(item).strip()
        if not name:
            continue
        normalized.append(name)

    has_do = any(_map_plot_variable_name(v) == 'DO' for v in normalized)
    has_aou = any(_map_plot_variable_name(v) == 'AOU' for v in normalized)

    # 先放 DO/AOU 组，确保在第一张图。
    plot_variables: list[str] = []
    if combine_do_aou:
        if has_do:
            plot_variables.append('DO')
        elif has_aou:
            plot_variables.append('AOU')
    else:
        if has_do:
            plot_variables.append('DO')
        if has_aou:
            plot_variables.append('AOU')

    seen = set(_map_plot_variable_name(v) for v in plot_variables)
    for raw_name in normalized:
        mapped = _map_plot_variable_name(raw_name)
        if mapped in ('DO', 'AOU'):
            continue
        if mapped in seen:
            continue
        plot_variables.append(raw_name)
        seen.add(mapped)

    if not plot_variables:
        raise ValueError("variables cannot be empty")

    overlay_aou_on_do = combine_do_aou and has_do and has_aou
    return plot_variables, overlay_aou_on_do


def _has_plottable_profile_variable(profile_rows: pd.DataFrame, value_col: str) -> bool:
    """判断剖面是否具备可绘制变量（AOU 支持由 DO/Temperature/Salinity 推导）。"""
    if value_col in profile_rows.columns:
        return True
    if value_col == 'AOU':
        required = ['DO', 'Temperature', 'Salinity']
        return all(c in profile_rows.columns for c in required)
    return False


def _apply_basic_argo_qc(profile_rows: pd.DataFrame, value_col: str) -> pd.DataFrame:
    """对单个变量执行基础 Argo 质控（Flag 过滤 + DO 规则过滤）。"""
    rows_qc = profile_rows.copy()
    qc_column_name = f"{value_col}_Flag"
    if qc_column_name in rows_qc.columns:
        good_qc_flags = ['1', '2', '5', '8', 1, 2, 5, 8]
        bad_qc_mask = ~rows_qc[qc_column_name].isin(good_qc_flags)
        rows_qc.loc[bad_qc_mask, value_col] = np.nan
    if value_col == 'DO':
        rows_qc.loc[rows_qc[value_col] <= 1.0, value_col] = np.nan
    return rows_qc


def _basic_argo_qc_bad_mask(profile_rows: pd.DataFrame, value_col: str) -> pd.Series:
    """返回基础 QC 下的异常点掩膜（True 表示该点不通过 QC）。"""
    bad_mask = pd.Series(False, index=profile_rows.index)
    qc_column_name = f"{value_col}_Flag"

    if qc_column_name in profile_rows.columns:
        good_qc_flags = ['1', '2', '5', '8', 1, 2, 5, 8]
        bad_mask = bad_mask | (~profile_rows[qc_column_name].isin(good_qc_flags))

    if value_col == 'DO' and value_col in profile_rows.columns:
        do_vals = pd.to_numeric(profile_rows[value_col], errors='coerce')
        bad_mask = bad_mask | (do_vals <= 1.0)

    return bad_mask.fillna(False)


def _aou_qc_bad_mask(profile_rows: pd.DataFrame) -> pd.Series:
    """返回 AOU 的基础 QC 异常掩膜（由 DO/Temperature/Salinity 任一异常触发）。"""
    bad_mask = pd.Series(False, index=profile_rows.index)
    for base_col in ('DO', 'Temperature', 'Salinity'):
        if base_col in profile_rows.columns:
            bad_mask = bad_mask | _basic_argo_qc_bad_mask(profile_rows, base_col)
    return bad_mask.fillna(False)


def _compute_aou_for_plot(
    profile_rows: pd.DataFrame,
    *,
    remove_outliers: bool,
) -> tuple[pd.Series, pd.Series]:
    """为绘图准备 AOU 序列与异常掩膜（与 DO 同级 QC 口径）。"""
    rows_for_aou = profile_rows.copy()
    base_bad_mask = _aou_qc_bad_mask(profile_rows)
    if remove_outliers:
        for base_col in ('DO', 'Temperature', 'Salinity'):
            if base_col in rows_for_aou.columns:
                rows_for_aou = _apply_basic_argo_qc(rows_for_aou, base_col)
        aou_vals = pd.to_numeric(_compute_profile_aou(rows_for_aou), errors='coerce')
        # 即使源数据存在 AOU 列，也强制应用与 DO/Temp/Sal 一致的 QC 掩膜，避免跨坏点连线。
        aou_vals = aou_vals.where(~base_bad_mask.reindex(aou_vals.index, fill_value=False), np.nan)
        bad_mask = base_bad_mask.reindex(rows_for_aou.index, fill_value=False)
    else:
        aou_vals = pd.to_numeric(_compute_profile_aou(rows_for_aou), errors='coerce')
        bad_mask = base_bad_mask.reindex(rows_for_aou.index, fill_value=False)

    return aou_vals, bad_mask.fillna(False)


def _compute_profile_aou(profile_rows: pd.DataFrame) -> pd.Series:
    """计算单剖面的 AOU（若已有 AOU 列则优先直接使用）。"""
    if 'AOU' in profile_rows.columns:
        return pd.to_numeric(profile_rows['AOU'], errors='coerce')

    required = ['DO', 'Temperature', 'Salinity']
    if any(c not in profile_rows.columns for c in required):
        return pd.Series(np.nan, index=profile_rows.index, dtype=float)

    do_vals = pd.to_numeric(profile_rows['DO'], errors='coerce').to_numpy(dtype=float)
    temp_vals = pd.to_numeric(profile_rows['Temperature'], errors='coerce').to_numpy(dtype=float)
    sal_vals = pd.to_numeric(profile_rows['Salinity'], errors='coerce').to_numpy(dtype=float)

    valid = np.isfinite(do_vals) & np.isfinite(temp_vals) & np.isfinite(sal_vals)
    if not valid.any():
        return pd.Series(np.nan, index=profile_rows.index, dtype=float)

    sat_vals = np.full(do_vals.shape, np.nan, dtype=float)
    try:
        if hasattr(gsw, 'O2sol_SP_pt'):
            sat_valid = gsw.O2sol_SP_pt(sal_vals[valid], temp_vals[valid])
        else:
            depth_vals = (
                pd.to_numeric(profile_rows['Depth'], errors='coerce').to_numpy(dtype=float)
                if 'Depth' in profile_rows.columns else np.zeros_like(do_vals)
            )
            lon_vals = (
                pd.to_numeric(profile_rows['Longitude'], errors='coerce').to_numpy(dtype=float)
                if 'Longitude' in profile_rows.columns else np.zeros_like(do_vals)
            )
            lat_vals = (
                pd.to_numeric(profile_rows['Latitude'], errors='coerce').to_numpy(dtype=float)
                if 'Latitude' in profile_rows.columns else np.zeros_like(do_vals)
            )
            depth_vals = np.nan_to_num(depth_vals, nan=0.0)
            lon_vals = np.nan_to_num(lon_vals, nan=0.0)
            lat_vals = np.nan_to_num(lat_vals, nan=0.0)
            sa_vals = gsw.SA_from_SP(sal_vals[valid], depth_vals[valid], lon_vals[valid], lat_vals[valid])
            ct_vals = gsw.CT_from_pt(sa_vals, temp_vals[valid])
            sat_valid = gsw.O2sol(sa_vals, ct_vals, depth_vals[valid], lon_vals[valid], lat_vals[valid])

        sat_vals[valid] = np.asarray(sat_valid, dtype=float)
    except Exception:
        return pd.Series(np.nan, index=profile_rows.index, dtype=float)

    return pd.Series(sat_vals - do_vals, index=profile_rows.index, dtype=float)


def _isolated_valid_mask(valid_mask: np.ndarray | pd.Series) -> np.ndarray:
    """识别孤立有效点：本点有效且上下相邻点均无效。"""
    arr = np.asarray(valid_mask, dtype=bool)
    if arr.size == 0:
        return arr
    prev_valid = np.r_[False, arr[:-1]]
    next_valid = np.r_[arr[1:], False]
    return arr & (~prev_valid) & (~next_valid)


def _anomaly_colors_for_variable(db_variable_name: str) -> tuple[str, str]:
    """返回(异常桥接线颜色, 异常点颜色)。"""
    if db_variable_name == 'AOU':
        # 与 DO 的异常红保持同色系，但差异更明显
        return ('#ad1457', '#ad1457')
    return ('#d62728', '#d62728')


_ANOMALY_POINT_SIZE = 40
_SINGLETON_POINT_SIZE = _ANOMALY_POINT_SIZE


def _overlay_aou_top_axis(
    ax,
    depth_vals: np.ndarray,
    aou_vals: np.ndarray,
    *,
    alpha: float = 0.7,
    color: str = '#1f77b4',
    bad_mask: np.ndarray | pd.Series | None = None,
    anomaly_color: str = '#c83f5a',
    show_normal_scatter: bool = True,
) -> bool:
    """在 DO 子图顶部叠加 AOU 横轴与曲线（可选 QC 异常段高亮）。"""
    depth_arr = np.asarray(depth_vals, dtype=float)
    aou_arr = np.asarray(aou_vals, dtype=float)
    finite = np.isfinite(depth_arr) & np.isfinite(aou_arr)
    if not finite.any():
        return False

    if bad_mask is None:
        bad_arr = np.zeros_like(finite, dtype=bool)
    else:
        bad_arr = np.asarray(bad_mask, dtype=bool)
        if bad_arr.shape != finite.shape:
            bad_arr = np.zeros_like(finite, dtype=bool)
    bad_arr = bad_arr & finite
    good_arr = finite & (~bad_arr)

    aou_ax = getattr(ax, '_aou_top_axis', None)
    if aou_ax is None:
        aou_ax = ax.twiny()
        aou_ax.xaxis.set_ticks_position('top')
        aou_ax.xaxis.set_label_position('top')
        aou_ax.set_xlabel('AOU', fontsize=16, color=color)
        aou_ax.tick_params(axis='x', labelsize=12, colors=color)
        aou_ax.grid(False)
        setattr(ax, '_aou_top_axis', aou_ax)
        setattr(ax, '_aou_xlim', None)

    blue_alpha = min(1.0, alpha + 0.05)
    if good_arr.any():
        aou_plot = np.where(good_arr, aou_arr, np.nan)
        aou_ax.plot(aou_plot, depth_arr, color=color, alpha=blue_alpha, linewidth=1.4, zorder=2)
    else:
        aou_plot = np.where(finite, aou_arr, np.nan)
        aou_ax.plot(aou_plot, depth_arr, color=color, alpha=blue_alpha, linewidth=1.4, zorder=2)

    singleton_good = _isolated_valid_mask(good_arr)
    if show_normal_scatter and singleton_good.any():
        aou_ax.scatter(
            aou_arr[singleton_good],
            depth_arr[singleton_good],
            color=color,
            marker='o',
            s=_SINGLETON_POINT_SIZE,
            linewidths=0.6,
            alpha=min(1.0, alpha + 0.15),
            zorder=6,
        )

    if bad_arr.any():
        valid_positions = np.flatnonzero(good_arr)
        bridge_alpha = min(1.0, alpha + 0.2)
        for left_pos, right_pos in zip(valid_positions[:-1], valid_positions[1:]):
            if right_pos - left_pos <= 1:
                continue
            if not bool(bad_arr[left_pos + 1:right_pos].any()):
                continue
            segment_depth = depth_arr[left_pos:right_pos + 1]
            segment_aou = aou_arr[left_pos:right_pos + 1]
            segment_finite = np.isfinite(segment_depth) & np.isfinite(segment_aou)
            if not segment_finite.any():
                continue
            aou_ax.plot(
                np.where(segment_finite, segment_aou, np.nan),
                np.where(segment_finite, segment_depth, np.nan),
                color=anomaly_color,
                alpha=bridge_alpha,
                linewidth=1.8,
                zorder=4,
            )

        flagged = bad_arr & finite
        if flagged.any():
            aou_ax.scatter(
                aou_arr[flagged],
                depth_arr[flagged],
                color=anomaly_color,
                marker='o',
                s=_ANOMALY_POINT_SIZE,
                linewidths=1.2,
                alpha=min(1.0, alpha + 0.2),
                zorder=5,
            )

    xmin = float(np.nanmin(aou_arr[finite]))
    xmax = float(np.nanmax(aou_arr[finite]))
    old_xlim = getattr(ax, '_aou_xlim', None)
    if old_xlim is None:
        new_xlim = (xmin, xmax)
    else:
        new_xlim = (min(old_xlim[0], xmin), max(old_xlim[1], xmax))

    if np.isclose(new_xlim[0], new_xlim[1]):
        pad = max(1.0, abs(new_xlim[0]) * 0.05)
    else:
        pad = (new_xlim[1] - new_xlim[0]) * 0.05
    aou_ax.set_xlim(new_xlim[0] - pad, new_xlim[1] + pad)
    setattr(ax, '_aou_xlim', new_xlim)
    return True


def _plot_single_argo_profile_line(
    ax,
    profile_rows: pd.DataFrame,
    var_name: str,
    color,
    *,
    remove_outliers: bool = True,
    show_normal_scatter: bool = True,
    do_aux_layers: list[str] | tuple[str, ...] | str | None = None,
    aou_aux_color: str = '#1f77b4',
    alpha: float = 0.7,
) -> bool:
    """在指定坐标轴上绘制单个 Argo 剖面单变量曲线，返回是否成功绘制。

    行为:
        - remove_outliers=True: 按基础 QC 剔除异常值后绘制。
        - remove_outliers=False: 保留 QC 通过段为原色，并将本应断开的跨段连接用红线桥接；
            同时用红色圆点标记不通过基础 QC 的点。
        - show_normal_scatter=True: 额外绘制“正常值孤立点”的散点标记；False 时不绘制该标记。
        - do_aux_layers 包含 'aou' 且变量为 DO 时：在顶部横轴叠加 AOU 曲线。
        - var_name='AOU' 时：优先使用 AOU 列；若缺失则由 DO/Temperature/Salinity 现场推导，
            并沿用同一套 QC 过滤与异常段高亮逻辑。
    """
    aux_layers_norm: tuple[str, ...]
    if do_aux_layers is None:
        aux_layers_norm = tuple()
    elif isinstance(do_aux_layers, str):
        aux_layers_norm = (str(do_aux_layers).strip().lower(),)
    else:
        aux_layers_norm = tuple(str(x).strip().lower() for x in do_aux_layers)
    overlay_aou = ('aou' in aux_layers_norm)

    db_variable_name = _map_plot_variable_name(var_name)
    if 'Depth' not in profile_rows.columns:
        return False
    if not _has_plottable_profile_variable(profile_rows, db_variable_name):
        return False

    rows_to_plot = profile_rows.copy()
    if db_variable_name == 'AOU':
        # 先用原始 AOU（或由原始 DO/T/S 推导）建立可绘制骨架，
        # 避免在 remove_outliers=True 时提前删掉坏点导致跨段相连。
        rows_to_plot['AOU'] = pd.to_numeric(_compute_profile_aou(rows_to_plot), errors='coerce')

    rows_to_plot = rows_to_plot.dropna(subset=[db_variable_name, 'Depth']).copy()
    if rows_to_plot.empty:
        return False
    rows_to_plot = rows_to_plot.sort_values('Depth').reset_index(drop=True)

    if remove_outliers:
        if db_variable_name == 'AOU':
            rows_qc = rows_to_plot.copy()
            rows_qc['AOU'], _ = _compute_aou_for_plot(rows_qc, remove_outliers=True)
            rows_to_plot = rows_qc
        else:
            rows_to_plot = _apply_basic_argo_qc(rows_to_plot, db_variable_name)

        if not rows_to_plot[db_variable_name].notna().any():
            return False

        ax.plot(rows_to_plot[db_variable_name], rows_to_plot['Depth'], color=color, alpha=alpha)
        valid_mask = rows_to_plot[db_variable_name].notna() & rows_to_plot['Depth'].notna()
        singleton_valid = _isolated_valid_mask(valid_mask.to_numpy())
        if show_normal_scatter and singleton_valid.any():
            ax.scatter(
                rows_to_plot.loc[singleton_valid, db_variable_name],
                rows_to_plot.loc[singleton_valid, 'Depth'],
                color=color,
                marker='o',
                s=_SINGLETON_POINT_SIZE,
                linewidths=0.7,
                alpha=min(1.0, alpha + 0.15),
                zorder=6,
            )
        if overlay_aou and db_variable_name == 'DO':
            aou_vals, _ = _compute_aou_for_plot(rows_to_plot, remove_outliers=True)
            _overlay_aou_top_axis(
                ax,
                rows_to_plot['Depth'].to_numpy(dtype=float),
                aou_vals.to_numpy(dtype=float),
                alpha=alpha,
                color=aou_aux_color,
                anomaly_color=_anomaly_colors_for_variable('AOU')[0],
                show_normal_scatter=show_normal_scatter,
            )
        return True

    if db_variable_name == 'AOU':
        bad_mask = _aou_qc_bad_mask(rows_to_plot)
    else:
        bad_mask = _basic_argo_qc_bad_mask(rows_to_plot, db_variable_name)
    valid_mask = (~bad_mask) & rows_to_plot[db_variable_name].notna() & rows_to_plot['Depth'].notna()

    plotted_any = False

    # 先画“QC 通过”的原色曲线（坏点处自动断开）
    x_valid = rows_to_plot[db_variable_name].where(valid_mask)
    if x_valid.notna().any():
        ax.plot(x_valid, rows_to_plot['Depth'], color=color, alpha=alpha)
        plotted_any = True

    singleton_valid = _isolated_valid_mask(valid_mask.to_numpy())
    if show_normal_scatter and singleton_valid.any():
        ax.scatter(
            rows_to_plot.loc[singleton_valid, db_variable_name],
            rows_to_plot.loc[singleton_valid, 'Depth'],
            color=color,
            marker='o',
            s=_SINGLETON_POINT_SIZE,
            linewidths=0.7,
            alpha=min(1.0, alpha + 0.15),
            zorder=6,
        )
        plotted_any = True

    # 再用红线重绘“本应断开”的跨段路径（包含中间坏点，保证红点落在红线上）
    valid_positions = np.flatnonzero(valid_mask.to_numpy())
    bridge_alpha = min(1.0, alpha + 0.15)
    bridge_color, point_color = _anomaly_colors_for_variable(db_variable_name)
    for left_pos, right_pos in zip(valid_positions[:-1], valid_positions[1:]):
        if right_pos - left_pos <= 1:
            continue
        if not bool(bad_mask.iloc[left_pos + 1:right_pos].any()):
            continue
        segment = rows_to_plot.iloc[left_pos:right_pos + 1]
        ax.plot(
            segment[db_variable_name],
            segment['Depth'],
            color=bridge_color,
            alpha=bridge_alpha,
            linewidth=2.0,
            zorder=4,
        )
        plotted_any = True

    if bad_mask.any():
        flagged = rows_to_plot[bad_mask & rows_to_plot[db_variable_name].notna() & rows_to_plot['Depth'].notna()]
        if not flagged.empty:
            ax.scatter(
                flagged[db_variable_name],
                flagged['Depth'],
                color=point_color,
                marker='o',
                s=_ANOMALY_POINT_SIZE,
                linewidths=1.2,
                alpha=min(1.0, alpha + 0.2),
                zorder=5,
            )
            plotted_any = True

    if overlay_aou and db_variable_name == 'DO':
        aou_vals, aou_bad_mask = _compute_aou_for_plot(rows_to_plot, remove_outliers=False)
        _overlay_aou_top_axis(
            ax,
            rows_to_plot['Depth'].to_numpy(dtype=float),
            aou_vals.to_numpy(dtype=float),
            alpha=alpha,
            bad_mask=aou_bad_mask.to_numpy(dtype=bool),
            color=aou_aux_color,
            anomaly_color=_anomaly_colors_for_variable('AOU')[0],
            show_normal_scatter=show_normal_scatter,
        )

    return plotted_any


def _apply_vertical_profile_axis_style(ax, var_name: str):
    """为 Argo 垂向曲线图应用通用坐标轴样式（可被多个绘图函数复用）。"""
    db_variable_name = _map_plot_variable_name(var_name)
    ax.set_ylim(-50, 2050)
    if db_variable_name == 'DO':
        ax.set_xlim(0, 350)
    elif db_variable_name == 'AOU':
        ax.set_xlim(-50, 350)
    elif db_variable_name == 'Temperature':
        ax.set_xlim(-2, 32)
    elif db_variable_name == 'Salinity':
        ax.set_xlim(32.5, 36.5)

    ax.set_xlabel(var_name, fontsize=20)
    ax.tick_params(axis='x', labelsize=16)
    ax.grid(True)

def plot_vertical(
    DS: list,
    no: int,
    show_fig: bool = False,
    save_fig: bool = False,
    color_mode: str = 'distance',
    variables: list = ['DO', 'Temp', 'Salinity'],
    show_colorbar: bool = False,
    remove_outliers: bool = True,
    plot_normal_scatter: bool = False,
    aggregated: bool = False,
    argo_required: list | None = None,
    year_required: list | None = None,
    month_required: list | None = None,
    day_required: list | None = None,
):
    '''
    根据涡旋轨迹与匹配到的 Argo 剖面，绘制变量-深度的垂直剖面。

    为 variables 中每个变量创建一个子图，按剖面绘制变量随深度变化的曲线，曲线颜色可按与涡旋中心的相对
    距离（distance）或采样时间（time）变化，可选显示颜色条并保存/显示图片；variables 同含 DO 与 AOU 时
    绘制两张独立子图（不做 DO 顶部 AOU 叠加）。

    参数:
        - DS (list): 涡旋轨迹数据集。
        - no (int): 涡旋编号。
        - show_fig (bool): 是否显示图片，默认 False。
        - save_fig (bool): 是否保存图片，默认 False。
        - color_mode (str): 颜色模式，'distance' 或 'time'，默认 'distance'。
        - variables (list): 需要绘制的变量名称，默认 ['DO', 'Temp', 'Salinity']。
        - show_colorbar (bool): 是否显示颜色条，默认 False。
        - remove_outliers (bool): True 时执行 QC 过滤与规则法去极值；False 时保留 QC 通过段为原色、断点用红线桥接并用红色圆点标记 QC 异常值，默认 True。
        - plot_normal_scatter (bool): 是否绘制正常值的孤立散点标记，默认 False。
        - aggregated (bool): 是否进行跨平台聚合绘制，默认 False。
        - argo_required (list | None): 平台过滤；None 表示不过滤，传入平台编号列表时仅保留指定平台。
        - year_required (list | None): 年份过滤；None 表示不过滤，传入年份列表时仅保留指定年份。
        - month_required (list | None): 月份过滤；聚合模式 None 表示使用所有可用月份，逐平台模式 None 表示不过滤。
        - day_required (list | None): 日期过滤（按日 1-31）；None 表示不过滤，传入日期列表时仅保留指定日期。
    输出:
        - 逐平台图（save_fig=True, aggregated=False）：`plot_outputs/shared/<region>/plot_vertical_profiles/{数据集}{编号}_Platform_{平台号}.png`
        - 聚合图（save_fig=True, aggregated=True）：`plot_outputs/shared/<region>/plot_vertical_monthly_aggregated/{数据集}{编号}_months_{月份}_aggregated.png`

    说明:
        模式差异:

            - aggregated=False（逐平台）：为每个浮标平台单独出图，每图含 variables 中各变量的一个子图。
            - aggregated=True（聚合）：所有平台剖面在同一张图上聚合绘制（每个变量一个子图）。

        质量控制:

            - remove_outliers=True 时执行基础 QC（仅保留 Flag {1,2,5,8}，DO<=1 置 NaN）。
            - remove_outliers=False 时不剔除点：QC 通过段保持原色，应断开的跨段连接用红线显示，QC 不通过的观测用红色圆点标记。
            - DO 子图绘制 AOU 时，AOU 复用 DO/Temp/Sal 的同等 QC 逻辑（True 同样剔除，False 同样以红色桥接线+红点标注）。
    '''
    try:
        track_df, ds_names, ds_source_for_filter = _resolve_track_context(DS, no, include_contours=False)
    except Exception as exc:
        print(f"  - Error: {exc}")
        return

    argo_data_filtered = filtered_float_data(ds_source_for_filter, no, track=track_df)

    if argo_data_filtered.empty:
        msg = "plot vertical profiles." if not aggregated else "to plot aggregated vertical profiles."
        print(f"No Argo data found for eddy {ds_names}{no} {msg}")
        return

    plot_variables, overlay_aou_on_do = _prepare_vertical_plot_variables(
        variables,
        combine_do_aou=False,
    )

    # 统一的月份和平台过滤（在两种模式下共享）
    # 先处理平台过滤（argo_required）
    if argo_required is not None:
        try:
            argo_required_set = set(int(x) for x in argo_required)
        except Exception:
            argo_required_set = set(argo_required)
        before_cnt = len(argo_data_filtered)
        argo_data_filtered = argo_data_filtered[argo_data_filtered['Platform_number'].isin(argo_required_set)]
        if argo_data_filtered.empty:
            print(f"No Argo data left after filtering by platforms {sorted(list(argo_required_set))}.")
            return

    # 处理年月日过滤（year/month/day）
    # 注意：aggregated=True 时，按原逻辑在未提供月份时默认使用所有可用月份；
    #       aggregated=False 时，仅在提供了 month_required 时进行过滤。
    #       year_required/day_required 在两种模式下仅当提供时进行过滤。
    
    # Year
    argo_data_filtered['Year'] = pd.to_numeric(argo_data_filtered['Year'], errors='coerce')
    argo_data_filtered.dropna(subset=['Year'], inplace=True)
    argo_data_filtered['Year'] = argo_data_filtered['Year'].astype(int)

    # Month
    argo_data_filtered['Month'] = pd.to_numeric(argo_data_filtered['Month'], errors='coerce')
    argo_data_filtered.dropna(subset=['Month'], inplace=True)
    argo_data_filtered['Month'] = argo_data_filtered['Month'].astype(int)

    # Day
    argo_data_filtered['Day'] = pd.to_numeric(argo_data_filtered['Day'], errors='coerce')
    argo_data_filtered.dropna(subset=['Day'], inplace=True)
    argo_data_filtered['Day'] = argo_data_filtered['Day'].astype(int)

    if aggregated:
        # 若未指定月份，采用所有可用月份（保持原 plot_vertical_monthly 行为）
        if not month_required:
            all_available_months = sorted(argo_data_filtered['Month'].unique().tolist())
            if not all_available_months:
                print(f"No valid months found in data for eddy {ds_names}{no}.")
                return
            print(f"No months specified. Defaulting to all available months: {all_available_months}")
            month_required = all_available_months
        # 根据年月日过滤
        if month_required:
            argo_data_filtered = argo_data_filtered[argo_data_filtered['Month'].isin(month_required)].copy()
        if year_required:
            argo_data_filtered = argo_data_filtered[argo_data_filtered['Year'].isin(year_required)].copy()
        if day_required:
            argo_data_filtered = argo_data_filtered[argo_data_filtered['Day'].isin(day_required)].copy()
        if argo_data_filtered.empty:
            print(f"No data found for eddy {ds_names}{no} after applying filters: year={year_required}, month={month_required}, day={day_required}.")
            return
    else:
        # 非聚合模式：仅当提供了对应过滤条件才过滤
        if year_required:
            argo_data_filtered = argo_data_filtered[argo_data_filtered['Year'].isin(year_required)].copy()
            if argo_data_filtered.empty:
                print(f"No data found for eddy {ds_names}{no} in years {year_required} (per-platform mode).")
                return
        if month_required:
            argo_data_filtered = argo_data_filtered[argo_data_filtered['Month'].isin(month_required)].copy()
            if argo_data_filtered.empty:
                print(f"No data found for eddy {ds_names}{no} in months {month_required} (per-platform mode).")
                return
        if day_required:
            argo_data_filtered = argo_data_filtered[argo_data_filtered['Day'].isin(day_required)].copy()
            if argo_data_filtered.empty:
                print(f"No data found for eddy {ds_names}{no} in days {day_required} (per-platform mode).")
                return

    # =========================
    # 分支一：非聚合（原 plot_vertical 行为，新增月份/平台筛选）
    # =========================
    if not aggregated:
        for platform_id_val, platform_data in argo_data_filtered.groupby("Platform_number"):
            profile_num_agg = platform_data['Profile_number'].agg(['min', 'max'])
            min_profile_num = profile_num_agg['min']
            max_profile_num = profile_num_agg['max']

            num_variables = len(plot_variables)
            fig, axes = plt.subplots(1, num_variables, figsize=(10 * num_variables, 20), sharey=True)
            if num_variables == 1:
                axes = [axes]

            cmap = plt.cm.coolwarm
            profile_dates_for_title = []

            for i, var_name in enumerate(plot_variables):
                ax = axes[i]
                plot_variable_name = var_name
                is_do_panel = (_map_plot_variable_name(plot_variable_name) == 'DO')
                do_aux_layers = ('aou',) if (is_do_panel and overlay_aou_on_do) else tuple()
                db_variable_name = _map_plot_variable_name(plot_variable_name)

                if not _has_plottable_profile_variable(platform_data, db_variable_name):
                    ax.text(0.5, 0.5, f"Variable '{db_variable_name}'\nnot found in data.",
                            ha='center', va='center', transform=ax.transAxes, fontsize=16)
                    ax.set_title(f"Variable: {plot_variable_name}", fontsize=20)
                    continue

                for profile_num, rows in platform_data.groupby("Profile_number"):
                    if rows.empty:
                        continue

                    try:
                        current_profile_date = pd.Timestamp(year=int(rows.iloc[0]['Year']),
                                                            month=int(rows.iloc[0]['Month']),
                                                            day=int(rows.iloc[0]['Day']))
                    except (ValueError, TypeError):
                        continue

                    color_value_normalized = 0.5

                    if color_mode == 'distance':
                        if 'Longitude' in rows.iloc[0] and 'Latitude' in rows.iloc[0] and track_df is not None and not track_df.empty:
                            track_dates_converted = track_df['date'] if 'date' in track_df.columns else convert_date(track_df['time'])
                            idx_track_list = [j for j, td in enumerate(track_dates_converted) if hasattr(td, 'date') and td.date() == current_profile_date.date()]
                            if idx_track_list:
                                idx_track = idx_track_list[0]
                                center_lon = float(track_df.iloc[idx_track]['center_lon'])
                                center_lat = float(track_df.iloc[idx_track]['center_lat'])
                                radius = float(track_df.iloc[idx_track]['radius'])
                                if radius > 1e-6:
                                    dist_m = adaptive_distance_m(
                                        rows.iloc[0]['Longitude'], rows.iloc[0]['Latitude'],
                                        center_lon, center_lat
                                    )
                                    distance = dist_m / radius
                                    color_value_normalized = 1.0 - np.clip(distance, 0.0, 1.0)
                    elif color_mode == 'time':
                        if max_profile_num > min_profile_num:
                            color_value_normalized = (profile_num - min_profile_num) / (max_profile_num - min_profile_num)
                        else:
                            color_value_normalized = 0.0

                    color = cmap(np.clip(color_value_normalized, 0.0, 1.0))
                    plotted = _plot_single_argo_profile_line(
                        ax,
                        rows,
                        plot_variable_name,
                        color,
                        remove_outliers=remove_outliers,
                        show_normal_scatter=plot_normal_scatter,
                        do_aux_layers=do_aux_layers,
                        alpha=0.7,
                    )
                    if plotted and i == 0:
                        profile_dates_for_title.append(current_profile_date)

                # 子图属性
                _apply_vertical_profile_axis_style(ax, plot_variable_name)

                if show_colorbar:
                    norm_for_cbar = Normalize(vmin=0, vmax=1)
                    scalar_mappable = ScalarMappable(cmap=cmap, norm=norm_for_cbar)
                    scalar_mappable.set_array([])
                    cbar = plt.colorbar(scalar_mappable, ax=ax, orientation='vertical', fraction=0.046, pad=0.04)
                    if color_mode == 'distance':
                        current_ticks = cbar.get_ticks()
                        new_tick_labels = [f"{1.0 - t:.1f}" for t in current_ticks]
                        cbar.set_ticks(current_ticks)
                        cbar.set_ticklabels(new_tick_labels)
                        cbar.set_label('Normalized Distance from Eddy Center (0=center, 1=edge)', fontsize=14)
                    elif color_mode == 'time':
                        cbar.set_label(f'Normalized Profile Sequence (Range: {min_profile_num} to {max_profile_num})', fontsize=14)

            # 整体属性
            if not profile_dates_for_title:
                plt.close(fig)
                continue

            date_start_platform = min(profile_dates_for_title)
            date_end_platform = max(profile_dates_for_title)

            axes[0].set_ylabel("Depth/m", fontsize=20)
            axes[0].tick_params(axis='y', labelsize=16)
            axes[0].invert_yaxis()

            fig.suptitle(f"{ds_names}{no}, Platform: {int(platform_id_val)}, {date_start_platform.date()}~{date_end_platform.date()}", fontsize=24, y=0.95)
            plt.tight_layout(rect=[0, 0, 1, 0.93])

            if save_fig:
                region_slug = _current_region_key()
                output_dir = _shared_output_dir("plot_vertical_profiles", region_slug)
                output_dir.mkdir(parents=True, exist_ok=True)
                save_path = output_dir / f"{ds_names}{no}_Platform_{int(platform_id_val)}.png"
                plt.savefig(save_path, dpi=300, bbox_inches='tight')
                print(f"Figure saved to: {save_path}")

            if show_fig:
                plt.show()
            plt.close(fig)
        return

    # =========================
    # 分支二：聚合（等价于原 plot_vertical_monthly）
    # =========================
    # 收集各剖面的基本信息
    profiles_to_plot = []
    try:
        argo_data_filtered['date_ts'] = pd.to_datetime(argo_data_filtered[['Year', 'Month', 'Day']])
    except (ValueError, TypeError) as e:
        print(f"Could not create timestamps for all rows due to invalid date components: {e}.")
        argo_data_filtered = argo_data_filtered.dropna(subset=['Year', 'Month', 'Day'])
        argo_data_filtered['date_ts'] = pd.to_datetime(argo_data_filtered[['Year', 'Month', 'Day']])

    for _, profile_rows in argo_data_filtered.groupby(['Platform_number', 'Profile_number']):
        if not profile_rows.empty:
            profiles_to_plot.append({
                'rows': profile_rows,
                'date': profile_rows.iloc[0]['date_ts'],
                'lon': profile_rows.iloc[0]['Longitude'],
                'lat': profile_rows.iloc[0]['Latitude']
            })

    if not profiles_to_plot:
        print(f"No data found for eddy {ds_names}{no} in months {month_required}.")
        return

    all_dates = [p['date'] for p in profiles_to_plot]
    min_time_for_norm, max_time_for_norm = (min(all_dates), max(all_dates)) if all_dates else (None, None)

    num_variables = len(plot_variables)
    fig, axes = plt.subplots(1, num_variables, figsize=(10 * num_variables, 20), sharey=True)
    if num_variables == 1:
        axes = [axes]

    cmap = plt.cm.coolwarm

    for i, var_name in enumerate(plot_variables):
        ax = axes[i]
        plot_variable_name = var_name
        is_do_panel = (_map_plot_variable_name(plot_variable_name) == 'DO')
        do_aux_layers = ('aou',) if (is_do_panel and overlay_aou_on_do) else tuple()
        db_variable_name = _map_plot_variable_name(plot_variable_name)

        if not _has_plottable_profile_variable(argo_data_filtered, db_variable_name):
            ax.text(0.5, 0.5, f"Variable '{db_variable_name}'\nnot found in data.",
                    ha='center', va='center', transform=ax.transAxes, fontsize=16)
            ax.set_xlabel(plot_variable_name, fontsize=20)
            ax.grid(True)
            continue

        for profile_info in profiles_to_plot:
            rows = profile_info['rows']
            current_date = profile_info['date']
            color_value_normalized = 0.5

            if color_mode == 'distance':
                if track_df is not None and not track_df.empty and 'lon' in profile_info and 'lat' in profile_info:
                    track_dates_converted = track_df['date'] if 'date' in track_df.columns else convert_date(track_df['time'])
                    idx_track_list = [j for j, td in enumerate(track_dates_converted) if hasattr(td, 'date') and td.date() == current_date.date()]
                    if idx_track_list:
                        idx_track = idx_track_list[0]
                        center_lon = float(track_df.iloc[idx_track]['center_lon'])
                        center_lat = float(track_df.iloc[idx_track]['center_lat'])
                        radius = float(track_df.iloc[idx_track]['radius'])
                        if radius > 1e-6:
                            dist_m = adaptive_distance_m(
                                profile_info['lon'], profile_info['lat'],
                                center_lon, center_lat
                            )
                            distance = dist_m / radius
                            color_value_normalized = 1.0 - np.clip(distance, 0.0, 1.0)
            elif color_mode == 'time':
                if min_time_for_norm and max_time_for_norm and max_time_for_norm > min_time_for_norm:
                    total_delta = (max_time_for_norm - min_time_for_norm).total_seconds()
                    current_delta = (current_date - min_time_for_norm).total_seconds()
                    color_value_normalized = current_delta / total_delta if total_delta > 0 else 0.0
                else:
                    color_value_normalized = 0.0

            color = cmap(np.clip(color_value_normalized, 0.0, 1.0))
            _plot_single_argo_profile_line(
                ax,
                rows,
                plot_variable_name,
                color,
                remove_outliers=remove_outliers,
                show_normal_scatter=plot_normal_scatter,
                do_aux_layers=do_aux_layers,
                alpha=0.7,
            )

        # 子图属性
        _apply_vertical_profile_axis_style(ax, plot_variable_name)

        if show_colorbar:
            norm = Normalize(vmin=0, vmax=1)
            sm = ScalarMappable(cmap=cmap, norm=norm)
            sm.set_array([])
            cbar = plt.colorbar(sm, ax=ax, orientation='vertical', fraction=0.046, pad=0.04)
            if color_mode == 'distance':
                cbar.set_ticks([0, 0.5, 1])
                cbar.set_ticklabels(['1.0', '0.5', '0.0'])
                cbar.set_label('Normalized Distance (0=center, 1=edge)', fontsize=14)
            elif color_mode == 'time' and min_time_for_norm and max_time_for_norm:
                cbar.set_label(f'Normalized Time\n({min_time_for_norm.strftime("%Y-%m-%d")} to {max_time_for_norm.strftime("%Y-%m-%d")})', fontsize=12)

    # 整体属性
    axes[0].set_ylabel("Depth/m", fontsize=20)
    axes[0].tick_params(axis='y', labelsize=16)
    axes[0].invert_yaxis()

    month_str = "All" if month_required and len(month_required) > 6 else (", ".join(map(str, month_required)) if month_required else "All")
    date_range_str = f"{min_time_for_norm.date()}~{max_time_for_norm.date()}" if all_dates else "No date range"
    fig.suptitle(f"{ds_names}{no}, Months: {month_str}, Data: {date_range_str}", fontsize=24, y=0.95)
    plt.tight_layout(rect=[0, 0, 1, 0.93])

    if save_fig:
        region_slug = _current_region_key()
        output_dir = _shared_output_dir("plot_vertical_monthly_aggregated", region_slug)
        output_dir.mkdir(parents=True, exist_ok=True)
        month_suffix = "all" if not month_required or (month_required and len(month_required) > 6) else "_".join(map(str, month_required))
        filename = f"{ds_names}{no}_months_{month_suffix}_aggregated.png"
        save_path = output_dir / filename
        plt.savefig(save_path, dpi=300, bbox_inches='tight')
        print(f"Figure saved to: {save_path}")

    if show_fig:
        plt.show()
    plt.close(fig)
        
def plot_relative_position(
    DS: list,
    no: int,
    show_fig: bool = False,
    save_fig: bool = False,
    color_mode: str = 'distance',
    show_colorbar: bool = False,
    aggregated: bool = False,
    argo_required: list | None = None,
    year_required: list | None = None,
    month_required: list | None = None,
    day_required: list | None = None
):
    '''
    根据涡旋轨迹和浮标数据，绘制浮标在单位圆涡旋中的相对位置分布图。

    计算剖面代表点在单位圆涡旋中的相对位置并以散点标注（中心×、单位圆圈），横纵轴刻度兼含
    真实经纬度与相对坐标，可选显示颜色条并保存/显示图片，支持按月份与平台编号筛选。颜色模式
    'distance' 按距离中心归一化（0=中心，1=边缘），'time' 按时间顺序归一化。

    参数:
        - DS (list): 涡旋轨迹数据集。
        - no (int): 涡旋编号。
        - show_fig (bool): 是否显示图片，默认 False。
        - save_fig (bool): 是否保存图片，默认 False。
        - color_mode (str): 'distance' 或 'time'，默认 'distance'。
        - show_colorbar (bool): 是否显示颜色条，默认 False。
        - aggregated (bool): 是否聚合所有平台于一图，默认 False。
        - argo_required (list | None): 平台筛选；None 表示不过滤；传入平台编号列表时仅保留指定平台。
        - year_required (list | None): 年份筛选；None 表示不过滤。
        - month_required (list | None): 月份筛选；聚合模式 None 表示使用所有可用月份；逐平台模式 None 表示不过滤。
        - day_required (list | None): 日期筛选（按日 1-31）；None 表示不过滤。
    输出:
        - 逐平台图（save_fig=True, aggregated=False）：`plot_outputs/shared/<region>/plot_relative_position/{数据集}{编号}RP{平台号}.png`
        - 聚合图（save_fig=True, aggregated=True）：`plot_outputs/shared/<region>/plot_relative_position_monthly_aggregated/{数据集}{编号}_RP_months_{月份}_aggregated.png`

    说明:
        模式差异:

            - aggregated=False（逐平台）：对每个平台分别出图，点内数字为该平台内部剖面时序（从 1 递增）。
            - aggregated=True（聚合）：所有平台代表点聚合到一图，点内数字为相对所选月份范围起始日的累积天数（如数据自 7 月 29 日起，则该日为 29、7 月 30 日为 30、8 月 1 日为 32）。
    '''
    try:
        track_df, ds_names, ds_source_for_filter = _resolve_track_context(DS, no, include_contours=False)
    except Exception as exc:
        print(f"  - Error: {exc}")
        return

    argo_data_filtered = filtered_float_data(ds_source_for_filter, no, track=track_df)
    ds_names = ds_names.upper()

    if argo_data_filtered.empty:
        print(f"No Argo data found for eddy {ds_names}{no} to plot relative positions.")
        return

    # 平台筛选（与 plot_vertical 一致）
    if argo_required is not None:
        try:
            argo_required_set = set(int(x) for x in argo_required)
        except Exception:
            argo_required_set = set(argo_required)
        argo_data_filtered = argo_data_filtered[argo_data_filtered['Platform_number'].isin(argo_required_set)]
        if argo_data_filtered.empty:
            print(f"No Argo data left after filtering by platforms {sorted(list(argo_required_set))}.")
            return

    # 年月日筛选（与 plot_vertical 一致）
    argo_data_filtered['Year'] = pd.to_numeric(argo_data_filtered['Year'], errors='coerce')
    argo_data_filtered.dropna(subset=['Year'], inplace=True)
    argo_data_filtered['Year'] = argo_data_filtered['Year'].astype(int)

    argo_data_filtered['Month'] = pd.to_numeric(argo_data_filtered['Month'], errors='coerce')
    argo_data_filtered.dropna(subset=['Month'], inplace=True)
    argo_data_filtered['Month'] = argo_data_filtered['Month'].astype(int)

    argo_data_filtered['Day'] = pd.to_numeric(argo_data_filtered['Day'], errors='coerce')
    argo_data_filtered.dropna(subset=['Day'], inplace=True)
    argo_data_filtered['Day'] = argo_data_filtered['Day'].astype(int)

    if aggregated:
        if not month_required:
            all_months = sorted(argo_data_filtered['Month'].unique().tolist())
            if not all_months:
                print(f"No valid months found in data for eddy {ds_names}{no}.")
                return
            print(f"No months specified for relative position plot. Defaulting to all available months: {all_months}")
            month_required = all_months
        # 按提供的年/月/日过滤（月份在聚合模式下为必有列表）
        if month_required:
            argo_data_filtered = argo_data_filtered[argo_data_filtered['Month'].isin(month_required)].copy()
        if year_required:
            argo_data_filtered = argo_data_filtered[argo_data_filtered['Year'].isin(year_required)].copy()
        if day_required:
            argo_data_filtered = argo_data_filtered[argo_data_filtered['Day'].isin(day_required)].copy()
        if argo_data_filtered.empty:
            print(f"No data found for eddy {ds_names}{no} after applying filters: year={year_required}, month={month_required}, day={day_required}.")
            return
    else:
        if year_required:
            argo_data_filtered = argo_data_filtered[argo_data_filtered['Year'].isin(year_required)].copy()
            if argo_data_filtered.empty:
                print(f"No data found for eddy {ds_names}{no} in years {year_required} (per-platform mode).")
                return
        if month_required:
            argo_data_filtered = argo_data_filtered[argo_data_filtered['Month'].isin(month_required)].copy()
            if argo_data_filtered.empty:
                print(f"No data found for eddy {ds_names}{no} in months {month_required} (per-platform mode).")
                return
        if day_required:
            argo_data_filtered = argo_data_filtered[argo_data_filtered['Day'].isin(day_required)].copy()
            if argo_data_filtered.empty:
                print(f"No data found for eddy {ds_names}{no} in days {day_required} (per-platform mode).")
                return

    # =========================
    # 分支一：逐平台模式（aggregated=False）
    # =========================
    if not aggregated:
        for platform_id_val, platform_data in argo_data_filtered.groupby("Platform_number"):
        # 获取每个剖面的第一行数据作为代表点
        # 原文: needed_data = platform.groupby("Profile_number").apply(lambda group: group.iloc[0])
        #       needed_data.index.name = None
        # 使用 .first() 更简洁，并且可以直接用 Profile_number 作为索引或后续处理
            profile_first_rows = platform_data.groupby("Profile_number").first().reset_index() # reset_index 使 Profile_number 成为列

            if profile_first_rows.empty:
            # print(f"No profile data for platform {platform_id_val}.")
                continue

            # 准备每个剖面点对应的涡旋轨迹数据
            points_for_this_platform = []
            track_info_for_this_platform = [] # 用于计算该平台下的平均涡旋参数
        
            track_dates_converted = track_df['date'] if 'date' in track_df.columns else convert_date(track_df['time'])

            for i, p_row in profile_first_rows.iterrows(): # i 将用作顺序编号
                try:
                    current_date_profile = pd.Timestamp(year=int(p_row['Year']),
                                                        month=int(p_row['Month']),
                                                        day=int(p_row['Day']))
                except (ValueError, TypeError):
                    # print(f"Skipping profile {p_row.get('Profile_number')} for platform {platform_id_val} due to invalid date.")
                    continue

                center_lon, center_lat, radius = None, None, None
                if track_df is not None and not track_df.empty:
                    matches = [k for k, td in enumerate(track_dates_converted) if hasattr(td, 'date') and td.date() == current_date_profile.date()]
                    if matches:
                        idx_track = matches[0]
                        center_lon = float(track_df.iloc[idx_track]['center_lon'])
                        center_lat = float(track_df.iloc[idx_track]['center_lat'])
                        radius = float(track_df.iloc[idx_track]['radius'])
            
                if center_lon is not None and radius is not None and radius > 1e-6:
                    if 'Longitude' not in p_row or 'Latitude' not in p_row:
                        # print(f"Skipping point on {current_date_profile.date()} due to missing Longitude/Latitude.")
                        continue
                    scale = approximate_degree_length(center_lat)
                    dlon_deg = _minimal_lon_diff_deg(p_row['Longitude'], center_lon)
                    dx_m = dlon_deg * scale['meters_per_degree_lon']
                    dy_m = (p_row['Latitude'] - center_lat) * scale['meters_per_degree_lat']
                    rel_x = dx_m / radius
                    rel_y = dy_m / radius
                    points_for_this_platform.append({
                        'rel_x': rel_x, 'rel_y': rel_y, 'date': current_date_profile,
                        'profile_num_original_idx': p_row['Profile_number'], # 用于 'time' mode
                        'sequence_label': i + 1 # 平台内时序标签
                    })
                    track_info_for_this_platform.append([center_lon, center_lat, radius])
            # else: 涡旋数据不匹配或半径过小，则跳过该点
        
            if not points_for_this_platform:
            # print(f"No valid points with track data for platform {platform_id_val}.")
                continue

            fig, ax = plt.subplots(figsize=(30, 20)) # 原为 (30,20) 但相对位置图常用正方形，可考虑 (20,20) 或 (15,15)
            cmap = plt.cm.coolwarm

            # 准备 'time' 模式颜色归一化 (基于 Profile_number)
            min_prof_num_platform = min(p['profile_num_original_idx'] for p in points_for_this_platform)
            max_prof_num_platform = max(p['profile_num_original_idx'] for p in points_for_this_platform)

            for point_data in points_for_this_platform:
                rel_x = point_data['rel_x']
                rel_y = point_data['rel_y']
            
                color_value_normalized = 0.5
                if color_mode == 'distance':
                    distance = np.sqrt(rel_x**2 + rel_y**2)
                    color_value_normalized = 1.0 - np.clip(distance, 0.0, 1.0)
                elif color_mode == 'time':
                    if max_prof_num_platform > min_prof_num_platform:
                        color_value_normalized = (point_data['profile_num_original_idx'] - min_prof_num_platform) / \
                                                 (max_prof_num_platform - min_prof_num_platform)
                    else:
                        color_value_normalized = 0.0
            
                color = cmap(np.clip(color_value_normalized, 0.0, 1.0))
                ax.scatter(rel_x, rel_y, color=color, s=300, zorder=5)
                ax.text(rel_x, rel_y, str(point_data['sequence_label']), weight='bold', fontsize=9, color='black', ha='center', va='center', zorder=6)
            
            ax.plot(0, 0, marker='x', color='black', markersize=16, markeredgewidth=3, label='Eddy Center (Relative)', zorder=3)
            circle = plt.Circle((0, 0), 1, color='gray', fill=False, linestyle='--', linewidth=2, label='Unit Eddy Boundary', zorder=2)
            ax.add_patch(circle)
            ax.set_aspect('equal')

            # 标题和坐标轴标签
            platform_dates = [p['date'] for p in points_for_this_platform]
            date_start_platform = min(platform_dates)
            date_end_platform = max(platform_dates)
            ax.set_title(f"{ds_names}{no}, Platform: {int(platform_id_val)}, {date_start_platform.date()}~{date_end_platform.date()}, Points: {len(points_for_this_platform)}", fontsize=20)
            ax.set_xlabel('Relative X (Eddy Radii)', fontsize=20)
            ax.set_ylabel('Relative Y (Eddy Radii)', fontsize=20)
            ax.tick_params(axis='both', which='major', labelsize=16) # 使用tick_params统一设置

            # 设置坐标轴刻度 (与聚合版一致)
            if track_info_for_this_platform:
                mean_center_lon = np.mean([info[0] for info in track_info_for_this_platform])
                mean_center_lat = np.mean([info[1] for info in track_info_for_this_platform])
                mean_radius = np.mean([info[2] for info in track_info_for_this_platform])

                if not np.isnan(mean_center_lon) and not np.isnan(mean_center_lat) and not np.isnan(mean_radius) and mean_radius > 1e-6:
                    scale_mean = approximate_degree_length(mean_center_lat)
                    mean_deg_x = mean_radius / scale_mean['meters_per_degree_lon']
                    mean_deg_y = mean_radius / scale_mean['meters_per_degree_lat']
                
                    tick_locs = [-1, -0.5, 0, 0.5, 1] # 使用更详细的刻度
                
                    x_tick_labels = [f"{(mean_center_lon + tick_loc * mean_deg_x):.2f}°\n({tick_loc:.1f})" for tick_loc in tick_locs]
                    ax.set_xticks(tick_locs)
                    ax.set_xticklabels(x_tick_labels)

                    y_tick_labels = [f"{(mean_center_lat + tick_loc * mean_deg_y):.2f}°\n({tick_loc:.1f})" for tick_loc in tick_locs]
                    ax.set_yticks(tick_locs)
                    ax.set_yticklabels(y_tick_labels)
        
            ax.set_xlim([-1.25, 1.25])
            ax.set_ylim([-1.25, 1.25])
            # ax.legend(fontsize=14) # 原代码legend注释掉了，保持一致

            # 添加颜色条
            if show_colorbar:
                norm_for_cbar = Normalize(vmin=0, vmax=1)
                scalar_mappable = ScalarMappable(cmap=cmap, norm=norm_for_cbar)
                scalar_mappable.set_array([])
                cbar = plt.colorbar(scalar_mappable, ax=ax, orientation='vertical', fraction=0.046, pad=0.04)

                if color_mode == 'distance':
                    current_ticks = cbar.get_ticks()
                    new_tick_labels = [f"{1.0 - t:.1f}" for t in current_ticks]
                    cbar.set_ticks(current_ticks)
                    cbar.set_ticklabels(new_tick_labels)
                    cbar.set_label('Normalized Distance from Eddy Center (0=center, 1=edge)', fontsize=14)
                elif color_mode == 'time':
                    cbar.set_label(f'Normalized Profile Sequence (Platform Range: {min_prof_num_platform} to {max_prof_num_platform})', fontsize=14)
        
            if save_fig:
                region_slug = _current_region_key()
                output_dir = _shared_output_dir("plot_relative_position", region_slug)
                output_dir.mkdir(parents=True, exist_ok=True)
                save_path = output_dir / f"{ds_names}{no}RP{int(platform_id_val)}.png"
                plt.savefig(save_path, dpi=300, bbox_inches='tight')
                print(f"Figure saved to: {save_path}")
        
            if show_fig:
                plt.show()
            plt.close(fig)
        return

    # =========================
    # 分支二：聚合模式（aggregated=True）
    # =========================
    # 1) 收集聚合绘图所需点
    points_to_process = []  # {date, data_row}
    monthly_filtered_data = argo_data_filtered.copy()
    if month_required:
        monthly_filtered_data = monthly_filtered_data[monthly_filtered_data['Month'].isin(month_required)].copy()

    if not monthly_filtered_data.empty:
        # 获取每个剖面的第一行数据作为代表点
        profile_first_rows = monthly_filtered_data.groupby(["Platform_number", "Profile_number"]).first().reset_index()
        for _, p_row in profile_first_rows.iterrows():
            try:
                current_date_profile = pd.Timestamp(year=int(p_row['Year']), month=int(p_row['Month']), day=int(p_row['Day']))
            except (ValueError, TypeError):
                continue
            if not month_required or current_date_profile.month in month_required:
                points_to_process.append({'date': current_date_profile, 'data_row': p_row})

    if not points_to_process:
        print(f"No data found for eddy {ds_names}{no} in months {month_required}.")
        return

    # 2) 计算日期标签的参考起点（按所选月份的最小月的一号）
    min_plot_date_overall = min(p['date'] for p in points_to_process)
    ref_month = min(month_required) if month_required else min_plot_date_overall.month
    reference_start_date_for_labels = pd.Timestamp(year=min_plot_date_overall.year, month=ref_month, day=1)

    points_to_plot = []
    all_track_info_for_overall_mean = []
    all_profile_dates_for_title = []
    all_profile_timestamps_for_time_mode = []

    track_dates_converted = track_df['date'] if 'date' in track_df.columns else convert_date(track_df['time'])

    for point_info in points_to_process:
        current_date = point_info['date']
        p_row = point_info['data_row']

        day_label = (current_date - reference_start_date_for_labels).days + 1

        center_lon, center_lat, radius = None, None, None
        if track_df is not None and not track_df.empty:
            matches = [i for i, td in enumerate(track_dates_converted) if hasattr(td, 'date') and td.date() == current_date.date()]
            if matches:
                idx_track = matches[0]
                center_lon = float(track_df.iloc[idx_track]['center_lon'])
                center_lat = float(track_df.iloc[idx_track]['center_lat'])
                radius = float(track_df.iloc[idx_track]['radius'])

        if center_lon is not None and radius is not None and radius > 1e-6:
            if 'Longitude' not in p_row or 'Latitude' not in p_row:
                continue

            scale = approximate_degree_length(center_lat)
            dlon_deg = _minimal_lon_diff_deg(p_row['Longitude'], center_lon)
            dx_m = dlon_deg * scale['meters_per_degree_lon']
            dy_m = (p_row['Latitude'] - center_lat) * scale['meters_per_degree_lat']
            rel_x = dx_m / radius
            rel_y = dy_m / radius

            points_to_plot.append({'rel_x': rel_x, 'rel_y': rel_y, 'date': current_date, 'day_label': day_label})
            all_track_info_for_overall_mean.append([center_lon, center_lat, radius])
            all_profile_dates_for_title.append(current_date)
            if color_mode == 'time':
                all_profile_timestamps_for_time_mode.append(current_date)

    if not points_to_plot:
        print(f"No valid points with track data found for eddy {ds_names}{no} in months {month_required}.")
        return

    # 3) 确定时间归一化范围（仅用于 color_mode='time'）
    min_time_for_norm, max_time_for_norm = (None, None)
    if color_mode == 'time' and all_profile_timestamps_for_time_mode:
        min_time_for_norm = min(all_profile_timestamps_for_time_mode)
        max_time_for_norm = max(all_profile_timestamps_for_time_mode)
        if min_time_for_norm == max_time_for_norm and len(all_profile_timestamps_for_time_mode) > 1:
            max_time_for_norm = min_time_for_norm + pd.Timedelta(days=1)

    date_start_overall = min(all_profile_dates_for_title) if all_profile_dates_for_title else None
    date_end_overall = max(all_profile_dates_for_title) if all_profile_dates_for_title else None

    # 4) 绘图
    fig, ax = plt.subplots(figsize=(30, 20))
    cmap = plt.cm.coolwarm

    for point in points_to_plot:
        rel_x, rel_y = point['rel_x'], point['rel_y']
        current_date, day_label = point['date'], point['day_label']

        color_value_normalized = 0.5
        if color_mode == 'distance':
            distance_from_center = np.sqrt(rel_x**2 + rel_y**2)
            color_value_normalized = 1.0 - np.clip(distance_from_center, 0.0, 1.0)
        elif color_mode == 'time':
            if min_time_for_norm and max_time_for_norm and min_time_for_norm < max_time_for_norm:
                total_delta = (max_time_for_norm - min_time_for_norm).total_seconds()
                current_delta = (current_date - min_time_for_norm).total_seconds()
                color_value_normalized = (current_delta / total_delta) if total_delta > 0 else 0.0
            elif min_time_for_norm and max_time_for_norm and min_time_for_norm == max_time_for_norm:
                color_value_normalized = 0.0

        color = cmap(np.clip(color_value_normalized, 0.0, 1.0))
        ax.scatter(rel_x, rel_y, color=color, s=300, zorder=5)
        ax.text(rel_x, rel_y, str(day_label), weight='bold', fontsize=9, color='black', ha='center', va='center', zorder=6)

    # 中心与单位圆
    ax.plot(0, 0, marker='x', color='black', markersize=16, markeredgewidth=3, label='Eddy Center (Relative)', zorder=3)
    circle = plt.Circle((0, 0), 1, color='gray', fill=False, linestyle='--', linewidth=2, label='Unit Eddy Boundary', zorder=2)
    ax.add_patch(circle)
    ax.set_aspect('equal')

    # 标题与坐标轴
    month_str = "All" if (month_required and len(month_required) > 6) else ", ".join(map(str, month_required)) if month_required else "All"
    title_str = f"{ds_names}{no}, Months: {month_str}, Relative Positions"
    if date_start_overall and date_end_overall:
        title_str += f"\nData: {date_start_overall.date()}~{date_end_overall.date()}, Total Points: {len(points_to_plot)}"
    ax.set_title(title_str, fontsize=20)
    ax.set_xlabel('Relative X (Eddy Radii)', fontsize=20)
    ax.set_ylabel('Relative Y (Eddy Radii)', fontsize=20)
    ax.tick_params(axis='both', which='major', labelsize=16)

    # 地理刻度（均值）
    if all_track_info_for_overall_mean:
        mean_center_lon = np.mean([info[0] for info in all_track_info_for_overall_mean])
        mean_center_lat = np.mean([info[1] for info in all_track_info_for_overall_mean])
        mean_radius = np.mean([info[2] for info in all_track_info_for_overall_mean])
        if not np.isnan(mean_center_lon) and not np.isnan(mean_center_lat) and not np.isnan(mean_radius) and mean_radius > 1e-6:
            scale_mean = approximate_degree_length(mean_center_lat)
            mean_deg_x = mean_radius / scale_mean['meters_per_degree_lon']
            mean_deg_y = mean_radius / scale_mean['meters_per_degree_lat']
            tick_locs = [-1, -0.5, 0, 0.5, 1]
            x_tick_labels = [f"{(mean_center_lon + t * mean_deg_x):.2f}°\n({t})" for t in tick_locs]
            y_tick_labels = [f"{(mean_center_lat + t * mean_deg_y):.2f}°\n({t})" for t in tick_locs]
            ax.set_xticks(tick_locs)
            ax.set_xticklabels(x_tick_labels)
            ax.set_yticks(tick_locs)
            ax.set_yticklabels(y_tick_labels)

    ax.set_xlim([-1.25, 1.25])
    ax.set_ylim([-1.25, 1.25])

    # 颜色条
    if show_colorbar:
        norm_for_cbar = Normalize(vmin=0, vmax=1)
        scalar_mappable = ScalarMappable(cmap=cmap, norm=norm_for_cbar)
        scalar_mappable.set_array([])
        cbar = plt.colorbar(scalar_mappable, ax=ax, orientation='vertical', fraction=0.046, pad=0.04)
        if color_mode == 'distance':
            current_ticks = cbar.get_ticks()
            new_tick_labels = [f"{1.0 - t:.1f}" for t in current_ticks]
            cbar.set_ticks(current_ticks)
            cbar.set_ticklabels(new_tick_labels)
            cbar.set_label('Normalized Distance from Eddy Center (0=center, 1=edge)', fontsize=14)
        elif color_mode == 'time':
            if min_time_for_norm and max_time_for_norm:
                cbar.set_label(f'Normalized Time ({min_time_for_norm.strftime("%Y-%m-%d")} to {max_time_for_norm.strftime("%Y-%m-%d")})', fontsize=14)
            else:
                cbar.set_label('Normalized Time', fontsize=14)

    # 保存/显示
    if save_fig:
        region_slug = _current_region_key()
        output_dir = _shared_output_dir("plot_relative_position_monthly_aggregated", region_slug)
        output_dir.mkdir(parents=True, exist_ok=True)
        month_suffix = "all" if not month_required or (month_required and len(month_required) > 6) else "_".join(map(str, month_required))
        base_filename = f"{ds_names}{no}_RP_months_{month_suffix}_aggregated.png"
        save_path = output_dir / base_filename
        plt.savefig(save_path, dpi=300, bbox_inches='tight')
        print(f"Figure saved to: {save_path}")

    if show_fig:
        plt.show()
    plt.close(fig)

def plot_vertical_monthly(
    DS: list,
    no: int,
    month_required: list = None,
    show_fig: bool = False,
    save_fig: bool = False,
    color_mode: str = 'distance',
    variables: list = ['DO', 'Temp', 'Salinity'],
    show_colorbar: bool = False,
    remove_outliers: bool = True,
    argo_required: list | None = None
):
    '''
    兼容封装：保持旧接口，内部以 aggregated=True 调用统一的 plot_vertical。

    参数:
        - DS (list): 涡旋轨迹数据集。
        - no (int): 涡旋编号。
        - month_required (list | None): 月份筛选；None 时使用所有可用月份。
        - show_fig (bool): 是否显示图片，默认 False。
        - save_fig (bool): 是否保存图片，默认 False。
        - color_mode (str): 着色方式，默认 'distance'。
        - variables (list): 绘制变量，默认 ['DO', 'Temp', 'Salinity']。
        - show_colorbar (bool): 是否显示颜色条，默认 False。
        - remove_outliers (bool): 是否执行基础 QC，默认 True。
        - argo_required (list | None): 平台筛选；None 时不过滤。
    '''
    return plot_vertical(
        DS=DS,
        no=no,
        show_fig=show_fig,
        save_fig=save_fig,
        color_mode=color_mode,
        variables=variables,
        show_colorbar=show_colorbar,
        remove_outliers=remove_outliers,
        aggregated=True,
        month_required=month_required,
        argo_required=argo_required
    )
    
def plot_relative_position_monthly(
    DS: list,
    no: int,
    show_fig: bool = False,
    save_fig: bool = False,
    color_mode: str = 'distance',
    show_colorbar: bool = False,
    month_required: list | None = None,
    argo_required: list | None = None
):
    '''
    兼容封装：保持旧接口，内部以 aggregated=True 调用统一的 plot_relative_position。

    参数:
        - DS (list): 涡旋轨迹数据集。
        - no (int): 涡旋编号。
        - show_fig (bool): 是否显示图片，默认 False。
        - save_fig (bool): 是否保存图片，默认 False。
        - color_mode (str): 着色方式，默认 'distance'。
        - show_colorbar (bool): 是否显示颜色条，默认 False。
        - month_required (list | None): 月份筛选；None 时使用所有可用月份。
        - argo_required (list | None): 平台筛选；None 时不过滤。
    '''
    return plot_relative_position(
        DS=DS,
        no=no,
        show_fig=show_fig,
        save_fig=save_fig,
        color_mode=color_mode,
        show_colorbar=show_colorbar,
        aggregated=True,
        month_required=month_required,
        argo_required=argo_required,
    )

def get_glorys_filepath(date)-> str:
    '''
    根据给定日期返回对应的 GLORYS NetCDF 文件路径。

    构造目标日期的文件夹路径与文件名模式并查找匹配文件：恰有一个匹配时返回其完整路径，否则抛异常。

    参数:
        - date (pd.Timestamp): 由 convert_date 得出的需要查找的日期。

    返回:
        - str: 匹配的 GLORYS NetCDF 文件完整路径。

    说明:
        - 抛出 RuntimeError：找到多个匹配文件。
        - 抛出 FileNotFoundError：没有找到匹配文件。
    '''
    nc_path = os.path.join(Glorys_path, date.strftime("%Y"), date.strftime("%m"), 'mercatorglorys12v1_gl12_mean_')
    nc_path += date.strftime("%Y%m%d")
    matches = glob.glob(nc_path + '_R*')
    if matches:
        if len(matches) > 1:
            raise RuntimeError(f"Multiple matching files found for {date.strftime('%Y-%m-%d')}: {matches}")
        nc_path = matches[0]
        #print("Matching file:", nc_path)
        return nc_path
    else:
        raise FileNotFoundError(f"No matching file found for {date.strftime('%Y-%m-%d')}.")
        
def print_glorys_variable(nc_path: str):
    '''
    打印指定路径 NetCDF 文件中的所有变量名称、标准名称、维度和形状信息。

    参数:
        - nc_path (str): NetCDF 文件路径。
    '''
    nc_file = Dataset(nc_path, 'r')
    
    with Dataset(nc_path, 'r') as nc_file:
        for var in nc_file.variables:
            print(var, nc_file.variables[var].standard_name)
            print(nc_file.variables[var].dimensions, nc_file.variables[var].shape)

def find_track_glorys_filepath(DS: list, no: int) -> dict:
    '''
    根据涡旋编号在 GLORYS 数据集中查找对应的轨迹文件路径。

    参数:
        - DS (list): 涡旋轨迹数据集。
        - no (int): 涡旋编号。

    返回:
        - dict: {date: glorys_filepath} 字典；未找到对应轨迹或文件路径时返回空字典。
    '''
    try:
        track_df, _, _ = _resolve_track_context(DS, no, include_contours=False)
    except Exception as exc:
        print(f"未找到涡旋 {no} 的轨迹数据：{exc}")
        return {}
    
    glorys_filepaths_dict = {}
    for _, track_point in track_df.iterrows():
        try:
            date = convert_date(track_point['time'])  
            glorys_filepath = get_glorys_filepath(date)
            glorys_filepaths_dict[date] = glorys_filepath          
        except (RuntimeError, FileNotFoundError) as e:
            print(f"为涡旋 {no} 在日期 {track_point.get('time', '未知')} (转换后: {date if 'date' in locals() else '未知'}) 查找 GLORYS 文件时出错: {e}")
        except IndexError:
            print(f"涡旋 {no} 的轨迹点数据格式不正确: {track_point}")
        except Exception as e: # 捕获其他可能的 convert_date 或 get_glorys_filepath 异常
            print(f"处理涡旋 {no} 在日期 {track_point.get('time', '未知')} 时发生未知错误: {e}")

    if not glorys_filepaths_dict:
        print(f"未找到涡旋 {no} 的 GLORYS 文件。")
        return {}
        
    return glorys_filepaths_dict

class LineDrawer:
    """
    一个用于在matplotlib图像上通过点击两点来交互式绘制直线的辅助类。
    新增了撤回功能：按 'z' 键可以撤回上一个点击的点或已绘制的直线。
    """
    def __init__(self, ax, legend_loc: str = 'best'):
        self.ax = ax
        self.legend_loc = legend_loc
        self.points = []  # 用于存储用户点击的坐标
        self.lines = []   # 用于存储已绘制的直线对象
        self.markers = [] # 用于存储用户点击时产生的临时视觉标记
        # 将键盘事件与 onkey 方法连接
        self.cid_key = self.ax.figure.canvas.mpl_connect('key_press_event', self.onkey)
        
        print("\n--- 交互模式已激活 (稳定版) ---")
        print("请在图像上点击两个点来定义一条新的直线。")
        print("提示：按 'z' 键可以撤回上一个点或刚绘制的直线。")

    def onclick(self, event):
        """鼠标点击事件的处理函数。"""
        # 确保点击事件发生在指定的坐标轴内
        if event.inaxes != self.ax: return
        
        # 记录点击的坐标点
        self.points.append((event.xdata, event.ydata))
        # 在点击位置添加一个临时的"+"标记，给予用户视觉反馈
        marker = self.ax.plot(event.xdata, event.ydata, 'm+', markersize=12, markeredgewidth=2)
        self.markers.extend(marker)
        # 触发一次完整的重绘，确保标记点被永久画上
        self.ax.figure.canvas.draw_idle()

        # 当记录的点达到两个时，开始绘制直线
        if len(self.points) == 2:
            self.draw_line()

    def onkey(self, event):
        """键盘按键事件的处理函数，用于实现撤回功能。"""
        if event.key == 'z':
            # 如果当前正在定义一条线（已点击了一个点）
            if self.markers:
                # 移除记录的点和对应的标记
                self.points.pop()
                last_marker = self.markers.pop()
                last_marker.remove()
                self.ax.figure.canvas.draw()
                print("上一个点已撤回。")
            # 如果要撤回一条已完成的线
            elif self.lines:
                # 移除上一条绘制的直线
                last_line = self.lines.pop()
                last_line.remove()
                self.ax.legend(loc=self.legend_loc) # 更新图例
                self.ax.figure.canvas.draw_idle()
                print("上一条线已撤回。")
            else:
                print("没有可撤回的操作。")

    def draw_line(self):
        """根据记录的两个点计算并绘制直线。"""
        x1, y1 = self.points[0]
        x2, y2 = self.points[1]

        # 计算直线方程 y = kx + b
        if abs(x2 - x1) < 1e-6: # 处理斜率不存在的垂直线情况
            k, b, eq_text = np.inf, np.nan, f"x = {x1:.2f}"
        else:
            k = (y2 - y1) / (x2 - x1)
            b = y1 - k * x1
            eq_text = f"y = {k:.4f}x{b:+.4f}"

        # 根据坐标轴的当前范围，计算直线的端点并绘制
        x_vals = np.array(self.ax.get_xlim())
        if np.isinf(k): # 绘制垂直线
             y_vals = np.array(self.ax.get_ylim())
             line = self.ax.plot([x1, x1], y_vals, 'purple', linestyle='--', linewidth=2, label=f'Interactive: {eq_text}')
        else: # 绘制普通斜率的直线
            y_vals = k * x_vals + b
            line = self.ax.plot(x_vals, y_vals, 'purple', linestyle='--', linewidth=2, label=f'Interactive: {eq_text}')
        
        self.lines.extend(line)
        
        # 清理本次操作的临时标记点和坐标
        for marker in self.markers:
            marker.remove()
        self.markers.clear()
        self.points.clear()

        # 更新图例并重绘图像
        self.ax.legend(loc=self.legend_loc)
        self.ax.figure.canvas.draw_idle()
        print("\n准备就绪，可继续点击绘制下一条直线。")

def _reduce_argo_profiles_by_anomaly(
    rows: pd.DataFrame,
    detection_config: DetectionConfig | None = None,
    min_depth: float | None = None,
) -> pd.DataFrame:
    """按 DetectionConfig 筛选异常，并为每个 Profile 保留最强代表点。"""
    if rows is None or rows.empty:
        return pd.DataFrame()

    overrides = {}
    if min_depth is not None:
        overrides['anomaly_min_depth'] = float(min_depth)
    cfg = _resolve_detection_config(detection_config, **overrides)

    deltas = calculate_delta_do(
        rows,
        detection_config=cfg,
        include_aou=(cfg.method != 'do'),
        remove_outliers=True,
        verbose=False,
    )
    if deltas is None or deltas.empty:
        return pd.DataFrame()

    deltas = _keep_best_anomaly_per_profile(deltas, cfg)
    return deltas.rename(columns={
        'depth': 'Depth',
        'do_value': 'DO',
        'temperature_value': 'Temperature',
        'salinity_value': 'Salinity',
    })

def _compute_horizontal_glorys_field(
    variable: str,
    needed_depth: float | int,
    loader,
) -> tuple[np.ndarray, np.ndarray, np.ndarray, np.ndarray | np.ma.MaskedArray]:
    """共享的水平背景场提取逻辑。loader(variables, depth) 由上层注入。"""
    if variable == 'vorticity':
        glorys_lon, glorys_lat, glorys_depth, glorys_vars = loader(['u', 'v'], needed_depth)
        zeta, f = calculate_vorticity(glorys_lon, glorys_lat, glorys_vars['u'], glorys_vars['v'])
        field = zeta / f
        return glorys_lon, glorys_lat, glorys_depth, field

    glorys_lon, glorys_lat, glorys_depth, glorys_vars = loader([variable], needed_depth)
    key_map = {
        'so': 'salinity',
        'uo': 'u',
        'vo': 'v',
        'zos': 'ssh',
        'mlotst': 'mlt',
    }
    var_key = key_map.get(variable, variable)
    if var_key not in glorys_vars:
        raise KeyError(
            f"Variable '{variable}' resolved to '{var_key}', but available keys are "
            f"{list(glorys_vars.keys())}"
        )
    return glorys_lon, glorys_lat, glorys_depth, glorys_vars[var_key]

def _style_horizontal_colorbar(
    pc,
    cbar,
    variable: str,
    cbar_label_fs: float,
):
    """统一水平图色标文本与默认范围设置。"""
    if variable == 'vorticity':
        cbar.set_label(r'$\zeta/f$', fontsize=cbar_label_fs)
        pc.set_clim(-0.7, 0.7)
    elif variable == 'thetao':
        cbar.set_label('Temperature (°C)', fontsize=cbar_label_fs)
    elif variable in ['so', 'salinity']:
        cbar.set_label('Salinity (psu)', fontsize=cbar_label_fs)
    elif variable in ['u', 'uo']:
        cbar.set_label('Zonal Velocity (m/s)', fontsize=cbar_label_fs)
    elif variable in ['v', 'vo']:
        cbar.set_label('Meridional Velocity (m/s)', fontsize=cbar_label_fs)
    elif variable in ['ssh', 'zos']:
        cbar.set_label('Sea Surface Height (m)', fontsize=cbar_label_fs)
    else:
        cbar.set_label(variable, fontsize=cbar_label_fs)

def _plot_horizontal_profile_lines(
    ax,
    k: float | list[float] | None,
    b: float | list[float] | None,
    lon_min: float,
    lon_max: float,
    line_lw: float = 2.0,
):
    """在水平图上绘制一条或多条剖面线 y=kx+b。"""
    if k is None or b is None:
        return

    k_list = [k] if isinstance(k, (int, float)) else k
    b_list = [b] if isinstance(b, (int, float)) else b
    if len(k_list) != len(b_list):
        raise ValueError("The lists for k and b must have the same length.")

    line_x = np.array([lon_min, lon_max], dtype=float)
    for i, (k_val, b_val) in enumerate(zip(k_list, b_list)):
        line_y = k_val * line_x + b_val
        ax.plot(
            line_x,
            line_y,
            color='purple',
            linestyle='-',
            linewidth=line_lw,
            label=f'Profile Line {i+1}: y={k_val:.2f}x{b_val:+.2f}',
        )

def plot_track_horizontal_glorys(DS: list, no: int, needed_date: str | pd.Timestamp, variable: str = 'vorticity',
                                 show_fig: bool = True, save_fig: bool = False,
                                 k: float | list[float] | None = None, b: float | list[float] | None = None,
                                 needed_depth: float | int = 0, inline_mode: bool = True,
                                 argo_detection_config: DetectionConfig | None = None,
                                 argo_min_depth: float | None = None,
                                 verbose: bool = True):
    '''
    绘制指定涡旋在指定日期的 GLORYS 水平快照，并叠加同日 Argo 异常点。

    track 场景下的水平可视化入口：背景场取目标日期、目标深度的 GLORYS 子区域，叠加涡旋轨迹、当日有效
    半径与轮廓，并对同日匹配 Argo 按 argo_detection_config 识别异常后仅保留每个 Profile 的最强异常代表点。

    参数:
        - DS (list | str | tuple | dict): 轨迹数据输入，常见可用值 'acs'/'acl'/'cs'/'cl'，或 legacy 轨迹列表结构。
        - no (int): 轨迹编号（track id）。
        - needed_date (str | pd.Timestamp): 目标日期，支持 'YYYY-MM-DD' 字符串或 pandas 时间戳，按日精度匹配轨迹时间。
        - variable (str): 背景变量名，常用 'vorticity'/'thetao'/'so'/'u'/'v'/'ssh'，默认 'vorticity'。
        - show_fig (bool): 是否显示图像，默认 True。
        - save_fig (bool): 是否保存图像到输出目录，默认 False。
        - k (float | list[float] | None): 剖面线斜率，满足 y = kx + b；传入等长列表可叠加多条线，仅传其一或长度不一致会报错。
        - b (float | list[float] | None): 剖面线截距，与 k 配套使用。
        - needed_depth (float | int): GLORYS 读取深度（m），默认 0（表层）。
        - inline_mode (bool): 是否使用内联静态模式，默认 True；详见“说明”。
        - argo_detection_config (DetectionConfig | None): Argo 异常点筛选配置；None 时使用默认。
        - argo_min_depth (float | None): Argo 筛选最小深度阈值（m），None 时回退全局配置。
        - verbose (bool): 是否打印保存路径与提示信息，默认 True。

    返回:
        - tuple: (fig, ax)，便于调用侧继续叠加绘图或交互操作。

    说明:
        显示模式:

            - inline_mode=True（默认）：静态出图模式，资源占用更可控，函数结束会关闭 figure 释放内存，适合脚本批量出图。
            - inline_mode=False：交互模式，保留图窗句柄，可配合 LineDrawer 手动点选剖面线。

        异常:

            - 抛出 ValueError：needed_date 无法解析，或该日期不在目标轨迹时间范围内。
    '''
    argo_detection_config = _resolve_detection_config(
        argo_detection_config,
        anomaly_min_depth=argo_min_depth,
    )

    # 若从 widget 切回 inline，清理遗留的交互式 figure，避免被 inline 后端顺带渲染
    if inline_mode:
        plt.close('all')

    # 统一两种模式的视觉参数
    figsize = (12, 10)
    title_fs, label_fs, tick_fs, legend_fs = 16, 14, 12, 10
    cbar_label_fs, cbar_tick_fs = 12, 10
    argo_text_fs = 6
    track_lw, contour_lw, circle_lw, line_lw = 1.0, 1.0, 1.0, 2.0
    cbar_pad = 0.18  # 增大两条水平色标间距

    # 加载轨迹（DataFrame）并准备常用列
    try:
        track_df, ds_name, ds_source_for_filter = _resolve_track_context(DS, no, include_contours=True)
    except Exception as exc:
        print(f"  - Error: {exc}")
        return
    if track_df.empty:
        print(f"  - Track for eddy {no} is empty.")
        return
    dates = track_df['date'] if 'date' in track_df.columns else convert_date(track_df['time'])
    dates = pd.to_datetime(dates, errors='coerce')
    try:
        target_ts = pd.Timestamp(needed_date).normalize()
    except Exception as exc:
        raise ValueError(f"needed_date={needed_date!r} is not a valid date.") from exc

    same_day_idx = np.nonzero(dates.dt.normalize().to_numpy() == target_ts.to_datetime64())[0]
    if same_day_idx.size == 0:
        raise ValueError(f"Date {target_ts.strftime('%Y-%m-%d')} not found in track {no}.")
    needed_idx = int(same_day_idx[0])
    needed_date = pd.Timestamp(dates.iloc[needed_idx])
    center_lon_arr = track_df['center_lon'].to_numpy()
    center_lat_arr = track_df['center_lat'].to_numpy()
    radius_arr = track_df['radius'].to_numpy()
    plot_anchor_lon = float(center_lon_arr[needed_idx])
    center_lon_plot = plot_anchor_lon + _minimal_lon_diff_deg(np.asarray(center_lon_arr, dtype=float), plot_anchor_lon)
    # 当前时刻（needed_idx）的轮廓坐标
    curr_contour_lon = track_df.iloc[needed_idx]['contour_lon']
    curr_contour_lat = track_df.iloc[needed_idx]['contour_lat']
    curr_contour_lon = np.asarray(curr_contour_lon, dtype=float).ravel()
    curr_contour_lat = np.asarray(curr_contour_lat, dtype=float).ravel()
    contour_now_valid = (
        np.isfinite(curr_contour_lon)
        & np.isfinite(curr_contour_lat)
        & (curr_contour_lon != 180.0)
        & (curr_contour_lat != 0.0)
    )
    curr_contour_lon_plot = plot_anchor_lon + _minimal_lon_diff_deg(curr_contour_lon[contour_now_valid], plot_anchor_lon)
    curr_contour_lat_plot = curr_contour_lat[contour_now_valid]

    # 获取Argo浮标数据（复用已读取轨迹，避免重复 I/O）
    argo_data_filtered = filtered_float_data(ds_source_for_filter, no, track=track_df)
    argo_data_filtered = argo_data_filtered[pd.to_datetime(argo_data_filtered[['Year', 'Month', 'Day']]) == dates.iloc[needed_idx]]
    
    if argo_data_filtered.empty:
        needed_data = pd.DataFrame(columns=argo_data_filtered.columns)
    else:
        deltas = _reduce_argo_profiles_by_anomaly(
            argo_data_filtered,
            detection_config=argo_detection_config,
        )
        if deltas.empty:
            if verbose:
                print(
                    f"Warning: No Argo points pass thresholds "
                    f"({argo_detection_config.threshold_label()}, depth>={float(argo_detection_config.anomaly_min_depth):g}m)."
                )
            needed_data = pd.DataFrame(columns=argo_data_filtered.columns)
        else:
            needed_data = deltas
            needed_data.index.name = None

    # 获取区域边界（覆盖整条轨迹，优先用全程轮廓，否则退回中心轨迹）
    all_contour_lon = []
    all_contour_lat = []
    for lon_arr, lat_arr in zip(track_df['contour_lon'], track_df['contour_lat']):
        try:
            lon_np = np.asarray(lon_arr, dtype=float).ravel()
            lat_np = np.asarray(lat_arr, dtype=float).ravel()
        except Exception:
            continue
        if lon_np.size and lat_np.size:
            all_contour_lon.append(lon_np)
            all_contour_lat.append(lat_np)

    pad_deg = 0.5
    if all_contour_lon and all_contour_lat:
        lon_stack = np.concatenate(all_contour_lon)
        lat_stack = np.concatenate(all_contour_lat)
        valid_stack = (
            np.isfinite(lon_stack)
            & np.isfinite(lat_stack)
            & (lon_stack != 180.0)
            & (lat_stack != 0.0)
        )
        lon_stack = lon_stack[valid_stack]
        lat_stack = lat_stack[valid_stack]
        if lon_stack.size > 0 and lat_stack.size > 0:
            lon_stack_plot = plot_anchor_lon + _minimal_lon_diff_deg(lon_stack, plot_anchor_lon)
            glorys_lon_min = float(np.nanmin(lon_stack_plot) - pad_deg)
            glorys_lon_max = float(np.nanmax(lon_stack_plot) + pad_deg)
            glorys_lat_min = float(np.nanmin(lat_stack) - pad_deg)
            glorys_lat_max = float(np.nanmax(lat_stack) + pad_deg)
        else:
            glorys_lon_min = float(np.nanmin(center_lon_plot) - pad_deg)
            glorys_lon_max = float(np.nanmax(center_lon_plot) + pad_deg)
            glorys_lat_min = float(np.nanmin(center_lat_arr) - pad_deg)
            glorys_lat_max = float(np.nanmax(center_lat_arr) + pad_deg)
    else:
        glorys_lon_min = float(np.nanmin(center_lon_plot) - pad_deg)
        glorys_lon_max = float(np.nanmax(center_lon_plot) + pad_deg)
        glorys_lat_min = float(np.nanmin(center_lat_arr) - pad_deg)
        glorys_lat_max = float(np.nanmax(center_lat_arr) + pad_deg)

    # 获取背景场数据（共享底层）
    glorys_lon_filtered, glorys_lat_filtered, glorys_depth_filtered, glorys_variable_filtered = _compute_horizontal_glorys_field(
        variable=variable,
        needed_depth=needed_depth,
        loader=lambda vars_req, depth_req: get_track_area_glorys(
            DS,
            no,
            needed_date,
            variables=vars_req,
            depth=depth_req,
        ),
    )
    glorys_lon_plot = plot_anchor_lon + _minimal_lon_diff_deg(np.asarray(glorys_lon_filtered, dtype=float), plot_anchor_lon)

    ds_names = ds_name.upper() if isinstance(ds_name, str) else "UNKNOWN"
    colors_cycle = plt.rcParams['axes.prop_cycle'].by_key()['color']
    colors = colors_cycle[1] if ds_names in {'ACS', 'ACL'} else colors_cycle[0]
    world = _load_world_geodataframe()

    fig, ax = plt.subplots(figsize=figsize)
    ax.set_title(f'Track {ds_names}{no} at {glorys_depth_filtered[0]:.2f}m, {dates.iloc[needed_idx].strftime("%Y-%m-%d")}', fontsize=title_fs)
    ax.set_xlabel('Longitude', fontsize=label_fs)
    ax.set_ylabel('Latitude', fontsize=label_fs)
    world.plot(color='green', ax=ax)

    ax.tick_params(axis='both', which='major', labelsize=tick_fs)

    # 绘制涡旋轨迹
    ax.plot(center_lon_plot, center_lat_arr, color=colors, linewidth=track_lw, label='Center Track')
    ax.plot(center_lon_plot[0], center_lat_arr[0], marker='o', color=colors, markersize=8)
    ax.plot(center_lon_plot[-1], center_lat_arr[-1], marker='x', color=colors, markersize=8)

    # 绘制背景场
    pc = ax.pcolormesh(glorys_lon_plot, glorys_lat_filtered, glorys_variable_filtered, cmap='seismic', shading='auto', alpha=0.5)
    cbar = plt.colorbar(pc, ax=ax, orientation='horizontal', fraction=0.046, pad=0.06)
    _style_horizontal_colorbar(pc, cbar, variable, cbar_label_fs)
    cbar.ax.tick_params(labelsize=cbar_tick_fs)

    # 绘制Argo浮标数据
    if not needed_data.empty:
        needed_lon_plot = plot_anchor_lon + _minimal_lon_diff_deg(pd.to_numeric(needed_data['Longitude'], errors='coerce').to_numpy(dtype=float), plot_anchor_lon)
        needed_lat_plot = pd.to_numeric(needed_data['Latitude'], errors='coerce').to_numpy(dtype=float)
        needed_do_plot = pd.to_numeric(needed_data['DO'], errors='coerce').to_numpy(dtype=float)
        needed_depth_plot = pd.to_numeric(needed_data['Depth'], errors='coerce').to_numpy(dtype=float)
        valid_argo = np.isfinite(needed_lon_plot) & np.isfinite(needed_lat_plot)
        sc = ax.scatter(needed_lon_plot[valid_argo], needed_lat_plot[valid_argo], c=needed_do_plot[valid_argo], cmap = 'bwr', s=120,
                        vmin=150, vmax=240, edgecolors='black', linewidths=0.5,
                        label=f'Argo anomaly point ({argo_detection_config.threshold_label()}, depth>={float(argo_detection_config.anomaly_min_depth):g}m)', zorder=5)
        cbar2 = plt.colorbar(sc, ax=ax, orientation='horizontal', fraction=0.046, pad=cbar_pad)
        cbar2.set_label('DO/μmol·kg⁻¹', fontsize=cbar_label_fs)
        cbar2.ax.tick_params(labelsize=cbar_tick_fs)
        for lon_i, lat_i, dep_i, ok_i in zip(needed_lon_plot, needed_lat_plot, needed_depth_plot, valid_argo):
            if not ok_i or not np.isfinite(dep_i):
                continue
            ax.text(lon_i, lat_i, f"{int(dep_i)}", fontsize=argo_text_fs, fontweight='bold', ha='center', va='center', color='black', zorder=6)
    else:
        if verbose:
            print(f"No Argo data available for eddy {ds_names}{no} on {dates.iloc[needed_idx].strftime('%Y-%m-%d')}.")

    # 绘制当前时刻涡旋
    scale_now = approximate_degree_length(center_lat_arr[needed_idx])
    deg_h = radius_arr[needed_idx] / scale_now['meters_per_degree_lat']
    deg_w = radius_arr[needed_idx] / scale_now['meters_per_degree_lon']
    ell_now = Ellipse((center_lon_plot[needed_idx], center_lat_arr[needed_idx]), width=2*deg_w, height=2*deg_h,
                      edgecolor='r', facecolor='none', linestyle='--', alpha=0.2, linewidth=circle_lw, label='Effective Radius')
    ax.add_patch(ell_now)
    ax.scatter(center_lon_plot[needed_idx], center_lat_arr[needed_idx], color='black', s=16, label='Eddy Center', zorder=5)
    ax.plot(curr_contour_lon_plot, curr_contour_lat_plot, color=colors, linewidth=contour_lw, alpha=0.5, label='Effective Contour')

    _plot_horizontal_profile_lines(ax, k, b, glorys_lon_min, glorys_lon_max, line_lw=line_lw)

    ax.legend(fontsize=legend_fs)
    ax.set_xlim(glorys_lon_min, glorys_lon_max)
    ax.set_ylim(glorys_lat_min, glorys_lat_max)
    # 显示层统一为 [-180, 180] 经度习惯，底层仍使用连续经度以保证几何连贯。
    ax.xaxis.set_major_formatter(lambda x, _pos: f"{float(_normalize_lon_array(x)):.0f}")
    lon_span = abs(float(glorys_lon_max - glorys_lon_min))
    lat_span = abs(float(glorys_lat_max - glorys_lat_min))
    if lat_span > 0 and (lon_span / lat_span) > 20.0:
        # 全球长条窗口下保持 equal 会触发超宽画布，导致 Agg 渲染失败。
        ax.set_aspect('auto')
    else:
        ax.set_aspect('equal')

    # 紧凑布局，消除多余空白
    plt.tight_layout()

    # 保存图片
    if save_fig:
        region_slug = _current_region_key()
        run_tag = argo_detection_config.file_stem()
        output_dir = argo_detection_config.output_dir("plot_track_horizontal_glorys", region_slug)
        output_dir.mkdir(parents=True, exist_ok=True)
        base_filename = (
            f"{ds_names}{no}_{glorys_depth_filtered[0]:.2f}m_{variable}_"
            f"{dates.iloc[needed_idx].strftime('%Y%m%d')}_{run_tag}.png"
        )
        save_path = output_dir / base_filename
        plt.savefig(save_path, dpi=300, bbox_inches='tight')
        if verbose:
            print(f"\nFigure saved to: {save_path}")

    # 显示图像
    if show_fig:
        # 仅在非内联模式（即交互模式）下，激活直线绘制器
        if not inline_mode:
            line_drawer = LineDrawer(ax, legend_loc='best')
            cid_click = fig.canvas.mpl_connect('button_press_event', line_drawer.onclick)
            # 绑定到 figure，避免局部变量被回收导致交互回调偶发失效
            fig._interactive_line_drawer = line_drawer
            fig._interactive_line_drawer_click_cid = cid_click
            try:
                plt.show(block=False)
            except TypeError:
                plt.show()
        else:
            plt.show()

    # 只有在静态内联模式下，才在函数结束时关闭图像以释放内存
    if inline_mode:
        plt.close(fig)

    # 返回句柄便于调用侧自行管理或重复展示
    return fig, ax

def _resolve_argo_profile_center(
    profile_number: int,
    profile_time: int | str | pd.Timestamp,
    platform_number: int | None = None,
    argo_data_dir: str | Path | None = None,
) -> dict:
    """按 profile + 时间输入（年份或日期）定位单个 Argo 剖面的中心位置与日期。"""
    if argo_data_dir is None:
        argo_data_dir = argo_path

    target_date: pd.Timestamp | None = None
    if isinstance(profile_time, (int, np.integer)):
        val = int(profile_time)
        if 1000 <= val <= 9999:
            year = val
        else:
            parsed_time = convert_date(val)
            if pd.isna(parsed_time):
                raise ValueError(f"profile_time={val!r} 无法解析为日期。")
            year = int(parsed_time.year)
            target_date = parsed_time.normalize()
    elif isinstance(profile_time, str) and profile_time.strip().isdigit() and len(profile_time.strip()) == 4:
        year = int(profile_time.strip())
    else:
        try:
            parsed_time = pd.Timestamp(profile_time)
        except Exception as exc:
            raise ValueError(
                f"profile_time={profile_time!r} is not valid. "
                "Use year (e.g. 2014) or a date/timestamp (e.g. '2014-05-09')."
            ) from exc
        if pd.isna(parsed_time):
            raise ValueError(
                f"profile_time={profile_time!r} is not valid. "
                "Use year (e.g. 2014) or a date/timestamp (e.g. '2014-05-09')."
            )
        year = int(parsed_time.year)
        target_date = parsed_time.normalize()

    df_year = load_argo_data(int(year), data_dir=argo_data_dir)
    if df_year.empty:
        raise ValueError(f"No Argo data for year {year}.")

    work = df_year.copy()
    work['Profile_number'] = pd.to_numeric(work.get('Profile_number'), errors='coerce')
    work = work[work['Profile_number'] == int(profile_number)].copy()
    if work.empty:
        raise ValueError(f"Profile_number={profile_number} not found in year={year}.")
    if platform_number is not None:
        work = work[pd.to_numeric(work.get('Platform_number'), errors='coerce') == int(platform_number)].copy()

    if work.empty:
        raise ValueError("No rows left after month/day/platform filters.")

    for col in ['Longitude', 'Latitude', 'Depth']:
        if col in work.columns:
            work[col] = pd.to_numeric(work[col], errors='coerce')

    work['_date'] = pd.to_datetime(work[['Year', 'Month', 'Day']], errors='coerce').dt.normalize()
    work = work.dropna(subset=['Longitude', 'Latitude', '_date']).copy()
    if work.empty:
        raise ValueError("Profile rows have no valid Longitude/Latitude/Date.")

    if target_date is not None:
        work = work[work['_date'] == target_date].copy()
        if work.empty:
            raise ValueError(
                f"Profile_number={profile_number} not found on {target_date.strftime('%Y-%m-%d')} in year={year}."
            )

    if platform_number is None and 'Platform_number' in work.columns:
        pvals = pd.to_numeric(work['Platform_number'], errors='coerce').dropna().astype(int).unique()
        if pvals.size > 0:
            work = work[pd.to_numeric(work['Platform_number'], errors='coerce') == int(pvals[0])].copy()

    unique_dates = sorted(pd.to_datetime(work['_date'], errors='coerce').dropna().dt.normalize().unique())
    if len(unique_dates) == 0:
        raise ValueError("Cannot resolve target date from profile rows.")
    if target_date is None and len(unique_dates) > 1:
        raise ValueError(
            f"Profile_number={profile_number} in year={year} spans multiple dates. "
            "Please pass a specific date/timestamp as profile_time."
        )

    if target_date is None:
        target_date = pd.Timestamp(unique_dates[0]).normalize()
    rows_day = work[work['_date'] == target_date].copy()
    if rows_day.empty:
        raise ValueError("No rows for resolved target date.")

    if 'Depth' in rows_day.columns:
        rows_day = rows_day.sort_values('Depth', kind='mergesort')
    center_row = rows_day.iloc[0]

    platform_val = None
    if 'Platform_number' in center_row.index:
        try:
            platform_val = int(pd.to_numeric(center_row['Platform_number'], errors='coerce'))
        except Exception:
            platform_val = None

    return {
        'year_df': df_year,
        'profile_rows': rows_day,
        'center_lon': float(center_row['Longitude']),
        'center_lat': float(center_row['Latitude']),
        'target_date': target_date,
        'platform_number': platform_val,
        'profile_number': int(profile_number),
    }

def _window_bounds_from_center_km(center_lon: float, center_lat: float, window_half_size_km: float) -> tuple[float, float, float, float]:
    """由中心点与半窗口尺寸（km）生成局地经纬窗口边界。"""
    half_km = float(window_half_size_km)
    if (not np.isfinite(half_km)) or half_km <= 0:
        raise ValueError(f"window_half_size_km must be a positive finite number, got {window_half_size_km}.")

    scale = approximate_degree_length(float(center_lat))
    lon_half_deg = (half_km * 1000.0) / float(scale['meters_per_degree_lon'])
    lat_half_deg = (half_km * 1000.0) / float(scale['meters_per_degree_lat'])
    return (
        float(center_lon - lon_half_deg),
        float(center_lon + lon_half_deg),
        float(center_lat - lat_half_deg),
        float(center_lat + lat_half_deg),
    )

def _load_glorys_window_by_center(
    needed_date: str | pd.Timestamp,
    center_lon_ref: float,
    lon_min_local: float,
    lon_max_local: float,
    lat_min: float,
    lat_max: float,
    variables: list,
    depth: float | int | None = None,
):
    """按中心点局地窗口直接读取 GLORYS 子区域（不依赖 META 轨迹）。"""
    try:
        target_ts = pd.Timestamp(needed_date)
    except Exception as exc:
        raise ValueError(f"needed_date={needed_date!r} is not a valid date.") from exc

    needed_glorys_data = Dataset(get_glorys_filepath(target_ts), 'r')
    try:
        glorys_lon = needed_glorys_data.variables['longitude'][:]
        glorys_lat = needed_glorys_data.variables['latitude'][:]
        glorys_depth = needed_glorys_data.variables['depth'][:]

        glorys_lon_local = center_lon_ref + _minimal_lon_diff_deg(glorys_lon, center_lon_ref)
        glorys_lon_mask = (glorys_lon_local >= lon_min_local) & (glorys_lon_local <= lon_max_local)
        glorys_lat_mask = (glorys_lat >= lat_min) & (glorys_lat <= lat_max)

        if depth is not None:
            glorys_depth_mask = np.zeros_like(glorys_depth, dtype=bool)
            if glorys_depth.size > 0:
                k = np.argmin(np.abs(glorys_depth - depth))
                glorys_depth_mask[k] = True
        else:
            glorys_depth_mask = (glorys_depth >= 0) & (glorys_depth <= 2000)

        if not np.any(glorys_lon_mask) or not np.any(glorys_lat_mask):
            raise ValueError("No GLORYS grid points fall inside the requested Argo-centered window.")

        glorys_lon_filtered = glorys_lon_local[glorys_lon_mask]
        glorys_lat_filtered = glorys_lat[glorys_lat_mask]
        glorys_depth_filtered = glorys_depth[glorys_depth_mask]

        lon_order = np.argsort(glorys_lon_filtered)
        glorys_lon_filtered = glorys_lon_filtered[lon_order]

        glorys_variables_filtered = {}
        for var in variables:
            if var == 'thetao':
                arr = needed_glorys_data.variables['thetao'][0, glorys_depth_mask, glorys_lat_mask, glorys_lon_mask]
                if depth is not None:
                    arr = arr[0, :, :]
                glorys_variables_filtered['thetao'] = arr[..., lon_order]
            elif var == 'salinity' or var == 'so':
                arr = needed_glorys_data.variables['so'][0, glorys_depth_mask, glorys_lat_mask, glorys_lon_mask]
                if depth is not None:
                    arr = arr[0, :, :]
                glorys_variables_filtered['salinity'] = arr[..., lon_order]
            elif var == 'u' or var == 'uo':
                arr = needed_glorys_data.variables['uo'][0, glorys_depth_mask, glorys_lat_mask, glorys_lon_mask]
                if depth is not None:
                    arr = arr[0, :, :]
                glorys_variables_filtered['u'] = arr[..., lon_order]
            elif var == 'v' or var == 'vo':
                arr = needed_glorys_data.variables['vo'][0, glorys_depth_mask, glorys_lat_mask, glorys_lon_mask]
                if depth is not None:
                    arr = arr[0, :, :]
                glorys_variables_filtered['v'] = arr[..., lon_order]
            elif var == 'ssh' or var == 'zos':
                arr = needed_glorys_data.variables['zos'][0, glorys_lat_mask, glorys_lon_mask]
                glorys_variables_filtered['ssh'] = arr[..., lon_order]
            elif var == 'mlt' or var == 'mlotst':
                arr = needed_glorys_data.variables['mlotst'][0, glorys_lat_mask, glorys_lon_mask]
                glorys_variables_filtered['mlt'] = arr[..., lon_order]
            else:
                raise ValueError(
                    f"Unsupported variable: {var}. Supported variables are: "
                    "'thetao', 'so', 'uo', 'vo', 'zos', 'mlotst'."
                )

        return glorys_lon_filtered, glorys_lat_filtered, glorys_depth_filtered, glorys_variables_filtered
    finally:
        needed_glorys_data.close()

def get_vertical_glorys_from_center(
    center_lon: float,
    center_lat: float,
    needed_date: str | pd.Timestamp,
    k: float | list[float] | None = None,
    b: float | list[float] | None = None,
    *,
    variables: list = ['vorticity'],
    x_min_km: float | None = None,
    x_max_km: float | None = None,
    profile_spacing_km: float | None = None,
    interpolate_z: bool = True,
    profile_depth_spacing_m: float | None = None,
    window_half_size_km: float | None = None,
    profile_id: int | None = None,
    ds_name: str = 'ARGO',
) -> list[dict]:
    """按给定中心点与剖面线参数计算 GLORYS 垂向剖面数据包。

    与 get_vertical_glorys 同构，但以显式中心点 (center_lon, center_lat) 而非涡旋编号定位剖面；k/b 为
    None 时默认取纬向剖面线（k=0、b=center_lat）。返回结构与 get_vertical_glorys 一致。

    参数:
        - center_lon (float): 剖面中心经度。
        - center_lat (float): 剖面中心纬度。
        - needed_date (str | pd.Timestamp): 目标日期（'YYYY-MM-DD' 或时间戳）。
        - k (float | list[float] | None): 直线方程 y=kx+b 的斜率或斜率列表；None 时取 0。
        - b (float | list[float] | None): 截距或截距列表；None 时取 center_lat。
        - variables (list): 需要提取的 GLORYS 物理变量列表，默认 ['vorticity']。
        - x_min_km (float | None): 横坐标采样下界（km，中心为 0）；与 x_max_km 都为 None 时用全范围。
        - x_max_km (float | None): 横坐标采样上界（km），与 x_min_km 配套。
        - profile_spacing_km (float | None): 剖面采样步长（km）；None 时用配置默认值。
        - interpolate_z (bool): 是否将 z 轴重采样为等间距网格，默认 True。
        - profile_depth_spacing_m (float | None): z 轴重采样步长（m）；None 时用配置默认值。
        - window_half_size_km (float | None): GLORYS 子区域读取窗口半宽（km）；None 时按采样范围自动确定。
        - profile_id (int | None): 可选剖面标识，写入结果 metadata。
        - ds_name (str): 数据源名称标签，默认 'ARGO'。

    返回:
        - list[dict]: 剖面结果字典列表，结构同 get_vertical_glorys（含 profile_data/y_coords/z_coords/lon_coords/lat_coords/projections/metadata）。
    """
    if k is None or b is None:
        k_list, b_list = [0.0], [float(center_lat)]
    else:
        k_list = [k] if isinstance(k, (int, float)) else k
        b_list = [b] if isinstance(b, (int, float)) else b
    if len(k_list) != len(b_list):
        raise ValueError("k 和 b 的列表长度必须一致。")

    if profile_spacing_km is None:
        profile_spacing_km = _default_vertical_profile_spacing_km
    profile_spacing_km = float(profile_spacing_km)
    if (not np.isfinite(profile_spacing_km)) or profile_spacing_km <= 0:
        raise ValueError(f"profile_spacing_km must be a positive finite number, got {profile_spacing_km}.")

    interpolate_z = bool(interpolate_z)
    if interpolate_z:
        if profile_depth_spacing_m is None:
            profile_depth_spacing_m = _default_vertical_profile_depth_spacing_m
        profile_depth_spacing_m = float(profile_depth_spacing_m)
        if (not np.isfinite(profile_depth_spacing_m)) or profile_depth_spacing_m <= 0:
            raise ValueError(
                f"profile_depth_spacing_m must be a positive finite number, got {profile_depth_spacing_m}."
            )

    try:
        needed_ts = pd.Timestamp(needed_date).normalize()
    except Exception as exc:
        raise ValueError(f"needed_date={needed_date!r} is not a valid date.") from exc

    if x_min_km is not None and x_max_km is not None:
        half_from_x = max(abs(float(x_min_km)), abs(float(x_max_km)))
    else:
        half_from_x = 0.0

    if window_half_size_km is None:
        window_half_size_km = half_from_x if half_from_x > 0 else 400.0
    window_half_size_km = float(window_half_size_km)
    if (not np.isfinite(window_half_size_km)) or window_half_size_km <= 0:
        raise ValueError(f"window_half_size_km must be a positive finite number, got {window_half_size_km}.")
    if half_from_x > window_half_size_km:
        window_half_size_km = half_from_x

    alias_map = {
        'thetao': 'thetao',
        'salinity': 'salinity', 'so': 'salinity',
        'density': 'sigma', 'sigma': 'sigma', 'sigma0': 'sigma',
        'u': 'u', 'uo': 'u',
        'v': 'v', 'vo': 'v',
        'ssh': 'ssh', 'zos': 'ssh',
        'mlt': 'mlt', 'mlotst': 'mlt',
        'vorticity': 'vorticity',
        'pv': 'pv',
    }
    var_dims = {
        'thetao': 3, 'salinity': 3, 'sigma': 3, 'u': 3, 'v': 3, 'vorticity': 3,
        'pv': 3,
        'ssh': 2, 'mlt': 2
    }

    raw_vars_to_fetch = set()
    for var in variables:
        standard_name = alias_map.get(var, var)
        if standard_name == 'vorticity':
            raw_vars_to_fetch.update(['u', 'v'])
        elif standard_name == 'pv':
            raw_vars_to_fetch.update(['u', 'v', 'salinity', 'thetao'])
        elif standard_name == 'sigma':
            raw_vars_to_fetch.update(['salinity', 'thetao'])
        else:
            raw_vars_to_fetch.add(var)

    lon_min_local, lon_max_local, lat_min, lat_max = _window_bounds_from_center_km(
        float(center_lon),
        float(center_lat),
        window_half_size_km,
    )

    glorys_lon_raw, glorys_lat_raw, glorys_depth_raw, glorys_data_raw = _load_glorys_window_by_center(
        needed_ts,
        float(center_lon),
        lon_min_local,
        lon_max_local,
        lat_min,
        lat_max,
        variables=list(raw_vars_to_fetch),
        depth=None,
    )

    if glorys_depth_raw.size == 0 and not all(var_dims.get(alias_map.get(v, v)) == 2 for v in variables):
        return [{} for _ in k_list]

    radius_m = window_half_size_km * 1000.0
    scale_center = approximate_degree_length(float(center_lat))
    lon_radius_deg = radius_m / float(scale_center['meters_per_degree_lon'])
    lat_radius_deg = radius_m / float(scale_center['meters_per_degree_lat'])
    theta = np.linspace(0.0, 2.0 * np.pi, 72, endpoint=False)
    contour_lon = np.asarray(float(center_lon) + lon_radius_deg * np.cos(theta), dtype=float)
    contour_lat = np.asarray(float(center_lat) + lat_radius_deg * np.sin(theta), dtype=float)

    sigma_3d_cache = None
    if any(alias_map.get(v, v) in ('sigma', 'pv') for v in variables):
        sal_3d = glorys_data_raw.get('salinity')
        theta_3d = glorys_data_raw.get('thetao')
        if sal_3d is not None and theta_3d is not None:
            sal_ma = np.ma.array(sal_3d, copy=False)
            theta_ma = np.ma.array(theta_3d, copy=False)
            if sal_ma.ndim == 2:
                sal_ma = sal_ma[np.newaxis, :, :]
            if theta_ma.ndim == 2:
                theta_ma = theta_ma[np.newaxis, :, :]

            if sal_ma.shape == theta_ma.shape and sal_ma.ndim == 3:
                try:
                    z3, lat3, lon3 = np.meshgrid(
                        glorys_depth_raw,
                        glorys_lat_raw,
                        _normalize_lon_array(glorys_lon_raw),
                        indexing='ij'
                    )
                    sal_filled = np.ma.filled(sal_ma, np.nan)
                    theta_filled = np.ma.filled(theta_ma, np.nan)
                    p3 = gsw.p_from_z(-z3, lat3)
                    SA = gsw.SA_from_SP(sal_filled, p3, lon3, lat3)
                    CT = gsw.CT_from_pt(SA, theta_filled)
                    sigma0 = gsw.sigma0(SA, CT)

                    input_mask = np.ma.getmaskarray(sal_ma) | np.ma.getmaskarray(theta_ma)
                    sigma_3d_cache = np.ma.masked_invalid(sigma0)
                    if input_mask.any():
                        sigma_3d_cache = np.ma.array(
                            sigma_3d_cache,
                            mask=np.ma.getmaskarray(sigma_3d_cache) | input_mask,
                            copy=False,
                        )
                except Exception:
                    sigma_3d_cache = None

    glorys_lon_min = float(np.min(glorys_lon_raw))
    glorys_lon_max = float(np.max(glorys_lon_raw))
    glorys_lat_min = float(np.min(glorys_lat_raw))
    glorys_lat_max = float(np.max(glorys_lat_raw))

    all_profiles_data = []
    for k_val, b_val in zip(k_list, b_list):
        intersection_pts: list[tuple[float, float]] = []
        tol = 1e-10

        if np.isclose(k_val, 0.0, atol=1e-12):
            if glorys_lat_min - tol <= b_val <= glorys_lat_max + tol:
                intersection_pts = [(float(glorys_lon_min), float(b_val)), (float(glorys_lon_max), float(b_val))]
        else:
            lat_left = k_val * glorys_lon_min + b_val
            if glorys_lat_min - tol <= lat_left <= glorys_lat_max + tol:
                intersection_pts.append((float(glorys_lon_min), float(lat_left)))

            lat_right = k_val * glorys_lon_max + b_val
            if glorys_lat_min - tol <= lat_right <= glorys_lat_max + tol:
                intersection_pts.append((float(glorys_lon_max), float(lat_right)))

            lon_bottom = (glorys_lat_min - b_val) / k_val
            if glorys_lon_min - tol <= lon_bottom <= glorys_lon_max + tol:
                intersection_pts.append((float(lon_bottom), float(glorys_lat_min)))

            lon_top = (glorys_lat_max - b_val) / k_val
            if glorys_lon_min - tol <= lon_top <= glorys_lon_max + tol:
                intersection_pts.append((float(lon_top), float(glorys_lat_max)))

        unique_pts: list[tuple[float, float]] = []
        for p_lon, p_lat in intersection_pts:
            duplicated = any(
                abs(float(_minimal_lon_diff_deg(p_lon, q_lon))) < 1e-8 and abs(p_lat - q_lat) < 1e-8
                for q_lon, q_lat in unique_pts
            )
            if not duplicated:
                unique_pts.append((p_lon, p_lat))

        if len(unique_pts) < 2:
            all_profiles_data.append({})
            continue

        if len(unique_pts) == 2:
            p0, p1 = unique_pts[0], unique_pts[1]
        else:
            best_pair = None
            best_dist = -np.inf
            for i in range(len(unique_pts) - 1):
                for j in range(i + 1, len(unique_pts)):
                    d_ij = local_xy_distance_m(
                        unique_pts[j][0],
                        unique_pts[j][1],
                        unique_pts[i][0],
                        unique_pts[i][1],
                        wrap_dateline=True,
                    )
                    if np.isfinite(d_ij) and d_ij > best_dist:
                        best_dist = float(d_ij)
                        best_pair = (unique_pts[i], unique_pts[j])

            if best_pair is None:
                all_profiles_data.append({})
                continue
            p0, p1 = best_pair

        dlon_p0_to_p1 = float(_minimal_lon_diff_deg(p1[0], p0[0]))
        if dlon_p0_to_p1 < 0 or (abs(dlon_p0_to_p1) <= 1e-12 and p1[1] < p0[1]):
            p0, p1 = p1, p0

        segment_len_m = local_xy_distance_m(p1[0], p1[1], p0[0], p0[1], wrap_dateline=True)
        if (not np.isfinite(segment_len_m)) or segment_len_m <= 0:
            all_profiles_data.append({})
            continue

        n_samples = max(2, int(np.ceil(segment_len_m / (profile_spacing_km * 1000.0))) + 1)
        t = np.linspace(0.0, 1.0, n_samples)
        dlon_total = float(_minimal_lon_diff_deg(p1[0], p0[0]))
        if dlon_total < 0:
            p0, p1 = p1, p0
            dlon_total = float(_minimal_lon_diff_deg(p1[0], p0[0]))

        profile_lons_full = p0[0] + t * dlon_total
        profile_lats_full = p0[1] + t * (p1[1] - p0[1])

        dlat_deg = np.diff(profile_lats_full)
        dlon_deg = _minimal_lon_diff_deg(profile_lons_full[1:], profile_lons_full[:-1])
        mid_lats_deg = (profile_lats_full[:-1] + profile_lats_full[1:]) / 2
        scale_mid = approximate_degree_length(mid_lats_deg)
        dist_segments = np.hypot(
            dlon_deg * scale_mid['meters_per_degree_lon'],
            dlat_deg * scale_mid['meters_per_degree_lat']
        )
        y_coords_raw_full = np.insert(np.cumsum(dist_segments), 0, 0) / 1000.0

        current_center_lon = float(center_lon)
        current_center_lat = float(center_lat)
        if k_val == 0:
            xp, yp = current_center_lon, b_val
        else:
            xp = (current_center_lon + k_val * current_center_lat - k_val * b_val) / (1 + k_val ** 2)
            yp = k_val * xp + b_val
        center_idx_on_profile = np.argmin(_minimal_lon_diff_deg(profile_lons_full, xp) ** 2 + (profile_lats_full - yp) ** 2)
        y_coords_recenter_full = y_coords_raw_full - y_coords_raw_full[center_idx_on_profile]

        profile_lons = profile_lons_full
        profile_lats = profile_lats_full
        y_coords_recenter = y_coords_recenter_full
        if x_min_km is not None and x_max_km is not None:
            x_left = float(min(x_min_km, x_max_km))
            x_right = float(max(x_min_km, x_max_km))
            span_km = x_right - x_left
            n_steps = int(np.floor(span_km / profile_spacing_km))
            if n_steps < 1:
                all_profiles_data.append({})
                continue

            y_target = x_left + np.arange(n_steps + 1) * profile_spacing_km
            y_full_min = float(np.nanmin(y_coords_recenter_full))
            y_full_max = float(np.nanmax(y_coords_recenter_full))
            valid_target = (y_target >= y_full_min) & (y_target <= y_full_max)
            if np.count_nonzero(valid_target) < 2:
                all_profiles_data.append({})
                continue

            y_coords_recenter = y_target[valid_target]
            profile_lons = np.interp(y_coords_recenter, y_coords_recenter_full, profile_lons_full)
            profile_lats = np.interp(y_coords_recenter, y_coords_recenter_full, profile_lats_full)

        z_coords_native = glorys_depth_raw
        profile_data_dict = {}
        for var in variables:
            standard_name = alias_map.get(var, var)
            dimension = var_dims.get(standard_name, 3)

            if dimension == 2:
                data_2d = glorys_data_raw.get(standard_name)
                if data_2d is None or data_2d.size == 0 or np.all(np.ma.getmask(data_2d)):
                    profile_data_dict[standard_name] = np.ma.masked_all(len(profile_lons))
                    continue

                data_2d_filled = np.ma.filled(np.ma.array(data_2d, copy=False), np.nan)
                interp_func_2d = RegularGridInterpolator(
                    (glorys_lat_raw, glorys_lon_raw),
                    data_2d_filled,
                    method='linear',
                    bounds_error=False,
                    fill_value=np.nan,
                )
                interpolated_1d = interp_func_2d(list(zip(profile_lats, profile_lons)))
                profile_data_dict[standard_name] = np.ma.masked_invalid(interpolated_1d)
                continue

            glorys_variable_3d = None
            if standard_name == 'vorticity':
                u, v = glorys_data_raw.get('u'), glorys_data_raw.get('v')
                if u is not None and v is not None and u.size > 0 and v.size > 0:
                    if u.ndim == 2:
                        u, v = u[np.newaxis, :, :], v[np.newaxis, :, :]
                    zeta_3d, f_3d = calculate_vorticity(glorys_lon_raw, glorys_lat_raw, u, v)
                    glorys_variable_3d = zeta_3d / f_3d
            elif standard_name == 'sigma':
                glorys_variable_3d = sigma_3d_cache
            elif standard_name == 'pv':
                u = glorys_data_raw.get('u')
                v = glorys_data_raw.get('v')
                if (u is not None and v is not None and u.size > 0 and v.size > 0
                        and sigma_3d_cache is not None):
                    if u.ndim == 2:
                        u, v = u[np.newaxis, :, :], v[np.newaxis, :, :]
                    zeta_3d, f_3d = calculate_vorticity(glorys_lon_raw, glorys_lat_raw, u, v)
                    z3 = glorys_depth_raw[:, np.newaxis, np.newaxis]
                    dz = np.gradient(z3, axis=0)
                    with np.errstate(divide='ignore', invalid='ignore'):
                        dsigma_dz = np.gradient(sigma_3d_cache, axis=0) / np.where(dz != 0, dz, np.nan)
                    N2 = (9.81 / 1025.0) * dsigma_dz
                    N2 = np.maximum(N2, 0.0)  # clip convective instability (N² < 0)
                    glorys_variable_3d = (f_3d + zeta_3d) * N2 / 9.81
            else:
                glorys_variable_3d = glorys_data_raw.get(standard_name)

            if glorys_variable_3d is None or glorys_variable_3d.size == 0 or np.all(np.ma.getmask(glorys_variable_3d)):
                profile_data_dict[standard_name] = np.ma.masked_all((len(z_coords_native), len(profile_lons)))
                continue

            if glorys_variable_3d.ndim == 2:
                glorys_variable_3d = glorys_variable_3d[np.newaxis, :, :]

            query_depths, query_lats = np.meshgrid(z_coords_native, profile_lats, indexing='ij')
            _, query_lons = np.meshgrid(z_coords_native, profile_lons, indexing='ij')
            xi_points = np.vstack([query_depths.ravel(), query_lats.ravel(), query_lons.ravel()]).T

            variable_3d_filled = np.ma.filled(np.ma.array(glorys_variable_3d, copy=False), np.nan)
            interp_func = RegularGridInterpolator(
                (z_coords_native, glorys_lat_raw, glorys_lon_raw),
                variable_3d_filled,
                method='linear',
                bounds_error=False,
                fill_value=np.nan,
            )
            interpolated_values_flat = interp_func(xi_points)
            profile_data_dict[standard_name] = np.ma.masked_invalid(
                interpolated_values_flat.reshape(len(z_coords_native), len(profile_lons))
            )

        z_coords = z_coords_native
        if interpolate_z and z_coords_native.size >= 2:
            z_min = float(z_coords_native[0])
            z_max = float(z_coords_native[-1])
            z_span = z_max - z_min
            n_steps_z = int(np.floor(z_span / profile_depth_spacing_m))
            if n_steps_z >= 1:
                z_coords_target = z_min + np.arange(n_steps_z + 1) * profile_depth_spacing_m

                zz_grid, yy_grid = np.meshgrid(z_coords_target, y_coords_recenter, indexing='ij')
                query_points_2d = np.vstack([zz_grid.ravel(), yy_grid.ravel()]).T

                for key_name, var_data in list(profile_data_dict.items()):
                    var_ma = np.ma.array(var_data, copy=False)
                    if var_ma.ndim != 2 or var_ma.shape[0] != len(z_coords_native):
                        continue
                    if np.ma.getmaskarray(var_ma).all():
                        profile_data_dict[key_name] = np.ma.masked_all((len(z_coords_target), len(y_coords_recenter)))
                        continue

                    interp_func_2d = RegularGridInterpolator(
                        (z_coords_native, y_coords_recenter),
                        np.ma.filled(var_ma, np.nan),
                        method='linear',
                        bounds_error=False,
                        fill_value=np.nan,
                    )
                    new_vals_flat = interp_func_2d(query_points_2d)
                    profile_data_dict[key_name] = np.ma.masked_invalid(
                        new_vals_flat.reshape(len(z_coords_target), len(y_coords_recenter))
                    )

                z_coords = z_coords_target

        scale_line = approximate_degree_length(current_center_lat)
        semi_axis_lon_deg = radius_m / float(scale_line['meters_per_degree_lon'])
        semi_axis_lat_deg = radius_m / float(scale_line['meters_per_degree_lat'])

        radius_intersections_lon: list[float] = []
        if semi_axis_lon_deg > 0 and semi_axis_lat_deg > 0 and np.isfinite(semi_axis_lon_deg) and np.isfinite(semi_axis_lat_deg):
            inv_a2 = 1.0 / (semi_axis_lon_deg ** 2)
            inv_b2 = 1.0 / (semi_axis_lat_deg ** 2)
            dy0 = b_val - current_center_lat

            A = inv_a2 + (k_val ** 2) * inv_b2
            B = -2.0 * current_center_lon * inv_a2 + 2.0 * k_val * dy0 * inv_b2
            C = (current_center_lon ** 2) * inv_a2 + (dy0 ** 2) * inv_b2 - 1.0

            discriminant = B ** 2 - 4.0 * A * C
            if discriminant >= -1e-12 and A > 0:
                discriminant = max(0.0, float(discriminant))
                sqrt_disc = float(np.sqrt(discriminant))
                radius_intersections_lon = [
                    (-B + sqrt_disc) / (2.0 * A),
                    (-B - sqrt_disc) / (2.0 * A),
                ]

        radius_proj_dists = [
            y_coords_raw_full[np.argmin(_minimal_lon_diff_deg(profile_lons_full, lon_i) ** 2 + (profile_lats_full - (k_val * lon_i + b_val)) ** 2)]
            - y_coords_raw_full[center_idx_on_profile]
            for lon_i in radius_intersections_lon
        ]

        curr_contour_lon = np.asarray(contour_lon, dtype=float).ravel()
        curr_contour_lat = np.asarray(contour_lat, dtype=float).ravel()
        valid_mask = np.isfinite(curr_contour_lon) & np.isfinite(curr_contour_lat)
        curr_contour_lon = curr_contour_lon[valid_mask]
        curr_contour_lat = curr_contour_lat[valid_mask]
        curr_contour_lon = current_center_lon + _minimal_lon_diff_deg(curr_contour_lon, current_center_lon)
        contour_intersections_xy = find_polygon_line_intersections(curr_contour_lon, curr_contour_lat, profile_lons_full, profile_lats_full)
        contour_proj_dists = [
            y_coords_raw_full[np.argmin(_minimal_lon_diff_deg(profile_lons_full, lon_i) ** 2 + (profile_lats_full - lat_i) ** 2)]
            - y_coords_raw_full[center_idx_on_profile]
            for lon_i, lat_i in contour_intersections_xy
        ]

        projections_dict = {'radius': sorted(radius_proj_dists), 'contour': sorted(contour_proj_dists)}

        single_profile_result = {
            'profile_data': profile_data_dict,
            'y_coords': y_coords_recenter,
            'z_coords': z_coords,
            'lon_coords': profile_lons,
            'lat_coords': profile_lats,
            'projections': projections_dict,
            'metadata': {
                'eddy_no': int(profile_id) if profile_id is not None else -1,
                'date_str': needed_ts.strftime("%Y-%m-%d"),
                'k': k_val,
                'b': b_val,
                'ds_name': 'Argo' if str(ds_name).lower() == 'argo' else str(ds_name).upper(),
                'entity_label': f"Profile {int(profile_id)}" if profile_id is not None else 'Profile',
                'draw_reference_lines': False,
            }
        }
        all_profiles_data.append(single_profile_result)

    return all_profiles_data

def plot_argo_horizontal_glorys(
    profile_number: int,
    profile_time: int | str | pd.Timestamp,
    platform_number: int | None = None,
    variable: str = 'vorticity',
    show_fig: bool = True,
    save_fig: bool = False,
    k: float | list[float] | None = None,
    b: float | list[float] | None = None,
    needed_depth: float | int = 0,
    inline_mode: bool = True,
    xmin: float = -400.0,
    xmax: float = 400.0,
    argo_detection_config: DetectionConfig | None = None,
    argo_min_depth: float | None = None,
    argo_data_dir: str | Path | None = None,
    output_dir: str | Path | None = None,
    verbose: bool = True,
):
    """以单个 Argo 剖面为中心绘制 GLORYS 水平快照图。

    先根据 profile_number 与 profile_time 定位目标剖面中心，再在给定窗口内读取同日 GLORYS 场并叠加同
    口径筛选后的 Argo 异常点。

    参数:
        - profile_number (int): 目标 Argo 剖面编号。
        - profile_time (int | str | pd.Timestamp): 时间输入，支持年份（如 2014 或 "2014"）或具体日期/时间戳（如 "2014-05-09"）；仅给年份且该剖面在该年对应多个日期时会提示需传具体日期。
        - platform_number (int | None): 可选平台号过滤；None 时自动选择该剖面对应平台。
        - variable (str): 背景变量名，常用 'vorticity'/'thetao'/'so'/'u'/'v'/'ssh'，默认 'vorticity'。
        - show_fig (bool): 是否显示图像，默认 True。
        - save_fig (bool): 是否保存图像到输出目录，默认 False。
        - k (float | list[float] | None): 剖面线斜率，按 y = kx + b 叠加；传列表需等长，逐条绘制。
        - b (float | list[float] | None): 剖面线截距，与 k 配套。
        - needed_depth (float | int): GLORYS 读取深度（m），默认 0（表层）。
        - inline_mode (bool): 是否使用内联静态模式，默认 True；详见“说明”。
        - xmin (float): 局地窗口横向下界（km，中心为 0），用于确定经纬子区域读取范围，默认 -400.0。
        - xmax (float): 局地窗口横向上界（km），默认 400.0。
        - argo_detection_config (DetectionConfig | None): 叠加点异常筛选配置；None 时使用默认。
        - argo_min_depth (float | None): 叠加点最小深度阈值（m）；None 时回退配置项。
        - argo_data_dir (str | Path | None): Argo 年度 parquet 目录；None 时使用配置默认目录。
        - output_dir (str | Path | None): 保存目录覆盖；None 时使用默认输出目录。
        - verbose (bool): 是否打印保存路径与提示信息，默认 True。

    返回:
        - tuple: (fig, ax)，便于调用侧进一步自定义或复用。

    说明:
        显示模式:

            - inline_mode=True（默认）：静态出图模式，资源占用更可控，函数结束会关闭 figure 释放内存，适合脚本批量出图。
            - inline_mode=False：交互模式，保留图窗句柄，适合 Notebook 实时查看与二次操作。
    """
    argo_detection_config = _resolve_detection_config(
        argo_detection_config,
        anomaly_min_depth=argo_min_depth,
    )

    if not inline_mode:
        backend_name = str(plt.get_backend()).lower()
        if 'inline' in backend_name:
            if verbose:
                print("Warning: current matplotlib backend is inline; click interaction may appear unresponsive. "
                      "Use %matplotlib widget or a GUI backend for stable interaction.")

    if inline_mode:
        plt.close('all')

    info = _resolve_argo_profile_center(
        profile_number=int(profile_number),
        profile_time=profile_time,
        platform_number=platform_number,
        argo_data_dir=argo_data_dir,
    )
    center_lon = float(info['center_lon'])
    center_lat = float(info['center_lat'])
    needed_date = pd.Timestamp(info['target_date'])
    df_year = info['year_df']

    if xmin is None or xmax is None:
        xmin = -400.0
        xmax = 400.0
    x_left = float(min(xmin, xmax))
    x_right = float(max(xmin, xmax))
    window_half_size_km = max(abs(x_left), abs(x_right))

    lon_min_local, lon_max_local, lat_min, lat_max = _window_bounds_from_center_km(
        center_lon,
        center_lat,
        window_half_size_km,
    )

    glorys_lon_filtered, glorys_lat_filtered, glorys_depth_filtered, glorys_variable_filtered = _compute_horizontal_glorys_field(
        variable=variable,
        needed_depth=needed_depth,
        loader=lambda vars_req, depth_req: _load_glorys_window_by_center(
            needed_date,
            center_lon,
            lon_min_local,
            lon_max_local,
            lat_min,
            lat_max,
            variables=vars_req,
            depth=depth_req,
        ),
    )

    # 同日 Argo，按窗口粗筛后沿用当前 DetectionConfig 的异常筛选与“每剖面取最强异常”规则
    day_ts = pd.to_datetime(df_year[['Year', 'Month', 'Day']], errors='coerce').dt.normalize()
    day_rows = df_year.loc[day_ts == needed_date.normalize()].copy()
    if not day_rows.empty:
        day_rows['Longitude'] = pd.to_numeric(day_rows['Longitude'], errors='coerce')
        day_rows['Latitude'] = pd.to_numeric(day_rows['Latitude'], errors='coerce')
        day_rows = day_rows.dropna(subset=['Longitude', 'Latitude'])

    needed_data = pd.DataFrame()
    if not day_rows.empty:
        day_lon_local = center_lon + _minimal_lon_diff_deg(day_rows['Longitude'].to_numpy(dtype=float), center_lon)
        mask_window = (
            (day_lon_local >= lon_min_local)
            & (day_lon_local <= lon_max_local)
            & (day_rows['Latitude'].to_numpy(dtype=float) >= lat_min)
            & (day_rows['Latitude'].to_numpy(dtype=float) <= lat_max)
        )
        day_window = day_rows.loc[mask_window].copy()
        if not day_window.empty:
            deltas = _reduce_argo_profiles_by_anomaly(
                day_window,
                detection_config=argo_detection_config,
            )
            if not deltas.empty:
                needed_data = deltas
                needed_data['Longitude_local'] = center_lon + _minimal_lon_diff_deg(
                    pd.to_numeric(needed_data['Longitude'], errors='coerce').to_numpy(dtype=float),
                    center_lon,
                )

    figsize = (12, 10)
    title_fs, label_fs, tick_fs, legend_fs = 16, 14, 12, 10
    cbar_label_fs, cbar_tick_fs = 12, 10
    argo_text_fs = 6
    line_lw = 2.0

    fig, ax = plt.subplots(figsize=figsize)
    world = _load_world_geodataframe()
    world.plot(color='green', ax=ax)
    ax.tick_params(axis='both', which='major', labelsize=tick_fs)

    pc = ax.pcolormesh(glorys_lon_filtered, glorys_lat_filtered, glorys_variable_filtered, cmap='seismic', shading='auto', alpha=0.5)
    cbar = plt.colorbar(pc, ax=ax, orientation='horizontal', fraction=0.046, pad=0.06)
    _style_horizontal_colorbar(pc, cbar, variable, cbar_label_fs)
    cbar.ax.tick_params(labelsize=cbar_tick_fs)

    _plot_horizontal_profile_lines(ax, k, b, lon_min_local, lon_max_local, line_lw=line_lw)

    # 叠加通过同标准筛选的 Argo
    if not needed_data.empty:
        sc = ax.scatter(
            needed_data['Longitude_local'],
            needed_data['Latitude'],
            c=needed_data['DO'],
            cmap='bwr',
            s=120,
            vmin=150,
            vmax=240,
            edgecolors='black',
            linewidths=0.5,
            label=f'Argo anomaly point ({argo_detection_config.threshold_label()}, depth>={float(argo_detection_config.anomaly_min_depth):g}m)',
            zorder=7,
        )
        cbar2 = plt.colorbar(sc, ax=ax, orientation='horizontal', fraction=0.046, pad=0.16)
        cbar2.set_label('DO/μmol·kg⁻¹', fontsize=cbar_label_fs)
        cbar2.ax.tick_params(labelsize=cbar_tick_fs)
        for _, row in needed_data.iterrows():
            ax.text(
                row['Longitude_local'],
                row['Latitude'],
                f"{int(row['Depth'])}",
                fontsize=argo_text_fs,
                fontweight='bold',
                ha='center',
                va='center',
                color='black',
                zorder=8,
            )
    else:
        if verbose:
            print(
                f"No Argo anomalies in window on {needed_date.strftime('%Y-%m-%d')} "
                f"({argo_detection_config.threshold_label()}, depth>={float(argo_detection_config.anomaly_min_depth):g}m)."
            )

    ax.set_title(
        f"Profile {int(profile_number)} GLORYS snapshot at {float(glorys_depth_filtered[0]):.2f}m, "
        f"{needed_date.strftime('%Y-%m-%d')}",
        fontsize=title_fs,
    )
    ax.set_xlabel('Longitude', fontsize=label_fs)
    ax.set_ylabel('Latitude', fontsize=label_fs)
    ax.legend(fontsize=legend_fs, loc='best')
    ax.set_xlim(lon_min_local, lon_max_local)
    ax.set_ylim(lat_min, lat_max)
    ax.set_aspect('equal')
    plt.tight_layout()

    if save_fig:
        region_slug = _current_region_key()
        run_tag = argo_detection_config.file_stem()
        save_dir = (
            Path(output_dir)
            if output_dir is not None
            else argo_detection_config.output_dir("plot_argo_horizontal_glorys", region_slug)
        )
        save_dir.mkdir(parents=True, exist_ok=True)
        fname = (
            f"Argo_{needed_date.strftime('%Y%m%d')}_P{int(profile_number)}_"
            f"{float(glorys_depth_filtered[0]):.2f}m_{variable}_{run_tag}.png"
        )
        save_path = save_dir / fname
        plt.savefig(save_path, dpi=300, bbox_inches='tight')
        if verbose:
            print(f"Figure saved to: {save_path}")

    if show_fig:
        if not inline_mode:
            line_drawer = LineDrawer(ax, legend_loc='best')
            cid_click = fig.canvas.mpl_connect('button_press_event', line_drawer.onclick)
            # 绑定到 figure，避免局部变量被回收导致交互回调偶发失效
            fig._interactive_line_drawer = line_drawer
            fig._interactive_line_drawer_click_cid = cid_click
            try:
                plt.show(block=False)
            except TypeError:
                plt.show()
        else:
            plt.show()

    if inline_mode:
        plt.close(fig)

    return fig, ax

def get_track_area_glorys(DS: list, no: int, needed_date: str | pd.Timestamp, variables: list = ['thetao'],
                          depth: float | int | None = None,
                          lon_min_local: float | None = None,
                          lon_max_local: float | None = None,
                          lat_min: float | None = None,
                          lat_max: float | None = None):
    '''
    获取指定涡旋在特定时间点周围的 GLORYS 数据。

    根据涡旋轮廓确定一个矩形区域，并从相应 GLORYS 文件中提取该区域内的一个或多个物理变量。

    参数:
        - DS (list): 包含涡旋轨迹信息的数据集。
        - no (int): 涡旋唯一编号。
        - needed_date (str | pd.Timestamp): 需要提取数据的日期（'YYYY-MM-DD' 或时间戳）。
        - variables (list): 需要提取的变量列表，默认 ['thetao']，可选 'salinity'/'u'/'v'/'ssh'/'mlt'。
        - depth (float | int | None): 指定时提取该深度的 GLORYS 数据；None 时提取 2000 米以内的所有深度。
        - lon_min_local (float | None): 可选局地经度下界（以轨迹中心经度为参考的连续经度坐标）。
        - lon_max_local (float | None): 可选局地经度上界。
        - lat_min (float | None): 可选纬度下界（与轮廓窗口取并集）。
        - lat_max (float | None): 可选纬度上界。

    返回:
        - tuple: (经度数组, 纬度数组, 深度数组, 变量数据字典)，字典存储所有请求变量的数据。
    '''
    wanted_track = find_track(DS, no)
    # 展平轮廓数组并过滤无效值，兼容 DataFrame/ndarray/object 列
    contour_lon_col = wanted_track['contour_lon'].values
    contour_lat_col = wanted_track['contour_lat'].values
    lon_flat: list[np.ndarray] = []
    lat_flat: list[np.ndarray] = []
    for lon_arr, lat_arr in zip(contour_lon_col, contour_lat_col):
        try:
            lon_np = np.asarray(lon_arr, dtype=float).ravel()
            lat_np = np.asarray(lat_arr, dtype=float).ravel()
        except Exception:
            continue
        if lon_np.size and lat_np.size:
            lon_flat.append(lon_np)
            lat_flat.append(lat_np)

    if not lon_flat or not lat_flat:
        raise ValueError("No valid contour coordinates found for the requested track.")

    contour_lon = np.concatenate(lon_flat)
    contour_lat = np.concatenate(lat_flat)

    try:
        target_ts = pd.Timestamp(needed_date)
    except Exception as exc:
        raise ValueError(f"needed_date={needed_date!r} is not a valid date.") from exc

    track_dates = wanted_track['date'] if 'date' in wanted_track.columns else convert_date(wanted_track['time'])
    track_dates = pd.to_datetime(track_dates, errors='coerce')
    same_day_idx = np.nonzero(track_dates.dt.normalize().to_numpy() == target_ts.normalize().to_datetime64())[0]
    if same_day_idx.size == 0:
        raise ValueError(f"Date {target_ts.strftime('%Y-%m-%d')} not found in track {no}.")

    ref_idx = int(same_day_idx[0])
    center_lon_ref = float(wanted_track.iloc[ref_idx]['center_lon'])
    target_date = pd.Timestamp(track_dates.iloc[ref_idx])
    needed_glorys_data = Dataset(get_glorys_filepath(target_date), 'r')

    contour_lon_arr = np.asarray(contour_lon, dtype=float)
    contour_lat_arr = np.asarray(contour_lat, dtype=float)
    contour_valid_mask = (
        np.isfinite(contour_lon_arr)
        & np.isfinite(contour_lat_arr)
        & (contour_lon_arr != 180.0)
        & (contour_lat_arr != 0.0)
    )
    if not np.any(contour_valid_mask):
        needed_glorys_data.close()
        raise ValueError("No valid contour coordinates remain after cleaning.")

    contour_lon_valid = contour_lon_arr[contour_valid_mask]
    contour_lat_valid = contour_lat_arr[contour_valid_mask]
    contour_lon_local = center_lon_ref + _minimal_lon_diff_deg(contour_lon_valid, center_lon_ref)

    pad_deg = 0.5
    glorys_lon_min = np.min(contour_lon_local) - pad_deg
    glorys_lon_max = np.max(contour_lon_local) + pad_deg
    glorys_lat_min = np.min(contour_lat_valid) - pad_deg
    glorys_lat_max = np.max(contour_lat_valid) + pad_deg

    if lon_min_local is not None:
        glorys_lon_min = min(glorys_lon_min, float(lon_min_local))
    if lon_max_local is not None:
        glorys_lon_max = max(glorys_lon_max, float(lon_max_local))
    if lat_min is not None:
        glorys_lat_min = min(glorys_lat_min, float(lat_min))
    if lat_max is not None:
        glorys_lat_max = max(glorys_lat_max, float(lat_max))

    glorys_lon = needed_glorys_data.variables['longitude'][:]
    glorys_lat = needed_glorys_data.variables['latitude'][:]
    glorys_depth = needed_glorys_data.variables['depth'][:]
    glorys_lon_local = center_lon_ref + _minimal_lon_diff_deg(glorys_lon, center_lon_ref)

    glorys_lon_mask = (glorys_lon_local >= glorys_lon_min) & (glorys_lon_local <= glorys_lon_max)
    glorys_lat_mask = (glorys_lat >= glorys_lat_min) & (glorys_lat <= glorys_lat_max)
    if depth is not None:
        glorys_depth_mask = np.zeros_like(glorys_depth, dtype=bool)
        if glorys_depth.size > 0:
            k = np.argmin(np.abs(glorys_depth - depth))
            glorys_depth_mask[k] = True
    else:
        glorys_depth_mask = (glorys_depth >= 0) & (glorys_depth <= 2000)

    if not np.any(glorys_lon_mask) or not np.any(glorys_lat_mask):
        needed_glorys_data.close()
        raise ValueError("No GLORYS grid points fall inside the dateline-safe local window.")
        
    glorys_lon_filtered = glorys_lon[glorys_lon_mask]
    glorys_lat_filtered = glorys_lat[glorys_lat_mask]
    glorys_depth_filtered = glorys_depth[glorys_depth_mask]

    # 存储多个变量的字典
    glorys_variables_filtered = {}
    for var in variables:
        if var == 'thetao':
            glorys_variables_filtered['thetao'] = needed_glorys_data.variables['thetao'][0, glorys_depth_mask, glorys_lat_mask, glorys_lon_mask]
            if depth is not None:
                glorys_variables_filtered['thetao'] = glorys_variables_filtered['thetao'][0, :, :]
        elif var == 'salinity' or var == 'so':
            glorys_variables_filtered['salinity'] = needed_glorys_data.variables['so'][0, glorys_depth_mask, glorys_lat_mask, glorys_lon_mask]
            if depth is not None:
                glorys_variables_filtered['salinity'] = glorys_variables_filtered['salinity'][0, :, :]
        elif var == 'u' or var == 'uo':
            glorys_variables_filtered['u'] = needed_glorys_data.variables['uo'][0, glorys_depth_mask, glorys_lat_mask, glorys_lon_mask]
            if depth is not None:
                glorys_variables_filtered['u'] = glorys_variables_filtered['u'][0, :, :]
        elif var == 'v' or var == 'vo':
            glorys_variables_filtered['v'] = needed_glorys_data.variables['vo'][0, glorys_depth_mask, glorys_lat_mask, glorys_lon_mask]
            if depth is not None:
                glorys_variables_filtered['v'] = glorys_variables_filtered['v'][0, :, :]
        elif var == 'ssh' or var == 'zos':
            glorys_variables_filtered['ssh'] = needed_glorys_data.variables['zos'][0, glorys_lat_mask, glorys_lon_mask]
        elif var == 'mlt' or var == 'mlotst':
            glorys_variables_filtered['mlt'] = needed_glorys_data.variables['mlotst'][0, glorys_lat_mask, glorys_lon_mask]
        else:
            raise ValueError(f"Unsupported variable: {var}. Supported variables are: 'thetao', 'so', 'uo', 'vo', 'zos', 'mlotst'.")

    needed_glorys_data.close()
    
    return glorys_lon_filtered, glorys_lat_filtered, glorys_depth_filtered, glorys_variables_filtered

def calculate_vorticity(lon, lat, u, v):
    '''
    计算给定速度场的相对涡度（zeta）和科里奥利参数（f）。

    可智能处理 2D/3D 以及 Masked Array 输入：u、v 维度应为 (latitude, longitude) 或
    (depth, latitude, longitude)；输入为 Masked Array 时输出也是 Masked Array，mask 边缘计算不准确的点
    会被自动 mask 掉。

    参数:
        - lon (np.ndarray): 一维经度数组（度）。
        - lat (np.ndarray): 一维纬度数组（度）。
        - u (np.ndarray | np.ma.MaskedArray): Zonal（纬向）速度数组。
        - v (np.ndarray | np.ma.MaskedArray): Meridional（经向）速度数组。

    返回:
        - tuple: (zeta, f) —— zeta 为相对涡度数组，f 为科里奥利参数数组（类型随输入，可为 MaskedArray）。
    '''
    # --- 1. 输入校验和维度处理 ---
    if u.shape[-2:] != (len(lat), len(lon)) or u.shape != v.shape:
        raise ValueError("速度数组的最后两个维度必须与经纬度数组的长度匹配，且u,v数组形状需一致。")
    if u.ndim not in [2, 3]:
        raise ValueError("输入速度场必须是 2D 或 3D 数组。")

    # --- 2. 智能处理 Masked Array ---
    # 检查输入是否为 masked array
    is_masked_input = np.ma.is_masked(u)

    if is_masked_input:
        # 如果是，用 NaN 填充被 mask 的位置。梯度计算会正确传播NaN。
        u_data = u.filled(np.nan)
        v_data = v.filled(np.nan)
    else:
        # 如果不是，直接使用原始数据
        u_data = u
        v_data = v

    # --- 3. 统一升维处理 ---
    original_ndim = u.ndim
    if original_ndim == 2:
        u_proc = u_data[np.newaxis, :, :]
        v_proc = v_data[np.newaxis, :, :]
    else:
        u_proc = u_data
        v_proc = v_data

    # --- 4. 计算物理坐标间距 (dx, dy) 和科里奥利参数 (f) ---
    Omega = 7.2921e-5  # 地球自转角速度 (弧度/秒)

    # lon, lat 是 1D 数组
    scale = approximate_degree_length(lat) # shape: (n_lat,)
    m_per_deg_lat = scale['meters_per_degree_lat']
    m_per_deg_lon = scale['meters_per_degree_lon']
    
    # 广播到 2D 网格 (lat, lon)
    # m_per_deg_lat 只随 lat 变化，沿 lon 轴广播
    # m_per_deg_lon 只随 lat 变化，沿 lon 轴广播
    m_per_deg_lat_2d = m_per_deg_lat[:, np.newaxis]
    m_per_deg_lon_2d = m_per_deg_lon[:, np.newaxis]
    
    # 计算梯度 (单位: 度)
    # np.gradient(lat) 返回 lat 方向的梯度 (dlat)
    # np.gradient(lon) 返回 lon 方向的梯度 (dlon)
    dlat_grid = np.gradient(lat)[:, np.newaxis] # shape (n_lat, 1)
    dlon_grid = np.gradient(lon)[np.newaxis, :] # shape (1, n_lon)
    
    # dy: 沿 axis=0 (lat) 的距离变化
    dy = dlat_grid * m_per_deg_lat_2d
    
    # dx: 沿 axis=1 (lon) 的距离变化
    dx = dlon_grid * m_per_deg_lon_2d
    
    # 计算科里奥利参数 f
    lon_rad, lat_rad = np.meshgrid(np.deg2rad(lon), np.deg2rad(lat))
    f_2d = 2 * Omega * np.sin(lat_rad)

    # 广播 f_2d 到 u_proc 和 v_proc 的深度维度，以便最终输出的 f 形状与 zeta 保持一致
    if original_ndim == 3:
        f_proc = np.tile(f_2d[np.newaxis, :, :], (u_proc.shape[0], 1, 1))
    else: # original_ndim == 2
        f_proc = f_2d # f_2d 已经是 (latitude, longitude) 形状

    # --- 5. 核心计算逻辑 ---
    vorticity_list = []
    for k_slice_idx in range(u_proc.shape[0]):
        u_slice = u_proc[k_slice_idx, :, :]
        v_slice = v_proc[k_slice_idx, :, :]
        dvdx = np.gradient(v_slice, axis=1) / dx
        dudy = np.gradient(u_slice, axis=0) / dy
        vorticity_list.append(dvdx - dudy)
    vorticity_result = np.stack(vorticity_list, axis=0)
    
    # --- 6. 根据输入类型，决定最终输出 ---
    if is_masked_input:
        # 如果输入是 masked，将结果中的 NaN 转回为 mask
        zeta_final = np.ma.masked_invalid(vorticity_result, copy=False)
        f_final = np.ma.masked_invalid(f_proc, copy=False) # f 也可能是 masked array
    else:
        zeta_final = vorticity_result
        f_final = f_proc

    # 如果原始输入是 2D，则降维回去
    if original_ndim == 2:
        return zeta_final.squeeze(axis=0), f_final
    else:
        return zeta_final, f_final
    
def get_idx(DS: list, no: int, start_date: str, end_date: str = None) -> int | list | None:
    '''
    获取指定涡旋编号在给定时间或时间范围内的索引。

    只提供 start_date 时返回该日期对应的单个索引；同时提供 start_date 与 end_date 时返回该时间范围内的
    索引列表。

    参数:
        - DS (list): 涡旋轨迹数据集。
        - no (int): 涡旋编号。
        - start_date (str): 起始日期，格式 'YYYY-MM-DD'。
        - end_date (str | None): 结束日期，格式 'YYYY-MM-DD'，默认 None。

    返回:
        - int | list | None: end_date 为 None 时返回单个整数索引或 None（未找到）；提供 end_date 时返回整数索引列表。
    '''
    wanted_track = find_track(DS, no)
    if wanted_track is None or len(wanted_track) == 0:
        print(f"未找到涡旋 {no} 的轨迹数据。")
        # 根据调用模式返回正确的空值
        return None if end_date is None else []

    start_date_dt = pd.to_datetime(start_date)

    # 情况一：只查找单个日期的索引
    if end_date is None:
        for idx, track_point in wanted_track.iterrows():
            track_date = convert_date(track_point['time'])
            # 精确匹配日期，忽略时间部分
            if track_date.date() == start_date_dt.date():
                return idx  # 找到后立即返回单个索引
        
        # 如果循环结束仍未找到匹配的日期
        print(f"在涡旋 {no} 的轨迹中未找到日期 {start_date}。")
        return None

    # 情况二：查找一个日期范围内的索引列表
    else:
        end_date_dt = pd.to_datetime(end_date)
        idx_list = []
        for idx, track_point in wanted_track.iterrows():
            track_date = convert_date(track_point['time'])
            if start_date_dt <= track_date <= end_date_dt:
                idx_list.append(idx)
        
        return idx_list

def get_vertical_glorys(DS: list, no: int, needed_date: str | pd.Timestamp,
                        k: float | list[float] | None = None,
                        b: float | list[float] | None = None,
                        variables: list = ['vorticity'],
                        x_min_km: float | None = None,
                        x_max_km: float | None = None,
                        profile_spacing_km: float | None = None,
                        interpolate_z: bool = True,
                        profile_depth_spacing_m: float | None = None) -> list[dict]:
    '''
    计算并返回指定涡旋在特定时刻沿一条或多条 y=kx+b 剖面的物理量数据。

    封装从 GLORYS 数据场提取剖面数据的核心插值计算，以字典列表返回；能正确处理三维变量（如温度，返回
    二维垂直剖面）和二维变量（如混合层深度，返回一维水平剖面）。派生变量 density/sigma/sigma0 统一表示
    势密度异常 σ0（kg/m³）；无论输入变量用何种别名，输出字典的键都是标准化变量名。

    参数:
        - DS (list): 包含涡旋轨迹信息的数据集。
        - no (int): 涡旋唯一编号。
        - needed_date (str | pd.Timestamp): 涡旋轨迹日期（'YYYY-MM-DD' 或时间戳）。
        - k (float | list[float]): 直线方程 y=kx+b 的斜率或斜率列表。
        - b (float | list[float]): 直线方程 y=kx+b 的截距或截距列表。
        - variables (list): 需要提取的 GLORYS 物理变量列表，默认 ['vorticity']。
        - x_min_km (float | None): 横坐标采样下界（km，中心投影为 0）；与 x_max_km 都为 None 时用剖面线在局地窗口内的全范围，都给定时先裁剪采样点再插值。
        - x_max_km (float | None): 横坐标采样上界（km），与 x_min_km 配套。
        - profile_spacing_km (float | None): 剖面采样步长（km）；None 时用 processing.yml 的 vertical_profile_spacing_km，给定正数则覆盖。
        - interpolate_z (bool): 是否将 z 轴重采样为等间距网格，默认 True。
        - profile_depth_spacing_m (float | None): z 轴重采样步长（m）；interpolate_z=True 且为 None 时用 processing.yml 的 vertical_profile_depth_spacing_m，给定正数则覆盖。

    返回:
        - list[dict]: 一个或多个剖面结果字典的列表，每个字典含以下键：

            - 'profile_data' (dict)：键为标准化物理量名（如 'salinity'、'mlt'）；三维变量（如 'vorticity'、'thetao'）的值为二维 MaskedArray（深度层数 × 剖面水平点数）的垂直剖面，二维变量（如 'mlt'、'ssh'）的值为一维 ndarray（长度=剖面水平点数）的水平分布。
            - 'y_coords' (np.ndarray)：剖面横坐标轴（物理距离，km），0 点对应涡旋中心在剖面线上的投影。
            - 'z_coords' (np.ndarray)：剖面纵坐标轴（深度，m）；interpolate_z=True 时为等间距，否则为 GLORYS 原始深度层。
            - 'lon_coords' / 'lat_coords' (np.ndarray)：剖面线上每个点（对应 y_coords）的经度/纬度。
            - 'projections' (dict)：涡旋边界在 y_coords 上的投影；键 'radius'（有效半径）/'contour'（有效轮廓），值为交点位置（km）列表。
            - 'metadata' (dict)：涡旋元数据，含 'eddy_no'(int)、'date_str'(str, 'YYYY-MM-DD')、'k'(float)、'b'(float)。
    '''
    # --- 0. 准备工作：统一输入格式并获取公共数据 ---
    if k is None and b is None:
        k_list, b_list = [], []
    elif k is None or b is None:
        raise ValueError("k 和 b 必须同时提供，或同时省略。")
    else:
        k_list = [k] if isinstance(k, (int, float)) else k
        b_list = [b] if isinstance(b, (int, float)) else b

    if len(k_list) != len(b_list):
        raise ValueError("k 和 b 的列表长度必须一致。")

    if profile_spacing_km is None:
        profile_spacing_km = _default_vertical_profile_spacing_km
    profile_spacing_km = float(profile_spacing_km)
    if (not np.isfinite(profile_spacing_km)) or profile_spacing_km <= 0:
        raise ValueError(f"profile_spacing_km must be a positive finite number, got {profile_spacing_km}.")

    interpolate_z = bool(interpolate_z)
    if interpolate_z:
        if profile_depth_spacing_m is None:
            profile_depth_spacing_m = _default_vertical_profile_depth_spacing_m
        profile_depth_spacing_m = float(profile_depth_spacing_m)
        if (not np.isfinite(profile_depth_spacing_m)) or profile_depth_spacing_m <= 0:
            raise ValueError(
                f"profile_depth_spacing_m must be a positive finite number, got {profile_depth_spacing_m}."
            )

    # 建立别名到标准名，以及变量维度的映射
    alias_map = {
        'thetao': 'thetao',
        'salinity': 'salinity', 'so': 'salinity',
        'density': 'sigma', 'sigma': 'sigma', 'sigma0': 'sigma',
        'u': 'u', 'uo': 'u',
        'v': 'v', 'vo': 'v',
        'ssh': 'ssh', 'zos': 'ssh',
        'mlt': 'mlt', 'mlotst': 'mlt',
        'vorticity': 'vorticity',
        'pv': 'pv',
    }
    var_dims = {
        'thetao': 3, 'salinity': 3, 'sigma': 3, 'u': 3, 'v': 3, 'vorticity': 3,
        'pv': 3,
        'ssh': 2, 'mlt': 2
    }

    raw_vars_to_fetch = set()
    for var in variables:
        standard_name = alias_map.get(var, var)
        if standard_name == 'vorticity':
            raw_vars_to_fetch.update(['u', 'v'])
        elif standard_name == 'pv':
            raw_vars_to_fetch.update(['u', 'v', 'salinity', 'thetao'])
        elif standard_name == 'sigma':
            raw_vars_to_fetch.update(['salinity', 'thetao'])
        else:
            raw_vars_to_fetch.add(var)

    try:
        track_df, ds_name, _ds_source_for_filter = _resolve_track_context(DS, no, include_contours=True)
    except Exception as exc:
        print(f"  - Error: {exc}")
        return [{} for _ in k_list]
    if track_df is None or track_df.empty:
        print(f"未找到涡旋 {no} 的轨迹数据。")
        return [{} for _ in k_list]

    ds_name_upper = ds_name.upper() if isinstance(ds_name, str) else "UNKNOWN"

    dates = track_df['date'] if 'date' in track_df.columns else convert_date(track_df['time'])
    dates = pd.to_datetime(dates, errors='coerce')
    try:
        needed_ts = pd.Timestamp(needed_date).normalize()
    except Exception as exc:
        raise ValueError(f"needed_date={needed_date!r} is not a valid date.") from exc

    same_day_idx = np.nonzero(dates.dt.normalize().to_numpy() == needed_ts.to_datetime64())[0]
    if same_day_idx.size == 0:
        raise ValueError(f"Date {needed_ts.strftime('%Y-%m-%d')} not found in track {no}.")
    needed_idx = int(same_day_idx[0])
    contour_lon_col = track_df['contour_lon'].to_numpy()
    contour_lat_col = track_df['contour_lat'].to_numpy()
    center_lon_arr = track_df['center_lon'].to_numpy()
    center_lat_arr = track_df['center_lat'].to_numpy()
    radius_arr = track_df['radius'].to_numpy()

    if not k_list and not b_list:
        k_list, b_list = [0.0], [float(center_lat_arr[int(needed_idx)])]

    lon_flat: list[np.ndarray] = []
    lat_flat: list[np.ndarray] = []
    current_center_lon_ref = float(center_lon_arr[int(needed_idx)])
    for lon_arr, lat_arr in zip(contour_lon_col, contour_lat_col):
        try:
            lon_np = np.asarray(lon_arr, dtype=float).ravel()
            lat_np = np.asarray(lat_arr, dtype=float).ravel()
        except Exception:
            continue
        if lon_np.size and lat_np.size:
            lon_local = current_center_lon_ref + _minimal_lon_diff_deg(lon_np, current_center_lon_ref)
            lon_flat.append(lon_local)
            lat_flat.append(lat_np)

    pad_deg = 0.5
    if lon_flat and lat_flat:
        lon_stack = np.concatenate(lon_flat)
        lat_stack = np.concatenate(lat_flat)
        lon_stack = lon_stack[lon_stack != 180.0]
        lat_stack = lat_stack[lat_stack != 0.0]
        glorys_lon_min = lon_stack.min() - pad_deg if lon_stack.size else center_lon_arr.min() - pad_deg
        glorys_lon_max = lon_stack.max() + pad_deg if lon_stack.size else center_lon_arr.max() + pad_deg
        glorys_lat_min = lat_stack.min() - pad_deg if lat_stack.size else center_lat_arr.min() - pad_deg
        glorys_lat_max = lat_stack.max() + pad_deg if lat_stack.size else center_lat_arr.max() + pad_deg
    else:
        glorys_lon_min = center_lon_arr.min() - pad_deg
        glorys_lon_max = center_lon_arr.max() + pad_deg
        glorys_lat_min = center_lat_arr.min() - pad_deg
        glorys_lat_max = center_lat_arr.max() + pad_deg

    lon_min_local_override = None
    lon_max_local_override = None
    lat_min_override = None
    lat_max_override = None
    if x_min_km is not None and x_max_km is not None:
        x_left = float(min(x_min_km, x_max_km))
        x_right = float(max(x_min_km, x_max_km))
        max_abs_range_km = max(abs(x_left), abs(x_right))
        if max_abs_range_km > 0:
            current_center_lat_ref = float(center_lat_arr[int(needed_idx)])
            window_pad_deg = 0.2
            for k_val, b_val in zip(k_list, b_list):
                if np.isclose(k_val, 0.0, atol=1e-12):
                    xp, yp = current_center_lon_ref, float(b_val)
                else:
                    xp = (current_center_lon_ref + k_val * current_center_lat_ref - k_val * b_val) / (1 + k_val ** 2)
                    yp = k_val * xp + b_val

                scale = approximate_degree_length(yp)
                meters_per_degree_lon = float(scale['meters_per_degree_lon'])
                meters_per_degree_lat = float(scale['meters_per_degree_lat'])
                ds_per_dlon = np.hypot(meters_per_degree_lon, k_val * meters_per_degree_lat)
                if (not np.isfinite(ds_per_dlon)) or ds_per_dlon <= 0:
                    continue

                dlon_half = (max_abs_range_km * 1000.0) / ds_per_dlon
                lon_a = xp - dlon_half
                lon_b = xp + dlon_half
                lat_a = k_val * lon_a + b_val
                lat_b = k_val * lon_b + b_val

                cur_lon_min = min(lon_a, lon_b) - window_pad_deg
                cur_lon_max = max(lon_a, lon_b) + window_pad_deg
                cur_lat_min = min(lat_a, lat_b) - window_pad_deg
                cur_lat_max = max(lat_a, lat_b) + window_pad_deg

                lon_min_local_override = cur_lon_min if lon_min_local_override is None else min(lon_min_local_override, cur_lon_min)
                lon_max_local_override = cur_lon_max if lon_max_local_override is None else max(lon_max_local_override, cur_lon_max)
                lat_min_override = cur_lat_min if lat_min_override is None else min(lat_min_override, cur_lat_min)
                lat_max_override = cur_lat_max if lat_max_override is None else max(lat_max_override, cur_lat_max)
    
    glorys_lon_raw, glorys_lat_raw, glorys_depth_raw, glorys_data_raw = get_track_area_glorys(
        DS,
        no,
        needed_ts,
        variables=list(raw_vars_to_fetch),
        lon_min_local=lon_min_local_override,
        lon_max_local=lon_max_local_override,
        lat_min=lat_min_override,
        lat_max=lat_max_override,
    )

    # 在当前中心经度参考系下构造连续经度轴，避免跨日界线导致插值坐标断裂
    if glorys_lon_raw.size:
        glorys_lon_local = current_center_lon_ref + _minimal_lon_diff_deg(glorys_lon_raw, current_center_lon_ref)
        lon_order = np.argsort(glorys_lon_local)
        glorys_lon_raw = glorys_lon_local[lon_order]

        for _name, _arr in list(glorys_data_raw.items()):
            if _arr is None:
                continue
            arr_ma = np.ma.array(_arr, copy=False)
            if arr_ma.ndim >= 2 and arr_ma.shape[-1] == len(lon_order):
                glorys_data_raw[_name] = arr_ma[..., lon_order]

    if glorys_lon_raw.size:
        glorys_lon_min = float(np.min(glorys_lon_raw))
        glorys_lon_max = float(np.max(glorys_lon_raw))
    if glorys_lat_raw.size:
        glorys_lat_min = float(np.min(glorys_lat_raw))
        glorys_lat_max = float(np.max(glorys_lat_raw))
    
    if glorys_depth_raw.size == 0 and not all(var_dims.get(alias_map.get(v, v)) == 2 for v in variables):
        return [{} for _ in k_list]

    sigma_3d_cache = None
    if any(alias_map.get(v, v) in ('sigma', 'pv') for v in variables):
        sal_3d = glorys_data_raw.get('salinity')
        theta_3d = glorys_data_raw.get('thetao')
        if sal_3d is not None and theta_3d is not None:
            sal_ma = np.ma.array(sal_3d, copy=False)
            theta_ma = np.ma.array(theta_3d, copy=False)
            if sal_ma.ndim == 2:
                sal_ma = sal_ma[np.newaxis, :, :]
            if theta_ma.ndim == 2:
                theta_ma = theta_ma[np.newaxis, :, :]

            if sal_ma.shape == theta_ma.shape and sal_ma.ndim == 3:
                try:
                    z3, lat3, lon3 = np.meshgrid(
                        glorys_depth_raw,
                        glorys_lat_raw,
                        _normalize_lon_array(glorys_lon_raw),
                        indexing='ij'
                    )
                    sal_filled = np.ma.filled(sal_ma, np.nan)
                    theta_filled = np.ma.filled(theta_ma, np.nan)
                    p3 = gsw.p_from_z(-z3, lat3)
                    SA = gsw.SA_from_SP(sal_filled, p3, lon3, lat3)
                    CT = gsw.CT_from_pt(SA, theta_filled)
                    sigma0 = gsw.sigma0(SA, CT)

                    input_mask = np.ma.getmaskarray(sal_ma) | np.ma.getmaskarray(theta_ma)
                    sigma_3d_cache = np.ma.masked_invalid(sigma0)
                    if input_mask.any():
                        sigma_3d_cache = np.ma.array(
                            sigma_3d_cache,
                            mask=np.ma.getmaskarray(sigma_3d_cache) | input_mask,
                            copy=False,
                        )
                except Exception:
                    sigma_3d_cache = None

    all_profiles_data = []

    # --- 开始循环，为每一对 k, b 计算一个剖面 ---
    for k_val, b_val in zip(k_list, b_list):
        # --- 1. 计算水平剖面线：先求与窗口交线，再按固定距离步长采样 ---
        intersection_pts: list[tuple[float, float]] = []
        tol = 1e-10

        if np.isclose(k_val, 0.0, atol=1e-12):
            if glorys_lat_min - tol <= b_val <= glorys_lat_max + tol:
                intersection_pts = [(float(glorys_lon_min), float(b_val)), (float(glorys_lon_max), float(b_val))]
        else:
            # 与左右边界相交
            lat_left = k_val * glorys_lon_min + b_val
            if glorys_lat_min - tol <= lat_left <= glorys_lat_max + tol:
                intersection_pts.append((float(glorys_lon_min), float(lat_left)))

            lat_right = k_val * glorys_lon_max + b_val
            if glorys_lat_min - tol <= lat_right <= glorys_lat_max + tol:
                intersection_pts.append((float(glorys_lon_max), float(lat_right)))

            # 与上下边界相交
            lon_bottom = (glorys_lat_min - b_val) / k_val
            if glorys_lon_min - tol <= lon_bottom <= glorys_lon_max + tol:
                intersection_pts.append((float(lon_bottom), float(glorys_lat_min)))

            lon_top = (glorys_lat_max - b_val) / k_val
            if glorys_lon_min - tol <= lon_top <= glorys_lon_max + tol:
                intersection_pts.append((float(lon_top), float(glorys_lat_max)))

        unique_pts: list[tuple[float, float]] = []
        for p_lon, p_lat in intersection_pts:
            duplicated = any(
                abs(float(_minimal_lon_diff_deg(p_lon, q_lon))) < 1e-8 and abs(p_lat - q_lat) < 1e-8
                for q_lon, q_lat in unique_pts
            )
            if not duplicated:
                unique_pts.append((p_lon, p_lat))

        if len(unique_pts) < 2:
            all_profiles_data.append({})
            continue

        if len(unique_pts) == 2:
            p0, p1 = unique_pts[0], unique_pts[1]
        else:
            best_pair = None
            best_dist = -np.inf
            for i in range(len(unique_pts) - 1):
                for j in range(i + 1, len(unique_pts)):
                    d_ij = local_xy_distance_m(
                        unique_pts[j][0],
                        unique_pts[j][1],
                        unique_pts[i][0],
                        unique_pts[i][1],
                        wrap_dateline=True,
                    )
                    if np.isfinite(d_ij) and d_ij > best_dist:
                        best_dist = float(d_ij)
                        best_pair = (unique_pts[i], unique_pts[j])

            if best_pair is None:
                all_profiles_data.append({})
                continue
            p0, p1 = best_pair

        # 固定剖面方向：左侧对应更小经度，避免同一条剖面在不同场景下镜像翻转。
        dlon_p0_to_p1 = float(_minimal_lon_diff_deg(p1[0], p0[0]))
        if dlon_p0_to_p1 < 0 or (abs(dlon_p0_to_p1) <= 1e-12 and p1[1] < p0[1]):
            p0, p1 = p1, p0

        segment_len_m = local_xy_distance_m(p1[0], p1[1], p0[0], p0[1], wrap_dateline=True)
        if (not np.isfinite(segment_len_m)) or segment_len_m <= 0:
            all_profiles_data.append({})
            continue

        n_samples = max(2, int(np.ceil(segment_len_m / (profile_spacing_km * 1000.0))) + 1)
        t = np.linspace(0.0, 1.0, n_samples)
        dlon_total = float(_minimal_lon_diff_deg(p1[0], p0[0]))
        if dlon_total < 0:
            p0, p1 = p1, p0
            dlon_total = float(_minimal_lon_diff_deg(p1[0], p0[0]))
        profile_lons_full = p0[0] + t * dlon_total
        profile_lats_full = p0[1] + t * (p1[1] - p0[1])

        dlat_deg = np.diff(profile_lats_full)
        dlon_deg = _minimal_lon_diff_deg(profile_lons_full[1:], profile_lons_full[:-1])
        mid_lats_deg = (profile_lats_full[:-1] + profile_lats_full[1:]) / 2
        
        scale_mid = approximate_degree_length(mid_lats_deg)
        dist_segments = np.hypot(
            dlon_deg * scale_mid['meters_per_degree_lon'],
            dlat_deg * scale_mid['meters_per_degree_lat']
        )
        y_coords_raw_full = np.insert(np.cumsum(dist_segments), 0, 0) / 1000.0

        current_center_lon, current_center_lat = center_lon_arr[needed_idx], center_lat_arr[needed_idx]
        if k_val == 0:
            xp, yp = current_center_lon, b_val
        else:
            xp = (current_center_lon + k_val * current_center_lat - k_val * b_val) / (1 + k_val**2)
            yp = k_val * xp + b_val
        center_idx_on_profile = np.argmin(_minimal_lon_diff_deg(profile_lons_full, xp)**2 + (profile_lats_full - yp)**2)
        y_coords_recenter_full = y_coords_raw_full - y_coords_raw_full[center_idx_on_profile]

        profile_lons = profile_lons_full
        profile_lats = profile_lats_full
        y_coords_recenter = y_coords_recenter_full
        if x_min_km is not None and x_max_km is not None:
            x_left = float(min(x_min_km, x_max_km))
            x_right = float(max(x_min_km, x_max_km))
            span_km = x_right - x_left
            n_steps = int(np.floor(span_km / profile_spacing_km))
            if n_steps < 1:
                all_profiles_data.append({})
                continue
            y_target = x_left + np.arange(n_steps + 1) * profile_spacing_km

            y_full_min = float(np.nanmin(y_coords_recenter_full))
            y_full_max = float(np.nanmax(y_coords_recenter_full))
            valid_target = (y_target >= y_full_min) & (y_target <= y_full_max)
            if np.count_nonzero(valid_target) < 2:
                all_profiles_data.append({})
                continue

            y_coords_recenter = y_target[valid_target]
            profile_lons = np.interp(y_coords_recenter, y_coords_recenter_full, profile_lons_full)
            profile_lats = np.interp(y_coords_recenter, y_coords_recenter_full, profile_lats_full)

        # --- 2. 区分维度，计算剖面数据 ---
        z_coords_native = glorys_depth_raw
        profile_data_dict = {}
        for var in variables:
            standard_name = alias_map.get(var, var)
            dimension = var_dims.get(standard_name, 3)

            # --- 2.1 处理2D变量 (如 mlt, ssh) ---
            if dimension == 2:
                data_2d = glorys_data_raw.get(standard_name)
                if data_2d is None or data_2d.size == 0 or np.all(np.ma.getmask(data_2d)):
                    profile_data_dict[standard_name] = np.ma.masked_all(len(profile_lons))
                    continue
                
                data_2d_filled = np.ma.filled(np.ma.array(data_2d, copy=False), np.nan)
                interp_func_2d = RegularGridInterpolator((glorys_lat_raw, glorys_lon_raw), data_2d_filled,
                                                         method='linear', bounds_error=False, fill_value=np.nan)
                interpolated_1d = interp_func_2d(list(zip(profile_lats, profile_lons)))
                profile_data_dict[standard_name] = np.ma.masked_invalid(interpolated_1d)
                continue

            # --- 2.2 处理3D变量 (如 thetao, vorticity) ---
            glorys_variable_3d = None
            if standard_name == 'vorticity':
                u, v = glorys_data_raw.get('u'), glorys_data_raw.get('v')
                if u is not None and v is not None and u.size > 0 and v.size > 0:
                    if u.ndim == 2: u, v = u[np.newaxis, :, :], v[np.newaxis, :, :]
                    zeta_3d, f_3d = calculate_vorticity(glorys_lon_raw, glorys_lat_raw, u, v)
                    glorys_variable_3d = zeta_3d / f_3d
            elif standard_name == 'sigma':
                glorys_variable_3d = sigma_3d_cache
            elif standard_name == 'pv':
                u = glorys_data_raw.get('u')
                v = glorys_data_raw.get('v')
                if (u is not None and v is not None and u.size > 0 and v.size > 0
                        and sigma_3d_cache is not None):
                    if u.ndim == 2:
                        u, v = u[np.newaxis, :, :], v[np.newaxis, :, :]
                    zeta_3d, f_3d = calculate_vorticity(glorys_lon_raw, glorys_lat_raw, u, v)
                    z3 = glorys_depth_raw[:, np.newaxis, np.newaxis]
                    dz = np.gradient(z3, axis=0)
                    with np.errstate(divide='ignore', invalid='ignore'):
                        dsigma_dz = np.gradient(sigma_3d_cache, axis=0) / np.where(dz != 0, dz, np.nan)
                    N2 = (9.81 / 1025.0) * dsigma_dz
                    N2 = np.maximum(N2, 0.0)  # clip convective instability (N² < 0)
                    glorys_variable_3d = (f_3d + zeta_3d) * N2 / 9.81
            else:
                glorys_variable_3d = glorys_data_raw.get(standard_name)

            if glorys_variable_3d is None or glorys_variable_3d.size == 0 or np.all(np.ma.getmask(glorys_variable_3d)):
                profile_data_dict[standard_name] = np.ma.masked_all((len(z_coords_native), len(profile_lons)))
                continue

            if glorys_variable_3d.ndim == 2: glorys_variable_3d = glorys_variable_3d[np.newaxis, :, :]
            
            query_depths, query_lats = np.meshgrid(z_coords_native, profile_lats, indexing='ij')
            _, query_lons = np.meshgrid(z_coords_native, profile_lons, indexing='ij')
            xi_points = np.vstack([query_depths.ravel(), query_lats.ravel(), query_lons.ravel()]).T
            
            variable_3d_filled = np.ma.filled(np.ma.array(glorys_variable_3d, copy=False), np.nan)
            interp_func = RegularGridInterpolator((z_coords_native, glorys_lat_raw, glorys_lon_raw), variable_3d_filled, method='linear', bounds_error=False, fill_value=np.nan)
            interpolated_values_flat = interp_func(xi_points)
            profile_data_dict[standard_name] = np.ma.masked_invalid(interpolated_values_flat.reshape(len(z_coords_native), len(profile_lons)))

        z_coords = z_coords_native
        if interpolate_z and z_coords_native.size >= 2:
            z_min = float(z_coords_native[0])
            z_max = float(z_coords_native[-1])
            z_span = z_max - z_min
            n_steps_z = int(np.floor(z_span / profile_depth_spacing_m))
            if n_steps_z >= 1:
                z_coords_target = z_min + np.arange(n_steps_z + 1) * profile_depth_spacing_m

                zz_grid, yy_grid = np.meshgrid(z_coords_target, y_coords_recenter, indexing='ij')
                query_points_2d = np.vstack([zz_grid.ravel(), yy_grid.ravel()]).T

                for key_name, var_data in list(profile_data_dict.items()):
                    var_ma = np.ma.array(var_data, copy=False)
                    if var_ma.ndim != 2 or var_ma.shape[0] != len(z_coords_native):
                        continue
                    if np.ma.getmaskarray(var_ma).all():
                        profile_data_dict[key_name] = np.ma.masked_all((len(z_coords_target), len(y_coords_recenter)))
                        continue

                    interp_func_2d = RegularGridInterpolator(
                        (z_coords_native, y_coords_recenter),
                        np.ma.filled(var_ma, np.nan),
                        method='linear',
                        bounds_error=False,
                        fill_value=np.nan,
                    )
                    new_vals_flat = interp_func_2d(query_points_2d)
                    profile_data_dict[key_name] = np.ma.masked_invalid(
                        new_vals_flat.reshape(len(z_coords_target), len(y_coords_recenter))
                    )

                z_coords = z_coords_target

        # --- 3. 计算边界投影 ---
        # 红线投影与 horizontal 图保持同一几何：使用经向/纬向不同尺度的等效椭圆求交
        scale_line = approximate_degree_length(current_center_lat)
        semi_axis_lon_deg = radius_arr[needed_idx] / float(scale_line['meters_per_degree_lon'])
        semi_axis_lat_deg = radius_arr[needed_idx] / float(scale_line['meters_per_degree_lat'])

        radius_intersections_lon: list[float] = []
        if semi_axis_lon_deg > 0 and semi_axis_lat_deg > 0 and np.isfinite(semi_axis_lon_deg) and np.isfinite(semi_axis_lat_deg):
            inv_a2 = 1.0 / (semi_axis_lon_deg ** 2)
            inv_b2 = 1.0 / (semi_axis_lat_deg ** 2)
            dy0 = b_val - current_center_lat

            # ((x-x0)^2/a^2) + ((k*x+b-y0)^2/b^2) = 1
            A = inv_a2 + (k_val ** 2) * inv_b2
            B = -2.0 * current_center_lon * inv_a2 + 2.0 * k_val * dy0 * inv_b2
            C = (current_center_lon ** 2) * inv_a2 + (dy0 ** 2) * inv_b2 - 1.0

            discriminant = B ** 2 - 4.0 * A * C
            if discriminant >= -1e-12 and A > 0:
                discriminant = max(0.0, float(discriminant))
                sqrt_disc = float(np.sqrt(discriminant))
                radius_intersections_lon = [
                    (-B + sqrt_disc) / (2.0 * A),
                    (-B - sqrt_disc) / (2.0 * A),
                ]

        radius_proj_dists = [
            y_coords_raw_full[np.argmin(_minimal_lon_diff_deg(profile_lons_full, lon_i)**2 + (profile_lats_full - (k_val * lon_i + b_val))**2)]
            - y_coords_raw_full[center_idx_on_profile]
            for lon_i in radius_intersections_lon
        ]

        curr_contour_lon = np.asarray(contour_lon_col[needed_idx], dtype=float).ravel()
        curr_contour_lat = np.asarray(contour_lat_col[needed_idx], dtype=float).ravel()
        valid_mask = (curr_contour_lon != 180.0) & (curr_contour_lat != 0.0)
        curr_contour_lon = curr_contour_lon[valid_mask]
        curr_contour_lat = curr_contour_lat[valid_mask]
        curr_contour_lon = current_center_lon + _minimal_lon_diff_deg(curr_contour_lon, current_center_lon)
        contour_intersections_xy = find_polygon_line_intersections(curr_contour_lon, curr_contour_lat, profile_lons_full, profile_lats_full)
        contour_proj_dists = [
            y_coords_raw_full[np.argmin(_minimal_lon_diff_deg(profile_lons_full, lon_i)**2 + (profile_lats_full - lat_i)**2)]
            - y_coords_raw_full[center_idx_on_profile]
            for lon_i, lat_i in contour_intersections_xy
        ]

        projections_dict = {'radius': sorted(radius_proj_dists), 'contour': sorted(contour_proj_dists)}
        
        # --- 4. 整合单个剖面的结果 ---
        single_profile_result = {
            'profile_data': profile_data_dict,
            'y_coords': y_coords_recenter,
            'z_coords': z_coords,
            'lon_coords': profile_lons,
            'lat_coords': profile_lats,
            'projections': projections_dict,
            'metadata': {
                'eddy_no': no,
                'date_str': dates[needed_idx].strftime("%Y-%m-%d"),
                'k': k_val,
                'b': b_val,
                'ds_name': ds_name_upper,
            }
        }
        all_profiles_data.append(single_profile_result)

    # --- 5. 返回所有剖面的结果列表 ---
    return all_profiles_data


def _draw_isolines(
    ax,
    Y_mesh: np.ndarray,
    Z_mesh: np.ndarray,
    data_filled: np.ndarray,
    clim: tuple[float, float] | None,
    *,
    isoline_levels: int | list[float] | np.ndarray | None = None,
    isoline_color: str = 'black',
    isoline_linewidth: float = 0.8,
    isoline_alpha: float = 0.45,
    label_isolines: bool = True,
    clabel_fontsize: int = 9,
    clabel_fmt: str = '%.3g',
) -> None:
    """在色斑图上叠加等值线，等值线级别自动适应数据范围。"""
    finite_vals = data_filled[np.isfinite(data_filled)]
    if finite_vals.size == 0:
        return

    levels_to_use = None
    if isoline_levels is None:
        lo = float(np.nanmin(finite_vals))
        hi = float(np.nanmax(finite_vals))
        if clim is not None:
            lo = max(lo, clim[0])
            hi = min(hi, clim[1])
        if hi > lo:
            levels_to_use = np.linspace(lo, hi, 9)
    elif isinstance(isoline_levels, (int, np.integer)):
        n_levels = int(isoline_levels)
        if clim is not None and n_levels >= 2:
            levels_to_use = np.linspace(clim[0], clim[1], n_levels)
    else:
        levels_to_use = np.asarray(isoline_levels, dtype=float)

    if levels_to_use is None:
        return
    levels_to_use = np.asarray(levels_to_use, dtype=float)
    levels_to_use = levels_to_use[np.isfinite(levels_to_use)]
    levels_to_use = np.unique(levels_to_use)
    if levels_to_use.size < 2:
        return

    cs = ax.contour(
        Y_mesh, Z_mesh, data_filled,
        levels=levels_to_use,
        colors=isoline_color,
        linewidths=float(isoline_linewidth),
        alpha=float(isoline_alpha),
        zorder=5,
    )
    if label_isolines:
        ax.clabel(cs, inline=True, fontsize=clabel_fontsize, fmt=clabel_fmt)


def _plot_vertical_glorys_core(DS: list | str | tuple | dict | None, no: int, needed_date: str | pd.Timestamp,
                     k: float | list[float], b: float | list[float],
                     variable: str = 'vorticity',
                     show_fig: bool = True, save_fig: bool = False,
                     xmin: float = -400.0, xmax: float = 400.0,
                     ymin: float = 0.0, ymax: float = 1000.0,
                     color_vmin: float | None = None,
                     color_vmax: float | None = None,
                     plot_mlt: bool = False,
                     plot_argo_projection: bool = True,
                     argo_projection_config: DetectionConfig | None = None,
                     argo_projection_min_depth: float | None = None,
                     plot_isolines: bool = True,
                     isoline_levels: int | list[float] | np.ndarray | None = None,
                     isoline_color: str = 'black',
                     isoline_linewidth: float = 0.8,
                     isoline_alpha: float = 0.45,
                     label_isolines: bool = False,
                     profile_spacing_km: float | None = None,
                     interpolate_z: bool = True,
                     profile_depth_spacing_m: float | None = None,
                     precomputed_data_packages: list[dict] | None = None,
                     precomputed_projected_argo_rows: pd.DataFrame | None = None,
                     projection_distance_scale_km: float | None = None,
                     x_axis_label: str = 'Distance from Eddy Center Projection (km)',
                     save_subdir: str = 'plot_track_vertical_glorys',
                     title_subject: str = 'Track'):
    '''
    共享的垂向剖面绘图核心：获取剖面数据并完成可视化。

    供 track/argo 两类入口函数复用，避免上层函数彼此委托。
    当 k 和 b 为列表时，会为每一对 (k, b) 分别绘制一张图。

    参数:
        DS (list): 涡旋数据集或类型标识（如 'acl'、'cl'）。
        no (int): 涡旋编号。
        needed_date (str | pd.Timestamp): 轨迹日期（'YYYY-MM-DD' 或时间戳）。
        k (float | list[float]): 剖面线斜率 y = kx + b。
        b (float | list[float]): 剖面线截距 y = kx + b。
        variable (str): 主图变量名（如 'vorticity'、'thetao'、'salinity'、'u'、'v'、'sigma'）。
        show_fig (bool): 是否显示图像。
        save_fig (bool): 是否保存图像。
        xmin, xmax (float | None): 横坐标范围（km）。当两者同时给定时：
            - 采样阶段先按该范围构造目标 y 网格；
            - 绘图阶段再应用 x 轴范围。
        ymin, ymax (float | None): 纵坐标显示范围（m，向下为正）。
        color_vmin, color_vmax (float | None): 主色斑图色标下/上限。
            - 默认 None：按当前剖面数据自适应；
            - 指定后：覆盖自动范围（可只给其中一个，另一个保持自动）。
        plot_mlt (bool): 是否叠加混合层深度线（MLD）。
        plot_argo_projection (bool): 是否叠加“同日期、涡旋内部”的 Argo 点投影。
        argo_projection_config: Argo 映射层异常筛选配置。
        argo_projection_min_depth (float | None): Argo 映射层使用的最小深度阈值；None 回退 `_cfg_anomaly_min_depth`。
        plot_isolines (bool): 是否在色斑图上叠加变量等值线。默认 True。
        isoline_levels (int | list[float] | np.ndarray | None): 等值线级别。
            - None: 自动按色标范围生成 9 条线；
            - int: 生成该数量的等间隔级别；
            - list/ndarray: 直接使用给定级别。
        isoline_color (str): 等值线颜色，默认 'black'。
        isoline_linewidth (float): 等值线线宽，默认 0.8。
        isoline_alpha (float): 等值线透明度，默认 0.45。
        label_isolines (bool): 是否标注等值线数值，默认 False。
        profile_spacing_km (float | None): 剖面采样步长（km），None 时使用 processing.yml 默认。
        interpolate_z (bool): 是否将 z 轴重采样为等间距网格。
            - True: 使用等间距深度网格绘图；
            - False: 使用 GLORYS 原始深度层。
        profile_depth_spacing_m (float | None): z 轴重采样步长（m），None 时使用 processing.yml 默认。

    返回:
        None。函数直接完成绘图、保存与显示，不返回图像对象。
    '''
    # --- 1. 获取所有计算好的剖面数据包 ---
    
    vars_to_fetch = {variable}
    if plot_mlt:
        vars_to_fetch.add('mlt')

    sampling_xmin = xmin if (xmin is not None and xmax is not None) else None
    sampling_xmax = xmax if (xmin is not None and xmax is not None) else None
    track_df_for_date = pd.DataFrame()
    needed_idx = 0

    try:
        needed_ts = pd.Timestamp(needed_date).normalize()
    except Exception as exc:
        raise ValueError(f"needed_date={needed_date!r} is not a valid date.") from exc

    if precomputed_data_packages is not None:
        all_data_packages = precomputed_data_packages
    else:
        track_df_for_date, _ds_name_for_date, _ds_source_for_date = _resolve_track_context(DS, no, include_contours=True)
        if 'date' in track_df_for_date.columns:
            track_dates_for_date = pd.to_datetime(track_df_for_date['date'], errors='coerce')
        else:
            track_dates_for_date = pd.to_datetime(convert_date(track_df_for_date['time']), errors='coerce')

        same_day_idx = np.nonzero(track_dates_for_date.dt.normalize().to_numpy() == needed_ts.to_datetime64())[0]
        if same_day_idx.size == 0:
            raise ValueError(f"Date {needed_ts.strftime('%Y-%m-%d')} not found in track {no}.")
        needed_idx = int(same_day_idx[0])
        needed_ts = pd.Timestamp(track_dates_for_date.iloc[needed_idx])

        all_data_packages = get_vertical_glorys(
            DS,
            no,
            needed_ts,
            k,
            b,
            variables=list(vars_to_fetch),
            x_min_km=sampling_xmin,
            x_max_km=sampling_xmax,
            profile_spacing_km=profile_spacing_km,
            interpolate_z=interpolate_z,
            profile_depth_spacing_m=profile_depth_spacing_m,
        )

    if not all_data_packages:
        print(f"警告: get_vertical_glorys 未能返回任何数据。绘图已取消。")
        return

    # Argo 投影层视觉参数
    _ARGO_MARKER_MAX_SIZE = 180.0
    _ARGO_MARKER_MIN_SIZE = 10.0

    # Argo 投影层阈值参数
    argo_projection_config = _resolve_detection_config(
        argo_projection_config,
        anomaly_min_depth=argo_projection_min_depth,
    )

    # --- 1.1 按需准备 Argo 映射数据（仅针对 needed_date 当天） ---
    projected_argo_rows = pd.DataFrame()
    default_distance_scale_km: float | None = (
        float(projection_distance_scale_km)
        if projection_distance_scale_km is not None and np.isfinite(float(projection_distance_scale_km)) and float(projection_distance_scale_km) > 0
        else None
    )

    if precomputed_projected_argo_rows is not None:
        projected_argo_rows = precomputed_projected_argo_rows.copy()
        if not projected_argo_rows.empty:
            projected_argo_rows['Longitude'] = pd.to_numeric(projected_argo_rows['Longitude'], errors='coerce')
            projected_argo_rows['Latitude'] = pd.to_numeric(projected_argo_rows['Latitude'], errors='coerce')
            projected_argo_rows['Depth'] = pd.to_numeric(projected_argo_rows['Depth'], errors='coerce')
            projected_argo_rows['DO'] = pd.to_numeric(projected_argo_rows.get('DO', np.nan), errors='coerce')
            projected_argo_rows = projected_argo_rows.dropna(subset=['Longitude', 'Latitude', 'Depth'])

    elif plot_argo_projection and precomputed_data_packages is None:
        try:
            track_df_overlay, _ds_name_overlay, ds_source_for_filter = _resolve_track_context(
                DS,
                no,
                include_contours=True
            )

            if not track_df_overlay.empty:
                if 'date' in track_df_overlay.columns:
                    track_dates_overlay = pd.to_datetime(track_df_overlay['date'], errors='coerce')
                else:
                    track_dates_overlay = pd.to_datetime(convert_date(track_df_overlay['time']), errors='coerce')

                target_date = None
                idx_int = int(needed_idx)
                if 0 <= idx_int < len(track_df_overlay):
                    target_date = pd.Timestamp(track_dates_overlay.iloc[idx_int]).normalize()
                    if default_distance_scale_km is None:
                        radius_m = pd.to_numeric(track_df_overlay['radius'], errors='coerce').iloc[idx_int]
                        if pd.notna(radius_m) and float(radius_m) > 0:
                            default_distance_scale_km = float(radius_m) / 1000.0

                argo_all = filtered_float_data(ds_source_for_filter, no, track=track_df_overlay)
                if not argo_all.empty:
                    argo_all = argo_all.copy()
                    argo_all['date'] = pd.to_datetime(argo_all[['Year', 'Month', 'Day']], errors='coerce').dt.normalize()
                    if target_date is not None:
                        projected_argo_rows = argo_all[argo_all['date'] == target_date].copy()
                    else:
                        projected_argo_rows = argo_all.copy()

                    projected_argo_rows['Longitude'] = pd.to_numeric(projected_argo_rows['Longitude'], errors='coerce')
                    projected_argo_rows['Latitude'] = pd.to_numeric(projected_argo_rows['Latitude'], errors='coerce')
                    projected_argo_rows['Depth'] = pd.to_numeric(projected_argo_rows['Depth'], errors='coerce')
                    projected_argo_rows['DO'] = pd.to_numeric(projected_argo_rows.get('DO', np.nan), errors='coerce')
                    projected_argo_rows = projected_argo_rows.dropna(subset=['Longitude', 'Latitude', 'Depth'])

                    # 每个 Profile 仅保留一个代表点：按当前异常配置筛选后取 anomaly_score 最强一条。
                    if not projected_argo_rows.empty and 'Profile_number' in projected_argo_rows.columns:
                        projected_argo_rows['Profile_number'] = pd.to_numeric(projected_argo_rows['Profile_number'], errors='coerce')
                        projected_argo_rows = projected_argo_rows.dropna(subset=['Profile_number']).copy()

                        if not projected_argo_rows.empty:
                            deltas = _reduce_argo_profiles_by_anomaly(
                                projected_argo_rows,
                                detection_config=argo_projection_config,
                            )
                            if not deltas.empty:
                                projected_argo_rows = deltas
                            else:
                                projected_argo_rows = pd.DataFrame()
        except Exception as exc:
            print(f"注意: Argo 剖面映射准备失败，已跳过投影层。原因: {exc}")
            projected_argo_rows = pd.DataFrame()

    # --- 开始循环，为每一个数据包生成一张图 ---
    for data_package in all_data_packages:
        if not data_package:
            print(f"警告: 收到一个空的数据包，跳过此剖面的绘图。")
            continue
            
        profile_variable_2d = data_package['profile_data'].get(variable)
        if profile_variable_2d is None:
            alias_map = {
                'so': 'salinity',
                'uo': 'u',
                'vo': 'v',
                'density': 'sigma',
                'sigma0': 'sigma',
            }
            standard_name = alias_map.get(variable, variable)
            profile_variable_2d = data_package['profile_data'].get(standard_name)

        if profile_variable_2d is None or np.all(np.ma.getmaskarray(np.ma.array(profile_variable_2d, copy=False))):
            k_meta, b_meta = data_package.get('metadata', {}).get('k'), data_package.get('metadata', {}).get('b')
            print(f"警告: 变量 '{variable}' 在剖面 k={k_meta}, b={b_meta} 上的数据无效。绘图已取消。")
            continue

        # --- 2. 准备绘图 ---
        y_coords = data_package['y_coords']
        z_coords = data_package['z_coords']
        projections = data_package['projections']
        metadata = data_package['metadata']

        ds_name = metadata.get('ds_name')
        if not ds_name:
            try:
                _, ds_name, _ = _resolve_track_context(DS, no, include_contours=False)
            except Exception:
                ds_name = "UNKNOWN"
        ds_name = ('Argo' if ds_name.lower() == 'argo' else ds_name.upper()) if isinstance(ds_name, str) else str(ds_name)

        prop_colors = plt.rcParams['axes.prop_cycle'].by_key()['color']
        eddy_color = prop_colors[1] if 'AC' in ds_name else prop_colors[0]
        
        if profile_variable_2d.shape[1] != len(y_coords):
            y_coords = y_coords[:profile_variable_2d.shape[1]]

        Y_mesh, Z_mesh = np.meshgrid(y_coords, z_coords)
        
        # 设置变量相关的绘图属性
        clim = None
        if variable == 'vorticity': cbar_label, cmap, clim = r'$\zeta/f$', 'seismic', (-0.3, 0.3)
        elif variable in ['thetao']: cbar_label, cmap = 'Temperature (°C)', 'rainbow'
        elif variable in ['salinity', 'so']: cbar_label, cmap = 'Salinity (psu)', 'viridis'
        elif variable in ['density', 'sigma', 'sigma0']: cbar_label, cmap = 'Potential Density Anomaly (σ0, kg/m³)', 'RdBu_r'
        elif variable in ['u', 'v', 'uo', 'vo']: cbar_label, cmap = 'Velocity (m/s)', 'RdBu_r'
        else: cbar_label, cmap = variable, 'viridis'
        
        if clim is None:
            prof_ma = np.ma.array(profile_variable_2d, copy=False)
            valid_values = prof_ma.compressed()
            clim = (valid_values.min(), valid_values.max()) if valid_values.size > 0 else (0,1)
            if variable in ['u', 'v', 'uo', 'vo']:
                max_abs = np.max(np.abs(valid_values)) if valid_values.size > 0 else 1
                clim = (-max_abs, max_abs)

        # 手动色标覆盖（默认保持自适应）
        vmin = float(color_vmin) if color_vmin is not None else float(clim[0])
        vmax = float(color_vmax) if color_vmax is not None else float(clim[1])
        if not np.isfinite(vmin) or not np.isfinite(vmax) or vmin >= vmax:
            raise ValueError(f"Invalid color range: color_vmin={color_vmin}, color_vmax={color_vmax}")
        clim = (vmin, vmax)

        # --- 3. 执行绘图 ---
        fig, ax = plt.subplots(figsize=(20, 15))
        
        date_str = metadata['date_str']
        entity_label = metadata.get('entity_label')
        if not entity_label:
            entity_label = f"{ds_name}{metadata['eddy_no']}"
        subject_label = f"{title_subject} {entity_label}".strip()
        title = (
            f"Vertical Profile of {cbar_label} for {subject_label} "
            f"on {date_str}, y={metadata['k']:.2f}x{metadata['b']:+.2f}"
        )
        ax.set_title(title, fontsize=20)
        ax.set_xlabel(x_axis_label, fontsize=18)
        ax.set_ylabel('Depth (m)', fontsize=18)
        ax.tick_params(labelsize=14)
        
        v_field_filled = np.ma.filled(np.ma.array(profile_variable_2d, copy=False), np.nan)
        pc = ax.pcolormesh(Y_mesh, Z_mesh, v_field_filled, cmap=cmap, shading='auto', vmin=clim[0], vmax=clim[1])

        if plot_isolines:
            _draw_isolines(ax, Y_mesh, Z_mesh, v_field_filled, clim,
                           isoline_levels=isoline_levels,
                           isoline_color=isoline_color,
                           isoline_linewidth=isoline_linewidth,
                           isoline_alpha=isoline_alpha,
                           label_isolines=label_isolines,
                           clabel_fontsize=10,
                           clabel_fmt='%.2g')

        cbar = fig.colorbar(pc, ax=ax, orientation='vertical', fraction=0.046, pad=0.04)
        cbar.set_label(cbar_label, fontsize=18)
        cbar.ax.tick_params(labelsize=14)

        draw_reference_lines = bool(metadata.get('draw_reference_lines', True))
        if draw_reference_lines:
            ax.axvline(0, color='black', linestyle='--', linewidth=2, label='Eddy Center Projection')
            for i, dist in enumerate(projections['radius']):
                ax.axvline(dist, color='r', linestyle='--', linewidth=2, label='Effective Radius Projection' if i == 0 else "")
            for i, dist in enumerate(projections['contour']):
                ax.axvline(dist, color=eddy_color, linestyle=':', linewidth=2, label='Effective Contour Projection' if i == 0 else "")
            
        # 绘制混合层深度
        mld_lines = () # 创建一个空元组，用于存放MLD的两条线
        if plot_mlt:
            mlt_data = data_package['profile_data'].get('mlt')
            if mlt_data is not None and not np.all(np.ma.getmaskarray(np.ma.array(mlt_data, copy=False))):
                # **捕获两条线的艺术家对象**
                # 注意 plot 返回的是一个列表，所以我们用 l, 来解包
                black_line, = ax.plot(y_coords, mlt_data, color='black', linewidth=2.5, zorder=3)
                white_line, = ax.plot(y_coords, mlt_data, color='white', linewidth=1.5, zorder=4, label='Mixed Layer Depth')
                mld_lines = (black_line, white_line) # 将两条线存入元组
            else:
                print(f"注意: 未能在剖面 k={metadata['k']}, b={metadata['b']} 上找到有效的混合层深度数据。")

        # 按需绘制 Argo 三维点在剖面平面的投影
        argo_scatter = None
        argo_has_do = False
        if plot_argo_projection and not projected_argo_rows.empty:
            try:
                k_line = float(metadata.get('k'))
                b_line = float(metadata.get('b'))

                line_lons = np.asarray(data_package.get('lon_coords', []), dtype=float)
                line_lats = np.asarray(data_package.get('lat_coords', []), dtype=float)
                line_y = np.asarray(y_coords, dtype=float)

                if line_lons.size > 1 and line_lats.size == line_lons.size and line_y.size >= line_lons.size:
                    line_y = line_y[:line_lons.size]

                    pts_lon = projected_argo_rows['Longitude'].to_numpy(dtype=float)
                    pts_lat = projected_argo_rows['Latitude'].to_numpy(dtype=float)
                    pts_depth = projected_argo_rows['Depth'].to_numpy(dtype=float)
                    pts_do = projected_argo_rows['DO'].to_numpy(dtype=float)

                    # 将地理点正交投影到剖面直线 y = kx + b（经纬度坐标）
                    denom = 1.0 + k_line ** 2
                    proj_lon = (pts_lon + k_line * (pts_lat - b_line)) / denom
                    proj_lat = k_line * proj_lon + b_line

                    # 在采样剖面线上寻找最近点，以读取剖面横轴坐标（km）
                    dist2 = (line_lons[None, :] - proj_lon[:, None]) ** 2 + (line_lats[None, :] - proj_lat[:, None]) ** 2
                    nearest_idx = np.argmin(dist2, axis=1)
                    proj_y = line_y[nearest_idx]

                    scale_ref = approximate_degree_length(np.nanmean(line_lats))
                    dlon_deg = _minimal_lon_diff_deg(pts_lon, proj_lon)
                    dx_m = dlon_deg * scale_ref['meters_per_degree_lon']
                    dy_m = (pts_lat - proj_lat) * scale_ref['meters_per_degree_lat']
                    dist_to_line_km = np.hypot(dx_m, dy_m) / 1000.0

                    scale_km = default_distance_scale_km
                    if scale_km is None or (not np.isfinite(scale_km)) or scale_km <= 0:
                        finite_span = np.asarray(line_y[np.isfinite(line_y)], dtype=float)
                        if finite_span.size:
                            scale_km = max(20.0, float(np.nanmax(np.abs(finite_span))))
                        else:
                            scale_km = 50.0

                    marker_max = max(float(_ARGO_MARKER_MAX_SIZE), 1.0)
                    marker_min = max(float(_ARGO_MARKER_MIN_SIZE), 1.0)
                    if marker_min > marker_max:
                        marker_min, marker_max = marker_max, marker_min

                    weight = 1.0 - np.clip(dist_to_line_km / scale_km, 0.0, 1.0)
                    marker_sizes = marker_min + (marker_max - marker_min) * (weight ** 1.2)

                    valid_pts = np.isfinite(proj_y) & np.isfinite(pts_depth)
                    if np.any(valid_pts):
                        if np.isfinite(pts_do[valid_pts]).any():
                            argo_scatter = ax.scatter(
                                proj_y[valid_pts],
                                pts_depth[valid_pts],
                                c=pts_do[valid_pts],
                                cmap='bwr',
                                vmin=150,
                                vmax=240,
                                s=marker_sizes[valid_pts],
                                edgecolors='black',
                                linewidths=0.4,
                                alpha=0.9,
                                zorder=6,
                                label='Projected Argo (one point/profile)',
                            )
                            argo_has_do = True
                        else:
                            argo_scatter = ax.scatter(
                                proj_y[valid_pts],
                                pts_depth[valid_pts],
                                color='blue',
                                s=marker_sizes[valid_pts],
                                edgecolors='black',
                                linewidths=0.4,
                                alpha=0.9,
                                zorder=6,
                                label='Projected Argo (one point/profile)',
                            )
            except Exception as exc:
                print(f"注意: Argo 映射层绘制失败 (k={metadata.get('k')}, b={metadata.get('b')}): {exc}")

        ax.set_ylim(z_coords.max(), z_coords.min())
        if xmin is not None and xmax is not None: ax.set_xlim(xmin, xmax)
        if ymin is not None and ymax is not None: ax.set_ylim(ymax, ymin)

        # Argo DO 色标（与水平图一致：bwr + 150~240）
        if argo_scatter is not None and argo_has_do:
            cbar2 = fig.colorbar(argo_scatter, ax=ax, orientation='horizontal', fraction=0.046, pad=0.12)
            cbar2.set_label('DO/μmol·kg⁻¹', fontsize=14)
            cbar2.ax.tick_params(labelsize=12)
        
        # --- 构建并自定义图例 ---
        handles, labels = ax.get_legend_handles_labels()
        by_label = dict(zip(labels, handles)) # 创建一个去重的标签-句柄字典

        # 如果我们画了MLD线，就用我们创建的复合句柄替换掉自动生成的单个白线句柄
        if plot_mlt and mld_lines:
            by_label['Mixed Layer Depth'] = mld_lines
            
        ax.legend(by_label.values(), by_label.keys(), fontsize=18)

        # --- 4. 保存和显示 ---
        if save_fig:
            region_slug = _current_region_key()
            run_tag = argo_projection_config.file_stem()
            output_dir = argo_projection_config.output_dir(str(save_subdir), region_slug)
            output_dir.mkdir(parents=True, exist_ok=True)
            date_fn = date_str.replace('-', '')
            if ds_name == 'Argo':
                base_filename = (
                    f"Argo_{date_fn}_profile{metadata['eddy_no']}_vertical_{variable}_"
                    f"k{metadata['k']:.2f}b{metadata['b']:.2f}_{run_tag}.png"
                )
            else:
                base_filename = (f"{ds_name}{metadata['eddy_no']}_vertical_{variable}_{date_fn}_"
                                 f"k{metadata['k']:.2f}b{metadata['b']:.2f}_{run_tag}.png")
            save_path = output_dir / base_filename
            plt.savefig(save_path, dpi=300, bbox_inches='tight')
            print(f"Figure saved to: {save_path}")

        if show_fig:
            plt.show()

        plt.close(fig)

def plot_track_vertical_glorys(DS: list, no: int, needed_date: str | pd.Timestamp,
                               k: float | list[float] | None = None,
                               b: float | list[float] | None = None,
                               variable: str = 'vorticity',
                               show_fig: bool = True, save_fig: bool = False,
                               xmin: float = -400.0, xmax: float = 400.0,
                               ymin: float = 0.0, ymax: float = 1000.0,
                               color_vmin: float | None = None,
                               color_vmax: float | None = None,
                               plot_mlt: bool = False,
                               plot_argo_projection: bool = True,
                               argo_projection_config: DetectionConfig | None = None,
                                         argo_projection_min_depth: float | None = None,
                               plot_isolines: bool = True,
                               isoline_levels: int | list[float] | np.ndarray | None = None,
                               isoline_color: str = 'black',
                               isoline_linewidth: float = 0.8,
                               isoline_alpha: float = 0.45,
                               label_isolines: bool = False,
                               profile_spacing_km: float | None = None,
                               interpolate_z: bool = True,
                               profile_depth_spacing_m: float | None = None):
    """绘制指定涡旋在指定日期的 GLORYS 垂向剖面图（实际业务入口）。

    基于轨迹数据（META/legacy）的垂向可视化入口，内部复用共享绘图核心；与 Argo 入口相比会保留 eddy
    语义的参考信息（中心/半径/轮廓投影线）。

    参数:
        - DS (list | str | tuple | dict): 轨迹数据输入，常见 'acs'/'acl'/'cs'/'cl' 或 legacy 列表结构。
        - no (int): 轨迹编号（track id）。
        - needed_date (str | pd.Timestamp): 目标日期。
        - k (float | list[float] | None): 剖面线斜率，y = kx + b；可传多条线。
        - b (float | list[float] | None): 剖面线截距，与 k 配套。
        - variable (str): 主绘变量，默认 'vorticity'。
        - show_fig (bool): 是否显示图像，默认 True。
        - save_fig (bool): 是否保存图像，默认 False。
        - xmin (float): 横向显示与采样下界（km），默认 -400.0。
        - xmax (float): 横向显示与采样上界（km），默认 400.0。
        - ymin (float): 纵向深度显示上界（m），默认 0.0。
        - ymax (float): 纵向深度显示下界（m），默认 1000.0。
        - color_vmin (float | None): 主色斑图色标下限覆盖；None 时自动。
        - color_vmax (float | None): 主色斑图色标上限覆盖；None 时自动。
        - plot_mlt (bool): 是否叠加混合层深度线，默认 False。
        - plot_argo_projection (bool): 是否叠加同日匹配 Argo 点投影层，默认 True。
        - argo_projection_config (DetectionConfig | None): 投影点异常筛选配置；None 时使用默认。
        - argo_projection_min_depth (float | None): 投影点最小深度阈值（m）；None 时回退配置。
        - plot_isolines (bool): 是否叠加变量等值线，默认 True。
        - isoline_levels (int | list[float] | np.ndarray | None): 等值线级别数或显式级别；None 时自动。
        - isoline_color (str): 等值线颜色，默认 'black'。
        - isoline_linewidth (float): 等值线线宽，默认 0.8。
        - isoline_alpha (float): 等值线透明度，默认 0.45。
        - label_isolines (bool): 是否标注等值线数值，默认 False。
        - profile_spacing_km (float | None): 水平采样步长（km）；None 时用配置默认。
        - interpolate_z (bool): 是否将深度轴重采样到等间距网格，默认 True。
        - profile_depth_spacing_m (float | None): 深度重采样步长（m）；None 时用配置默认。
    """
    return _plot_vertical_glorys_core(
        DS,
        no,
        needed_date,
        k=k,
        b=b,
        variable=variable,
        show_fig=show_fig,
        save_fig=save_fig,
        xmin=xmin,
        xmax=xmax,
        ymin=ymin,
        ymax=ymax,
        color_vmin=color_vmin,
        color_vmax=color_vmax,
        plot_mlt=plot_mlt,
        plot_argo_projection=plot_argo_projection,
        argo_projection_config=argo_projection_config,
        argo_projection_min_depth=argo_projection_min_depth,
        plot_isolines=plot_isolines,
        isoline_levels=isoline_levels,
        isoline_color=isoline_color,
        isoline_linewidth=isoline_linewidth,
        isoline_alpha=isoline_alpha,
        label_isolines=label_isolines,
        profile_spacing_km=profile_spacing_km,
        interpolate_z=interpolate_z,
        profile_depth_spacing_m=profile_depth_spacing_m,
        save_subdir='plot_track_vertical_glorys',
        title_subject='Track',
    )

def plot_argo_vertical_glorys(
    profile_number: int,
    profile_time: int | str | pd.Timestamp,
    k: float | list[float] | None = None,
    b: float | list[float] | None = None,
    platform_number: int | None = None,
    variable: str = 'vorticity',
    show_fig: bool = True,
    save_fig: bool = False,
    xmin: float = -400.0,
    xmax: float = 400.0,
    ymin: float = 0.0,
    ymax: float = 1000.0,
    color_vmin: float | None = None,
    color_vmax: float | None = None,
    plot_mlt: bool = False,
    plot_argo_projection: bool = True,
    argo_projection_config: DetectionConfig | None = None,
    argo_projection_min_depth: float | None = None,
    plot_isolines: bool = True,
    isoline_levels: int | list[float] | np.ndarray | None = None,
    isoline_color: str = 'black',
    isoline_linewidth: float = 0.8,
    isoline_alpha: float = 0.45,
    label_isolines: bool = False,
    profile_spacing_km: float | None = None,
    interpolate_z: bool = True,
    profile_depth_spacing_m: float | None = None,
    argo_data_dir: str | Path | None = None,
):
    """以单个 Argo 剖面为中心绘制 GLORYS 垂向剖面图。

    基于 Argo 剖面中心和剖面线 y = kx + b 计算垂向切片；图中默认不绘制 eddy 语义参考竖线（中心/半径/轮廓）。

    参数:
        - profile_number (int): 目标 Argo 剖面编号。
        - profile_time (int | str | pd.Timestamp): 时间输入，支持年份（在该年内定位该剖面，对应多个日期时要求给具体日期）或具体日期/时间戳（直接定位当日剖面）。
        - k (float | list[float] | None): 剖面线斜率，满足 y = kx + b；可传标量或等长列表（多条剖面线分别出图）。
        - b (float | list[float] | None): 剖面线截距，与 k 配套。
        - platform_number (int | None): 可选平台号过滤。
        - variable (str): 主绘变量，常用 'vorticity'/'thetao'/'so'/'u'/'v'/'density'，默认 'vorticity'。
        - show_fig (bool): 是否显示图像，默认 True。
        - save_fig (bool): 是否保存图像，默认 False。
        - xmin (float): 横向显示与采样下界（km），默认 -400.0。
        - xmax (float): 横向显示与采样上界（km），默认 400.0。
        - ymin (float): 纵向深度显示上界（m），默认 0.0。
        - ymax (float): 纵向深度显示下界（m），默认 1000.0。
        - color_vmin (float | None): 主色斑图色标下限覆盖；None 时自动。
        - color_vmax (float | None): 主色斑图色标上限覆盖；None 时自动。
        - plot_mlt (bool): 是否叠加混合层深度线，默认 False。
        - plot_argo_projection (bool): 是否叠加同日 Argo 点投影层，默认 True。
        - argo_projection_config (DetectionConfig | None): 投影点异常筛选配置；None 时使用默认。
        - argo_projection_min_depth (float | None): 投影点最小深度阈值（m）；None 时回退配置。
        - plot_isolines (bool): 是否叠加变量等值线，默认 True。
        - isoline_levels (int | list[float] | np.ndarray | None): 等值线级别数或显式级别；None 时自动。
        - isoline_color (str): 等值线颜色，默认 'black'。
        - isoline_linewidth (float): 等值线线宽，默认 0.8。
        - isoline_alpha (float): 等值线透明度，默认 0.45。
        - label_isolines (bool): 是否标注等值线数值，默认 False。
        - profile_spacing_km (float | None): 水平采样步长（km）；None 时用配置默认。
        - interpolate_z (bool): 是否将深度轴重采样到等间距网格，默认 True。
        - profile_depth_spacing_m (float | None): 深度重采样步长（m）；None 时用配置默认。
        - argo_data_dir (str | Path | None): Argo 年度 parquet 目录；None 时使用配置默认目录。
    """
    info = _resolve_argo_profile_center(
        profile_number=int(profile_number),
        profile_time=profile_time,
        platform_number=platform_number,
        argo_data_dir=argo_data_dir,
    )

    center_lon = float(info['center_lon'])
    center_lat = float(info['center_lat'])
    target_date = pd.Timestamp(info['target_date'])
    df_year = info['year_df']

    if xmin is None or xmax is None:
        xmin = -400.0
        xmax = 400.0
    window_half_size_km = max(abs(float(xmin)), abs(float(xmax)))

    all_data_packages = get_vertical_glorys_from_center(
        center_lon=center_lon,
        center_lat=center_lat,
        needed_date=target_date,
        k=k,
        b=b,
        variables=[variable] + (['mlt'] if plot_mlt else []),
        x_min_km=xmin,
        x_max_km=xmax,
        profile_spacing_km=profile_spacing_km,
        interpolate_z=interpolate_z,
        profile_depth_spacing_m=profile_depth_spacing_m,
        window_half_size_km=float(window_half_size_km),
        profile_id=int(profile_number),
        ds_name='ARGO',
    )

    projected_argo_rows = pd.DataFrame()
    if plot_argo_projection:
        argo_projection_config = _resolve_detection_config(
            argo_projection_config,
            anomaly_min_depth=argo_projection_min_depth,
        )

        day_ts = pd.to_datetime(df_year[['Year', 'Month', 'Day']], errors='coerce').dt.normalize()
        day_rows = df_year.loc[day_ts == target_date.normalize()].copy()
        if not day_rows.empty:
            day_rows['Longitude'] = pd.to_numeric(day_rows['Longitude'], errors='coerce')
            day_rows['Latitude'] = pd.to_numeric(day_rows['Latitude'], errors='coerce')
            day_rows['Depth'] = pd.to_numeric(day_rows.get('Depth'), errors='coerce')
            day_rows['DO'] = pd.to_numeric(day_rows.get('DO', np.nan), errors='coerce')
            day_rows = day_rows.dropna(subset=['Longitude', 'Latitude', 'Depth'])

            lon_min_local, lon_max_local, lat_min, lat_max = _window_bounds_from_center_km(
                center_lon,
                center_lat,
                float(window_half_size_km),
            )
            day_lon_local = center_lon + _minimal_lon_diff_deg(day_rows['Longitude'].to_numpy(dtype=float), center_lon)
            mask_window = (
                (day_lon_local >= lon_min_local)
                & (day_lon_local <= lon_max_local)
                & (day_rows['Latitude'].to_numpy(dtype=float) >= lat_min)
                & (day_rows['Latitude'].to_numpy(dtype=float) <= lat_max)
            )
            day_window = day_rows.loc[mask_window].copy()
            if not day_window.empty:
                if 'Profile_number' in day_window.columns:
                    day_window['Profile_number'] = pd.to_numeric(day_window['Profile_number'], errors='coerce')
                    day_window = day_window.dropna(subset=['Profile_number'])
                deltas = _reduce_argo_profiles_by_anomaly(
                    day_window,
                    detection_config=argo_projection_config,
                )
                if not deltas.empty:
                    projected_argo_rows = deltas

    return _plot_vertical_glorys_core(
        None,
        int(profile_number),
        target_date,
        k=k,
        b=b,
        variable=variable,
        show_fig=show_fig,
        save_fig=save_fig,
        xmin=xmin,
        xmax=xmax,
        ymin=ymin,
        ymax=ymax,
        color_vmin=color_vmin,
        color_vmax=color_vmax,
        plot_mlt=plot_mlt,
        plot_argo_projection=plot_argo_projection,
        argo_projection_config=argo_projection_config,
        argo_projection_min_depth=argo_projection_min_depth,
        plot_isolines=plot_isolines,
        isoline_levels=isoline_levels,
        isoline_color=isoline_color,
        isoline_linewidth=isoline_linewidth,
        isoline_alpha=isoline_alpha,
        label_isolines=label_isolines,
        profile_spacing_km=profile_spacing_km,
        interpolate_z=interpolate_z,
        profile_depth_spacing_m=profile_depth_spacing_m,
        precomputed_data_packages=all_data_packages,
        precomputed_projected_argo_rows=projected_argo_rows,
        projection_distance_scale_km=float(window_half_size_km),
        x_axis_label='Distance from Profile Center (km)',
        save_subdir='plot_argo_vertical_glorys',
        title_subject='',
    )


def _normalize_profile_lines(
    k: float | list[float] | None = None,
    b: float | list[float] | None = None,
) -> tuple[list[float], list[float]]:
    """标准化剖面线参数，返回等长的 k/b 列表。k/b 均为 None 时返回空列表。"""
    if k is None and b is None:
        return [], []
    if k is None or b is None:
        raise ValueError("k 和 b 必须同时提供，或同时省略。")

    k_list = [float(k)] if isinstance(k, (int, float, np.integer, np.floating)) else [float(v) for v in k]
    b_list = [float(b)] if isinstance(b, (int, float, np.integer, np.floating)) else [float(v) for v in b]
    if len(k_list) != len(b_list):
        raise ValueError("k 和 b 的列表长度必须一致。")
    if len(k_list) == 0:
        raise ValueError("至少需要一条剖面线参数。")
    return k_list, b_list


def _normalize_overview_vertical_variables(variables: list[str] | None) -> list[str]:
    """规范化 overview 的 2x2 vertical 变量顺序。"""
    base_order = ['vorticity', 'sigma', 'thetao', 'salinity']
    if variables is None:
        return base_order

    alias_map = {
        'vorticity': 'vorticity',
        'thetao': 'thetao',
        'temp': 'thetao',
        'temperature': 'thetao',
        'density': 'sigma',
        'sigma': 'sigma',
        'sigma0': 'sigma',
        'salinity': 'salinity',
        'so': 'salinity',
    }
    mapped = {
        alias_map.get(str(var).strip().lower())
        for var in variables
    }
    mapped.discard(None)

    ordered = [k for k in base_order if k in mapped]
    for k in base_order:
        if k not in ordered:
            ordered.append(k)
    return ordered[:4]


def _overview_var_style(var_key: str) -> tuple[str, str, str, tuple[float, float] | None]:
    """返回变量的标题、色标标签、配色和固定色标范围。

    所有变量均使用固定色标范围，以确保跨图颜色可比：
    - pv:        (-5e-9, 5e-9) QG PV (m⁻¹s⁻¹)
    - vorticity: (-0.7, 0.7) ζ/f
    - thetao:    (1, 27) °C
    - sigma:     (23, 28) kg/m³
    - salinity:  (33, 36) psu
    """
    if var_key == 'pv':
        return ('QG PV', 'QG PV (m⁻¹s⁻¹)', 'RdBu_r', (-5e-9, 5e-9))
    if var_key == 'vorticity':
        return ('Vorticity', r'$\zeta/f$', 'seismic', (-0.7, 0.7))
    if var_key == 'thetao':
        return ('Temperature', 'Temperature (°C)', 'rainbow', (1.0, 27.0))
    if var_key == 'sigma':
        return ('Density', 'Potential Density Anomaly (kg/m³)', 'RdBu_r', (23.0, 28.0))
    if var_key == 'salinity':
        return ('Salinity', 'Salinity (psu)', 'viridis', (33.0, 36.0))
    if var_key == 'z_of_sigma':
        return ('Isopycnal Depth', 'Depth (m)', 'terrain', None)
    if var_key == 'argo_weight':
        return ('Argo Coverage', 'Effective Weight', 'YlOrRd', None)
    return (var_key, var_key, 'viridis', None)


def _auto_clim(data: np.ndarray | np.ma.MaskedArray, fallback: tuple[float, float] = (0.0, 1.0)) -> tuple[float, float]:
    """按有效数据自动计算色标范围。"""
    arr = np.ma.array(data, copy=False)
    vals = arr.compressed()
    if vals.size == 0:
        return fallback
    vmin = float(np.nanmin(vals))
    vmax = float(np.nanmax(vals))
    if not np.isfinite(vmin) or not np.isfinite(vmax):
        return fallback
    if np.isclose(vmin, vmax):
        span = abs(vmin) * 0.05 + 1e-6
        return (vmin - span, vmax + span)
    return (vmin, vmax)


def _project_argo_rows_to_profile_for_overview(
    data_package: dict,
    projected_argo_rows: pd.DataFrame,
    *,
    distance_scale_km: float | None = None,
) -> pd.DataFrame:
    """将 Argo 代表点投影到指定剖面，返回用于绘制的坐标与点大小。"""
    if projected_argo_rows is None or projected_argo_rows.empty:
        return pd.DataFrame()

    line_lons_raw = np.asarray(data_package.get('lon_coords', []), dtype=float)
    line_lats = np.asarray(data_package.get('lat_coords', []), dtype=float)
    line_y = np.asarray(data_package.get('y_coords', []), dtype=float)
    metadata = data_package.get('metadata', {})

    if line_lons_raw.size < 2 or line_lats.size != line_lons_raw.size or line_y.size < line_lons_raw.size:
        return pd.DataFrame()

    valid_line_lons = np.isfinite(line_lons_raw)
    if not np.any(valid_line_lons):
        return pd.DataFrame()
    lon_anchor = float(np.nanmean(line_lons_raw[valid_line_lons]))
    # 统一经度域，避免 0~360 与 -180~180 混用导致投影偏离。
    line_lons = lon_anchor + _minimal_lon_diff_deg(line_lons_raw, lon_anchor)

    line_y = line_y[:line_lons.size]
    try:
        k_line = float(metadata.get('k'))
        b_line = float(metadata.get('b'))
    except Exception:
        return pd.DataFrame()

    rows = projected_argo_rows.copy()
    rows['Longitude'] = pd.to_numeric(rows['Longitude'], errors='coerce')
    rows['Latitude'] = pd.to_numeric(rows['Latitude'], errors='coerce')
    rows['Depth'] = pd.to_numeric(rows['Depth'], errors='coerce')
    rows['DO'] = pd.to_numeric(rows.get('DO', np.nan), errors='coerce')
    rows = rows.dropna(subset=['Longitude', 'Latitude', 'Depth'])
    if rows.empty:
        return pd.DataFrame()

    pts_lon_raw = rows['Longitude'].to_numpy(dtype=float)
    pts_lon = lon_anchor + _minimal_lon_diff_deg(pts_lon_raw, lon_anchor)
    pts_lat = rows['Latitude'].to_numpy(dtype=float)
    pts_depth = rows['Depth'].to_numpy(dtype=float)
    pts_do = rows['DO'].to_numpy(dtype=float)

    denom = 1.0 + k_line ** 2
    proj_lon = (pts_lon + k_line * (pts_lat - b_line)) / denom
    proj_lat = k_line * proj_lon + b_line

    dist2 = (line_lons[None, :] - proj_lon[:, None]) ** 2 + (line_lats[None, :] - proj_lat[:, None]) ** 2
    nearest_idx = np.argmin(dist2, axis=1)
    proj_y = line_y[nearest_idx]

    scale_ref = approximate_degree_length(np.nanmean(line_lats))
    dlon_deg = _minimal_lon_diff_deg(pts_lon, proj_lon)
    dx_m = dlon_deg * scale_ref['meters_per_degree_lon']
    dy_m = (pts_lat - proj_lat) * scale_ref['meters_per_degree_lat']
    dist_to_line_km = np.hypot(dx_m, dy_m) / 1000.0

    scale_km = None
    if distance_scale_km is not None and np.isfinite(float(distance_scale_km)) and float(distance_scale_km) > 0:
        scale_km = float(distance_scale_km)
    if scale_km is None:
        finite_span = np.asarray(line_y[np.isfinite(line_y)], dtype=float)
        scale_km = max(20.0, float(np.nanmax(np.abs(finite_span)))) if finite_span.size else 50.0

    marker_max = 126.0
    marker_min = 7.0
    weight = 1.0 - np.clip(dist_to_line_km / scale_km, 0.0, 1.0)
    marker_sizes = marker_min + (marker_max - marker_min) * (weight ** 1.2)

    out = pd.DataFrame({
        'proj_y': proj_y,
        'Depth': pts_depth,
        'DO': pts_do,
        'marker_size': marker_sizes,
    })
    out = out[np.isfinite(out['proj_y']) & np.isfinite(out['Depth'])].copy()
    return out


def _project_argo_rows_to_sigma_for_overview(
    z_package: dict,
    projected_argo_rows: pd.DataFrame,
) -> pd.DataFrame:
    """将 z 坐标 Argo 投影点转换到 σ 坐标，用于在 σ 坐标 overview 上叠加。

    在每个 Argo 点的投影位置，从 GLORYS sigma 场中插值出对应深度的 σ₀，
    返回 ``(proj_y, sigma, DO, marker_size)`` 格式的 DataFrame。
    """
    if projected_argo_rows is None or projected_argo_rows.empty:
        return pd.DataFrame()

    profile_data = z_package.get('profile_data', {})
    sigma_2d = profile_data.get('sigma')
    z_coords = np.asarray(z_package.get('z_coords', []), dtype=float)
    y_coords = np.asarray(z_package.get('y_coords', []), dtype=float)

    if sigma_2d is None or z_coords.size < 2 or y_coords.size < 2:
        return pd.DataFrame()

    sigma_ma = np.ma.array(sigma_2d, copy=False)

    rows = projected_argo_rows.copy()
    proj_y = rows['proj_y'].to_numpy(dtype=float)
    depths = rows['Depth'].to_numpy(dtype=float)
    do_vals = rows['DO'].to_numpy(dtype=float)
    marker_sizes = rows['marker_size'].to_numpy(dtype=float)

    sigma_at_argo = np.full(len(rows), np.nan)

    for i, (y_i, depth_i) in enumerate(zip(proj_y, depths)):
        # 找到最近的 y 索引
        j = int(np.argmin(np.abs(y_coords - y_i)))
        sigma_col = np.asarray(np.ma.filled(sigma_ma[:, j], np.nan), dtype=float)
        valid = np.isfinite(sigma_col)
        if valid.sum() < 2:
            continue
        sigma_at_argo[i] = np.interp(depth_i, z_coords[valid], sigma_col[valid],
                                     left=np.nan, right=np.nan)

    out = pd.DataFrame({
        'proj_y': proj_y,
        'sigma': sigma_at_argo,
        'DO': do_vals,
        'marker_size': marker_sizes,
    })
    return out[np.isfinite(out['proj_y']) & np.isfinite(out['sigma'])].copy()


def _prepare_overview_projection_rows(
    rows: pd.DataFrame,
    *,
    detection_config: DetectionConfig,
) -> pd.DataFrame:
    """清洗 Argo 行并按统一异常配置筛选代表点。"""
    if rows is None or rows.empty:
        return pd.DataFrame()

    cleaned = rows.copy()
    cleaned['Longitude'] = pd.to_numeric(cleaned.get('Longitude'), errors='coerce')
    cleaned['Latitude'] = pd.to_numeric(cleaned.get('Latitude'), errors='coerce')
    cleaned['Depth'] = pd.to_numeric(cleaned.get('Depth'), errors='coerce')
    cleaned['DO'] = pd.to_numeric(cleaned.get('DO', np.nan), errors='coerce')
    cleaned = cleaned.dropna(subset=['Longitude', 'Latitude', 'Depth'])
    if cleaned.empty:
        return pd.DataFrame()

    if 'Profile_number' in cleaned.columns:
        cleaned['Profile_number'] = pd.to_numeric(cleaned['Profile_number'], errors='coerce')
        cleaned = cleaned.dropna(subset=['Profile_number'])
        if cleaned.empty:
            return pd.DataFrame()

    deltas = _reduce_argo_profiles_by_anomaly(
        cleaned,
        detection_config=detection_config,
    )
    return deltas if not deltas.empty else pd.DataFrame()


def _as_profile_2d_array(values, expected_shape: tuple[int, int] | None = None) -> np.ndarray | None:
    """将剖面变量转为 float 2D ndarray，masked 值填为 NaN。"""
    if values is None:
        return None
    arr = np.ma.filled(np.ma.array(values, copy=False), np.nan)
    arr = np.asarray(arr, dtype=float)
    if arr.ndim != 2:
        return None
    if expected_shape is not None and arr.shape != expected_shape:
        return None
    return arr


def _nan_stat(values: np.ndarray, op: str) -> float:
    """NaN-safe scalar statistic helper."""
    vals = np.asarray(values, dtype=float)
    vals = vals[np.isfinite(vals)]
    if vals.size == 0:
        return np.nan
    if op == 'mean':
        return float(np.nanmean(vals))
    if op == 'median':
        return float(np.nanmedian(vals))
    if op == 'max':
        return float(np.nanmax(vals))
    if op == 'p95':
        return float(np.nanpercentile(vals, 95))
    if op == 'min':
        return float(np.nanmin(vals))
    raise ValueError(f"Unknown statistic op: {op}")


def _format_scientific_mathtext(value: float, precision: int = 2) -> str:
    """Format a scalar as Matplotlib mathtext scientific notation."""
    try:
        value_f = float(value)
    except Exception:
        return r"\mathrm{NaN}"
    if not np.isfinite(value_f):
        return r"\mathrm{NaN}"

    mantissa, exponent_text = f"{value_f:.{precision}e}".split("e")
    exponent = int(exponent_text)
    return rf"{mantissa}\times10^{{{exponent}}}"


def calculate_glorys_vertical_profile_diagnostics(
    data_package: dict,
    *,
    projection_x_km: float = 0.0,
    projection_depth_m: float | None = None,
    x_window_km: float = 25.0,
    z_window_m: float | None = 100.0,
    depth_range_m: tuple[float, float] | None = (0.0, 1000.0),
    heave_search_range: float = _heave_search_range,
    heave_depth_threshold: float = _heave_depth_threshold,
    heave_threshold: float = _heave_magnitude_threshold,
    heave_x_window_km: float = _heave_x_window_km,
    heave_local_x_window_km: float = _heave_local_x_window_km,
    heave_z_search_m: float | None = _heave_z_search_m,
) -> dict:
    """计算等密面出露指数 OI（Outcrop Index）与垂向位移 Heave。

    OI 是一个二值诊断量，判定 Argo 异常深度处的等密面（σ_argo）在水平窗口内是否「出露」（outcrop）到
    近表层，由 Heave（等密面垂向位移）与 Ventilation（通风深度）双准则联合判定，用于区分深层高 DO 能否
    经局地通风追溯至表层（Type 1）还是只能停留深层（Type 2）。

    参数:
        - data_package (dict): 含 GLORYS 垂向切片的数据包（get_vertical_glorys 输出）。
        - projection_x_km (float): Argo 在剖面线上的投影 x 坐标（km），默认 0.0。
        - projection_depth_m (float | None): Argo 异常深度（m）；None 时用深度范围中点。
        - x_window_km (float): 局地水平窗口半宽（km，用于 valid_fraction），默认 25.0。
        - z_window_m (float | None): 局地垂向窗口半宽（m，用于 valid_fraction），默认 100.0。
        - depth_range_m (tuple[float, float] | None): 限制搜索的深度范围（m），默认 (0.0, 1000.0)。
        - heave_search_range (float): 从 σ_argo 向上搜索的 σ 跨度（kg/m³），默认来自 processing.yml。
        - heave_depth_threshold (float): 通风判定深度（m），z_min 浅于此值视为「接近海表」，默认来自 processing.yml。
        - heave_threshold (float): Heave 判定阈值（m），等密面起伏 ≥ 此值视为显著位移，默认来自 processing.yml。
        - heave_x_window_km (float): 搜索通风/出露的水平窗口半宽（km），默认来自 processing.yml。
        - heave_local_x_window_km (float): 计算 Heave 时局部最深的搜索半宽（km），默认来自 processing.yml。
        - heave_z_search_m (float | None): 等密线连通性垂向范围（m），仅搜索 Argo 点 ± 该距离内存在过的 σ 面，默认来自 processing.yml。

    返回:
        - dict: 诊断指标字典，含以下键：

            - glorys_heave_m (float)：Heave 幅度（m），等密线在局地凹底到窗口最浅点的垂直距离。
            - glorys_heave_zmin (float)：heave 峰值 σ 面在窗口内的最浅深度（m），与 σ_peak 同源。
            - glorys_heave_sigma_argo (float)：Argo 异常点的 σ（kg/m³）。
            - glorys_heave_sigma_peak (float)：Heave 峰值所在 σ 面（kg/m³）。
            - heave_valid_fraction (float)：局地窗口内 σ 有效数据占比。

    说明:
        Heave（等密面垂向位移）:

            - 在 Argo 附近 ±local_x_window_km 内找到 σ ≈ σ_argo 等密线的最深点（凹底），与全窗口（±x_window_km）内该等密线最浅点之间的垂直距离；Heave ≥ heave_threshold 说明等密面发生显著中尺度起伏。

        Ventilation（通风深度）:

            - Heave 峰值 σ 面（σ_peak）在 ±x_window_km 内的最浅深度 z_min；z_min < heave_depth_threshold 说明该等密面在窗口内接近海表，存在通风/潜沉的物理通道。

        联合判定:

            - OI=True（Type 1）：Heave ≥ threshold 且 z_min < depth_threshold，深层高 DO 可经等密面通风追溯至表层。
            - OI=False（Type 2）：任一条件不满足，等密线全程停留深层，高 DO 来源不能用局地通风解释。
    """
    try:
        x0 = float(projection_x_km)
    except Exception:
        x0 = np.nan
    try:
        depth0 = float(projection_depth_m) if projection_depth_m is not None else np.nan
    except Exception:
        depth0 = np.nan

    heave_x_half = abs(float(heave_x_window_km))
    heave_sr = abs(float(heave_search_range))
    heave_dt = abs(float(heave_depth_threshold))
    heave_ht = abs(float(heave_threshold))

    out: dict[str, object] = {
        'projection_x_km': x0 if np.isfinite(x0) else np.nan,
        'projection_depth_m': depth0 if np.isfinite(depth0) else np.nan,
        'heave_x_window_km': float(heave_x_half),
        'heave_search_range': float(heave_sr),
        'heave_depth_threshold': float(heave_dt),
        'heave_threshold': float(heave_ht),
        'heave_valid_fraction': np.nan,
        'glorys_heave_zmin': np.nan,
        'glorys_heave_m': np.nan,
        'glorys_heave_sigma_argo': np.nan,
        'glorys_heave_sigma_peak': np.nan,
        'heave_error': None,
    }

    if not data_package:
        out['heave_error'] = 'empty_data_package'
        return out

    y_coords = np.asarray(data_package.get('y_coords', []), dtype=float)
    z_coords = np.asarray(data_package.get('z_coords', []), dtype=float)
    if y_coords.size < 2 or z_coords.size < 2:
        out['heave_error'] = 'insufficient_coordinates'
        return out

    finite_y = np.isfinite(y_coords)
    finite_z = np.isfinite(z_coords)
    if np.count_nonzero(finite_y) < 2 or np.count_nonzero(finite_z) < 2:
        out['heave_error'] = 'nonfinite_coordinates'
        return out

    y_idx = np.nonzero(finite_y)[0]
    z_idx = np.nonzero(finite_z)[0]
    y_use = y_coords[y_idx]
    z_use = z_coords[z_idx]
    y_order = np.argsort(y_use)
    z_order = np.argsort(z_use)
    y_idx = y_idx[y_order]
    z_idx = z_idx[z_order]
    y_use = y_use[y_order]
    z_use = z_use[z_order]

    expected_shape = (len(z_coords), len(y_coords))
    sigma_full = _as_profile_2d_array(data_package.get('profile_data', {}).get('sigma'), expected_shape=expected_shape)
    if sigma_full is None:
        out['heave_error'] = 'sigma_unavailable'
        return out

    sigma = sigma_full[np.ix_(z_idx, y_idx)]
    if sigma.shape != (z_use.size, y_use.size):
        out['heave_error'] = 'sigma_shape_mismatch'
        return out

    if not np.isfinite(x0):
        x0 = 0.0

    # 确定 Argo 投影点的 σ 值
    nearest_y_idx = int(np.nanargmin(np.abs(y_use - x0)))
    if np.isfinite(depth0):
        nearest_z_idx = int(np.nanargmin(np.abs(z_use - depth0)))
    else:
        if depth_range_m is not None and len(depth_range_m) == 2:
            z_min_d = float(min(depth_range_m))
            z_max_d = float(max(depth_range_m))
            z_in_range = np.nonzero((z_use >= z_min_d) & (z_use <= z_max_d))[0]
        else:
            z_in_range = np.arange(len(z_use))
        if z_in_range.size == 0:
            z_in_range = np.arange(len(z_use))
        nearest_z_idx = int(z_in_range[len(z_in_range) // 2])

    sigma_argo = float(sigma[nearest_z_idx, nearest_y_idx])
    out['glorys_heave_sigma_argo'] = sigma_argo if np.isfinite(sigma_argo) else np.nan

    if not np.isfinite(sigma_argo):
        out['heave_error'] = 'sigma_argo_nan'
        return out

    # 局地窗口 valid_fraction
    try:
        x_half = abs(float(x_window_km))
    except Exception:
        x_half = 25.0
    x_mask = np.abs(y_use - x0) <= x_half
    if not np.any(x_mask):
        x_mask[np.nanargmin(np.abs(y_use - x0))] = True

    if np.isfinite(depth0) and z_window_m is not None:
        z_half = abs(float(z_window_m))
        z_mask = np.abs(z_use - depth0) <= z_half
        if not np.any(z_mask):
            z_mask[np.nanargmin(np.abs(z_use - depth0))] = True
    elif depth_range_m is not None and len(depth_range_m) == 2:
        z_min = float(min(depth_range_m))
        z_max = float(max(depth_range_m))
        z_mask = (z_use >= z_min) & (z_use <= z_max)
        if not np.any(z_mask):
            z_mask = np.ones_like(z_use, dtype=bool)
    else:
        z_mask = np.ones_like(z_use, dtype=bool)

    local_mask = z_mask[:, None] & x_mask[None, :]
    local_sigma = sigma[local_mask]
    out['heave_valid_fraction'] = (
        float(np.count_nonzero(np.isfinite(local_sigma)) / local_sigma.size)
        if local_sigma.size > 0 else np.nan
    )

    # --- Heave + OI 联合算法 ---
    # 水平搜索窗口
    heave_x_mask = np.abs(y_use - x0) <= heave_x_half
    if not np.any(heave_x_mask):
        heave_x_mask[np.nanargmin(np.abs(y_use - x0))] = True

    sigma_lower = sigma_argo - heave_sr

    # 等密线连通性约束：取 Argo 投影列 ±z_connect_m 内的 σ 范围，与原 σ 窗取交集
    sigma_argo_col = sigma[:, nearest_y_idx]
    if heave_z_search_m is not None and np.isfinite(depth0):
        z_connect_mask = np.abs(z_use - depth0) <= float(heave_z_search_m)
        sigma_connect = sigma_argo_col[z_connect_mask]
        sigma_connect = sigma_connect[np.isfinite(sigma_connect)]
        if sigma_connect.size < 2:
            out['heave_error'] = 'no_valid_sigma_near_argo'
            return out
        sigma_lower = max(sigma_lower, float(np.nanmin(sigma_connect)))

    sigma_search_range = max(0.05, sigma_argo - sigma_lower)
    n_search = max(8, int(sigma_search_range / 0.01))
    sigma_levels = np.linspace(sigma_lower, sigma_argo, n_search)

    # Heave: 等密线在 Argo 附近的最大深度 − 窗口内最浅深度
    # 用 local_z_max (±50km) 而非 x=0 点值，避免 Argo 恰好落在凸顶上时 heave 被低估
    local_x_half = abs(float(heave_local_x_window_km))
    local_x_mask = np.abs(y_use - x0) <= local_x_half
    if not np.any(local_x_mask):
        local_x_mask[np.nanargmin(np.abs(y_use - x0))] = True

    max_heave = 0.0
    max_heave_sigma = np.nan
    max_heave_zmin = np.nan

    tol = max(0.02, 1.5 * (sigma_levels[1] - sigma_levels[0]))

    for sigma_i in sigma_levels:
        sigma_diff = np.abs(sigma - sigma_i)
        close_mask = (sigma_diff < tol) & heave_x_mask[None, :]
        close_local = (sigma_diff < tol) & local_x_mask[None, :]

        # z_min：当前 σ 面在 ±200km 窗口内的最浅深度
        if not np.any(close_mask):
            continue
        matched_z = z_use[np.nonzero(close_mask)[0]]
        if matched_z.size < 2:
            continue
        min_z_i = float(np.nanmin(matched_z))

        # local_z_max：当前 σ 面在 Argo 附近 ±50km 的最深点（凹底）
        if not np.any(close_local):
            continue
        local_z = z_use[np.nonzero(close_local)[0]]
        if local_z.size == 0:
            continue
        local_z_max_i = float(np.nanmax(local_z))

        if np.isfinite(local_z_max_i) and np.isfinite(min_z_i):
            heave_i = float(local_z_max_i - min_z_i)
            if heave_i > max_heave:
                max_heave = heave_i
                max_heave_sigma = sigma_i
                max_heave_zmin = min_z_i

    out['glorys_heave_m'] = float(max_heave) if max_heave > 0 else np.nan
    out['glorys_heave_sigma_peak'] = float(max_heave_sigma) if np.isfinite(max_heave_sigma) else np.nan
    out['glorys_heave_zmin'] = float(max_heave_zmin) if np.isfinite(max_heave_zmin) else np.nan

    return out


def _remap_vertical_package_to_sigma(
    vertical_package: dict,
    sigma_min: float = 23.0,
    sigma_max: float = 28.0,
    sigma_step: float = 0.05,
) -> dict:
    """将 z 坐标 vertical_package 的 3D 变量重映射到 σ 坐标。

    对每个水平列，从混合层底以下的稳定层化区提取 σ 单调段，
    再插值到目标 σ 级别上。返回与 z 坐标包结构相同的字典，
    其中 ``z_coords`` 变为 σ 级别，``profile_data`` 新增 ``z_of_sigma``。
    """
    profile_data = vertical_package.get('profile_data', {})
    z_coords = np.asarray(vertical_package.get('z_coords', []), dtype=float)
    y_coords = np.asarray(vertical_package.get('y_coords', []), dtype=float)

    sigma_2d = profile_data.get('sigma')
    if sigma_2d is None:
        raise ValueError("vertical_package 中缺少 'sigma' 变量，无法重映射到 σ 坐标")

    sigma_ma = np.ma.array(sigma_2d, copy=False)
    nz, ny = sigma_ma.shape
    if nz != len(z_coords) or ny != len(y_coords):
        raise ValueError("sigma 维度与 z_coords / y_coords 不匹配")

    sigma_targets = np.arange(sigma_min, sigma_max + sigma_step * 0.5, sigma_step)

    # 收集需要重映射的 3D 变量（排除 sigma 和 2D 变量如 mlt）
    vars_3d = {}
    for key, val in profile_data.items():
        if key == 'sigma':
            continue
        val_arr = np.ma.array(val, copy=False)
        if val_arr.ndim == 2 and val_arr.shape[0] == nz and val_arr.shape[1] == ny:
            vars_3d[key] = val_arr

    out_profile_data: dict[str, np.ma.MaskedArray] = {}
    for key in vars_3d:
        out_profile_data[key] = np.ma.masked_all((len(sigma_targets), ny), dtype=float)
    z_of_sigma = np.ma.masked_all((len(sigma_targets), ny), dtype=float)

    for j in range(ny):
        sigma_col = sigma_ma[:, j]
        valid = ~sigma_col.mask if np.ma.is_masked(sigma_col) else np.ones(nz, dtype=bool)
        if valid.sum() < 2:
            continue

        z_valid = z_coords[valid]
        sigma_valid = np.asarray(sigma_col[valid], dtype=float)

        # 找到稳定层化起点: dσ/dz > threshold
        dsigma_dz = np.gradient(sigma_valid, z_valid)
        strat_start = 0
        for idx in range(len(dsigma_dz)):
            if dsigma_dz[idx] > 0.001:
                strat_start = idx
                break

        z_strat = z_valid[strat_start:]
        sigma_strat = sigma_valid[strat_start:]
        if len(sigma_strat) < 2:
            continue

        # 确保严格单调递增: 去除 dσ <= 0 的扰动点
        mono_mask = np.ones(len(sigma_strat), dtype=bool)
        for idx in range(1, len(sigma_strat)):
            if sigma_strat[idx] <= sigma_strat[idx - 1]:
                mono_mask[idx] = False
        z_mono = z_strat[mono_mask]
        sigma_mono = sigma_strat[mono_mask]
        if len(sigma_mono) < 2:
            continue

        # 裁剪到目标 σ 范围内
        z_mono_interp = np.interp(sigma_targets, sigma_mono, z_mono,
                                  left=np.nan, right=np.nan)
        z_of_sigma[:, j] = np.ma.masked_invalid(z_mono_interp)

        for key, var_arr in vars_3d.items():
            var_col = var_arr[:, j]
            var_valid = np.asarray(var_col[valid][strat_start:][mono_mask], dtype=float)
            var_interp = np.interp(sigma_targets, sigma_mono, var_valid,
                                   left=np.nan, right=np.nan)
            out_profile_data[key][:, j] = np.ma.masked_invalid(var_interp)

    # 将 z_of_sigma 加入导出变量
    out_profile_data['z_of_sigma'] = z_of_sigma

    return {
        'profile_data': out_profile_data,
        'y_coords': y_coords.copy(),
        'z_coords': sigma_targets.copy(),
        'lon_coords': np.asarray(vertical_package.get('lon_coords', []), dtype=float).copy(),
        'lat_coords': np.asarray(vertical_package.get('lat_coords', []), dtype=float).copy(),
        'projections': copy.deepcopy(vertical_package.get('projections', {})),
        'metadata': copy.deepcopy(vertical_package.get('metadata', {})),
        'is_sigma_coords': True,
    }


def _plot_glorys_overview_vertical_2x2(
    *,
    vertical_package: dict,
    variables: list[str],
    k_val: float,
    b_val: float,
    subject_label: str,
    date_label: str,
    xmin: float | None,
    xmax: float | None,
    ymin: float | None,
    ymax: float | None,
    projected_argo_profile: pd.DataFrame | None = None,
    plot_mlt: bool = False,
    plot_isolines: bool = True,
    isoline_levels: int | list[float] | np.ndarray | None = None,
    isoline_color: str = 'black',
    isoline_linewidth: float = 0.8,
    isoline_alpha: float = 0.45,
    label_isolines: bool = True,
    title_extra: str = '',
    source_label: str = 'GLORYS',
):
    """绘制 overview 的 2x2 vertical 图。"""
    fig, axes = plt.subplots(2, 2, figsize=(15, 10), constrained_layout=True)
    axes_flat = axes.ravel()

    y_coords = np.asarray(vertical_package.get('y_coords', []), dtype=float)
    z_coords = np.asarray(vertical_package.get('z_coords', []), dtype=float)
    projections = vertical_package.get('projections', {})
    metadata = vertical_package.get('metadata', {})
    draw_reference_lines = bool(metadata.get('draw_reference_lines', True))
    mlt_profile = None
    if plot_mlt:
        mlt_raw = vertical_package.get('profile_data', {}).get('mlt')
        if mlt_raw is not None:
            mlt_arr = np.ma.array(mlt_raw, copy=False)
            if mlt_arr.ndim == 1:
                mlt_profile = np.asarray(np.ma.filled(mlt_arr, np.nan), dtype=float)
            elif mlt_arr.ndim == 2 and 1 in mlt_arr.shape:
                mlt_profile = np.asarray(np.ma.filled(mlt_arr.reshape(-1), np.nan), dtype=float)

    for i, var_key in enumerate(variables[:4]):
        ax = axes_flat[i]
        title_txt, cbar_label, cmap_name, fixed_clim = _overview_var_style(var_key)

        v_field = np.ma.array(vertical_package.get('profile_data', {}).get(var_key), copy=False)
        if v_field.ndim != 2 or y_coords.size < 2 or z_coords.size < 2:
            ax.text(0.5, 0.5, f"No valid {title_txt}", ha='center', va='center', transform=ax.transAxes)
            ax.set_title(f"Vertical {title_txt}")
            continue

        nz = min(v_field.shape[0], len(z_coords))
        ny = min(v_field.shape[1], len(y_coords))
        v_field = v_field[:nz, :ny]
        z_plot = z_coords[:nz]
        y_plot = y_coords[:ny]

        Y_mesh, Z_mesh = np.meshgrid(y_plot, z_plot)
        v_clim = fixed_clim if fixed_clim is not None else _auto_clim(v_field)
        v_field_filled = np.ma.filled(np.ma.array(v_field, copy=False), np.nan)
        pc = ax.pcolormesh(Y_mesh, Z_mesh, v_field_filled, cmap=cmap_name, shading='auto', vmin=v_clim[0], vmax=v_clim[1])

        if plot_isolines:
            _draw_isolines(ax, Y_mesh, Z_mesh, v_field_filled, v_clim,
                           isoline_levels=isoline_levels,
                           isoline_color=isoline_color,
                           isoline_linewidth=isoline_linewidth,
                           isoline_alpha=isoline_alpha,
                           label_isolines=label_isolines)

        if draw_reference_lines:
            ax.axvline(0.0, color='black', linestyle='--', linewidth=1.6, label='Center Projection' if i == 0 else None)
            for j, dist in enumerate(projections.get('radius', [])):
                ax.axvline(dist, color='r', linestyle='--', linewidth=1.2, label='Radius Projection' if (i == 0 and j == 0) else None)
            for j, dist in enumerate(projections.get('contour', [])):
                ax.axvline(dist, color='tab:blue', linestyle=':', linewidth=1.2, label='Contour Projection' if (i == 0 and j == 0) else None)

        mlt_drawn = False
        if plot_mlt and mlt_profile is not None and mlt_profile.size >= ny:
            mlt_line = mlt_profile[:ny]
            valid_mlt = np.isfinite(mlt_line)
            if np.any(valid_mlt):
                ax.plot(y_plot[valid_mlt], mlt_line[valid_mlt], color='black', linewidth=2.0, zorder=6)
                ax.plot(
                    y_plot[valid_mlt],
                    mlt_line[valid_mlt],
                    color='white',
                    linewidth=1.2,
                    zorder=7,
                    label='Mixed Layer Depth' if i == 0 else None,
                )
                mlt_drawn = True

        if projected_argo_profile is not None and not projected_argo_profile.empty:
            valid_pts = np.isfinite(projected_argo_profile['proj_y'].to_numpy(dtype=float)) & np.isfinite(projected_argo_profile['Depth'].to_numpy(dtype=float))
            if np.any(valid_pts):
                do_vals = projected_argo_profile['DO'].to_numpy(dtype=float)
                if np.isfinite(do_vals[valid_pts]).any():
                    ax.scatter(
                        projected_argo_profile['proj_y'].to_numpy(dtype=float)[valid_pts],
                        projected_argo_profile['Depth'].to_numpy(dtype=float)[valid_pts],
                        c=do_vals[valid_pts],
                        cmap='bwr',
                        vmin=150,
                        vmax=240,
                        s=projected_argo_profile['marker_size'].to_numpy(dtype=float)[valid_pts],
                        edgecolors='black',
                        linewidths=0.35,
                        alpha=0.9,
                        zorder=6,
                        label='Projected Argo' if i == 0 else None,
                    )
                else:
                    ax.scatter(
                        projected_argo_profile['proj_y'].to_numpy(dtype=float)[valid_pts],
                        projected_argo_profile['Depth'].to_numpy(dtype=float)[valid_pts],
                        color='blue',
                        s=projected_argo_profile['marker_size'].to_numpy(dtype=float)[valid_pts],
                        edgecolors='black',
                        linewidths=0.35,
                        alpha=0.9,
                        zorder=6,
                        label='Projected Argo' if i == 0 else None,
                    )

        ax.set_ylim(z_plot.max(), z_plot.min())
        if xmin is not None and xmax is not None:
            ax.set_xlim(float(xmin), float(xmax))
        if ymin is not None and ymax is not None:
            ax.set_ylim(float(ymax), float(ymin))

        ax.set_title(f"Vertical {title_txt}", fontsize=12)
        ax.set_xlabel('Distance (km)')
        ax.set_ylabel('Depth (m)')
        cbar = fig.colorbar(pc, ax=ax, orientation='vertical', fraction=0.045, pad=0.02)
        cbar.set_label(cbar_label, fontsize=10)
        cbar.ax.tick_params(labelsize=9)
        if i == 0 and (draw_reference_lines or mlt_drawn or (projected_argo_profile is not None and not projected_argo_profile.empty)):
            ax.legend(fontsize=9, loc='best')

    extra_text = f", {title_extra}" if title_extra else ""
    fig.suptitle(
        f"{subject_label} {source_label} Vertical Overview on {date_label}, y={k_val:.2f}x{b_val:+.2f}{extra_text}",
        fontsize=16,
        y=1.02,
    )
    return fig


def _plot_glorys_overview_vertical_2x2_sigma(
    *,
    vertical_package: dict,
    k_val: float,
    b_val: float,
    subject_label: str,
    date_label: str,
    xmin: float | None = None,
    xmax: float | None = None,
    ymin: float = 23.0,
    ymax: float = 28.0,
    projected_argo_profile: pd.DataFrame | None = None,
    plot_isolines: bool = True,
    isoline_levels: int | list[float] | np.ndarray | None = None,
    isoline_color: str = 'black',
    isoline_linewidth: float = 0.8,
    isoline_alpha: float = 0.45,
    label_isolines: bool = True,
    title_extra: str = '',
):
    """绘制 σ 坐标的 2x2 垂向总览图：PV / Z(σ) / θ / S。

    y 轴为 σ₀ (kg/m³)，向下增大以对应海洋密度结构。
    ``plot_isolines`` 控制各面板自身变量的等值线叠加，
    其中 Z(σ) 面板天然为等深度线。
    """
    sigma_vars = ['pv', 'z_of_sigma', 'thetao', 'salinity']

    fig, axes = plt.subplots(2, 2, figsize=(15, 10), constrained_layout=True)
    axes_flat = axes.ravel()

    y_coords = np.asarray(vertical_package.get('y_coords', []), dtype=float)
    sigma_coords = np.asarray(vertical_package.get('z_coords', []), dtype=float)
    projections = vertical_package.get('projections', {})
    metadata = vertical_package.get('metadata', {})
    draw_reference_lines = bool(metadata.get('draw_reference_lines', True))

    for i, var_key in enumerate(sigma_vars):
        ax = axes_flat[i]
        title_txt, cbar_label, cmap_name, fixed_clim = _overview_var_style(var_key)

        v_field = np.ma.array(
            vertical_package.get('profile_data', {}).get(var_key), copy=False,
        )
        if v_field.ndim != 2 or y_coords.size < 2 or sigma_coords.size < 2:
            ax.text(0.5, 0.5, f"No valid {title_txt}", ha='center', va='center',
                    transform=ax.transAxes)
            ax.set_title(f"{title_txt} ($\\sigma$ coord)")
            continue

        ns = min(v_field.shape[0], len(sigma_coords))
        ny = min(v_field.shape[1], len(y_coords))
        v_field = v_field[:ns, :ny]
        s_plot = sigma_coords[:ns]
        y_plot = y_coords[:ny]

        Y_mesh, S_mesh = np.meshgrid(y_plot, s_plot)
        v_clim = fixed_clim if fixed_clim is not None else _auto_clim(v_field)
        v_field_filled = np.ma.filled(np.ma.array(v_field, copy=False), np.nan)
        pc = ax.pcolormesh(Y_mesh, S_mesh, v_field_filled, cmap=cmap_name, shading='auto',
                           vmin=v_clim[0], vmax=v_clim[1])

        if plot_isolines:
            _draw_isolines(ax, Y_mesh, S_mesh, v_field_filled, v_clim,
                           isoline_levels=isoline_levels,
                           isoline_color=isoline_color,
                           isoline_linewidth=isoline_linewidth,
                           isoline_alpha=isoline_alpha,
                           label_isolines=label_isolines)

        if draw_reference_lines:
            ax.axvline(0.0, color='black', linestyle='--', linewidth=1.6,
                       label='Center Projection' if i == 0 else None)
            for j, dist in enumerate(projections.get('radius', [])):
                ax.axvline(dist, color='r', linestyle='--', linewidth=1.2,
                           label='Radius Projection' if (i == 0 and j == 0) else None)
            for j, dist in enumerate(projections.get('contour', [])):
                ax.axvline(dist, color='tab:blue', linestyle=':', linewidth=1.2,
                           label='Contour Projection' if (i == 0 and j == 0) else None)

        if (projected_argo_profile is not None and not projected_argo_profile.empty
                and 'sigma' in projected_argo_profile.columns):
            valid_pts = (
                np.isfinite(projected_argo_profile['proj_y'].to_numpy(dtype=float))
                & np.isfinite(projected_argo_profile['sigma'].to_numpy(dtype=float))
            )
            if np.any(valid_pts):
                do_vals = projected_argo_profile['DO'].to_numpy(dtype=float)
                if np.isfinite(do_vals[valid_pts]).any():
                    ax.scatter(
                        projected_argo_profile['proj_y'].to_numpy(dtype=float)[valid_pts],
                        projected_argo_profile['sigma'].to_numpy(dtype=float)[valid_pts],
                        c=do_vals[valid_pts],
                        cmap='bwr',
                        vmin=150,
                        vmax=240,
                        s=projected_argo_profile['marker_size'].to_numpy(dtype=float)[valid_pts],
                        edgecolors='black',
                        linewidths=0.35,
                        alpha=0.9,
                        zorder=6,
                        label='Projected Argo' if i == 0 else None,
                    )
                else:
                    ax.scatter(
                        projected_argo_profile['proj_y'].to_numpy(dtype=float)[valid_pts],
                        projected_argo_profile['sigma'].to_numpy(dtype=float)[valid_pts],
                        color='blue',
                        s=projected_argo_profile['marker_size'].to_numpy(dtype=float)[valid_pts],
                        edgecolors='black',
                        linewidths=0.35,
                        alpha=0.9,
                        zorder=6,
                        label='Projected Argo' if i == 0 else None,
                    )

        ax.set_ylim(float(ymax), float(ymin))
        if xmin is not None and xmax is not None:
            ax.set_xlim(float(xmin), float(xmax))

        ax.set_title(f"{title_txt} ($\\sigma$ coord)", fontsize=12)
        ax.set_xlabel('Distance (km)')
        ax.set_ylabel(r'$\sigma_0$ (kg/m$^3$)')
        cbar = fig.colorbar(pc, ax=ax, orientation='vertical', fraction=0.045, pad=0.02)
        cbar.set_label(cbar_label, fontsize=10)
        cbar.ax.tick_params(labelsize=9)
        if i == 0 and draw_reference_lines:
            ax.legend(fontsize=9, loc='best')

    extra_text = f", {title_extra}" if title_extra else ""
    fig.suptitle(
        f"{subject_label} GLORYS $\\sigma$ PV Overview on {date_label}, "
        f"y={k_val:.2f}x{b_val:+.2f}{extra_text}",
        fontsize=16,
        y=1.02,
    )
    return fig


def compute_spiciness_anomaly(
    bg_theta: np.ndarray,
    bg_sal: np.ndarray,
    fg_theta: np.ndarray,
    fg_sal: np.ndarray,
    *,
    center_lat: float = 30.0,
    center_lon: float = 0.0,
    sigma_bandwidth: float = 0.25,
    min_effective_weight: float = 50.0,
) -> tuple[np.ndarray, np.ndarray]:
    """计算前景 T-S 点相对背景分布的带符号 spiciness 异常。

    用高斯核在 σ₀ 空间加权所有背景点，返回 π 的加权中位数偏差 δπ 和加权百分位；P < 10 为显著冷鲜（T-S
    图偏左下），P > 90 为显著暖咸（T-S 图偏右）。百分位不受 σ₀ 范围宽窄影响，中纬高纬通用。

    参数:
        - bg_theta (np.ndarray): 背景位温一维数组。
        - bg_sal (np.ndarray): 背景盐度一维数组。
        - fg_theta (np.ndarray): 前景位温一维数组。
        - fg_sal (np.ndarray): 前景盐度一维数组。
        - center_lat (float): 参考纬度（仅影响 SA 计算，对本函数影响极小），默认 30.0。
        - center_lon (float): 参考经度（同上），默认 0.0。
        - sigma_bandwidth (float): σ₀ 高斯核带宽（kg/m³），越大越平滑，默认 0.25。
        - min_effective_weight (float): 有效权重下限，不足则返回 NaN，默认 50.0。

    返回:
        - tuple: (delta_pi, percentile)，等长一维数组（无法计算处为 NaN）；delta_pi 为带符号 spiciness 偏差，percentile 为 σ₀ 加权背景 π 分布中的百分位（0–100）。
    """
    bg_theta = np.asarray(bg_theta, dtype=float)
    bg_sal = np.asarray(bg_sal, dtype=float)
    fg_theta = np.asarray(fg_theta, dtype=float)
    fg_sal = np.asarray(fg_sal, dtype=float)

    finite_bg = np.isfinite(bg_theta) & np.isfinite(bg_sal)
    bg_theta = bg_theta[finite_bg]
    bg_sal = bg_sal[finite_bg]
    if len(bg_theta) < 30:
        return np.full_like(fg_theta, np.nan), np.full_like(fg_theta, np.nan)

    p_ref = 0.0  # surface-referenced; bg 无深度信息，fg 同理以保持一致
    SA_bg = gsw.SA_from_SP(bg_sal, p_ref, center_lon, center_lat)
    CT_bg = gsw.CT_from_pt(SA_bg, bg_theta)
    sigma0_bg = gsw.sigma0(SA_bg, CT_bg)
    pi_bg = gsw.spiciness0(SA_bg, CT_bg)

    valid_bg = np.isfinite(sigma0_bg) & np.isfinite(pi_bg)
    sigma0_bg = sigma0_bg[valid_bg]
    pi_bg = pi_bg[valid_bg]
    if len(sigma0_bg) < 30:
        return np.full_like(fg_theta, np.nan), np.full_like(fg_theta, np.nan)

    # Sort bg by π ascending once — all fg points share this sorted order
    sort_idx = np.argsort(pi_bg)
    pi_bg_sorted = pi_bg[sort_idx]
    sigma0_bg_sorted = sigma0_bg[sort_idx]

    SA_fg = gsw.SA_from_SP(fg_sal, p_ref, center_lon, center_lat)
    CT_fg = gsw.CT_from_pt(SA_fg, fg_theta)
    sigma0_fg = gsw.sigma0(SA_fg, CT_fg)
    pi_fg = gsw.spiciness0(SA_fg, CT_fg)

    delta_pi = np.full_like(fg_theta, np.nan)
    percentile = np.full_like(fg_theta, np.nan)
    inv_two_s2 = -0.5 / (float(sigma_bandwidth) ** 2)

    for j in range(len(fg_theta)):
        sf = sigma0_fg[j]
        pf = pi_fg[j]
        if not (np.isfinite(sf) and np.isfinite(pf)):
            continue

        # Gaussian weights in σ₀ space
        dsigma = sigma0_bg_sorted - sf
        weights = np.exp(dsigma * dsigma * inv_two_s2)
        total_weight = np.sum(weights)
        if total_weight < float(min_effective_weight):
            continue

        # Weighted median → δπ
        cum_weights = np.cumsum(weights)
        med_pos = np.searchsorted(cum_weights, total_weight / 2.0)
        pi_median_w = float(pi_bg_sorted[min(med_pos, len(pi_bg_sorted) - 1)])
        delta_pi[j] = float(pf - pi_median_w)

        # Weighted percentile of π_fg in bg π distribution
        rank = np.searchsorted(pi_bg_sorted, pf, side='right')
        if rank <= 0:
            percentile[j] = 0.0
        elif rank >= len(pi_bg_sorted):
            percentile[j] = 100.0
        else:
            percentile[j] = float(cum_weights[rank - 1]) / float(total_weight) * 100.0

    return delta_pi, percentile


def _plot_ts_diagram_core(
    bg_theta: np.ndarray,
    bg_sal: np.ndarray,
    argo_rows: pd.DataFrame,
    *,
    center_lat: float = 30.0,
    center_lon: float = 0.0,
    anomaly_depth: float | None = None,
    anomaly_peaks: pd.DataFrame | None = None,
    subject_label: str = '',
    date_label: str = '',
    bg_label: str = 'Background',
    color_by: str = 'depth',
    show_fig: bool = True,
    save_fig: bool = False,
    save_path: str | Path | None = None,
    sigma_contour_levels: list[float] | None = None,
    contour_color: str = 'black',
    contour_linewidth: float = 0.6,
    contour_alpha: float = 0.45,
    label_contours: bool = True,
    annotate_spice: bool = True,
) -> None:
    """T-S 图内核：背景散点 + σ₀ 等值线 + Argo 剖面叠加。

    ``bg_theta`` 与 ``bg_sal`` 为**等长一维**配对数组（调用方负责展平与掩膜过滤）。
    ``argo_rows`` 需含 Temperature / Salinity / Depth 列；
    当 ``color_by='do'`` 时需含 DO 列；当 ``color_by='month'`` 时需含 Month 列（int 1-12）。
    """
    if bg_theta is None or bg_sal is None or len(bg_theta) == 0:
        print("[TS] 背景 θ/S 数据为空，跳过 T-S 图。")
        return None
    if len(bg_theta) != len(bg_sal):
        raise ValueError(f"bg_theta ({len(bg_theta)}) 与 bg_sal ({len(bg_sal)}) 长度不一致。")
    theta_vals = np.asarray(bg_theta, dtype=float)
    sal_vals = np.asarray(bg_sal, dtype=float)
    finite_bg = np.isfinite(theta_vals) & np.isfinite(sal_vals)
    theta_vals = theta_vals[finite_bg]
    sal_vals = sal_vals[finite_bg]
    if theta_vals.size < 5:
        print("[TS] 有效背景 θ/S 点不足 5，跳过 T-S 图。")
        return None
    z_ref = 0.0  # surface-referenced σ₀

    argo_valid = argo_rows.dropna(subset=['Temperature', 'Salinity', 'Depth']) if not argo_rows.empty else argo_rows
    argo_theta = argo_valid['Temperature'].to_numpy(dtype=float) if not argo_valid.empty else np.array([])
    argo_sal = argo_valid['Salinity'].to_numpy(dtype=float) if not argo_valid.empty else np.array([])
    argo_depth = argo_valid['Depth'].to_numpy(dtype=float) if not argo_valid.empty else np.array([])
    has_argo = len(argo_valid) > 0

    all_sal = np.concatenate([sal_vals, argo_sal]) if has_argo else sal_vals
    all_theta = np.concatenate([theta_vals, argo_theta]) if has_argo else theta_vals
    s_min, s_max = float(np.nanmin(all_sal)), float(np.nanmax(all_sal))
    t_min, t_max = float(np.nanmin(all_theta)), float(np.nanmax(all_theta))
    s_pad = max(0.1, (s_max - s_min) * 0.05)
    t_pad = max(0.1, (t_max - t_min) * 0.05)
    s_min, s_max = s_min - s_pad, s_max + s_pad
    t_min, t_max = t_min - t_pad, t_max + t_pad

    n_grid = 100
    S_grid, T_grid = np.meshgrid(
        np.linspace(s_min, s_max, n_grid),
        np.linspace(t_min, t_max, n_grid),
    )
    p_ref = gsw.p_from_z(-z_ref, center_lat)
    SA_grid = gsw.SA_from_SP(S_grid, p_ref, center_lon, center_lat)
    CT_grid = gsw.CT_from_pt(SA_grid, T_grid)
    sigma_grid = gsw.sigma0(SA_grid, CT_grid)

    if sigma_contour_levels is None:
        sigma_valid = sigma_grid[np.isfinite(sigma_grid)]
        if sigma_valid.size < 2:
            sigma_contour_levels = np.linspace(20, 28, 9)
        else:
            lo = np.floor(np.nanmin(sigma_valid))
            hi = np.ceil(np.nanmax(sigma_valid))
            sigma_contour_levels = np.arange(lo, hi + 0.5, 0.5)

    fig, ax = plt.subplots(figsize=(8, 7), constrained_layout=True)

    ax.scatter(sal_vals, theta_vals, c='lightgray', s=2, alpha=0.5,
               rasterized=True, zorder=1, label=bg_label)

    cs = ax.contour(S_grid, T_grid, sigma_grid, levels=sigma_contour_levels,
                    colors=contour_color, linewidths=contour_linewidth,
                    alpha=contour_alpha, zorder=2)
    if label_contours:
        ax.clabel(cs, inline=True, fontsize=8, fmt='%.1f')

    if has_argo:
        if color_by == 'do' and 'DO' in argo_valid.columns:
            c_arr = argo_valid['DO'].to_numpy(dtype=float)
            cmap, cbar_label = 'RdBu_r', 'DO (μmol/kg)'
        elif color_by == 'month' and 'Month' in argo_valid.columns:
            c_arr = argo_valid['Month'].to_numpy(dtype=float)
            cmap, cbar_label = 'hsv', 'Month'
        elif color_by == 'none':
            c_arr, cmap, cbar_label = 'steelblue', None, None
        else:  # 'depth' or fallback
            c_arr, cmap, cbar_label = argo_depth, 'viridis', 'Depth (m)'

        sc = ax.scatter(argo_sal, argo_theta, c=c_arr, cmap=cmap,
                        s=30, edgecolors='black', linewidths=0.3,
                        zorder=3, label='Argo')
        if cbar_label:
            cbar = fig.colorbar(sc, ax=ax, orientation='vertical', fraction=0.045, pad=0.02)
            cbar.set_label(cbar_label, fontsize=10)

    if anomaly_peaks is not None and not anomaly_peaks.empty:
        peaks = anomaly_peaks.dropna(subset=['Salinity', 'Temperature'])
        if not peaks.empty:
            ax.scatter(peaks['Salinity'], peaks['Temperature'], marker='*',
                       c='red', s=100, edgecolors='black', linewidths=0.5,
                       zorder=5, label='Anomaly peak')
    elif anomaly_depth is not None and np.isfinite(anomaly_depth) and has_argo:
        argo_depths_sorted = np.argsort(argo_depth)
        ad_theta = float(np.interp(anomaly_depth, argo_depth[argo_depths_sorted],
                                    argo_theta[argo_depths_sorted], left=np.nan, right=np.nan))
        ad_sal = float(np.interp(anomaly_depth, argo_depth[argo_depths_sorted],
                                  argo_sal[argo_depths_sorted], left=np.nan, right=np.nan))
        if np.isfinite(ad_theta) and np.isfinite(ad_sal):
            ax.scatter([ad_sal], [ad_theta], marker='*', c='red', s=200,
                       edgecolors='black', linewidths=0.5, zorder=5,
                       label=f'Anomaly ({anomaly_depth:.0f} m)')

    spice_label = ''
    if annotate_spice and anomaly_peaks is not None and not anomaly_peaks.empty:
        peaks = anomaly_peaks.dropna(subset=['Salinity', 'Temperature'])
        if not peaks.empty:
            dp_arr, pct_arr = compute_spiciness_anomaly(
                theta_vals, sal_vals,
                peaks['Temperature'].to_numpy(dtype=float),
                peaks['Salinity'].to_numpy(dtype=float),
                center_lat=center_lat, center_lon=center_lon,
            )
            valid = np.isfinite(dp_arr) & np.isfinite(pct_arr)
            if valid.any():
                dp_v = dp_arr[valid]
                pct_v = pct_arr[valid]
                if len(pct_v) == 1:
                    spice_label = f'δπ={dp_v[0]:+.3f}, {pct_v[0]:.0f}%'
                else:
                    spice_label = (
                        f'δπ={np.mean(dp_v):+.3f}, '
                        f'{np.min(pct_v):.0f}%–{np.max(pct_v):.0f}%'
                    )
    elif annotate_spice and anomaly_depth is not None and np.isfinite(anomaly_depth) and has_argo:
        ad_sal_arr = np.array([ad_sal], dtype=float)
        ad_theta_arr = np.array([ad_theta], dtype=float)
        dp_single, pct_single = compute_spiciness_anomaly(
            theta_vals, sal_vals,
            ad_theta_arr, ad_sal_arr,
            center_lat=center_lat, center_lon=center_lon,
        )
        if np.isfinite(pct_single[0]) and np.isfinite(dp_single[0]):
            spice_label = f'δπ={dp_single[0]:+.3f}, {pct_single[0]:.0f}%'

    ax.set_xlabel('Salinity (psu)', fontsize=12)
    ax.set_ylabel('Temperature (°C)', fontsize=12)
    title_parts = []
    if subject_label:
        title_parts.append(subject_label)
    title_parts.append('T-S Diagram')
    if date_label:
        title_parts.append(date_label)
    if spice_label:
        title_parts.append(spice_label)
    ax.set_title(' | '.join(title_parts), fontsize=14)
    ax.legend(fontsize=9, loc='best', markerscale=0.8)
    ax.grid(True, alpha=0.2)

    if save_fig and save_path is not None:
        Path(save_path).parent.mkdir(parents=True, exist_ok=True)
        fig.savefig(save_path, dpi=300, bbox_inches='tight')
    if show_fig:
        plt.show()
    plt.close(fig)


def plot_argo_ts_diagram(
    profile_number: int,
    profile_time: int | str | pd.Timestamp,
    k: float | list[float] | None = None,
    b: float | list[float] | None = None,
    *,
    platform_number: int | None = None,
    detection_config: DetectionConfig | None = None,
    color_by: str = 'depth',
    xmin: float = -400.0,
    xmax: float = 400.0,
    profile_spacing_km: float | None = None,
    interpolate_z: bool = True,
    profile_depth_spacing_m: float | None = None,
    show_fig: bool = True,
    save_fig: bool = False,
    output_dir: str | Path | None = None,
    sigma_contour_levels: list[float] | None = None,
    contour_color: str = 'black',
    contour_linewidth: float = 0.6,
    contour_alpha: float = 0.45,
    label_contours: bool = True,
    annotate_spice: bool = True,
) -> None:
    """绘制单个 Argo 剖面的温盐图（GLORYS 垂向切片 + Argo 叠加）。

    给定剖面编号和观测日期，自动加载 Argo 剖面数据与沿 GLORYS 垂向切片的 θ/S 背景场，叠加 σ₀ 等密度线，
    适合快速查看某个特定剖面的水团属性与异常深度位置。

    参数:
        - profile_number (int): Argo 剖面编号。
        - profile_time (int | str | pd.Timestamp): 剖面日期（int YYYYMMDD / 'YYYY-MM-DD' / Timestamp）。
        - k (float | list[float] | None): GLORYS 垂向剖面线斜率，y = kx + b（支持 list，每条线出一张图；None 时默认纬向线 k=0、b=center_lat）。
        - b (float | list[float] | None): 剖面线截距，与 k 配套。
        - platform_number (int | None): 浮标平台编号，辅助定位（可选）。
        - detection_config (DetectionConfig | None): 异常检测配置，传入后在异常峰值深度叠加 ★ 标记。
        - color_by (str): Argo 点着色方式 'depth'/'do'/'month'/'none'，默认 'depth'。
        - xmin (float): 垂向剖面窗口下界（km），默认 -400.0。
        - xmax (float): 垂向剖面窗口上界（km），默认 400.0。
        - profile_spacing_km (float | None): 剖面内采样间距（km），None 时回退默认。
        - interpolate_z (bool): 是否对 GLORYS 垂向插值，默认 True。
        - profile_depth_spacing_m (float | None): 垂向插值间距（m）。
        - show_fig (bool): 是否显示图片，默认 True。
        - save_fig (bool): 是否保存图片，默认 False。
        - output_dir (str | Path | None): 图片输出目录，None 时使用默认路径。
        - sigma_contour_levels (list[float] | None): σ₀ 等密度线层级，None 时按数据范围自动推断。
        - contour_color (str): 等密度线颜色，默认 'black'。
        - contour_linewidth (float): 等密度线线宽，默认 0.6。
        - contour_alpha (float): 等密度线透明度，默认 0.45。
        - label_contours (bool): 是否在等密度线上标注 σ₀ 值，默认 True。
        - annotate_spice (bool): 是否在标题标注异常峰值的 spiciness 异常 δπ，默认 True。
    """
    info = _resolve_argo_profile_center(
        profile_number=int(profile_number),
        profile_time=profile_time,
        platform_number=platform_number,
    )
    center_lon = float(info['center_lon'])
    center_lat = float(info['center_lat'])
    target_date = pd.Timestamp(info['target_date'])
    argo_rows = info['profile_rows'].copy()

    window_half_size_km = max(abs(float(xmin)), abs(float(xmax)))
    pkgs = get_vertical_glorys_from_center(
        center_lon, center_lat, target_date, k, b,
        variables=['thetao', 'salinity'],
        x_min_km=xmin, x_max_km=xmax,
        profile_spacing_km=profile_spacing_km,
        interpolate_z=interpolate_z,
        profile_depth_spacing_m=profile_depth_spacing_m,
        window_half_size_km=window_half_size_km,
        profile_id=int(profile_number),
        ds_name='ARGO',
    )
    pkg = pkgs[0] if pkgs else {}
    pd_data = pkg.get('profile_data', {})
    theta_gl = pd_data.get('thetao')
    sal_gl = pd_data.get('salinity')
    if theta_gl is None or sal_gl is None:
        print("[TS] GLORYS 缺少 thetao 或 salinity，跳过 T-S 图。")
        return

    theta_ma = np.ma.array(theta_gl, copy=False)
    sal_ma = np.ma.array(sal_gl, copy=False)
    valid_gl = ~np.ma.getmaskarray(theta_ma) & ~np.ma.getmaskarray(sal_ma)
    if valid_gl.sum() < 5:
        print("[TS] GLORYS 有效 θ/S 点不足，跳过 T-S 图。")
        return
    bg_theta = np.asarray(theta_ma[valid_gl], dtype=float)
    bg_sal = np.asarray(sal_ma[valid_gl], dtype=float)

    anomaly_peaks = None
    if detection_config is not None:
        anomaly_rows = _reduce_argo_profiles_by_anomaly(argo_rows, detection_config=detection_config)
        if not anomaly_rows.empty:
            anomaly_peaks = anomaly_rows

    k_val = k if isinstance(k, (int, float)) else (k[0] if k else 0.0)
    b_val = b if isinstance(b, (int, float)) else (b[0] if b else float(center_lat))
    date_str = target_date.strftime('%Y-%m-%d')
    save_path = None
    if save_fig:
        region_slug = _current_region_key()
        out_dir = Path(output_dir) if output_dir is not None else Path('plot_outputs') / 'shared' / region_slug / 'plot_ts_diagram'
        filename = f"Argo_P{int(profile_number)}_{date_str}_k{k_val:.2f}b{b_val:+.2f}.png"
        save_path = out_dir / filename

    _plot_ts_diagram_core(
        bg_theta, bg_sal, argo_rows,
        center_lat=center_lat, center_lon=center_lon,
        anomaly_peaks=anomaly_peaks,
        subject_label=f"Profile {int(profile_number)}",
        date_label=date_str,
        bg_label='GLORYS',
        color_by=color_by,
        show_fig=show_fig, save_fig=save_fig, save_path=save_path,
        sigma_contour_levels=sigma_contour_levels,
        contour_color=contour_color, contour_linewidth=contour_linewidth,
        contour_alpha=contour_alpha, label_contours=label_contours,
        annotate_spice=annotate_spice,
    )


def plot_track_ts_diagram(
    DS: str,
    no: int,
    start_date: str | pd.Timestamp,
    end_date: str | pd.Timestamp,
    *,
    background: str = 'argo',
    detection_config: DetectionConfig | None = None,
    color_by: str = 'depth',
    show_fig: bool = True,
    save_fig: bool = False,
    output_dir: str | Path | None = None,
    annotate_spice: bool = True,
) -> None:
    """沿涡旋轨迹的聚合温盐图（拉格朗日视角）。

    追踪指定涡旋在时间窗口内的移动路径，收集路径附近所有匹配 Argo 剖面叠加在背景 T-S 散点之上；背景可选
    Argo（区域内全量剖面）或 GLORYS（逐月收集涡旋有效半径内 θ/S 散点），适合分析特定涡旋演化过程中水团
    属性的系统性变化。

    参数:
        - DS (str): kind 字符串（'acs'|'acl'|'cs'|'cl'）。
        - no (int): 涡旋编号。
        - start_date (str | pd.Timestamp): 时间窗口起始。
        - end_date (str | pd.Timestamp): 时间窗口结束。
        - background (str): 背景源，'argo' 用全量 Argo，'glorys' 用涡旋半径内 GLORYS 逐月聚合，默认 'argo'。
        - detection_config (DetectionConfig | None): 异常检测配置，传入后仅异常剖面在前台叠加、峰值标 ★。
        - color_by (str): Argo 点着色方式 'depth'/'do'/'month'/'none'，默认 'depth'。
        - show_fig (bool): 是否显示图片，默认 True。
        - save_fig (bool): 是否保存图片，默认 False。
        - output_dir (str | Path | None): 图片输出目录，None 时使用默认路径。
        - annotate_spice (bool): 是否在标题标注异常峰值的 spiciness 异常 δπ，默认 True。
    """
    # --- 1. 加载轨迹并按时间过滤 ---
    track_df = find_track(DS, no)
    if track_df is None or (isinstance(track_df, pd.DataFrame) and track_df.empty):
        print(f"[TS] Track for eddy {no} not found, returning None.")
        return None
    track_df = track_df.copy()
    if 'date' not in track_df.columns:
        track_df['date'] = convert_date(track_df['time'])
    start_ts, end_ts = pd.Timestamp(start_date), pd.Timestamp(end_date)
    track_df = track_df[(track_df['date'] >= start_ts) & (track_df['date'] <= end_ts)]
    if track_df.empty:
        print(f"[TS] Track {no} has no data in [{start_date}, {end_date}].")
        return None

    # --- 2. 匹配 Argo 浮标 ---
    print(f"[TS] Loading Argo floats for eddy {no} in [{start_date}, {end_date}]...")
    argo_data = filtered_float_data(DS, no, track=track_df)
    if argo_data.empty:
        print("[TS] No matching Argo floats found.")
        return None
    if 'date' not in argo_data.columns:
        argo_data['date'] = pd.to_datetime(argo_data[['Year', 'Month', 'Day']])
    argo_data = argo_data[(argo_data['date'] >= start_ts) & (argo_data['date'] <= end_ts)]
    if argo_data.empty:
        print("[TS] No Argo data in the specified date range.")
        return None

    # --- 3. 准备背景数据（全量 Argo 或 GLORYS）---
    bg_theta: np.ndarray
    bg_sal: np.ndarray
    if background == 'argo':
        argo_full = argo_data.dropna(subset=['Temperature', 'Salinity'])
        if len(argo_full) < 50:
            print("[TS] Fewer than 50 valid θ/S points; background scatter will be skipped.")
        bg_theta = argo_full['Temperature'].to_numpy(dtype=float)
        bg_sal = argo_full['Salinity'].to_numpy(dtype=float)
    elif background == 'glorys':
        # 逐月收集涡旋半径 + 轮廓多边形内的 GLORYS θ/S
        circle_factor = globals().get('circle_enlargement_factor', 1.2)
        radii_m = pd.to_numeric(track_df['radius'], errors='coerce')
        clats = pd.to_numeric(track_df['center_lat'], errors='coerce')
        radius_deg = radii_m / (111320.0 * np.cos(np.radians(clats.clip(-80, 80))))
        track_df['_r_deg'] = radius_deg * float(circle_factor)
        track_df['_ym'] = track_df['date'].dt.to_period('M')

        bg_parts_theta, bg_parts_sal = [], []
        for _ym, group in track_df.groupby('_ym', sort=True):
            mid_idx = len(group) // 2
            ref_date = group['date'].iloc[mid_idx]
            ref_lon = float(_normalize_lon_array(group['center_lon'].mean()))
            ref_lat = float(group['center_lat'].mean())
            # 窗口半宽 = max(圆形半径, 轮廓边界框)
            half_deg = float(group['_r_deg'].max()) + 0.25
            for _, row in group.iterrows():
                clon_arr = np.asarray(row['contour_lon'], dtype=float)
                clat_arr = np.asarray(row['contour_lat'], dtype=float)
                if clon_arr.size < 3 or np.all(np.abs(clon_arr) >= 179.9):
                    continue
                clon_norm = ref_lon + _minimal_lon_diff_deg(clon_arr, ref_lon)
                half_deg = max(half_deg,
                               float(np.abs(clon_norm - ref_lon).max()) + 0.25,
                               float(np.abs(clat_arr - ref_lat).max()) + 0.25)
            lon_lo = ref_lon - half_deg
            lon_hi = ref_lon + half_deg
            lat_lo = ref_lat - half_deg
            lat_hi = ref_lat + half_deg
            try:
                _gl, _ga, _gd, gv = _load_glorys_window_by_center(
                    ref_date, ref_lon, lon_lo, lon_hi, lat_lo, lat_hi,
                    ['thetao', 'salinity'],
                )
            except (ValueError, FileNotFoundError):
                continue
            th = gv.get('thetao')
            sa = gv.get('salinity')
            if th is None or sa is None:
                continue
            th_ma = np.ma.array(th, copy=False)
            sa_ma = np.ma.array(sa, copy=False)
            valid_data = ~np.ma.getmaskarray(th_ma) & ~np.ma.getmaskarray(sa_ma)
            if valid_data.sum() == 0:
                continue
            # 构建网格坐标
            glon_1d = np.asarray(_gl, dtype=float)
            glat_1d = np.asarray(_ga, dtype=float)
            glon_2d, glat_2d = np.meshgrid(glon_1d, glat_1d)
            spatial_mask = np.zeros(glon_2d.shape, dtype=bool)
            # 逐日累积：圆形 + 轮廓多边形
            for _, row in group.iterrows():
                day_lon = float(row['center_lon'])
                day_lat = float(row['center_lat'])
                day_radius_m = float(row['radius'])
                dist_m = adaptive_distance_m(
                    glon_2d.ravel(), glat_2d.ravel(),
                    day_lon, day_lat, wrap_dateline=True,
                ).reshape(glon_2d.shape)
                day_mask = dist_m <= day_radius_m * float(circle_factor)
                # 轮廓多边形
                clon = np.asarray(row['contour_lon'], dtype=float)
                clat = np.asarray(row['contour_lat'], dtype=float)
                if clon.size >= 3 and not np.all(np.abs(clon) >= 179.9):
                    clon_n = day_lon + _minimal_lon_diff_deg(clon, day_lon)
                    verts = np.column_stack([clon_n, clat])
                    try:
                        poly = MplPath(verts)
                        day_mask |= poly.contains_points(
                            np.column_stack([glon_2d.ravel(), glat_2d.ravel()])
                        ).reshape(glon_2d.shape)
                    except Exception:
                        pass
                spatial_mask |= day_mask
            # 应用到所有深度层
            full_mask = valid_data & spatial_mask[np.newaxis, :, :]
            if full_mask.sum() > 0:
                bg_parts_theta.append(np.asarray(th_ma[full_mask], dtype=float))
                bg_parts_sal.append(np.asarray(sa_ma[full_mask], dtype=float))

        if not bg_parts_theta:
            print("[TS] GLORYS 未收集到有效 θ/S 点，无法生成背景。")
            return None
        bg_theta = np.concatenate(bg_parts_theta)
        bg_sal = np.concatenate(bg_parts_sal)
        if bg_theta.size < 50:
            print(f"[TS] GLORYS 有效 θ/S 配对点仅 {bg_theta.size}，过少。")
            return None
    else:
        raise ValueError(f"background 须为 'argo' 或 'glorys'，收到 '{background}'。")

    # --- 4. 前台：异常剖面全深度 + 峰值 ★ ---
    argo_all = argo_full if background == 'argo' else argo_data
    anomaly_rows = _reduce_argo_profiles_by_anomaly(argo_all, detection_config=detection_config)
    if anomaly_rows.empty:
        print("[TS] No anomalous profiles detected.")
        return None
    anomaly_pns = anomaly_rows['Profile_number'].unique()
    argo_fg = argo_all[argo_all['Profile_number'].isin(anomaly_pns)]

    # --- 5. 绘图 ---
    region_slug = _current_region_key()
    out_dir = Path(output_dir) if output_dir is not None else Path('plot_outputs') / 'shared' / region_slug / 'plot_ts_diagram'

    kind = DS.lower() if isinstance(DS, str) else 'unknown'
    sd_tag = start_ts.strftime('%Y%m%d')
    ed_tag = end_ts.strftime('%Y%m%d')
    filename = f"{kind}{no}_ts_{sd_tag}_{ed_tag}.png"
    save_path = out_dir / filename if save_fig else None

    date_label = f"{start_ts.strftime('%Y-%m-%d')} – {end_ts.strftime('%Y-%m-%d')}"
    _plot_ts_diagram_core(
        bg_theta, bg_sal, argo_fg,
        center_lat=float(track_df['center_lat'].mean()),
        center_lon=float(track_df['center_lon'].mean()),
        anomaly_peaks=anomaly_rows,
        subject_label=f"{kind.upper()} {no}",
        date_label=date_label,
        bg_label='Argo (bkg)' if background == 'argo' else 'GLORYS',
        color_by=color_by,
        show_fig=show_fig, save_fig=save_fig, save_path=save_path,
        annotate_spice=annotate_spice,
    )


def plot_regional_ts_diagram(
    lon_range: tuple[float, float],
    lat_range: tuple[float, float],
    start_date: str | pd.Timestamp,
    end_date: str | pd.Timestamp,
    *,
    background: str = 'argo',
    detection_config: DetectionConfig | None = None,
    color_by: str = 'depth',
    show_fig: bool = True,
    save_fig: bool = False,
    output_dir: str | Path | None = None,
    annotate_spice: bool = True,
) -> None:
    """固定区域的聚合温盐图（欧拉视角）。

    在指定经纬度范围内收集时间窗口内所有 Argo 剖面叠加在背景 T-S 散点之上；背景可选 Argo（区域内全量
    剖面）或 GLORYS（时间窗口中点的区域场），适合分析特定海域的水团结构及其异常剖面的 T-S 分布特征。

    参数:
        - lon_range (tuple[float, float]): (lon_min, lon_max)，允许跨日界线（如 170, -170）。
        - lat_range (tuple[float, float]): (lat_min, lat_max)。
        - start_date (str | pd.Timestamp): 时间窗口起始。
        - end_date (str | pd.Timestamp): 时间窗口结束。
        - background (str): 背景源，'argo' 用全量 Argo，'glorys' 用 GLORYS 区域场，默认 'argo'。
        - detection_config (DetectionConfig | None): 异常检测配置，传入后仅异常剖面在前台叠加、峰值标 ★。
        - color_by (str): Argo 点着色方式 'depth'/'do'/'month'/'none'，默认 'depth'。
        - show_fig (bool): 是否显示图片，默认 True。
        - save_fig (bool): 是否保存图片，默认 False。
        - output_dir (str | Path | None): 图片输出目录，None 时使用默认路径。
        - annotate_spice (bool): 是否在标题标注异常峰值的 spiciness 异常 δπ，默认 True。
    """
    lon_min, lon_max = float(lon_range[0]), float(lon_range[1])
    lat_min, lat_max = float(lat_range[0]), float(lat_range[1])
    start_ts, end_ts = pd.Timestamp(start_date), pd.Timestamp(end_date)

    # --- 1. 按年加载 Argo 并空间过滤 ---
    print(f"[TS] Loading Argo data for region ({lon_min}, {lon_max}), ({lat_min}, {lat_max})...")
    all_years = []
    for year in range(start_ts.year, end_ts.year + 1):
        try:
            yearly = load_argo_data(year)
        except FileNotFoundError:
            continue
        if yearly.empty:
            continue
        lon_mask = _region_lon_mask(yearly['Longitude'].to_numpy(dtype=float), lon_min, lon_max)
        lat_vals = yearly['Latitude'].to_numpy(dtype=float)
        lat_mask = (lat_vals >= lat_min) & (lat_vals <= lat_max)
        yearly = yearly[lon_mask & lat_mask]
        if not yearly.empty:
            all_years.append(yearly)

    if not all_years:
        print("[TS] No Argo data found in the specified region.")
        return None
    argo_data = pd.concat(all_years, ignore_index=True)

    # --- 2. 时间过滤 ---
    if 'date' not in argo_data.columns:
        argo_data['date'] = pd.to_datetime(argo_data[['Year', 'Month', 'Day']])
    argo_data = argo_data[(argo_data['date'] >= start_ts) & (argo_data['date'] <= end_ts)]
    n_profiles = argo_data['Profile_number'].nunique() if 'Profile_number' in argo_data.columns else len(argo_data)
    print(f"[TS] Found {n_profiles} unique profiles in the region and time window.")
    if argo_data.empty:
        print("[TS] No Argo data in the specified date range.")
        return None

    # --- 3. 准备背景数据（全量 Argo 或 GLORYS）---
    bg_theta: np.ndarray
    bg_sal: np.ndarray
    if background == 'argo':
        argo_full = argo_data.dropna(subset=['Temperature', 'Salinity'])
        bg_theta = argo_full['Temperature'].to_numpy(dtype=float)
        bg_sal = argo_full['Salinity'].to_numpy(dtype=float)
    elif background == 'glorys':
        mid_date = start_ts + (end_ts - start_ts) / 2
        # 跨日线安全：用 _minimal_lon_diff_deg 计算区间中点
        center_lon_ref = float(_normalize_lon_array(
            lon_min + float(_minimal_lon_diff_deg(lon_max, lon_min)) / 2.0
        ))
        half_span = abs(float(_minimal_lon_diff_deg(lon_max, lon_min))) / 2.0
        # _load_glorys_window_by_center 的 lon_min/max 是绝对经度，非偏移量
        lon_min_local = center_lon_ref - half_span
        lon_max_local = center_lon_ref + half_span
        try:
            _g_lon, _g_lat, _g_depth, g_vars = _load_glorys_window_by_center(
                mid_date, center_lon_ref, lon_min_local, lon_max_local,
                lat_min, lat_max, ['thetao', 'salinity'],
            )
        except (ValueError, FileNotFoundError) as exc:
            print(f"[TS] GLORYS 加载失败: {exc}")
            return None
        theta_gl = g_vars.get('thetao')
        sal_gl = g_vars.get('salinity')
        if theta_gl is None or sal_gl is None:
            print("[TS] GLORYS 缺少 thetao 或 salinity。")
            return None
        theta_ma = np.ma.array(theta_gl, copy=False)
        sal_ma = np.ma.array(sal_gl, copy=False)
        valid_mask = ~np.ma.getmaskarray(theta_ma) & ~np.ma.getmaskarray(sal_ma)
        bg_theta = np.asarray(theta_ma[valid_mask], dtype=float)
        bg_sal = np.asarray(sal_ma[valid_mask], dtype=float)
        if bg_theta.size < 50:
            print(f"[TS] GLORYS 有效 θ/S 配对点仅 {bg_theta.size}，过少。")
            return None
    else:
        raise ValueError(f"background 须为 'argo' 或 'glorys'，收到 '{background}'。")

    # --- 4. 前台：异常剖面全深度 + 峰值 ★ ---
    argo_all = argo_full if background == 'argo' else argo_data
    anomaly_rows = _reduce_argo_profiles_by_anomaly(argo_all, detection_config=detection_config)
    if anomaly_rows.empty:
        print("[TS] No anomalous profiles detected.")
        return None
    anomaly_pns = anomaly_rows['Profile_number'].unique()
    argo_fg = argo_all[argo_all['Profile_number'].isin(anomaly_pns)]

    # --- 5. 绘图 ---
    # 跨日线安全：用区域几何中心（而非 Argo 坐标均值）作为 σ₀ 参考点
    center_lat = (lat_min + lat_max) / 2.0
    center_lon = float(_normalize_lon_array(
        lon_min + float(_minimal_lon_diff_deg(lon_max, lon_min)) / 2.0
    ))

    region_slug = _current_region_key()
    out_dir = Path(output_dir) if output_dir is not None else Path('plot_outputs') / 'shared' / region_slug / 'plot_ts_diagram'
    sd_tag = start_ts.strftime('%Y%m%d')
    ed_tag = end_ts.strftime('%Y%m%d')
    lon0_s = f"{abs(lon_min):.0f}{'E' if lon_min >= 0 else 'W'}"
    lon1_s = f"{abs(lon_max):.0f}{'E' if lon_max >= 0 else 'W'}"
    lat0_s = f"{abs(lat_min):.0f}{'N' if lat_min >= 0 else 'S'}"
    lat1_s = f"{abs(lat_max):.0f}{'N' if lat_max >= 0 else 'S'}"
    filename = f"regional_ts_{lon0_s}-{lon1_s}_{lat0_s}-{lat1_s}_{sd_tag}_{ed_tag}.png"
    save_path = out_dir / filename if save_fig else None

    date_label = f"{start_ts.strftime('%Y-%m-%d')} – {end_ts.strftime('%Y-%m-%d')}"
    _plot_ts_diagram_core(
        bg_theta, bg_sal, argo_fg,
        center_lat=center_lat, center_lon=center_lon,
        anomaly_peaks=anomaly_rows,
        subject_label=f"Regional ({lon_min:.0f}°–{lon_max:.0f}°, {lat_min:.0f}°–{lat_max:.0f}°)",
        date_label=date_label,
        bg_label='Argo (bkg)' if background == 'argo' else 'GLORYS',
        color_by=color_by,
        show_fig=show_fig, save_fig=save_fig, save_path=save_path,
        annotate_spice=annotate_spice,
    )


def _run_vertical_overview_batch(
    *,
    vertical_packages: list[dict],
    k_list: list[float],
    b_list: list[float],
    vertical_vars: list[str],
    target_date: pd.Timestamp,
    subject_label: str,
    save_name_prefix: str,
    save_subdir: str,
    xmin: float,
    xmax: float,
    ymin: float,
    ymax: float,
    projected_argo_rows: pd.DataFrame,
    projection_distance_scale_km: float | None,
    plot_argo_projection: bool,
    plot_mlt: bool,
    plot_isolines: bool,
    isoline_levels: int | list[float] | np.ndarray | None,
    isoline_color: str,
    isoline_linewidth: float,
    isoline_alpha: float,
    label_isolines: bool,
    detection_config: DetectionConfig | None,
    z_overview: bool = True,
    sigma_overview: bool = False,
    sigma_ymin: float = 23.0,
    sigma_ymax: float = 28.0,
    ts_diagram: bool = False,
    argo_profile_rows: pd.DataFrame | None = None,
    annotate_heave: bool = True,
    heave_projection_depth_m: float | None = None,
    heave_x_window_km: float = 25.0,
    heave_z_window_m: float | None = 100.0,
    heave_search_range: float = _heave_search_range,
    heave_depth_threshold: float = _heave_depth_threshold,
    heave_z_search_m: float | None = _heave_z_search_m,
    annotate_spice: bool = True,
    anomaly_sal: float | None = None,
    anomaly_theta: float | None = None,
    show_fig: bool,
    save_fig: bool,
    output_dir: str | Path | None,
    verbose: bool,
) -> list[dict]:
    """按多条 k/b 批量绘制 vertical overview，并统一处理保存与显示。

    ``z_overview=True`` 绘制 z 坐标 2x2 总览图，
    ``sigma_overview=True`` 绘制 σ 坐标 2x2 总览图（PV / Z(σ) / θ / S），
    ``ts_diagram=True`` 绘制 T-S 图（需传入 ``argo_profile_rows``）。
    ``annotate_heave`` / ``annotate_spice`` 控制对应诊断量的计算与结果输出。
    """
    results: list[dict] = []
    region_slug = _current_region_key()
    cfg = _resolve_detection_config(detection_config)
    run_tag = cfg.file_stem()
    out_dir = Path(output_dir) if output_dir is not None else cfg.output_dir(save_subdir, region_slug)
    if save_fig:
        out_dir.mkdir(parents=True, exist_ok=True)

    date_str = target_date.strftime('%Y-%m-%d')
    date_tag = target_date.strftime('%Y%m%d')

    for i, (k_val, b_val) in enumerate(zip(k_list, b_list)):
        pkg = vertical_packages[i] if i < len(vertical_packages) else {}
        if not pkg:
            print(f"Warning: empty vertical package for line {i+1} (k={k_val}, b={b_val}), skipped.")
            continue

        proj_profile = _project_argo_rows_to_profile_for_overview(
            pkg,
            projected_argo_rows,
            distance_scale_km=projection_distance_scale_km,
        ) if plot_argo_projection else pd.DataFrame()

        heave_diag = calculate_glorys_vertical_profile_diagnostics(
            pkg,
            projection_x_km=0.0,
            projection_depth_m=heave_projection_depth_m,
            x_window_km=heave_x_window_km,
            z_window_m=heave_z_window_m,
            heave_search_range=heave_search_range,
            heave_depth_threshold=heave_depth_threshold,
            heave_z_search_m=heave_z_search_m,
        ) if annotate_heave else {}
        heave_title_extra = ''
        sigma_title_extra = ''
        if annotate_heave and heave_diag:
            heave_val = heave_diag.get('glorys_heave_m')
            zmin_val = heave_diag.get('glorys_heave_zmin')
            parts = []
            if heave_val is not None and np.isfinite(float(heave_val)):
                parts.append(rf"H={heave_val:.0f}\,\mathrm{{m}}")
            if zmin_val is not None and np.isfinite(float(zmin_val)):
                parts.append(rf"z_{{\mathrm{{min}}}}={zmin_val:.0f}\,\mathrm{{m}}")
            if parts:
                heave_title_extra = rf"${', '.join(parts)}$"
            # σ 坐标标题：显示 heave 峰值密度面
            sigma_peak = heave_diag.get('glorys_heave_sigma_peak')
            if sigma_peak is not None and np.isfinite(float(sigma_peak)):
                sigma_title_extra = (
                    rf"$\sigma_{{\mathrm{{peak}}}}={float(sigma_peak):.2f}\,\mathrm{{kg/m^3}}$"
                )

        spice_anomaly_val = np.nan
        spice_percentile_val = np.nan
        if annotate_spice and anomaly_sal is not None and anomaly_theta is not None:
            pd_data = pkg.get('profile_data', {})
            theta_gl = pd_data.get('thetao')
            sal_gl = pd_data.get('salinity')
            if theta_gl is not None and sal_gl is not None:
                theta_ma = np.ma.array(theta_gl, copy=False)
                sal_ma = np.ma.array(sal_gl, copy=False)
                valid_gl = ~np.ma.getmaskarray(theta_ma) & ~np.ma.getmaskarray(sal_ma)
                bg_theta = np.asarray(theta_ma[valid_gl], dtype=float)
                bg_sal = np.asarray(sal_ma[valid_gl], dtype=float)
                center_lat_sp = float(np.nanmean(np.asarray(pkg.get('lat_coords', [30.0]), dtype=float)))
                center_lon_sp = float(np.nanmean(np.asarray(pkg.get('lon_coords', [0.0]), dtype=float)))
                dp_arr, pct_arr = compute_spiciness_anomaly(
                    bg_theta, bg_sal,
                    np.array([anomaly_theta]), np.array([anomaly_sal]),
                    center_lat=center_lat_sp, center_lon=center_lon_sp,
                )
                if np.isfinite(dp_arr[0]) and np.isfinite(pct_arr[0]):
                    spice_anomaly_val = float(dp_arr[0])
                    spice_percentile_val = float(pct_arr[0])

        title_extra = heave_title_extra
        save_path = None
        if z_overview:
            fig = _plot_glorys_overview_vertical_2x2(
                vertical_package=pkg,
                variables=vertical_vars,
                k_val=k_val,
                b_val=b_val,
                subject_label=subject_label,
                date_label=date_str,
                xmin=xmin,
                xmax=xmax,
                ymin=ymin,
                ymax=ymax,
                projected_argo_profile=proj_profile,
                plot_mlt=plot_mlt,
                plot_isolines=plot_isolines,
                isoline_levels=isoline_levels,
                isoline_color=isoline_color,
                isoline_linewidth=isoline_linewidth,
                isoline_alpha=isoline_alpha,
                label_isolines=label_isolines,
                title_extra=title_extra,
            )

            if save_fig:
                if '{date}' in save_name_prefix:
                    filename_core = save_name_prefix.format(date=date_tag)
                    filename = f"{filename_core}_z_k{k_val:.2f}b{b_val:+.2f}_{run_tag}.png"
                else:
                    filename = f"{save_name_prefix}_z_{date_tag}_k{k_val:.2f}b{b_val:+.2f}_{run_tag}.png"
                save_path = out_dir / filename
                fig.savefig(save_path, dpi=300, bbox_inches='tight')
                if verbose:
                    print(f"Figure saved to: {save_path}")

            if show_fig:
                plt.show()
            plt.close(fig)

        sigma_save_path = None
        if sigma_overview and 'sigma' in pkg.get('profile_data', {}):
            try:
                sigma_pkg = _remap_vertical_package_to_sigma(pkg)
                sigma_proj = _project_argo_rows_to_sigma_for_overview(
                    pkg, proj_profile,
                ) if plot_argo_projection else pd.DataFrame()

                sigma_fig = _plot_glorys_overview_vertical_2x2_sigma(
                    vertical_package=sigma_pkg,
                    k_val=k_val,
                    b_val=b_val,
                    subject_label=subject_label,
                    date_label=date_str,
                    xmin=xmin,
                    xmax=xmax,
                    ymin=sigma_ymin,
                    ymax=sigma_ymax,
                    projected_argo_profile=sigma_proj if not sigma_proj.empty else None,
                    plot_isolines=plot_isolines,
                    isoline_levels=isoline_levels,
                    isoline_color=isoline_color,
                    isoline_linewidth=isoline_linewidth,
                    isoline_alpha=isoline_alpha,
                    label_isolines=label_isolines,
                    title_extra=sigma_title_extra,
                )

                if save_fig:
                    if '{date}' in save_name_prefix:
                        filename_core = save_name_prefix.format(date=date_tag)
                        sigma_filename = f"{filename_core}_sigma_k{k_val:.2f}b{b_val:+.2f}_{run_tag}.png"
                    else:
                        sigma_filename = (
                            f"{save_name_prefix}_sigma_{date_tag}_k{k_val:.2f}b{b_val:+.2f}_{run_tag}.png"
                        )
                    sigma_save_path = out_dir / sigma_filename
                    sigma_fig.savefig(sigma_save_path, dpi=300, bbox_inches='tight')
                    if verbose:
                        print(f"Sigma overview saved to: {sigma_save_path}")

                if show_fig:
                    plt.show()
                plt.close(sigma_fig)
            except (ValueError, RuntimeError, TypeError, KeyError):
                if verbose:
                    print(f"Warning: sigma overview failed for k={k_val}, b={b_val}:")
                    traceback.print_exc()

        ts_save_path = None
        if ts_diagram and argo_profile_rows is not None and not argo_profile_rows.empty:
            if save_fig:
                if '{date}' in save_name_prefix:
                    filename_core = save_name_prefix.format(date=date_tag)
                    ts_filename = f"{filename_core}_ts_k{k_val:.2f}b{b_val:+.2f}_{run_tag}.png"
                else:
                    ts_filename = f"{save_name_prefix}_ts_{date_tag}_k{k_val:.2f}b{b_val:+.2f}_{run_tag}.png"
                ts_save_path = out_dir / ts_filename
            try:
                pd_data = pkg.get('profile_data', {})
                theta_gl = pd_data.get('thetao')
                sal_gl = pd_data.get('salinity')
                if theta_gl is not None and sal_gl is not None:
                    theta_ma = np.ma.array(theta_gl, copy=False)
                    sal_ma = np.ma.array(sal_gl, copy=False)
                    valid_gl = ~np.ma.getmaskarray(theta_ma) & ~np.ma.getmaskarray(sal_ma)
                    bg_theta = np.asarray(theta_ma[valid_gl], dtype=float)
                    bg_sal = np.asarray(sal_ma[valid_gl], dtype=float)
                    center_lat = float(np.nanmean(np.asarray(pkg.get('lat_coords', [30.0]), dtype=float)))
                    center_lon = float(np.nanmean(np.asarray(pkg.get('lon_coords', [0.0]), dtype=float)))
                    _plot_ts_diagram_core(
                        bg_theta, bg_sal, argo_profile_rows,
                        center_lat=center_lat, center_lon=center_lon,
                        anomaly_depth=heave_projection_depth_m,
                        subject_label=subject_label, date_label=date_str,
                        bg_label='GLORYS',
                        show_fig=show_fig, save_fig=save_fig, save_path=ts_save_path,
                        annotate_spice=annotate_spice,
                    )
            except (ValueError, RuntimeError, TypeError, KeyError):
                if verbose:
                    print(f"Warning: T-S diagram failed for k={k_val}, b={b_val}:")
                    traceback.print_exc()

        result_item = {'k': k_val, 'b': b_val, 'save_path': str(save_path) if save_path else None}
        if sigma_save_path is not None:
            result_item['sigma_save_path'] = str(sigma_save_path)
        if ts_save_path is not None:
            result_item['ts_save_path'] = str(ts_save_path)
        if heave_diag:
            result_item.update(heave_diag)
        if annotate_spice:
            result_item['spice_anomaly'] = spice_anomaly_val
            result_item['spice_percentile'] = spice_percentile_val
        results.append(result_item)

    return results


def plot_track_vertical_glorys_overview(
    DS: list | str | tuple | dict,
    no: int,
    needed_date: str | pd.Timestamp,
    k: float | list[float] | None = None,
    b: float | list[float] | None = None,
    *,
    variables: list[str] | None = None,
    needed_depth: float | int = 0,
    xmin: float = -400.0,
    xmax: float = 400.0,
    ymin: float = 0.0,
    ymax: float = 1000.0,
    profile_spacing_km: float | None = None,
    interpolate_z: bool = True,
    profile_depth_spacing_m: float | None = None,
    plot_mlt: bool = False,
    plot_argo_projection: bool = True,
    argo_projection_config: DetectionConfig | None = None,
    argo_projection_min_depth: float | None = None,
    plot_isolines: bool = True,
    isoline_levels: int | list[float] | np.ndarray | None = None,
    isoline_color: str = 'black',
    isoline_linewidth: float = 0.8,
    isoline_alpha: float = 0.45,
    label_isolines: bool = True,
    show_fig: bool = True,
    save_fig: bool = False,
    output_dir: str | Path | None = None,
    verbose: bool = True,
    heave_projection_depth_m: float | None = None,
    heave_x_window_km: float = 25.0,
    heave_z_window_m: float | None = 100.0,
    heave_search_range: float = _heave_search_range,
    heave_depth_threshold: float = _heave_depth_threshold,
    heave_z_search_m: float | None = _heave_z_search_m,
    annotate_heave: bool = True,
    z_overview: bool = True,
    sigma_overview: bool = False,
    ts_diagram: bool = False,
) -> list[dict]:
    """绘制 track 场景 GLORYS vertical 2×2 总览图。

    沿涡旋剖面线生成 z 坐标、σ 坐标垂向总览与 T-S 图，并可叠加 Argo 投影与 heave/OI 诊断；k/b 为 None 时
    默认取纬向剖面线（k=0、b=center_lat），z_overview/sigma_overview/ts_diagram 独立控制三类输出。

    参数:
        - DS (list | str | tuple | dict): 轨迹数据输入，常见 'acs'/'acl'/'cs'/'cl' 或 legacy 列表结构。
        - no (int): 轨迹编号（track id）。
        - needed_date (str | pd.Timestamp): 目标日期。
        - k (float | list[float] | None): 剖面线斜率，y = kx + b；None 时取纬向线 k=0。
        - b (float | list[float] | None): 剖面线截距；None 时取 center_lat。
        - variables (list[str] | None): 需要绘制的变量列表；None 时用默认集。
        - needed_depth (float | int): 水平参考深度（m），默认 0（表层）。
        - xmin (float): 横向显示与采样下界（km），默认 -400.0。
        - xmax (float): 横向显示与采样上界（km），默认 400.0。
        - ymin (float): 纵向深度显示上界（m），默认 0.0。
        - ymax (float): 纵向深度显示下界（m），默认 1000.0。
        - profile_spacing_km (float | None): 水平采样步长（km）；None 时用配置默认。
        - interpolate_z (bool): 是否将深度轴重采样到等间距网格，默认 True。
        - profile_depth_spacing_m (float | None): 深度重采样步长（m）；None 时用配置默认。
        - plot_mlt (bool): 是否叠加混合层深度线，默认 False。
        - plot_argo_projection (bool): 是否叠加同日匹配 Argo 点投影层，默认 True。
        - argo_projection_config (DetectionConfig | None): 投影点异常筛选配置；None 时使用默认。
        - argo_projection_min_depth (float | None): 投影点最小深度阈值（m）；None 时回退配置。
        - plot_isolines (bool): 是否叠加变量等值线，默认 True。
        - isoline_levels (int | list[float] | np.ndarray | None): 等值线级别数或显式级别；None 时自动。
        - isoline_color (str): 等值线颜色，默认 'black'。
        - isoline_linewidth (float): 等值线线宽，默认 0.8。
        - isoline_alpha (float): 等值线透明度，默认 0.45。
        - label_isolines (bool): 是否标注等值线数值，默认 True。
        - show_fig (bool): 是否显示图像，默认 True。
        - save_fig (bool): 是否保存图像，默认 False。
        - output_dir (str | Path | None): 图片输出目录，None 时使用默认路径。
        - verbose (bool): 是否打印进度与保存提示，默认 True。
        - heave_projection_depth_m (float | None): heave 诊断的投影深度（m）；None 时自动。
        - heave_x_window_km (float): heave 诊断水平窗口半宽（km），默认 25.0。
        - heave_z_window_m (float | None): heave 诊断垂向窗口半宽（m），默认 100.0。
        - heave_search_range (float): 从 σ_argo 向上搜索的 σ 跨度（kg/m³），默认来自 processing.yml。
        - heave_depth_threshold (float): 通风判定深度（m），默认来自 processing.yml。
        - heave_z_search_m (float | None): 等密线连通性垂向范围（m），默认来自 processing.yml。
        - annotate_heave (bool): 是否在图上标注 heave/OI 诊断，默认 True。
        - z_overview (bool): 是否输出 z 坐标总览图，默认 True。
        - sigma_overview (bool): 是否输出 σ 坐标总览图，默认 False。
        - ts_diagram (bool): 是否输出 T-S 图，默认 False。

    返回:
        - list[dict]: 每条剖面线一个结果字典，含保存路径、heave/OI 诊断与 spice 异常等字段。
    """
    vertical_vars = _normalize_overview_vertical_variables(variables)
    k_list, b_list = _normalize_profile_lines(k, b)
    vars_to_fetch = set(vertical_vars)
    if annotate_heave:
        vars_to_fetch.add('sigma')
    if sigma_overview:
        vars_to_fetch.update(['pv', 'sigma'])
    if plot_mlt:
        vars_to_fetch.add('mlt')

    track_df, ds_name, ds_source_for_filter = _resolve_track_context(DS, no, include_contours=True)
    if track_df.empty:
        raise ValueError(f"Track {no} has no data.")

    dates = track_df['date'] if 'date' in track_df.columns else convert_date(track_df['time'])
    dates = pd.to_datetime(dates, errors='coerce')
    target_ts = pd.Timestamp(needed_date).normalize()
    same_day_idx = np.nonzero(dates.dt.normalize().to_numpy() == target_ts.to_datetime64())[0]
    if same_day_idx.size == 0:
        raise ValueError(f"Date {target_ts.strftime('%Y-%m-%d')} not found in track {no}.")
    needed_idx = int(same_day_idx[0])
    target_date = pd.Timestamp(dates.iloc[needed_idx])

    if not k_list and not b_list:
        center_lats = pd.to_numeric(track_df['center_lat'], errors='coerce')
        k_list, b_list = [0.0], [float(center_lats.iloc[needed_idx])]

    radius_arr = pd.to_numeric(track_df['radius'], errors='coerce').to_numpy(dtype=float)
    projection_distance_scale_km = None
    if 0 <= needed_idx < len(radius_arr) and np.isfinite(radius_arr[needed_idx]) and float(radius_arr[needed_idx]) > 0:
        projection_distance_scale_km = float(radius_arr[needed_idx]) / 1000.0

    vertical_packages = get_vertical_glorys(
        DS,
        no,
        target_date,
        k_list,
        b_list,
        variables=list(vars_to_fetch),
        x_min_km=xmin,
        x_max_km=xmax,
        profile_spacing_km=profile_spacing_km,
        interpolate_z=interpolate_z,
        profile_depth_spacing_m=profile_depth_spacing_m,
    )

    projected_argo_rows = pd.DataFrame()
    ts_argo_rows = pd.DataFrame()
    if plot_argo_projection or ts_diagram:
        argo_all = filtered_float_data(ds_source_for_filter, no, track=track_df)
        if not argo_all.empty:
            argo_all = argo_all.copy()
            argo_all['date'] = pd.to_datetime(argo_all[['Year', 'Month', 'Day']], errors='coerce').dt.normalize()
            same_day_rows = argo_all[argo_all['date'] == target_date.normalize()].copy()
            if ts_diagram:
                ts_argo_rows = same_day_rows.copy()
            if plot_argo_projection and not same_day_rows.empty:
                argo_projection_config = _resolve_detection_config(
                    argo_projection_config,
                    anomaly_min_depth=argo_projection_min_depth,
                )
                projected_argo_rows = _prepare_overview_projection_rows(
                    same_day_rows,
                    detection_config=argo_projection_config,
                )

    ds_name_upper = ds_name.upper() if isinstance(ds_name, str) else "UNKNOWN"

    return _run_vertical_overview_batch(
        vertical_packages=vertical_packages,
        k_list=k_list,
        b_list=b_list,
        vertical_vars=vertical_vars,
        target_date=target_date,
        subject_label=f"Track {ds_name_upper}{int(no)}",
        save_name_prefix=f"{ds_name_upper}{int(no)}",
        save_subdir='plot_track_vertical_glorys_overview',
        xmin=xmin,
        xmax=xmax,
        ymin=ymin,
        ymax=ymax,
        projected_argo_rows=projected_argo_rows,
        projection_distance_scale_km=projection_distance_scale_km,
        plot_argo_projection=plot_argo_projection,
        plot_mlt=plot_mlt,
        plot_isolines=plot_isolines,
        isoline_levels=isoline_levels,
        isoline_color=isoline_color,
        isoline_linewidth=isoline_linewidth,
        isoline_alpha=isoline_alpha,
        label_isolines=label_isolines,
        detection_config=argo_projection_config,
        show_fig=show_fig,
        save_fig=save_fig,
        output_dir=output_dir,
        verbose=verbose,
        heave_projection_depth_m=heave_projection_depth_m,
        heave_x_window_km=heave_x_window_km,
        heave_z_window_m=heave_z_window_m,
        heave_search_range=heave_search_range,
        heave_depth_threshold=heave_depth_threshold,
        heave_z_search_m=heave_z_search_m,
        annotate_heave=annotate_heave,
        z_overview=z_overview,
        sigma_overview=sigma_overview,
        ts_diagram=ts_diagram,
        argo_profile_rows=ts_argo_rows if ts_diagram else None,
    )


def plot_argo_vertical_glorys_overview(
    profile_number: int,
    profile_time: int | str | pd.Timestamp,
    k: float | list[float] | None = None,
    b: float | list[float] | None = None,
    *,
    platform_number: int | None = None,
    variables: list[str] | None = None,
    needed_depth: float | int = 0,
    xmin: float = -400.0,
    xmax: float = 400.0,
    ymin: float = 0.0,
    ymax: float = 1000.0,
    profile_spacing_km: float | None = None,
    interpolate_z: bool = True,
    profile_depth_spacing_m: float | None = None,
    plot_mlt: bool = False,
    plot_argo_projection: bool = True,
    argo_projection_config: DetectionConfig | None = None,
    argo_projection_min_depth: float | None = None,
    plot_isolines: bool = True,
    isoline_levels: int | list[float] | np.ndarray | None = None,
    isoline_color: str = 'black',
    isoline_linewidth: float = 0.8,
    isoline_alpha: float = 0.45,
    label_isolines: bool = True,
    argo_data_dir: str | Path | None = None,
    show_fig: bool = True,
    save_fig: bool = False,
    output_dir: str | Path | None = None,
    verbose: bool = True,
    heave_projection_depth_m: float | None = None,
    heave_x_window_km: float = 25.0,
    heave_z_window_m: float | None = 100.0,
    heave_search_range: float = _heave_search_range,
    heave_depth_threshold: float = _heave_depth_threshold,
    heave_z_search_m: float | None = _heave_z_search_m,
    annotate_heave: bool = True,
    annotate_spice: bool = True,
    anomaly_sal: float | None = None,
    anomaly_theta: float | None = None,
    z_overview: bool = True,
    sigma_overview: bool = False,
    ts_diagram: bool = False,
) -> list[dict]:
    """绘制 Argo 场景 GLORYS vertical 2×2 总览图。

    以单个 Argo 剖面为中心沿剖面线生成 z 坐标、σ 坐标垂向总览与 T-S 图，并可叠加 Argo 投影、heave/OI 与
    spiciness 诊断；k/b 为 None 时默认取纬向剖面线（k=0、b=center_lat），z_overview/sigma_overview/
    ts_diagram 独立控制三类输出。

    参数:
        - profile_number (int): 目标 Argo 剖面编号。
        - profile_time (int | str | pd.Timestamp): 剖面日期（int YYYYMMDD / 'YYYY-MM-DD' / Timestamp）。
        - k (float | list[float] | None): 剖面线斜率，y = kx + b；None 时取纬向线 k=0。
        - b (float | list[float] | None): 剖面线截距；None 时取 center_lat。
        - platform_number (int | None): 浮标平台编号，辅助定位（可选）。
        - variables (list[str] | None): 需要绘制的变量列表；None 时用默认集。
        - needed_depth (float | int): 水平参考深度（m），默认 0（表层）。
        - xmin (float): 横向显示与采样下界（km），默认 -400.0。
        - xmax (float): 横向显示与采样上界（km），默认 400.0。
        - ymin (float): 纵向深度显示上界（m），默认 0.0。
        - ymax (float): 纵向深度显示下界（m），默认 1000.0。
        - profile_spacing_km (float | None): 水平采样步长（km）；None 时用配置默认。
        - interpolate_z (bool): 是否将深度轴重采样到等间距网格，默认 True。
        - profile_depth_spacing_m (float | None): 深度重采样步长（m）；None 时用配置默认。
        - plot_mlt (bool): 是否叠加混合层深度线，默认 False。
        - plot_argo_projection (bool): 是否叠加同日匹配 Argo 点投影层，默认 True。
        - argo_projection_config (DetectionConfig | None): 投影点异常筛选配置；None 时使用默认。
        - argo_projection_min_depth (float | None): 投影点最小深度阈值（m）；None 时回退配置。
        - plot_isolines (bool): 是否叠加变量等值线，默认 True。
        - isoline_levels (int | list[float] | np.ndarray | None): 等值线级别数或显式级别；None 时自动。
        - isoline_color (str): 等值线颜色，默认 'black'。
        - isoline_linewidth (float): 等值线线宽，默认 0.8。
        - isoline_alpha (float): 等值线透明度，默认 0.45。
        - label_isolines (bool): 是否标注等值线数值，默认 True。
        - argo_data_dir (str | Path | None): Argo 年度 parquet 目录；None 时使用配置默认目录。
        - show_fig (bool): 是否显示图像，默认 True。
        - save_fig (bool): 是否保存图像，默认 False。
        - output_dir (str | Path | None): 图片输出目录，None 时使用默认路径。
        - verbose (bool): 是否打印进度与保存提示，默认 True。
        - heave_projection_depth_m (float | None): heave 诊断的投影深度（m）；None 时自动。
        - heave_x_window_km (float): heave 诊断水平窗口半宽（km），默认 25.0。
        - heave_z_window_m (float | None): heave 诊断垂向窗口半宽（m），默认 100.0。
        - heave_search_range (float): 从 σ_argo 向上搜索的 σ 跨度（kg/m³），默认来自 processing.yml。
        - heave_depth_threshold (float): 通风判定深度（m），默认来自 processing.yml。
        - heave_z_search_m (float | None): 等密线连通性垂向范围（m），默认来自 processing.yml。
        - annotate_heave (bool): 是否在图上标注 heave/OI 诊断，默认 True。
        - annotate_spice (bool): 是否在标题标注 spiciness 异常 δπ，默认 True。
        - anomaly_sal (float | None): 异常点盐度，用于 spiciness 诊断；None 时自动取。
        - anomaly_theta (float | None): 异常点位温，用于 spiciness 诊断；None 时自动取。
        - z_overview (bool): 是否输出 z 坐标总览图，默认 True。
        - sigma_overview (bool): 是否输出 σ 坐标总览图，默认 False。
        - ts_diagram (bool): 是否输出 T-S 图，默认 False。

    返回:
        - list[dict]: 每条剖面线一个结果字典，含保存路径、heave/OI 诊断与 spice 异常等字段。
    """
    vertical_vars = _normalize_overview_vertical_variables(variables)
    k_list, b_list = _normalize_profile_lines(k, b)
    vars_to_fetch = set(vertical_vars)
    if annotate_heave:
        vars_to_fetch.add('sigma')
    if sigma_overview:
        vars_to_fetch.update(['pv', 'sigma'])
    if plot_mlt:
        vars_to_fetch.add('mlt')

    info = _resolve_argo_profile_center(
        profile_number=int(profile_number),
        profile_time=profile_time,
        platform_number=platform_number,
        argo_data_dir=argo_data_dir,
    )
    center_lon = float(info['center_lon'])
    center_lat = float(info['center_lat'])
    target_date = pd.Timestamp(info['target_date'])

    if not k_list and not b_list:
        k_list, b_list = [0.0], [center_lat]

    window_half_size_km = max(abs(float(xmin)), abs(float(xmax)))

    vertical_packages = get_vertical_glorys_from_center(
        center_lon=center_lon,
        center_lat=center_lat,
        needed_date=target_date,
        k=k_list,
        b=b_list,
        variables=list(vars_to_fetch),
        x_min_km=xmin,
        x_max_km=xmax,
        profile_spacing_km=profile_spacing_km,
        interpolate_z=interpolate_z,
        profile_depth_spacing_m=profile_depth_spacing_m,
        window_half_size_km=window_half_size_km,
        profile_id=int(profile_number),
        ds_name='ARGO',
    )

    projected_argo_rows = pd.DataFrame()
    if plot_argo_projection:
        argo_projection_config = _resolve_detection_config(
            argo_projection_config,
            anomaly_min_depth=argo_projection_min_depth,
        )

        df_year = info['year_df']
        day_ts = pd.to_datetime(df_year[['Year', 'Month', 'Day']], errors='coerce').dt.normalize()
        day_rows = df_year.loc[day_ts == target_date.normalize()].copy()
        if not day_rows.empty:
            day_rows['Longitude'] = pd.to_numeric(day_rows['Longitude'], errors='coerce')
            day_rows['Latitude'] = pd.to_numeric(day_rows['Latitude'], errors='coerce')
            day_rows = day_rows.dropna(subset=['Longitude', 'Latitude'])

            lon_min_local, lon_max_local, lat_min_local, lat_max_local = _window_bounds_from_center_km(
                center_lon,
                center_lat,
                float(window_half_size_km),
            )
            day_lon_local = center_lon + _minimal_lon_diff_deg(day_rows['Longitude'].to_numpy(dtype=float), center_lon)
            mask_window = (
                (day_lon_local >= lon_min_local)
                & (day_lon_local <= lon_max_local)
                & (day_rows['Latitude'].to_numpy(dtype=float) >= lat_min_local)
                & (day_rows['Latitude'].to_numpy(dtype=float) <= lat_max_local)
            )
            day_window = day_rows.loc[mask_window].copy()
            if not day_window.empty:
                projected_argo_rows = _prepare_overview_projection_rows(
                    day_window,
                    detection_config=argo_projection_config,
                )

    return _run_vertical_overview_batch(
        vertical_packages=vertical_packages,
        k_list=k_list,
        b_list=b_list,
        vertical_vars=vertical_vars,
        target_date=target_date,
        subject_label=f"Profile {int(profile_number)}",
        save_name_prefix=f"Argo_{{date}}_P{int(profile_number)}",
        save_subdir='plot_argo_vertical_glorys_overview',
        xmin=xmin,
        xmax=xmax,
        ymin=ymin,
        ymax=ymax,
        projected_argo_rows=projected_argo_rows,
        projection_distance_scale_km=float(window_half_size_km),
        plot_argo_projection=plot_argo_projection,
        plot_mlt=plot_mlt,
        plot_isolines=plot_isolines,
        isoline_levels=isoline_levels,
        isoline_color=isoline_color,
        isoline_linewidth=isoline_linewidth,
        isoline_alpha=isoline_alpha,
        label_isolines=label_isolines,
        detection_config=argo_projection_config,
        show_fig=show_fig,
        save_fig=save_fig,
        output_dir=output_dir,
        verbose=verbose,
        heave_projection_depth_m=heave_projection_depth_m,
        heave_x_window_km=heave_x_window_km,
        heave_z_window_m=heave_z_window_m,
        heave_search_range=heave_search_range,
        heave_depth_threshold=heave_depth_threshold,
        heave_z_search_m=heave_z_search_m,
        annotate_heave=annotate_heave,
        z_overview=z_overview,
        sigma_overview=sigma_overview,
        ts_diagram=ts_diagram,
        argo_profile_rows=info.get('profile_rows') if ts_diagram else None,
        annotate_spice=annotate_spice,
        anomaly_sal=anomaly_sal,
        anomaly_theta=anomaly_theta,
    )


def collect_argo_pool(
    lon_range: tuple[float, float],
    lat_range: tuple[float, float],
    start_date: str | pd.Timestamp,
    end_date: str | pd.Timestamp,
    *,
    max_depth: float = 2000.0,
) -> pd.DataFrame:
    """收集指定经纬度范围与时间窗口内的 Argo 剖面池。

    仅保留 T/S 均有效的深度行，不处理日界线穿越。

    参数:
        - lon_range (tuple[float, float]): (lon_min, lon_max) 经度范围，不支持跨日界线。
        - lat_range (tuple[float, float]): (lat_min, lat_max) 纬度范围。
        - start_date (str | pd.Timestamp): 时间窗口起始（含端点）。
        - end_date (str | pd.Timestamp): 时间窗口结束（含端点）。
        - max_depth (float): 最大采样深度（m），默认 2000.0。

    返回:
        - pd.DataFrame: 含剖面 T/S/Depth/位置/时间列；无匹配数据时返回空 DataFrame。
    """
    lon_min, lon_max = float(lon_range[0]), float(lon_range[1])
    lat_min, lat_max = float(lat_range[0]), float(lat_range[1])
    t_min, t_max = pd.Timestamp(start_date), pd.Timestamp(end_date)

    needed_cols = [
        'Year', 'Month', 'Day', 'Longitude', 'Latitude', 'Depth',
        'Temp_Adjusted', 'PSAL_Adjusted', 'Profile_number', 'Platform_number',
    ]
    dfs = []
    for yr in range(t_min.year, t_max.year + 1):
        fpath = Path(argo_path) / f'Argo{yr}.parquet'
        if not fpath.exists():
            continue
        df = pd.read_parquet(fpath, columns=needed_cols)
        mask = (
            df['Latitude'].between(lat_min, lat_max)
            & df['Longitude'].between(lon_min, lon_max)
            & (df['Depth'] <= max_depth)
        )
        dfs.append(df[mask])

    if not dfs:
        return pd.DataFrame()

    pool = pd.concat(dfs, ignore_index=True)
    pool['date'] = pd.to_datetime(
        {'year': pool['Year'], 'month': pool['Month'], 'day': pool['Day']}
    )
    pool = pool[(pool['date'] >= t_min) & (pool['date'] <= t_max)]
    pool = pool.dropna(subset=['Temp_Adjusted', 'PSAL_Adjusted'])
    return pool.reset_index(drop=True)


def _build_argo_3d_depth_chunk(args: tuple) -> tuple[int, np.ndarray, np.ndarray, np.ndarray]:
    """Worker for _build_argo_3d_field: processes a contiguous slice of depth levels.

    Vectorizes the lat dimension in chunks to eliminate the inner Python loop while
    keeping per-chunk memory bounded to ~50 MB.
    """
    (iz_start, z_chunk,
     p_lons, p_lats, p_depths, p_temps, p_sals,
     lon_grid, lat_grid,
     inv2_h, inv2_z, h_cut, z_cut, min_weight, lat_chunk_size) = args

    nz_c = len(z_chunk)
    nlat = len(lat_grid)
    nlon = len(lon_grid)
    T_c = np.full((nz_c, nlat, nlon), np.nan)
    S_c = np.full((nz_c, nlat, nlon), np.nan)
    W_c = np.zeros((nz_c, nlat, nlon))

    for ic, zc in enumerate(z_chunk):
        depth_mask = np.abs(p_depths - zc) < z_cut
        if not depth_mask.any():
            continue
        d_lons = p_lons[depth_mask]
        d_lats = p_lats[depth_mask]
        d_temps = p_temps[depth_mask]
        d_sals = p_sals[depth_mask]
        w_depth = np.exp(inv2_z * (p_depths[depth_mask] - zc) ** 2)

        for lat_start in range(0, nlat, lat_chunk_size):
            lat_end = min(lat_start + lat_chunk_size, nlat)
            lat_c = lat_grid[lat_start:lat_end]  # (nc,)
            cos_lats = np.cos(np.radians(lat_c))

            lat_lo = lat_c[0] - h_cut / 111.32
            lat_hi = lat_c[-1] + h_cut / 111.32
            f_mask = (d_lats >= lat_lo) & (d_lats <= lat_hi)
            if not f_mask.any():
                continue

            d_lons_f = d_lons[f_mask]
            d_lats_f = d_lats[f_mask]
            d_temps_f = d_temps[f_mask]
            d_sals_f = d_sals[f_mask]
            w_depth_f = w_depth[f_mask]

            dy_km = (lat_c[:, None] - d_lats_f[None, :]) * 111.32          # (nc, n_f)
            dx_km = (lon_grid[None, :, None] - d_lons_f[None, None, :]) * 111.32 * cos_lats[:, None, None]
            r2 = dx_km ** 2 + dy_km[:, None, :] ** 2                        # (nc, nlon, n_f)
            w = np.exp(inv2_h * r2) * w_depth_f[None, None, :]              # (nc, nlon, n_f)
            w_sum = w.sum(axis=2)                                            # (nc, nlon)
            valid = w_sum >= min_weight
            if not valid.any():
                continue

            T_num = (w * d_temps_f[None, None, :]).sum(axis=2)
            S_num = (w * d_sals_f[None, None, :]).sum(axis=2)
            T_c[ic, lat_start:lat_end] = np.where(valid, T_num / w_sum, np.nan)
            S_c[ic, lat_start:lat_end] = np.where(valid, S_num / w_sum, np.nan)
            W_c[ic, lat_start:lat_end] = np.where(valid, w_sum, 0.0)

    return iz_start, T_c, S_c, W_c


def _build_argo_3d_field(
    pool: pd.DataFrame,
    lon_range: tuple[float, float],
    lat_range: tuple[float, float],
    *,
    h_bw: float = _argo_recon_h_bw_km,
    depth_bw: float = _argo_recon_depth_bw_m,
    h_spacing_deg: float = _argo_recon_h_spacing_deg,
    z_max_m: float = _argo_recon_z_max_m,
    z_spacing_m: float = _argo_recon_z_spacing_m,
    min_weight: float = _argo_recon_min_weight,
    n_jobs: int | None = None,
    verbose: bool = True,
) -> dict:
    """用 Argo 剖面池三维高斯核重建 (lon × lat × depth) 网格场。

    核函数：w = exp(−0.5·(r²/h_bw² + dz²/depth_bw²))，截断半径 3σ。
    权重低于 min_weight 的格点保持 NaN，不做外推填充。
    深度维度按 n_jobs 分块，各块通过 multiprocessing.Pool 并行。

    参数:
        pool (DataFrame): collect_argo_pool 返回的剖面数据。
        lon_range / lat_range (tuple): 重建区域经纬度范围（°）。
        h_bw (float): 水平高斯核带宽（km），默认 60 km。
        depth_bw (float): 垂向高斯核带宽（m），默认 25 m。
        h_spacing_deg (float): 水平网格间距（°），默认 0.1°。
        z_max_m (float): 最大重建深度（m），默认 1500 m。
        z_spacing_m (float): 垂向网格间距（m），默认 10 m。
        min_weight (float): 最小累积权重阈值，低于此值格点标为 NaN，默认 3.0。
        n_jobs (int | None): 并行进程数，None 时取 min(cpu_count, 8)。
        verbose (bool): 是否打印进度信息。

    返回:
        dict，键包括 thetao / salinity / sigma0 / weight（均为 (nz, nlat, nlon) ndarray）
        及 lon / lat / depth 坐标数组和 attrs 元数据字典。
    """
    lon_min, lon_max = float(lon_range[0]), float(lon_range[1])
    lat_min, lat_max = float(lat_range[0]), float(lat_range[1])

    lon_grid = np.arange(lon_min, lon_max + h_spacing_deg / 2.0, h_spacing_deg)
    lat_grid = np.arange(lat_min, lat_max + h_spacing_deg / 2.0, h_spacing_deg)
    z_grid = np.arange(0.0, z_max_m + z_spacing_m / 2.0, z_spacing_m)
    nlon, nlat, nz = len(lon_grid), len(lat_grid), len(z_grid)

    _attrs = {
        'h_bw_km': h_bw, 'depth_bw_m': depth_bw,
        'h_spacing_deg': h_spacing_deg, 'z_spacing_m': z_spacing_m,
        'min_weight': min_weight,
    }
    if pool.empty:
        return {
            'thetao': np.full((nz, nlat, nlon), np.nan),
            'salinity': np.full((nz, nlat, nlon), np.nan),
            'sigma0': np.full((nz, nlat, nlon), np.nan),
            'weight': np.zeros((nz, nlat, nlon)),
            'lon': lon_grid, 'lat': lat_grid, 'depth': z_grid, 'attrs': _attrs,
        }

    p_lons = pool['Longitude'].values
    p_lats = pool['Latitude'].values
    p_depths = pool['Depth'].values
    p_temps = pool['Temp_Adjusted'].values
    p_sals = pool['PSAL_Adjusted'].values

    inv2_h = -0.5 / h_bw ** 2
    inv2_z = -0.5 / depth_bw ** 2
    h_cut = 3.0 * h_bw
    z_cut = 3.0 * depth_bw

    lat_chunk_size = max(1, nlat // 8)  # ~8 chunks per depth level

    _default_jobs = min(os.cpu_count() or 1, 8)
    n_workers = min(max(1, n_jobs if n_jobs is not None else _default_jobs), nz)
    z_chunks = np.array_split(z_grid, n_workers)
    iz_starts = [int(sum(len(c) for c in z_chunks[:i])) for i in range(n_workers)]

    shared = (p_lons, p_lats, p_depths, p_temps, p_sals,
              lon_grid, lat_grid, inv2_h, inv2_z, h_cut, z_cut, min_weight, lat_chunk_size)
    task_args = [(iz_s, zc) + shared for iz_s, zc in zip(iz_starts, z_chunks)]

    if n_workers == 1:
        results = [_build_argo_3d_depth_chunk(task_args[0])]
    else:
        with multiprocessing.Pool(processes=n_workers) as mp_pool:
            results = mp_pool.map(_build_argo_3d_depth_chunk, task_args)

    T_out = np.full((nz, nlat, nlon), np.nan)
    S_out = np.full((nz, nlat, nlon), np.nan)
    W_out = np.zeros((nz, nlat, nlon))
    for iz_s, T_c, S_c, W_c in results:
        iz_e = iz_s + len(T_c)
        T_out[iz_s:iz_e] = T_c
        S_out[iz_s:iz_e] = S_c
        W_out[iz_s:iz_e] = W_c

    center_lat = float(lat_grid.mean())
    center_lon = float(lon_grid.mean())
    p_from_z = gsw.p_from_z(-z_grid, center_lat)
    p_3d = p_from_z[:, None, None]
    valid3d = ~np.isnan(T_out)
    sigma_out = np.full_like(T_out, np.nan)
    if valid3d.any():
        SA = np.where(valid3d, gsw.SA_from_SP(S_out, np.broadcast_to(p_3d, T_out.shape), center_lon, center_lat), np.nan)
        CT = np.where(valid3d, gsw.CT_from_t(SA, T_out, np.broadcast_to(p_3d, T_out.shape)), np.nan)
        sigma_out = np.where(valid3d, gsw.sigma0(SA, CT), np.nan)

    if verbose:
        n_valid = int(valid3d.sum())
        n_total = nz * nlat * nlon
        # 顶层 1000m 覆盖率：避开深层（多数 Argo 浮标到不了）拉低的部分，更贴近常见绘图深度
        z_shallow = z_grid <= 1000.0
        n_valid_top = int(valid3d[z_shallow].sum())
        n_total_top = int(z_shallow.sum()) * nlat * nlon
        print(
            f'  3D field: ({nz}, {nlat}, {nlon}), valid={n_valid}/{n_total} '
            f'({100 * n_valid / n_total:.1f}% full, {100 * n_valid_top / n_total_top:.1f}% ≤1000m)'
        )

    return {
        'thetao': T_out, 'salinity': S_out, 'sigma0': sigma_out, 'weight': W_out,
        'lon': lon_grid, 'lat': lat_grid, 'depth': z_grid, 'attrs': _attrs,
    }


def _estimate_argo_top_coverage(
    pool: pd.DataFrame,
    lon_range: tuple[float, float],
    lat_range: tuple[float, float],
    *,
    h_bw: float = _argo_recon_h_bw_km,
    depth_bw: float = _argo_recon_depth_bw_m,
    h_spacing_deg: float = _argo_recon_h_spacing_deg,
    min_weight: float = _argo_recon_min_weight,
    z_top_m: float = 1000.0,
    probe_spacing_m: float = _argo_recon_coverage_probe_spacing_m,
    n_jobs: int | None = 1,
) -> float:
    """粗深度探针预估 ≤z_top_m 的重建覆盖率（批处理预筛用，避免全量 build）。

    复用 _build_argo_3d_field，但只在 0..z_top_m 间按 probe_spacing_m 取稀疏深度层
    （默认 100 m，约为正式 10 m 网格的 1/10 成本），返回有效格点占比。Argo 在
    ≤1000 m 垂向连续密集，稀疏深度采样即可忠实反映真实 ≤1000m 覆盖率。
    """
    if pool.empty:
        return 0.0
    field = _build_argo_3d_field(
        pool, lon_range, lat_range,
        h_bw=h_bw, depth_bw=depth_bw, h_spacing_deg=h_spacing_deg,
        z_max_m=z_top_m, z_spacing_m=probe_spacing_m, min_weight=min_weight,
        n_jobs=n_jobs, verbose=False,
    )
    valid = ~np.isnan(field['thetao'])
    if valid.size == 0:
        return 0.0
    return float(valid.mean())


def _save_argo_3d_field(field: dict, zarr_path: str | Path) -> None:
    """将 _build_argo_3d_field 返回的 3D 场保存为 zarr 格式。"""
    root = zarr.open_group(str(zarr_path), mode='w')
    for key in ('thetao', 'salinity', 'sigma0', 'weight', 'lon', 'lat', 'depth'):
        root[key] = field[key]
    root.attrs.update(field.get('attrs', {}))


def load_argo_3d_field(zarr_path: str | Path) -> dict:
    """加载 zarr 格式的 3D Argo 重建场。

    参数:
        - zarr_path (str | Path): 3D 场的 zarr 路径。

    返回:
        - dict: 含 thetao/salinity/sigma0/weight/lon/lat/depth 数组及 attrs 元信息的字典。
    """
    root = zarr.open_group(str(zarr_path), mode='r')
    return {
        **{key: root[key][:] for key in ('thetao', 'salinity', 'sigma0', 'weight', 'lon', 'lat', 'depth')},
        'attrs': dict(root.attrs),
    }


def slice_section_from_argo_field(
    field: dict,
    k: float,
    center_lon: float,
    center_lat: float,
    *,
    x_min_km: float | None = None,
    x_max_km: float | None = None,
    x_spacing_km: float = _argo_recon_x_spacing_km,
) -> list[dict]:
    """从 3D Argo 场沿测线切取 2D 垂向断面。

    用 RegularGridInterpolator 双线性插值，返回与 get_vertical_glorys 格式兼容的 list[dict]。

    参数:
        - field (dict): _build_argo_3d_field / load_argo_3d_field 返回的 3D 场字典。
        - k (float): 测线斜率 Δlat/Δlon；0.0 为纯纬向断面。
        - center_lon (float): 测线 x=0 的参考经度。
        - center_lat (float): 测线 x=0 的参考纬度。
        - x_min_km (float | None): 断面 x 轴下界（km）；None 时从 field 经度范围自动推导以覆盖完整宽度。
        - x_max_km (float | None): 断面 x 轴上界（km）；None 时自动推导。
        - x_spacing_km (float): 断面水平采样间距（km），默认 5。

    返回:
        - list[dict]: 与 get_vertical_glorys 兼容的断面结果，供 _plot_glorys_overview_vertical_2x2 使用。
    """
    cos_lat = np.cos(np.radians(center_lat))
    if x_min_km is None or x_max_km is None:
        _half_km = (float(field['lon'][-1]) - float(field['lon'][0])) / 2.0 * 111.32 * cos_lat
        x_min_km = x_min_km if x_min_km is not None else -_half_km
        x_max_km = x_max_km if x_max_km is not None else _half_km
    slope_km = k / cos_lat if cos_lat > 1e-6 else k
    dir_norm = np.sqrt(1.0 + slope_km ** 2)
    ux, uy = 1.0 / dir_norm, slope_km / dir_norm

    x_grid = np.arange(x_min_km, x_max_km + x_spacing_km / 2.0, x_spacing_km)
    lon_sec = center_lon + x_grid * ux / (111.32 * cos_lat if cos_lat > 1e-6 else 111.32)
    lat_sec = center_lat + x_grid * uy / 111.32
    nx = len(x_grid)

    z_grid = field['depth']
    lat_grid = field['lat']
    lon_grid = field['lon']
    nz = len(z_grid)

    z_pts = np.repeat(z_grid, nx)
    lat_pts = np.tile(lat_sec, nz)
    lon_pts = np.tile(lon_sec, nz)
    query = np.column_stack([z_pts, lat_pts, lon_pts])

    def _interp(data3d):
        rgi = RegularGridInterpolator(
            (z_grid, lat_grid, lon_grid), data3d,
            method='linear', bounds_error=False, fill_value=np.nan,
        )
        return rgi(query).reshape(nz, nx)

    T_2d = _interp(field['thetao'])
    S_2d = _interp(field['salinity'])
    sigma_2d = _interp(field['sigma0'])
    W_2d = _interp(field['weight'])

    return [{
        'profile_data': {
            'thetao': np.ma.masked_invalid(T_2d),
            'salinity': np.ma.masked_invalid(S_2d),
            'sigma': np.ma.masked_invalid(sigma_2d),
            'argo_weight': W_2d,
        },
        'y_coords': x_grid,
        'z_coords': z_grid,
        'lon_coords': lon_sec,
        'lat_coords': lat_sec,
        'projections': {},
        'metadata': {'draw_reference_lines': False, 'source': 'argo_reconstruction'},
    }]


def plot_regional_vertical_argo_overview(
    lon_range: tuple[float, float],
    lat_range: tuple[float, float],
    start_date: str | pd.Timestamp | None = None,
    end_date: str | pd.Timestamp | None = None,
    *,
    k: float = 0.0,
    center_lon: float | None = None,
    center_lat: float | None = None,
    h_bw: float = _argo_recon_h_bw_km,
    depth_bw: float = _argo_recon_depth_bw_m,
    h_spacing_deg: float = _argo_recon_h_spacing_deg,
    z_max_m: float = _argo_recon_z_max_m,
    z_spacing_m: float = _argo_recon_z_spacing_m,
    min_weight: float = _argo_recon_min_weight,
    x_min_km: float | None = None,
    x_max_km: float | None = None,
    x_spacing_km: float = _argo_recon_x_spacing_km,
    ymin: float = 0.0,
    ymax: float = 1000.0,
    plot_isolines: bool = True,
    isoline_levels: int | list[float] | np.ndarray | None = None,
    isoline_color: str = 'black',
    isoline_linewidth: float = 0.8,
    isoline_alpha: float = 0.45,
    label_isolines: bool = True,
    field: dict | None = None,
    save_field: bool = False,
    field_path: str | Path | None = None,
    n_jobs: int | None = None,
    show_fig: bool = True,
    save_fig: bool = False,
    output_dir: str | Path | None = None,
    verbose: bool = True,
) -> None:
    """用 Argo 数据重建并绘制区域垂向断面 2x2 总览图（σ / θ / S / 覆盖权重）。

    先建立 3D 高斯核重建场（_build_argo_3d_field），再用
    slice_section_from_argo_field 切取 2D 断面绘图。3D 场可选以 zarr 保存
    供后续直接切片，避免每次重建。n_jobs>1 时重建阶段按深度并行。

    参数:
        - lon_range (tuple[float, float]): 区域经度范围（°）。
        - lat_range (tuple[float, float]): 区域纬度范围（°）。
        - start_date (str | pd.Timestamp | None): 时间窗口起始（含端点），与 field 互斥；field 为 None 时必填。
        - end_date (str | pd.Timestamp | None): 时间窗口结束（含端点），与 field 互斥；field 为 None 时必填。
        - k (float): 断面测线斜率 Δlat/Δlon，0.0 为纬向，默认 0.0。
        - center_lon (float | None): 测线参考经度，None 时取区域经度中心。
        - center_lat (float | None): 测线参考纬度，None 时取区域纬度中心。
        - h_bw (float): 水平高斯核带宽（km），默认来自 processing.yml:argo_reconstruction。
        - depth_bw (float): 垂向高斯核带宽（m），默认来自配置。
        - h_spacing_deg (float): 重建网格水平间距（°），默认来自配置。
        - z_max_m (float): 最大重建深度（m），默认来自配置。
        - z_spacing_m (float): 垂向网格间距（m），默认来自配置。
        - min_weight (float): 最小累积权重阈值，低于此值格点显示为 NaN，默认来自配置。
        - x_min_km (float | None): 断面 x 轴下界（km），None 时自动取区域全宽。
        - x_max_km (float | None): 断面 x 轴上界（km），None 时自动取区域全宽。
        - x_spacing_km (float): 断面水平采样间距（km），默认来自配置。
        - ymin (float): 图纵轴深度上界（m），默认 0.0。
        - ymax (float): 图纵轴深度下界（m），默认 1000.0；3D 场仍建到 z_max_m。
        - plot_isolines (bool): 是否叠加 σ₀ 等值线，默认 True。
        - isoline_levels (int | list[float] | np.ndarray | None): 等值线级别数或显式级别；None 时自动。
        - isoline_color (str): 等值线颜色，默认 'black'。
        - isoline_linewidth (float): 等值线线宽，默认 0.8。
        - isoline_alpha (float): 等值线透明度，默认 0.45。
        - label_isolines (bool): 是否标注等值线数值，默认 True。
        - field (dict | None): 预建 3D 场，传入后跳过 collect_argo_pool + _build_argo_3d_field，与 start_date/end_date 互斥，日期从 field['attrs'] 读取。
        - save_field (bool): 是否将 3D 场保存为 zarr，默认 False。
        - field_path (str | Path | None): zarr 保存路径，None 时按区域和时间窗口自动生成。
        - n_jobs (int | None): _build_argo_3d_field 并行进程数，None 时取 min(cpu_count, 8)。
        - show_fig (bool): 是否显示图片，默认 True。
        - save_fig (bool): 是否保存图片，默认 False。
        - output_dir (str | Path | None): 图片输出目录，None 时使用默认路径。
        - verbose (bool): 是否打印进度信息，默认 True。
    输出:
        - 断面图（save_fig=True，或指定 output_dir 时）：`plot_outputs/shared/<region>/plot_regional_vertical_argo_overview/argo_regional_{经度范围}E_{纬度范围}N_{起止日期}_hbw{带宽}km.png`
        - 3D 重建场（save_field=True 时）：zarr，路径为 `field_path` 或按区域与时间窗口自动命名。
    """
    if field is not None and (start_date is not None or end_date is not None):
        raise ValueError(
            "Provide either field= (pre-built 3D field) or start_date/end_date — not both."
        )
    if field is None and (start_date is None or end_date is None):
        raise ValueError("start_date and end_date are required when field= is not provided.")

    lon_min, lon_max = float(lon_range[0]), float(lon_range[1])
    lat_min, lat_max = float(lat_range[0]), float(lat_range[1])

    if field is None:
        t_min, t_max = pd.Timestamp(start_date), pd.Timestamp(end_date)
    else:
        _s = field['attrs'].get('start_date')
        _e = field['attrs'].get('end_date')
        t_min = pd.Timestamp(_s) if _s else None
        t_max = pd.Timestamp(_e) if _e else None

    clon = center_lon if center_lon is not None else (lon_min + lon_max) / 2.0
    clat = center_lat if center_lat is not None else (lat_min + lat_max) / 2.0
    b = clat - k * clon

    cos_lat_c = np.cos(np.radians(clat))
    half_lon_km = (lon_max - lon_min) / 2.0 * 111.32 * cos_lat_c
    xmin = x_min_km if x_min_km is not None else -half_lon_km
    xmax = x_max_km if x_max_km is not None else half_lon_km
    z_plot_max = ymax

    if field is None:
        if verbose:
            print(
                f'[Argo3D] ({lon_min:.1f},{lon_max:.1f}) / ({lat_min:.1f},{lat_max:.1f})  '
                f'{t_min.date()} – {t_max.date()}'
            )
        pool = collect_argo_pool(
            lon_range, lat_range, t_min, t_max,
            max_depth=z_max_m + 200.0,
        )
        if pool.empty:
            print('No Argo data found.')
            return
        n_prof = pool['Profile_number'].nunique()
        if verbose:
            print(f'  {n_prof} profiles, {len(pool)} depth rows')
        field = _build_argo_3d_field(
            pool, lon_range, lat_range,
            h_bw=h_bw, depth_bw=depth_bw, h_spacing_deg=h_spacing_deg,
            z_max_m=z_max_m, z_spacing_m=z_spacing_m, min_weight=min_weight,
            n_jobs=n_jobs, verbose=verbose,
        )
        field['attrs']['start_date'] = t_min.strftime('%Y-%m-%d')
        field['attrs']['end_date'] = t_max.strftime('%Y-%m-%d')
    else:
        n_prof = None

    if save_field:
        fp = Path(field_path) if field_path else (
            _shared_output_dir('argo_3d_fields')
            / (
                f'argo3d_{lon_min:.1f}_{lon_max:.1f}_{lat_min:.1f}_{lat_max:.1f}'
                f'_{t_min.strftime("%Y%m%d") if t_min else "unknown"}'
                f'_{t_max.strftime("%Y%m%d") if t_max else "unknown"}'
                f'_hbw{h_bw:.0f}_dbw{depth_bw:.0f}.zarr'
            )
        )
        fp.parent.mkdir(parents=True, exist_ok=True)
        _save_argo_3d_field(field, fp)
        if verbose:
            print(f'  Saved 3D field: {fp}')

    vertical_packages = slice_section_from_argo_field(
        field, k, clon, clat,
        x_min_km=xmin, x_max_km=xmax, x_spacing_km=x_spacing_km,
    )
    if not vertical_packages:
        return

    if t_min and t_max:
        date_label = f'{t_min.strftime("%Y-%m-%d")} – {t_max.strftime("%Y-%m-%d")}'
    else:
        date_label = 'unknown'
    subject_label = f'({lon_min:.1f}–{lon_max:.1f}°E, {lat_min:.1f}–{lat_max:.1f}°N)'
    n_label = f'  N={n_prof}' if n_prof is not None else ''
    title_extra = f'  h_bw={h_bw:.0f}km{n_label}'

    for vp in vertical_packages:
        _plot_glorys_overview_vertical_2x2(
            vertical_package=vp,
            variables=['argo_weight', 'sigma', 'thetao', 'salinity'],
            k_val=k,
            b_val=b,
            subject_label=subject_label,
            date_label=date_label,
            xmin=xmin,
            xmax=xmax,
            ymin=ymin,
            ymax=z_plot_max,
            plot_isolines=plot_isolines,
            isoline_levels=isoline_levels,
            isoline_color=isoline_color,
            isoline_linewidth=isoline_linewidth,
            isoline_alpha=isoline_alpha,
            label_isolines=label_isolines,
            title_extra=title_extra,
            source_label='Argo',
        )
        if save_fig:
            out_dir = Path(output_dir) if output_dir else _shared_output_dir('plot_regional_vertical_argo_overview')
            out_dir.mkdir(parents=True, exist_ok=True)
            fname = (
                f'argo_regional_{lon_min:.1f}E_{lon_max:.1f}E_'
                f'{lat_min:.1f}N_{lat_max:.1f}N_'
                f'{t_min.strftime("%Y%m%d") if t_min else "unknown"}_'
                f'{t_max.strftime("%Y%m%d") if t_max else "unknown"}'
                f'_hbw{h_bw:.0f}km.png'
            )
            plt.savefig(out_dir / fname, dpi=150, bbox_inches='tight')
            if verbose:
                print(f'  Saved: {out_dir / fname}')
        if show_fig:
            plt.show()
        plt.close()


def _plot_center_vertical_argo_overview(
    center_lon: float,
    center_lat: float,
    target_date: pd.Timestamp,
    *,
    subject_label: str,
    save_name_prefix: str,
    save_subdir: str,
    n_prof: int | None = None,
    k: float = 0.0,
    radius_km: float = _argo_recon_radius_km,
    day_window: int = _argo_recon_day_window,
    h_bw: float = _argo_recon_h_bw_km,
    depth_bw: float = _argo_recon_depth_bw_m,
    h_spacing_deg: float = _argo_recon_h_spacing_deg,
    z_max_m: float = _argo_recon_z_max_m,
    z_spacing_m: float = _argo_recon_z_spacing_m,
    min_weight: float = _argo_recon_min_weight,
    x_spacing_km: float = _argo_recon_x_spacing_km,
    ymin: float = 0.0,
    ymax: float = 1000.0,
    plot_isolines: bool = True,
    isoline_levels: int | list[float] | np.ndarray | None = None,
    isoline_color: str = 'black',
    isoline_linewidth: float = 0.8,
    isoline_alpha: float = 0.45,
    label_isolines: bool = True,
    field: dict | None = None,
    save_field: bool = False,
    field_path: str | Path | None = None,
    n_jobs: int | None = None,
    show_fig: bool = True,
    save_fig: bool = False,
    output_dir: str | Path | None = None,
    save_name: str | None = None,
    verbose: bool = True,
) -> None:
    """以中心点 + 半径重建并绘制 Argo 垂向断面 2x2 总览图（track / argo 变体共用）。

    将 center_lon/center_lat 加 ±radius_km 转成经纬度盒子，喂给 collect_argo_pool
    与 _build_argo_3d_field，再沿过中心的测线切取 ±radius_km 断面绘图，时间窗取
    target_date ± day_window 天。与 plot_regional_vertical_argo_overview 共用同一套
    重建 / 切片 / 绘图核心，区别仅在于以中心半径而非显式经纬度边界圈定范围。
    """
    clon = float(_normalize_lon_array(center_lon))
    clat = float(center_lat)
    target_date = pd.Timestamp(target_date).normalize()

    lon_min, lon_max, lat_min, lat_max = _window_bounds_from_center_km(clon, clat, radius_km)
    lon_range = (lon_min, lon_max)
    lat_range = (lat_min, lat_max)
    b = clat - k * clon
    xmin, xmax = -float(radius_km), float(radius_km)
    z_plot_max = ymax

    if field is None:
        t_min = target_date - pd.Timedelta(days=int(day_window))
        t_max = target_date + pd.Timedelta(days=int(day_window))
        if verbose:
            print(
                f'[Argo3D] {subject_label}  center ({clon:.2f}, {clat:.2f})  '
                f'{target_date.date()} ±{int(day_window)}d  r={radius_km:.0f}km'
            )
        pool = collect_argo_pool(
            lon_range, lat_range, t_min, t_max,
            max_depth=z_max_m + 200.0,
        )
        if pool.empty:
            print('No Argo data found.')
            return
        n_prof = pool['Profile_number'].nunique()
        if verbose:
            print(f'  {n_prof} profiles, {len(pool)} depth rows')
        field = _build_argo_3d_field(
            pool, lon_range, lat_range,
            h_bw=h_bw, depth_bw=depth_bw, h_spacing_deg=h_spacing_deg,
            z_max_m=z_max_m, z_spacing_m=z_spacing_m, min_weight=min_weight,
            n_jobs=n_jobs, verbose=verbose,
        )
        field['attrs']['start_date'] = t_min.strftime('%Y-%m-%d')
        field['attrs']['end_date'] = t_max.strftime('%Y-%m-%d')
    # field 预建（如 track 拉格朗日场）时跳过采集，n_prof 由调用方经参数传入

    if save_field:
        fp = Path(field_path) if field_path else (
            _shared_output_dir('argo_3d_fields')
            / (
                f'argo3d_{save_name_prefix}_{target_date.strftime("%Y%m%d")}'
                f'_hbw{h_bw:.0f}_dbw{depth_bw:.0f}.zarr'
            )
        )
        fp.parent.mkdir(parents=True, exist_ok=True)
        _save_argo_3d_field(field, fp)
        if verbose:
            print(f'  Saved 3D field: {fp}')

    vertical_packages = slice_section_from_argo_field(
        field, k, clon, clat,
        x_min_km=xmin, x_max_km=xmax, x_spacing_km=x_spacing_km,
    )
    if not vertical_packages:
        return

    date_label = f'{target_date.strftime("%Y-%m-%d")} ±{int(day_window)}d'
    n_label = f'  N={n_prof}' if n_prof is not None else ''
    title_extra = f'  h_bw={h_bw:.0f}km  r={radius_km:.0f}km{n_label}'

    for vp in vertical_packages:
        _plot_glorys_overview_vertical_2x2(
            vertical_package=vp,
            variables=['argo_weight', 'sigma', 'thetao', 'salinity'],
            k_val=k,
            b_val=b,
            subject_label=subject_label,
            date_label=date_label,
            xmin=xmin,
            xmax=xmax,
            ymin=ymin,
            ymax=z_plot_max,
            plot_isolines=plot_isolines,
            isoline_levels=isoline_levels,
            isoline_color=isoline_color,
            isoline_linewidth=isoline_linewidth,
            isoline_alpha=isoline_alpha,
            label_isolines=label_isolines,
            title_extra=title_extra,
            source_label='Argo',
        )
        if save_fig:
            out_dir = Path(output_dir) if output_dir else _shared_output_dir(save_subdir)
            out_dir.mkdir(parents=True, exist_ok=True)
            fname = save_name or (
                f'{save_name_prefix}_argo_{target_date.strftime("%Y%m%d")}'
                f'_hbw{h_bw:.0f}km.png'
            )
            plt.savefig(out_dir / fname, dpi=150, bbox_inches='tight')
            if verbose:
                print(f'  Saved: {out_dir / fname}')
        if show_fig:
            plt.show()
        plt.close()


def _build_lagrangian_argo_field(
    track_df: pd.DataFrame,
    ref_lon: float,
    ref_lat: float,
    ref_date: pd.Timestamp,
    *,
    radius_km: float = _argo_recon_radius_km,
    day_window: int = _argo_recon_day_window,
    h_bw: float = _argo_recon_h_bw_km,
    depth_bw: float = _argo_recon_depth_bw_m,
    h_spacing_deg: float = _argo_recon_h_spacing_deg,
    z_max_m: float = _argo_recon_z_max_m,
    z_spacing_m: float = _argo_recon_z_spacing_m,
    min_weight: float = _argo_recon_min_weight,
    n_jobs: int | None = None,
    verbose: bool = True,
) -> tuple[dict | None, int]:
    """构建随涡（拉格朗日）Argo 重建场，返回 (field, n_prof)。

    将 ref_date ± day_window 窗口内每条 Argo 剖面按其观测当天的涡心（来自
    track_df）重投影到涡旋相对坐标，再以参考涡心（ref_date 当天）为锚点建场，
    从而抵消涡旋在时间窗内的平移、沿轨迹累积更多剖面。窗口内无可用数据时
    返回 (None, 0)。
    """
    ref_lon = float(_normalize_lon_array(ref_lon))
    ref_lat = float(ref_lat)
    ref_date = pd.Timestamp(ref_date).normalize()
    t_min = ref_date - pd.Timedelta(days=int(day_window))
    t_max = ref_date + pd.Timedelta(days=int(day_window))

    tdates = pd.to_datetime(
        track_df['date'] if 'date' in track_df.columns else convert_date(track_df['time']),
        errors='coerce',
    ).dt.normalize().to_numpy()
    elon = pd.to_numeric(track_df['center_lon'], errors='coerce').to_numpy(dtype=float)
    elat = pd.to_numeric(track_df['center_lat'], errors='coerce').to_numpy(dtype=float)
    win = (
        (tdates >= t_min.to_datetime64()) & (tdates <= t_max.to_datetime64())
        & np.isfinite(elon) & np.isfinite(elat)
    )
    if not win.any():
        return None, 0
    centre_by_day = {
        pd.Timestamp(d).normalize(): (float(lo), float(la))
        for d, lo, la in zip(tdates[win], elon[win], elat[win])
    }

    win_lon = np.array([v[0] for v in centre_by_day.values()], dtype=float)
    win_lat = np.array([v[1] for v in centre_by_day.values()], dtype=float)
    pad_lon_deg = float(np.max(np.abs((win_lon - ref_lon + 180.0) % 360.0 - 180.0)))
    pad_lat_deg = float(np.max(np.abs(win_lat - ref_lat)))

    # 采集盒子在场盒子上外扩涡心最大漂移量，保证去平移后落入 ±radius 的剖面不漏采
    fb_lon_lo, fb_lon_hi, fb_lat_lo, fb_lat_hi = _window_bounds_from_center_km(ref_lon, ref_lat, radius_km)
    coll_lon = (fb_lon_lo - pad_lon_deg, fb_lon_hi + pad_lon_deg)
    coll_lat = (fb_lat_lo - pad_lat_deg, fb_lat_hi + pad_lat_deg)

    pool = collect_argo_pool(coll_lon, coll_lat, t_min, t_max, max_depth=z_max_m + 200.0)
    if pool.empty:
        return None, 0

    pool['_d'] = pool['date'].dt.normalize()
    cb = pd.DataFrame({
        '_d': list(centre_by_day.keys()),
        '_elon': [v[0] for v in centre_by_day.values()],
        '_elat': [v[1] for v in centre_by_day.values()],
    })
    pool = pool.merge(cb, on='_d', how='left')
    eddy_lon_d = pool['_elon'].to_numpy(dtype=float)
    eddy_lat_d = pool['_elat'].to_numpy(dtype=float)
    keep = np.isfinite(eddy_lon_d) & np.isfinite(eddy_lat_d)
    pool = pool.loc[keep].copy()
    if pool.empty:
        return None, 0
    eddy_lon_d = eddy_lon_d[keep]
    eddy_lat_d = eddy_lat_d[keep]

    # 去平移：每条剖面相对其当天涡心的偏移，移到参考涡心系（涡旋相对坐标）
    rel_lon = (pool['Longitude'].to_numpy(dtype=float) - eddy_lon_d + 180.0) % 360.0 - 180.0
    rel_lat = pool['Latitude'].to_numpy(dtype=float) - eddy_lat_d
    pool['Longitude'] = ref_lon + rel_lon
    pool['Latitude'] = ref_lat + rel_lat

    n_prof = int(pool['Profile_number'].nunique())
    if verbose:
        scale = approximate_degree_length(ref_lat)
        drift_km = max(
            pad_lon_deg * float(scale['meters_per_degree_lon']),
            pad_lat_deg * float(scale['meters_per_degree_lat']),
        ) / 1000.0
        print(
            f'  Lagrangian composite: {n_prof} profiles, {ref_date.date()} ±{int(day_window)}d, '
            f'eddy drift ≤ {drift_km:.0f} km'
        )

    field = _build_argo_3d_field(
        pool, (fb_lon_lo, fb_lon_hi), (fb_lat_lo, fb_lat_hi),
        h_bw=h_bw, depth_bw=depth_bw, h_spacing_deg=h_spacing_deg,
        z_max_m=z_max_m, z_spacing_m=z_spacing_m, min_weight=min_weight,
        n_jobs=n_jobs, verbose=verbose,
    )
    field['attrs']['start_date'] = t_min.strftime('%Y-%m-%d')
    field['attrs']['end_date'] = t_max.strftime('%Y-%m-%d')
    field['attrs']['lagrangian'] = True
    return field, n_prof


def plot_track_vertical_argo_overview(
    DS: list | str | tuple | dict,
    no: int,
    needed_date: str | pd.Timestamp,
    *,
    k: float = 0.0,
    radius_km: float = _argo_recon_radius_km,
    day_window: int = _argo_recon_day_window,
    h_bw: float = _argo_recon_h_bw_km,
    depth_bw: float = _argo_recon_depth_bw_m,
    h_spacing_deg: float = _argo_recon_h_spacing_deg,
    z_max_m: float = _argo_recon_z_max_m,
    z_spacing_m: float = _argo_recon_z_spacing_m,
    min_weight: float = _argo_recon_min_weight,
    x_spacing_km: float = _argo_recon_x_spacing_km,
    ymin: float = 0.0,
    ymax: float = 1000.0,
    plot_isolines: bool = True,
    isoline_levels: int | list[float] | np.ndarray | None = None,
    isoline_color: str = 'black',
    isoline_linewidth: float = 0.8,
    isoline_alpha: float = 0.45,
    label_isolines: bool = True,
    field: dict | None = None,
    save_field: bool = False,
    field_path: str | Path | None = None,
    n_jobs: int | None = None,
    show_fig: bool = True,
    save_fig: bool = False,
    output_dir: str | Path | None = None,
    verbose: bool = True,
) -> None:
    """随涡（拉格朗日）重建并绘制 track 垂向断面 2x2 总览图。

    取 needed_date 当天的涡旋中心（来自 find_track）为参考涡心，对 ±day_window
    窗口内每条 Argo 剖面按其观测当天的涡心重投影到涡旋相对坐标后合成，从而抵消
    涡旋平移、沿轨迹累积更多剖面（窗口越宽收益越大）；再沿过涡心的测线切取
    ±radius_km 断面。与 plot_track_vertical_glorys_overview 对应，但垂向场来自
    随涡 Argo 重建而非 GLORYS；与单日欧拉式的 plot_argo_vertical_argo_overview 互补。

    参数:
        - DS (list | str | tuple | dict): 涡旋数据源（kind 串 'acl'|'cl'|'cs'|'acs'，或 list/tuple/dict）。
        - no (int): 涡旋 track_id。
        - needed_date (str | pd.Timestamp): 参考日期，须是该 track 中存在的某天，决定参考涡心位置。
        - k (float): 断面测线斜率 Δlat/Δlon，0.0 为纬向，默认 0.0。
        - radius_km (float): 中心半径（km），同时作为采集盒子半宽与断面 x 轴半宽，默认来自配置。
        - day_window (int): 时间窗半宽（天），围绕 needed_date 取 ±day_window，默认来自配置。
        - h_bw (float): 水平高斯核带宽（km，保留中尺度结构），默认来自配置。
        - depth_bw (float): 垂向高斯核带宽（m），默认来自配置。
        - h_spacing_deg (float): 重建网格水平间距（°），默认来自配置。
        - z_max_m (float): 最大重建深度（m），默认来自配置。
        - z_spacing_m (float): 垂向网格间距（m），默认来自配置。
        - min_weight (float): 最小累积权重阈值，低于此值格点显示为 NaN，默认来自配置。
        - x_spacing_km (float): 断面水平采样间距（km），默认来自配置。
        - ymin (float): 图纵轴深度上界（m），默认 0.0。
        - ymax (float): 图纵轴深度下界（m），默认 1000.0；3D 场仍建到 z_max_m。
        - plot_isolines (bool): 是否叠加 σ₀ 等值线，默认 True。
        - isoline_levels (int | list[float] | np.ndarray | None): 等值线级别数或显式级别；None 时自动。
        - isoline_color (str): 等值线颜色，默认 'black'。
        - isoline_linewidth (float): 等值线线宽，默认 0.8。
        - isoline_alpha (float): 等值线透明度，默认 0.45。
        - label_isolines (bool): 是否标注等值线数值，默认 True。
        - field (dict | None): 预建（拉格朗日）3D 场，传入后跳过重建直接切片绘图。
        - save_field (bool): 是否将 3D 场保存为 zarr，默认 False。
        - field_path (str | Path | None): zarr 保存路径，None 时按涡旋与日期自动生成。
        - n_jobs (int | None): 重建并行进程数，None 时取 min(cpu_count, 8)。
        - show_fig (bool): 是否显示图片，默认 True。
        - save_fig (bool): 是否保存图片，默认 False。
        - output_dir (str | Path | None): 图片输出目录，None 时使用默认路径。
        - verbose (bool): 是否打印进度信息，默认 True。
    """
    track_df, ds_name, _ = _resolve_track_context(DS, no, include_contours=True)

    dates = pd.to_datetime(
        track_df['date'] if 'date' in track_df.columns else convert_date(track_df['time']),
        errors='coerce',
    )
    target_ts = pd.Timestamp(needed_date).normalize()
    same_day_idx = np.nonzero(dates.dt.normalize().to_numpy() == target_ts.to_datetime64())[0]
    if same_day_idx.size == 0:
        raise ValueError(f"Date {target_ts.strftime('%Y-%m-%d')} not found in track {no}.")
    needed_idx = int(same_day_idx[0])

    center_lon = float(pd.to_numeric(track_df['center_lon'], errors='coerce').iloc[needed_idx])
    center_lat = float(pd.to_numeric(track_df['center_lat'], errors='coerce').iloc[needed_idx])
    ds_name_upper = ds_name.upper() if isinstance(ds_name, str) else 'UNKNOWN'

    n_prof = None
    if field is None:
        if verbose:
            print(
                f'[Argo3D-L] Track {ds_name_upper}{int(no)}  ref涡心 ({center_lon:.2f}, {center_lat:.2f})  '
                f'{target_ts.date()} ±{int(day_window)}d  r={radius_km:.0f}km'
            )
        field, n_prof = _build_lagrangian_argo_field(
            track_df, center_lon, center_lat, target_ts,
            radius_km=radius_km, day_window=day_window,
            h_bw=h_bw, depth_bw=depth_bw, h_spacing_deg=h_spacing_deg,
            z_max_m=z_max_m, z_spacing_m=z_spacing_m, min_weight=min_weight,
            n_jobs=n_jobs, verbose=verbose,
        )
        if field is None:
            print('No Argo data found.')
            return

    _plot_center_vertical_argo_overview(
        center_lon, center_lat, target_ts,
        subject_label=f'Track {ds_name_upper}{int(no)} (Lagrangian)',
        save_name_prefix=f'{ds_name_upper}{int(no)}_L{int(day_window)}d',
        save_subdir='plot_track_vertical_argo_overview',
        n_prof=n_prof,
        k=k, radius_km=radius_km, day_window=day_window,
        h_bw=h_bw, depth_bw=depth_bw, h_spacing_deg=h_spacing_deg,
        z_max_m=z_max_m, z_spacing_m=z_spacing_m, min_weight=min_weight,
        x_spacing_km=x_spacing_km, ymin=ymin, ymax=ymax,
        plot_isolines=plot_isolines, isoline_levels=isoline_levels,
        isoline_color=isoline_color, isoline_linewidth=isoline_linewidth,
        isoline_alpha=isoline_alpha, label_isolines=label_isolines,
        field=field, save_field=save_field, field_path=field_path, n_jobs=n_jobs,
        show_fig=show_fig, save_fig=save_fig, output_dir=output_dir, verbose=verbose,
    )


def plot_argo_vertical_argo_overview(
    profile_number: int,
    profile_time: int | str | pd.Timestamp,
    *,
    platform_number: int | None = None,
    k: float = 0.0,
    radius_km: float = _argo_recon_radius_km,
    day_window: int = _argo_recon_day_window,
    h_bw: float = _argo_recon_h_bw_km,
    depth_bw: float = _argo_recon_depth_bw_m,
    h_spacing_deg: float = _argo_recon_h_spacing_deg,
    z_max_m: float = _argo_recon_z_max_m,
    z_spacing_m: float = _argo_recon_z_spacing_m,
    min_weight: float = _argo_recon_min_weight,
    x_spacing_km: float = _argo_recon_x_spacing_km,
    ymin: float = 0.0,
    ymax: float = 1000.0,
    plot_isolines: bool = True,
    isoline_levels: int | list[float] | np.ndarray | None = None,
    isoline_color: str = 'black',
    isoline_linewidth: float = 0.8,
    isoline_alpha: float = 0.45,
    label_isolines: bool = True,
    argo_data_dir: str | Path | None = None,
    field: dict | None = None,
    save_field: bool = False,
    field_path: str | Path | None = None,
    n_jobs: int | None = None,
    show_fig: bool = True,
    save_fig: bool = False,
    output_dir: str | Path | None = None,
    verbose: bool = True,
) -> None:
    """以 Argo 剖面位置为中心，用 Argo 数据重建并绘制垂向断面 2x2 总览图。

    取目标剖面（profile_number + profile_time）的位置与日期，以 ±radius_km 圈定
    范围、±day_window 天为时间窗，重建 3D Argo 高斯核场并沿过该剖面的测线切取断面
    （单日中心固定的欧拉快照，不随涡平移）。与 plot_argo_vertical_glorys_overview
    对应，但垂向场来自 Argo 重建而非 GLORYS；与随涡的 plot_track_vertical_argo_overview 互补。

    参数:
        - profile_number (int): 目标 Argo 剖面编号。
        - profile_time (int | str | pd.Timestamp): 年份（如 2014）或具体日期（如 '2014-05-09'），用于定位剖面。
        - platform_number (int | None): 浮标编号，剖面编号同年重复时用于消歧，默认 None。
        - k (float): 断面测线斜率 Δlat/Δlon，0.0 为纬向，默认 0.0。
        - radius_km (float): 中心半径（km），同时作为采集盒子半宽与断面 x 轴半宽，默认来自配置。
        - day_window (int): 时间窗半宽（天），围绕剖面日期取 ±day_window，默认来自配置。
        - h_bw (float): 水平高斯核带宽（km，保留中尺度结构），默认来自配置。
        - depth_bw (float): 垂向高斯核带宽（m），默认来自配置。
        - h_spacing_deg (float): 重建网格水平间距（°），默认来自配置。
        - z_max_m (float): 最大重建深度（m），默认来自配置。
        - z_spacing_m (float): 垂向网格间距（m），默认来自配置。
        - min_weight (float): 最小累积权重阈值，低于此值格点显示为 NaN，默认来自配置。
        - x_spacing_km (float): 断面水平采样间距（km），默认来自配置。
        - ymin (float): 图纵轴深度上界（m），默认 0.0。
        - ymax (float): 图纵轴深度下界（m），默认 1000.0；3D 场仍建到 z_max_m。
        - plot_isolines (bool): 是否叠加 σ₀ 等值线，默认 True。
        - isoline_levels (int | list[float] | np.ndarray | None): 等值线级别数或显式级别；None 时自动。
        - isoline_color (str): 等值线颜色，默认 'black'。
        - isoline_linewidth (float): 等值线线宽，默认 0.8。
        - isoline_alpha (float): 等值线透明度，默认 0.45。
        - label_isolines (bool): 是否标注等值线数值，默认 True。
        - argo_data_dir (str | Path | None): Argo 数据目录，None 时用默认 argo_path。
        - field (dict | None): 预建 3D 场，传入后跳过重建直接切片绘图。
        - save_field (bool): 是否将 3D 场保存为 zarr，默认 False。
        - field_path (str | Path | None): zarr 保存路径，None 时按剖面与日期自动生成。
        - n_jobs (int | None): 重建并行进程数，None 时取 min(cpu_count, 8)。
        - show_fig (bool): 是否显示图片，默认 True。
        - save_fig (bool): 是否保存图片，默认 False。
        - output_dir (str | Path | None): 图片输出目录，None 时使用默认路径。
        - verbose (bool): 是否打印进度信息，默认 True。
    """
    info = _resolve_argo_profile_center(
        profile_number=int(profile_number),
        profile_time=profile_time,
        platform_number=platform_number,
        argo_data_dir=argo_data_dir,
    )
    target_date = pd.Timestamp(info['target_date'])
    _plot_center_vertical_argo_overview(
        float(info['center_lon']), float(info['center_lat']), target_date,
        subject_label=f'Profile {int(profile_number)}',
        save_name_prefix=f'P{int(profile_number)}',
        save_subdir='plot_argo_vertical_argo_overview',
        k=k, radius_km=radius_km, day_window=day_window,
        h_bw=h_bw, depth_bw=depth_bw, h_spacing_deg=h_spacing_deg,
        z_max_m=z_max_m, z_spacing_m=z_spacing_m, min_weight=min_weight,
        x_spacing_km=x_spacing_km, ymin=ymin, ymax=ymax,
        plot_isolines=plot_isolines, isoline_levels=isoline_levels,
        isoline_color=isoline_color, isoline_linewidth=isoline_linewidth,
        isoline_alpha=isoline_alpha, label_isolines=label_isolines,
        field=field, save_field=save_field, field_path=field_path, n_jobs=n_jobs,
        show_fig=show_fig, save_fig=save_fig, output_dir=output_dir, verbose=verbose,
    )


# 帮助函数：判断三个点 (p, q, r) 的方向（共线，顺时针，逆时针）
def _orientation(p, q, r):
    val = (q[1] - p[1]) * (r[0] - q[0]) - (q[0] - p[0]) * (r[1] - q[1])
    if val == 0: return 0  # Collinear
    return 1 if val > 0 else 2 # Clockwise or Counterclockwise (1 for clockwise, 2 for counterclockwise)

# 帮助函数：判断点 q 是否在线段 pr 上
def _on_segment(p, q, r):
    return (q[0] <= max(p[0], r[0]) and q[0] >= min(p[0], r[0]) and
            q[1] <= max(p[1], r[1]) and q[1] >= min(p[1], r[1]))

# 核心帮助函数：计算两条线段的交点
def _line_segment_intersect(p1, q1, p2, q2):
    """
    计算两条线段 (p1, q1) 与 (p2, q2) 的交点。
    若相交则返回 (x, y)，否则返回 None；共线且重叠的情况会返回重叠端点。
    """
    o1 = _orientation(p1, q1, p2)
    o2 = _orientation(p1, q1, q2)
    o3 = _orientation(p2, q2, p1)
    o4 = _orientation(p2, q2, q1)

    # General case: Non-collinear and non-parallel intersection
    if o1 != o2 and o3 != o4:
        denom = (p1[0] - q1[0]) * (p2[1] - q2[1]) - (p1[1] - q1[1]) * (p2[0] - q2[0])
        if denom == 0:
            return None 

        t = ((p1[0] - p2[0]) * (p2[1] - q2[1]) - (p1[1] - p2[1]) * (p2[0] - q2[0])) / denom
        u = -((p1[0] - q1[0]) * (p1[1] - p2[1]) - (p1[1] - q1[1]) * (p1[0] - p2[0])) / denom

        if (0 <= t <= 1) and (0 <= u <= 1):
            intersection_x = p1[0] + t * (q1[0] - p1[0])
            intersection_y = p1[1] + t * (q1[1] - p1[1])
            return (intersection_x, intersection_y)
        return None

    # Special Cases for collinear segments
    if o1 == 0 and _on_segment(p1, p2, q1): return p2
    if o2 == 0 and _on_segment(p1, q2, q1): return q2
    if o3 == 0 and _on_segment(p2, p1, q2): return p1
    if o4 == 0 and _on_segment(p2, q1, q2): return q1

    return None


# 寻找多边形和线段的交点
def find_polygon_line_intersections(polygon_lon, polygon_lat, line_lons, line_lats, tolerance=1e-6):
    """
    计算一条折线与闭合多边形边界的交点。

    参数:
        - polygon_lon (array): 多边形顶点经度。
        - polygon_lat (array): 多边形顶点纬度。
        - line_lons (array): 折线点经度。
        - line_lats (array): 折线点纬度。
        - tolerance (float): 判定重复交点的容差，默认 1e-6。

    返回:
        - list: 交点的 (lon, lat) 元组列表。
    """
    intersections = []
    
    if len(polygon_lon) < 3: # 需要至少三个点才能构成多边形
        return []

    polygon_points_closed = list(zip(polygon_lon, polygon_lat))
    if not (np.isclose(polygon_points_closed[0][0], polygon_points_closed[-1][0], atol=tolerance) and
            np.isclose(polygon_points_closed[0][1], polygon_points_closed[-1][1], atol=tolerance)):
        polygon_points_closed.append(polygon_points_closed[0])

    line_points = list(zip(line_lons, line_lats))

    for i in range(len(polygon_points_closed) - 1):
        poly_seg_p1 = polygon_points_closed[i]
        poly_seg_p2 = polygon_points_closed[i+1]
        
        for j in range(len(line_points) - 1):
            line_seg_p1 = line_points[j]
            line_seg_p2 = line_points[j+1]
            
            intersect_pt = _line_segment_intersect(poly_seg_p1, poly_seg_p2, line_seg_p1, line_seg_p2)
            if intersect_pt:
                is_duplicate = False
                for existing_pt in intersections:
                    if np.isclose(existing_pt[0], intersect_pt[0], atol=tolerance) and \
                       np.isclose(existing_pt[1], intersect_pt[1], atol=tolerance):
                        is_duplicate = True
                        break
                if not is_duplicate:
                    intersections.append(intersect_pt)
    
    if len(intersections) > 0:
        if np.abs(line_lons[0] - line_lons[-1]) > np.abs(line_lats[0] - line_lats[-1]):
            intersections.sort(key=lambda p: p[0])
        else:
            intersections.sort(key=lambda p: p[1])
            
    return intersections

def regrid_vertical_slice(data_package: dict, dy: float = None, dz: float = None) -> dict:
    '''
    将垂直剖面数据包插值到新的等间距网格上。

    接收 get_vertical_glorys 的输出，根据用户指定的水平（dy）或垂直（dz）间距生成网格化更规整的新数据包。

    参数:
        - data_package (dict): 从 get_vertical_glorys 获取的原始数据包。
        - dy (float | None): 新的水平（y 轴）网格间距（km）；None 时不改变水平网格。
        - dz (float | None): 新的垂直（z 轴）网格间距（m）；None 时不改变垂直网格。

    返回:
        - dict: 结构与输入相同但数据已插值到新网格的新数据包，含以下键：

            - 'profile_data' (dict)：内部结构同输入，但每个二维数组为插值后的结果。
            - 'y_coords' (np.ndarray)：提供 dy 时为新生成的等间距数组，否则同输入。
            - 'z_coords' (np.ndarray)：提供 dz 时为新生成的等间距数组，否则同输入。
            - 'lon_coords' / 'lat_coords' (np.ndarray)：水平坐标被重采样时为新插值数组，否则同输入。
            - 'projections' (dict)：从原始数据包原样复制，数值仍对应原始坐标系。
            - 'metadata' (dict)：从原始数据包原样复制。
    '''
    # 如果用户未指定任何新的网格间距，则直接返回原始数据包的深拷贝
    if dy is None and dz is None:
        return copy.deepcopy(data_package)

    # --- 1. 提取原始数据和坐标 ---
    original_y = data_package['y_coords']
    original_z = data_package['z_coords']
    original_profiles = data_package['profile_data']
    original_lon = data_package['lon_coords']
    original_lat = data_package['lat_coords']
    
    # RegularGridInterpolator 要求坐标必须是严格递增的。
    if np.any(np.diff(original_z) <= 0):
        raise ValueError("原始z坐标（深度）必须是严格递增的才能进行插值。")
    if np.any(np.diff(original_y) <= 0):
        raise ValueError("原始y坐标（距离）必须是严格递增的才能进行插值。")

    # --- 2. 创建新的网格坐标 ---
    new_y = np.arange(original_y.min(), original_y.max(), dy) if dy is not None else original_y
    if dz is not None:
        z_min, z_max = original_z.min(), original_z.max()
        new_z = np.arange(z_min, z_max, dz)
    else:
        new_z = original_z

    # --- 3. 对每个变量的剖面数据及经纬度进行插值 ---
    # 插值经纬度 (1D)，仅当水平坐标被重采样时执行
    if dy is not None:
        new_lon = np.interp(new_y, original_y, original_lon)
        new_lat = np.interp(new_y, original_y, original_lat)
    else:
        new_lon, new_lat = original_lon, original_lat

    # 插值剖面数据 (2D)
    new_profiles = {}
    # 创建查询点网格，用于高效插值, 'ij'索引确保维度顺序正确
    zz_grid, yy_grid = np.meshgrid(new_z, new_y, indexing='ij')
    query_points = np.vstack([zz_grid.ravel(), yy_grid.ravel()]).T

    for var, data_array in original_profiles.items():
        if data_array.mask.all():
            new_profiles[var] = np.ma.masked_all((len(new_z), len(new_y)))
            continue

        # 创建插值器实例
        interp_func = RegularGridInterpolator(
            (original_z, original_y), 
            data_array.filled(np.nan),
            method='linear', 
            bounds_error=False, 
            fill_value=np.nan
        )
        
        # 在新网格上执行插值
        new_data_flat = interp_func(query_points)
        
        # 将一维插值结果重塑为二维网格
        new_data = new_data_flat.reshape(len(new_z), len(new_y))
        
        # 将结果转回 masked array 并存入新字典
        new_profiles[var] = np.ma.masked_invalid(new_data)

    # --- 4. 构建并返回新的数据包 ---
    new_data_package = {
        'profile_data': new_profiles,
        'y_coords': new_y,
        'z_coords': new_z,
        'lon_coords': new_lon,
        'lat_coords': new_lat,
        'projections': copy.deepcopy(data_package['projections']),
        'metadata': copy.deepcopy(data_package['metadata'])
    }

    return new_data_package

def plot_data_package(data_package: dict, DS: list, variable: str,
                      show_fig: bool = False, save_fig: bool = False, xmin: float = None, xmax: float = None,
                      ymin: float = None, ymax: float = None):
    '''
    根据一个数据包和原始涡旋数据集，绘制单一物理量的垂直剖面图。

    灵活的可视化接口：接收数据包并根据传入的涡旋数据集（DS）确定绘图风格（颜色、标题），适用于
    get_vertical_glorys 或 regrid_vertical_slice 的输出。

    参数:
        - data_package (dict): 从 get_vertical_glorys 或 regrid_vertical_slice 获取的数据包。
        - DS (list): 原始涡旋数据集（用于解析涡旋类型显示样式）。
        - variable (str): 需要绘制的变量名。
        - show_fig (bool): 是否显示图像，默认 False。
        - save_fig (bool): 是否保存图像，默认 False。
        - xmin (float | None): 横坐标下界（km）；与 xmax 同时给定时按当前步长在该范围内重建目标 y 网格并裁剪到有效覆盖区间，减少仅 set_xlim 导致的空白。
        - xmax (float | None): 横坐标上界（km），与 xmin 配套。
        - ymin (float | None): 纵坐标下界（m）。
        - ymax (float | None): 纵坐标上界（m）。
    输出:
        - 图像（save_fig=True 时）：`plot_outputs/shared/<region>/plot_track_vertical_glorys/{数据集}{编号}_vertical_{变量}_YYYYMMDD_k*b*.png`
    '''
    # --- 1. 验证输入并解包数据 ---
    required_keys = ['profile_data', 'y_coords', 'z_coords', 'lon_coords', 'lat_coords', 'projections', 'metadata']
    if not data_package or not all(k in data_package for k in required_keys):
        print(f"错误: 输入的 data_package 格式不完整。缺少键。")
        return
        
    profile_variable_2d = data_package['profile_data'].get(variable)
    if profile_variable_2d is None or np.all(np.ma.getmaskarray(np.ma.array(profile_variable_2d, copy=False))):
        print(f"警告: 变量 '{variable}' 的剖面数据无效，无法绘图。")
        return

    y_coords = np.asarray(data_package['y_coords'], dtype=float)
    z_coords = np.asarray(data_package['z_coords'], dtype=float)
    profile_variable_2d = np.ma.array(profile_variable_2d, copy=False)

    if profile_variable_2d.ndim != 2:
        print(f"警告: 变量 '{variable}' 不是二维剖面数据，无法绘图。")
        return

    # 对齐坐标和数据维度，避免因尺寸不一致引发绘图或插值异常。
    nz = min(profile_variable_2d.shape[0], len(z_coords))
    ny = min(profile_variable_2d.shape[1], len(y_coords))
    if nz < 2 or ny < 2:
        print(f"警告: 变量 '{variable}' 的有效网格点不足（至少需要 2x2）。")
        return

    profile_variable_2d = profile_variable_2d[:nz, :ny]
    z_coords = z_coords[:nz]
    y_coords = y_coords[:ny]

    if np.any(np.diff(z_coords) <= 0):
        z_order = np.argsort(z_coords)
        z_coords = z_coords[z_order]
        profile_variable_2d = profile_variable_2d[z_order, :]

    if np.any(np.diff(y_coords) <= 0):
        y_order = np.argsort(y_coords)
        y_coords = y_coords[y_order]
        profile_variable_2d = profile_variable_2d[:, y_order]

    # 与 plot_track_vertical_glorys 一致：当给定 x 范围时，先在该范围重建等步长 y 网格再绘图。
    xlim_to_apply = None
    if xmin is not None and xmax is not None:
        x_left = float(min(xmin, xmax))
        x_right = float(max(xmin, xmax))

        dy_candidates = np.diff(y_coords)
        dy_candidates = dy_candidates[np.isfinite(dy_candidates) & (dy_candidates > 1e-12)]
        if dy_candidates.size > 0 and x_right > x_left:
            target_dy = float(np.median(dy_candidates))
            n_steps = int(np.floor((x_right - x_left) / target_dy))
            if n_steps >= 1:
                y_target = x_left + np.arange(n_steps + 1) * target_dy
                valid_target = (y_target >= float(np.min(y_coords))) & (y_target <= float(np.max(y_coords)))
                if np.count_nonzero(valid_target) >= 2:
                    y_target_valid = y_target[valid_target]
                    zz_grid, yy_grid = np.meshgrid(z_coords, y_target_valid, indexing='ij')
                    query_points = np.vstack([zz_grid.ravel(), yy_grid.ravel()]).T

                    interp_func = RegularGridInterpolator(
                        (z_coords, y_coords),
                        np.ma.filled(profile_variable_2d, np.nan),
                        method='linear',
                        bounds_error=False,
                        fill_value=np.nan,
                    )
                    interp_flat = interp_func(query_points)
                    profile_variable_2d = np.ma.masked_invalid(
                        interp_flat.reshape(len(z_coords), len(y_target_valid))
                    )
                    y_coords = y_target_valid

        x_plot_min = max(x_left, float(np.min(y_coords)))
        x_plot_max = min(x_right, float(np.max(y_coords)))
        if np.isfinite(x_plot_min) and np.isfinite(x_plot_max) and x_plot_max > x_plot_min:
            xlim_to_apply = (x_plot_min, x_plot_max)

    projections = data_package['projections']
    metadata = data_package['metadata']

    # --- 2. 在当前上下文中计算与DS相关的元数据 ---
    ds_name = metadata.get('ds_name')
    if not ds_name:
        track_id_for_resolve = metadata.get('eddy_no')
        if track_id_for_resolve is not None:
            try:
                _, ds_name, _ = _resolve_track_context(DS, track_id_for_resolve, include_contours=False)
            except Exception:
                ds_name = "UNKNOWN"
        else:
            ds_name = "UNKNOWN"
    ds_name = ('Argo' if ds_name.lower() == 'argo' else ds_name.upper()) if isinstance(ds_name, str) else str(ds_name)

    prop_colors = plt.rcParams['axes.prop_cycle'].by_key()['color']
    eddy_color = prop_colors[1] if 'AC' in ds_name else prop_colors[0]

    # --- 3. 准备绘图元素 ---

    Y_mesh, Z_mesh = np.meshgrid(y_coords, z_coords)
    
    clim = None
    if variable == 'vorticity': cbar_label, cmap, clim = r'$\zeta/f$', 'seismic', (-0.3, 0.3)
    elif variable in ['thetao']: cbar_label, cmap = 'Temperature (°C)', 'rainbow'
    elif variable in ['salinity', 'so']: cbar_label, cmap = 'Salinity (psu)', 'viridis'
    elif variable in ['density', 'sigma', 'sigma0']: cbar_label, cmap = 'Potential Density Anomaly (σ0, kg/m³)', 'RdBu_r'
    elif variable in ['u', 'v', 'uo', 'vo']: cbar_label, cmap = 'Velocity (m/s)', 'RdBu_r'
    else: cbar_label, cmap = variable, 'viridis'
    
    if clim is None:
        prof_ma = np.ma.array(profile_variable_2d, copy=False)
        valid_values = prof_ma.compressed()
        clim = (valid_values.min(), valid_values.max()) if valid_values.size > 0 else (0,1)
        if variable in ['u', 'v', 'uo', 'vo']:
            max_abs = np.max(np.abs(valid_values)) if valid_values.size > 0 else 1
            clim = (-max_abs, max_abs)

    # --- 4. 执行绘图 ---
    fig, ax = plt.subplots(figsize=(20, 15))
    
    k_val, b_val = metadata.get('k'), metadata.get('b')
    if k_val is not None and b_val is not None:
        title = (f"Vertical Profile of {cbar_label} for Track {ds_name}{metadata['eddy_no']} "
                 f"on {metadata['date_str']}, y={k_val:.2f}x{b_val:+.2f}")
    else:
        title = (f"Vertical Profile of {cbar_label} for Track {ds_name}{metadata['eddy_no']} "
                 f"on {metadata['date_str']}")

    ax.set_title(title, fontsize=20)
    ax.set_xlabel('Distance from Eddy Center Projection (km)', fontsize=18)
    ax.set_ylabel('Depth (m)', fontsize=18)
    ax.tick_params(labelsize=14)
    
    pc = ax.pcolormesh(Y_mesh, Z_mesh, profile_variable_2d, cmap=cmap, shading='auto', vmin=clim[0], vmax=clim[1])
    cbar = fig.colorbar(pc, ax=ax, orientation='vertical', fraction=0.046, pad=0.04)
    cbar.set_label(cbar_label, fontsize=18)
    cbar.ax.tick_params(labelsize=14)

    ax.axvline(0, color='black', linestyle='--', linewidth=2, label='Eddy Center Projection')
    for i, dist in enumerate(projections['radius']): 
        ax.axvline(dist, color='r', linestyle='--', linewidth=2, label='Effective Radius Projection' if i == 0 else "")
    for i, dist in enumerate(projections['contour']): 
        ax.axvline(dist, color=eddy_color, linestyle=':', linewidth=2, label='Effective Contour Projection' if i == 0 else "")

    ax.set_ylim(z_coords.max(), z_coords.min())
    if xlim_to_apply is not None:
        ax.set_xlim(xlim_to_apply[0], xlim_to_apply[1])
    elif xmin is not None and xmax is not None:
        ax.set_xlim(min(xmin, xmax), max(xmin, xmax))
    if ymin is not None and ymax is not None: ax.set_ylim(ymax, ymin)
    
    handles, labels = ax.get_legend_handles_labels()
    by_label = dict(zip(labels, handles))
    ax.legend(by_label.values(), by_label.keys(), fontsize=18)

    # --- 5. 保存和显示 ---
    if save_fig:
        region_slug = _current_region_key()
        output_dir = _shared_output_dir("plot_track_vertical_glorys", region_slug)
        output_dir.mkdir(parents=True, exist_ok=True)
        date_fn = metadata['date_str'].replace('-', '')
        
        if k_val is not None and b_val is not None:
            k_str = f"k{k_val:.2f}"
            b_str = f"b{b_val:.2f}"
            if ds_name == 'Argo':
                base_filename = (
                    f"Argo_{date_fn}_profile{metadata['eddy_no']}_vertical_{variable}_{k_str}{b_str}.png"
                )
            else:
                base_filename = (f"{ds_name}{metadata['eddy_no']}_vertical_{variable}_{date_fn}_{k_str}{b_str}.png")
        else:
            if ds_name == 'Argo':
                base_filename = f"Argo_{date_fn}_profile{metadata['eddy_no']}_vertical_{variable}.png"
            else:
                base_filename = (f"{ds_name}{metadata['eddy_no']}_vertical_{variable}_{date_fn}.png")

        save_path = output_dir / base_filename
        plt.savefig(save_path, dpi=300, bbox_inches='tight')
        print(f"Figure saved to: {save_path}")

    if show_fig:
        plt.show()

    plt.close(fig)

def get_raytraceR_inputs(data_package: dict,
                                f_coriolis: float, # 科里奥利频率 f (例如 1.454e-4)。**必须提供**。
                                rho0: float = 1025.0, # 参考密度 (kg/m^3)。
                                g: float = 9.81, # 重力加速度 (m/s^2)。
                                m0: float = 1/100.0, # 初始垂直波数。
                                omega_factor: float = 0.97, # 固有频率 omega = omega_factor * f。
                                v0_amplitude: float = 0.1, # 初始波速振幅。
                                thresh_val: float = 0.1e-2, # 内部反射阈值。
                                chstart_val: int = 1, # 初始特征线 (1或2)。
                                filter_sigma_z: float = 0.5, # 沿Z方向高斯平滑的标准差（用于导数计算，防止噪声放大）。
                                filter_sigma_y: float = 0.5 # 沿Y方向高斯平滑的标准差。
                               ) -> dict:
    """
    将来自 track.regrid_vertical_slice 的数据包转换为 MATLAB raytraceR 函数所需的输入格式。

    参数:
        - data_package (dict): GLORYS 垂直切片数据（通常是 regrid_vertical_slice 输出），结构见“说明”。
        - f_coriolis (float): 科里奥利频率 f（如 1.454e-4 rad/s），必须提供。
        - rho0 (float): 参考密度（kg/m³）。
        - g (float): 重力加速度（m/s²）。
        - m0 (float): 初始垂直波数。
        - omega_factor (float): 固有频率倍数，omega = omega_factor · f。
        - v0_amplitude (float): 初始波速振幅。
        - thresh_val (float): 内部反射阈值。
        - chstart_val (int): 初始特征线（1 或 2）。
        - filter_sigma_z (float): 沿深度方向高斯平滑标准差（用于导数计算）。
        - filter_sigma_y (float): 沿横流方向高斯平滑标准差。

    返回:
        - dict: 含 raytraceR 所需所有输入参数的字典。

    说明:
        输入 data_package 结构（来自 get_vertical_glorys + regrid_vertical_slice）:

            - z_coords (np.ndarray)：等间距深度 Z 坐标（m）。
            - y_coords (np.ndarray)：等间距横流 Y 坐标（km）。
            - lon_coords / lat_coords (np.ndarray)：剖面线上每个点的经度/纬度。
            - profile_data (dict)：含 'u'（纬向速度 uo）、'v'（经向速度 vo）、'salinity'（实用盐度 SP）、'thetao'（位温 pt）四个二维数组。
    """

    # 1. 提取和映射输入数据
    # raytraceR 的 Z 是垂直/深度，Y 是水平横流。
    z_g = data_package['z_coords'] # 深度坐标，单位：米 (m)
    y_g_km = data_package['y_coords'] # 横流坐标，单位：公里 (km)
    lon_coords_1d = data_package['lon_coords'] # 剖面经度 (1D array)
    lat_coords_1d = data_package['lat_coords'] # 剖面纬度 (1D array)

    # 将横流坐标从公里转换为米，以保持 raytraceR 内部单位一致性
    y_g = y_g_km * 1000.0 # 横流坐标，单位：米 (m)

    # 从 profile_data 中获取核心物理量
    # 它们已经是 (深度点数, 横流点数) 的形状
    # .filled(np.nan) 用于处理 MaskedArray，将其掩码值填充为 NaN
    u_data = data_package['profile_data']['u'].filled(np.nan)
    v_data = data_package['profile_data']['v'].filled(np.nan)
    salinity_data = data_package['profile_data']['salinity'].filled(np.nan) # Practical Salinity (SP)
    thetao_data = data_package['profile_data']['thetao'].filled(np.nan) # Potential Temperature (pt)

    # 检查维度和数据有效性
    expected_shape_2d = (len(z_g), len(y_g_km))
    if not (u_data.shape == salinity_data.shape == thetao_data.shape == expected_shape_2d):
        raise ValueError(f"Profile data shapes are inconsistent or do not match expected ({len(z_g)}, {len(y_g_km)}).")
    if not np.isfinite(f_coriolis):
        raise ValueError("Coriolis frequency 'f_coriolis' must be a finite number.")

    # 2. 计算网格间距
    dz = np.mean(np.diff(z_g)) # regrid_vertical_slice 保证了等间距
    dy = np.mean(np.diff(y_g)) # 横流坐标已经转换为米，这里计算米为单位的间距

    # 3. 对物理量进行平滑处理 (减少噪声对导数计算的影响)
    # 确保 sigma 值与数据尺寸兼容
    u_smoothed = gaussian_filter(u_data, sigma=(filter_sigma_z, filter_sigma_y))
    v_smoothed = gaussian_filter(v_data, sigma=(filter_sigma_z, filter_sigma_y))
    salinity_smoothed = gaussian_filter(salinity_data, sigma=(filter_sigma_z, filter_sigma_y))
    thetao_smoothed = gaussian_filter(thetao_data, sigma=(filter_sigma_z, filter_sigma_y))
    
    # 处理平滑后可能出现的NaN（高斯滤波的边界效应），尤其是在mask区域外推的NaN
    u_smoothed[np.isnan(u_smoothed)] = 0.0 # 假设用0填充，或选择其他填充值
    v_smoothed[np.isnan(v_smoothed)] = 0.0
    salinity_smoothed[np.isnan(salinity_smoothed)] = np.nanmean(salinity_smoothed) if np.all(np.isnan(salinity_smoothed)) else salinity_smoothed[np.isnan(salinity_smoothed)].mean() # 用平均值填充，避免gsw出错
    thetao_smoothed[np.isnan(thetao_smoothed)] = np.nanmean(thetao_smoothed) if np.all(np.isnan(thetao_smoothed)) else thetao_smoothed[np.isnan(thetao_smoothed)].mean() # 同上

    # 4. 计算核心物理量: 密度, 浮力 (bg), N2, ug, F2, S2, s_M, omegamin
    
    # 4.1. 转换为 gsw 所需的 Absolute Salinity (SA) 和 Conservative Temperature (CT)
    # gsw 需要经度和纬度网格与数据形状匹配
    Z_mesh_m, Y_mesh_m = np.meshgrid(z_g, y_g, indexing='ij') # Z_mesh_m (深度, 米), Y_mesh_m (横流，米)
    
    # 将 1D lon/lat_coords 扩展为 2D 网格，匹配 (深度点数, 横流点数) 形状
    # 假设 lon_coords_1d 和 lat_coords_1d 是与 y_g_km 对应的，即沿横流方向的经纬度
    # 那么它们需要沿着深度方向广播
    lon_2d = np.tile(lon_coords_1d[np.newaxis, :], (len(z_g), 1))
    lat_2d = np.tile(lat_coords_1d[np.newaxis, :], (len(z_g), 1))

    # 计算压力（p_from_z 需要负深度）
    pressure_2d = gsw.p_from_z(-Z_mesh_m, lat_2d)

    # 实用盐度 -> 绝对盐度
    SA_processed = gsw.SA_from_SP(salinity_smoothed, pressure_2d, lon_2d, lat_2d)
    
    # 位温 -> 保守温度
    CT_processed = gsw.CT_from_pt(SA_processed, thetao_smoothed)

    # 计算位势密度 (referenced to 0 dbar)
    rho_potential_processed = gsw.rho(SA_processed, CT_processed, 0)
    
    # 计算浮力 bg = -g * (rho_potential - rho0) / rho0
    bg_processed = -g * (rho_potential_processed - rho0) / rho0

    # 4.2. 计算 N2 (浮力频率的平方)
    # gsw.Nsquared(SA, CT, p, lat=None, axis=0) 返回 (N2, p_mid)
    # axis=0 表示沿深度（第一个）维度计算梯度
    N2_output, p_mid_N2 = gsw.Nsquared(SA_processed, CT_processed, pressure_2d, lat=lat_2d, axis=0)
    
    # N2_output 的形状比原始数据在深度维度上少一个点 (因为是 mid-point 导数)
    # 需要将其插值或映射回原始深度网格的尺寸
    # 常用方法是插值回原始深度点或在第一个点进行外推/复制
    # 这里我们采用在第一个深度点复制 N2_output 的第一行，使其维度匹配
    N2_processed = np.vstack([N2_output[0:1, :], N2_output])
    # 再次裁剪以确保与原始深度维度完全一致（防止因浮点数问题多/少一行）
    N2_processed = N2_processed[:u_data.shape[0], :] 
    
    # 处理 N2 中可能出现的 NaN 或负值（物理上 N2 应为正或零）
    N2_processed[np.isnan(N2_processed)] = 0.0 # 用 0 填充 NaN
    N2_processed[N2_processed < 0] = 0.0 # 强制非负

    # 4.3. 计算 ug (背景地转流场)
    # 论文中 ug 是背景地转流的 x 分量。这里直接使用平滑后的 u 分量作为 ug。
    ug_processed = u_smoothed

    # 4.4. 计算 F2, S2 (根据 Whitt and Thomas 2013 论文定义)
    
    # zeta_g = -du_g/dy
    # np.gradient 自动处理 MaskedArray，但我们已填充为 NaN
    # 注意 dy 已经是米
    du_dy_smoothed = np.gradient(ug_processed, dy, axis=1) # du_g/dy
    zeta_g_processed = -du_dy_smoothed
    
    # F2 = f * (f + zeta_g)
    F2_processed = f_coriolis * (f_coriolis + zeta_g_processed)

    # S2 = f * du_g/dz
    # 注意 dz 已经是米
    du_dz_smoothed = np.gradient(ug_processed, dz, axis=0) # du_g/dz
    S2_processed = f_coriolis * du_dz_smoothed
    
    # 4.5. 计算 s_M = F2 / S2
    # 避免除以零或 NaN
    s_M_processed = np.divide(F2_processed, S2_processed, out=np.zeros_like(F2_processed), where=S2_processed!=0)
    s_M_processed[np.isnan(s_M_processed)] = 0.0 # 填充 NaN
    s_M_processed[np.isinf(s_M_processed)] = 0.0 # 填充 Inf

    # 4.6. 计算 omegamin
    # omegamin = sqrt(q/N2)
    # q = F2 * N2 - S2^4
    PV_q = F2_processed * N2_processed - S2_processed**4
    
    # 避免根号下出现负值或除以零
    omegamin_squared = np.divide(PV_q, N2_processed, out=np.zeros_like(PV_q), where=N2_processed!=0)
    omegamin_squared = np.maximum(0, omegamin_squared) # 确保非负

    omegamin_processed = np.sqrt(omegamin_squared)
    omegamin_processed = np.real(omegamin_processed) # 确保是实数
    omegamin_processed[np.isnan(omegamin_processed)] = 0.0 # 填充 NaN
    omegamin_processed[np.isinf(omegamin_processed)] = 0.0 # 填充 Inf

    # 5. 二阶导数 (根据论文示例，默认可以设为零，或者在这里计算)
    # 鉴于计算二阶导数对噪声敏感且需要额外代码，我们这里保持为零，
    # 除非用户明确要求并提供计算方法。
    d2udy2_processed = np.zeros_like(u_data)
    d2udz2_processed = np.zeros_like(u_data)
    d2bdy2_processed = np.zeros_like(u_data)
    d2bdz2_processed = np.zeros_like(u_data)
    d2bdzdy_processed = np.zeros_like(u_data)
    d2udzdy_processed = np.zeros_like(u_data)

    # 6. 计算领域长度和波参数
    omega = omega_factor * f_coriolis
    Lz = z_g.max() - z_g.min()
    Ly = y_g.max() - y_g.min() # y_g 已经是米

    # m0 = 1/minv，所以 minv = 1/m0
    minv_val = 1.0 / m0

    # 准备输出字典，包含 raytraceR 所需所有参数
    raytraceR_inputs = {
        'Z': -Z_mesh_m, # 2D 深度网格 (深度, 横流)，单位：米 (m)
        'dz': dz, # 深度间距 (米)
        'Y': Y_mesh_m, # 2D 横流网格 (深度, 横流)，单位：米 (m)
        'dy': dy, # 横流间距 (米)
        'F2': F2_processed, # 2D 矩阵 (深度, 横流)
        'S2': S2_processed, # 2D 矩阵 (深度, 横流)
        'N2': N2_processed, # 2D 矩阵 (深度, 横流)
        'ug': ug_processed, # 2D 矩阵 (深度, 横流)
        'bg': bg_processed, # 2D 矩阵 (深度, 横流)
        's_M': s_M_processed, # 2D 矩阵 (深度, 横流)
        'omegamin': omegamin_processed, # 2D 矩阵 (深度, 横流)
        'f': f_coriolis, # 标量
        'g': g, # 标量
        'rho0': rho0, # 标量
        'omega': omega, # 标量
        'minv': minv_val, # 标量
        'v0': v0_amplitude, # 标量
        
        # 二阶导数 (目前设为零)
        'd2udy2': d2udy2_processed,
        'd2udz2': d2udz2_processed,
        'd2bdy2': d2bdy2_processed,
        'd2bdz2': d2bdz2_processed,
        'd2bdzdy': d2bdzdy_processed,
        'd2udzdy': d2udzdy_processed,
        
        # 领域长度 (根据 Z, Y 范围计算)
        'Lz': Lz,
        'Ly': Ly,
        
        # 其他 raytraceR 需要但通常在调用时设定的参数
        'thresh': thresh_val,
        'm0': m0, # 与 minv 互补，此处也返回
        'chstart': chstart_val,
        # projections数据用于Matlab中绘图
        'eddy_radius': data_package['projections'].get('radius', []), # 半径投影
        'eddy_contour': data_package['projections'].get('contour', []), # 等值线投影
    }
    
    print("\n--- Data Preparation Summary for raytraceR ---")
    print(f"Z (Depth) meshgrid shape: {Z_mesh_m.shape}, min/max: {Z_mesh_m.min():.1f}/{Z_mesh_m.max():.1f} m")
    print(f"Y (Cross-stream) meshgrid shape: {Y_mesh_m.shape}, min/max: {Y_mesh_m.min()/1000:.1f}/{Y_mesh_m.max()/1000:.1f} km")
    print(f"dz (Depth spacing): {dz:.3f} m")
    print(f"dy (Cross-stream spacing): {dy:.3f} m")
    print(f"Calculated F2 shape: {F2_processed.shape}, min/max: {np.nanmin(F2_processed):.2e}/{np.nanmax(F2_processed):.2e}")
    print(f"Calculated S2 shape: {S2_processed.shape}, min/max: {np.nanmin(S2_processed):.2e}/{np.nanmax(S2_processed):.2e}")
    print(f"Calculated N2 shape: {N2_processed.shape}, min/max: {np.nanmin(N2_processed):.2e}/{np.nanmax(N2_processed):.2e}")
    print(f"Calculated ug shape: {ug_processed.shape}, min/max: {np.nanmin(ug_processed):.2e}/{np.nanmax(ug_processed):.2e}")
    print(f"Calculated bg shape: {bg_processed.shape}, min/max: {np.nanmin(bg_processed):.2e}/{np.nanmax(bg_processed):.2e}")
    print(f"Calculated omegamin shape: {omegamin_processed.shape}, min/max: {np.nanmin(omegamin_processed):.2e}/{np.nanmax(omegamin_processed):.2e}")
    print(f"Coriolis frequency (f): {f_coriolis:.4e} rad/s")
    print(f"Wave frequency (omega): {omega:.4e} rad/s")
    print("--------------------------------------------\n")

    return raytraceR_inputs

def init_worker(eddy_data_shared: dict):
    """
    multiprocessing 子进程的初始化器：为每个工作进程设置共享的只读数据。

    在每个工作进程启动时仅调用一次，接收大数据集（如涡旋数据字典）并设为该进程的全局变量，避免在任务
    间重复传输大数据，是性能优化的关键；同时将 OpenMP/MKL/OpenBLAS 线程数限制为 1，避免多进程×多线程
    争抢 CPU。

    参数:
        - eddy_data_shared (dict): 供子进程只读共享的涡旋数据字典。
    """
    # 限制 OpenMP/MKL 线程数，避免多进程 x 多线程导致 CPU 争抢
    # 注意：这只影响子进程环境，不影响主进程
    os.environ['OMP_NUM_THREADS'] = '1'
    os.environ['MKL_NUM_THREADS'] = '1'
    os.environ['OPENBLAS_NUM_THREADS'] = '1'

    # 强制切换 Matplotlib 后端为非交互式 (Agg)，避免多进程绘图死锁
    # 注意：init_worker 仅在子进程中运行，此处修改不会影响 Jupyter 主进程的交互式绘图
    try:
        plt.switch_backend('Agg')
    except Exception:
        pass
    
    global worker_eddy_datasets
    worker_eddy_datasets = eddy_data_shared

def check_single_track(
    track_data,
    argo_by_date,
    start_date,
    end_date,
    ds_name,
    circle_enlargement_factor: float | None = None,
    use_adaptive_circle: bool = False,
    adaptive_lat_threshold: float = 70.0,
    adaptive_distance_threshold_km: float = 300.0,
    force_great_circle_circle: bool = False,
    save_interacting_argo: bool = False,
):
    """
    检查单个涡旋轨迹是否与 Argo 数据有交集（内部辅助函数）。

    纯计算函数，为所有在时间范围内的涡旋返回结果，并附带布尔标志说明其是否与 Argo 浮标有交集。

    参数:
        - track_data (list): 单条涡旋的轨迹数据（list of lists）。
        - argo_by_date (dict): 按日期组织的 Argo 明细，形如 {date: list[dict]}，每个 dict 至少含 'Longitude'/'Latitude'，可附带 Profile_number、Year/Month/Day、delta_do、do_value/DO、Anomaly_depth 等元数据。
        - start_date (pd.Timestamp): 检查的开始日期。
        - end_date (pd.Timestamp): 检查的结束日期。
        - ds_name (str): 数据集名称（如 'ACS'）。
        - circle_enlargement_factor (float | None): 半径放大因子，None 时回退全局配置。
        - use_adaptive_circle (bool): True 时半径距离用 adaptive_distance_m 自适应大圆，默认 False。
        - adaptive_lat_threshold (float): |lat| 超过此值触发大圆距离，默认 70.0。
        - adaptive_distance_threshold_km (float): 平面距离超过此值（km）触发大圆距离，默认 300.0。
        - force_great_circle_circle (bool): 强制使用大圆距离（忽略阈值），默认 False。
        - save_interacting_argo (bool): True 时收集并返回每个命中的 Argo 点明细（含 method/track 等），False 时为性能在首个命中即停止当日迭代且不返回点明细，默认 False。

    返回:
        - dict | None: 涡旋在时间范围内时返回含绘图/判定信息的字典，否则返回 None；主要键：

            - 'track_id'：轨迹编号。
            - 'has_interaction'：是否与 Argo 交互。
            - 'in_range_segments'：连续片段用于绘图。
            - 'contours_to_plot'：命中多边形时用于绘图的等值线。
            - 'candidate_dates_for_contour'：圆命中待进一步二次判定的日期。
            - 'dates_in_range'：轨迹在时间窗内的日期。
            - 'text_info'：绘图标签信息。
            - 'is_ace'：是否反气旋。
            - 'interacting_argo'：save_interacting_argo=True 时返回 list[dict]（每个至少含 date/lon/lat/method(poly|circle)/ds_name，并附带 Profile_number/指标等），否则为空列表。
    """
    num, time, center_lon, center_lat, _, _, contour_lon, contour_lat, radius, _, _ = zip(*track_data)
    dates = convert_date(time)
    
    indices_in_range = np.where((dates >= start_date) & (dates <= end_date))[0]
    if indices_in_range.size == 0:
        return None

    if circle_enlargement_factor is None:
        circle_enlargement_factor = globals().get('circle_enlargement_factor', 1.2)

    has_interaction = False
    contours_to_plot = []
    candidate_dates_for_contour = set()
    interacting_points: list[dict] = []
    for i in indices_in_range:
        current_date = dates[i].normalize()
        if current_date in argo_by_date:
            center_lon_i = float(center_lon[i])
            center_lat_i = float(center_lat[i])
            radius_i = float(radius[i])

            # 预计算等值线多边形（若存在）
            contour_poly = None
            contour_lon_norm = None
            try:
                contour_lon_i = contour_lon[i]
                contour_lat_i = contour_lat[i]
                if contour_lon_i is not None and contour_lat_i is not None:
                    contour_lon_arr = np.asarray(contour_lon_i, dtype=float)
                    contour_lat_arr = np.asarray(contour_lat_i, dtype=float)
                    if contour_lon_arr.size >= 3 and contour_lat_arr.size >= 3 and contour_lon_arr.shape == contour_lat_arr.shape:
                        contour_lon_norm = center_lon_i + _minimal_lon_diff_deg(contour_lon_arr, center_lon_i)
                        contour_poly = MplPath(list(zip(contour_lon_norm, contour_lat_arr)))
            except Exception:
                contour_poly = None

            points_today = argo_by_date[current_date]

            any_point_inside_poly = False
            added_contour_for_today = False
            for idx_pt, argo_point in enumerate(points_today):
                # argo_point 为 dict，至少含 Longitude/Latitude；兼容 lon/lat 键
                point_lon = float(argo_point.get('Longitude', argo_point.get('lon')))
                point_lat = float(argo_point.get('Latitude', argo_point.get('lat')))
                if use_adaptive_circle:
                    distance_m = adaptive_distance_m(
                        point_lon,
                        point_lat,
                        center_lon_i,
                        center_lat_i,
                        wrap_dateline=True,
                        gc_lat_threshold=adaptive_lat_threshold,
                        gc_distance_threshold_km=adaptive_distance_threshold_km,
                        force_great_circle=force_great_circle_circle
                    )
                else:
                    scale_ci = approximate_degree_length(center_lat_i)
                    dlon_deg = _minimal_lon_diff_deg(point_lon, center_lon_i)
                    dx_m = dlon_deg * scale_ci['meters_per_degree_lon']
                    dy_m = (point_lat - center_lat_i) * scale_ci['meters_per_degree_lat']
                    distance_m = np.hypot(dx_m, dy_m)
                inside_circle = distance_m <= radius_i * circle_enlargement_factor

                inside_poly_point = False
                if contour_poly is not None:
                    point_lon_norm = center_lon_i + _minimal_lon_diff_deg(point_lon, center_lon_i)
                    try:
                        inside_poly_point = contour_poly.contains_point((point_lon_norm, point_lat))
                    except Exception:
                        inside_poly_point = False
                if inside_poly_point:
                    any_point_inside_poly = True

                if inside_poly_point or inside_circle:
                    has_interaction = True
                    if inside_circle:
                        candidate_dates_for_contour.add(current_date)
                    # 仅在首次发现多边形命中时加入对应等值线以供绘制
                    if inside_poly_point and (not added_contour_for_today) and contour_poly is not None and contour_lon[i] is not None and contour_lat[i] is not None:
                        contours_to_plot.append((contour_lon[i], contour_lat[i]))
                        added_contour_for_today = True
                    if save_interacting_argo:
                        rec = {
                            'method': 'poly' if inside_poly_point else 'circle',
                            'ds_name': ds_name,
                        }
                        if isinstance(argo_point, dict):
                            rec.update(argo_point)
                        # 若缺失日期拆分列，使用当前日期补齐
                        if 'Year' not in rec or pd.isna(rec.get('Year')):
                            rec['Year'] = int(current_date.year)
                        if 'Month' not in rec or pd.isna(rec.get('Month')):
                            rec['Month'] = int(current_date.month)
                        if 'Day' not in rec or pd.isna(rec.get('Day')):
                            rec['Day'] = int(current_date.day)
                        # 移除与原始字段重复的派生列
                        rec.pop('lon', None)
                        rec.pop('lat', None)
                        rec.pop('date', None)
                        interacting_points.append(rec)
                    # 若不需要收集全部点，则命中一个即可跳出当天循环
                    if not save_interacting_argo:
                        break
    
    in_range_segments = []
    splits = np.where(np.diff(indices_in_range) != 1)[0] + 1
    contiguous_blocks = np.split(indices_in_range, splits)
    for block in contiguous_blocks:
        if block.size > 0:
            in_range_segments.append((np.array(center_lon)[block], np.array(center_lat)[block]))

    start_idx = indices_in_range[0]
    # 统一使用旧版结构首字段（已在新路径中改为 track_id）
    _tid_label = int(num[0])
    text_info = {"text": f"{ds_name}{_tid_label}", "lon": center_lon[start_idx], "lat": center_lat[start_idx]}

    return {
        "ds_name": ds_name,
        "track_id": int(_tid_label) if len(num) else None,
        "center_lon": center_lon,
        "center_lat": center_lat,
        "in_range_segments": in_range_segments,
        "contours_to_plot": contours_to_plot,
        "candidate_dates_for_contour": sorted(candidate_dates_for_contour),
        "dates_in_range": [pd.Timestamp(d).normalize() for d in dates[indices_in_range]],
        "text_info": text_info,
        "is_ace": 'AC' in ds_name.upper(),
        "has_interaction": has_interaction,
        "interacting_argo": interacting_points if save_interacting_argo else [],
    }

def plot_all_tracks_in_range(
    start_date_str: str,
    end_date_str: str,
    eddy_datasets: dict | list[str] | tuple[str, ...] | None = None,
    plot_unrelated_eddies: bool = False,
    plot_unrelated_argo: bool = True,
    save_fig: bool = False,
    skip_save_if_empty: bool = False,
    show_labels: bool = True,
    show_fig: bool = True,
    circle_enlargement_factor: float | None = None,
    detection_config: DetectionConfig | None = None,
    anomaly_color_by: str = 'auto',
    fix_colorbar: bool = True,
    cbar_min: float | None = None,
    cbar_max: float | None = None,
    cbar_ticks: list | None = None,
    meta_output_root: str | Path | None = None,
    save_interacted_eddies: bool = False,
    save_interacting_argo: bool = False,
):
    """指定时间段内涡旋轨迹 + Argo 异常代表点的核心绘图（支持 do/aou/trim）。

    依赖 Cartopy 制图，自动处理跨国际日期变更线的轨迹连线：装载时间范围内 Argo 数据、过滤地理范围、按
    detection_config 计算异常（深度限制由 DetectionConfig 统一管理，每个剖面保留 anomaly_score 最强一条），
    anomaly_color_by='auto' 时按当前方法主变量着色。

    参数:
        - start_date_str (str): 起始日期。
        - end_date_str (str): 结束日期。
        - eddy_datasets (dict | list[str] | tuple[str, ...] | None): 涡旋数据集，支持三种形式：

            - 兼容旧版 dict，如 {'ACS': acs, ...}，每个值是轨迹列表。
            - 新版便捷字符串列表/元组，如 ['acs','acl','cs','cl']，按配置从 META_tracks 读取对应时间段与区域内的轨迹。
            - None：并行 worker 模式下默认使用全局 worker_eddy_datasets。
        - plot_unrelated_eddies (bool): 是否绘制未与 Argo 交互的涡旋，默认 False。
        - plot_unrelated_argo (bool): 是否额外绘制所有 Argo 剖面位置（空心圆）作基准分布背景，默认 True。
        - save_fig (bool): 是否保存图像，默认 False。
        - skip_save_if_empty (bool): True 且本图未绘制任何涡旋时跳过保存，默认 False。
        - show_labels (bool): 是否绘制轨迹文本标签（如 ACLXXXX），默认 True。
        - show_fig (bool): 是否显示图像，默认 True。
        - circle_enlargement_factor (float | None): 涡旋边界放大系数；None 时从配置读取。
        - detection_config (DetectionConfig | None): 异常识别配置；None 时使用 processing.yml 默认值。
        - anomaly_color_by (str): 'auto'、'primary_value'、'anomaly_score' 或异常表中任意数值列名，默认 'auto'。
        - fix_colorbar (bool): 是否固定异常主变量色标，默认 True。
        - cbar_min (float | None): 色标下限；None 时自动。
        - cbar_max (float | None): 色标上限；None 时自动。
        - cbar_ticks (list | None): 色标刻度；None 时自动。
        - meta_output_root (str | Path | None): META_tracks 根目录（可覆盖配置默认）。
        - save_interacted_eddies (bool): True 时保存本期交互涡旋标签（NPY），默认 False。
        - save_interacting_argo (bool): True 时保存本期交互 Argo 明细（Parquet），默认 False，输出目录按 detection_config.file_stem() 划分。

    输出:
        保存到 `plot_outputs/<method>/<region>/plot_all_tracks_in_range/<detection_config.file_stem()>/`：

            - `All_Tracks_{start}_to_{end}.png`
            - `Interacted_Eddies_{start}_{end}_{file_stem}.npy`（save_interacted_eddies=True 时）
            - `Interacting_Argo_{start}_{end}_{file_stem}.parquet`（save_interacting_argo=True 时）
    """
    # --- 0. 确定数据源 ---
    local_eddy_datasets = eddy_datasets
    if circle_enlargement_factor is None:
        circle_enlargement_factor = globals().get('circle_enlargement_factor', 1.2)
    cfg = _resolve_detection_config(
        detection_config,
        cbar_min=cbar_min,
        cbar_max=cbar_max,
        cbar_ticks=cbar_ticks,
    )
    method_name = cfg.method
    # 若未显式提供，则在并行 worker 中从全局共享获得
    if local_eddy_datasets is None:
        try:
            global worker_eddy_datasets
            local_eddy_datasets = worker_eddy_datasets
        except NameError:
            local_eddy_datasets = None

    # 新版便捷用法：若传入的是字符串列表/元组（kinds），先不加载等值线以加速第一轮判定
    lazy_contour_mode = False
    if isinstance(local_eddy_datasets, (list, tuple)):
        kinds = [str(k).lower() for k in local_eddy_datasets]
        local_eddy_datasets = _load_eddy_datasets_for_range(
            kinds=kinds,
            start_date=pd.to_datetime(start_date_str),
            end_date=pd.to_datetime(end_date_str),
            output_root=meta_output_root,
            include_contours=False,
        )
        lazy_contour_mode = True

    start_date = pd.to_datetime(start_date_str)
    end_date = pd.to_datetime(end_date_str)
    
    # --- 1. 动态加载Argo数据 ---
    min_year, max_year = start_date.year, end_date.year
    argo_data_list = []
    for year in range(min_year, max_year + 1):
        try:
            argo_data_list.append(load_argo_data(year, data_dir=argo_path))
        except FileNotFoundError:
            print(f"Warning: Argo data for year {year} not found, skipping.")
    if not argo_data_list:
        print(f"No Argo data found for the period {start_date_str} to {end_date_str}.")
        return

    argo_data = pd.concat(argo_data_list, ignore_index=True)
    
    # --- 2. 预处理Argo数据 ---
    argo_dates = pd.to_datetime(argo_data[['Year', 'Month', 'Day']])
    argo_in_range = argo_data[(argo_dates >= start_date) & (argo_dates <= end_date)]
    needed_argo_data = pd.DataFrame()
    base_argo_positions = pd.DataFrame()
    if not argo_in_range.empty:
        lon_vals = argo_in_range['Longitude'].to_numpy(dtype=float, copy=False)
        lat_vals = argo_in_range['Latitude'].to_numpy(dtype=float, copy=False)
        lon_mask = _region_lon_mask(lon_vals, lonmin, lonmax)
        lat_mask = (lat_vals >= latmin) & (lat_vals <= latmax)
        geo_mask = lon_mask & lat_mask
        argo_in_geo_range = argo_in_range[geo_mask].copy()
        if not argo_in_geo_range.empty:
            # 所有剖面基础位置（避免按深度重复）
            base_argo_positions = (
                argo_in_geo_range.sort_values(['Profile_number','Depth'])
                .groupby('Profile_number', as_index=False)
                .first()[['Profile_number','Longitude','Latitude']]
            )
            anomalies = calculate_delta_do(
                argo_in_geo_range,
                detection_config=cfg,
                remove_outliers=True,
                verbose=False
            )
            if not anomalies.empty:
                anomalies_unique = _keep_best_anomaly_per_profile(anomalies, cfg)
                needed_argo_data = anomalies_unique.rename(columns={'depth': 'Anomaly_depth'})

    argo_by_date = defaultdict(list)
    if not needed_argo_data.empty:
        for _, row in needed_argo_data.iterrows():
            date_key = pd.Timestamp(year=int(row['Year']), month=int(row['Month']), day=int(row['Day']))
            row_method = row.get('detection_method', method_name)
            if pd.isna(row_method):
                row_method = method_name
            argo_by_date[date_key].append({
                'Profile_number': row.get('Profile_number'),
                'Longitude': float(row.get('Longitude')),
                'Latitude': float(row.get('Latitude')),
                'Year': int(row.get('Year')),
                'Month': int(row.get('Month')),
                'Day': int(row.get('Day')),
                'delta_do': float(row.get('delta_do')) if 'delta_do' in row else np.nan,
                'delta_aou': float(row.get('delta_aou')) if 'delta_aou' in row else np.nan,
                'trim_score': float(row.get('trim_score')) if 'trim_score' in row else np.nan,
                'anomaly_score': float(row.get('anomaly_score')) if 'anomaly_score' in row else np.nan,
                'do_value': float(row.get('do_value')) if 'do_value' in row else (float(row.get('DO')) if 'DO' in row else np.nan),
                'Anomaly_depth': float(row.get('Anomaly_depth')) if 'Anomaly_depth' in row else np.nan,
                'detection_method': str(row_method).lower(),
            })

    # --- 3. 检查所有涡旋轨迹 ---
    # 兼容两种形态：
    #  - 旧形态：ds_data 为 [track_list, ...]
    #  - 新形态：ds_data 为 [(track_list, track_id), ...]
    all_tracks_with_names: list[tuple[list, str, int | None]] = []
    for ds_name, ds_data in local_eddy_datasets.items():
        for item in ds_data:
            if isinstance(item, tuple) and len(item) == 2 and isinstance(item[1], (int, np.integer)):
                track_list, tid = item
            else:
                track_list, tid = item, None
            all_tracks_with_names.append((track_list, ds_name, tid))
    # 第一轮判定时的圆半径：使用原始 circle_enlargement_factor（不再放大）。
    _cef_first_pass = circle_enlargement_factor
    results = [
        check_single_track(
            track,
            argo_by_date,
            start_date,
            end_date,
            ds_name,
            circle_enlargement_factor=_cef_first_pass,
            save_interacting_argo=bool(save_interacting_argo),
        )
        for track, ds_name, tid in all_tracks_with_names
    ]

    # --- 4. 绘图（第一轮：不画等值线） ---
    crosses_dateline = bool(_REGION_CFG.get('crosses_dateline') and (lonmax < lonmin))
    central_lon = 180 if crosses_dateline else 0
    data_crs = ccrs.PlateCarree()
    map_crs = ccrs.PlateCarree(central_longitude=central_lon)

    criteria_label = cfg.threshold_label()

    fig = plt.figure(figsize=(40, 30))
    ax = fig.add_subplot(1, 1, 1, projection=map_crs)
    ax.set_title(
        f"Eddy Tracks and Argo Anomalies ({cfg.method}; {criteria_label}; {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')})",
        fontsize=20
    )

    # Basemap features（自然地理背景）
    base_ocean = _BASEMAP_COLORS['ocean']
    base_land = _BASEMAP_COLORS['land']
    coast_color = _BASEMAP_COLORS['coastline']
    grid_color = _BASEMAP_COLORS['grid']

    ax.set_facecolor(base_ocean)
    ax.add_feature(cfeature.OCEAN, facecolor=base_ocean, zorder=0)
    ax.add_feature(cfeature.LAND, facecolor=base_land, edgecolor=coast_color, linewidth=0.5, zorder=0)
    ax.add_feature(cfeature.COASTLINE, linewidth=0.7, edgecolor=coast_color, zorder=1)

    gl = ax.gridlines(draw_labels=True, linewidth=0.4, color=grid_color, alpha=0.45, linestyle='--')
    gl.top_labels = False
    gl.right_labels = False

    # 地图范围处理（跨日界线时扩展 longitude）
    lon_extent_min = lonmin
    lon_extent_max = lonmax
    if crosses_dateline:
        if lon_extent_max < lon_extent_min:
            lon_extent_max += 360
    ax.set_extent([lon_extent_min, lon_extent_max, latmin, latmax], crs=data_crs)
    
    prop_colors = plt.rcParams['axes.prop_cycle'].by_key()['color']
    eddy_colors = {'ACE': prop_colors[1], 'CE': prop_colors[0]}

    any_eddy_drawn = False
    label_texts = []
    label_points_x = []
    label_points_y = []
    for result in filter(None, results):
        has_interaction = result['has_interaction']
        if not has_interaction and not plot_unrelated_eddies:
            continue
        
        is_ace = result['is_ace']
        color = eddy_colors['ACE'] if is_ace else eddy_colors['CE']
        text_color = 'red' if has_interaction else 'black'
        
        ax.plot(
            result['center_lon'],
            result['center_lat'],
            color=color,
            alpha=0.4,
            linestyle='--',
            zorder=4,
            transform=ccrs.Geodetic(),
        )
        any_eddy_drawn = True
        for lon_seg, lat_seg in result['in_range_segments']:
            ax.plot(
                lon_seg,
                lat_seg,
                color=color,
                alpha=0.8,
                linestyle='-',
                zorder=4,
                transform=ccrs.Geodetic(),
            )
        
        if show_labels:
            info = result['text_info']
            text_obj = ax.text(
                info['lon'],
                info['lat'],
                info['text'],
                fontsize=12,
                color=text_color,
                weight='bold',
                zorder=5,
                clip_on=False,
                transform=data_crs,
            )
            label_texts.append(text_obj)
            label_points_x.append(info['lon'])
            label_points_y.append(info['lat'])

        # 第一轮不画等值线；若不是 lazy 模式但已有 contours_to_plot，则保持现状
        if (not lazy_contour_mode) and has_interaction:
            for contour_lon, contour_lat in result['contours_to_plot']:
                ax.plot(
                    contour_lon,
                    contour_lat,
                    color=color,
                    linewidth=1,
                    alpha=0.5,
                    zorder=4,
                    linestyle=':',
                    transform=ccrs.Geodetic(),
                )

    if plot_unrelated_argo and not base_argo_positions.empty:
        ax.scatter(
            base_argo_positions['Longitude'], base_argo_positions['Latitude'],
            facecolors='none', edgecolors='gray', linewidths=0.8, s=36,
            label='All Argo Profiles (baseline)', zorder=2,
            transform=data_crs,
        )

    anomaly_legend_label = criteria_label

    if not needed_argo_data.empty:
        scatter_kwargs = {}
        if fix_colorbar:
            cbar_lo, cbar_hi = cfg.resolved_cbar()
            scatter_kwargs.update(dict(vmin=cbar_lo, vmax=cbar_hi))
        depth_label = (
            f' @ depth ≥ {cfg.anomaly_min_depth} m'
            if cfg.anomaly_min_depth is not None and cfg.anomaly_min_depth > 0
            else ''
        )
        color_values, _, color_label, cmap_name = _color_values_for_anomalies(needed_argo_data, cfg)
        if anomaly_color_by not in {'auto', None} and anomaly_color_by in needed_argo_data.columns:
            color_values = pd.to_numeric(needed_argo_data[anomaly_color_by], errors='coerce')
            color_label = anomaly_color_by
        if color_values is None:
            color_values = pd.Series(np.arange(len(needed_argo_data)), index=needed_argo_data.index)
        sc = ax.scatter(
            needed_argo_data['Longitude'], needed_argo_data['Latitude'],
            c=color_values, cmap=cmap_name, s=70,
            edgecolors='black', linewidths=0.5,
            label=f'{anomaly_legend_label}{depth_label}', zorder=3,
            transform=data_crs,
            **scatter_kwargs
        )
        cbar = plt.colorbar(sc, ax=ax, orientation='horizontal', fraction=0.046, pad=0.04)
        cbar.set_label(color_label, fontsize=20); cbar.ax.tick_params(labelsize=14)
        if fix_colorbar:
            _apply_detection_colorbar_ticks(cbar, cfg, cbar_lo, cbar_hi)

    legend_elements = [
        Line2D([0], [0], color=eddy_colors['ACE'], lw=2, label='ACE Track'),
        Line2D([0], [0], color=eddy_colors['CE'], lw=2, label='CE Track')
    ]
    handles, labels = ax.get_legend_handles_labels()
    added = None
    for h, lab in zip(handles, labels):
        if lab.startswith(anomaly_legend_label) or lab == 'Argo Anomaly Profiles':
            added = h
            break
    if added is None:
        for h, lab in zip(handles, labels):
            if lab == 'All Argo Profiles (baseline)':
                added = h
                break
    if added is not None:
        ax.legend(handles=legend_elements + [added], fontsize=18, loc='upper left')
    else:
        ax.legend(handles=legend_elements, fontsize=18, loc='upper left')

    # --- 4.2 第二轮：按需补充绘制等值线（仅对 lazy 模式下有交互的涡旋；使用真实 track_id） ---
    if lazy_contour_mode:
        try:
            need_items = [r for r in results if r and r.get('has_interaction')]
            if need_items:
                need_by_ds: dict[str, dict[int, set[pd.Timestamp]]] = {}
                for r in need_items:
                    ds = r['ds_name']
                    tid = r.get('track_id')
                    if tid is None:
                        continue
                    # 仅针对“半径命中”的候选日期进一步检查等值线
                    dates_req = set(r.get('candidate_dates_for_contour', []) or r.get('dates_in_range', []))
                    if not dates_req:
                        continue
                    need_by_ds.setdefault(ds, {})
                    need_by_ds[ds].setdefault(tid, set()).update(dates_req)

                for ds_name, tid_map in need_by_ds.items():
                    kind_l = ds_name.lower()
                    rich = find_track(
                        kind_l,
                        sorted(tid_map.keys()),
                        region=_current_region_key(),
                        include_contours=True,
                        return_list=False,
                    )
                    # 统一为 {track_id: DataFrame}
                    grouped: dict[int, pd.DataFrame]
                    if isinstance(rich, dict):
                        grouped = {
                            k: pd.DataFrame(v, columns=['track_id','time','center_lon','center_lat','max_lon','max_lat','contour_lon','contour_lat','radius','speed_contour_lon','speed_contour_lat'])
                            for k, v in rich.items()
                        }
                    else:
                        df_rich = pd.DataFrame(rich) if isinstance(rich, list) else rich
                        if isinstance(df_rich, pd.DataFrame) and 'track_id' in df_rich.columns:
                            grouped = {tid: g for tid, g in df_rich.groupby('track_id')}
                        else:
                            only_tid = next(iter(tid_map.keys())) if len(tid_map)==1 else None
                            grouped = {only_tid: df_rich if isinstance(df_rich, pd.DataFrame) else pd.DataFrame(df_rich)}

                    for tid, dates_req in tid_map.items():
                        df_t = grouped.get(tid)
                        if df_t is None or df_t.empty:
                            continue
                        if 'date' not in df_t.columns:
                            if 'time' in df_t.columns:
                                df_t = df_t.copy(); df_t['date'] = convert_date(df_t['time'])
                            else:
                                continue
                        mask = df_t['date'].isin(dates_req)
                        df_needed = df_t[mask]
                        if df_needed.empty:
                            continue
                        color2 = eddy_colors['ACE'] if 'AC' in ds_name else eddy_colors['CE']
                        for _, row in df_needed.iterrows():
                            cl = row.get('contour_lon'); ct = row.get('contour_lat')
                            if isinstance(cl, (list, np.ndarray)) and isinstance(ct, (list, np.ndarray)) and len(cl) >= 3 and len(ct) >= 3:
                                # 二次判定：仅当 Argo 点在等值线内时才绘制该 contour
                                try:
                                    contour_lon_arr = np.asarray(cl, dtype=float)
                                    contour_lat_arr = np.asarray(ct, dtype=float)
                                    center_lon_i = float(row.get('center_lon', contour_lon_arr[0]))
                                    contour_lon_norm = center_lon_i + _minimal_lon_diff_deg(contour_lon_arr, center_lon_i)
                                    contour_lon_norm = np.asarray(contour_lon_norm, dtype=float)
                                    
                                    path = MplPath(list(zip(contour_lon_norm, contour_lat_arr)))

                                    date_norm = row['date'].normalize() if isinstance(row['date'], pd.Timestamp) else pd.Timestamp(row['date']).normalize()
                                    entries = argo_by_date.get(date_norm, [])
                                    draw_it = False
                                    for entry in entries:
                                        try:
                                            pt_lon = float(entry.get('Longitude', entry.get('lon')))
                                            pt_lat = float(entry.get('Latitude', entry.get('lat')))
                                        except Exception:
                                            continue
                                        point_lon_norm = float(center_lon_i + _minimal_lon_diff_deg(pt_lon, center_lon_i))
                                        
                                        if path.contains_point((point_lon_norm, pt_lat)):
                                            draw_it = True
                                            break
                                except Exception:
                                    draw_it = False
                                if draw_it:
                                    ax.plot(
                                        contour_lon_norm,
                                        contour_lat_arr,
                                        color=color2,
                                        linewidth=1,
                                        alpha=0.6,
                                        zorder=4,
                                        linestyle=':',
                                        transform=ccrs.Geodetic(),
                                    )
        except Exception as e:
            print(f"[LazyContour] failed to draw contours lazily: {e}")

    # --- 4.3 标签避让 ---
    orig_xlim = ax.get_xlim()
    orig_ylim = ax.get_ylim()
    region_is_global = _current_region_key().lower() == 'global'
    if region_is_global and adjust_text is not None and label_texts:
        try:
            span_x = max(abs(orig_xlim[1] - orig_xlim[0]), 1e-6)
            span_y = max(abs(orig_ylim[1] - orig_ylim[0]), 1e-6)
            span_ratio_x = span_x / 360.0
            span_ratio_y = span_y / 180.0
            span_ratio = max(span_ratio_x, span_ratio_y, 0.02)
            force_mag = 0.6 * span_ratio
            axis_force = 0.05 * span_ratio
            limit_steps = max(200, int(800 * min(span_ratio, 1.2)))
            adjust_text(
                label_texts,
                x=label_points_x,
                y=label_points_y,
                ax=ax,
                arrowprops=dict(arrowstyle='->', color='gray', lw=0.5, alpha=0.6, shrinkA=6, shrinkB=6),
                expand_text=(1.25, 1.25),
                expand_points=(1.6, 1.6),
                expand_axes=(1.04, 1.04),
                force_text=(force_mag, force_mag),
                force_axes=(axis_force, axis_force),
                lim=limit_steps,
                only_move={'text': 'xy'},
            )
        except Exception as e:
            print(f"[plot_all_tracks_in_range] adjust_text failed: {e}")
    ax.set_xlim(orig_xlim)
    ax.set_ylim(orig_ylim)

    # --- 5. 输出控制 ---
    region_slug_for_path = _current_region_key()
    run_tag = cfg.file_stem()
    run_output_dir = cfg.output_dir("plot_all_tracks_in_range", region_slug_for_path) / run_tag
    if save_fig:
        output_dir = run_output_dir
        output_dir.mkdir(exist_ok=True, parents=True)
        base_filename = f"All_Tracks_{start_date_str}_to_{end_date_str}.png"
        save_path = output_dir / base_filename
        # 若开启跳过空图保存且确认为空（没有任何涡旋轨迹绘制），则跳过
        if skip_save_if_empty and not any_eddy_drawn:
            print(f"Skip saving empty figure for {start_date_str} to {end_date_str} (no eddies plotted).")
        else:
            plt.savefig(save_path, dpi=300, bbox_inches='tight')
            print(f"Figure saved to: {save_path}")
    
    if show_fig:
        plt.show()
    
    plt.close(fig)

    # 汇总输出
    interacted_eddies = []
    interacting_argo_records = []
    for r in filter(None, results):
        if r['has_interaction']:
            interacted_eddies.append(r['text_info']['text'])
        if save_interacting_argo and r.get('interacting_argo'):
            # 补充轨迹标签信息，便于后续统计
            track_label = r['text_info']['text']
            for rec in r['interacting_argo']:
                rec2 = dict(rec)
                rec2['track_label'] = track_label
                rec2['track_id'] = r.get('track_id')
                interacting_argo_records.append(rec2)

    if save_interacting_argo and interacting_argo_records:
        out_dir = run_output_dir
        out_dir.mkdir(exist_ok=True, parents=True)
        fname_pq = out_dir / f"Interacting_Argo_{start_date_str}_to_{end_date_str}_{run_tag}.parquet"
        try:
            df_out = pd.DataFrame(interacting_argo_records)
            df_out.to_parquet(fname_pq, index=False)
            print(f"Interacting Argo saved to: {fname_pq}")
        except Exception as e:
            print(f"[WARN] Failed to save interacting Argo parquet: {e}")

    # 保存交互涡旋标签（每期）为 NPY，并不返回
    try:
        out_dir = run_output_dir
        out_dir.mkdir(exist_ok=True, parents=True)
        if save_interacted_eddies and interacted_eddies:
            eddies_npy = out_dir / f"Interacted_Eddies_{start_date_str}_to_{end_date_str}_{run_tag}.npy"
            # 保存为标准 Unicode 字符串数组，避免 object 导致读取需 allow_pickle=True
            labels = sorted(set(interacted_eddies))
            np.save(eddies_npy, np.array(labels, dtype=str))
            print(f"Interacted eddies saved to: {eddies_npy}")
    except Exception as e:
        print(f"[WARN] Failed to save interacted eddies npy: {e}")

    return None

def _load_eddy_datasets_for_range(
    *,
    kinds: list[str],
    start_date: pd.Timestamp,
    end_date: pd.Timestamp,
    region_key: str | None = None,
    output_root: str | Path | None = None,
    include_contours: bool = False,
) -> dict:
    """根据 kinds（如 ['acs','acl']）与时间/区域，从 Parquet+Zarr 动态装载需要的涡旋轨迹列表。

    返回: dict，如 {'ACS': [(track_list, track_id), ...], 'ACL': [...]}。
    其中 track_list 为“旧版轨迹结构”的 list[list]，track_id 为真实的轨迹ID（与 parquet 的 track_id 对齐），
    便于后续懒加载等值线阶段精准定位。
    """
    region_slug = region_key or _current_region_key()
    root = _ensure_meta_tracks_root(output_root)
    region_dir = Path(root) / region_slug
    if not region_dir.exists():
        raise FileNotFoundError(f"区域目录不存在：{region_dir}")

    def _locate_daily_source(kind_l: str) -> tuple[Path, str]:
        daily_file = region_dir / f"{kind_l}_daily.parquet"
        daily_tmp_dir = region_dir / f"{kind_l}_daily_tmp"
        daily_dir = region_dir / f"{kind_l}_daily.parquet"  # 目录形式
        if daily_file.exists() and daily_file.is_file():
            return daily_file, 'file'
        if daily_tmp_dir.exists() and daily_tmp_dir.is_dir():
            return daily_tmp_dir, 'dir'
        if daily_dir.exists() and daily_dir.is_dir():
            return daily_dir, 'dir'
        raise FileNotFoundError(f"未找到 daily 数据：{daily_file} 或 {daily_tmp_dir} / {daily_dir}")

    # dateline 兼容：使用全局 lonmin/lonmax/latmin/latmax
    def _normalize_lon(val: np.ndarray | float) -> np.ndarray | float:
        return (np.asarray(val, dtype=float) + 180.0) % 360.0 - 180.0

    lon_min_cfg = float(lonmin)
    lon_max_cfg = float(lonmax)
    lat_min_cfg = float(latmin)
    lat_max_cfg = float(latmax)
    lon_min_eff = float(_normalize_lon(lon_min_cfg))
    lon_max_eff = float(_normalize_lon(lon_max_cfg))

    raw_span = abs(lon_max_cfg - lon_min_cfg)
    eff_span = (lon_max_eff - lon_min_eff) % 360.0
    is_global_lon = (raw_span >= 359.5) or (eff_span >= 359.5) or np.isclose(eff_span, 0.0, atol=1e-6)

    crosses_cfg = bool(_REGION_CFG.get('crosses_dateline'))
    simple_interval = (not crosses_cfg) and (not is_global_lon) and (lon_min_eff <= lon_max_eff)

    def _build_lon_mask(lon_vals: np.ndarray) -> np.ndarray:
        if lon_vals.size == 0 or is_global_lon:
            return np.ones(lon_vals.size, dtype=bool)
        lon_norm = _normalize_lon(lon_vals)
        if lon_min_eff <= lon_max_eff:
            return (lon_norm >= lon_min_eff) & (lon_norm <= lon_max_eff)
        return (lon_norm >= lon_min_eff) | (lon_norm <= lon_max_eff)

    def _geo_mask(df: pd.DataFrame) -> np.ndarray:
        if df.empty:
            return np.zeros(0, dtype=bool)
        lon_vals = df['center_lon'].to_numpy(dtype=float, copy=False)
        lat_vals = df['center_lat'].to_numpy(dtype=float, copy=False)
        lat_mask = (lat_vals >= latmin) & (lat_vals <= latmax)
        lon_mask = _build_lon_mask(lon_vals)
        return lat_mask & lon_mask

    out: dict[str, list] = {}
    # 预先计算时间边界（闭开区间），避免逐行日期转换：time ∈ [start_ts, end_exclusive)
    start_ts = pd.Timestamp(start_date).normalize()
    end_exclusive = pd.Timestamp(end_date).normalize() + pd.Timedelta(days=1)
    for kind in kinds:
        kind_l = kind.lower()
        if kind_l not in {'acs','acl','cs','cl'}:
            continue
        daily_source, source_type = _locate_daily_source(kind_l)

        # 收集时间+区域内的 track_id
        track_ids: set[int] = set()
        if source_type == 'file':
            pf = pq.ParquetFile(daily_source)
            for rg in range(pf.num_row_groups):
                # 先用统计信息按时间裁剪（使用 'time' 列的 min/max）
                # 同时在不跨日界线时，尝试用经纬度 min/max 做粗滤
                read_rg = True
                try:
                    schema_names = pf.schema.names
                    # time stats
                    if 'time' in schema_names:
                        t_idx = schema_names.index('time')
                        t_stats = pf.metadata.row_group(rg).column(t_idx).statistics
                        if t_stats and t_stats.has_min_max:
                            tmin = pd.to_datetime(t_stats.min)
                            tmax = pd.to_datetime(t_stats.max)
                            # 跳过与 [start_ts, end_exclusive) 无交集的行组
                            if (tmax < start_ts) or (tmin >= end_exclusive):
                                read_rg = False
                    # lat/lon stats（仅当使用简单区间时启用）
                    if read_rg:
                        try:
                            lat_idx = schema_names.index('center_lat')
                            lat_stats = pf.metadata.row_group(rg).column(lat_idx).statistics
                            if lat_stats and lat_stats.has_min_max:
                                lat_min, lat_max_ = float(lat_stats.min), float(lat_stats.max)
                                if (lat_max_ < lat_min_cfg) or (lat_min > lat_max_cfg):
                                    read_rg = False
                        except Exception:
                            pass
                    if read_rg and simple_interval:
                        try:
                            lon_idx = schema_names.index('center_lon')
                            lon_stats = pf.metadata.row_group(rg).column(lon_idx).statistics
                            if lon_stats and lon_stats.has_min_max:
                                lon_min_rg = float(lon_stats.min)
                                lon_max_rg = float(lon_stats.max)
                                lon_norm_min = float(_normalize_lon(lon_min_rg))
                                lon_norm_max = float(_normalize_lon(lon_max_rg))
                                if lon_norm_min <= lon_norm_max:
                                    if (lon_norm_max < lon_min_eff) or (lon_norm_min > lon_max_eff):
                                        read_rg = False
                                else:
                                    # 无法可靠判断跨日期线的行组，保留
                                    pass
                        except Exception:
                            pass
                except Exception:
                    pass
                if not read_rg:
                    continue

                table = pf.read_row_group(rg, columns=['track_id','time','center_lon','center_lat'])
                df = table.to_pandas()
                # 直接用时间戳闭开区间过滤，避免生成 'date'
                mask_time = (df['time'] >= start_ts) & (df['time'] < end_exclusive)
                mask_geo = _geo_mask(df)
                df_sel = df.loc[mask_time.to_numpy() & mask_geo]
                if not df_sel.empty:
                    track_ids.update(df_sel['track_id'].astype(int).unique().tolist())
        else:
            # 目录形式：逐个 part 文件按 row-group 读取与裁剪，避免整文件加载与逐行日期转换
            for part in sorted(daily_source.glob('*.parquet')):
                try:
                    ppf = pq.ParquetFile(part)
                except Exception:
                    continue
                schema_names = ppf.schema.names
                for rg in range(ppf.num_row_groups):
                    read_rg = True
                    try:
                        # time stats
                        if 'time' in schema_names:
                            t_idx = schema_names.index('time')
                            t_stats = ppf.metadata.row_group(rg).column(t_idx).statistics
                            if t_stats and t_stats.has_min_max:
                                tmin = pd.to_datetime(t_stats.min)
                                tmax = pd.to_datetime(t_stats.max)
                                if (tmax < start_ts) or (tmin >= end_exclusive):
                                    read_rg = False
                        # lat/lon stats prune（不跨日界线时）
                        if read_rg:
                            try:
                                lat_idx = schema_names.index('center_lat')
                                lat_stats = ppf.metadata.row_group(rg).column(lat_idx).statistics
                                if lat_stats and lat_stats.has_min_max:
                                    lat_min, lat_max_ = float(lat_stats.min), float(lat_stats.max)
                                    if (lat_max_ < latmin) or (lat_min > latmax):
                                        read_rg = False
                            except Exception:
                                pass
                        if read_rg and simple_interval:
                            try:
                                lon_idx = schema_names.index('center_lon')
                                lon_stats = ppf.metadata.row_group(rg).column(lon_idx).statistics
                                if lon_stats and lon_stats.has_min_max:
                                    lon_min_rg = float(lon_stats.min)
                                    lon_max_rg = float(lon_stats.max)
                                    lon_norm_min = float(_normalize_lon(lon_min_rg))
                                    lon_norm_max = float(_normalize_lon(lon_max_rg))
                                    if lon_norm_min <= lon_norm_max:
                                        if (lon_norm_max < lon_min_eff) or (lon_norm_min > lon_max_eff):
                                            read_rg = False
                                # 若归一化后出现 wrap（lon_norm_min > lon_norm_max），不做截断
                            except Exception:
                                pass
                    except Exception:
                        pass
                    if not read_rg:
                        continue
                    tbl = ppf.read_row_group(rg, columns=['track_id','time','center_lon','center_lat'])
                    df = tbl.to_pandas()
                    mask_time = (df['time'] >= start_ts) & (df['time'] < end_exclusive)
                    mask_geo = _geo_mask(df)
                    df_sel = df.loc[mask_time.to_numpy() & mask_geo]
                    if not df_sel.empty:
                        track_ids.update(df_sel['track_id'].astype(int).unique().tolist())

        # 组装旧版轨迹结构（批量提取以减少重复 I/O；失败时回退逐个提取）
        tracks: list = []
        if track_ids:
            try:
                batch = find_track(
                    kind_l,
                    sorted(track_ids),
                    region=region_key,
                    output_root=output_root,
                    include_contours=include_contours,
                    return_list=True
                )
                # 批量返回为 {track_id: list}
                tracks = [(batch[tid], int(tid)) for tid in sorted(batch.keys())]
            except Exception:
                for tid in sorted(track_ids):
                    try:
                        track_list = find_track(
                            kind_l,
                            tid,
                            region=region_key,
                            output_root=output_root,
                            include_contours=include_contours,
                            return_list=True
                        )
                        tracks.append((track_list, int(tid)))
                    except Exception:
                        continue

        out[kind_l.upper()] = tracks

    return out

def worker_wrapper(args: tuple):
    """multiprocessing worker 包装函数；结果由 plot_all_tracks_in_range 写盘保存。

    参数:
        - args (tuple): (start_date_str, end_date_str, plot_unrelated_eddies, skip_save_if_empty, show_labels, save_interacting_argo, save_interacted_eddies, cfg)。
    """
    start_d, end_d, unrelated_flag, skip_empty, show_labels, save_interacting_argo_flag, save_eddies_flag, cfg = args
    try:
        plot_all_tracks_in_range(
            start_date_str=start_d,
            end_date_str=end_d,
            plot_unrelated_eddies=unrelated_flag,
            save_fig=True,
            skip_save_if_empty=skip_empty,
            show_labels=show_labels,
            show_fig=False,
            save_interacting_argo=bool(save_interacting_argo_flag),
            save_interacted_eddies=bool(save_eddies_flag),
            detection_config=cfg,
        )
        return None
    except Exception as e:
        print(f"!!! ERROR processing period {start_d}: {e}")
        return None

def run_batch_plotting_multiprocessing(
    start_date_str: str,
    end_date_str: str,
    eddy_datasets: dict,
    num_workers: int,
    plot_unrelated_eddies: bool = False,
    skip_save_if_empty: bool = True,
    show_labels: bool | None = None,
    detection_config: DetectionConfig | None = None,
    save_interacted_eddies: bool = False,
    save_interacting_argo: bool = False,
):
    """
    使用 multiprocessing 启动一个扁平化的并行绘图作业，并带进度条（批处理控制器）。

    创建横跨指定核心数的进程池，用 initializer 高效将涡旋数据共享给所有工作进程，按月份切分任务并用
    tqdm 实时显示进度。

    参数:
        - start_date_str (str): 批处理的开始日期 'YYYY-MM-DD'。
        - end_date_str (str): 批处理的结束日期 'YYYY-MM-DD'。
        - eddy_datasets (dict): 【已加载】的、将被共享给所有进程的涡旋数据集。
        - num_workers (int): 需要启动的并行工作进程数（核心数）。
        - plot_unrelated_eddies (bool): 是否在批处理中绘制无关涡旋，默认 False。
        - skip_save_if_empty (bool): 批处理默认 True（空图不保存）；传 False 时透传至 worker，空图也会保存。
        - show_labels (bool | None): 是否绘制轨迹标签；None 时智能判定（全球且 plot_unrelated_eddies=True 时默认 False）。
        - detection_config (DetectionConfig | None): 异常识别配置；None 时使用 processing.yml 默认值。
        - save_interacted_eddies (bool): True 时各月份写出 `Interacted_Eddies_*.npy` 并在末尾汇总，默认 False。
        - save_interacting_argo (bool): True 时各月份写出 `Interacting_Argo_*.parquet` 并在末尾聚合，默认 False。

    输出:
        - 每月图像写入 `plot_outputs/<method>/<region>/plot_all_tracks_in_range/<file_stem>/`。
        - save_interacted_eddies=True 时整期交互涡旋标签汇总为 `.../eddy_list_<file_stem>.npy`。
        - save_interacting_argo=True 时整期交互 Argo 汇总为 `.../interacting_argo_all_<file_stem>.parquet`。

    说明:
        - 批处理开始前仅清空当前 detection_config.file_stem() 子目录，不会清空其它方法或参数目录。
    """
    print("="*60)
    print("      Multiprocessing Batch Plotting with Progress Bar      ")
    print("="*60)
    print(f"Strategy: Creating a single pool of {num_workers} workers.")
    
    # --- 1. 创建按月切分的任务列表 ---
    month_starts = pd.date_range(start=start_date_str, end=end_date_str, freq='MS')
    region_slug_for_path = _current_region_key()
    cfg = _resolve_detection_config(detection_config)
    print(f"Mode: {cfg.method}")
    run_tag = cfg.file_stem()
    effective_show_labels = show_labels
    if show_labels is None:
        if plot_unrelated_eddies and region_slug_for_path.lower() in {'global', 'world', 'global_ocean', 'all'}:
            effective_show_labels = False
        else:
            effective_show_labels = True

    tasks = [
        (
            start_date.strftime('%Y-%m-%d'),
            (start_date + pd.tseries.offsets.MonthEnd(1)).strftime('%Y-%m-%d'),
            plot_unrelated_eddies,
            skip_save_if_empty,
            effective_show_labels,
            save_interacting_argo,
            save_interacted_eddies,
            cfg,
        )
        for start_date in month_starts
    ]
    
    print(f"[*] Created {len(tasks)} monthly plotting tasks to be processed by {num_workers} cores.")

    # 在批量任务开始前清理输出目录
    base_output_dir = cfg.output_dir("plot_all_tracks_in_range", region_slug_for_path)
    output_dir = base_output_dir / run_tag
    try:
        if output_dir.exists():
            shutil.rmtree(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
        print(f"[*] Cleared threshold-specific output directory: {output_dir}")
    except Exception as e:
        print(f"[WARN] Failed to clear output directory {output_dir}: {e}")
    
    # --- 2. 启动进程池并执行任务 ---
    start_time_total = tm.time()
    
    # 使用 initializer 来高效地传递一次大的涡旋数据（子进程写盘，不收集返回）
    # maxtasksperchild=1: 强制每个子进程处理完一个任务后重启，彻底释放 Matplotlib 内存和文件句柄，防止死锁和内存泄漏
    with multiprocessing.Pool(processes=num_workers, initializer=init_worker, initargs=(eddy_datasets,), maxtasksperchild=1) as pool:
        print("[*] Processing tasks...")
        for _ in tqdm(pool.imap_unordered(worker_wrapper, tasks), total=len(tasks)):
            pass
        
    end_time_total = tm.time()
    total_duration_minutes = (end_time_total - start_time_total) / 60
    
    print("\n" + "="*60)
    print("--- All Plotting Tasks Have Finished ---")
    print(f"Total execution time: {total_duration_minutes:.2f} minutes.")
    # 聚合交互涡旋标签（从每月 NPY 汇总）
    interacted_labels: set[str] = set()
    try:
        monthly_eddy_files = sorted(output_dir.glob("Interacted_Eddies_*.npy"))
        for f in monthly_eddy_files:
            try:
                # 优先以非 pickle 方式读取；兼容历史 object 数组文件回退到允许 pickle
                try:
                    arr = np.load(f)
                except Exception:
                    arr = np.load(f, allow_pickle=True)
                interacted_labels.update([str(x) for x in np.asarray(arr).ravel().tolist()])
            except Exception:
                pass
        unique_interacted = sorted(interacted_labels)
        print(f"Total interacted eddies: {len(unique_interacted)}")
        if unique_interacted:
            preview = ", ".join(unique_interacted[:20])
            print(f"Sample (first 20): {preview}{' ...' if len(unique_interacted)>20 else ''}")
        eddy_list_suffixed = output_dir / f"eddy_list_{run_tag}.npy"
        np.save(eddy_list_suffixed, np.array(unique_interacted, dtype=str))
        print(f"Eddy list saved to: {eddy_list_suffixed}")
    except Exception as e:
        print(f"[WARN] Failed to aggregate eddy labels: {e}")

    # 聚合交互 Argo（Parquet）
    if save_interacting_argo:
        try:
            monthly_argo_files = sorted(output_dir.glob("Interacting_Argo_*.parquet"))
            df_parts = []
            for f in monthly_argo_files:
                try:
                    dfp = pd.read_parquet(f)
                    if isinstance(dfp, pd.DataFrame) and not dfp.empty:
                        df_parts.append(dfp)
                except Exception:
                    pass
            if df_parts:
                df_all = pd.concat(df_parts, ignore_index=True)
                if 'Profile_number' in df_all.columns:
                    if 'date' in df_all.columns:
                        df_all.sort_values(by=['Profile_number', 'date'], inplace=True)
                        df_all = df_all.drop_duplicates(subset=['Profile_number', 'date', 'track_label'], keep='first')
                    elif all(col in df_all.columns for col in ('Year', 'Month', 'Day')):
                        temp_date = pd.to_datetime(df_all[['Year', 'Month', 'Day']])
                        df_all = df_all.assign(_date=temp_date)
                        df_all.sort_values(by=['Profile_number', '_date'], inplace=True)
                        df_all = df_all.drop_duplicates(subset=['Profile_number', '_date', 'track_label'], keep='first')
                        df_all.drop(columns=['_date'], inplace=True)
                argo_parquet = output_dir / f"interacting_argo_all_{run_tag}.parquet"
                df_all.to_parquet(argo_parquet, index=False)
                print(f"Interacting Argo (all) saved to: {argo_parquet}")
        except Exception as e:
            print(f"[WARN] Failed to aggregate/save interacting Argo parquet: {e}")

def load_combined_eddy_list(
    region: str | None = None,
    detection_config: DetectionConfig | None = None,
    thr_dir: str | Path | None = None,
    plots_root: str | Path | None = None,
    include_monthly: bool = True,
    deduplicate: bool = True,
    save_path: str | Path | None = None,
) -> list[str]:
    """读取并合并单个阈值目录下的 `eddy_list*.npy`（及可选月度文件）。

    参数:
        - region (str | None): 区域 slug；None 时复用当前 switch_region 的配置。
        - detection_config (DetectionConfig | None): 异常识别配置；None 时使用 processing.yml 默认值。
        - thr_dir (str | Path | None): 参数子目录名称或路径；None 时自动扫描 do/aou/trim 参数目录并要求只存在一个候选目录。
        - plots_root (str | Path | None): 自定义 plot_outputs 根路径；None 时使用配置 plots_output_root。
        - include_monthly (bool): True 时若目录缺少汇总 `eddy_list*.npy`，退回加载 `Interacted_Eddies_*.npy`（逐月文件），默认 True。
        - deduplicate (bool): True 返回去重并排序后的唯一列表，False 按读取顺序返回，默认 True。
        - save_path (str | Path | None): 可选输出路径；相对路径时会解析到目标阈值目录中。

    返回:
        - list[str]: 聚合后的涡旋标签列表；无匹配文件时返回空列表。
    """
    region_slug = region or _current_region_key()
    cfg = _resolve_detection_config(detection_config)
    plots_base = Path(plots_root) if plots_root is not None else Path(plots_output_root)
    base_dir = plots_base / cfg.method / region_slug / "plot_all_tracks_in_range"
    if not base_dir.exists():
        raise FileNotFoundError(f"Plot outputs directory not found: {base_dir}")

    if thr_dir is not None:
        target_dir = Path(thr_dir)
        if not target_dir.is_absolute():
            target_dir = base_dir / target_dir
        if not target_dir.is_dir():
            raise FileNotFoundError(f"阈值目录不存在：{target_dir}")
    else:
        candidate_dirs = sorted(
            p for p in base_dir.iterdir()
            if p.is_dir() and p.name.lower().split('_', 1)[0] in {'do', 'aou', 'trim'}
        )
        if not candidate_dirs:
            print(f"[load_combined_eddy_list] No threshold directories found under {base_dir}.")
            return []
        if len(candidate_dirs) > 1:
            raise ValueError(f"检测到多个阈值目录，请通过 thr_dir 指定其一：{[p.name for p in candidate_dirs]}")
        target_dir = candidate_dirs[0]

    collected_labels: list[str] = []
    npy_files = sorted(target_dir.glob('eddy_list*.npy'))
    if not npy_files and include_monthly:
        npy_files = sorted(target_dir.glob('Interacted_Eddies_*.npy'))
    for npy_path in npy_files:
        try:
            try:
                arr = np.load(npy_path)
            except Exception:
                arr = np.load(npy_path, allow_pickle=True)
            flattened = np.asarray(arr).ravel().tolist()
            labels = [str(item) for item in flattened if str(item)]
            collected_labels.extend(labels)
        except Exception as exc:
            print(f"[load_combined_eddy_list] Failed to read {npy_path}: {exc}")

    if not collected_labels:
        print("[load_combined_eddy_list] No eddy labels were loaded.")
        return []

    combined = sorted(set(collected_labels)) if deduplicate else collected_labels

    if save_path is None:
        save_path = target_dir / "eddy_list_combined.npy"
    else:
        save_path = Path(save_path)
        if not save_path.is_absolute():
            save_path = target_dir / save_path
    save_path.parent.mkdir(parents=True, exist_ok=True)
    np.save(save_path, np.array(combined, dtype=str))
    print(f"[load_combined_eddy_list] Combined eddy list saved to: {save_path}")

    return combined

def plot_argo_hotspots(
    start_year: int,
    end_year: int,
    detection_config: DetectionConfig | None = None,
    plot_unrelated_argo: bool = True,
    fix_colorbar: bool = True,
    cbar_min: float | None = None,
    cbar_max: float | None = None,
    cbar_ticks: list | None = None,
    save_fig: bool = False,
    show_fig: bool = True,
    save_data: bool = True,
    dask_scheduler: str | None = None,
    dask_workers: int | None = None,
    dask_memory_limit: str | None = None,
    use_interacting_argo: bool = False,
    use_glorys_heave: bool = False,
    argo_glorys_summary_data_path: str | Path | None = None,
    split_plots: bool | str = False,
    hotspot_type_heave_threshold: float | None = _heave_depth_threshold,
) -> dict | None:
    """以 DetectionConfig 指定的异常识别方法绘制多年期 Argo 异常分布。

    逐年加载 Argo 年度数据并合并（可利用全局 lonmin/latmin/lonmax/latmax 做空间裁剪），用
    `calculate_delta_do` 按 detection_config 检测每个剖面的潜在异常（深度限制由 DetectionConfig
    统一管理），每个剖面只保留 anomaly_score 最强的一条记录，再绘制异常散点（可选固定色标范围），
    并可选叠加所有匹配剖面的基线位置（空心灰圈）。

    参数:
        - start_year / end_year (int): 年度范围（闭区间）。
        - detection_config (DetectionConfig | None): 异常识别配置；None 时使用 processing.yml 默认值。
        - plot_unrelated_argo (bool): 是否绘制所有匹配剖面基线（被筛掉或无异常的）。
        - fix_colorbar / cbar_min / cbar_max / cbar_ticks: 异常主变量色标控制。
        - save_fig / show_fig (bool): 输出控制。
        - save_data (bool): True 时保存 anomalies 为 Parquet，False 不保存数据。默认 True。
        - dask_scheduler (str | None): Dask 调度器，`'threads'`|`'processes'`|`'single'`，None 默认 `'processes'`。
        - dask_workers (int | None): Dask worker 数量；None 自动取 min(年度数, CPU)。
        - dask_memory_limit (str | None): LocalCluster 模式下的单 worker 内存限制，如 `'4GB'`；None 不限制。
        - use_interacting_argo (bool): 是否读取 `run_batch_plotting_multiprocessing` 生成的交互 Argo 文件，并在图中区分交互/非交互 Argo，同时统计交互比例。默认 False。
        - use_glorys_heave (bool): 是否读取 `plot_hotspot_anomaly_argo_glorys_overviews` 保存的 summary parquet，用 GLORYS OI 将非近岸异常进一步分为 hotspot_type 1/2。找不到文件或 OI 时退回「近岸第 3 类 / 非第 3 类」绘图。
        - argo_glorys_summary_data_path (str | Path | None): 可选 GLORYS overview summary parquet 路径；None 时按默认命名自动定位。
        - split_plots (bool | str): False 时只绘制合并图；True 时按默认规则拆图（use_glorys_heave=True 或 use_interacting_argo=False 时默认 `'hotspot_type'`，仅 use_interacting_argo=True 时默认 `'eddy_interaction'`）。也可直接传拆图模式字符串：

            - `'eddy_interaction'`：原有 META 交互/非交互拆图；
            - `'hotspot_type'`：按 hotspot_type=1/2/3 拆图；
            - `'spice_type'`：按 spice_type=1/2/3 拆图，需 summary parquet 含 spice_type；
            - `'cross'`：按 hotspot_type{1,2} × spice_type{1,2} 叉乘 + type 3 OMZ 拆图。
        - hotspot_type_heave_threshold (float | None): 当 anomalies 已含 `glorys_heave_zmin` 时，用于生成 hotspot_type 的出露深度阈值。
    返回:
        - dict | None: 含 `summary`（剖面计数/比例与异常、深度统计等汇总 dict）、`figure_paths`（写出的图像路径列表）、`anomalies_path`（异常 Parquet 路径，未保存时为 None）；筛选后无数据时整体返回 None。
    输出:
        - 图像（可选）：`plot_outputs/<method>/<region>/plot_argo_hotspots/Argo_Anomaly_Hotspots_*.png`
        - 异常数据（Parquet，可选）：`plot_outputs/<method>/<region>/plot_argo_hotspots/anomalies_{start}_{end}_{detection_config.file_stem()}.parquet`
    """
    cfg = _resolve_detection_config(
        detection_config,
        cbar_min=cbar_min,
        cbar_max=cbar_max,
        cbar_ticks=cbar_ticks,
    )
    method_name = cfg.method
    run_tag = cfg.file_stem()
    split_mode_explicit = isinstance(split_plots, str)
    if isinstance(split_plots, str):
        split_mode_raw = split_plots
        split_plots = True
    else:
        split_mode_raw = None

    if use_interacting_argo and use_glorys_heave and not split_mode_explicit:
        print("[Plot Info] Both use_interacting_argo and use_glorys_heave are enabled; plotting uses hotspot_type by default.")

    split_mode = None
    if split_plots:
        if split_mode_raw is None:
            split_mode_raw = 'hotspot_type' if (use_glorys_heave or not use_interacting_argo) else 'eddy_interaction'
        split_mode_key = str(split_mode_raw).strip().lower().replace('-', '_')
        split_aliases = {
            'interaction': 'eddy_interaction',
            'interacting': 'eddy_interaction',
            'eddy': 'eddy_interaction',
            'eddy_interaction': 'eddy_interaction',
            'meta': 'eddy_interaction',
            'type': 'hotspot_type',
            'types': 'hotspot_type',
            'hotspot_type': 'hotspot_type',
            'cross': 'cross',
            'spice_type': 'spice_type',
            'spice': 'spice_type',
        }
        split_mode = split_aliases.get(split_mode_key, split_mode_key)
        if split_mode not in {'eddy_interaction', 'hotspot_type', 'cross', 'spice_type'}:
            raise ValueError("split_plots must be bool or one of 'eddy_interaction', 'hotspot_type', 'cross', 'spice_type'.")

    # --- 尝试加载交互 Argo 文件（若启用） ---
    interacting_argo_ids: set[int] = set()
    if use_interacting_argo:
        region_slug_for_path = _current_region_key()
        interacting_file = (
            cfg.output_dir("plot_all_tracks_in_range", region_slug_for_path)
            / run_tag
            / f"interacting_argo_all_{run_tag}.parquet"
        )
        
        if interacting_file.exists():
            print(f"[*] Loading interacting Argo from: {interacting_file}")
            try:
                df_int = pd.read_parquet(interacting_file)
                if 'detection_method' in df_int.columns:
                    method_mask = df_int['detection_method'].astype(str).str.lower().eq(method_name)
                    total_count = len(df_int)
                    mismatch_count = int((~method_mask).sum())
                    if mismatch_count > 0:
                        print(f"[WARN] Mixed detection_method found: expected={method_name}, mismatched={mismatch_count}/{total_count}.")
                    else:
                        print(f"[*] Method={method_name}, records={total_count}")
                    df_int = df_int[method_mask].copy()
                if 'Profile_number' in df_int.columns:
                    interacting_argo_ids = set(df_int['Profile_number'].unique())
                print(f"[*] Loaded {len(interacting_argo_ids)} unique interacting profiles.")
            except Exception as e:
                print(f"[WARN] Failed to read interacting Argo file: {e}")
        else:
            print(f"[WARN] Interacting Argo file not found: {interacting_file}")

    print(f"--- Building Argo anomaly map {start_year}-{end_year} (method={method_name}) ---")

    # --- 按年份加载策略：串行或并行 ---
    years = list(range(start_year, end_year + 1))
    # Worker 参数打包列表（避免闭包，支持多进程 pickling）
    # 捕获当前区域边界到参数中，避免在 Dask 子进程中重新导入模块后回退默认区域
    current_lon_min = float(lonmin)
    current_lon_max = float(lonmax)
    current_lat_min = float(latmin)
    current_lat_max = float(latmax)
    worker_args_list = [
        (
            y,
            cfg,
            current_lon_min,
            current_lon_max,
            current_lat_min,
            current_lat_max,
        )
        for y in years
    ]

    # Dask-only backend
    # 结果累积容器
    baselines_list: list[pd.DataFrame] = []
    anomalies_list: list[pd.DataFrame] = []

    sched = dask_scheduler or 'processes'
    worker_count = dask_workers or min(len(years), os.cpu_count() or 1)
    print(f"[*] Hotspots Dask mode: scheduler={sched}, workers={worker_count}, years={len(years)}")
    cluster = None
    client = None
    if sched == 'processes':
        try:
            cluster = LocalCluster(n_workers=worker_count, threads_per_worker=1,
                                   memory_limit=dask_memory_limit or None,
                                   silence_logs='CRITICAL')
            client = Client(cluster)
        except Exception as e:
            print(f"[WARN] 创建 LocalCluster 失败，改用 dask.compute scheduler='{sched}': {e}")
            cluster = None
            client = None

    delayed_tasks = [delayed(_hotspot_year_worker)(args) for args in worker_args_list]
    # 显示进度：优先使用 distributed.as_completed + tqdm；否则回退到 ProgressBar
    if client is not None:
        try:
            futures = client.compute(delayed_tasks)
            for fut in tqdm(as_completed(futures), total=len(futures), desc="hotspots(dask)"):
                try:
                    baseline_year, anomalies_year = fut.result()
                except Exception as e:
                    print(f"[hotspots] Dask task failed: {e}")
                    continue
                if not baseline_year.empty:
                    baselines_list.append(baseline_year)
                if not anomalies_year.empty:
                    anomalies_list.append(anomalies_year)
        finally:
            if client:
                client.close()
            if cluster:
                cluster.close()
    else:
        # 非 distributed client 情况：使用 dask.diagnostics.ProgressBar
        try:
            with ProgressBar():
                results = compute(*delayed_tasks, scheduler=sched)
        except Exception:
            results = compute(*delayed_tasks, scheduler=sched)
        for baseline_year, anomalies_year in results:
            if not baseline_year.empty:
                baselines_list.append(baseline_year)
            if not anomalies_year.empty:
                anomalies_list.append(anomalies_year)

    if not baselines_list and not anomalies_list:
        print("No data loaded after filtering; abort.")
        return None

    baseline_profiles = pd.concat(baselines_list, ignore_index=True) if baselines_list else pd.DataFrame()
    anomalies = pd.concat(anomalies_list, ignore_index=True) if anomalies_list else pd.DataFrame()

    # 基线与异常已在加载阶段完成区域过滤，这里只再检查是否为空
    if baseline_profiles.empty:
        print("No baseline profiles after filtering.")
    if anomalies.empty:
        print("No anomalies detected.")

    glorys_heave_available = False
    glorys_summary_path_used = None
    if use_glorys_heave and not anomalies.empty:
        region_slug_for_path = _current_region_key()
        glorys_summary_candidates: list[Path] = []
        if argo_glorys_summary_data_path is not None:
            glorys_summary_candidates.append(Path(argo_glorys_summary_data_path))
        glorys_summary_candidates.append(
            cfg.output_dir("plot_hotspot_anomaly_argo_glorys_overviews", region_slug_for_path)
            / f"hotspot_anomaly_argo_glorys_overviews_summary_{start_year}_{end_year}_{run_tag}.parquet"
        )
        for path_obj in glorys_summary_candidates:
            if path_obj.exists():
                glorys_summary_path_used = path_obj
                break

        if glorys_summary_path_used is None:
            print(
                "[WARN] GLORYS heave summary parquet not found; "
                "falling back to nearshore/non-nearshore hotspot types."
            )
        else:
            try:
                glorys_summary_df = pd.read_parquet(glorys_summary_path_used)
                anomalies = _merge_hotspot_glorys_summary_fields(anomalies, glorys_summary_df)
                heave_vals = pd.to_numeric(anomalies.get('glorys_heave_zmin'), errors='coerce')
                matched_outcrop = int(heave_vals.notna().sum())
                if matched_outcrop > 0:
                    glorys_heave_available = True
                    print(
                        f"[*] Loaded GLORYS OI hotspot type fields from {glorys_summary_path_used} "
                        f"(matched {matched_outcrop}/{len(anomalies)} anomalies)."
                    )
                else:
                    print(
                        f"[WARN] GLORYS summary parquet has no matched heave/OI fields: {glorys_summary_path_used}; "
                        "falling back to nearshore/non-nearshore hotspot types."
                    )
            except Exception as exc:
                print(
                    f"[WARN] Failed to read GLORYS heave summary parquet {glorys_summary_path_used}: {exc}; "
                    "falling back to nearshore/non-nearshore hotspot types."
                )

    # 绘图
    crosses_dateline = bool(_REGION_CFG.get('crosses_dateline') and (lonmax < lonmin))
    central_lon = 180 if crosses_dateline else 0
    data_crs = ccrs.PlateCarree()
    map_crs = ccrs.PlateCarree(central_longitude=central_lon)

    # 准备数据分组
    anom_interacting = pd.DataFrame()
    anom_others = pd.DataFrame()
    scatter_kwargs = {}
    
    if not anomalies.empty:
        if fix_colorbar:
            cbar_lo, cbar_hi = cfg.resolved_cbar()
            scatter_kwargs.update(dict(vmin=cbar_lo, vmax=cbar_hi))

        anom_others = anomalies.copy()
        if use_interacting_argo and interacting_argo_ids:
            is_interacting = anomalies['Profile_number'].isin(interacting_argo_ids)
            anom_interacting = anomalies[is_interacting].copy()
            anom_others = anomalies[~is_interacting].copy()
            
            # 统计输出
            total_anom = len(anomalies)
            count_int = len(anom_interacting)
            pct = (count_int / total_anom * 100) if total_anom > 0 else 0.0
            print(f"[Plot Info] Anomalies: {total_anom}, Interacting: {count_int} ({pct:.1f}%)")

    # 定义绘图任务
    base_anomaly_label = cfg.threshold_label()
    title_threshold_label = cfg.threshold_label()
    plots_to_generate = []
    if split_plots and split_mode == 'hotspot_type':
        if glorys_heave_available:
            type_vals = (
                pd.to_numeric(anomalies['hotspot_type'], errors='coerce')
                if (not anomalies.empty and 'hotspot_type' in anomalies.columns)
                else pd.Series(dtype=float)
            )
            if not anomalies.empty:
                counts = {
                    type_id: int((type_vals == type_id).sum())
                    for type_id in (1, 2, 3)
                }
                untyped = int(type_vals.isna().sum())
                print(
                    "[Plot Info] Hotspot types: "
                    f"type1={counts[1]}, type2={counts[2]}, type3={counts[3]}, "
                    f"untyped={untyped}"
                )

            type_specs = [
                {
                    'type_id': 1,
                    'name': 'hotspot_type_1',
                    'title_extra': ' (Type 1: Ventilated)',
                    'file_suffix': '_type1',
                    'marker': 'o',
                    'edgecolor': _HOTSPOT_TYPE_COLORS['type_1'],
                    'label': f'Type 1 - ventilated ({base_anomaly_label})',
                    's': 60,
                    'zorder': 3,
                },
                {
                    'type_id': 2,
                    'name': 'hotspot_type_2',
                    'title_extra': ' (Type 2: Deep only)',
                    'file_suffix': '_type2',
                    'marker': 'o',
                    'edgecolor': _HOTSPOT_TYPE_COLORS['type_2'],
                    'label': f'Type 2 - deep only ({base_anomaly_label})',
                    's': 60,
                    'zorder': 3,
                },
                {
                    'type_id': 3,
                    'name': 'hotspot_type_3',
                    'title_extra': ' (Type 3: Nearshore DO Dip)',
                    'file_suffix': '_type3',
                    'marker': '^',
                    'edgecolor': _HOTSPOT_TYPE_COLORS['type_3'],
                    'label': f'Type 3 - nearshore DO dip ({base_anomaly_label})',
                    's': 80,
                    'zorder': 4,
                },
            ]
            for spec in type_specs:
                type_data = (
                    anomalies.loc[type_vals == spec['type_id']].copy()
                    if not anomalies.empty else pd.DataFrame()
                )
                plots_to_generate.append({
                    'name': spec['name'],
                    'title_extra': spec['title_extra'],
                    'file_suffix': spec['file_suffix'],
                    'data_list': [
                        {
                            'data': type_data,
                            'marker': spec['marker'],
                            'edgecolor': spec['edgecolor'],
                            'label': spec['label'],
                            's': spec['s'],
                            'zorder': spec['zorder'],
                        }
                    ],
                })
        else:
            nearshore_vals = (
                anomalies['nearshore_do_dip'].astype('boolean').fillna(False)
                if (not anomalies.empty and 'nearshore_do_dip' in anomalies.columns)
                else pd.Series(False, index=anomalies.index, dtype=bool)
            )
            nearshore_count = int(nearshore_vals.sum()) if not anomalies.empty else 0
            non_nearshore_count = int((~nearshore_vals).sum()) if not anomalies.empty else 0
            if not anomalies.empty:
                print(
                    "[Plot Info] Hotspot type groups: "
                    f"non_nearshore={non_nearshore_count}, type3_nearshore={nearshore_count}"
                )
            hotspot_type_split_specs = [
                {
                    'data': anomalies.loc[~nearshore_vals].copy() if not anomalies.empty else pd.DataFrame(),
                    'name': 'non_nearshore',
                    'title_extra': ' (Non-nearshore)',
                    'file_suffix': '_non_nearshore',
                    'marker': 'o',
                    'edgecolor': _HOTSPOT_TYPE_COLORS['type_1'],
                    'label': f'Non-nearshore ({base_anomaly_label})',
                    's': 60,
                    'zorder': 3,
                },
                {
                    'data': anomalies.loc[nearshore_vals].copy() if not anomalies.empty else pd.DataFrame(),
                    'name': 'hotspot_type_3',
                    'title_extra': ' (Type 3: Nearshore DO Dip)',
                    'file_suffix': '_type3',
                    'marker': '^',
                    'edgecolor': _HOTSPOT_TYPE_COLORS['type_3'],
                    'label': f'Type 3 - nearshore DO dip ({base_anomaly_label})',
                    's': 80,
                    'zorder': 4,
                },
            ]
            for spec in hotspot_type_split_specs:
                plots_to_generate.append({
                    'name': spec['name'],
                    'title_extra': spec['title_extra'],
                    'file_suffix': spec['file_suffix'],
                    'data_list': [
                        {
                            'data': spec['data'],
                            'marker': spec['marker'],
                            'edgecolor': spec['edgecolor'],
                            'label': spec['label'],
                            's': spec['s'],
                            'zorder': spec['zorder'],
                        }
                    ],
                })
    elif split_plots and split_mode == 'spice_type':
        if (glorys_heave_available and not anomalies.empty
                and 'spice_type' in anomalies.columns):
            spice_vals = pd.to_numeric(anomalies['spice_type'], errors='coerce')
            counts = {s: int((spice_vals == s).sum()) for s in (1, 2, 3)}
            untyped = int(spice_vals.isna().sum())
            print(
                "[Plot Info] Spice type counts: "
                f"cold-fresh={counts[1]}, background={counts[2]}, "
                f"warm-salty={counts[3]}, untyped={untyped}"
            )

            spice_specs = [
                {
                    'type_id': 1,
                    'name': 'spice_cold_fresh',
                    'title_extra': ' (Cold-Fresh)',
                    'file_suffix': '_spice_cold_fresh',
                    'marker': 'o', 'edgecolor': _HOTSPOT_SPICE_COLORS['cold_fresh'],
                    'label': f'Cold-Fresh ({base_anomaly_label})',
                    's': 60, 'zorder': 3,
                },
                {
                    'type_id': 2,
                    'name': 'spice_background',
                    'title_extra': ' (Background-Consistent)',
                    'file_suffix': '_spice_background',
                    'marker': 's', 'edgecolor': _HOTSPOT_SPICE_COLORS['background'],
                    'label': f'Background-Consistent ({base_anomaly_label})',
                    's': 60, 'zorder': 3,
                },
                {
                    'type_id': 3,
                    'name': 'spice_warm_salty',
                    'title_extra': ' (Warm-Salty)',
                    'file_suffix': '_spice_warm_salty',
                    'marker': '^', 'edgecolor': _HOTSPOT_SPICE_COLORS['warm_salty'],
                    'label': f'Warm-Salty ({base_anomaly_label})',
                    's': 80, 'zorder': 4,
                },
            ]
            for spec in spice_specs:
                spice_data = anomalies.loc[spice_vals == spec['type_id']].copy()
                plots_to_generate.append({
                    'name': spec['name'],
                    'title_extra': spec['title_extra'],
                    'file_suffix': spec['file_suffix'],
                    'data_list': [
                        {
                            'data': spice_data,
                            'marker': spec['marker'],
                            'edgecolor': spec['edgecolor'],
                            'label': spec['label'],
                            's': spec['s'],
                            'zorder': spec['zorder'],
                        }
                    ],
                })
        else:
            print(
                "[WARN] spice_type mode requires GLORYS overview parquet with matched profiles "
                "and spice_type column; falling back to combined plot."
            )
            combined_data = []
            if not anomalies.empty:
                combined_data.append({
                    'data': anom_others if not anom_others.empty else anomalies.copy(),
                    'marker': 'o', 'edgecolor': 'black',
                    'label': f'{base_anomaly_label}',
                    's': 60, 'zorder': 3,
                })
            plots_to_generate.append({
                'name': 'combined',
                'title_extra': '',
                'file_suffix': '',
                'data_list': combined_data,
            })

    elif split_plots and split_mode == 'cross':
        if (glorys_heave_available and not anomalies.empty
                and 'hotspot_type' in anomalies.columns
                and 'spice_type' in anomalies.columns):
            type_vals = pd.to_numeric(anomalies['hotspot_type'], errors='coerce')
            spice_vals = pd.to_numeric(anomalies['spice_type'], errors='coerce')

            counts = {}
            for t in (1, 2):
                for s in (1, 2):
                    counts[(t, s)] = int(((type_vals == t) & (spice_vals == s)).sum())
            count_t3 = int((type_vals == 3).sum())
            untyped = int((~type_vals.isin([1, 2, 3])).sum())
            nospice = int(((type_vals.isin([1, 2])) & spice_vals.isna()).sum())
            print(
                "[Plot Info] Cross counts: "
                f"T1×cold-fresh={counts[(1, 1)]}, T1×bg={counts[(1, 2)]}, "
                f"T2×cold-fresh={counts[(2, 1)]}, T2×bg={counts[(2, 2)]}, "
                f"T3={count_t3}, untyped={untyped}, no-spice={nospice}"
            )

            cross_specs = [
                {
                    'type_id': 1, 'spice_id': 1,
                    'name': 'cross_T1_cold_fresh',
                    'title_extra': ' (Ventilated × Cold-Fresh)',
                    'file_suffix': '_T1_cold_fresh',
                    'marker': 'o', 'edgecolor': _HOTSPOT_CROSS_COLORS['T1_cold_fresh'],
                    'label': f'T1 Ventilated × Cold-Fresh ({base_anomaly_label})',
                    's': 60, 'zorder': 3,
                },
                {
                    'type_id': 1, 'spice_id': 2,
                    'name': 'cross_T1_background',
                    'title_extra': ' (Ventilated × Background)',
                    'file_suffix': '_T1_background',
                    'marker': 'o', 'edgecolor': _HOTSPOT_CROSS_COLORS['T1_background'],
                    'label': f'T1 Ventilated × Background ({base_anomaly_label})',
                    's': 60, 'zorder': 3,
                },
                {
                    'type_id': 2, 'spice_id': 1,
                    'name': 'cross_T2_cold_fresh',
                    'title_extra': ' (Isolated × Cold-Fresh)',
                    'file_suffix': '_T2_cold_fresh',
                    'marker': 's', 'edgecolor': _HOTSPOT_CROSS_COLORS['T2_cold_fresh'],
                    'label': f'T2 Isolated × Cold-Fresh ({base_anomaly_label})',
                    's': 60, 'zorder': 3,
                },
                {
                    'type_id': 2, 'spice_id': 2,
                    'name': 'cross_T2_background',
                    'title_extra': ' (Isolated × Background)',
                    'file_suffix': '_T2_background',
                    'marker': 's', 'edgecolor': _HOTSPOT_CROSS_COLORS['T2_background'],
                    'label': f'T2 Isolated × Background ({base_anomaly_label})',
                    's': 60, 'zorder': 3,
                },
                {
                    'type_id': 3, 'spice_id': None,
                    'name': 'cross_T3_OMZ',
                    'title_extra': ' (Type 3 OMZ)',
                    'file_suffix': '_T3_OMZ',
                    'marker': '^', 'edgecolor': _HOTSPOT_CROSS_COLORS['T3_OMZ'],
                    'label': f'T3 OMZ - nearshore DO dip ({base_anomaly_label})',
                    's': 80, 'zorder': 4,
                },
            ]
            for spec in cross_specs:
                if spec['spice_id'] is not None:
                    cross_data = anomalies.loc[
                        (type_vals == spec['type_id']) & (spice_vals == spec['spice_id'])
                    ].copy()
                else:
                    cross_data = anomalies.loc[type_vals == spec['type_id']].copy()
                plots_to_generate.append({
                    'name': spec['name'],
                    'title_extra': spec['title_extra'],
                    'file_suffix': spec['file_suffix'],
                    'data_list': [
                        {
                            'data': cross_data,
                            'marker': spec['marker'],
                            'edgecolor': spec['edgecolor'],
                            'label': spec['label'],
                            's': spec['s'],
                            'zorder': spec['zorder'],
                        }
                    ],
                })
        else:
            raise ValueError(
                "split_plots='cross' 需要 GLORYS overview parquet 含 spice_type "
                "且有匹配的 heave 记录，请先运行 plot_hotspot_anomaly_argo_glorys_overviews。"
            )

    elif split_plots and split_mode == 'eddy_interaction' and use_interacting_argo:
        plots_to_generate.append({
            'name': 'interacting',
            'title_extra': ' (Interacting)',
            'file_suffix': '_interacting',
            'data_list': [
                {'data': anom_interacting, 'marker': 'D', 'edgecolor': 'blue', 'label': f'Interacting ({base_anomaly_label})', 's': 100, 'zorder': 4}
            ]
        })
        plots_to_generate.append({
            'name': 'non_interacting',
            'title_extra': ' (Non-interacting)',
            'file_suffix': '_non_interacting',
            'data_list': [
                {'data': anom_others, 'marker': 'o', 'edgecolor': 'black', 'label': f'Non-interacting ({base_anomaly_label})', 's': 60, 'zorder': 3}
            ]
        })
    else:
        if split_plots and split_mode == 'eddy_interaction' and not use_interacting_argo:
            print("[WARN] split_plots='eddy_interaction' requires use_interacting_argo=True; drawing combined plot.")
        # 合并模式
        combined_data = []
        use_hotspot_type_combined = (
            not anomalies.empty
            and 'nearshore_do_dip' in anomalies.columns
            and (use_glorys_heave or not use_interacting_argo)
        )
        if use_hotspot_type_combined and glorys_heave_available:
            type_vals = pd.to_numeric(anomalies.get('hotspot_type'), errors='coerce')
            hotspot_type_combined_specs = [
                (1, 'o', _HOTSPOT_TYPE_COLORS['type_1'], f'Type 1 - ventilated ({base_anomaly_label})', 60, 3),
                (2, 'o', _HOTSPOT_TYPE_COLORS['type_2'], f'Type 2 - deep only ({base_anomaly_label})', 60, 3),
                (3, '^', _HOTSPOT_TYPE_COLORS['type_3'], f'Type 3 - nearshore DO dip ({base_anomaly_label})', 80, 4),
            ]
            for type_id, marker, edgecolor, label, size, zorder in hotspot_type_combined_specs:
                type_data = anomalies.loc[type_vals == type_id].copy()
                if not type_data.empty:
                    combined_data.append({
                        'data': type_data,
                        'marker': marker,
                        'edgecolor': edgecolor,
                        'label': label,
                        's': size,
                        'zorder': zorder,
                    })
        elif use_hotspot_type_combined:
            nearshore_vals = anomalies['nearshore_do_dip'].astype('boolean').fillna(False)
            non_nearshore = anomalies.loc[~nearshore_vals].copy()
            nearshore = anomalies.loc[nearshore_vals].copy()
            if not non_nearshore.empty:
                combined_data.append({
                    'data': non_nearshore,
                    'marker': 'o',
                    'edgecolor': _HOTSPOT_TYPE_COLORS['type_1'],
                    'label': f'Non-nearshore ({base_anomaly_label})',
                    's': 60,
                    'zorder': 3,
                })
            if not nearshore.empty:
                combined_data.append({
                    'data': nearshore,
                    'marker': '^',
                    'edgecolor': _HOTSPOT_TYPE_COLORS['type_3'],
                    'label': f'Type 3 - nearshore DO dip ({base_anomaly_label})',
                    's': 80,
                    'zorder': 4,
                })
        else:
            if not anom_others.empty:
                label_str = base_anomaly_label
                if use_interacting_argo and interacting_argo_ids:
                    label_str = f'Non-interacting ({base_anomaly_label})'
                combined_data.append({'data': anom_others, 'marker': 'o', 'edgecolor': 'black', 'label': label_str, 's': 60, 'zorder': 3})
            
            if not anom_interacting.empty:
                combined_data.append({'data': anom_interacting, 'marker': 'D', 'edgecolor': 'blue', 'label': f'Interacting ({base_anomaly_label})', 's': 100, 'zorder': 4})
            
        plots_to_generate.append({
            'name': 'combined',
            'title_extra': '',
            'file_suffix': '',
            'data_list': combined_data
        })

    saved_figure_paths: list[str] = []

    for p_cfg in plots_to_generate:
        fig = plt.figure(figsize=(40, 30))
        ax = fig.add_subplot(1, 1, 1, projection=map_crs)
        
        depth_title = (
            f' (depth ≥ {cfg.anomaly_min_depth} m)'
            if cfg.anomaly_min_depth is not None and cfg.anomaly_min_depth > 0 else ''
        )
        thr_title = f' ({title_threshold_label})'
        ax.set_title(f'Argo anomalies {start_year}-{end_year} (method={cfg.method}){thr_title}{depth_title}{p_cfg["title_extra"]}', fontsize=20)

        # Basemap features
        base_ocean = _BASEMAP_COLORS['ocean']
        base_land = _BASEMAP_COLORS['land']
        coast_color = _BASEMAP_COLORS['coastline']
        grid_color = _BASEMAP_COLORS['grid']
        ax.set_facecolor(base_ocean)
        ax.add_feature(cfeature.OCEAN, facecolor=base_ocean, zorder=0)
        ax.add_feature(cfeature.LAND, facecolor=base_land, edgecolor=coast_color, linewidth=0.5, zorder=0)
        ax.add_feature(cfeature.COASTLINE, linewidth=0.7, edgecolor=coast_color, zorder=1)
        gl = ax.gridlines(draw_labels=True, linewidth=0.4, color=grid_color, alpha=0.45, linestyle='--')
        gl.top_labels = False
        gl.right_labels = False

        # 设定范围（处理跨日界线）
        lon_extent_min = lonmin
        lon_extent_max = lonmax
        if crosses_dateline and lon_extent_max < lon_extent_min:
            lon_extent_max += 360
        ax.set_extent([lon_extent_min, lon_extent_max, latmin, latmax], crs=data_crs)

        if plot_unrelated_argo and not baseline_profiles.empty:
            ax.scatter(
                baseline_profiles['Longitude'], baseline_profiles['Latitude'],
                facecolors='none', edgecolors='gray', linewidths=0.7, s=25,
                label='All Argo Profiles (baseline)', zorder=2, transform=data_crs
            )

        sc = None
        has_anom_plot = False
        for d_cfg in p_cfg['data_list']:
            data = d_cfg['data']
            if data.empty: continue
            has_anom_plot = True
            color_values, _, color_label, cmap_name = _color_values_for_anomalies(data, cfg)
            if color_values is None:
                color_values = pd.Series(np.arange(len(data)), index=data.index)
            
            sc_curr = ax.scatter(
                data['Longitude'], data['Latitude'],
                c=color_values, cmap=cmap_name, s=d_cfg['s'],
                marker=d_cfg['marker'],
                edgecolors=d_cfg['edgecolor'], linewidths=0.5 if d_cfg['marker']=='o' else 1.0,
                label=d_cfg['label'], zorder=d_cfg['zorder'],
                transform=data_crs,
                **scatter_kwargs
            )
            sc = sc_curr

        if not has_anom_plot:
            empty_msg = 'No anomalies' if anomalies.empty else 'No anomalies for this split'
            ax.text(
                0.5,
                0.5,
                empty_msg,
                transform=ax.transAxes,
                ha='center',
                va='center',
                fontsize=24,
                color='red',
            )

        if sc is not None:
            cbar = plt.colorbar(sc, ax=ax, orientation='horizontal', fraction=0.046, pad=0.05)
            cbar.set_label(color_label, fontsize=20); cbar.ax.tick_params(labelsize=14)
            if fix_colorbar:
                _apply_detection_colorbar_ticks(cbar, cfg, cbar_lo, cbar_hi)

        ax.legend(fontsize=18, loc='upper left')

        if save_fig:
            region_slug_for_path = _current_region_key()
            out_dir = cfg.output_dir("plot_argo_hotspots", region_slug_for_path)
            out_dir.mkdir(exist_ok=True, parents=True)
            fname = out_dir / f"Argo_Anomaly_Hotspots_{start_year}_{end_year}_{run_tag}{p_cfg['file_suffix']}.png"
            plt.savefig(fname, dpi=300, bbox_inches='tight')
            saved_figure_paths.append(str(fname))
            print(f"Figure saved: {fname}")
        if show_fig:
            plt.show()
        plt.close(fig)

    # 保存 anomalies 为 Parquet（高效压缩存储）
    saved_anomalies_path: str | None = None
    if save_data:
        region_slug_for_path = _current_region_key()
        out_dir = cfg.output_dir("plot_argo_hotspots", region_slug_for_path)
        out_dir.mkdir(exist_ok=True, parents=True)
        pq_path = out_dir / f"anomalies_{start_year}_{end_year}_{run_tag}.parquet"
        try:
            anomalies_out = anomalies.copy()
            preferred_cols = _hotspot_anomaly_output_columns()
            if anomalies_out.empty:
                anomalies_out = pd.DataFrame(columns=preferred_cols)
            else:
                ordered_cols = [c for c in preferred_cols if c in anomalies_out.columns]
                extra_cols = [c for c in anomalies_out.columns if c not in ordered_cols]
                anomalies_out = anomalies_out[ordered_cols + extra_cols]
            anomalies_out.to_parquet(pq_path, index=False)
            saved_anomalies_path = str(pq_path)
            print(f"Anomalies saved to: {pq_path}")
        except Exception as e:
            print(f"[WARN] Failed to save anomalies parquet: {e}")

    total_argo_profiles = int(len(baseline_profiles))
    selected_argo_profiles = int(len(anomalies))
    selected_ratio = (
        float(selected_argo_profiles) / float(total_argo_profiles)
        if total_argo_profiles > 0 else np.nan
    )

    selected_delta_do_max = np.nan
    selected_delta_do_mean = np.nan
    selected_depth_max = np.nan
    selected_depth_mean = np.nan
    if not anomalies.empty:
        if 'delta_do' in anomalies.columns:
            delta_vals = pd.to_numeric(anomalies['delta_do'], errors='coerce').to_numpy(dtype=float)
            valid_delta = np.isfinite(delta_vals)
            if valid_delta.any():
                selected_delta_do_max = float(np.nanmax(delta_vals[valid_delta]))
                selected_delta_do_mean = float(np.nanmean(delta_vals[valid_delta]))
        if 'depth' in anomalies.columns:
            depth_vals = pd.to_numeric(anomalies['depth'], errors='coerce').to_numpy(dtype=float)
            valid_depth = np.isfinite(depth_vals)
            if valid_depth.any():
                selected_depth_max = float(np.nanmax(depth_vals[valid_depth]))
                selected_depth_mean = float(np.nanmean(depth_vals[valid_depth]))

    summary = {
        'total_argo_profiles': total_argo_profiles,
        'selected_argo_profiles': selected_argo_profiles,
        'selected_ratio': float(selected_ratio) if np.isfinite(selected_ratio) else np.nan,
        'selected_delta_do_max': selected_delta_do_max,
        'selected_delta_do_mean': selected_delta_do_mean,
        'selected_depth_max': selected_depth_max,
        'selected_depth_mean': selected_depth_mean,
        'all_argo_depth_mean': np.nan,
        'all_argo_do_mean': np.nan,
        'use_glorys_heave': bool(use_glorys_heave),
        'glorys_heave_available': bool(glorys_heave_available),
        'argo_glorys_summary_data_path': str(glorys_summary_path_used) if glorys_summary_path_used is not None else None,
    }

    return {
        'summary': summary,
        'figure_paths': saved_figure_paths,
        'anomalies_path': saved_anomalies_path,
    }


def _hotspot_profile_to_int_or_none(val):
    try:
        if pd.isna(val):
            return None
        return int(val)
    except Exception:
        return None


def _plot_hotspot_vertical_profile_from_record(
    row: dict,
    df_year: pd.DataFrame,
    *,
    cfg: DetectionConfig,
    run_tag: str,
    plot_variables: list,
    overlay_aou_on_do: bool,
    remove_outliers: bool,
    plot_normal_scatter: bool,
    annotate_delta_ts: bool,
    save_fig: bool,
    show_fig: bool,
    output_dir: str | Path,
) -> dict:
    """绘制单个 hotspot anomaly 的 Argo 垂向剖面。"""
    fig = None
    year_val = int(row['_year'])
    profile_num = int(row['_profile'])
    result = {
        'year': year_val,
        'profile_number': profile_num,
        'plotted': False,
        'skipped': True,
        'save_path': None,
        'warning': None,
    }

    try:
        if df_year.empty:
            result['warning'] = f"Empty Argo yearly data for {year_val}, skip profile {profile_num}."
            return result

        profile_rows = df_year[df_year['Profile_number'] == profile_num].copy()
        if profile_rows.empty:
            result['warning'] = f"No raw profile found for Year={year_val}, Profile_number={profile_num}."
            return result

        month_val = _hotspot_profile_to_int_or_none(row.get('Month'))
        day_val = _hotspot_profile_to_int_or_none(row.get('Day'))
        if month_val is not None and day_val is not None:
            day_rows = profile_rows[
                (pd.to_numeric(profile_rows['Month'], errors='coerce') == month_val)
                & (pd.to_numeric(profile_rows['Day'], errors='coerce') == day_val)
            ].copy()
            if not day_rows.empty:
                profile_rows = day_rows

        platform_val = None
        row_platform_val = _hotspot_profile_to_int_or_none(row.get('Platform_number'))
        if row_platform_val is not None and 'Platform_number' in profile_rows.columns:
            platform_rows = profile_rows[
                pd.to_numeric(profile_rows['Platform_number'], errors='coerce') == row_platform_val
            ].copy()
            if not platform_rows.empty:
                profile_rows = platform_rows
                platform_val = row_platform_val

        if 'Platform_number' in profile_rows.columns:
            platforms = pd.to_numeric(profile_rows['Platform_number'], errors='coerce').dropna().astype(int).unique()
            if platforms.size > 0:
                platform_val = platform_val if platform_val is not None else int(platforms[0])
                if platforms.size > 1:
                    result['warning'] = (
                        f"Multiple platforms found for Year={year_val}, Profile={profile_num}; "
                        f"using Platform_number={platform_val}."
                    )
                    profile_rows = profile_rows[
                        pd.to_numeric(profile_rows['Platform_number'], errors='coerce') == platform_val
                    ].copy()

        if profile_rows.empty:
            return result

        profile_rows = profile_rows.sort_values('Depth').copy()

        annotation_text = ""
        depth_text = ""
        if annotate_delta_ts:
            annotation_text, depth_text = _annotation_text_from_anomaly_record(row, cfg)

        num_variables = len(plot_variables)
        fig, axes = plt.subplots(1, num_variables, figsize=(10 * num_variables, 20), sharey=True)
        if num_variables == 1:
            axes = [axes]

        line_color = plt.cm.coolwarm(0.15)
        any_plotted = False

        for var_name, ax in zip(plot_variables, axes):
            plot_variable_name = var_name
            is_do_panel = (_map_plot_variable_name(plot_variable_name) == 'DO')
            do_aux_layers = ('aou',) if (is_do_panel and overlay_aou_on_do) else tuple()

            db_variable_name = _map_plot_variable_name(plot_variable_name)
            plot_line_color = '#ff8c00' if db_variable_name == 'AOU' else line_color

            if not _has_plottable_profile_variable(profile_rows, db_variable_name):
                ax.text(
                    0.5,
                    0.5,
                    f"Variable '{db_variable_name}'\nnot found in data.",
                    ha='center',
                    va='center',
                    transform=ax.transAxes,
                    fontsize=16,
                )
                _apply_vertical_profile_axis_style(ax, var_name)
                continue

            did_plot = _plot_single_argo_profile_line(
                ax,
                profile_rows,
                plot_variable_name,
                plot_line_color,
                remove_outliers=remove_outliers,
                show_normal_scatter=plot_normal_scatter,
                do_aux_layers=do_aux_layers,
                aou_aux_color='#ff8c00',
                alpha=0.9,
            )
            if not did_plot:
                ax.text(
                    0.5,
                    0.5,
                    "No valid data after QC.",
                    ha='center',
                    va='center',
                    transform=ax.transAxes,
                    fontsize=14,
                )
            else:
                any_plotted = True

            _apply_vertical_profile_axis_style(ax, plot_variable_name)

        axes[0].set_ylabel("Depth/m", fontsize=20)
        axes[0].tick_params(axis='y', labelsize=16)
        axes[0].invert_yaxis()

        row_month = _hotspot_profile_to_int_or_none(row.get('Month'))
        row_day = _hotspot_profile_to_int_or_none(row.get('Day'))
        if row_month is not None and row_day is not None:
            date_text = f"{year_val:04d}-{row_month:02d}-{row_day:02d}"
        else:
            first_row = profile_rows.iloc[0]
            first_year = _hotspot_profile_to_int_or_none(first_row.get('Year'))
            first_month = _hotspot_profile_to_int_or_none(first_row.get('Month'))
            first_day = _hotspot_profile_to_int_or_none(first_row.get('Day'))
            if first_year is not None and first_month is not None and first_day is not None:
                date_text = f"{first_year:04d}-{first_month:02d}-{first_day:02d}"
            else:
                date_text = str(year_val)

        platform_text = f", Platform={platform_val}" if platform_val is not None else ""
        lon_text = ""
        lat_text = ""
        lon_val = pd.to_numeric(pd.Series([row.get('Longitude')]), errors='coerce').iloc[0]
        lat_val = pd.to_numeric(pd.Series([row.get('Latitude')]), errors='coerce').iloc[0]
        if not np.isfinite(lon_val) or not np.isfinite(lat_val):
            first_row = profile_rows.iloc[0]
            lon_val = pd.to_numeric(pd.Series([first_row.get('Longitude')]), errors='coerce').iloc[0]
            lat_val = pd.to_numeric(pd.Series([first_row.get('Latitude')]), errors='coerce').iloc[0]
        if np.isfinite(lon_val):
            lon_text = f"Lon={float(_normalize_lon_array(lon_val)):.3f}"
        if np.isfinite(lat_val):
            lat_text = f"Lat={float(lat_val):.3f}"
        location_text = ", ".join([t for t in [lon_text, lat_text] if t])

        title_line1 = (
            f"Hotspots Profile {profile_num}{platform_text}, "
            f"{date_text}{annotation_text}{depth_text}"
        )
        title_line2 = location_text if location_text else "Lon/Lat unavailable"
        fig.suptitle(
            f"{title_line1}\n{title_line2}",
            fontsize=24,
            y=0.97,
        )
        plt.tight_layout(rect=[0, 0, 1, 0.90])

        if save_fig and any_plotted:
            output_dir = Path(output_dir)
            output_dir.mkdir(parents=True, exist_ok=True)
            file_date = date_text.replace('-', '')
            platform_suffix = f"_platform{platform_val}" if platform_val is not None else ""
            save_path = output_dir / f"hotspot_{file_date}_P{profile_num}{platform_suffix}_{run_tag}.png"
            plt.savefig(save_path, dpi=300, bbox_inches='tight')
            result['save_path'] = str(save_path)

        if show_fig:
            plt.show()

        result['plotted'] = bool(any_plotted)
        result['skipped'] = not bool(any_plotted)
        return result
    except Exception as exc:
        result['warning'] = f"Failed plotting Year={year_val}, Profile={profile_num}: {exc}"
        return result
    finally:
        if fig is not None:
            plt.close(fig)


def _plot_hotspot_vertical_profiles_year_worker(args: dict) -> dict:
    """按年份绘制 hotspot Argo 垂向剖面的 multiprocessing worker。"""
    year_val = int(args['year'])
    records = list(args.get('records') or [])
    result = {
        'year': year_val,
        'total': len(records),
        'plotted': 0,
        'skipped': 0,
        'save_paths': [],
        'warnings': [],
    }
    if bool(args.get('force_agg_backend', False)):
        try:
            plt.switch_backend('Agg')
        except Exception:
            pass
    if not bool(args.get('show_fig', False)):
        plt.ioff()

    try:
        df_year = load_argo_data(year_val, data_dir=args.get('argo_data_dir'))
    except FileNotFoundError:
        result['skipped'] = len(records)
        result['warnings'].append(f"Missing Argo yearly file for {year_val}, skip {len(records)} profiles.")
        return result
    except Exception as exc:
        result['skipped'] = len(records)
        result['warnings'].append(f"Failed loading Argo {year_val}: {exc}")
        return result

    for row in records:
        item = _plot_hotspot_vertical_profile_from_record(
            row,
            df_year,
            cfg=args['detection_config'],
            run_tag=args['run_tag'],
            plot_variables=args['plot_variables'],
            overlay_aou_on_do=bool(args.get('overlay_aou_on_do', False)),
            remove_outliers=bool(args.get('remove_outliers', True)),
            plot_normal_scatter=bool(args.get('plot_normal_scatter', True)),
            annotate_delta_ts=bool(args.get('annotate_delta_ts', False)),
            save_fig=bool(args.get('save_fig', True)),
            show_fig=bool(args.get('show_fig', False)),
            output_dir=args['output_dir'],
        )
        if item.get('plotted'):
            result['plotted'] += 1
        else:
            result['skipped'] += 1
        if item.get('save_path'):
            result['save_paths'].append(item['save_path'])
        if item.get('warning'):
            result['warnings'].append(item['warning'])

    plt.close('all')
    return result


def plot_hotspot_anomaly_vertical_profiles(
    start_year: int | None = None,
    end_year: int | None = None,
    anomalies_path: str | Path | None = None,
    detection_config: DetectionConfig | None = None,
    variables: list = ['DO', 'AOU', 'Temp', 'Salinity'],
    remove_outliers: bool = True,
    plot_normal_scatter: bool = True,
    annotate_delta_ts: bool = True,
    save_fig: bool = True,
    show_fig: bool = False,
    clear_output_dir: bool = True,
    max_profiles: int | None = None,
    argo_data_dir: str | Path | None = None,
    use_multiprocessing: bool = True,
    num_workers: int | None = None,
) -> dict:
    """基于 hotspots 异常文件批量绘制 Argo 垂向剖面。

    参数:
        - start_year (int | None): anomalies_path=None 时用于定位默认 anomalies 文件。
        - end_year (int | None): 同上，结束年份。
        - anomalies_path (str | Path | None): 指定 anomalies parquet 路径；None 时按 plot_argo_hotspots 命名规则自动定位。
        - detection_config (DetectionConfig | None): 异常识别配置；用于自动定位模式子目录与 file_stem 文件名。
        - variables (list): 每幅图绘制的变量列表，默认 ['DO','AOU','Temp','Salinity']。
        - remove_outliers (bool): True 时按基础 QC 剔除异常值；False 时保留 QC 通过段为原色、断点用红线桥接并用红色圆点标记 QC 异常，默认 True。
        - plot_normal_scatter (bool): 是否绘制正常值的孤立散点标记，默认 True。
        - annotate_delta_ts (bool): 是否在标题标注当前异常方法的判别变量、辅助 ΔT/ΔS 及深度，默认 True。
        - save_fig (bool): 是否保存图像，默认 True。
        - show_fig (bool): 是否显示图像，默认 False。
        - clear_output_dir (bool): 保存图片时是否在本次运行开始前清空输出目录，默认 True。
        - max_profiles (int | None): 最多绘制多少个异常剖面；None 表示全部。
        - argo_data_dir (str | Path | None): Argo 年数据目录；None 时使用配置默认路径。
        - use_multiprocessing (bool): 是否按年份并行绘图，默认 True；show_fig=True 时自动退回串行。
        - num_workers (int | None): 并行 worker 数；None 时自动取 min(年份数, CPU数, 8)。

    返回:
        - dict: 含 total_candidates/plotted_profiles/skipped_profiles/output_dir/anomalies_path。

    说明:
        - 输入 anomalies parquet 仅含异常摘要，不含完整剖面；本函数按 Year+Profile_number 回查原始 Argo 年数据。
        - 绘图阶段复用 plot_vertical 的单剖面画线内核，可通过 remove_outliers 控制基础 QC。
    """

    cfg = _resolve_detection_config(detection_config)
    method_name = cfg.method
    run_tag = cfg.file_stem()
    if argo_data_dir is None:
        argo_data_dir = argo_path

    plot_variables, overlay_aou_on_do = _prepare_vertical_plot_variables(variables)

    if anomalies_path is None:
        if start_year is None or end_year is None:
            raise ValueError("anomalies_path 为空时，必须提供 start_year 与 end_year。")
        region_slug = _current_region_key()
        mode_path = cfg.output_dir("plot_argo_hotspots", region_slug) / f"anomalies_{start_year}_{end_year}_{run_tag}.parquet"
        anomalies_path = mode_path
    else:
        anomalies_path = Path(anomalies_path)

    if not Path(anomalies_path).exists():
        raise FileNotFoundError(f"Anomalies file not found: {anomalies_path}")

    anomalies = pd.read_parquet(anomalies_path)
    if 'detection_method' in anomalies.columns:
        method_mask = anomalies['detection_method'].astype(str).str.lower().eq(method_name)
        total_count = len(anomalies)
        mismatch_count = int((~method_mask).sum())
        if mismatch_count > 0:
            print(f"[WARN] Mixed detection_method found: expected={method_name}, mismatched={mismatch_count}/{total_count}.")
        else:
            print(f"[*] Method={method_name}, records={total_count}")
        anomalies = anomalies[method_mask].copy()
    if anomalies.empty:
        print(f"[*] No anomalies in file: {anomalies_path}")
        return {
            'total_candidates': 0,
            'plotted_profiles': 0,
            'skipped_profiles': 0,
            'output_dir': None,
            'anomalies_path': str(anomalies_path),
        }

    required_cols = ['Year', 'Profile_number']
    missing_cols = [c for c in required_cols if c not in anomalies.columns]
    if missing_cols:
        raise ValueError(f"Anomalies file missing required columns: {missing_cols}")

    work = anomalies.copy()
    work['_year'] = pd.to_numeric(work['Year'], errors='coerce')
    work['_profile'] = pd.to_numeric(work['Profile_number'], errors='coerce')
    work = work.dropna(subset=['_year', '_profile']).copy()
    work['_year'] = work['_year'].astype(int)
    work['_profile'] = work['_profile'].astype(int)

    if max_profiles is not None and max_profiles > 0:
        if cfg.score_col() in work.columns:
            work = work.sort_values(cfg.score_col(), ascending=False).head(int(max_profiles)).copy()
        elif 'delta_do' in work.columns:
            work = work.sort_values('delta_do', ascending=False).head(int(max_profiles)).copy()
        else:
            work = work.head(int(max_profiles)).copy()

    region_slug = _current_region_key()
    output_dir = cfg.output_dir("plot_argo_hotspots_vertical_profiles", region_slug)
    if save_fig:
        if clear_output_dir and output_dir.exists():
            try:
                shutil.rmtree(output_dir)
            except Exception as exc:
                print(f"[WARN] Failed to clear output directory {output_dir}: {exc}")
        output_dir.mkdir(parents=True, exist_ok=True)

    total_candidates = int(len(work))
    plotted_profiles = 0
    skipped_profiles = 0
    saved_figure_paths: list[str] = []
    warnings_list: list[str] = []
    active_num_workers = 1

    records_by_year = [
        (int(year_val), group.to_dict('records'))
        for year_val, group in work.groupby('_year', sort=True)
    ]

    if show_fig and use_multiprocessing:
        print("[WARN] show_fig=True is not compatible with multiprocessing; falling back to serial plotting.")
        use_multiprocessing = False

    worker_args = [
        {
            'year': year_val,
            'records': records,
            'detection_config': cfg,
            'run_tag': run_tag,
            'plot_variables': plot_variables,
            'overlay_aou_on_do': overlay_aou_on_do,
            'remove_outliers': remove_outliers,
            'plot_normal_scatter': plot_normal_scatter,
            'annotate_delta_ts': annotate_delta_ts,
            'save_fig': save_fig,
            'show_fig': show_fig,
            'output_dir': str(output_dir),
            'argo_data_dir': str(argo_data_dir),
            'force_agg_backend': bool(not show_fig),
        }
        for year_val, records in records_by_year
    ]

    def _consume_year_result(res: dict) -> None:
        nonlocal plotted_profiles, skipped_profiles, saved_figure_paths, warnings_list
        plotted_profiles += int(res.get('plotted', 0))
        skipped_profiles += int(res.get('skipped', 0))
        saved_figure_paths.extend([str(p) for p in res.get('save_paths', []) if p])
        warnings_list.extend([str(w) for w in res.get('warnings', []) if w])

    if use_multiprocessing and len(worker_args) > 1:
        worker_count = int(num_workers) if num_workers is not None else min(len(worker_args), os.cpu_count() or 1, 8)
        worker_count = max(1, min(worker_count, len(worker_args)))
        active_num_workers = worker_count
        print(
            f"[*] Hotspots vertical profile multiprocessing: "
            f"workers={worker_count}, years={len(worker_args)}, profiles={total_candidates}."
        )
        with multiprocessing.Pool(processes=worker_count, maxtasksperchild=1) as pool:
            progress = tqdm(total=total_candidates, desc="hotspot vertical profiles", unit="profile")
            try:
                for res in pool.imap_unordered(_plot_hotspot_vertical_profiles_year_worker, worker_args):
                    _consume_year_result(res)
                    progress.update(int(res.get('total', 0)))
            finally:
                progress.close()
    else:
        progress = tqdm(total=total_candidates, desc="hotspot vertical profiles", unit="profile")
        try:
            for args in worker_args:
                res = _plot_hotspot_vertical_profiles_year_worker(args)
                _consume_year_result(res)
                progress.update(int(res.get('total', 0)))
        finally:
            progress.close()

    for msg in warnings_list:
        print(f"[WARN] {msg}")

    print(
        f"[*] Hotspots profile plotting complete: "
        f"total={total_candidates}, plotted={plotted_profiles}, skipped={skipped_profiles}."
    )

    return {
        'total_candidates': total_candidates,
        'plotted_profiles': int(plotted_profiles),
        'skipped_profiles': int(skipped_profiles),
        'output_dir': str(output_dir) if save_fig else None,
        'anomalies_path': str(anomalies_path),
        'figure_paths': saved_figure_paths,
        'use_multiprocessing': bool(use_multiprocessing and len(worker_args) > 1),
        'num_workers': int(active_num_workers),
    }


def _plot_hotspot_argo_glorys_profile_worker(args: dict) -> dict:
    """单个 hotspot Argo profile 的 GLORYS 水平图与 vertical overview 绘图 worker。"""
    task_index = int(args.get('task_index', -1))
    year_val = int(args['year'])
    profile_num = int(args['profile_number'])
    month_val = args.get('month')
    day_val = args.get('day')
    platform_val = args.get('platform_number')
    heave_projection_depth_m = args.get('heave_projection_depth_m')

    if month_val is not None and day_val is not None:
        profile_time = pd.Timestamp(year=year_val, month=int(month_val), day=int(day_val))
    else:
        profile_time = year_val

    record = {
        'task_index': task_index,
        'year': year_val,
        'profile_number': profile_num,
        'platform_number': platform_val,
        'profile_time': str(profile_time),
        'line_strategy': 'zonal',
        'k': 0.0,
        'b': np.nan,
        'center_lon': np.nan,
        'center_lat': np.nan,
        'target_date': None,
        'anomaly_depth_m': heave_projection_depth_m,
        'nearshore_do_dip': args.get('nearshore_do_dip', pd.NA),
        'horizontal_status': 'not_started',
        'vertical_status': 'not_started',
        'vertical_save_paths': None,
        'status': 'failed',
        'error': None,
    }

    try:
        region_config_key = args.get('region_config_key')
        if region_config_key:
            switch_region(str(region_config_key), verbose=False)

        if bool(args.get('force_agg_backend', False)):
            try:
                plt.switch_backend('Agg')
            except Exception:
                pass
        if not bool(args.get('show_fig', False)):
            plt.ioff()
        plt.close('all')

        cfg = args['detection_config']
        info = _resolve_argo_profile_center(
            profile_number=profile_num,
            profile_time=profile_time,
            platform_number=platform_val,
            argo_data_dir=args.get('argo_data_dir'),
        )
        center_lon = float(info['center_lon'])
        center_lat = float(info['center_lat'])
        target_date = pd.Timestamp(info['target_date']).normalize()
        platform_resolved = info.get('platform_number')
        if platform_val is None and platform_resolved is not None:
            platform_val = int(platform_resolved)

        k_val = 0.0
        b_val = center_lat
        record.update({
            'platform_number': platform_val,
            'profile_time': target_date.strftime('%Y-%m-%d'),
            'b': float(b_val),
            'center_lon': center_lon,
            'center_lat': center_lat,
            'target_date': target_date.strftime('%Y-%m-%d'),
        })

        horiz_var = args.get('horizontal_variable', 'vorticity')
        horiz_depths = [0.0]  # always sea surface
        if bool(args.get('horizontal_argo_depth', False)):
            argo_d = args.get('heave_projection_depth_m')
            if argo_d is not None and np.isfinite(float(argo_d)) and float(argo_d) > 0:
                horiz_depths.append(float(argo_d))
        for hd in horiz_depths:
            plot_argo_horizontal_glorys(
                profile_number=profile_num,
                profile_time=target_date,
                platform_number=platform_val,
                variable=horiz_var,
                show_fig=bool(args.get('show_fig', False)),
                save_fig=bool(args.get('save_fig', True)),
                k=k_val,
                b=b_val,
                needed_depth=float(hd),
                inline_mode=bool(args.get('inline_mode', True)),
                xmin=args.get('xmin', -400.0),
                xmax=args.get('xmax', 400.0),
                argo_detection_config=cfg,
                argo_min_depth=args.get('argo_min_depth'),
                argo_data_dir=args.get('argo_data_dir'),
                output_dir=args.get('horizontal_output_dir'),
                verbose=bool(args.get('verbose', False)),
            )
        record['horizontal_status'] = 'ok'

        vertical_results = plot_argo_vertical_glorys_overview(
            profile_number=profile_num,
            profile_time=target_date,
            k=k_val,
            b=b_val,
            platform_number=platform_val,
            variables=args.get('vertical_variables'),
            needed_depth=args.get('needed_depth', 0),
            xmin=args.get('xmin', -400.0),
            xmax=args.get('xmax', 400.0),
            ymin=args.get('ymin', 0.0),
            ymax=args.get('ymax', 1000.0),
            profile_spacing_km=args.get('profile_spacing_km'),
            interpolate_z=bool(args.get('interpolate_z', True)),
            profile_depth_spacing_m=args.get('profile_depth_spacing_m'),
            plot_mlt=bool(args.get('plot_mlt', False)),
            plot_argo_projection=bool(args.get('plot_argo_projection', True)),
            argo_projection_config=cfg,
            argo_projection_min_depth=args.get('argo_projection_min_depth'),
            plot_isolines=bool(args.get('plot_isolines', True)),
            isoline_levels=args.get('isoline_levels'),
            isoline_color=args.get('isoline_color', 'black'),
            isoline_linewidth=float(args.get('isoline_linewidth', 0.8)),
            isoline_alpha=float(args.get('isoline_alpha', 0.45)),
            label_isolines=bool(args.get('label_isolines', False)),
            argo_data_dir=args.get('argo_data_dir'),
            show_fig=bool(args.get('show_fig', False)),
            save_fig=bool(args.get('save_fig', True)),
            output_dir=args.get('vertical_output_dir'),
            verbose=bool(args.get('verbose', False)),
            heave_projection_depth_m=heave_projection_depth_m,
            heave_x_window_km=float(args.get('heave_x_window_km', 25.0)),
            heave_z_window_m=args.get('heave_z_window_m', 100.0),
            heave_search_range=float(args.get('heave_search_range', 0.5)),
            heave_depth_threshold=float(args.get('heave_depth_threshold', 150.0)),
            heave_z_search_m=args.get('heave_z_search_m', 200.0),
            annotate_heave=bool(args.get('annotate_heave', False)),
            z_overview=bool(args.get('z_overview', True)),
            sigma_overview=bool(args.get('sigma_overview', False)),
            ts_diagram=bool(args.get('ts_diagram', False)),
            annotate_spice=bool(args.get('annotate_spice', True)),
            anomaly_sal=args.get('salinity_value'),
            anomaly_theta=args.get('temperature_value'),
        )
        record['vertical_status'] = 'ok' if vertical_results else 'empty'
        record['vertical_save_paths'] = ";".join(
            [str(item.get('save_path')) for item in vertical_results if item.get('save_path')]
        ) or None
        record['sigma_save_paths'] = ";".join(
            [str(item.get('sigma_save_path')) for item in vertical_results if item.get('sigma_save_path')]
        ) or None
        record['ts_save_paths'] = ";".join(
            [str(item.get('ts_save_path')) for item in vertical_results if item.get('ts_save_path')]
        ) or None
        if vertical_results:
            first_vertical = vertical_results[0]
            for key in [
                'projection_depth_m',
                'heave_x_window_km',
                'heave_z_window_m',
                'heave_valid_fraction',
                'glorys_heave_sigma_argo',
                'glorys_heave_sigma_peak',
                'glorys_heave_zmin',
                'glorys_heave_m',
                'heave_error',
                'spice_anomaly',
                'spice_percentile',
            ]:
                if key in first_vertical:
                    value = first_vertical[key]
                    record[key] = value
        record['status'] = 'ok' if record['vertical_status'] == 'ok' else 'partial'

    except Exception as exc:
        record['error'] = str(exc)
    finally:
        plt.close('all')

    return record


def plot_hotspot_anomaly_argo_glorys_overviews(
    start_year: int | None = None,
    end_year: int | None = None,
    anomalies_path: str | Path | None = None,
    detection_config: DetectionConfig | None = None,
    *,
    horizontal_variable: str = 'vorticity',
    horizontal_argo_depth: bool = False,
    vertical_variables: list[str] | None = None,
    needed_depth: float | int = 0,
    xmin: float = -400.0,
    xmax: float = 400.0,
    ymin: float = 0.0,
    ymax: float = 1000.0,
    profile_spacing_km: float | None = None,
    interpolate_z: bool = True,
    profile_depth_spacing_m: float | None = None,
    plot_mlt: bool = False,
    plot_argo_projection: bool = True,
    plot_isolines: bool = True,
    isoline_levels: int | list[float] | np.ndarray | None = None,
    isoline_color: str = 'black',
    isoline_linewidth: float = 0.8,
    isoline_alpha: float = 0.45,
    label_isolines: bool = True,
    argo_min_depth: float | None = None,
    argo_projection_min_depth: float | None = None,
    argo_data_dir: str | Path | None = None,
    output_dir: str | Path | None = None,
    clear_output_dir: bool = True,
    save_fig: bool = True,
    show_fig: bool = False,
    inline_mode: bool = True,
    verbose: bool = False,
    use_multiprocessing: bool = True,
    num_workers: int | None = None,
    maxtasksperchild: int | None = 4,
    heave_x_window_km: float = 25.0,
    heave_z_window_m: float | None = 100.0,
    heave_search_range: float = _heave_search_range,
    heave_depth_threshold: float = _heave_depth_threshold,
    heave_z_search_m: float | None = _heave_z_search_m,
    annotate_heave: bool = True,
    annotate_spice: bool = True,
    z_overview: bool = True,
    sigma_overview: bool = False,
    ts_diagram: bool = False,
    return_details: bool = False,
    save_summary_data: bool = True,
    summary_data_path: str | Path | None = None,
) -> dict:
    """为 hotspots 异常剖面批量绘制 Argo-centered GLORYS 水平图与垂向总览图。

    z_overview/sigma_overview 独立控制 z 坐标与 σ 坐标 2×2 总览图（默认 z_overview=True、sigma_overview=False，
    可同时开启）。垂向剖面线默认纬向（k=0、b=center_lat）；输入 anomalies parquet 的定位规则与
    plot_hotspot_anomaly_vertical_profiles 一致，每个剖面的 GLORYS 图分别交给 plot_argo_horizontal_glorys 与
    plot_argo_vertical_glorys_overview 生成。

    参数:
        - start_year (int | None): anomalies_path=None 时用于定位默认 anomalies 文件。
        - end_year (int | None): 同上，结束年份。
        - anomalies_path (str | Path | None): 指定 anomalies parquet 路径；None 时按 plot_argo_hotspots 命名规则自动定位。
        - detection_config (DetectionConfig | None): 异常识别配置；同时用于水平图异常点与垂向图投影点筛选。
        - horizontal_variable (str): 水平 GLORYS 背景变量，默认 'vorticity'。
        - horizontal_argo_depth (bool): 是否额外绘制 Argo 异常深度的水平图，默认 False（仅海表）。
        - vertical_variables (list[str] | None): 垂向总览变量；None 时用默认 ['vorticity','sigma','thetao','salinity']。
        - needed_depth (float | int): 水平图读取深度（m），默认 0。
        - xmin (float): 垂向图横向下界（km），同时决定 Argo-centered GLORYS 读取窗口，默认 -400.0。
        - xmax (float): 垂向图横向上界（km），默认 400.0。
        - ymin (float): 垂向图深度上界（m），默认 0.0。
        - ymax (float): 垂向图深度下界（m），默认 1000.0。
        - profile_spacing_km (float | None): 垂向 GLORYS 插值的水平采样步长（km）；None 时用配置默认。
        - interpolate_z (bool): 是否将深度轴重采样到等间距网格，默认 True。
        - profile_depth_spacing_m (float | None): 深度重采样步长（m）；None 时用配置默认。
        - plot_mlt (bool): z 坐标总览图是否叠加混合层深度线，默认 False。
        - plot_argo_projection (bool): z 坐标总览图是否叠加 Argo 投影层，默认 True。
        - plot_isolines (bool): 是否叠加变量等值线（z 与 σ 总览图共用），默认 True。
        - isoline_levels (int | list[float] | np.ndarray | None): 等值线级别数或显式级别；None 时自动。
        - isoline_color (str): 等值线颜色，默认 'black'。
        - isoline_linewidth (float): 等值线线宽，默认 0.8。
        - isoline_alpha (float): 等值线透明度，默认 0.45。
        - label_isolines (bool): 是否标注等值线数值，默认 True。
        - argo_min_depth (float | None): 覆盖水平图异常点最小深度阈值（m）；None 时用配置。
        - argo_projection_min_depth (float | None): 覆盖垂向投影点最小深度阈值（m）；None 时用配置。
        - argo_data_dir (str | Path | None): Argo 年数据目录；None 时使用配置默认路径。
        - output_dir (str | Path | None): 批处理专属输出根目录；None 时使用当前 method/region 默认目录。
        - clear_output_dir (bool): 保存图片时是否在本次运行开始前清空批处理专属输出目录，默认 True。
        - save_fig (bool): 是否保存图像，默认 True。
        - show_fig (bool): 是否显示图像，默认 False。
        - inline_mode (bool): 是否使用内联静态模式，默认 True。
        - verbose (bool): 是否打印底层单图保存路径等详细信息，默认 False。
        - use_multiprocessing (bool): 是否用多进程并行处理 profile，默认 True。
        - num_workers (int | None): worker 数；None 时自动取 min(profile数, CPU数, 4)。
        - maxtasksperchild (int | None): 每个 worker 处理多少任务后重启；None 表示不自动重启，默认 4。
        - heave_x_window_km (float): OI 局地水平窗口半宽（km），默认 25.0。
        - heave_z_window_m (float | None): OI 局地垂向窗口半宽（m），默认 100.0。
        - heave_search_range (float): 从 σ_argo 向上搜索的 σ 跨度（kg/m³），默认来自 processing.yml。
        - heave_depth_threshold (float): 通风判定深度（m），默认来自 processing.yml。
        - heave_z_search_m (float | None): 等密线连通性垂向范围（m），默认来自 processing.yml。
        - annotate_heave (bool): 是否在 vertical overview 标题显示出露深度，默认 True。
        - annotate_spice (bool): 是否在标题标注 spiciness 异常 δπ 并在 summary parquet 输出对应列，默认 True。
        - z_overview (bool): 是否绘制 z 坐标 2×2 垂向总览图，默认 True。
        - sigma_overview (bool): 是否绘制 σ 坐标 2×2 垂向总览图（PV/Z(σ)/θ/S），默认 False。
        - ts_diagram (bool): 是否额外绘制 T-S 图（仅 Argo 场景生效），默认 False。
        - return_details (bool): 是否返回每个 profile 的运行日志；默认 False，仅返回摘要。
        - save_summary_data (bool): 是否保存逐 profile GLORYS/Heave 诊断明细 parquet，默认 True。
        - summary_data_path (str | Path | None): 明细 parquet 保存路径；None 时使用批处理专属输出目录。

    返回:
        - dict: 默认含 total_candidates/processed_profiles/skipped_profiles/output_dir/anomalies_path/summary_data_path 等摘要。
    """

    def _to_int_or_none(val):
        try:
            if pd.isna(val):
                return None
            return int(val)
        except Exception:
            return None

    def _to_float_or_none(val):
        try:
            if pd.isna(val):
                return None
            out_val = float(val)
            return out_val if np.isfinite(out_val) else None
        except Exception:
            return None

    def _first_float_from_row(row: pd.Series, names: list[str]) -> float | None:
        for name in names:
            if name in row.index:
                value = _to_float_or_none(row.get(name))
                if value is not None:
                    return value
        return None

    def _infer_year_tag(path_obj: Path | None = None) -> str:
        if start_year is not None and end_year is not None:
            return f"{int(start_year)}_{int(end_year)}"
        if path_obj is not None:
            match = re.search(r'anomalies_(\d{4})_(\d{4})_', path_obj.name)
            if match:
                return f"{match.group(1)}_{match.group(2)}"
        return "custom"

    cfg = _resolve_detection_config(detection_config)
    method_name = cfg.method
    run_tag = cfg.file_stem()
    if argo_data_dir is None:
        argo_data_dir = argo_path
    region_slug = _current_region_key()
    batch_output_dir = (
        Path(output_dir)
        if output_dir is not None
        else cfg.output_dir("plot_hotspot_anomaly_argo_glorys_overviews", region_slug)
    )

    if anomalies_path is None:
        if start_year is None or end_year is None:
            raise ValueError("anomalies_path 为空时，必须提供 start_year 与 end_year。")
        anomalies_path = cfg.output_dir("plot_argo_hotspots", region_slug) / (
            f"anomalies_{start_year}_{end_year}_{run_tag}.parquet"
        )
    else:
        anomalies_path = Path(anomalies_path)

    if not Path(anomalies_path).exists():
        raise FileNotFoundError(f"Anomalies file not found: {anomalies_path}")

    year_tag = _infer_year_tag(Path(anomalies_path))
    resolved_summary_data_path = (
        Path(summary_data_path)
        if summary_data_path is not None
        else batch_output_dir / f"hotspot_anomaly_argo_glorys_overviews_summary_{year_tag}_{run_tag}.parquet"
    )

    anomalies = pd.read_parquet(anomalies_path)
    if 'detection_method' in anomalies.columns:
        method_mask = anomalies['detection_method'].astype(str).str.lower().eq(method_name)
        total_count = len(anomalies)
        mismatch_count = int((~method_mask).sum())
        if mismatch_count > 0:
            print(f"[WARN] Mixed detection_method found: expected={method_name}, mismatched={mismatch_count}/{total_count}.")
        elif verbose:
            print(f"[*] Method={method_name}, records={total_count}")
        anomalies = anomalies[method_mask].copy()

    if anomalies.empty:
        print(f"[*] No anomalies in file: {anomalies_path}")
        saved_summary_data_path = None
        if save_summary_data:
            try:
                resolved_summary_data_path.parent.mkdir(parents=True, exist_ok=True)
                empty_cols = [
                    'task_index',
                    'year',
                    'profile_number',
                    'platform_number',
                    'profile_time',
                    'line_strategy',
                    'k',
                    'b',
                    'center_lon',
                    'center_lat',
                    'target_date',
                    'anomaly_depth_m',
                    'horizontal_status',
                    'vertical_status',
                    'vertical_save_paths',
                    'status',
                    'error',
                    'projection_depth_m',
                    'heave_x_window_km',
                    'heave_z_window_m',
                    'heave_valid_fraction',
                    'glorys_heave_sigma_argo',
                    'glorys_heave_sigma_peak',
                    'glorys_heave_zmin',
                    'glorys_heave_m',
                    'glorys_heave_oi',
                    'heave_error',
                    'spice_anomaly',
                    'spice_percentile',
                    'spice_type',
                    'sigma_save_paths',
                ]
                pd.DataFrame(columns=empty_cols).to_parquet(resolved_summary_data_path, index=False)
                saved_summary_data_path = str(resolved_summary_data_path)
            except Exception as exc:
                print(f"[WARN] Failed to save empty GLORYS summary data: {exc}")
        empty_summary = {
            'total_candidates': 0,
            'processed_profiles': 0,
            'skipped_profiles': 0,
            'output_dir': str(batch_output_dir) if save_fig else None,
            'anomalies_path': str(anomalies_path),
            'summary_data_path': saved_summary_data_path,
            'heave_diagnostics': bool(annotate_heave),
            'annotate_heave': bool(annotate_heave),
            'annotate_spice': bool(annotate_spice),
        }
        if return_details:
            empty_summary['results'] = []
        return empty_summary

    required_cols = ['Year', 'Profile_number']
    missing_cols = [c for c in required_cols if c not in anomalies.columns]
    if missing_cols:
        raise ValueError(f"Anomalies file missing required columns: {missing_cols}")

    work = anomalies.copy()
    work['_year'] = pd.to_numeric(work['Year'], errors='coerce')
    work['_profile'] = pd.to_numeric(work['Profile_number'], errors='coerce')
    work = work.dropna(subset=['_year', '_profile']).copy()
    work['_year'] = work['_year'].astype(int)
    work['_profile'] = work['_profile'].astype(int)

    horizontal_output_dir = batch_output_dir / "horizontal"
    vertical_output_dir = batch_output_dir / "vertical_overview"
    if save_fig:
        if clear_output_dir and batch_output_dir.exists():
            try:
                shutil.rmtree(batch_output_dir)
            except Exception as exc:
                print(f"[WARN] Failed to clear output directory {batch_output_dir}: {exc}")
    if save_fig or save_summary_data:
        batch_output_dir.mkdir(parents=True, exist_ok=True)
    if save_fig:
        horizontal_output_dir.mkdir(parents=True, exist_ok=True)
        vertical_output_dir.mkdir(parents=True, exist_ok=True)

    total_candidates = int(len(work))
    worker_count = 1
    if use_multiprocessing and total_candidates > 1:
        worker_count = int(num_workers) if num_workers is not None else min(total_candidates, os.cpu_count() or 1, 4)
        worker_count = max(1, worker_count)

    region_config_key = _current_region_config_key()
    worker_args: list[dict] = []
    for task_index, (_, row) in enumerate(work.iterrows()):
        worker_args.append({
            'task_index': int(task_index),
            'year': int(row['_year']),
            'profile_number': int(row['_profile']),
            'month': _to_int_or_none(row.get('Month')),
            'day': _to_int_or_none(row.get('Day')),
            'platform_number': _to_int_or_none(row.get('Platform_number')),
            'nearshore_do_dip': row.get('nearshore_do_dip', pd.NA),
            'heave_projection_depth_m': _first_float_from_row(row, ['depth', 'Depth', 'Anomaly_depth']),
            'detection_config': cfg,
            'region_config_key': region_config_key,
            'horizontal_variable': horizontal_variable,
            'horizontal_argo_depth': horizontal_argo_depth,
            'vertical_variables': vertical_variables,
            'needed_depth': needed_depth,
            'xmin': xmin,
            'xmax': xmax,
            'ymin': ymin,
            'ymax': ymax,
            'profile_spacing_km': profile_spacing_km,
            'interpolate_z': interpolate_z,
            'profile_depth_spacing_m': profile_depth_spacing_m,
            'plot_mlt': plot_mlt,
            'plot_argo_projection': plot_argo_projection,
            'plot_isolines': plot_isolines,
            'isoline_levels': isoline_levels,
            'isoline_color': isoline_color,
            'isoline_linewidth': isoline_linewidth,
            'isoline_alpha': isoline_alpha,
            'label_isolines': label_isolines,
            'argo_min_depth': argo_min_depth,
            'argo_projection_min_depth': argo_projection_min_depth,
            'argo_data_dir': argo_data_dir,
            'save_fig': save_fig,
            'show_fig': show_fig,
            'inline_mode': inline_mode,
            'horizontal_output_dir': horizontal_output_dir if save_fig else None,
            'vertical_output_dir': vertical_output_dir if save_fig else None,
            'verbose': verbose,
            'force_agg_backend': bool(worker_count > 1 and not show_fig),
            'heave_x_window_km': heave_x_window_km,
            'heave_z_window_m': heave_z_window_m,
            'heave_search_range': float(heave_search_range),
            'heave_depth_threshold': heave_depth_threshold,
            'heave_z_search_m': heave_z_search_m,
            'annotate_heave': annotate_heave,
            'z_overview': z_overview,
            'sigma_overview': sigma_overview,
            'ts_diagram': ts_diagram,
            'annotate_spice': annotate_spice,
            'salinity_value': _to_float_or_none(row.get('salinity_value')),
            'temperature_value': _to_float_or_none(row.get('temperature_value')),
        })

    if worker_count > 1:
        print(f"[*] Hotspots Argo GLORYS multiprocessing: workers={worker_count}, profiles={total_candidates}.")
        pool_kwargs = {'processes': worker_count}
        if maxtasksperchild is not None:
            pool_kwargs['maxtasksperchild'] = max(1, int(maxtasksperchild))
        with multiprocessing.Pool(**pool_kwargs) as pool:
            results = list(tqdm(
                pool.imap_unordered(_plot_hotspot_argo_glorys_profile_worker, worker_args),
                total=total_candidates,
                desc="hotspot argo GLORYS",
                unit="profile",
            ))
        results = sorted(results, key=lambda item: int(item.get('task_index', 0)))
    else:
        results = [
            _plot_hotspot_argo_glorys_profile_worker(args)
            for args in tqdm(worker_args, total=total_candidates, desc="hotspot argo GLORYS", unit="profile")
        ]

    processed_profiles = int(sum(1 for item in results if item.get('status') in {'ok', 'partial'}))
    skipped_profiles = int(sum(1 for item in results if item.get('status') == 'failed'))
    for item in results:
        if item.get('status') == 'failed':
            print(
                f"[WARN] Failed hotspot Argo GLORYS for "
                f"Year={item.get('year')}, Profile={item.get('profile_number')}: {item.get('error')}"
            )

    print(
        f"[*] Hotspots Argo GLORYS plotting complete: "
        f"total={total_candidates}, processed={processed_profiles}, skipped={skipped_profiles}."
    )

    saved_summary_data_path = None
    if save_summary_data:
        try:
            resolved_summary_data_path.parent.mkdir(parents=True, exist_ok=True)
            summary_df = pd.DataFrame(results)
            # spice_type / hotspot_type 整数码一并落盘，使其成为一等列（导出仅映射成名称）
            summary_df['spice_type'] = _assign_spice_type(
                summary_df, percentile_threshold=cfg.spice_percentile_threshold)
            summary_df['hotspot_type'] = _assign_hotspot_type(
                summary_df,
                heave_z_threshold=float(heave_depth_threshold),
                heave_m_threshold=float(_heave_magnitude_threshold),
            )
            summary_df.to_parquet(resolved_summary_data_path, index=False)
            saved_summary_data_path = str(resolved_summary_data_path)
            if verbose:
                print(f"Summary data saved to: {resolved_summary_data_path}")
        except Exception as exc:
            print(f"[WARN] Failed to save GLORYS summary data: {exc}")

    summary = {
        'total_candidates': total_candidates,
        'processed_profiles': int(processed_profiles),
        'skipped_profiles': int(skipped_profiles),
        'output_dir': str(batch_output_dir) if save_fig else None,
        'horizontal_output_dir': str(horizontal_output_dir) if save_fig else None,
        'vertical_output_dir': str(vertical_output_dir) if save_fig else None,
        'anomalies_path': str(anomalies_path),
        'summary_data_path': saved_summary_data_path,
        'heave_diagnostics': bool(annotate_heave),
        'annotate_heave': bool(annotate_heave),
        'annotate_spice': bool(annotate_spice),
    }
    if return_details:
        summary['results'] = results
    return summary


def _plot_hotspot_argo_reconstruction_profile_worker(args: dict) -> dict:
    """单个 hotspot Argo profile 的点 Eulerian 重建 vertical overview worker（含覆盖率预筛）。"""
    task_index = int(args.get('task_index', -1))
    year_val = int(args['year'])
    profile_num = int(args['profile_number'])
    month_val = args.get('month')
    day_val = args.get('day')
    platform_val = args.get('platform_number')
    anomaly_depth_m = args.get('anomaly_depth_m')

    if month_val is not None and day_val is not None:
        profile_time = pd.Timestamp(year=year_val, month=int(month_val), day=int(day_val))
    else:
        profile_time = year_val

    record = {
        'task_index': task_index,
        'year': year_val,
        'profile_number': profile_num,
        'platform_number': platform_val,
        'profile_time': str(profile_time),
        'center_lon': np.nan,
        'center_lat': np.nan,
        'target_date': None,
        'anomaly_depth_m': anomaly_depth_m,
        'n_profiles': 0,
        'est_coverage_top': np.nan,
        'coverage_top': np.nan,
        'save_path': None,
        'skip_reason': None,
        'status': 'failed',
        'error': None,
    }

    try:
        region_config_key = args.get('region_config_key')
        if region_config_key:
            switch_region(str(region_config_key), verbose=False)
        if bool(args.get('force_agg_backend', False)):
            try:
                plt.switch_backend('Agg')
            except Exception:
                pass
        if not bool(args.get('show_fig', False)):
            plt.ioff()
        plt.close('all')

        info = _resolve_argo_profile_center(
            profile_number=profile_num,
            profile_time=profile_time,
            platform_number=platform_val,
            argo_data_dir=args.get('argo_data_dir'),
        )
        center_lon = float(info['center_lon'])
        center_lat = float(info['center_lat'])
        target_date = pd.Timestamp(info['target_date']).normalize()
        platform_resolved = info.get('platform_number')
        if platform_val is None and platform_resolved is not None:
            platform_val = int(platform_resolved)
        record.update({
            'platform_number': platform_val,
            'profile_time': target_date.strftime('%Y-%m-%d'),
            'center_lon': center_lon,
            'center_lat': center_lat,
            'target_date': target_date.strftime('%Y-%m-%d'),
        })

        radius_km = float(args['radius_km'])
        day_window = int(args['day_window'])
        h_bw = float(args['h_bw'])
        depth_bw = float(args['depth_bw'])
        h_spacing_deg = float(args['h_spacing_deg'])
        z_max_m = float(args['z_max_m'])
        z_spacing_m = float(args['z_spacing_m'])
        min_weight = float(args['min_weight'])
        inner_n_jobs = args.get('inner_n_jobs')

        lon_min, lon_max, lat_min, lat_max = _window_bounds_from_center_km(center_lon, center_lat, radius_km)
        lon_range = (lon_min, lon_max)
        lat_range = (lat_min, lat_max)
        t_min = target_date - pd.Timedelta(days=day_window)
        t_max = target_date + pd.Timedelta(days=day_window)

        pool = collect_argo_pool(lon_range, lat_range, t_min, t_max, max_depth=z_max_m + 200.0)
        n_prof = int(pool['Profile_number'].nunique()) if not pool.empty else 0
        record['n_profiles'] = n_prof

        # Stage 1：剖面数预筛（~毫秒）；不足直接跳过，不进入任何 build
        min_profiles = int(args['min_profiles'])
        if n_prof < min_profiles:
            record['skip_reason'] = f'n_profiles<{min_profiles}'
            record['status'] = 'skipped'
            return record

        # Stage 2：粗深度探针预估 ≤1000m 覆盖率（~亚秒）；不足直接跳过，不做全量 build
        est_cov = _estimate_argo_top_coverage(
            pool, lon_range, lat_range,
            h_bw=h_bw, depth_bw=depth_bw, h_spacing_deg=h_spacing_deg,
            min_weight=min_weight,
            probe_spacing_m=float(args['coverage_probe_spacing_m']),
            n_jobs=inner_n_jobs,
        )
        record['est_coverage_top'] = est_cov
        min_coverage = float(args['min_coverage_top1000'])
        if est_cov < min_coverage:
            record['skip_reason'] = f'est_coverage<{min_coverage:.2f}'
            record['status'] = 'skipped'
            return record

        field = _build_argo_3d_field(
            pool, lon_range, lat_range,
            h_bw=h_bw, depth_bw=depth_bw, h_spacing_deg=h_spacing_deg,
            z_max_m=z_max_m, z_spacing_m=z_spacing_m, min_weight=min_weight,
            n_jobs=inner_n_jobs, verbose=False,
        )
        field['attrs']['start_date'] = t_min.strftime('%Y-%m-%d')
        field['attrs']['end_date'] = t_max.strftime('%Y-%m-%d')
        z_top = field['depth'] <= 1000.0
        valid_top = ~np.isnan(field['thetao'][z_top])
        record['coverage_top'] = float(valid_top.mean()) if valid_top.size else np.nan

        out_dir = args.get('vertical_output_dir')
        # 与 GLORYS 总览同前缀 Argo_{日期}_P{剖面}，便于两批输出在目录里成对相邻对比
        save_name = f'Argo_{target_date.strftime("%Y%m%d")}_P{profile_num}_recon_hbw{h_bw:.0f}km.png'
        _plot_center_vertical_argo_overview(
            center_lon, center_lat, target_date,
            subject_label=f'Profile {profile_num}',
            save_name_prefix=f'P{profile_num}',
            save_name=save_name,
            save_subdir='plot_hotspot_anomaly_argo_reconstruction_overviews',
            n_prof=n_prof,
            k=float(args.get('k', 0.0)),
            radius_km=radius_km, day_window=day_window,
            h_bw=h_bw, depth_bw=depth_bw, h_spacing_deg=h_spacing_deg,
            z_max_m=z_max_m, z_spacing_m=z_spacing_m, min_weight=min_weight,
            x_spacing_km=float(args['x_spacing_km']),
            ymin=float(args.get('ymin', 0.0)), ymax=float(args.get('ymax', 1000.0)),
            plot_isolines=bool(args.get('plot_isolines', True)),
            isoline_levels=args.get('isoline_levels'),
            isoline_color=args.get('isoline_color', 'black'),
            isoline_linewidth=float(args.get('isoline_linewidth', 0.8)),
            isoline_alpha=float(args.get('isoline_alpha', 0.45)),
            label_isolines=bool(args.get('label_isolines', True)),
            field=field,
            show_fig=bool(args.get('show_fig', False)),
            save_fig=bool(args.get('save_fig', True)),
            output_dir=out_dir,
            verbose=bool(args.get('verbose', False)),
        )
        if bool(args.get('save_fig', True)) and out_dir is not None:
            record['save_path'] = str(Path(out_dir) / save_name)
        record['status'] = 'ok'

    except Exception as exc:
        record['error'] = str(exc)
    finally:
        plt.close('all')

    return record


def plot_hotspot_anomaly_argo_reconstruction_overviews(
    start_year: int | None = None,
    end_year: int | None = None,
    anomalies_path: str | Path | None = None,
    detection_config: DetectionConfig | None = None,
    *,
    k: float = 0.0,
    radius_km: float = _argo_recon_radius_km,
    day_window: int = _argo_recon_day_window,
    h_bw: float = _argo_recon_h_bw_km,
    depth_bw: float = _argo_recon_depth_bw_m,
    h_spacing_deg: float = _argo_recon_h_spacing_deg,
    z_max_m: float = _argo_recon_z_max_m,
    z_spacing_m: float = _argo_recon_z_spacing_m,
    min_weight: float = _argo_recon_min_weight,
    x_spacing_km: float = _argo_recon_x_spacing_km,
    ymin: float = 0.0,
    ymax: float = 1000.0,
    min_profiles: int = _argo_recon_min_profiles,
    min_coverage_top1000: float = _argo_recon_min_coverage_top,
    coverage_probe_spacing_m: float = _argo_recon_coverage_probe_spacing_m,
    plot_isolines: bool = True,
    isoline_levels: int | list[float] | np.ndarray | None = None,
    isoline_color: str = 'black',
    isoline_linewidth: float = 0.8,
    isoline_alpha: float = 0.45,
    label_isolines: bool = True,
    argo_data_dir: str | Path | None = None,
    output_dir: str | Path | None = None,
    clear_output_dir: bool = True,
    save_fig: bool = True,
    show_fig: bool = False,
    verbose: bool = False,
    use_multiprocessing: bool = True,
    num_workers: int | None = None,
    maxtasksperchild: int | None = 4,
    return_details: bool = False,
    save_summary_data: bool = True,
    summary_data_path: str | Path | None = None,
) -> dict:
    """为 hotspots 异常剖面批量重建并绘制点 Eulerian Argo 垂向断面 2x2 总览图。

    与 plot_hotspot_anomaly_argo_glorys_overviews 对标：读取同一份 anomalies parquet、
    按完全一致的逐行规则（_resolve_argo_profile_center）定位每条剖面的中心与日期，因此
    第 i 张重建图与该函数第 i 张 GLORYS 图一一对应，可并排比较。每条剖面以
    center ± radius_km 圈定盒子、target_date ± day_window 为时间窗，调用与
    plot_argo_vertical_argo_overview 相同的点 Eulerian 重建核心（argo 变体），输出
    [Argo Coverage / σ₀ / θ / S] 四联图。

    重建较重（单图约一分钟），故在 build 前做两段廉价预筛：先数盒子+时间窗内的 Argo
    剖面数（min_profiles），再用粗深度探针预估 ≤1000m 覆盖率（min_coverage_top1000，
    复用 _build_argo_3d_field 的稀疏深度层，约亚秒）；任一不达标即跳过、不做全量 build。
    被跳过的剖面在 summary parquet 中记录 n_profiles / est_coverage_top / skip_reason，
    便于事后审计与回调阈值。

    参数:
        - start_year (int | None): anomalies_path=None 时按命名规则定位默认 anomalies 文件。
        - end_year (int | None): 同上，结束年份。
        - anomalies_path (str | Path | None): 指定 anomalies parquet 路径；None 时按 plot_argo_hotspots 命名规则自动定位。
        - detection_config (DetectionConfig | None): 异常识别配置；用于筛选 detection_method 与定位输出目录。
        - k (float): 断面测线斜率 Δlat/Δlon，0.0 为纬向，默认 0.0。
        - radius_km (float): 中心半径（km），同时作为采集盒子半宽与断面 x 轴半宽，默认来自配置。
        - day_window (int): 时间窗半宽（天），围绕剖面日期取 ±day_window，默认来自配置。
        - h_bw (float): 水平高斯核带宽（km，保留中尺度结构），默认来自配置。
        - depth_bw (float): 垂向高斯核带宽（m），默认来自配置。
        - h_spacing_deg (float): 重建网格水平间距（°），默认来自配置。
        - z_max_m (float): 最大重建深度（m），默认来自配置。
        - z_spacing_m (float): 垂向网格间距（m），默认来自配置。
        - min_weight (float): 最小累积权重阈值，低于此值格点标为 NaN，默认来自配置。
        - x_spacing_km (float): 断面水平采样间距（km），默认来自配置。
        - ymin (float): 图纵轴深度上界（m），默认 0.0。
        - ymax (float): 图纵轴深度下界（m），默认 1000.0；3D 场仍建到 z_max_m。
        - min_profiles (int): 预筛一段，盒子+时间窗内剖面数下限，不足直接跳过，默认来自配置。
        - min_coverage_top1000 (float): 预筛二段，≤1000m 预估覆盖率下限，不足直接跳过，默认来自配置。
        - coverage_probe_spacing_m (float): 预估覆盖率的粗深度探针步长（m），默认来自配置。
        - plot_isolines (bool): 是否叠加 σ₀ 等值线，默认 True。
        - isoline_levels (int | list[float] | np.ndarray | None): 等值线级别数或显式级别；None 时自动。
        - isoline_color (str): 等值线颜色，默认 'black'。
        - isoline_linewidth (float): 等值线线宽，默认 0.8。
        - isoline_alpha (float): 等值线透明度，默认 0.45。
        - label_isolines (bool): 是否标注等值线数值，默认 True。
        - argo_data_dir (str | Path | None): Argo 年数据目录；None 时使用配置默认路径。
        - output_dir (str | Path | None): 批处理专属输出根目录；None 时使用当前 method/region 默认目录。
        - clear_output_dir (bool): 保存图片时是否在本次运行开始前清空批处理专属输出目录，默认 True。
        - save_fig (bool): 是否保存图像，默认 True。
        - show_fig (bool): 是否显示图像，默认 False。
        - verbose (bool): 是否打印底层单图进度等详细信息，默认 False。
        - use_multiprocessing (bool): 是否按剖面并行；并行时每个 worker 内部 build 强制单进程以避免嵌套，默认 True。
        - num_workers (int | None): worker 数；None 时自动取 min(剖面数, CPU数, 24)。
        - maxtasksperchild (int | None): 每个 worker 处理多少任务后重启；None 表示不自动重启，默认 4。
        - return_details (bool): 是否在返回值中附每条剖面的运行记录，默认 False。
        - save_summary_data (bool): 是否保存逐剖面预筛/覆盖率明细 parquet，默认 True。
        - summary_data_path (str | Path | None): 明细 parquet 保存路径；None 时使用批处理专属输出目录。

    返回:
        - dict: 含 total_candidates/processed_profiles/skipped_profiles/failed_profiles/output_dir/vertical_output_dir/anomalies_path/summary_data_path 等摘要；return_details=True 时附带每条剖面的 results 列表。
    """

    def _to_int_or_none(val):
        try:
            if pd.isna(val):
                return None
            return int(val)
        except Exception:
            return None

    def _first_float_from_row(row: pd.Series, names: list[str]) -> float | None:
        for name in names:
            if name not in row.index:
                continue
            try:
                if pd.isna(row.get(name)):
                    continue
                value = float(row.get(name))
                if np.isfinite(value):
                    return value
            except Exception:
                continue
        return None

    def _infer_year_tag(path_obj: Path | None = None) -> str:
        if start_year is not None and end_year is not None:
            return f"{int(start_year)}_{int(end_year)}"
        if path_obj is not None:
            match = re.search(r'anomalies_(\d{4})_(\d{4})_', path_obj.name)
            if match:
                return f"{match.group(1)}_{match.group(2)}"
        return "custom"

    cfg = _resolve_detection_config(detection_config)
    method_name = cfg.method
    run_tag = cfg.file_stem()
    if argo_data_dir is None:
        argo_data_dir = argo_path
    region_slug = _current_region_key()
    batch_output_dir = (
        Path(output_dir)
        if output_dir is not None
        else cfg.output_dir("plot_hotspot_anomaly_argo_reconstruction_overviews", region_slug)
    )

    if anomalies_path is None:
        if start_year is None or end_year is None:
            raise ValueError("anomalies_path 为空时，必须提供 start_year 与 end_year。")
        anomalies_path = cfg.output_dir("plot_argo_hotspots", region_slug) / (
            f"anomalies_{start_year}_{end_year}_{run_tag}.parquet"
        )
    else:
        anomalies_path = Path(anomalies_path)

    if not Path(anomalies_path).exists():
        raise FileNotFoundError(f"Anomalies file not found: {anomalies_path}")

    year_tag = _infer_year_tag(Path(anomalies_path))
    resolved_summary_data_path = (
        Path(summary_data_path)
        if summary_data_path is not None
        else batch_output_dir / f"hotspot_anomaly_argo_reconstruction_overviews_summary_{year_tag}_{run_tag}.parquet"
    )

    anomalies = pd.read_parquet(anomalies_path)
    if 'detection_method' in anomalies.columns:
        method_mask = anomalies['detection_method'].astype(str).str.lower().eq(method_name)
        mismatch_count = int((~method_mask).sum())
        if mismatch_count > 0:
            print(f"[WARN] Mixed detection_method found: expected={method_name}, mismatched={mismatch_count}/{len(anomalies)}.")
        anomalies = anomalies[method_mask].copy()

    required_cols = ['Year', 'Profile_number']
    missing_cols = [c for c in required_cols if c not in anomalies.columns]
    if missing_cols:
        raise ValueError(f"Anomalies file missing required columns: {missing_cols}")

    work = anomalies.copy()
    work['_year'] = pd.to_numeric(work['Year'], errors='coerce')
    work['_profile'] = pd.to_numeric(work['Profile_number'], errors='coerce')
    work = work.dropna(subset=['_year', '_profile']).copy()
    work['_year'] = work['_year'].astype(int)
    work['_profile'] = work['_profile'].astype(int)

    vertical_output_dir = batch_output_dir / "vertical_overview"
    if save_fig:
        if clear_output_dir and batch_output_dir.exists():
            try:
                shutil.rmtree(batch_output_dir)
            except Exception as exc:
                print(f"[WARN] Failed to clear output directory {batch_output_dir}: {exc}")
    if save_fig or save_summary_data:
        batch_output_dir.mkdir(parents=True, exist_ok=True)
    if save_fig:
        vertical_output_dir.mkdir(parents=True, exist_ok=True)

    total_candidates = int(len(work))
    if total_candidates == 0:
        print(f"[*] No anomalies in file: {anomalies_path}")
        empty_summary = {
            'total_candidates': 0,
            'processed_profiles': 0,
            'skipped_profiles': 0,
            'failed_profiles': 0,
            'output_dir': str(batch_output_dir) if save_fig else None,
            'vertical_output_dir': str(vertical_output_dir) if save_fig else None,
            'anomalies_path': str(anomalies_path),
            'summary_data_path': None,
        }
        if return_details:
            empty_summary['results'] = []
        return empty_summary

    worker_count = 1
    if use_multiprocessing and total_candidates > 1:
        # 纯本地 Argo parquet + CPU-bound build，无远程 GLORYS IO 争用，故默认放到物理核量级
        # （24 对应 i9-14900KF 的 24 物理核，避免 HT 超订；不沿用 GLORYS 那个为远程盘设的 4）
        worker_count = int(num_workers) if num_workers is not None else min(
            total_candidates, os.cpu_count() or 1, 24
        )
        worker_count = max(1, worker_count)
    # 跨剖面并行时，每个 worker 内部 build 强制单进程，避免 daemon 进程嵌套 multiprocessing
    inner_n_jobs = 1 if worker_count > 1 else None

    region_config_key = _current_region_config_key()
    worker_args: list[dict] = []
    for task_index, (_, row) in enumerate(work.iterrows()):
        worker_args.append({
            'task_index': int(task_index),
            'year': int(row['_year']),
            'profile_number': int(row['_profile']),
            'month': _to_int_or_none(row.get('Month')),
            'day': _to_int_or_none(row.get('Day')),
            'platform_number': _to_int_or_none(row.get('Platform_number')),
            'anomaly_depth_m': _first_float_from_row(row, ['depth', 'Depth', 'Anomaly_depth']),
            'region_config_key': region_config_key,
            'argo_data_dir': argo_data_dir,
            'k': k,
            'radius_km': radius_km,
            'day_window': day_window,
            'h_bw': h_bw,
            'depth_bw': depth_bw,
            'h_spacing_deg': h_spacing_deg,
            'z_max_m': z_max_m,
            'z_spacing_m': z_spacing_m,
            'min_weight': min_weight,
            'x_spacing_km': x_spacing_km,
            'ymin': ymin,
            'ymax': ymax,
            'min_profiles': min_profiles,
            'min_coverage_top1000': min_coverage_top1000,
            'coverage_probe_spacing_m': coverage_probe_spacing_m,
            'inner_n_jobs': inner_n_jobs,
            'plot_isolines': plot_isolines,
            'isoline_levels': isoline_levels,
            'isoline_color': isoline_color,
            'isoline_linewidth': isoline_linewidth,
            'isoline_alpha': isoline_alpha,
            'label_isolines': label_isolines,
            'save_fig': save_fig,
            'show_fig': show_fig,
            'vertical_output_dir': vertical_output_dir if save_fig else None,
            'verbose': verbose,
            'force_agg_backend': bool(worker_count > 1 and not show_fig),
        })

    if worker_count > 1:
        print(f"[*] Hotspots Argo reconstruction multiprocessing: workers={worker_count}, profiles={total_candidates}.")
        pool_kwargs = {'processes': worker_count}
        if maxtasksperchild is not None:
            pool_kwargs['maxtasksperchild'] = max(1, int(maxtasksperchild))
        with multiprocessing.Pool(**pool_kwargs) as mp_pool:
            results = list(tqdm(
                mp_pool.imap_unordered(_plot_hotspot_argo_reconstruction_profile_worker, worker_args),
                total=total_candidates,
                desc="hotspot argo recon",
                unit="profile",
            ))
        results = sorted(results, key=lambda item: int(item.get('task_index', 0)))
    else:
        results = [
            _plot_hotspot_argo_reconstruction_profile_worker(a)
            for a in tqdm(worker_args, total=total_candidates, desc="hotspot argo recon", unit="profile")
        ]

    processed_profiles = int(sum(1 for item in results if item.get('status') == 'ok'))
    skipped_profiles = int(sum(1 for item in results if item.get('status') == 'skipped'))
    failed_profiles = int(sum(1 for item in results if item.get('status') == 'failed'))
    for item in results:
        if item.get('status') == 'failed':
            print(
                f"[WARN] Failed hotspot Argo reconstruction for "
                f"Year={item.get('year')}, Profile={item.get('profile_number')}: {item.get('error')}"
            )

    print(
        f"[*] Hotspots Argo reconstruction complete: total={total_candidates}, "
        f"plotted={processed_profiles}, skipped(low-coverage)={skipped_profiles}, failed={failed_profiles}."
    )

    saved_summary_data_path = None
    if save_summary_data:
        try:
            resolved_summary_data_path.parent.mkdir(parents=True, exist_ok=True)
            pd.DataFrame(results).to_parquet(resolved_summary_data_path, index=False)
            saved_summary_data_path = str(resolved_summary_data_path)
            if verbose:
                print(f"Summary data saved to: {resolved_summary_data_path}")
        except Exception as exc:
            print(f"[WARN] Failed to save reconstruction summary data: {exc}")

    summary = {
        'total_candidates': total_candidates,
        'processed_profiles': processed_profiles,
        'skipped_profiles': skipped_profiles,
        'failed_profiles': failed_profiles,
        'output_dir': str(batch_output_dir) if save_fig else None,
        'vertical_output_dir': str(vertical_output_dir) if save_fig else None,
        'anomalies_path': str(anomalies_path),
        'summary_data_path': saved_summary_data_path,
    }
    if return_details:
        summary['results'] = results
    return summary


def _glorys_residual_core(args: dict) -> tuple[dict, dict | None]:
    """单条 Argo 剖面的 GLORYS 逐点残差核心（多进程入口）：同位残差 + ±窗最佳匹配，拆错位 vs 丢细节。

    把 GLORYS 插值到 Argo 剖面位置（同位），并在 ±经纬度/±天窗口内搜最小 RMS 的原生
    GLORYS 列（最佳匹配）；二者之差给出错位贡献，最佳匹配后的高通细结构比 <1 即真丢细节。
    返回 (record, curves)：record 为扁平标量明细，成功时 curves 含统一 z 网格上的三条潜温
    廓线（供单剖面对比图），跳过/失败时 curves 为 None。
    """
    task_index = int(args.get('task_index', -1))
    record = {
        'task_index': task_index,
        'year': int(args['year']),
        'profile_number': int(args['profile_number']),
        'lat': np.nan, 'lon': np.nan,
        'anomaly_depth_m': args.get('anomaly_depth_m'),
        'rms_same': np.nan, 'rms_best': np.nan, 'disp_rms': np.nan,
        'disp_km': np.nan, 'day_off': 0,
        'fine_argo': np.nan, 'fine_glorys_same': np.nan, 'fine_glorys_best': np.nan,
        'fine_ratio_same': np.nan, 'fine_ratio_best': np.nan,
        'status': 'failed', 'skip_reason': None, 'error': None,
    }
    curves = None
    try:
        region_config_key = args.get('region_config_key')
        if region_config_key:
            switch_region(str(region_config_key), verbose=False)

        clon = float(args['center_lon'])
        clat = float(args['center_lat'])
        record['lon'] = round(clon, 3)
        record['lat'] = round(clat, 3)
        year = int(args['year'])
        date = pd.Timestamp(year=year, month=int(args['month']), day=int(args['day']))

        z_max = float(args['z_max_m'])
        z_sp = float(args['z_spacing_m'])
        fine_win = int(args['fine_window'])
        zc = np.arange(0.0, z_max + 0.1, z_sp)

        def onto(z, v):
            m = np.isfinite(z) & np.isfinite(v)
            return np.interp(zc, z[m], v[m], left=np.nan, right=np.nan) if m.sum() > 3 else np.full(zc.size, np.nan)

        def fine_std(v):
            # 高通：减去滚动均值后取标准差，得到比窗口尺度更细的垂向结构强度
            s = pd.Series(v)
            resid = (s - s.rolling(fine_win, center=True, min_periods=5).mean()).to_numpy()
            resid = resid[np.isfinite(resid)]
            return float(np.std(resid)) if resid.size else np.nan

        def load_thetao(needed_date, rad_km):
            lo, hi, la, lb = _window_bounds_from_center_km(clon, clat, rad_km)
            glon, glat, gdep, gdat = _load_glorys_window_by_center(
                needed_date, clon, lo, hi, la, lb, variables=['thetao'], depth=None)
            return glon, glat, gdep, np.ma.filled(np.ma.array(gdat['thetao']), np.nan)

        if year not in _RESID_ARGO_CACHE:
            if len(_RESID_ARGO_CACHE) >= _RESID_ARGO_CACHE_MAX:
                _RESID_ARGO_CACHE.pop(next(iter(_RESID_ARGO_CACHE)))  # 淘汰最早进入的年份
            _RESID_ARGO_CACHE[year] = load_argo_data(year, data_dir=args.get('argo_data_dir'))
        prof = _RESID_ARGO_CACHE[year]
        prof = prof[prof['Profile_number'] == int(args['profile_number'])].sort_values('Depth')
        za = prof['Depth'].to_numpy(float)
        ti = prof['Temperature'].to_numpy(float)
        sa = prof['Salinity'].to_numpy(float)
        if za.size < 5:
            record['skip_reason'] = 'argo_profile<5pts'
            record['status'] = 'skipped'
            return record, None
        # Argo 现场温 → 潜温，与 GLORYS thetao 对齐
        pres = gsw.p_from_z(-za, clat)
        SA = gsw.SA_from_SP(sa, pres, clon, clat)
        pta = gsw.pt0_from_t(SA, ti, pres)
        PTa = onto(za, pta)
        fa = fine_std(PTa)
        if not np.isfinite(fa) or fa <= 0:
            record['skip_reason'] = 'argo_fine_std<=0'
            record['status'] = 'skipped'
            return record, None
        record['fine_argo'] = round(fa, 4)

        glon, glat, gdep, cube0 = load_thetao(date, float(args['sameloc_radius_km']))
        qlat = float(np.clip(clat, glat.min(), glat.max()))
        qlon = float(np.clip(clon, glon.min(), glon.max()))
        rgi = RegularGridInterpolator((gdep, glat, glon), cube0, bounds_error=False, fill_value=np.nan)
        ptg_same = rgi(np.column_stack([gdep, np.full(gdep.size, qlat), np.full(gdep.size, qlon)]))
        PTg_same = onto(gdep, ptg_same)
        in_z = zc <= z_max
        resid_same = (PTg_same - PTa)[in_z]
        resid_same = resid_same[np.isfinite(resid_same)]
        rms_same = float(np.std(resid_same)) if resid_same.size else np.nan
        f_same = fine_std(PTg_same)

        win_deg = float(args['match_window_deg'])
        win_days = int(args['match_window_days'])
        match_rad = float(args['match_radius_km'])
        match_min_cov = float(args['match_min_depth_coverage'])
        best = dict(rms=np.inf, dk=np.nan, dd=0, col=None, gdep=None)
        for dd in range(-win_days, win_days + 1):
            try:
                glon2, glat2, gdep2, cube = load_thetao(date + pd.Timedelta(days=dd), match_rad)
            except Exception:
                continue
            pta_on_g = np.interp(gdep2, zc, PTa, left=np.nan, right=np.nan)
            zmask = gdep2 <= z_max
            d = (cube - pta_on_g[:, None, None])[zmask]
            valid = np.isfinite(d)
            cnt = valid.sum(axis=0)
            denom = np.maximum(cnt, 1)
            # 逐列差值的「去均值标准差」，与同位 rms_same 同口径（只看结构、不计整体冷暖偏差，
            # 使 disp²=total²−resid² 的分解自洽）；计数保护避开全 NaN 海床列的空切片告警
            col_mean = np.where(valid, d, 0.0).sum(axis=0) / denom
            col_var = np.where(valid, d * d, 0.0).sum(axis=0) / denom - col_mean ** 2
            rmsmap = np.sqrt(np.where(cnt > 0, np.clip(col_var, 0.0, None), np.nan))
            # 候选列须覆盖 Argo 廓线深度范围的 match_min_cov 比例，否则浅海床列会以截断深度刷低误胜
            nz_argo = int(np.isfinite(pta_on_g[zmask]).sum())
            enough = cnt >= max(1, int(np.ceil(match_min_cov * nz_argo)))
            dlon = np.array([_minimal_lon_diff_deg(lon_val, clon) for lon_val in glon2])
            within = (np.abs(glat2[:, None] - clat) <= win_deg) & (np.abs(dlon[None, :]) <= win_deg)
            rmsmap = np.where(within & enough & np.isfinite(rmsmap), rmsmap, np.inf)
            if not np.isfinite(rmsmap).any():
                continue
            j, i = np.unravel_index(np.argmin(rmsmap), rmsmap.shape)
            if rmsmap[j, i] < best['rms']:
                sc = approximate_degree_length(clat)
                dk = np.hypot(dlon[i] * sc['meters_per_degree_lon'],
                              (glat2[j] - clat) * sc['meters_per_degree_lat']) / 1000.0
                best = dict(rms=float(rmsmap[j, i]), col=cube[:, j, i], gdep=gdep2, dk=float(dk), dd=int(dd))
        if best['col'] is None:
            record['skip_reason'] = 'no_glorys_match'
            record['status'] = 'skipped'
            return record, None
        PTg_best = onto(best['gdep'], best['col'])
        fb = fine_std(PTg_best)
        # 错位贡献按二次方差分离：总偏差² = 错位² + 残差²
        disp_rms = float(np.sqrt(max(rms_same ** 2 - best['rms'] ** 2, 0.0)))

        record.update({
            'rms_same': round(rms_same, 3), 'rms_best': round(best['rms'], 3),
            'disp_rms': round(disp_rms, 3), 'disp_km': round(best['dk'], 1), 'day_off': best['dd'],
            'fine_glorys_same': round(f_same, 4), 'fine_glorys_best': round(fb, 4),
            'fine_ratio_same': round(f_same / fa, 3), 'fine_ratio_best': round(fb / fa, 3),
            'status': 'ok',
        })
        curves = {
            'profile_number': int(args['profile_number']),
            'lat': clat, 'lon': clon,
            'date': date.strftime('%Y-%m-%d'),
            'zc': zc, 'PTa': PTa, 'PTg_same': PTg_same, 'PTg_best': PTg_best,
            'anomaly_depth_m': args.get('anomaly_depth_m'),
            'disp_km': float(best['dk']), 'day_off': int(best['dd']),
            'fine_ratio_best': record['fine_ratio_best'], 'rms_best': record['rms_best'],
        }
    except Exception as exc:
        record['error'] = str(exc)
    return record, curves


def _glorys_residual_curve_worker(args: dict) -> dict | None:
    """单剖面对比图的廓线 worker：只取廓线（reuse 时按需补算缓存缺失的剖面用）。"""
    return _glorys_residual_core(args)[1]


def _save_residual_curves(path: str | Path, curves: list[dict | None]) -> None:
    """把所有剖面的三条潜温廓线压成单个 npz（共享 z 网格），供绘图零 GLORYS 读地复用。"""
    cur = [c for c in curves if c is not None]
    if not cur:
        return
    np.savez_compressed(
        path,
        zc=cur[0]['zc'],
        profile_number=np.array([c['profile_number'] for c in cur], int),
        lat=np.array([c['lat'] for c in cur], float),
        anomaly_depth_m=np.array([np.nan if c['anomaly_depth_m'] is None else c['anomaly_depth_m']
                                  for c in cur], float),
        disp_km=np.array([c['disp_km'] for c in cur], float),
        day_off=np.array([c['day_off'] for c in cur], int),
        fine_ratio_best=np.array([c['fine_ratio_best'] for c in cur], float),
        PTa=np.vstack([c['PTa'] for c in cur]),
        PTg_same=np.vstack([c['PTg_same'] for c in cur]),
        PTg_best=np.vstack([c['PTg_best'] for c in cur]),
    )


def _load_residual_curves(path: str | Path) -> dict[int, dict]:
    """读回 npz 廓线缓存，返回 {profile_number: curve_dict}；文件缺失返回空字典。"""
    if not Path(path).exists():
        return {}
    z = np.load(path)
    zc = z['zc']
    store = {}
    for k, pn in enumerate(z['profile_number']):
        store[int(pn)] = {
            'profile_number': int(pn), 'lat': float(z['lat'][k]),
            'anomaly_depth_m': float(z['anomaly_depth_m'][k]),
            'disp_km': float(z['disp_km'][k]), 'day_off': int(z['day_off'][k]),
            'fine_ratio_best': float(z['fine_ratio_best'][k]),
            'zc': zc, 'PTa': z['PTa'][k], 'PTg_same': z['PTg_same'][k], 'PTg_best': z['PTg_best'][k],
        }
    return store


def _residual_basin_of(lat: float, lon: float, so_lat: float = -40.0) -> str:
    """按经纬度粗分海盆；纬度低于 so_lat 一律归为南大洋（SO）。"""
    if lat < so_lat:
        return 'SO'
    if lat > 66:
        return 'Arctic'
    lon360 = lon % 360
    if 20 <= lon360 < 120 and lat <= 30:
        return 'Indian'
    if 100 <= lon360 < 290:
        return 'Pacific'
    return 'Atlantic'


def _render_glorys_residual_figures(df: pd.DataFrame, output_dir: Path, so_lat: float) -> list[str]:
    """由逐点残差明细 df 渲染 4 张聚合图，返回保存路径列表。"""
    from scipy.stats import mannwhitneyu

    df = df[df['status'] == 'ok'].copy()
    df['is_SO'] = df['lat'] < so_lat
    # hotspot_type 分类整数码 → 英文名：1=通风 2=隔离 3=OMZ（氧最小带，早期 Argo 曲线分类）
    df['htype'] = df['hotspot_type'].map({1: 'ventilated', 2: 'isolated', 3: 'OMZ'})
    df['basin'] = [_residual_basin_of(la, lo, so_lat) for la, lo in zip(df['lat'], df['lon'])]
    # 异常筛选已排除 ≤300m，故深度分箱从 300 起（无 <300 一列）
    df['dbin'] = pd.cut(df['anomaly_depth_m'].astype(float), [300, 500, 800, 5000],
                        labels=['300-500', '500-800', '>800'])
    n_total = len(df)
    saved: list[str] = []

    # Fig 1：细结构比分布（深度 / 海盆 / hotspot 分类）
    htype_order = [t for t in ['ventilated', 'isolated', 'OMZ'] if t in set(df['htype'].dropna())]
    fig, axes = plt.subplots(1, 3, figsize=(16, 5))
    for ax, key, order, ttl in [
        (axes[0], 'dbin', ['300-500', '500-800', '>800'], 'by anomaly depth (m)'),
        (axes[1], 'basin', ['Pacific', 'Atlantic', 'Indian', 'SO'], 'by basin'),
        (axes[2], 'htype', htype_order, 'by hotspot type')]:
        data = [df[df[key] == o]['fine_ratio_best'].clip(0, 2).dropna() for o in order]
        bp = ax.boxplot(data, labels=[str(o) for o in order], showmeans=True, patch_artist=True)
        for patch in bp['boxes']:
            patch.set_facecolor('#9ecae1')
            patch.set_alpha(0.7)
        ax.axhline(1.0, color='grey', ls='--', lw=1)
        ax.set_title(ttl)
        ax.set_ylabel('fine-structure ratio (GLORYS/Argo, best-match)')
        ax.set_ylim(0, 1.6)
        ax.grid(alpha=0.3, axis='y')
    fig.suptitle('GLORYS retained vertical fine-structure  (<1 = detail lost; n=%d)' % n_total, fontsize=13)
    fig.tight_layout()
    fp = output_dir / 'fig1_fine_ratio_distributions.png'
    fig.savefig(fp, dpi=140, bbox_inches='tight')
    plt.close(fig)
    saved.append(str(fp))

    # Fig 2：错位 vs 丢细节占比
    fig, ax = plt.subplots(1, 2, figsize=(13, 5.2))
    order = [b for b in ['Pacific', 'Atlantic', 'Indian', 'SO'] if b in set(df['basin'])]
    gb = df.groupby('basin')
    disp = [gb.get_group(b)['disp_rms'].mean() for b in order]
    resid = [gb.get_group(b)['rms_best'].mean() for b in order]
    x = np.arange(len(order))
    ax[0].bar(x, disp, label='displacement (quad)', color='#fdae6b')
    ax[0].bar(x, resid, bottom=disp, label='residual detail-loss', color='#3182bd')
    ax[0].set_xticks(x)
    ax[0].set_xticklabels(order)
    ax[0].set_ylabel('pot-temp deviation std (°C)')
    ax[0].set_title('deviation decomposition: displacement vs detail-loss')
    ax[0].legend()
    sc = ax[1].scatter(df['rms_same'], df['rms_best'], c=df['disp_km'], cmap='viridis', s=18, alpha=0.7)
    hi = float(np.nanpercentile(df['rms_same'], 99))
    ax[1].plot([0, hi], [0, hi], 'k--', lw=1, label='no displacement gain')
    ax[1].set_xlim(0, hi)
    ax[1].set_ylim(0, hi)
    ax[1].set_xlabel('same-loc std (total, °C)')
    ax[1].set_ylabel('best-match std (residual, °C)')
    ax[1].set_title('best-match below diagonal => displacement removed')
    plt.colorbar(sc, ax=ax[1], label='best-match displacement (km)')
    ax[1].legend()
    fig.tight_layout()
    fp = output_dir / 'fig2_displacement_vs_detailloss.png'
    fig.savefig(fp, dpi=140, bbox_inches='tight')
    plt.close(fig)
    saved.append(str(fp))

    # Fig 3：大残差世界地图
    fig = plt.figure(figsize=(15, 7))
    ax = plt.axes(projection=ccrs.PlateCarree(central_longitude=180))
    ax.add_feature(cfeature.LAND, facecolor=_BASEMAP_COLORS['land'])
    ax.coastlines(color=_BASEMAP_COLORS['coastline'], lw=0.4)
    ax.set_global()
    loss = 1 - df['fine_ratio_best'].clip(0, 1)
    sc = ax.scatter(df['lon'], df['lat'], c=loss, s=20 + df['rms_best'] * 40, cmap='inferno_r',
                    vmin=0, vmax=1, transform=ccrs.PlateCarree(), edgecolor='k', linewidth=0.2)
    ax.plot([-180, 180], [so_lat, so_lat], color='cyan', ls='--', lw=1.2, transform=ccrs.PlateCarree())
    plt.colorbar(sc, ax=ax, orientation='horizontal', pad=0.04, shrink=0.6,
                 label='detail loss = 1 - fine_ratio_best  (size ~ residual std)')
    ax.set_title('Where GLORYS loses vertical fine-structure (cyan = %.0fS, SO boundary)' % abs(so_lat))
    fp = output_dir / 'fig3_residual_map.png'
    fig.savefig(fp, dpi=140, bbox_inches='tight')
    plt.close(fig)
    saved.append(str(fp))

    # Fig 4：南大洋专项
    so = df[df['is_SO']]
    rest = df[~df['is_SO']]
    fig, ax = plt.subplots(1, 3, figsize=(15, 5))
    for a, col, ttl in [(ax[0], 'fine_ratio_best', 'fine-structure ratio'),
                        (ax[1], 'rms_best', 'residual std (°C)'),
                        (ax[2], 'disp_km', 'best-match displacement (km)')]:
        rest_v = rest[col].clip(0, 2) if 'ratio' in col else rest[col]
        so_v = so[col].clip(0, 2) if 'ratio' in col else so[col]
        a.hist(rest_v.dropna(), bins=25, density=True, alpha=0.5,
               label='rest (n=%d)' % len(rest), color='grey')
        a.hist(so_v.dropna(), bins=15, density=True, alpha=0.6,
               label='SO (n=%d)' % len(so), color='red')
        a.axvline(rest[col].median(), color='grey', ls='--')
        a.axvline(so[col].median(), color='red', ls='--')
        try:
            _, pval = mannwhitneyu(so[col].dropna(), rest[col].dropna(), alternative='two-sided')
        except Exception:
            pval = np.nan
        a.set_title('%s\nSO med=%.2f  rest med=%.2f  p=%.2g' %
                    (ttl, so[col].median(), rest[col].median(), pval))
        a.legend()
    fig.suptitle('Southern Ocean (<%.0fS) vs rest' % abs(so_lat), fontsize=13)
    fig.tight_layout()
    fp = output_dir / 'fig4_southern_ocean.png'
    fig.savefig(fp, dpi=140, bbox_inches='tight')
    plt.close(fig)
    saved.append(str(fp))
    return saved


def _render_glorys_residual_profile_examples(curves_list: list[dict], output_dir: Path) -> str | None:
    """单剖面对比图：每个面板叠 Argo（黑）/ GLORYS 同位（红虚）/ GLORYS 最佳匹配（绿虚）潜温廓线。"""
    curves_list = [c for c in curves_list if c is not None]
    if not curves_list:
        return None
    n = len(curves_list)
    ncol = min(4, n)
    nrow = int(np.ceil(n / ncol))
    fig, axes = plt.subplots(nrow, ncol, figsize=(4.2 * ncol, 4.6 * nrow), sharey=True, squeeze=False)
    for ax in axes.ravel()[n:]:
        ax.axis('off')
    for ax, c in zip(axes.ravel(), curves_list):
        zc = c['zc']
        ax.plot(c['PTa'], zc, 'k-', lw=1.5, label='Argo (obs)')
        ax.plot(c['PTg_same'], zc, 'r--', lw=1.2, label='GLORYS same-loc')
        ax.plot(c['PTg_best'], zc, 'g--', lw=1.2, label='GLORYS best-match')
        adep = c.get('anomaly_depth_m')
        if adep is not None and np.isfinite(adep):
            ax.axhline(float(adep), color='blue', ls=':', lw=1, label='anomaly depth')
        ax.set_ylim(zc.max(), 0)
        ax.set_xlabel('potential temp (°C)')
        ax.grid(alpha=0.3)
        # 每格各自的位移/天偏移进标题（单一 legend 无法逐格显示）
        ax.set_title("P%d (%.1f°N)  kept=%.2f  disp=%.0fkm/%+dd" %
                     (c['profile_number'], c['lat'], c['fine_ratio_best'], c['disp_km'], c['day_off']),
                     fontsize=10)
    for r in range(nrow):
        axes[r, 0].set_ylabel('depth (m)')
    axes.ravel()[0].legend(fontsize=8, loc='upper left')
    fig.suptitle('Argo vs GLORYS profiles  (same-loc=red, best-match within search window=green)', fontsize=13)
    fig.tight_layout()
    fp = output_dir / 'fig5_profile_examples.png'
    fig.savefig(fp, dpi=140, bbox_inches='tight')
    plt.close(fig)
    return str(fp)


def plot_glorys_detail_loss_residual_atlas(
    start_year: int | None = None,
    end_year: int | None = None,
    anomalies_path: str | Path | None = None,
    detection_config: DetectionConfig | None = None,
    *,
    z_max_m: float = _glorys_resid_z_max_m,
    z_spacing_m: float = _glorys_resid_z_spacing_m,
    fine_struct_window: int = _glorys_resid_fine_window,
    sameloc_radius_km: float = _glorys_resid_sameloc_radius_km,
    match_radius_km: float = _glorys_resid_match_radius_km,
    match_window_deg: float = _glorys_resid_match_window_deg,
    match_window_days: int = _glorys_resid_match_window_days,
    match_min_depth_coverage: float = _glorys_resid_match_min_cov,
    so_lat_threshold: float = _glorys_resid_so_lat,
    argo_data_dir: str | Path | None = None,
    output_dir: str | Path | None = None,
    rows_path: str | Path | None = None,
    reuse_rows: bool = False,
    make_figures: bool = True,
    save_fig: bool = True,
    profile_examples: list[int] | None = None,
    n_example_profiles: int = 8,
    use_multiprocessing: bool = True,
    num_workers: int | None = None,
    maxtasksperchild: int | None = None,
    verbose: bool = False,
    return_details: bool = False,
) -> dict:
    """逐点统计 GLORYS 相对原始 Argo 剖面丢失的垂向细结构，并出全量聚合图。

    对每条 hotspot 异常剖面，把 GLORYS 插值到其 (lon,lat,date,depth) 与原始 Argo 潜温
    廓线逐点比较：``rms_same`` 为同位总偏差；在 ±``match_window_deg``/±``match_window_days``
    窗口内搜最小 RMS 的原生 GLORYS 列得 ``rms_best``，二者按二次方差分离出错位贡献
    （``disp_rms`` / ``disp_km``）；剔除错位后的高通细结构比 ``fine_ratio_best`` <1 即
    GLORYS 真丢细节。即便单条剖面也成立，故在 Argo 稀疏、重建不可行处（尤其南大洋）仍适用。
    聚合图按深度/海盆/Type 给出细结构比分布、错位 vs 丢细节占比、大残差世界地图，并单独
    对比南大洋；另出一张单剖面对比图，逐条叠 Argo / GLORYS 同位 / GLORYS 最佳匹配三条潜温廓线。

    参数:
        - start_year (int | None): anomalies_path=None 时用于定位默认 anomalies 文件。
        - end_year (int | None): 同上，结束年份。
        - anomalies_path (str | Path | None): anomalies parquet 路径；None 时按 plot_argo_hotspots 命名规则自动定位。
        - detection_config (DetectionConfig | None): 异常识别配置；决定 method/region 与默认输入输出路径。
        - z_max_m (float): 残差比较的最大深度（m），默认来自 processing.yml。
        - z_spacing_m (float): 潜温廓线统一插值的垂向间距（m），默认来自配置。
        - fine_struct_window (int): 高通细结构的滚动窗口点数（×z_spacing_m 为物理尺度），默认来自配置。
        - sameloc_radius_km (float): 同位 GLORYS 读取窗口半宽（km），默认来自配置。
        - match_radius_km (float): 最佳匹配搜索的 GLORYS 读取窗口半宽（km），默认来自配置。
        - match_window_deg (float): 最佳匹配的经纬度搜索半宽（°），默认来自配置。
        - match_window_days (int): 最佳匹配的时间搜索半宽（天），默认来自配置。
        - match_min_depth_coverage (float): 候选 GLORYS 列须覆盖 Argo 廓线深度范围的最低比例，剔除浅海床截断列以截断深度刷低 RMS 误胜的伪匹配，默认来自配置。
        - so_lat_threshold (float): 南大洋专项分界纬度（°），低于此值归为 SO，默认来自配置。
        - argo_data_dir (str | Path | None): Argo 年数据目录；None 时使用配置默认路径。
        - output_dir (str | Path | None): 批处理专属输出根目录；None 时使用当前 method/region 默认目录。
        - rows_path (str | Path | None): 逐剖面明细 parquet 路径；None 时落在 output_dir 下默认文件名；同目录另存同名 `_curves.npz`（全部廓线）供重绘单剖面对比图时零 GLORYS 读地复用。
        - reuse_rows (bool): 明细 parquet 已存在时直接复用、跳过计算（仅重绘图时用），此时单剖面图从 `_curves.npz` 取、缺失剖面才补算，默认 False。
        - make_figures (bool): 是否渲染聚合图，默认 True。
        - save_fig (bool): 是否把图保存到 output_dir，默认 True。
        - profile_examples (list[int] | None): 单剖面对比图（Argo/同位/最佳匹配三条廓线）要画的 Profile_number 列表；None 时按 fine_ratio_best 等距自动选样。
        - n_example_profiles (int): 自动选样时的剖面数（按细结构比从最差到最好等距取），<=0 关闭该图，默认 8。
        - use_multiprocessing (bool): 是否多进程并行计算 profile，默认 True。
        - num_workers (int | None): worker 数；None 时自动取 min(profile数, CPU数, 12)。
        - maxtasksperchild (int | None): 每个 worker 处理多少任务后重启；None 表示不重启（保留 Argo 年缓存）。
        - verbose (bool): 是否打印每条 profile 的失败明细，默认 False。
        - return_details (bool): 是否在返回中附带逐 profile 结果 DataFrame，默认 False。

    返回:
        - dict: 含 total_candidates/processed_profiles/skipped_profiles/rows_path/output_dir/figures/southern_ocean 等摘要；return_details=True 时附 'rows'（DataFrame）。
    """
    cfg = _resolve_detection_config(detection_config)
    method_name = cfg.method
    run_tag = cfg.file_stem()
    if argo_data_dir is None:
        argo_data_dir = argo_path
    region_slug = _current_region_key()
    region_config_key = _current_region_config_key()
    batch_output_dir = (
        Path(output_dir)
        if output_dir is not None
        else cfg.output_dir("plot_glorys_detail_loss_residual_atlas", region_slug)
    )

    if anomalies_path is None:
        if start_year is None or end_year is None:
            raise ValueError("anomalies_path 为空时，必须提供 start_year 与 end_year。")
        anomalies_path = cfg.output_dir("plot_argo_hotspots", region_slug) / (
            f"anomalies_{start_year}_{end_year}_{run_tag}.parquet"
        )
    else:
        anomalies_path = Path(anomalies_path)
    if not Path(anomalies_path).exists():
        raise FileNotFoundError(f"Anomalies file not found: {anomalies_path}")

    anomalies = pd.read_parquet(anomalies_path)
    if 'detection_method' in anomalies.columns:
        method_mask = anomalies['detection_method'].astype(str).str.lower().eq(method_name)
        anomalies = anomalies[method_mask].copy()
    required_cols = ['Year', 'Profile_number', 'Longitude', 'Latitude', 'Month', 'Day']
    missing_cols = [c for c in required_cols if c not in anomalies.columns]
    if missing_cols:
        raise ValueError(f"Anomalies file missing required columns: {missing_cols}")

    batch_output_dir.mkdir(parents=True, exist_ok=True)
    resolved_rows_path = (
        Path(rows_path) if rows_path is not None
        else batch_output_dir / f"glorys_residual_rows_{run_tag}.parquet"
    )
    resolved_curves_path = Path(resolved_rows_path).with_name(
        Path(resolved_rows_path).stem.replace('_rows', '_curves') + '.npz')

    depth_cols = ['heave_projection_depth_m', 'depth', 'Depth', 'Anomaly_depth']

    def _arg_for_row(task_index: int, row: pd.Series) -> dict:
        adep = next((float(row[c]) for c in depth_cols
                     if c in anomalies.columns and pd.notna(row.get(c))), np.nan)
        return {
            'task_index': int(task_index),
            'year': int(row['Year']),
            'profile_number': int(row['Profile_number']),
            'month': int(row['Month']),
            'day': int(row['Day']),
            'center_lon': float(row['Longitude']),
            'center_lat': float(row['Latitude']),
            'anomaly_depth_m': adep,
            'argo_data_dir': argo_data_dir,
            'region_config_key': region_config_key,
            'z_max_m': z_max_m, 'z_spacing_m': z_spacing_m,
            'fine_window': fine_struct_window,
            'sameloc_radius_km': sameloc_radius_km,
            'match_radius_km': match_radius_km,
            'match_window_deg': match_window_deg,
            'match_window_days': match_window_days,
            'match_min_depth_coverage': match_min_depth_coverage,
        }

    if reuse_rows and resolved_rows_path.exists():
        rows_df = pd.read_parquet(resolved_rows_path)
        curves_store = _load_residual_curves(resolved_curves_path)
        print(f"[*] GLORYS residual: reusing existing rows ({len(rows_df)}) from {resolved_rows_path}"
              f"{' + curves cache' if curves_store else ''}")
    else:
        # 按年份排序后分块派发，使每个 worker 处理相邻年份、Argo 缓存（上限 _RESID_ARGO_CACHE_MAX 年）命中率高
        worker_args = [_arg_for_row(i, row) for i, (_, row)
                       in enumerate(anomalies.sort_values('Year', kind='stable').iterrows())]

        total_candidates = len(worker_args)
        worker_count = 1
        if use_multiprocessing and total_candidates > 1:
            worker_count = int(num_workers) if num_workers is not None else min(total_candidates, os.cpu_count() or 1, 12)
            worker_count = max(1, worker_count)

        if worker_count > 1:
            print(f"[*] GLORYS residual multiprocessing: workers={worker_count}, profiles={total_candidates}.")
            pool_kwargs = {'processes': worker_count}
            if maxtasksperchild is not None:
                pool_kwargs['maxtasksperchild'] = max(1, int(maxtasksperchild))
            with multiprocessing.Pool(**pool_kwargs) as pool:
                pairs = list(tqdm(
                    pool.imap_unordered(_glorys_residual_core, worker_args, chunksize=8),
                    total=total_candidates, desc="glorys residual", unit="profile"))
            pairs = sorted(pairs, key=lambda rc: int(rc[0].get('task_index', 0)))
        else:
            pairs = [
                _glorys_residual_core(a)
                for a in tqdm(worker_args, total=total_candidates, desc="glorys residual", unit="profile")
            ]
        results = [rc[0] for rc in pairs]
        curves_store = {c['profile_number']: c for _, c in pairs if c is not None}

        if verbose:
            for item in results:
                if item.get('status') == 'failed':
                    print(f"[WARN] residual failed Year={item.get('year')} "
                          f"Profile={item.get('profile_number')}: {item.get('error')}")
        rows_df = pd.DataFrame(results)
        # 合并 hotspot 分类后再存盘，让 parquet 自洽（供聚合图分组与后续单独分析）
        rows_df = rows_df.merge(
            anomalies[['Profile_number', 'hotspot_type']].rename(columns={'Profile_number': 'profile_number'}),
            on='profile_number', how='left')
        rows_df.to_parquet(resolved_rows_path, index=False)
        _save_residual_curves(resolved_curves_path, list(curves_store.values()))
        print(f"[*] GLORYS residual rows saved: {resolved_rows_path}")

    # reuse 旧 rows 缺 hotspot_type 列时补上（新算的已在上面带好）
    if 'hotspot_type' not in rows_df.columns:
        rows_df = rows_df.merge(
            anomalies[['Profile_number', 'hotspot_type']].rename(columns={'Profile_number': 'profile_number'}),
            on='profile_number', how='left')

    ok = rows_df[rows_df['status'] == 'ok']
    processed_profiles = int(len(ok))
    skipped_profiles = int((rows_df['status'] != 'ok').sum())

    figures: list[str] = []
    if make_figures and save_fig and processed_profiles > 0:
        figures = _render_glorys_residual_figures(rows_df, batch_output_dir, so_lat_threshold)

        # 单剖面对比图：显式 profile_examples 优先，否则按 fine_ratio_best 等距取 n 条（最差→最好）
        example_pns: list[int] = []
        if profile_examples:
            example_pns = [int(p) for p in profile_examples]
        elif n_example_profiles and n_example_profiles > 0:
            okx = ok.dropna(subset=['fine_ratio_best']).sort_values('fine_ratio_best')
            if len(okx):
                idx = np.linspace(0, len(okx) - 1, min(int(n_example_profiles), len(okx))).round().astype(int)
                example_pns = okx.iloc[np.unique(idx)]['profile_number'].astype(int).tolist()
        if example_pns:
            # 优先用 npz 廓线缓存（零 GLORYS 读）；仅对缓存缺失的剖面按需补算
            example_curves = [curves_store[pn] for pn in example_pns if pn in curves_store]
            missing = [pn for pn in example_pns if pn not in curves_store]
            if missing:
                pn_to_row = {int(r['Profile_number']): r
                             for _, r in anomalies.drop_duplicates('Profile_number').iterrows()}
                miss_args = [_arg_for_row(-1, pn_to_row[pn]) for pn in missing if pn in pn_to_row]
                recomputed = [_glorys_residual_curve_worker(a) for a in miss_args]
                example_curves += [c for c in recomputed if c is not None]
            ex_fig = _render_glorys_residual_profile_examples(example_curves, batch_output_dir)
            if ex_fig:
                figures.append(ex_fig)
        print(f"[*] GLORYS residual figures saved to {batch_output_dir}")

    so_mask = ok['lat'] < so_lat_threshold
    summary = {
        'total_candidates': int(len(rows_df)),
        'processed_profiles': processed_profiles,
        'skipped_profiles': skipped_profiles,
        'rows_path': str(resolved_rows_path),
        'output_dir': str(batch_output_dir),
        'anomalies_path': str(anomalies_path),
        'figures': figures,
        'fine_ratio_best_median': float(ok['fine_ratio_best'].median()) if processed_profiles else None,
        'southern_ocean': {
            'n': int(so_mask.sum()),
            'fine_ratio_best_median': float(ok.loc[so_mask, 'fine_ratio_best'].median()) if so_mask.any() else None,
            'rest_fine_ratio_best_median': float(ok.loc[~so_mask, 'fine_ratio_best'].median()) if (~so_mask).any() else None,
        },
    }
    if return_details:
        summary['rows'] = rows_df
    return summary


def export_hotspot_anomaly_summary_table(
    vertical_profiles_result: dict | None = None,
    argo_glorys_result: dict | None = None,
    start_year: int | None = None,
    end_year: int | None = None,
    anomalies_path: str | Path | None = None,
    detection_config: DetectionConfig | None = None,
    *,
    argo_glorys_summary_data_path: str | Path | None = None,
    output_path: str | Path | None = None,
    output_format: str | None = None,
    save_table: bool = True,
    heave_z_threshold: float | None = _heave_depth_threshold,
    diagnose_nearshore_do: bool = True,
    nearshore_do_min_threshold: float = 50.0,
    nearshore_do_drop_threshold: float = 100.0,
    nearshore_do_recovery_threshold: float = 100.0,
    nearshore_do_depth_gap_threshold_m: float = 100.0,
    compact_columns: bool = True,
    require_glorys_details: bool = False,
    verbose: bool = True,
) -> dict:
    """导出 hotspot anomaly 检查用 summary 表。

    该函数不重新做异常检测或 GLORYS 绘图。基础剖面信息与异常指标来自 `plot_argo_hotspots` 保存的
    anomalies parquet；OI 与 GLORYS 绘图状态优先来自 `plot_hotspot_anomaly_argo_glorys_overviews`
    的返回值，若未传返回值则读取该函数默认保存的 summary parquet。若 heave 详情缺失，会在表中保留
    空值并给出提示。近岸 DO 形态指标优先读取 `plot_argo_hotspots` 新版 anomalies parquet 中已保存的
    列；旧版 parquet 缺列时才回查 Argo 年数据。

    `hotspot_type` 用于快速人工复核，输出为类别名（与整数码一一对应，整数码本身仍保留在 anomalies
    parquet）：`OMZ`（原 3）表示 Argo DO 剖面呈近岸型先快速降低再回升的形态；非近岸剖面中，heave
    通风型为 `ventilated`（原 1），否则为深层隔离型 `isolated`（原 2）。`spice_type` 是与 `hotspot_type`
    正交的 T-S 水团轴，与之同构：整数码（`1` cold-fresh / `2` background-consistent / `3` warm-salty，
    阈值 `detection_config.spice_percentile_threshold`）由 `plot_hotspot_anomaly_argo_glorys_overviews`
    写进 GLORYS overview summary parquet，导出时仅读码映射成名称；二者叉乘即 T-S × 通风四宫格。导出表
    把 `hotspot_type`、`spice_type` 连同身份-定位列前置到表头，其余为支撑证据列。

    参数:
        - vertical_profiles_result (dict | None): `plot_hotspot_anomaly_vertical_profiles` 的返回值。仅用于兼容旧调用：若含 `anomalies_path`，可用来定位输入 parquet。
        - argo_glorys_result (dict | None): `plot_hotspot_anomaly_argo_glorys_overviews` 的返回值。若含 `results`，会合并 GLORYS 状态和 OI 字段。
        - start_year / end_year / anomalies_path: 当返回值中没有 anomalies_path 时用于定位输入 parquet。
        - detection_config (DetectionConfig | None): 异常识别配置；None 时使用 processing.yml 默认值。
        - argo_glorys_summary_data_path (str | Path | None): `plot_hotspot_anomaly_argo_glorys_overviews` 保存的明细 parquet。
        - output_path (str | Path | None): 保存路径；None 时写到当前 method/region 的默认 summary 目录。
        - output_format (str | None): `'csv'` 或 `'xlsx'`；None 时从 output_path 后缀推断，默认 `'csv'`。
        - save_table (bool): 是否保存表格文件，默认 True。
        - heave_z_threshold (float | None): 通风深度阈值 (m)，用于 hotspot_type 分类；None 时不新增对应列。
        - diagnose_nearshore_do (bool): 是否基于 Argo DO 剖面诊断近岸型先降后升曲线，默认 True。
        - nearshore_do_min_threshold (float): 近岸型判据中，异常深度以上 DO 最小值阈值。
        - nearshore_do_drop_threshold (float): 近岸型判据中，表层参考 DO 到最小 DO 的最小降幅。
        - nearshore_do_recovery_threshold (float): 近岸型判据中，最小 DO 到异常深度 DO 的最小回升幅度。
        - nearshore_do_depth_gap_threshold_m (float): 近岸型判据中，最小 DO 与异常深度的最小垂向间隔。
        - compact_columns (bool): 是否删除人工检查中冗余的固定/重复列，默认 True。
        - require_glorys_details (bool): True 时若缺少 GLORYS heave/OI 信息则报错。
        - verbose (bool): 是否打印保存路径和缺失提示。
    返回:
        - dict: 包含 `summary_table`、`output_path`、`n_rows`、`warnings`。
    """

    def _result_path(result: dict | None, key: str) -> Path | None:
        if not isinstance(result, dict):
            return None
        val = result.get(key)
        if val is None or (isinstance(val, float) and pd.isna(val)):
            return None
        return Path(val)

    def _to_int_or_none(val) -> int | None:
        try:
            if pd.isna(val):
                return None
            return int(val)
        except Exception:
            return None

    def _date_key_from_parts(year_val, month_val=None, day_val=None) -> str | None:
        try:
            if month_val is None or day_val is None or pd.isna(month_val) or pd.isna(day_val):
                return None
            return pd.Timestamp(
                year=int(year_val),
                month=int(month_val),
                day=int(day_val),
            ).strftime('%Y-%m-%d')
        except Exception:
            return None

    def _date_key_from_value(val) -> str | None:
        try:
            if val is None or pd.isna(val):
                return None
        except Exception:
            pass
        try:
            return pd.Timestamp(val).normalize().strftime('%Y-%m-%d')
        except Exception:
            return None

    def _record_keys(year_val, profile_val, date_key=None, platform_val=None) -> list[tuple]:
        year_key = _to_int_or_none(year_val)
        profile_key = _to_int_or_none(profile_val)
        if year_key is None or profile_key is None:
            return []
        platform_key = _to_int_or_none(platform_val)
        keys = []
        if date_key and platform_key is not None:
            keys.append((year_key, profile_key, date_key, platform_key))
        if date_key:
            keys.append((year_key, profile_key, date_key, None))
        if platform_key is not None:
            keys.append((year_key, profile_key, None, platform_key))
        keys.append((year_key, profile_key, None, None))
        return keys

    def _infer_years_from_path(path_obj: Path | None) -> tuple[int | None, int | None]:
        if path_obj is None:
            return None, None
        match = re.search(r'anomalies_(\d{4})_(\d{4})_', path_obj.name)
        if not match:
            return None, None
        return int(match.group(1)), int(match.group(2))

    def _copy_first_available(src: pd.DataFrame, dst: pd.DataFrame, target: str, names: list[str]):
        for name in names:
            if name in src.columns:
                dst[target] = src[name]
                return
        dst[target] = np.nan

    def _read_first_existing_parquet(paths: list[Path]) -> tuple[list[dict], Path | None, str | None]:
        for path_obj in paths:
            if path_obj is None:
                continue
            path_obj = Path(path_obj)
            if not path_obj.exists():
                continue
            try:
                df_obj = pd.read_parquet(path_obj)
                return df_obj.to_dict('records'), path_obj, None
            except Exception as exc:
                return [], path_obj, str(exc)
        return [], None, None

    def _series_constant(series: pd.Series) -> bool:
        vals = series.dropna()
        if vals.empty:
            return True
        return vals.astype(str).nunique(dropna=True) <= 1

    def _numeric_columns_equal(left: pd.Series, right: pd.Series) -> bool:
        left_vals = pd.to_numeric(left, errors='coerce')
        right_vals = pd.to_numeric(right, errors='coerce')
        both_nan = left_vals.isna() & right_vals.isna()
        comparable = ~(both_nan)
        if not comparable.any():
            return True
        return bool(np.allclose(
            left_vals[comparable].to_numpy(dtype=float),
            right_vals[comparable].to_numpy(dtype=float),
            equal_nan=True,
            rtol=0.0,
            atol=1e-8,
        ))

    def _compact_summary_columns(table_in: pd.DataFrame) -> tuple[pd.DataFrame, list[str]]:
        table_out = table_in.copy()
        drop_cols: list[str] = []

        constant_drop_candidates = [
            'detection_method',
            'depth_interval_m',
            'glorys_status',
            'glorys_error',
            'glorys_horizontal_status',
            'glorys_vertical_status',
            'glorys_line_strategy',
            'glorys_k',
            'heave_x_window_km',
            'heave_z_window_m',
            'heave_error',
            'heave_z_threshold',
            'primary_metric',
            'do_threshold',
            'anomaly_min_depth_m',
            'anomaly_max_depth_m',
        ]
        for col_name in constant_drop_candidates:
            if col_name in table_out.columns and _series_constant(table_out[col_name]):
                drop_cols.append(col_name)

        # nearshore_do_dip ≡ (hotspot_type == 3) — 保留后者
        if ('nearshore_do_dip' in table_out.columns
                and 'hotspot_type' in table_out.columns):
            drop_cols.append('nearshore_do_dip')

        duplicate_pairs = [
            ('heave_projection_depth_m', 'anomaly_depth_m'),
            ('glorys_center_lon', 'Longitude'),
            ('glorys_center_lat', 'Latitude'),
            ('glorys_b', 'Latitude'),
            # primary_value/anomaly_score 是首选指标的通用镜像，与具体 delta_<metric> 数值重复
            ('primary_value', 'delta_do'),
            ('primary_value', 'delta_aou'),
            ('anomaly_score', 'delta_do'),
            ('anomaly_score', 'delta_aou'),
        ]
        for duplicate_col, source_col in duplicate_pairs:
            if (
                duplicate_col in table_out.columns
                and source_col in table_out.columns
                and _numeric_columns_equal(table_out[duplicate_col], table_out[source_col])
            ):
                drop_cols.append(duplicate_col)

        # Profile_number 全局唯一、date 已含年月日 → 精简模式 Year/Month/Day 冗余，只留可读的 date
        if 'date' in table_out.columns:
            drop_cols.extend(c for c in ('Year', 'Month', 'Day') if c in table_out.columns)

        if drop_cols:
            drop_cols = [c for c in dict.fromkeys(drop_cols) if c in table_out.columns]
            table_out = table_out.drop(columns=drop_cols)
        return table_out, drop_cols

    def _do_profile_shape_metrics(profile_rows: pd.DataFrame, target_depth, target_do) -> dict:
        return _nearshore_do_profile_shape_metrics(
            profile_rows,
            target_depth,
            target_do,
            nearshore_do_min_threshold=nearshore_do_min_threshold,
            nearshore_do_drop_threshold=nearshore_do_drop_threshold,
            nearshore_do_recovery_threshold=nearshore_do_recovery_threshold,
            nearshore_do_depth_gap_threshold_m=nearshore_do_depth_gap_threshold_m,
        )

    def _append_nearshore_do_diagnostics(table_in: pd.DataFrame) -> pd.DataFrame:
        table_out = table_in.copy()
        metric_cols = [
            'surface_do_ref',
            'pre_anomaly_do_min',
            'pre_anomaly_do_min_depth_m',
            'surface_to_min_do_drop',
            'min_to_anomaly_do_recovery',
            'min_to_anomaly_depth_gap_m',
            'do_v_shape_score',
            'nearshore_do_dip',
        ]
        for col_name in metric_cols:
            table_out[col_name] = np.nan
        table_out['nearshore_do_dip'] = pd.Series(pd.NA, index=table_out.index, dtype='boolean')

        required = {'Year', 'Month', 'Day', 'Profile_number', 'Platform_number', 'anomaly_depth_m'}
        if not required.issubset(table_out.columns):
            warnings_list.append("缺少 Argo DO 近岸诊断所需列，已跳过 nearshore_do_dip。")
            return table_out

        for year_val in sorted(pd.to_numeric(table_out['Year'], errors='coerce').dropna().astype(int).unique()):
            year_mask = pd.to_numeric(table_out['Year'], errors='coerce') == year_val
            year_rows = table_out.loc[year_mask].copy()
            if year_rows.empty:
                continue
            argo_file = Path(argo_path) / f'Argo{year_val}.parquet'
            if not argo_file.exists():
                warnings_list.append(f"Argo file not found for nearshore DO diagnostics: {argo_file}")
                continue

            try:
                parquet_cols = set(pq.read_schema(argo_file).names)
                read_cols = [
                    c for c in [
                        'Year', 'Month', 'Day', 'Depth', 'depth',
                        'DO', 'DOXY_Adjusted', 'DOXY', 'DO_Adjusted', 'DO_Raw',
                        'Profile_number', 'Platform_number',
                    ]
                    if c in parquet_cols
                ]
                argo_year = pd.read_parquet(argo_file, columns=read_cols)
            except Exception:
                try:
                    argo_year = load_argo_data(year_val, data_dir=argo_path, verbose=False)
                except Exception as exc:
                    warnings_list.append(f"读取 Argo{year_val} 失败，已跳过 nearshore DO diagnostics: {exc}")
                    continue

            if argo_year.empty:
                continue
            for col_name in ['Year', 'Month', 'Day', 'Profile_number', 'Platform_number']:
                if col_name in argo_year.columns:
                    argo_year[col_name] = pd.to_numeric(argo_year[col_name], errors='coerce').astype('Int64')

            needed_keys = set()
            for _, row in year_rows.iterrows():
                needed_keys.add((
                    _to_int_or_none(row.get('Profile_number')),
                    _to_int_or_none(row.get('Platform_number')),
                    _to_int_or_none(row.get('Month')),
                    _to_int_or_none(row.get('Day')),
                ))

            if {'Profile_number', 'Platform_number', 'Month', 'Day'}.issubset(argo_year.columns):
                key_series = list(zip(
                    argo_year['Profile_number'],
                    argo_year['Platform_number'],
                    argo_year['Month'],
                    argo_year['Day'],
                ))
                argo_year = argo_year.loc[[key in needed_keys for key in key_series]].copy()
            if argo_year.empty:
                continue

            grouped = {
                key: grp
                for key, grp in argo_year.groupby(
                    ['Profile_number', 'Platform_number', 'Month', 'Day'],
                    dropna=False,
                    sort=False,
                )
            }
            for idx, row in year_rows.iterrows():
                key = (
                    _to_int_or_none(row.get('Profile_number')),
                    _to_int_or_none(row.get('Platform_number')),
                    _to_int_or_none(row.get('Month')),
                    _to_int_or_none(row.get('Day')),
                )
                profile_rows = grouped.get(key)
                if profile_rows is None:
                    continue
                metrics = _do_profile_shape_metrics(
                    profile_rows,
                    row.get('anomaly_depth_m'),
                    row.get('do_value'),
                )
                for col_name, value in metrics.items():
                    table_out.at[idx, col_name] = value

        return table_out

    cfg = _resolve_detection_config(detection_config)
    method_name = cfg.method
    run_tag = cfg.file_stem()
    region_slug = _current_region_key()
    warnings_list: list[str] = []

    candidate_paths = []
    if anomalies_path is not None:
        candidate_paths.append(Path(anomalies_path))
    for result in (vertical_profiles_result, argo_glorys_result):
        path_val = _result_path(result, 'anomalies_path')
        if path_val is not None:
            candidate_paths.append(path_val)

    if not candidate_paths and start_year is not None and end_year is not None:
        candidate_paths.append(
            cfg.output_dir("plot_argo_hotspots", region_slug)
            / f"anomalies_{start_year}_{end_year}_{run_tag}.parquet"
        )

    if not candidate_paths:
        raise ValueError(
            "缺少 anomalies_path。请先运行 plot_argo_hotspots，"
            "或显式传入 anomalies_path/start_year/end_year。"
        )

    resolved_anomalies_path = candidate_paths[0]
    for path_val in candidate_paths[1:]:
        if Path(path_val) != resolved_anomalies_path:
            warnings_list.append(
                f"收到多个 anomalies_path，使用 {resolved_anomalies_path}，忽略 {path_val}。"
            )
            break

    if not resolved_anomalies_path.exists():
        raise FileNotFoundError(f"Anomalies file not found: {resolved_anomalies_path}")

    inferred_start, inferred_end = _infer_years_from_path(resolved_anomalies_path)
    start_tag = start_year if start_year is not None else inferred_start
    end_tag = end_year if end_year is not None else inferred_end
    year_tag = f"{start_tag}_{end_tag}" if start_tag is not None and end_tag is not None else "custom"

    anomalies = pd.read_parquet(resolved_anomalies_path)
    if 'detection_method' in anomalies.columns:
        method_mask = anomalies['detection_method'].astype(str).str.lower().eq(method_name)
        mismatch_count = int((~method_mask).sum())
        if mismatch_count > 0:
            warnings_list.append(
                f"anomalies 中存在非 {method_name} 方法记录，已过滤 {mismatch_count} 行。"
            )
        anomalies = anomalies[method_mask].copy()

    if anomalies.empty:
        warnings_list.append(f"No anomalies in file: {resolved_anomalies_path}")

    required_cols = ['Year', 'Profile_number']
    missing_cols = [c for c in required_cols if c not in anomalies.columns]
    if missing_cols:
        raise ValueError(f"Anomalies file missing required columns: {missing_cols}")

    work = anomalies.copy()
    work['_year'] = pd.to_numeric(work['Year'], errors='coerce')
    work['_profile'] = pd.to_numeric(work['Profile_number'], errors='coerce')
    work = work.dropna(subset=['_year', '_profile']).copy()
    work['_year'] = work['_year'].astype(int)
    work['_profile'] = work['_profile'].astype(int)

    table = pd.DataFrame(index=work.index)
    table['detection_method'] = str(method_name)
    table['date'] = [
        _date_key_from_parts(row.get('Year'), row.get('Month'), row.get('Day'))
        for _, row in work.iterrows()
    ]
    _copy_first_available(work, table, 'Year', ['Year'])
    _copy_first_available(work, table, 'Month', ['Month'])
    _copy_first_available(work, table, 'Day', ['Day'])
    _copy_first_available(work, table, 'Profile_number', ['Profile_number'])
    _copy_first_available(work, table, 'Platform_number', ['Platform_number'])
    _copy_first_available(work, table, 'Longitude', ['Longitude'])
    _copy_first_available(work, table, 'Latitude', ['Latitude'])
    _copy_first_available(work, table, 'anomaly_depth_m', ['depth', 'Depth', 'Anomaly_depth'])

    for col_name in [
        'primary_metric',
        'primary_value',
        'anomaly_score',
        'delta_do',
        'do_value',
        'delta_aou',
        'aou_value',
        'delta_pi',
        'pi_value',
        'trim_score',
        'trim_scale_res_rob_aou',
        'trim_scale_res_rob_abs_sal',
        'delta_temperature',
        'delta_salinity',
    ]:
        if col_name in work.columns:
            table[col_name] = work[col_name]

    for col_name in _NEARSHORE_DO_DIAGNOSTIC_COLUMNS:
        if col_name in work.columns:
            table[col_name] = work[col_name]

    table['anomaly_min_depth_m'] = (
        float(cfg.anomaly_min_depth)
        if cfg.anomaly_min_depth is not None else np.nan
    )
    table['anomaly_max_depth_m'] = (
        float(cfg.anomaly_max_depth)
        if cfg.anomaly_max_depth is not None else np.nan
    )
    table['depth_interval_m'] = float(cfg.depth_interval)

    if method_name == 'do':
        table['do_threshold'] = float(cfg.do_threshold)
        if cfg.salinity_threshold is not None and float(cfg.salinity_threshold) > 0:
            table['salinity_threshold'] = float(cfg.salinity_threshold)
        if cfg.temperature_threshold is not None and float(cfg.temperature_threshold) > 0:
            table['temperature_threshold'] = float(cfg.temperature_threshold)
    elif method_name == 'aou':
        table['aou_threshold'] = float(cfg.aou_threshold)
        table['pi_threshold'] = float(cfg.pi_threshold)
        table['aou_pi_depth_tolerance_m'] = float(cfg.aou_pi_depth_tolerance)
    elif method_name == 'trim':
        table['trim_cutoff_sigma'] = float(cfg.trim_cutoff)
        table['trim_depth_min_m'] = float(cfg.trim_depth_min)
        table['trim_depth_max_m'] = float(cfg.trim_depth_max)

    glorys_records = []
    glorys_summary_path_used = None
    if isinstance(argo_glorys_result, dict) and isinstance(argo_glorys_result.get('results'), list):
        glorys_records = argo_glorys_result.get('results') or []
        glorys_summary_path_used = _result_path(argo_glorys_result, 'summary_data_path')
    else:
        glorys_summary_candidates: list[Path] = []
        if argo_glorys_summary_data_path is not None:
            glorys_summary_candidates.append(Path(argo_glorys_summary_data_path))
        result_glorys_summary_path = _result_path(argo_glorys_result, 'summary_data_path')
        if result_glorys_summary_path is not None:
            glorys_summary_candidates.append(result_glorys_summary_path)
        glorys_summary_candidates.append(
            cfg.output_dir("plot_hotspot_anomaly_argo_glorys_overviews", region_slug)
            / f"hotspot_anomaly_argo_glorys_overviews_summary_{year_tag}_{run_tag}.parquet"
        )
        glorys_records, glorys_summary_path_used, glorys_summary_error = _read_first_existing_parquet(
            glorys_summary_candidates
        )
        if glorys_summary_error is not None:
            warnings_list.append(
                f"读取 GLORYS summary parquet 失败: {glorys_summary_path_used}: {glorys_summary_error}"
            )
        if not glorys_records and len(table) > 0:
            if glorys_summary_path_used is None and argo_glorys_result is None:
                warnings_list.append(
                    "未收到 plot_hotspot_anomaly_argo_glorys_overviews 的返回值，也未找到默认 GLORYS summary parquet；"
                    "OI 与 GLORYS 绘图状态将为空。"
                )
            elif glorys_summary_path_used is None:
                warnings_list.append(
                    "plot_hotspot_anomaly_argo_glorys_overviews 返回值中没有 results，且未找到可读 summary parquet。"
                )
            else:
                warnings_list.append(
                    f"已找到 GLORYS summary parquet，但其中没有逐 profile 记录: {glorys_summary_path_used}"
                )

    if require_glorys_details and len(table) > 0 and not glorys_records:
        raise ValueError(
            "缺少 GLORYS details。请先运行 "
            "plot_hotspot_anomaly_argo_glorys_overviews(..., annotate_heave=True)。"
        )
    glorys_by_key: dict[tuple, dict] = {}
    for rec in glorys_records:
        if not isinstance(rec, dict):
            continue
        date_key = _date_key_from_value(rec.get('target_date')) or _date_key_from_value(rec.get('profile_time'))
        for key in _record_keys(
            rec.get('year'),
            rec.get('profile_number'),
            date_key,
            rec.get('platform_number'),
        ):
            glorys_by_key.setdefault(key, rec)

    glorys_cols = [
        ('glorys_status', 'status'),
        ('glorys_error', 'error'),
        ('glorys_horizontal_status', 'horizontal_status'),
        ('glorys_vertical_status', 'vertical_status'),
        ('glorys_line_strategy', 'line_strategy'),
        ('glorys_k', 'k'),
        ('glorys_b', 'b'),
        ('glorys_center_lon', 'center_lon'),
        ('glorys_center_lat', 'center_lat'),
        ('heave_projection_depth_m', 'projection_depth_m'),
        ('heave_x_window_km', 'heave_x_window_km'),
        ('heave_z_window_m', 'heave_z_window_m'),
        ('heave_valid_fraction', 'heave_valid_fraction'),
        ('glorys_heave_sigma_argo', 'glorys_heave_sigma_argo'),
        ('glorys_heave_sigma_peak', 'glorys_heave_sigma_peak'),
        ('glorys_heave_zmin', 'glorys_heave_zmin'),
        ('glorys_heave_m', 'glorys_heave_m'),
        ('heave_error', 'heave_error'),
        ('spice_anomaly', 'spice_anomaly'),
        ('spice_percentile', 'spice_percentile'),
        ('spice_type', 'spice_type'),
        ('hotspot_type', 'hotspot_type'),
    ]

    matched_records = []
    for _, row in work.iterrows():
        date_key = _date_key_from_parts(row.get('Year'), row.get('Month'), row.get('Day'))
        matched = None
        for key in _record_keys(row.get('Year'), row.get('Profile_number'), date_key, row.get('Platform_number')):
            matched = glorys_by_key.get(key)
            if matched is not None:
                break
        matched_records.append(matched or {})

    for out_col, rec_col in glorys_cols:
        table[out_col] = [rec.get(rec_col, np.nan) for rec in matched_records]

    if glorys_records and table['glorys_heave_zmin'].isna().all():
        warnings_list.append(
            "GLORYS details 中没有可用 OI。若需要 OI，请使用 annotate_heave=True 重新运行 "
            "plot_hotspot_anomaly_argo_glorys_overviews。"
        )
        if require_glorys_details:
            raise ValueError(warnings_list[-1])

    if heave_z_threshold is not None:
        table['heave_z_threshold'] = float(heave_z_threshold)

    has_nearshore_diagnostics = (
        'nearshore_do_dip' in table.columns
        and table['nearshore_do_dip'].notna().any()
    )
    if diagnose_nearshore_do and not has_nearshore_diagnostics:
        table = _append_nearshore_do_diagnostics(table)
    elif not diagnose_nearshore_do:
        table['nearshore_do_dip'] = pd.Series(pd.NA, index=table.index, dtype='boolean')

    # 读取 overview_summary.parquet 已落盘的完整分类
    hotspot_type_codes = pd.to_numeric(table.get('hotspot_type'), errors='coerce').astype('Int64')
    if not hotspot_type_codes.isin([1, 2, 3]).any():
        raise ValueError(
            "overview_summary.parquet 中缺少有效 hotspot_type 分类码，"
            "请重新运行 plot_hotspot_anomaly_argo_glorys_overviews。"
        )
    table['hotspot_type'] = hotspot_type_codes.map(_HOTSPOT_TYPE_NAMES, na_action='ignore').astype('object')
    # spice_type：读已落盘的整数码（join 自 GLORYS overview summary）→ 映射成名称
    spice_type_codes = pd.to_numeric(table['spice_type'], errors='coerce').astype('Int64')
    table['spice_type'] = spice_type_codes.map(_SPICE_TYPE_NAMES, na_action='ignore').astype('object')

    # --- 基于 META 的涡旋位置辅助判定 ---
    table['meta_inside_eddy'] = pd.Series(pd.NA, index=table.index, dtype='boolean')
    table['meta_eddy_list'] = pd.Series('', index=table.index, dtype='object')
    try:
        meta_dir = _shared_output_dir("statistics", region_slug)
        meta_files = sorted(meta_dir.glob("all_interacting_argo_*.parquet"))
        if meta_files:
            meta_df = pd.read_parquet(meta_files[-1])
            required = {'Year', 'Profile_number'}
            if required.issubset(meta_df.columns):
                meta_df['_profile'] = pd.to_numeric(meta_df['Profile_number'], errors='coerce')
                meta_df['_year'] = pd.to_numeric(meta_df['Year'], errors='coerce')
                # 构建涡旋编号字符串：如 "ACL24906494, CL595940"
                ds_col = 'ds_name' if 'ds_name' in meta_df.columns else None
                tid_col = 'track_id' if 'track_id' in meta_df.columns else None
                if ds_col and tid_col:
                    meta_df['_eddy_tag'] = meta_df[ds_col].astype(str).str.upper() + meta_df[tid_col].astype(str)
                elif tid_col:
                    meta_df['_eddy_tag'] = meta_df[tid_col].astype(str)
                else:
                    meta_df['_eddy_tag'] = ''
                eddy_list = meta_df.groupby(['_year', '_profile'])['_eddy_tag'].apply(lambda x: ', '.join(sorted(set(x)))).reset_index(name='eddy_ids')
                table['_year_tmp'] = pd.to_numeric(table['Year'], errors='coerce')
                table['_profile_tmp'] = pd.to_numeric(table['Profile_number'], errors='coerce')
                merged = table[['_year_tmp', '_profile_tmp']].merge(
                    eddy_list, left_on=['_year_tmp', '_profile_tmp'],
                    right_on=['_year', '_profile'], how='left'
                )
                table['meta_eddy_list'] = merged['eddy_ids'].fillna('').astype(str)
                table['meta_inside_eddy'] = table['meta_eddy_list'] != ''
                table.drop(columns=['_year_tmp', '_profile_tmp'], inplace=True, errors='ignore')
    except Exception as exc:
        if verbose:
            print(f"[INFO] META eddy lookup skipped: {exc}")

    for col_name in [
        'Year', 'Month', 'Day', 'Profile_number', 'Platform_number',
        'Longitude', 'Latitude', 'anomaly_depth_m',
        'primary_value', 'anomaly_score', 'delta_do', 'do_value',
        'delta_aou', 'aou_value', 'delta_pi', 'pi_value',
        'trim_score', 'trim_scale_res_rob_aou', 'trim_scale_res_rob_abs_sal',
        'delta_temperature', 'delta_salinity',
        'glorys_k', 'glorys_b', 'glorys_center_lon', 'glorys_center_lat',
        'heave_projection_depth_m', 'heave_x_window_km', 'heave_z_window_m',
        'heave_valid_fraction', 'glorys_heave_sigma_argo', 'glorys_heave_sigma_peak',
        'glorys_heave_zmin', 'glorys_heave_m',
        'surface_do_ref', 'pre_anomaly_do_min', 'pre_anomaly_do_min_depth_m',
        'surface_to_min_do_drop', 'min_to_anomaly_do_recovery',
        'min_to_anomaly_depth_gap_m', 'do_v_shape_score',
    ]:
        if col_name in table.columns:
            table[col_name] = pd.to_numeric(table[col_name], errors='coerce')

    sort_cols = [c for c in ['Year', 'Month', 'Day', 'Profile_number', 'Platform_number'] if c in table.columns]
    if sort_cols:
        table = table.sort_values(sort_cols).reset_index(drop=True)
    else:
        table = table.reset_index(drop=True)

    dropped_summary_columns: list[str] = []
    if compact_columns:
        table, dropped_summary_columns = _compact_summary_columns(table)

    # 四宫格两列 hotspot_type/spice_type 连同身份-定位列前置，便于人工复核
    front_priority = [
        'date', 'Year', 'Month', 'Day', 'Profile_number', 'Platform_number',
        'Longitude', 'Latitude', 'anomaly_depth_m', 'hotspot_type', 'spice_type',
    ]
    front_cols = [c for c in front_priority if c in table.columns]
    table = table[front_cols + [c for c in table.columns if c not in front_cols]]

    saved_path = None
    if save_table:
        if output_path is None:
            fmt = str(output_format or 'csv').lower().lstrip('.')
            if fmt == 'excel':
                fmt = 'xlsx'
            if fmt not in {'csv', 'xlsx'}:
                raise ValueError("output_format must be 'csv' or 'xlsx'.")
            out_dir = cfg.output_dir("hotspot_anomaly_summary_table", region_slug)
            out_dir.mkdir(parents=True, exist_ok=True)
            output_path = out_dir / f"hotspot_anomaly_summary_{year_tag}_{run_tag}.{fmt}"
        else:
            output_path = Path(output_path)
            output_path.parent.mkdir(parents=True, exist_ok=True)
            fmt = str(output_format or output_path.suffix.lstrip('.') or 'csv').lower()
            if fmt == 'excel':
                fmt = 'xlsx'

        saved_path = Path(output_path)
        if fmt == 'csv':
            table.to_csv(saved_path, index=False)
        elif fmt == 'xlsx':
            try:
                from openpyxl.utils import get_column_letter

                with pd.ExcelWriter(saved_path, engine='openpyxl') as writer:
                    sheet_name = 'summary'
                    table.to_excel(writer, index=False, sheet_name=sheet_name)
                    worksheet = writer.sheets[sheet_name]
                    worksheet.freeze_panes = 'A2'

                    int_cols = {'Year', 'Month', 'Day', 'Profile_number', 'Platform_number'}
                    coord_cols = {'Longitude', 'Latitude', 'glorys_center_lon', 'glorys_center_lat', 'glorys_b'}
                    sci_cols = {'glorys_heave_sigma_argo', 'glorys_heave_sigma_peak', 'glorys_heave_zmin', 'heave_z_threshold'}
                    fraction_cols = {'heave_valid_fraction'}
                    one_decimal_cols = {
                        'surface_do_ref',
                        'pre_anomaly_do_min',
                        'pre_anomaly_do_min_depth_m',
                        'surface_to_min_do_drop',
                        'min_to_anomaly_do_recovery',
                        'min_to_anomaly_depth_gap_m',
                        'do_v_shape_score',
                        'glorys_heave_m',
                    }

                    def _excel_number_format(col_name: str) -> str | None:
                        col_lower = str(col_name).lower()
                        if col_name in int_cols:
                            return '0'
                        if col_name == 'date':
                            return 'yyyy-mm-dd'
                        if col_name in coord_cols:
                            return '0.000'
                        if col_name in sci_cols:
                            return '0.00E+00'
                        if col_name in fraction_cols:
                            return '0.00'
                        if col_name in one_decimal_cols:
                            return '0.0'
                        if 'depth' in col_lower:
                            return '0.0'
                        if any(token in col_lower for token in ['delta', 'value', 'score', 'threshold']):
                            return '0.000'
                        if col_name == 'glorys_k':
                            return '0.000'
                        return None

                    def _preview_excel_value(value, number_format: str | None) -> str:
                        if pd.isna(value):
                            return ''
                        if number_format == '0':
                            return f"{float(value):.0f}"
                        if number_format == 'yyyy-mm-dd':
                            try:
                                return pd.Timestamp(value).strftime('%Y-%m-%d')
                            except Exception:
                                return str(value)
                        if number_format == '0.000':
                            return f"{float(value):.3f}"
                        if number_format == '0.00E+00':
                            return f"{float(value):.2E}"
                        if number_format == '0.00':
                            return f"{float(value):.2f}"
                        if number_format == '0.0':
                            return f"{float(value):.1f}"
                        return str(value)

                    for col_idx, col_name in enumerate(table.columns, start=1):
                        values = table[col_name].dropna()
                        sample_values = values.head(1000)
                        number_format = _excel_number_format(str(col_name))
                        if number_format is not None:
                            for row_idx in range(2, len(table) + 2):
                                worksheet.cell(row=row_idx, column=col_idx).number_format = number_format
                        max_value_len = max(
                            [len(str(col_name))]
                            + [len(_preview_excel_value(value, number_format)) for value in sample_values]
                        )
                        width = min(max(max_value_len + 1, 6), 42)
                        worksheet.column_dimensions[get_column_letter(col_idx)].width = width
            except Exception as exc:
                warnings_list.append(f"设置 xlsx 列宽失败，已使用默认宽度保存: {exc}")
                table.to_excel(saved_path, index=False)
        else:
            raise ValueError("output_format must be 'csv' or 'xlsx'.")

        if verbose:
            print(f"Summary table saved to: {saved_path}")

    if verbose:
        for msg in warnings_list:
            print(f"[WARN] {msg}")

    return {
        'summary_table': table,
        'output_path': str(saved_path) if saved_path is not None else None,
        'anomalies_path': str(resolved_anomalies_path),
        'argo_glorys_summary_data_path': str(glorys_summary_path_used) if glorys_summary_path_used is not None else None,
        'n_rows': int(len(table)),
        'dropped_columns': dropped_summary_columns,
        'nearshore_do_thresholds': {
            'pre_anomaly_do_min': float(nearshore_do_min_threshold),
            'surface_to_min_do_drop': float(nearshore_do_drop_threshold),
            'min_to_anomaly_do_recovery': float(nearshore_do_recovery_threshold),
            'min_to_anomaly_depth_gap_m': float(nearshore_do_depth_gap_threshold_m),
        },
        'warnings': warnings_list,
    }


def plot_single_hotspot_profile(
    profile_number: int,
    profile_time: int | str | pd.Timestamp,
    platform_number: int | None = None,
    detection_config: DetectionConfig | None = None,
    variables: list = ['DO', 'AOU', 'Temp', 'Salinity'],
    xlim_overrides: dict[str, tuple[float, float]] | None = None,
    remove_outliers: bool = True,
    plot_normal_scatter: bool = True,
    annotate_delta_ts: bool = True,
    save_fig: bool = False,
    show_fig: bool = True,
    argo_data_dir: str | Path | None = None,
    output_dir: str | Path | None = None,
) -> dict:
    """重绘单个 Hotspots Profile，并支持按变量单独调节横轴范围。

    参数:
        - profile_number (int): 目标剖面编号（Profile_number）。
        - profile_time (int | str | pd.Timestamp): 时间输入，支持年份（如 2014）或日期/时间戳（如 '2014-05-09'）。
        - platform_number (int | None): 可选平台编号筛选。
        - detection_config (DetectionConfig | None): 异常识别配置；None 时使用 processing.yml 默认值。
        - variables (list): 绘制变量列表，默认 ['DO', 'AOU', 'Temp', 'Salinity']。
        - xlim_overrides (dict[str, tuple[float, float]] | None): 横轴范围覆盖，如 {'DO': (0, 300), 'Temp': (-2, 30), 'Salinity': (33, 36)}；键可用 'Temp' 或 'Temperature'。
        - remove_outliers (bool): 与 plot_vertical 同义；False 时显示 QC 异常标记与红色桥接线，默认 True。
        - plot_normal_scatter (bool): 是否绘制正常值的孤立散点标记，默认 True。
        - annotate_delta_ts (bool): 是否在标题追加当前异常方法的判别变量、辅助 ΔT/ΔS 及深度，默认 True。
        - save_fig (bool): 是否保存图像，默认 False。
        - show_fig (bool): 是否显示图像，默认 True。
        - argo_data_dir (str | Path | None): 年度 Argo parquet 目录；None 时使用配置默认。
        - output_dir (str | Path | None): 自定义输出目录；None 时使用 `plot_outputs/<method>/<region>/plot_argo_hotspots_vertical_profiles_single`。

    返回:
        - dict: 含 save_path/profile_number/platform_number/date/ado/atemp/asalinity。
    """
    if argo_data_dir is None:
        argo_data_dir = argo_path
    cfg = _resolve_detection_config(detection_config)

    info = _resolve_argo_profile_center(
        profile_number=int(profile_number),
        profile_time=profile_time,
        platform_number=platform_number,
        argo_data_dir=argo_data_dir,
    )

    profile_rows = info['profile_rows'].sort_values('Depth').copy()
    target_date = pd.Timestamp(info['target_date']).normalize()

    plot_variables, overlay_aou_on_do = _prepare_vertical_plot_variables(variables)

    # 若未指定平台，且仍有多个平台，默认取第一个以保持“单图单剖面”语义
    platform_val = None
    if 'Platform_number' in profile_rows.columns:
        platforms = pd.to_numeric(profile_rows['Platform_number'], errors='coerce').dropna().astype(int).unique()
        if platforms.size > 0:
            platform_val = int(platforms[0])
            if platform_number is None and platforms.size > 1:
                profile_rows = profile_rows[
                    pd.to_numeric(profile_rows['Platform_number'], errors='coerce') == platform_val
                ].copy()

    num_variables = len(plot_variables)
    fig, axes = plt.subplots(1, num_variables, figsize=(10 * num_variables, 20), sharey=True)
    if num_variables == 1:
        axes = [axes]

    line_color = plt.cm.coolwarm(0.15)
    any_plotted = False

    for var_name, ax in zip(plot_variables, axes):
        plot_variable_name = var_name
        is_do_panel = (_map_plot_variable_name(plot_variable_name) == 'DO')
        do_aux_layers = ('aou',) if (is_do_panel and overlay_aou_on_do) else tuple()

        db_variable_name = _map_plot_variable_name(plot_variable_name)
        plot_line_color = '#ff8c00' if db_variable_name == 'AOU' else line_color
        if not _has_plottable_profile_variable(profile_rows, db_variable_name):
            ax.text(
                0.5,
                0.5,
                f"Variable '{db_variable_name}'\\nnot found in data.",
                ha='center',
                va='center',
                transform=ax.transAxes,
                fontsize=16,
            )
            _apply_vertical_profile_axis_style(ax, var_name)
            continue

        did_plot = _plot_single_argo_profile_line(
            ax,
            profile_rows,
            plot_variable_name,
            plot_line_color,
            remove_outliers=remove_outliers,
            show_normal_scatter=plot_normal_scatter,
            do_aux_layers=do_aux_layers,
            aou_aux_color='#ff8c00',
            alpha=0.9,
        )
        if not did_plot:
            ax.text(
                0.5,
                0.5,
                "No valid data after QC.",
                ha='center',
                va='center',
                transform=ax.transAxes,
                fontsize=14,
            )
        else:
            any_plotted = True

        _apply_vertical_profile_axis_style(ax, plot_variable_name)

        if xlim_overrides:
            override = None
            if var_name in xlim_overrides:
                override = xlim_overrides[var_name]
            elif db_variable_name in xlim_overrides:
                override = xlim_overrides[db_variable_name]
            if override is not None and len(override) == 2:
                try:
                    ax.set_xlim(float(override[0]), float(override[1]))
                except Exception:
                    pass

    axes[0].set_ylabel("Depth/m", fontsize=20)
    axes[0].tick_params(axis='y', labelsize=16)
    axes[0].invert_yaxis()

    date_text = target_date.strftime('%Y-%m-%d')

    annotation_text = ""
    depth_text = ""
    delta_do_val = np.nan
    delta_temp_val = np.nan
    delta_salinity_val = np.nan
    delta_aou_val = np.nan
    delta_pi_val = np.nan
    trim_score_val = np.nan
    anomaly_score_val = np.nan
    primary_metric = None
    primary_value = np.nan
    picked_depth = np.nan
    if annotate_delta_ts:
        try:
            deltas = calculate_delta_do(
                profile_rows.copy(),
                detection_config=cfg,
                remove_outliers=remove_outliers,
                verbose=False,
            )
            if not deltas.empty:
                picked_df = _keep_best_anomaly_per_profile(deltas, cfg)
                picked = picked_df.sort_values(cfg.score_col(), ascending=False).iloc[0]
                annotation_text, depth_text = _annotation_text_from_anomaly_record(picked, cfg)

                delta_do_val = _num_from_record(picked, 'delta_do')
                delta_temp_val = _num_from_record(picked, 'delta_temperature')
                delta_salinity_val = _num_from_record(picked, 'delta_salinity')
                delta_aou_val = _num_from_record(picked, 'delta_aou')
                delta_pi_val = _num_from_record(picked, 'delta_pi')
                trim_score_val = _num_from_record(picked, 'trim_score')
                anomaly_score_val = _num_from_record(picked, 'anomaly_score')
                picked_depth = _num_from_record(picked, 'depth')
                primary_metric = picked.get('primary_metric')
                primary_value = _num_from_record(picked, 'primary_value')
        except Exception as exc:
            print(f"[WARN] Failed to compute anomaly annotation for profile {profile_number}: {exc}")

    platform_label = (
        int(platform_number) if platform_number is not None
        else (platform_val if platform_val is not None else None)
    )
    platform_text = f", Platform={platform_label}" if platform_label is not None else ""

    lon_text = ""
    lat_text = ""
    lon_val = pd.to_numeric(pd.Series([profile_rows.iloc[0].get('Longitude')]), errors='coerce').iloc[0]
    lat_val = pd.to_numeric(pd.Series([profile_rows.iloc[0].get('Latitude')]), errors='coerce').iloc[0]
    if np.isfinite(lon_val):
        lon_text = f"Lon={float(_normalize_lon_array(lon_val)):.3f}"
    if np.isfinite(lat_val):
        lat_text = f"Lat={float(lat_val):.3f}"
    location_text = ", ".join([t for t in [lon_text, lat_text] if t])

    title_line1 = (
        f"Hotspots Profile {int(profile_number)}{platform_text}, "
        f"{date_text}{annotation_text}{depth_text}"
    )
    title_line2 = location_text if location_text else "Lon/Lat unavailable"

    fig.suptitle(
        f"{title_line1}\n{title_line2}",
        fontsize=24,
        y=0.97,
    )
    plt.tight_layout(rect=[0, 0, 1, 0.90])

    saved_path = None
    if save_fig and any_plotted:
        if output_dir is None:
            region_slug = _current_region_key()
            output_dir = cfg.output_dir("plot_argo_hotspots_vertical_profiles_single", region_slug)
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
        run_tag = cfg.file_stem()
        platform_suffix = f"_platform{platform_label}" if platform_label is not None else ""
        filename = (
            f"hotspot_{date_text.replace('-', '')}_P{int(profile_number)}"
            f"{platform_suffix}_{run_tag}_single.png"
        )
        saved_path = output_dir / filename
        plt.savefig(saved_path, dpi=300, bbox_inches='tight')

    if show_fig:
        plt.show()
    plt.close(fig)

    return {
        'profile_number': int(profile_number),
        'platform_number': int(platform_label) if platform_label is not None else None,
        'date': date_text,
        'delta_do': float(delta_do_val) if np.isfinite(delta_do_val) else np.nan,
        'delta_temperature': float(delta_temp_val) if np.isfinite(delta_temp_val) else np.nan,
        'delta_salinity': float(delta_salinity_val) if np.isfinite(delta_salinity_val) else np.nan,
        'delta_aou': float(delta_aou_val) if np.isfinite(delta_aou_val) else np.nan,
        'delta_pi': float(delta_pi_val) if np.isfinite(delta_pi_val) else np.nan,
        'trim_score': float(trim_score_val) if np.isfinite(trim_score_val) else np.nan,
        'anomaly_score': float(anomaly_score_val) if np.isfinite(anomaly_score_val) else np.nan,
        'primary_metric': primary_metric,
        'primary_value': float(primary_value) if np.isfinite(primary_value) else np.nan,
        'detection_method': cfg.method,
        'save_path': str(saved_path) if saved_path is not None else None,
    }

_NEARSHORE_DO_DIAGNOSTIC_COLUMNS = [
    'surface_do_ref',
    'pre_anomaly_do_min',
    'pre_anomaly_do_min_depth_m',
    'surface_to_min_do_drop',
    'min_to_anomaly_do_recovery',
    'min_to_anomaly_depth_gap_m',
    'do_v_shape_score',
    'nearshore_do_dip',
]


def _first_existing_profile_column(columns: pd.Index, names: list[str]) -> str | None:
    for name in names:
        if name in columns:
            return name
    return None


def _nearshore_do_profile_shape_metrics(
    profile_rows: pd.DataFrame,
    target_depth,
    target_do=np.nan,
    *,
    nearshore_do_min_threshold: float = 50.0,
    nearshore_do_drop_threshold: float = 100.0,
    nearshore_do_recovery_threshold: float = 100.0,
    nearshore_do_depth_gap_threshold_m: float = 100.0,
) -> dict:
    """计算单个 Argo 剖面的近岸型 DO 先降后升诊断。"""
    out = {
        'surface_do_ref': np.nan,
        'pre_anomaly_do_min': np.nan,
        'pre_anomaly_do_min_depth_m': np.nan,
        'surface_to_min_do_drop': np.nan,
        'min_to_anomaly_do_recovery': np.nan,
        'min_to_anomaly_depth_gap_m': np.nan,
        'do_v_shape_score': np.nan,
        'nearshore_do_dip': pd.NA,
    }
    try:
        target_depth_f = float(target_depth)
    except Exception:
        return out
    if not np.isfinite(target_depth_f):
        return out

    depth_col = _first_existing_profile_column(profile_rows.columns, ['Depth', 'depth', 'DEPTH'])
    if depth_col is None:
        return out
    do_adj_col = _first_existing_profile_column(profile_rows.columns, ['DO', 'DOXY_Adjusted', 'DO_Adjusted'])
    do_raw_col = _first_existing_profile_column(profile_rows.columns, ['DOXY', 'DO_Raw'])
    if do_adj_col is None and do_raw_col is None:
        return out

    depth_vals = pd.to_numeric(profile_rows[depth_col], errors='coerce')
    do_vals = (
        pd.to_numeric(profile_rows[do_adj_col], errors='coerce')
        if do_adj_col is not None else pd.Series(np.nan, index=profile_rows.index)
    )
    if do_raw_col is not None:
        do_raw_vals = pd.to_numeric(profile_rows[do_raw_col], errors='coerce')
        do_vals = do_vals.where(do_vals.notna(), do_raw_vals)

    prof = pd.DataFrame({'Depth': depth_vals, 'DO': do_vals}).dropna().sort_values('Depth')
    prof = prof[np.isfinite(prof['Depth']) & np.isfinite(prof['DO'])].copy()
    if len(prof) < 5:
        return out

    pre = prof[(prof['Depth'] >= 0.0) & (prof['Depth'] <= target_depth_f)].copy()
    if len(pre) < 5:
        return out

    surface = pre[pre['Depth'] <= 100.0]
    if len(surface) >= 3:
        surface_ref = float(surface['DO'].median())
    else:
        surface_ref = float(pre.head(min(10, len(pre)))['DO'].median())

    min_idx = pre['DO'].idxmin()
    min_do = float(pre.loc[min_idx, 'DO'])
    min_depth = float(pre.loc[min_idx, 'Depth'])

    try:
        anomaly_do = float(target_do)
    except Exception:
        anomaly_do = np.nan
    if not np.isfinite(anomaly_do):
        nearest_idx = (prof['Depth'] - target_depth_f).abs().idxmin()
        anomaly_do = float(prof.loc[nearest_idx, 'DO'])

    drop_val = float(surface_ref - min_do)
    recovery_val = float(anomaly_do - min_do)
    depth_gap = float(target_depth_f - min_depth)
    score_val = float(min(drop_val, recovery_val)) if np.isfinite(drop_val) and np.isfinite(recovery_val) else np.nan
    nearshore_flag = (
        np.isfinite(min_do)
        and np.isfinite(drop_val)
        and np.isfinite(recovery_val)
        and np.isfinite(depth_gap)
        and min_do <= float(nearshore_do_min_threshold)
        and drop_val >= float(nearshore_do_drop_threshold)
        and recovery_val >= float(nearshore_do_recovery_threshold)
        and depth_gap >= float(nearshore_do_depth_gap_threshold_m)
    )

    out.update({
        'surface_do_ref': surface_ref,
        'pre_anomaly_do_min': min_do,
        'pre_anomaly_do_min_depth_m': min_depth,
        'surface_to_min_do_drop': drop_val,
        'min_to_anomaly_do_recovery': recovery_val,
        'min_to_anomaly_depth_gap_m': depth_gap,
        'do_v_shape_score': score_val,
        'nearshore_do_dip': bool(nearshore_flag),
    })
    return out


def _append_nearshore_do_diagnostics_from_profiles(
    anomalies: pd.DataFrame,
    profile_rows: pd.DataFrame,
    *,
    target_depth_col: str = 'depth',
    target_do_col: str = 'do_value',
    nearshore_do_min_threshold: float = 50.0,
    nearshore_do_drop_threshold: float = 100.0,
    nearshore_do_recovery_threshold: float = 100.0,
    nearshore_do_depth_gap_threshold_m: float = 100.0,
) -> pd.DataFrame:
    """把近岸型 DO 诊断附加到 anomaly 摘要表。"""
    out = anomalies.copy()
    for col_name in _NEARSHORE_DO_DIAGNOSTIC_COLUMNS:
        out[col_name] = np.nan
    out['nearshore_do_dip'] = pd.Series(pd.NA, index=out.index, dtype='boolean')

    if out.empty or profile_rows is None or profile_rows.empty:
        return out
    if 'Profile_number' not in out.columns or 'Profile_number' not in profile_rows.columns:
        return out
    if target_depth_col not in out.columns:
        return out

    key_cols = [
        c for c in ['Profile_number', 'Platform_number', 'Month', 'Day']
        if c in out.columns and c in profile_rows.columns
    ]
    if 'Profile_number' not in key_cols:
        key_cols.insert(0, 'Profile_number')
    key_cols = list(dict.fromkeys(key_cols))

    def _key_from_row(row) -> tuple:
        vals = []
        for col_name in key_cols:
            vals.append(_to_int_like(row.get(col_name)))
        return tuple(vals)

    prof = profile_rows.copy()
    for col_name in key_cols:
        prof[col_name] = pd.to_numeric(prof[col_name], errors='coerce').astype('Int64')

    def _normalize_group_key(key) -> tuple:
        if not isinstance(key, tuple):
            key = (key,)
        return tuple(_to_int_like(v) for v in key)

    grouped = {
        _normalize_group_key(key): grp
        for key, grp in prof.groupby(key_cols, dropna=False, sort=False)
    }

    for idx, row in out.iterrows():
        key = _key_from_row(row)
        profile_group = grouped.get(key)
        if profile_group is None:
            continue
        metrics = _nearshore_do_profile_shape_metrics(
            profile_group,
            row.get(target_depth_col),
            row.get(target_do_col, np.nan),
            nearshore_do_min_threshold=nearshore_do_min_threshold,
            nearshore_do_drop_threshold=nearshore_do_drop_threshold,
            nearshore_do_recovery_threshold=nearshore_do_recovery_threshold,
            nearshore_do_depth_gap_threshold_m=nearshore_do_depth_gap_threshold_m,
        )
        for col_name, value in metrics.items():
            out.at[idx, col_name] = value

    return out


# 导出表渲染用：hotspot_type 整数码→类别名（整数码仍在 anomalies parquet，拆图逻辑依赖它）
_HOTSPOT_TYPE_NAMES = {1: 'ventilated', 2: 'isolated', 3: 'OMZ'}
# spice_type 整数码→类别名（码落盘在 GLORYS overview summary parquet，导出时映射）
_SPICE_TYPE_NAMES = {1: 'cold-fresh', 2: 'background-consistent', 3: 'warm-salty'}


def _assign_hotspot_type(
    table: pd.DataFrame,
    *,
    heave_z_threshold: float | None = _heave_depth_threshold,
    heave_z_col: str = 'glorys_heave_zmin',
    heave_m_threshold: float | None = _heave_magnitude_threshold,
    heave_m_col: str = 'glorys_heave_m',
    nearshore_col: str = 'nearshore_do_dip',
) -> pd.Series:
    """按近岸 DO 诊断和 heave 联合判定生成 hotspot_type。

    - Type 1: 非近岸，heave H ≥ heave_m_threshold 且 z_min < heave_z_threshold（等密面显著位移且通风型）
    - Type 2: 非近岸，不满足 heave 联合判定（深层隔离型）
    - Type 3: 近岸 DO dip

    阈值默认值来自 processing.yml:processing.heave（运行时可覆盖），
    若阈值参数为 None 或表中无对应列，则只分 Type 3 / 未分类。
    """
    nearshore_vals = (
        table[nearshore_col].astype('boolean')
        if nearshore_col in table.columns
        else pd.Series(False, index=table.index, dtype='boolean')
    )
    hotspot_type = pd.Series(pd.NA, index=table.index, dtype='Int64')
    nearshore_mask = nearshore_vals.fillna(False)
    hotspot_type[nearshore_mask] = 3

    has_z = heave_z_threshold is not None and heave_z_col in table.columns
    has_m = heave_m_threshold is not None and heave_m_col in table.columns
    if not has_z or not has_m:
        return hotspot_type

    z_vals = pd.to_numeric(table[heave_z_col], errors='coerce')
    m_vals = pd.to_numeric(table[heave_m_col], errors='coerce')
    z_ok = z_vals.notna()
    m_ok = m_vals.notna()
    non_nearshore_mask = ~nearshore_mask

    pass_z = z_vals < float(heave_z_threshold)
    pass_m = m_vals >= float(heave_m_threshold)
    hotspot_type[non_nearshore_mask & z_ok & m_ok & pass_z & pass_m] = 1
    hotspot_type[non_nearshore_mask & z_ok & m_ok & ~(pass_z & pass_m)] = 2
    return hotspot_type


def _assign_spice_type(
    table: pd.DataFrame,
    *,
    percentile_threshold: float = _default_spice_percentile_threshold,
    percentile_col: str = 'spice_percentile',
) -> pd.Series:
    """按背景相对 spiciness 百分位生成 spice_type 整数码（与 hotspot_type 同构的一等列）。

    判据基于 compute_spiciness_anomaly 的 σ₀ 加权背景百分位 P（0–100），返回 Int64 码：
    - ``1`` 冷淡水团 cold-fresh：P < percentile_threshold（T-S 图偏左下）
    - ``3`` 暖咸水团 warm-salty：P > 100 − percentile_threshold（T-S 图偏右）
    - ``2`` 背景一致 background-consistent：两者之间
    ``spice_percentile`` 为 NaN（背景不足无法计算）时留 <NA>，与码 2 区分。
    名称映射见 ``_SPICE_TYPE_NAMES``；阈值默认来自 processing.yml:processing.spice.percentile_threshold。
    """
    codes = pd.Series(pd.NA, index=table.index, dtype='Int64')
    if percentile_col not in table.columns:
        return codes
    pct = pd.to_numeric(table[percentile_col], errors='coerce')
    valid = pct.notna()
    thr = float(percentile_threshold)
    codes[valid] = 2
    codes[valid & (pct < thr)] = 1
    codes[valid & (pct > 100.0 - thr)] = 3
    return codes


def _hotspot_date_key_from_parts(year_val, month_val=None, day_val=None) -> str | None:
    try:
        if month_val is None or day_val is None or pd.isna(month_val) or pd.isna(day_val):
            return None
        return pd.Timestamp(
            year=int(year_val),
            month=int(month_val),
            day=int(day_val),
        ).strftime('%Y-%m-%d')
    except Exception:
        return None


def _hotspot_date_key_from_value(val) -> str | None:
    try:
        if val is None or pd.isna(val):
            return None
    except Exception:
        pass
    try:
        return pd.Timestamp(val).normalize().strftime('%Y-%m-%d')
    except Exception:
        return None


def _hotspot_record_keys(year_val, profile_val, date_key=None, platform_val=None) -> list[tuple]:
    year_key = _to_int_like(year_val)
    profile_key = _to_int_like(profile_val)
    if year_key is None or profile_key is None:
        return []
    platform_key = _to_int_like(platform_val)
    keys = []
    if date_key and platform_key is not None:
        keys.append((year_key, profile_key, date_key, platform_key))
    if date_key:
        keys.append((year_key, profile_key, date_key, None))
    if platform_key is not None:
        keys.append((year_key, profile_key, None, platform_key))
    keys.append((year_key, profile_key, None, None))
    return keys


def _merge_hotspot_glorys_summary_fields(
    anomalies: pd.DataFrame,
    glorys_summary: pd.DataFrame,
) -> pd.DataFrame:
    """将 GLORYS overview summary parquet 中的 OI/状态字段合并到 hotspot anomalies。"""
    out = anomalies.copy()
    if out.empty or glorys_summary is None or glorys_summary.empty:
        return out

    glorys_cols = [
        ('glorys_status', 'status'),
        ('glorys_error', 'error'),
        ('glorys_horizontal_status', 'horizontal_status'),
        ('glorys_vertical_status', 'vertical_status'),
        ('glorys_line_strategy', 'line_strategy'),
        ('glorys_k', 'k'),
        ('glorys_b', 'b'),
        ('glorys_center_lon', 'center_lon'),
        ('glorys_center_lat', 'center_lat'),
        ('heave_projection_depth_m', 'projection_depth_m'),
        ('heave_x_window_km', 'heave_x_window_km'),
        ('heave_z_window_m', 'heave_z_window_m'),
        ('heave_valid_fraction', 'heave_valid_fraction'),
        ('glorys_heave_sigma_argo', 'glorys_heave_sigma_argo'),
        ('glorys_heave_sigma_peak', 'glorys_heave_sigma_peak'),
        ('glorys_heave_zmin', 'glorys_heave_zmin'),
        ('glorys_heave_m', 'glorys_heave_m'),
        ('heave_error', 'heave_error'),
        ('spice_anomaly', 'spice_anomaly'),
        ('spice_percentile', 'spice_percentile'),
        ('spice_type', 'spice_type'),
        ('hotspot_type', 'hotspot_type'),
    ]
    for out_col, _ in glorys_cols:
        if out_col not in out.columns:
            out[out_col] = np.nan

    glorys_by_key: dict[tuple, dict] = {}
    for rec in glorys_summary.to_dict('records'):
        if not isinstance(rec, dict):
            continue
        date_key = (
            _hotspot_date_key_from_value(rec.get('target_date'))
            or _hotspot_date_key_from_value(rec.get('profile_time'))
        )
        for key in _hotspot_record_keys(
            rec.get('year'),
            rec.get('profile_number'),
            date_key,
            rec.get('platform_number'),
        ):
            glorys_by_key.setdefault(key, rec)

    matched_records = []
    for _, row in out.iterrows():
        date_key = _hotspot_date_key_from_parts(row.get('Year'), row.get('Month'), row.get('Day'))
        matched = None
        for key in _hotspot_record_keys(row.get('Year'), row.get('Profile_number'), date_key, row.get('Platform_number')):
            matched = glorys_by_key.get(key)
            if matched is not None:
                break
        matched_records.append(matched or {})

    for out_col, rec_col in glorys_cols:
        out[out_col] = [rec.get(rec_col, np.nan) for rec in matched_records]
    return out


def _to_int_like(val) -> int | None:
    try:
        if pd.isna(val):
            return None
        return int(val)
    except Exception:
        return None


def _hotspot_anomaly_output_columns() -> list[str]:
    """返回 ``plot_argo_hotspots`` anomaly parquet 的首选列顺序。"""
    return [
        'Profile_number',
        'Platform_number',
        'Longitude',
        'Latitude',
        'Year',
        'Month',
        'Day',
        'depth',
        'delta_do',
        'do_value',
        'delta_aou',
        'aou_value',
        'delta_pi',
        'pi_value',
        'trim_score',
        'trim_scale_res_rob_aou',
        'trim_scale_res_rob_abs_sal',
        'delta_salinity',
        'salinity_value',
        'delta_temperature',
        'temperature_value',
        'pi_peak_type',
        'pi_peak_depth',
        'aou_peak_depth',
        'peak_depth_offset',
        'anomaly_score',
        'primary_metric',
        'primary_value',
        'surface_do_ref',
        'pre_anomaly_do_min',
        'pre_anomaly_do_min_depth_m',
        'surface_to_min_do_drop',
        'min_to_anomaly_do_recovery',
        'min_to_anomaly_depth_gap_m',
        'do_v_shape_score',
        'nearshore_do_dip',
        'heave_valid_fraction',
        'glorys_heave_sigma_argo',
        'glorys_heave_zmin',
        'heave_error',
        'hotspot_type',
        'detection_method',
    ]


def _hotspot_year_worker(args: tuple) -> tuple[pd.DataFrame, pd.DataFrame]:
    """模块级 worker，支持 multiprocessing pickling。

    参数 args: (
        year,
        cfg,
        lon_min_bound,
        lon_max_bound,
        lat_min_bound,
        lat_max_bound,
    )
    返回: (baseline_df, anomalies_df)
    baseline_df: 每个剖面第一条记录的基本信息
    anomalies_df: 该年筛选出的 Argo 异常（每剖面保留 anomaly_score 最强一条）
    """
    (
        year,
        cfg,
        lon_min_bound,
        lon_max_bound,
        lat_min_bound,
        lat_max_bound,
    ) = args
    try:
        df_y = load_argo_data(year=year)
    except FileNotFoundError:
        print(f"[hotspots] Missing year {year}, skip.")
        return pd.DataFrame(), pd.DataFrame()
    except Exception as e:
        print(f"[hotspots] Error loading year {year}: {e}")
        return pd.DataFrame(), pd.DataFrame()
    # 地理过滤
    lon_vals = df_y['Longitude'].to_numpy(dtype=float, copy=False)
    lat_vals = df_y['Latitude'].to_numpy(dtype=float, copy=False)
    lon_mask = _region_lon_mask(lon_vals, lon_min_bound, lon_max_bound)
    lat_mask = (lat_vals >= lat_min_bound) & (lat_vals <= lat_max_bound)
    geo_mask = lon_mask & lat_mask
    df_geo = df_y[geo_mask].copy()
    if df_geo.empty:
        return pd.DataFrame(), pd.DataFrame()
    # baseline
    baseline = (
        df_geo.sort_values(['Profile_number','Depth'])
        .groupby('Profile_number', as_index=False)
        .first()[['Profile_number','Longitude','Latitude','Year','Month','Day']]
    )
    anomalies_year = calculate_delta_do(
        df_geo,
        detection_config=cfg,
        remove_outliers=True,
        verbose=False
    )
    if anomalies_year.empty:
        return baseline, pd.DataFrame()
    anomalies_year = _keep_best_anomaly_per_profile(anomalies_year, cfg)
    anomalies_year = _append_nearshore_do_diagnostics_from_profiles(
        anomalies_year,
        df_geo,
        target_depth_col='depth',
        target_do_col='do_value',
    )
    anomalies_year['hotspot_type'] = _assign_hotspot_type(
        anomalies_year,
        heave_z_threshold=None,
        heave_m_threshold=None,
    )
    needed_cols = [
        c for c in _hotspot_anomaly_output_columns()
        if c in anomalies_year.columns
    ]
    anomalies_year = anomalies_year[needed_cols]
    return baseline, anomalies_year

def _build_euler_grid_edges(grid_step_deg: float = 1.0) -> tuple[np.ndarray, np.ndarray, dict]:
    """为当前区域构建欧拉统计网格边界（经纬度）。

    返回:
        lon_edges, lat_edges, grid_meta
    说明:
        - 对跨日界线区域，经度边界会转换为连续坐标（例如 137.5 -> 182.5）。
        - 采用 floor/ceil 对齐规则，保证不同统计图层使用同一网格口径。
    """
    if grid_step_deg <= 0:
        raise ValueError("grid_step_deg 必须大于 0")

    lon_min_cfg = float(lonmin)
    lon_max_cfg = float(lonmax)
    lat_min_cfg = float(latmin)
    lat_max_cfg = float(latmax)

    lon_min_norm = float(_normalize_lon_array(lon_min_cfg))
    lon_max_norm = float(_normalize_lon_array(lon_max_cfg))
    crosses_dateline = bool(_REGION_CFG.get('crosses_dateline') and (lon_max_norm < lon_min_norm))

    lon_start = lon_min_norm
    lon_end = lon_max_norm + (360.0 if crosses_dateline else 0.0)

    raw_span = abs(lon_max_cfg - lon_min_cfg)
    eff_span = (lon_max_norm - lon_min_norm) % 360.0
    is_global_lon = (raw_span >= 359.5) or (eff_span >= 359.5) or np.isclose(eff_span, 0.0, atol=1e-6)
    if is_global_lon:
        lon_start = -180.0
        lon_end = 180.0
        crosses_dateline = False

    lon_floor = math.floor(lon_start / grid_step_deg) * grid_step_deg
    lon_ceil = math.ceil(lon_end / grid_step_deg) * grid_step_deg
    lat_floor = math.floor(lat_min_cfg / grid_step_deg) * grid_step_deg
    lat_ceil = math.ceil(lat_max_cfg / grid_step_deg) * grid_step_deg

    lon_edges = np.arange(lon_floor, lon_ceil + grid_step_deg * 0.5, grid_step_deg, dtype=float)
    lat_edges = np.arange(lat_floor, lat_ceil + grid_step_deg * 0.5, grid_step_deg, dtype=float)

    if lon_edges.size < 2:
        lon_edges = np.array([lon_floor, lon_floor + grid_step_deg], dtype=float)
    if lat_edges.size < 2:
        lat_edges = np.array([lat_floor, lat_floor + grid_step_deg], dtype=float)

    meta = {
        'grid_step_deg': float(grid_step_deg),
        'crosses_dateline': bool(crosses_dateline),
        'lon_start_continuous': float(lon_floor),
        'lon_end_continuous': float(lon_ceil),
        'region_key': _current_region_key(),
    }
    return lon_edges, lat_edges, meta


def _to_continuous_lon(lon_vals: np.ndarray, lon_ref: float, crosses_dateline: bool) -> np.ndarray:
    """将经度转换到连续坐标，便于跨日界线网格分箱。"""
    lon_norm = _normalize_lon_array(lon_vals)
    if not crosses_dateline:
        return lon_norm
    lon_cont = lon_norm.copy()
    lon_cont[lon_cont < lon_ref] += 360.0
    return lon_cont


def _build_ocean_mask_for_grid(
    lon_edges: np.ndarray,
    lat_edges: np.ndarray,
    *,
    crosses_dateline: bool,
    lon_ref: float,
) -> np.ndarray:
    """基于 Natural Earth 陆地多边形构建海洋网格掩膜（True=海洋）。"""
    n_lat = max(int(lat_edges.size - 1), 0)
    n_lon = max(int(lon_edges.size - 1), 0)
    if n_lat == 0 or n_lon == 0:
        return np.zeros((n_lat, n_lon), dtype=bool)

    lon_centers = 0.5 * (lon_edges[:-1] + lon_edges[1:])
    lat_centers = 0.5 * (lat_edges[:-1] + lat_edges[1:])
    lon_2d, lat_2d = np.meshgrid(lon_centers, lat_centers)

    lon_plot = _normalize_lon_array(lon_2d.ravel()) if crosses_dateline else lon_2d.ravel()
    lat_plot = lat_2d.ravel()

    try:
        world = _load_world_geodataframe()[['geometry']].copy()
        if world.crs is not None and str(world.crs).lower() not in {'epsg:4326', '4326'}:
            world = world.to_crs('EPSG:4326')

        points = gpd.GeoDataFrame(
            {'idx': np.arange(lon_plot.size, dtype=int)},
            geometry=gpd.points_from_xy(lon_plot, lat_plot),
            crs='EPSG:4326',
        )

        try:
            joined = gpd.sjoin(points, world, how='left', predicate='within')
        except TypeError:
            joined = gpd.sjoin(points, world, how='left', op='within')

        land_flat = joined['index_right'].notna().to_numpy(dtype=bool)
        ocean_flat = ~land_flat
        return ocean_flat.reshape(n_lat, n_lon)
    except Exception as exc:
        print(f"[ocean_mask] Failed to build land/ocean mask, fallback to all-ocean: {exc}")
        return np.ones((n_lat, n_lon), dtype=bool)


def _grid_count_from_points(
    df: pd.DataFrame,
    *,
    lon_col: str,
    lat_col: str,
    lon_edges: np.ndarray,
    lat_edges: np.ndarray,
    crosses_dateline: bool,
    lon_ref: float,
    dedupe_cols: list[str] | None = None,
) -> tuple[np.ndarray, pd.DataFrame]:
    """将点数据按网格统计计数，可选按字段去重。"""
    if df.empty:
        shape = (max(lat_edges.size - 1, 0), max(lon_edges.size - 1, 0))
        return np.zeros(shape, dtype=int), pd.DataFrame(columns=['lon_bin', 'lat_bin', 'count'])

    work = df.copy()
    lon_vals = work[lon_col].to_numpy(dtype=float, copy=False)
    lat_vals = work[lat_col].to_numpy(dtype=float, copy=False)

    lon_cont = _to_continuous_lon(lon_vals, lon_ref=lon_ref, crosses_dateline=crosses_dateline)
    lon_bin = np.digitize(lon_cont, lon_edges, right=False) - 1
    lat_bin = np.digitize(lat_vals, lat_edges, right=False) - 1

    work['lon_bin'] = lon_bin
    work['lat_bin'] = lat_bin

    valid = (
        (work['lon_bin'] >= 0) & (work['lon_bin'] < lon_edges.size - 1)
        & (work['lat_bin'] >= 0) & (work['lat_bin'] < lat_edges.size - 1)
    )
    work = work[valid].copy()
    if work.empty:
        shape = (max(lat_edges.size - 1, 0), max(lon_edges.size - 1, 0))
        return np.zeros(shape, dtype=int), pd.DataFrame(columns=['lon_bin', 'lat_bin', 'count'])

    if dedupe_cols:
        keep_cols = [c for c in dedupe_cols if c in work.columns]
        keep_cols.extend(['lon_bin', 'lat_bin'])
        work = work.drop_duplicates(subset=keep_cols, keep='first')

    grouped = (
        work.groupby(['lat_bin', 'lon_bin'], as_index=False)
        .size()
        .rename(columns={'size': 'count'})
    )

    mat = np.zeros((lat_edges.size - 1, lon_edges.size - 1), dtype=int)
    if not grouped.empty:
        mat[grouped['lat_bin'].to_numpy(), grouped['lon_bin'].to_numpy()] = grouped['count'].to_numpy(dtype=int)
    return mat, grouped.rename(columns={'size': 'count'})


def _load_meta_daily_points_for_years(
    *,
    kind: str,
    start_year: int,
    end_year: int,
    meta_output_root: str | Path | None = None,
) -> pd.DataFrame:
    """读取指定 kind 的 META 日尺度记录，并按年份与当前区域做裁剪。"""
    kind_l = str(kind).lower()
    if kind_l not in {'acs', 'acl', 'cs', 'cl'}:
        raise ValueError("kind 必须是 'acs'|'acl'|'cs'|'cl'")

    region_slug = _current_region_key()
    root = _ensure_meta_tracks_root(meta_output_root)
    region_dir = Path(root) / region_slug
    if not region_dir.exists():
        raise FileNotFoundError(f"区域 META 目录不存在：{region_dir}")

    file_candidate = region_dir / f"{kind_l}_daily.parquet"
    dir_candidate = region_dir / f"{kind_l}_daily_tmp"
    if file_candidate.exists() and file_candidate.is_file():
        source = file_candidate
    elif file_candidate.exists() and file_candidate.is_dir():
        source = file_candidate
    elif dir_candidate.exists() and dir_candidate.is_dir():
        source = dir_candidate
    else:
        raise FileNotFoundError(f"未找到 daily 数据源：{file_candidate} 或 {dir_candidate}")

    cols = ['track_id', 'time', 'center_lon', 'center_lat']
    df = pd.read_parquet(source, columns=cols)
    if df.empty:
        return df

    if np.issubdtype(df['time'].dtype, np.number):
        time_int = pd.to_numeric(df['time'], errors='coerce').astype('Int64')
        yyyymmdd_min = int(f"{start_year:04d}0101")
        yyyymmdd_max = int(f"{end_year:04d}1231")
        mask_time = (time_int >= yyyymmdd_min) & (time_int <= yyyymmdd_max)
    else:
        date_series = pd.to_datetime(df['time'], errors='coerce')
        mask_time = (
            date_series >= pd.Timestamp(year=start_year, month=1, day=1)
        ) & (
            date_series <= pd.Timestamp(year=end_year, month=12, day=31)
        )

    df = df[mask_time].copy()
    if df.empty:
        return df

    date_series = convert_date(df['time'])
    if isinstance(date_series, pd.Timestamp):
        df['date'] = pd.Series([date_series] * len(df), index=df.index)
    else:
        df['date'] = pd.to_datetime(date_series, errors='coerce')
    df = df[df['date'].notna()].copy()
    if df.empty:
        return df

    lon_vals = df['center_lon'].to_numpy(dtype=float, copy=False)
    lat_vals = df['center_lat'].to_numpy(dtype=float, copy=False)
    lon_mask = _region_lon_mask(lon_vals, lonmin, lonmax)
    lat_mask = (lat_vals >= latmin) & (lat_vals <= latmax)
    df = df[lon_mask & lat_mask].copy()
    if df.empty:
        return df

    df['date'] = pd.to_datetime(df['date']).dt.normalize()
    return df[['track_id', 'date', 'center_lon', 'center_lat']]


def build_glorys_eke_native_grid(
    start_year: int = 2002,
    end_year: int = 2022,
    *,
    depth: float | int = 0.0,
    dask_scheduler: str = 'processes',
    dask_workers: int | None = None,
    dask_memory_limit: str | None = None,
    dask_batch_size: int = 32,
    verbose: bool = True,
) -> dict:
    """在 GLORYS 原始网格上计算时段平均 EKE。

    按天读取 GLORYS 的 uo/vo，在目标深度层计算 EKE = 0.5·(u'² + v'²) 并在时间维聚合。

    参数:
        - start_year (int): 起始年份（闭区间），默认 2002。
        - end_year (int): 结束年份（闭区间），默认 2022。
        - depth (float | int): 目标深度（米），按最近深度层提取，默认 0.0。
        - dask_scheduler (str): distributed 调度器（'threads'|'processes'），默认 'processes'。
        - dask_workers (int | None): worker 数；None 时自动估计。
        - dask_memory_limit (str | None): 每 worker 内存上限（如 '4GB'、'6GB'）。
        - dask_batch_size (int): 分批提交任务大小，限制同时在队列中的 futures 数量，默认 32。
        - verbose (bool): 是否打印进度，默认 True。

    返回:
        - dict: 含 'lon'、'lat'、'eke'、'count'、'meta'。
    """
    if end_year < start_year:
        raise ValueError("end_year 必须大于等于 start_year")

    dates = pd.date_range(
        start=pd.Timestamp(year=int(start_year), month=1, day=1),
        end=pd.Timestamp(year=int(end_year), month=12, day=31),
        freq='D',
    )

    file_pairs: list[tuple[pd.Timestamp, str]] = []
    for dt in dates:
        try:
            file_pairs.append((pd.Timestamp(dt), get_glorys_filepath(dt)))
        except Exception:
            continue

    if not file_pairs:
        raise RuntimeError("未找到可用 GLORYS 文件，无法计算 EKE")

    lon_raw = None
    lat_raw = None
    for _, nc_path in file_pairs:
        try:
            with Dataset(nc_path, 'r') as ds0:
                lon_raw = np.asarray(ds0.variables['longitude'][:], dtype=float)
                lat_raw = np.asarray(ds0.variables['latitude'][:], dtype=float)
                break
        except Exception:
            continue

    if lon_raw is None or lat_raw is None:
        raise RuntimeError("无法读取 GLORYS 网格坐标，无法计算 EKE")

    def _daily_stats(nc_path: str, target_depth: float):
        try:
            with Dataset(nc_path, 'r') as ds:
                gdep = np.asarray(ds.variables['depth'][:], dtype=float)
                if gdep.size == 0:
                    return None

                k = int(np.argmin(np.abs(gdep - float(target_depth))))
                u = np.asarray(np.ma.filled(ds.variables['uo'][0, k, :, :], np.nan), dtype=float)
                v = np.asarray(np.ma.filled(ds.variables['vo'][0, k, :, :], np.nan), dtype=float)

                valid = np.isfinite(u) & np.isfinite(v)
                if not np.any(valid):
                    return None

                su = np.zeros_like(u, dtype=float)
                sv = np.zeros_like(v, dtype=float)
                su2 = np.zeros_like(u, dtype=float)
                sv2 = np.zeros_like(v, dtype=float)
                cc = np.zeros_like(u, dtype=float)
                su[valid] = u[valid]
                sv[valid] = v[valid]
                su2[valid] = u[valid] * u[valid]
                sv2[valid] = v[valid] * v[valid]
                cc[valid] = 1.0
                return su, sv, su2, sv2, cc
        except Exception:
            return None

    sum_u = None
    sum_v = None
    sum_u2 = None
    sum_v2 = None
    count = None
    n_used = 0
    total_tasks = len(file_pairs)

    def _accumulate(item):
        nonlocal sum_u, sum_v, sum_u2, sum_v2, count, n_used
        if item is None:
            return
        su, sv, su2, sv2, cc = item
        if sum_u is None:
            sum_u = np.zeros_like(su, dtype=float)
            sum_v = np.zeros_like(sv, dtype=float)
            sum_u2 = np.zeros_like(su2, dtype=float)
            sum_v2 = np.zeros_like(sv2, dtype=float)
            count = np.zeros_like(cc, dtype=float)
        if su.shape != sum_u.shape:
            return
        sum_u += su
        sum_v += sv
        sum_u2 += su2
        sum_v2 += sv2
        count += cc
        n_used += 1

    depth_f = float(depth)
    batch_size = max(1, int(dask_batch_size))
    worker_count = int(dask_workers) if dask_workers is not None else max(1, (os.cpu_count() or 1))
    sched = str(dask_scheduler).lower()
    processes = (sched == 'processes')
    cluster = LocalCluster(
        n_workers=worker_count,
        threads_per_worker=1,
        processes=processes,
        memory_limit=dask_memory_limit or 'auto',
    )
    client = Client(cluster)
    pbar = tqdm(
        total=total_tasks,
        desc='EKE calculation',
        unit='day',
        dynamic_ncols=True,
        disable=not verbose,
    )
    try:
        done_tasks = 0
        for i in range(0, len(file_pairs), batch_size):
            batch = file_pairs[i:i + batch_size]
            futures = [client.submit(_daily_stats, path, depth_f) for _, path in batch]
            for fut in as_completed(futures):
                done_tasks += 1
                try:
                    _accumulate(fut.result())
                except Exception:
                    pass
                pbar.update(1)
    finally:
        pbar.close()
        client.close()
        cluster.close()

    if sum_u is None or count is None:
        raise RuntimeError("未读取到可用 GLORYS 数据，无法计算 EKE")

    with np.errstate(invalid='ignore', divide='ignore'):
        mean_u = np.where(count > 0, sum_u / count, np.nan)
        mean_v = np.where(count > 0, sum_v / count, np.nan)
        var_u = np.where(count > 0, sum_u2 / count - mean_u * mean_u, np.nan)
        var_v = np.where(count > 0, sum_v2 / count - mean_v * mean_v, np.nan)
        eke = 0.5 * (var_u + var_v)

    eke = np.where(eke >= 0, eke, np.nan)

    if verbose:
        print(
            f"[EKE] Native grid done: {start_year}-{end_year}, depth~{float(depth):g} m, "
            f"valid days={n_used}"
        )

    return {
        'lon': np.asarray(lon_raw, dtype=float),
        'lat': np.asarray(lat_raw, dtype=float),
        'eke': np.asarray(eke, dtype=float),
        'count': np.asarray(count, dtype=float),
        'meta': {
            'start_year': int(start_year),
            'end_year': int(end_year),
            'depth': float(depth),
            'n_days_used': int(n_used),
            'dask_scheduler': str(dask_scheduler),
            'dask_memory_limit': str(dask_memory_limit) if dask_memory_limit is not None else None,
        },
    }


def remap_glorys_eke_to_euler_grid(
    eke_native: dict,
    *,
    grid_step_deg: float = 1.0,
    method: str = 'linear',
) -> dict:
    """将 GLORYS 原网格 EKE 插值到当前区域的欧拉网格（默认 1°）。

    参数:
        - eke_native (dict): build_glorys_eke_native_grid 的输出（含 lon/lat/eke）。
        - grid_step_deg (float): 目标欧拉网格步长（°），默认 1.0。
        - method (str): RegularGridInterpolator 插值方法，默认 'linear'。

    返回:
        - dict: 含 'eke_grid'（区域外置 NaN）、'grid'（边与元信息）、'meta'。
    """
    lon_src = np.asarray(eke_native['lon'], dtype=float)
    lat_src = np.asarray(eke_native['lat'], dtype=float)
    eke_src = np.asarray(eke_native['eke'], dtype=float)

    if lon_src.ndim != 1 or lat_src.ndim != 1 or eke_src.ndim != 2:
        raise ValueError("eke_native 的 lon/lat/eke 维度不正确")
    if eke_src.shape != (lat_src.size, lon_src.size):
        raise ValueError("eke_native['eke'] 形状必须为 (n_lat, n_lon)")

    lon_edges, lat_edges, grid_meta = _build_euler_grid_edges(grid_step_deg=grid_step_deg)
    lon_ref = float(grid_meta['lon_start_continuous'])
    crosses_dateline = bool(grid_meta['crosses_dateline'])

    lon_norm = _normalize_lon_array(lon_src)
    if crosses_dateline:
        lon_work = lon_norm.copy()
        lon_work[lon_work < lon_ref] += 360.0
    else:
        lon_work = lon_norm.copy()

    sort_idx = np.argsort(lon_work)
    lon_sorted = lon_work[sort_idx]
    eke_sorted = eke_src[:, sort_idx]

    lat_centers = 0.5 * (lat_edges[:-1] + lat_edges[1:])
    lon_centers = 0.5 * (lon_edges[:-1] + lon_edges[1:])
    tgt_lon_2d, tgt_lat_2d = np.meshgrid(lon_centers, lat_centers)

    interp = RegularGridInterpolator(
        (lat_src, lon_sorted),
        eke_sorted,
        method=method,
        bounds_error=False,
        fill_value=np.nan,
    )
    pts = np.column_stack([tgt_lat_2d.ravel(), tgt_lon_2d.ravel()])
    eke_grid = interp(pts).reshape(tgt_lat_2d.shape)

    lon_mask = _region_lon_mask(lon_centers, lonmin, lonmax)
    lat_mask = (lat_centers >= latmin) & (lat_centers <= latmax)
    region_mask = lat_mask[:, None] & lon_mask[None, :]
    eke_grid = np.where(region_mask, eke_grid, np.nan)

    return {
        'eke_grid': np.asarray(eke_grid, dtype=float),
        'grid': {
            'lon_edges': lon_edges,
            'lat_edges': lat_edges,
            **grid_meta,
        },
        'meta': {
            **eke_native.get('meta', {}),
            'grid_step_deg': float(grid_step_deg),
            'remap_method': str(method),
        },
    }


def build_euler_eke_grid(
    start_year: int = 2002,
    end_year: int = 2022,
    *,
    depth: float | int = 0.0,
    grid_step_deg: float = 1.0,
    method: str = 'linear',
    dask_scheduler: str = 'processes',
    dask_workers: int | None = None,
    dask_memory_limit: str | None = None,
    dask_batch_size: int = 32,
    verbose: bool = True,
) -> dict:
    """一站式计算：GLORYS 原网格 EKE → 欧拉网格 EKE。

    内部依次调用 build_glorys_eke_native_grid 与 remap_glorys_eke_to_euler_grid。

    参数:
        - start_year (int): 起始年份（闭区间），默认 2002。
        - end_year (int): 结束年份（闭区间），默认 2022。
        - depth (float | int): 目标深度（米），默认 0.0。
        - grid_step_deg (float): 欧拉网格步长（°），默认 1.0。
        - method (str): 插值方法，默认 'linear'。
        - dask_scheduler (str): distributed 调度器，默认 'processes'。
        - dask_workers (int | None): worker 数；None 时自动估计。
        - dask_memory_limit (str | None): 每 worker 内存上限。
        - dask_batch_size (int): 分批提交任务大小，默认 32。
        - verbose (bool): 是否打印进度，默认 True。

    返回:
        - dict: 含 'native'（原网格结果）与 'euler'（重映射结果）。
    """
    native = build_glorys_eke_native_grid(
        start_year=start_year,
        end_year=end_year,
        depth=depth,
        dask_scheduler=dask_scheduler,
        dask_workers=dask_workers,
        dask_memory_limit=dask_memory_limit,
        dask_batch_size=dask_batch_size,
        verbose=verbose,
    )
    remapped = remap_glorys_eke_to_euler_grid(
        native,
        grid_step_deg=grid_step_deg,
        method=method,
    )
    return {
        'native': native,
        'euler': remapped,
    }


def export_glorys_eke_native_dask(
    start_year: int = 2002,
    end_year: int = 2022,
    *,
    depth: float | int = 0.0,
    output_path: str | Path | None = None,
    chunks: tuple[int, int] = (256, 256),
    dask_scheduler: str = 'processes',
    dask_workers: int | None = 8,
    dask_memory_limit: str | None = '6GB',
    dask_batch_size: int = 16,
    overwrite: bool = False,
    verbose: bool = True,
) -> str:
    """使用 Dask 计算 GLORYS 原始分辨率 EKE 并保存到本地 zarr。

    先调用 build_glorys_eke_native_grid 计算原始分辨率 EKE，再写入 zarr（默认 GLORYS_processed/eke.zarr）；
    默认参数为 8 核 + 内存受限预设。

    参数:
        - start_year (int): 起始年份（闭区间），默认 2002。
        - end_year (int): 结束年份（闭区间），默认 2022。
        - depth (float | int): 目标深度（米），默认 0.0。
        - output_path (str | Path | None): 输出 zarr 路径；None 时用默认目录。
        - chunks (tuple[int, int]): zarr 二维分块大小 (lat, lon)，默认 (256, 256)。
        - dask_scheduler (str): distributed 调度器，默认 'processes'。
        - dask_workers (int | None): worker 数，默认 8。
        - dask_memory_limit (str | None): 每 worker 内存上限，默认 '6GB'。
        - dask_batch_size (int): 分批提交任务大小，默认 16。
        - overwrite (bool): 输出已存在时是否覆盖，默认 False。
        - verbose (bool): 是否打印进度，默认 True。

    返回:
        - str: 保存的 zarr 路径。
    """
    if output_path is None:
        output_path = glorys_processed_root / 'eke.zarr'
    out_path = Path(output_path)
    if out_path.exists() and not overwrite:
        raise FileExistsError(f"EKE 文件已存在: {out_path}. 如需覆盖请设置 overwrite=True")

    native = build_glorys_eke_native_grid(
        start_year=start_year,
        end_year=end_year,
        depth=depth,
        dask_scheduler=dask_scheduler,
        dask_workers=dask_workers,
        dask_memory_limit=dask_memory_limit,
        dask_batch_size=dask_batch_size,
        verbose=verbose,
    )

    try:
        import zarr  # type: ignore
    except Exception as exc:
        raise RuntimeError("zarr 未安装，无法写入 zarr 格式。") from exc

    if out_path.exists():
        if out_path.is_dir():
            pass
        else:
            raise FileExistsError(f"输出路径已存在且不是目录: {out_path}")

    out_path.parent.mkdir(parents=True, exist_ok=True)
    root = zarr.open_group(out_path, mode='w')
    lon_arr = np.asarray(native['lon'], dtype=float)
    lat_arr = np.asarray(native['lat'], dtype=float)
    eke_arr = np.asarray(native['eke'], dtype=float)
    cnt_arr = np.asarray(native['count'], dtype=float)

    # 显式提供 shape/dtype 以兼容 zarr v3 的 create_dataset 签名。
    ds_lon = root.create_dataset('lon', shape=lon_arr.shape, dtype=lon_arr.dtype, chunks=(lon_arr.size,))
    ds_lat = root.create_dataset('lat', shape=lat_arr.shape, dtype=lat_arr.dtype, chunks=(lat_arr.size,))
    ds_eke = root.create_dataset('eke', shape=eke_arr.shape, dtype=eke_arr.dtype, chunks=chunks)
    ds_count = root.create_dataset('count', shape=cnt_arr.shape, dtype=cnt_arr.dtype, chunks=chunks)

    ds_lon[...] = lon_arr
    ds_lat[...] = lat_arr
    ds_eke[...] = eke_arr
    ds_count[...] = cnt_arr
    for key, val in native.get('meta', {}).items():
        root.attrs[str(key)] = val

    if verbose:
        print(f"[EKE] Saved native file: {out_path}")
    return str(out_path)


def load_glorys_eke_native(file_path: str | Path) -> dict:
    """从本地 zarr 文件加载 GLORYS 原始分辨率 EKE。

    参数:
        - file_path (str | Path): EKE zarr 目录路径（export_glorys_eke_native_dask 的输出）。

    返回:
        - dict: 含 lon/lat/eke/count 数组及 meta 元信息。
    """
    p = Path(file_path)
    if not p.exists():
        raise FileNotFoundError(
            f"未找到本地 EKE 文件: {p}。请先运行 export_glorys_eke_native_dask(...) 生成。"
        )
    if not (p.is_dir() or p.suffix.lower() == '.zarr'):
        raise ValueError("仅支持读取 zarr 目录作为 EKE 原始分辨率缓存。")

    try:
        import zarr  # type: ignore
    except Exception as exc:
        raise RuntimeError("zarr 未安装，无法读取 zarr 格式。") from exc
    root = zarr.open_group(p, mode='r')
    meta = dict(root.attrs)
    return {
        'lon': np.asarray(root['lon'], dtype=float),
        'lat': np.asarray(root['lat'], dtype=float),
        'eke': np.asarray(root['eke'], dtype=float),
        'count': np.asarray(root['count'], dtype=float),
        'meta': meta,
    }


def _euler_summary_year_worker(args: tuple) -> tuple[pd.DataFrame, pd.DataFrame]:
    """按年份处理 Euler summary 的 baseline 与统一异常子集（支持 Dask 并行）。"""
    (
        year,
        cfg,
        lon_min_bound,
        lon_max_bound,
        lat_min_bound,
        lat_max_bound,
    ) = args

    try:
        df_year = load_argo_data(int(year))
    except FileNotFoundError:
        return pd.DataFrame(), pd.DataFrame()
    except Exception as exc:
        print(f"[build_euler_grid_summary] 读取 Argo {year} 失败: {exc}")
        return pd.DataFrame(), pd.DataFrame()

    if df_year.empty:
        return pd.DataFrame(), pd.DataFrame()

    lon_vals = df_year['Longitude'].to_numpy(dtype=float, copy=False)
    lat_vals = df_year['Latitude'].to_numpy(dtype=float, copy=False)
    lon_mask = _region_lon_mask(lon_vals, lon_min_bound, lon_max_bound)
    lat_mask = (lat_vals >= lat_min_bound) & (lat_vals <= lat_max_bound)
    df_geo = df_year[lon_mask & lat_mask].copy()
    if df_geo.empty:
        return pd.DataFrame(), pd.DataFrame()

    baseline_year = (
        df_geo.sort_values(['Profile_number', 'Depth'])
        .groupby('Profile_number', as_index=False)
        .first()[['Profile_number', 'Longitude', 'Latitude', 'Year', 'Month', 'Day']]
    )
    baseline_year['profile_uid'] = (
        baseline_year['Year'].astype(int).astype(str) + '-'
        + baseline_year['Month'].astype(int).astype(str).str.zfill(2) + '-'
        + baseline_year['Day'].astype(int).astype(str).str.zfill(2) + '-'
        + baseline_year['Profile_number'].astype(str)
    )
    baseline_out = baseline_year[['profile_uid', 'Longitude', 'Latitude']]

    anomalies = calculate_delta_do(
        df_geo,
        detection_config=cfg,
        remove_outliers=True,
        verbose=False,
    )
    if anomalies.empty:
        return baseline_out, pd.DataFrame()

    depth_col = 'depth' if 'depth' in anomalies.columns else None
    if depth_col is None:
        return baseline_out, pd.DataFrame()

    if cfg.anomaly_min_depth is not None and cfg.anomaly_min_depth > 0:
        anomalies = anomalies[anomalies[depth_col] >= cfg.anomaly_min_depth].copy()
    if anomalies.empty:
        return baseline_out, pd.DataFrame()

    anomalies = _keep_best_anomaly_per_profile(anomalies, cfg)

    if all(c in anomalies.columns for c in ['Year', 'Month', 'Day']):
        anomalies['profile_uid'] = (
            anomalies['Year'].astype(int).astype(str) + '-'
            + anomalies['Month'].astype(int).astype(str).str.zfill(2) + '-'
            + anomalies['Day'].astype(int).astype(str).str.zfill(2) + '-'
            + anomalies['Profile_number'].astype(str)
        )
    else:
        anomalies['profile_uid'] = anomalies['Profile_number'].astype(str)

    keep_cols = [
        c for c in [
            'profile_uid', 'Profile_number', 'Longitude', 'Latitude', 'delta_do',
            'delta_aou', 'trim_score', 'anomaly_score', 'primary_metric',
            'primary_value', depth_col
        ]
        if c in anomalies.columns
    ]
    do_out = anomalies[keep_cols].rename(columns={depth_col: 'depth'})

    return baseline_out, do_out


def build_euler_grid_summary(
    start_year: int,
    end_year: int,
    *,
    grid_step_deg: float = 1.0,
    meta_output_root: str | Path | None = None,
    detection_config: DetectionConfig | None = None,
    use_dask: bool = True,
    dask_scheduler: str = 'processes',
    dask_workers: int | None = None,
    dask_show_progress: bool = True,
) -> dict:
    """构建欧拉网格汇总：ACE/CE 天数与 Argo 异常出现率。

    参数:
        - start_year (int): 起始年份（闭区间）。
        - end_year (int): 结束年份（闭区间）。
        - grid_step_deg (float): 欧拉网格步长（°），默认 1.0。
        - meta_output_root (str | Path | None): META 导出根目录；None 时用默认。
        - detection_config (DetectionConfig | None): 异常识别配置；None 时用默认。
        - use_dask (bool): 是否用 Dask 并行，默认 True。
        - dask_scheduler (str): distributed 调度器，默认 'processes'。
        - dask_workers (int | None): worker 数；None 时自动估计。
        - dask_show_progress (bool): 是否显示 Dask 进度，默认 True。

    返回:
        - dict: 含各类型网格天数、Argo baseline 与异常计数及网格元信息的汇总字典。

    说明:
        统计口径:

            - 同日同网格同类型只计 1 天（按 date+grid 去重）。
            - 异常剖面先按 Profile_number 去重，再按 profile_id+grid 计数，避免单剖面多峰值重复。
            - 使用同一套区域边界与网格边界，确保三图可直接比较。
            - 深度筛选由 DetectionConfig 统一管理。
            - 关联分析默认用异常出现率 anomaly_profiles / argo_baseline_profiles（仅 baseline>0 网格）。
    """
    if end_year < start_year:
        raise ValueError("end_year 必须大于等于 start_year")

    cfg = _resolve_detection_config(detection_config)

    lon_edges, lat_edges, grid_meta = _build_euler_grid_edges(grid_step_deg=grid_step_deg)
    lon_ref = float(grid_meta['lon_start_continuous'])
    crosses_dateline = bool(grid_meta['crosses_dateline'])

    # 1) 涡旋网格天数
    acl_df = _load_meta_daily_points_for_years(
        kind='acl',
        start_year=start_year,
        end_year=end_year,
        meta_output_root=meta_output_root,
    )
    acs_df = _load_meta_daily_points_for_years(
        kind='acs',
        start_year=start_year,
        end_year=end_year,
        meta_output_root=meta_output_root,
    )
    cl_df = _load_meta_daily_points_for_years(
        kind='cl',
        start_year=start_year,
        end_year=end_year,
        meta_output_root=meta_output_root,
    )
    cs_df = _load_meta_daily_points_for_years(
        kind='cs',
        start_year=start_year,
        end_year=end_year,
        meta_output_root=meta_output_root,
    )

    acl_grid, acl_grouped = _grid_count_from_points(
        acl_df,
        lon_col='center_lon',
        lat_col='center_lat',
        lon_edges=lon_edges,
        lat_edges=lat_edges,
        crosses_dateline=crosses_dateline,
        lon_ref=lon_ref,
        dedupe_cols=['date'],
    )
    acs_grid, acs_grouped = _grid_count_from_points(
        acs_df,
        lon_col='center_lon',
        lat_col='center_lat',
        lon_edges=lon_edges,
        lat_edges=lat_edges,
        crosses_dateline=crosses_dateline,
        lon_ref=lon_ref,
        dedupe_cols=['date'],
    )
    cl_grid, cl_grouped = _grid_count_from_points(
        cl_df,
        lon_col='center_lon',
        lat_col='center_lat',
        lon_edges=lon_edges,
        lat_edges=lat_edges,
        crosses_dateline=crosses_dateline,
        lon_ref=lon_ref,
        dedupe_cols=['date'],
    )
    cs_grid, cs_grouped = _grid_count_from_points(
        cs_df,
        lon_col='center_lon',
        lat_col='center_lat',
        lon_edges=lon_edges,
        lat_edges=lat_edges,
        crosses_dateline=crosses_dateline,
        lon_ref=lon_ref,
        dedupe_cols=['date'],
    )

    ace_grid = acl_grid + acs_grid
    ce_grid = cl_grid + cs_grid

    # 2) Argo 异常网格剖面数 + Argo 观测机会网格
    do_frames = []
    baseline_frames = []
    year_args = [
        (
            year,
            cfg,
            lonmin,
            lonmax,
            latmin,
            latmax,
        )
        for year in range(start_year, end_year + 1)
    ]

    year_results: list[tuple[pd.DataFrame, pd.DataFrame]] = []
    if use_dask and len(year_args) > 1:
        scheduler = str(dask_scheduler).lower().strip()
        if scheduler not in {'threads', 'processes', 'synchronous'}:
            scheduler = 'processes'
        compute_kwargs: dict = {'scheduler': scheduler}
        if dask_workers is not None and scheduler in {'threads', 'processes'}:
            compute_kwargs['num_workers'] = max(1, int(dask_workers))

        delayed_tasks = [delayed(_euler_summary_year_worker)(arg) for arg in year_args]
        try:
            if dask_show_progress:
                with ProgressBar():
                    year_results = list(compute(*delayed_tasks, **compute_kwargs))
            else:
                year_results = list(compute(*delayed_tasks, **compute_kwargs))
        except Exception as exc:
            print(f"[build_euler_grid_summary] Dask 并行失败，回退串行: {exc}")
            year_results = [_euler_summary_year_worker(arg) for arg in year_args]
    else:
        year_results = [_euler_summary_year_worker(arg) for arg in year_args]

    for baseline_year, do_year in year_results:
        if not baseline_year.empty:
            baseline_frames.append(baseline_year)
        if not do_year.empty:
            do_frames.append(do_year)

    do_df = pd.concat(do_frames, ignore_index=True) if do_frames else pd.DataFrame()
    baseline_df = pd.concat(baseline_frames, ignore_index=True) if baseline_frames else pd.DataFrame()

    baseline_grid, baseline_grouped = _grid_count_from_points(
        baseline_df,
        lon_col='Longitude',
        lat_col='Latitude',
        lon_edges=lon_edges,
        lat_edges=lat_edges,
        crosses_dateline=crosses_dateline,
        lon_ref=lon_ref,
        dedupe_cols=['profile_uid'],
    )

    do_grid, do_grouped = _grid_count_from_points(
        do_df,
        lon_col='Longitude',
        lat_col='Latitude',
        lon_edges=lon_edges,
        lat_edges=lat_edges,
        crosses_dateline=crosses_dateline,
        lon_ref=lon_ref,
        dedupe_cols=['profile_uid'],
    )

    high_do_occurrence_ratio = np.full_like(do_grid, np.nan, dtype=float)
    valid_baseline = baseline_grid > 0
    if np.any(valid_baseline):
        high_do_occurrence_ratio[valid_baseline] = do_grid[valid_baseline] / baseline_grid[valid_baseline]

    ocean_mask = _build_ocean_mask_for_grid(
        lon_edges,
        lat_edges,
        crosses_dateline=crosses_dateline,
        lon_ref=lon_ref,
    )
    observation_mask = baseline_grid > 0
    analysis_mask = ocean_mask & observation_mask

    return {
        'meta': {
            'region_key': _current_region_key(),
            'start_year': int(start_year),
            'end_year': int(end_year),
            'grid_step_deg': float(grid_step_deg),
            'detection_method': cfg.method,
            'detection_label': cfg.threshold_label(),
            'detection_file_stem': cfg.file_stem(),
            'anomaly_min_depth': float(cfg.anomaly_min_depth) if cfg.anomaly_min_depth is not None else np.nan,
            'do_threshold': float(cfg.do_threshold),
            'salinity_threshold': float(cfg.salinity_threshold),
            'temperature_threshold': float(cfg.temperature_threshold),
        },
        'grid': {
            'lon_edges': lon_edges,
            'lat_edges': lat_edges,
            **grid_meta,
        },
        'acl_days': acl_grid,
        'acs_days': acs_grid,
        'cl_days': cl_grid,
        'cs_days': cs_grid,
        'ace_days': ace_grid,
        'ce_days': ce_grid,
        'eddy_days_total': ace_grid + ce_grid,
        'anomaly_profiles': do_grid,
        'anomaly_occurrence_ratio': high_do_occurrence_ratio,
        'high_do_profiles': do_grid,
        'high_do_occurrence_ratio': high_do_occurrence_ratio,
        'argo_baseline_profiles': baseline_grid,
        'acl_table': acl_grouped,
        'acs_table': acs_grouped,
        'cl_table': cl_grouped,
        'cs_table': cs_grouped,
        'argo_baseline_table': baseline_grouped,
        'anomaly_table': do_grouped,
        'high_do_table': do_grouped,
        'ocean_mask': ocean_mask,
        'observation_mask': observation_mask,
        'analysis_mask': analysis_mask,
    }


def plot_euler_grid_summary(
    summary: dict,
    *,
    save_fig: bool = False,
    show_fig: bool = True,
    save_data: bool = False,
    output_dir: str | Path | None = None,
    output_prefix: str = 'EddyDays',
    cmap_eddy: str = 'YlOrRd',
    cmap_do: str = 'Reds',
) -> dict:
    """绘制欧拉网格三图：ACE 天数、CE 天数、Argo 异常出现率。

    参数:
        - summary (dict): build_euler_grid_summary 的输出。
        - save_fig (bool): 是否保存图像，默认 False。
        - show_fig (bool): 是否显示图像，默认 True。
        - save_data (bool): 是否保存底层网格数据，默认 False。
        - output_dir (str | Path | None): 输出目录；None 时使用默认路径。
        - output_prefix (str): 输出文件名前缀，默认 'EddyDays'。
        - cmap_eddy (str): 涡旋天数图色标，默认 'YlOrRd'。
        - cmap_do (str): 异常出现率图色标，默认 'Reds'。

    返回:
        - dict: 含图/数据保存路径等信息。
    """
    grid = summary['grid']
    meta = summary['meta']

    lon_edges = np.asarray(grid['lon_edges'], dtype=float)
    lat_edges = np.asarray(grid['lat_edges'], dtype=float)
    ace_days = np.asarray(summary['ace_days'])
    ce_days = np.asarray(summary['ce_days'])
    anomaly_rate = np.asarray(
        summary.get('anomaly_occurrence_ratio', summary.get('high_do_occurrence_ratio', summary['high_do_profiles'])),
        dtype=float,
    )

    crosses_dateline = bool(grid.get('crosses_dateline', False))
    central_lon = 180 if crosses_dateline else 0
    data_crs = ccrs.PlateCarree()
    map_crs = ccrs.PlateCarree(central_longitude=central_lon)

    fig, axes = plt.subplots(
        1,
        3,
        figsize=(30, 9),
        subplot_kw={'projection': map_crs},
        constrained_layout=True,
    )

    base_ocean = _BASEMAP_COLORS['ocean']
    base_land = _BASEMAP_COLORS['land']
    coast_color = _BASEMAP_COLORS['coastline']
    grid_color = _BASEMAP_COLORS['grid']

    vmax_eddy = float(max(np.nanmax(ace_days), np.nanmax(ce_days), 1.0))
    vmax_anom = float(np.nanmax(anomaly_rate)) if np.any(np.isfinite(anomaly_rate)) else 1.0
    if not np.isfinite(vmax_anom) or vmax_anom <= 0:
        vmax_anom = 1.0

    lon_extent_min = float(lon_edges[0])
    lon_extent_max = float(lon_edges[-1])

    lon_mesh, lat_mesh = np.meshgrid(lon_edges, lat_edges)
    panel_defs = [
        ('ACE Eddy Days', ace_days, cmap_eddy, Normalize(vmin=0.0, vmax=vmax_eddy)),
        ('CE Eddy Days', ce_days, cmap_eddy, Normalize(vmin=0.0, vmax=vmax_eddy)),
        ('Anomaly Occurrence Rate', anomaly_rate, cmap_do, Normalize(vmin=0.0, vmax=vmax_anom)),
    ]

    mappables = []
    for ax, (title, mat, cmap, norm) in zip(axes, panel_defs):
        ax.set_facecolor(base_ocean)
        ax.add_feature(cfeature.OCEAN, facecolor=base_ocean, zorder=0)
        ax.add_feature(cfeature.LAND, facecolor=base_land, edgecolor=coast_color, linewidth=0.5, zorder=0)
        ax.add_feature(cfeature.COASTLINE, linewidth=0.7, edgecolor=coast_color, zorder=1)

        gl = ax.gridlines(draw_labels=True, linewidth=0.4, color=grid_color, alpha=0.45, linestyle='--')
        gl.top_labels = False
        gl.right_labels = False

        hm = ax.pcolormesh(
            lon_mesh,
            lat_mesh,
            mat,
            cmap=cmap,
            norm=norm,
            shading='auto',
            transform=data_crs,
            zorder=2,
        )
        mappables.append(hm)
        ax.set_title(title, fontsize=14)
        ax.set_extent([lon_extent_min, lon_extent_max, float(lat_edges[0]), float(lat_edges[-1])], crs=data_crs)

    cbar1 = fig.colorbar(mappables[0], ax=axes[:2], orientation='horizontal', fraction=0.05, pad=0.08)
    cbar1.set_label('Eddy Days per Grid Cell', fontsize=12)
    cbar2 = fig.colorbar(mappables[2], ax=axes[2], orientation='horizontal', fraction=0.05, pad=0.08)
    cbar2.set_label('Anomaly Occurrence Rate per Grid Cell', fontsize=12)

    detection_label = meta.get('detection_label')
    if not detection_label:
        detection_label = f"ΔDO ≥ {meta.get('do_threshold', np.nan):g} μmol kg⁻¹"
    fig.suptitle(
        (
            f"EddyDays & Anomaly Summary ({meta['region_key']}, {meta['start_year']}-{meta['end_year']})\n"
            f"{detection_label}"
            f" | depth ≥ {meta.get('anomaly_min_depth', np.nan):g} m"
            f" | grid step = {meta['grid_step_deg']:.1f}°"
        ),
        fontsize=16,
    )

    if output_dir is None:
        output_dir = _detection_output_dir_from_meta('plot_euler_grid_summary', meta)
    else:
        output_dir = Path(output_dir)
    if save_fig or save_data:
        output_dir.mkdir(parents=True, exist_ok=True)

    prefix_safe = ''.join(ch if (ch.isalnum() or ch in {'_', '-'}) else '_' for ch in str(output_prefix).strip())
    if not prefix_safe:
        prefix_safe = 'EddyDays'

    out = {}
    run_tag = str(meta.get('detection_file_stem') or f"do{_format_detection_value(meta.get('do_threshold', 'NA'))}_depth{_format_detection_value(meta.get('anomaly_min_depth', 'NA'))}m")
    if save_fig:
        fname = output_dir / (
            f"{prefix_safe}_Summary_{meta['start_year']}_{meta['end_year']}_"
            f"{meta['grid_step_deg']:g}deg_{run_tag}.png"
        )
        fig.savefig(fname, dpi=300, bbox_inches='tight')
        out['figure'] = str(fname)
        print(f"Figure saved: {fname}")

    if show_fig:
        plt.show()
    plt.close(fig)

    if save_data:
        npz_path = output_dir / (
            f"{prefix_safe}_Summary_{meta['start_year']}_{meta['end_year']}_{meta['grid_step_deg']:g}deg_"
            f"{run_tag}.npz"
        )
        np.savez_compressed(
            npz_path,
            lon_edges=lon_edges,
            lat_edges=lat_edges,
            ace_days=ace_days,
            ce_days=ce_days,
            eddy_days_total=np.asarray(summary['eddy_days_total']),
            anomaly_profiles=np.asarray(summary.get('anomaly_profiles', summary.get('high_do_profiles'))),
            anomaly_occurrence_ratio=np.asarray(summary.get('anomaly_occurrence_ratio', summary.get('high_do_occurrence_ratio'))),
            high_do_profiles=np.asarray(summary.get('high_do_profiles', summary.get('anomaly_profiles'))),
            high_do_occurrence_ratio=np.asarray(summary.get('high_do_occurrence_ratio')),
            argo_baseline_profiles=np.asarray(summary['argo_baseline_profiles']),
            ocean_mask=np.asarray(summary['ocean_mask']),
            observation_mask=np.asarray(summary['observation_mask']),
            analysis_mask=np.asarray(summary['analysis_mask']),
        )
        out['npz'] = str(npz_path)
        print(f"Grid data saved: {npz_path}")

    return out


def _compute_grid_association_core(
    x_grid: np.ndarray,
    y_grid: np.ndarray,
    *,
    analysis_mask: np.ndarray | None = None,
    active_threshold: float | str | None = 'auto',
    active_quantile: float = 0.5,
    active_positive_only: bool = True,
    x_log_for_map: bool = True,
    y_log_for_map: bool = True,
) -> dict:
    """通用网格双变量关联核心统计。"""

    def _distance_correlation_1d(xv: np.ndarray, yv: np.ndarray) -> float:
        """计算一维样本的距离相关系数 dCor（范围 [0, 1]）。"""
        if xv.size < 3 or yv.size < 3:
            return np.nan
        a = np.abs(xv[:, None] - xv[None, :])
        b = np.abs(yv[:, None] - yv[None, :])

        A = a - a.mean(axis=0, keepdims=True) - a.mean(axis=1, keepdims=True) + a.mean()
        B = b - b.mean(axis=0, keepdims=True) - b.mean(axis=1, keepdims=True) + b.mean()

        dcov2 = float(np.mean(A * B))
        dvarx2 = float(np.mean(A * A))
        dvary2 = float(np.mean(B * B))
        if dvarx2 <= 0 or dvary2 <= 0:
            return np.nan

        dcor = np.sqrt(max(dcov2, 0.0) / np.sqrt(dvarx2 * dvary2))
        return float(dcor)

    x2 = np.asarray(x_grid, dtype=float)
    y2 = np.asarray(y_grid, dtype=float)
    finite_mask = np.isfinite(x2) & np.isfinite(y2)
    if analysis_mask is not None:
        am = np.asarray(analysis_mask, dtype=bool)
        if am.shape == finite_mask.shape:
            finite_mask = finite_mask & am

    x = x2[finite_mask]
    y = y2[finite_mask]
    if x.size == 0:
        raise RuntimeError("无可用网格用于关联统计")

    if active_threshold is None or str(active_threshold).lower() == 'auto':
        if active_positive_only:
            x_thr_base = x[x > 0]
        else:
            x_thr_base = x[np.isfinite(x)]
        if x_thr_base.size > 0:
            thr = float(np.quantile(x_thr_base, float(active_quantile)))
        else:
            thr = 0.0
    else:
        thr = float(active_threshold)

    active = x >= thr
    inactive = ~active
    mean_active = np.nanmean(y[active]) if np.any(active) else np.nan
    mean_inactive = np.nanmean(y[inactive]) if np.any(inactive) else np.nan

    uplift_ratio = np.nan
    uplift_pct = np.nan
    if np.isfinite(mean_active) and np.isfinite(mean_inactive) and mean_inactive > 0:
        uplift_ratio = mean_active / mean_inactive
        uplift_pct = (uplift_ratio - 1.0) * 100.0

    spearman_rho = np.nan
    spearman_p = np.nan
    pearson_r = np.nan
    pearson_p = np.nan
    distance_corr = np.nan
    nz_mask = (x > 0) | (y > 0)
    x_corr = x[nz_mask]
    y_corr = y[nz_mask]
    if x_corr.size >= 3 and np.nanstd(x_corr) > 0 and np.nanstd(y_corr) > 0:
        try:
            spearman_rho, spearman_p = spearmanr(x_corr, y_corr)
        except Exception:
            spearman_rho, spearman_p = np.nan, np.nan
        try:
            pearson_r, pearson_p = pearsonr(x_corr, y_corr)
        except Exception:
            pearson_r, pearson_p = np.nan, np.nan
        try:
            distance_corr = _distance_correlation_1d(x_corr, y_corr)
        except Exception:
            distance_corr = np.nan

    xmap = np.asarray(x2, dtype=float)
    ymap = np.asarray(y2, dtype=float)
    xx = np.log1p(np.where(np.isfinite(xmap), np.maximum(xmap, 0.0), np.nan)) if x_log_for_map else np.where(np.isfinite(xmap), xmap, np.nan)
    yy = np.log1p(np.where(np.isfinite(ymap), np.maximum(ymap, 0.0), np.nan)) if y_log_for_map else np.where(np.isfinite(ymap), ymap, np.nan)

    valid_map = np.isfinite(xx) & np.isfinite(yy)
    if analysis_mask is not None:
        am2 = np.asarray(analysis_mask, dtype=bool)
        if am2.shape == valid_map.shape:
            valid_map = valid_map & am2

    assoc_map = np.full_like(xx, np.nan, dtype=float)
    if np.any(valid_map):
        xmean = np.nanmean(xx[valid_map])
        ymean = np.nanmean(yy[valid_map])
        xstd = np.nanstd(xx[valid_map])
        ystd = np.nanstd(yy[valid_map])
        if xstd > 0 and ystd > 0:
            zx = (xx - xmean) / xstd
            zy = (yy - ymean) / ystd
            assoc_map[valid_map] = zx[valid_map] * zy[valid_map]

    return {
        'x': x,
        'y': y,
        'active_threshold': float(thr),
        'active_mask': active,
        'inactive_mask': inactive,
        'mean_y_active': float(mean_active) if np.isfinite(mean_active) else np.nan,
        'mean_y_inactive': float(mean_inactive) if np.isfinite(mean_inactive) else np.nan,
        'uplift_ratio': float(uplift_ratio) if np.isfinite(uplift_ratio) else np.nan,
        'uplift_pct': float(uplift_pct) if np.isfinite(uplift_pct) else np.nan,
        'spearman_rho': float(spearman_rho) if np.isfinite(spearman_rho) else np.nan,
        'spearman_p': float(spearman_p) if np.isfinite(spearman_p) else np.nan,
        'pearson_r': float(pearson_r) if np.isfinite(pearson_r) else np.nan,
        'pearson_p': float(pearson_p) if np.isfinite(pearson_p) else np.nan,
        'distance_corr': float(distance_corr) if np.isfinite(distance_corr) else np.nan,
        'association_map': assoc_map,
        'n_cells_total': int(x.size),
        'n_cells_active': int(np.count_nonzero(active)),
        'n_cells_inactive': int(np.count_nonzero(inactive)),
    }


def _plot_two_panel_association_figure(
    *,
    lon_edges: np.ndarray,
    lat_edges: np.ndarray,
    left_mat: np.ndarray,
    right_mat: np.ndarray,
    left_title: str,
    right_title: str,
    left_cmap: str,
    right_cmap: str,
    left_norm: Normalize,
    right_norm: Normalize,
    left_cbar_label: str,
    right_cbar_label: str,
    suptitle: str,
) -> tuple[plt.Figure, np.ndarray]:
    """通用双面板地理图绘制助手。"""
    crosses_dateline = bool(_REGION_CFG.get('crosses_dateline') and (lonmax < lonmin))
    central_lon = 180 if crosses_dateline else 0
    data_crs = ccrs.PlateCarree()
    map_crs = ccrs.PlateCarree(central_longitude=central_lon)

    fig, axes = plt.subplots(
        1,
        2,
        figsize=(20, 8),
        subplot_kw={'projection': map_crs},
        constrained_layout=True,
    )

    base_ocean = _BASEMAP_COLORS['ocean']
    base_land = _BASEMAP_COLORS['land']
    coast_color = _BASEMAP_COLORS['coastline']
    grid_color = _BASEMAP_COLORS['grid']

    lon_mesh, lat_mesh = np.meshgrid(lon_edges, lat_edges)
    panels = [
        (left_mat, left_title, left_cmap, left_norm),
        (right_mat, right_title, right_cmap, right_norm),
    ]
    mappables = []
    for ax, (mat, title, cmap, norm) in zip(axes, panels):
        ax.set_facecolor(base_ocean)
        ax.add_feature(cfeature.OCEAN, facecolor=base_ocean, zorder=0)
        ax.add_feature(cfeature.LAND, facecolor=base_land, edgecolor=coast_color, linewidth=0.5, zorder=0)
        ax.add_feature(cfeature.COASTLINE, linewidth=0.7, edgecolor=coast_color, zorder=1)
        gl = ax.gridlines(draw_labels=True, linewidth=0.4, color=grid_color, alpha=0.45, linestyle='--')
        gl.top_labels = False
        gl.right_labels = False
        hm = ax.pcolormesh(
            lon_mesh,
            lat_mesh,
            mat,
            cmap=cmap,
            norm=norm,
            shading='auto',
            transform=data_crs,
            zorder=2,
        )
        mappables.append(hm)
        ax.set_title(title, fontsize=12)
        ax.set_extent([float(lon_edges[0]), float(lon_edges[-1]), float(lat_edges[0]), float(lat_edges[-1])], crs=data_crs)

    cbar0 = fig.colorbar(mappables[0], ax=axes[0], orientation='horizontal', fraction=0.05, pad=0.08)
    cbar0.set_label(left_cbar_label, fontsize=11)
    cbar1 = fig.colorbar(mappables[1], ax=axes[1], orientation='horizontal', fraction=0.05, pad=0.08)
    cbar1.set_label(right_cbar_label, fontsize=11)

    fig.suptitle(suptitle, fontsize=14)
    return fig, axes


def _single_eddy_do_association(
    eddy_grid: np.ndarray,
    high_do_grid: np.ndarray,
    *,
    active_threshold: float = 1.0,
    do_rate_high_low_threshold: float | None = None,
    analysis_mask: np.ndarray | None = None,
    high_do_is_rate: bool = False,
) -> dict:
    """计算单一涡旋类型（ACE 或 CE）与 Argo 异常出现率的双向关联统计。"""

    def _distance_correlation_1d(xv: np.ndarray, yv: np.ndarray) -> float:
        if xv.size < 3 or yv.size < 3:
            return np.nan
        a = np.abs(xv[:, None] - xv[None, :])
        b = np.abs(yv[:, None] - yv[None, :])
        A = a - a.mean(axis=0, keepdims=True) - a.mean(axis=1, keepdims=True) + a.mean()
        B = b - b.mean(axis=0, keepdims=True) - b.mean(axis=1, keepdims=True) + b.mean()
        dcov2 = float(np.mean(A * B))
        dvarx2 = float(np.mean(A * A))
        dvary2 = float(np.mean(B * B))
        if dvarx2 <= 0 or dvary2 <= 0:
            return np.nan
        return float(np.sqrt(max(dcov2, 0.0) / np.sqrt(dvarx2 * dvary2)))

    x = np.asarray(eddy_grid, dtype=float).ravel()
    y = np.asarray(high_do_grid, dtype=float).ravel()
    finite_mask = np.isfinite(x) & np.isfinite(y)
    if analysis_mask is not None:
        am = np.asarray(analysis_mask, dtype=bool).ravel()
        if am.size == finite_mask.size:
            finite_mask = finite_mask & am
    x = x[finite_mask]
    y = y[finite_mask]

    active = x >= float(active_threshold)
    inactive = ~active

    mean_active = np.nanmean(y[active]) if np.any(active) else np.nan
    mean_inactive = np.nanmean(y[inactive]) if np.any(inactive) else np.nan

    uplift_ratio = np.nan
    uplift_pct = np.nan
    if np.isfinite(mean_inactive) and mean_inactive > 0 and np.isfinite(mean_active):
        uplift_ratio = mean_active / mean_inactive
        uplift_pct = (uplift_ratio - 1.0) * 100.0

    p_active = np.nanmean((y[active] > 0).astype(float)) if np.any(active) else np.nan
    p_inactive = np.nanmean((y[inactive] > 0).astype(float)) if np.any(inactive) else np.nan
    occurrence_ratio = np.nan
    occurrence_uplift_pct = np.nan
    if np.isfinite(p_inactive) and p_inactive > 0 and np.isfinite(p_active):
        occurrence_ratio = p_active / p_inactive
        occurrence_uplift_pct = (occurrence_ratio - 1.0) * 100.0

    # 反向指标：按异常出现率分组，比较两组涡旋天数均值
    if do_rate_high_low_threshold is None:
        positive_y = y[y > 0]
        do_rate_threshold = float(np.quantile(positive_y, 0.5)) if positive_y.size > 0 else np.nan
    else:
        do_rate_threshold = float(do_rate_high_low_threshold)
    high_do_rate_mask = (y >= do_rate_threshold) if np.isfinite(do_rate_threshold) else np.zeros_like(y, dtype=bool)
    low_do_rate_mask = ~high_do_rate_mask

    mean_eddy_high_do_rate = np.nanmean(x[high_do_rate_mask]) if np.any(high_do_rate_mask) else np.nan
    mean_eddy_low_do_rate = np.nanmean(x[low_do_rate_mask]) if np.any(low_do_rate_mask) else np.nan

    eddy_days_ratio_high_vs_low_do_rate = np.nan
    eddy_days_uplift_pct_high_vs_low_do_rate = np.nan
    if (
        np.isfinite(mean_eddy_low_do_rate)
        and mean_eddy_low_do_rate > 0
        and np.isfinite(mean_eddy_high_do_rate)
    ):
        eddy_days_ratio_high_vs_low_do_rate = mean_eddy_high_do_rate / mean_eddy_low_do_rate
        eddy_days_uplift_pct_high_vs_low_do_rate = (eddy_days_ratio_high_vs_low_do_rate - 1.0) * 100.0

    spearman_rho = np.nan
    spearman_p = np.nan
    pearson_r = np.nan
    pearson_p = np.nan
    distance_corr = np.nan
    nz_mask = (x > 0) | (y > 0)
    x_corr = x[nz_mask]
    y_corr = y[nz_mask]
    if x_corr.size >= 3 and np.nanstd(x_corr) > 0 and np.nanstd(y_corr) > 0:
        try:
            spearman_rho, spearman_p = spearmanr(x_corr, y_corr)
        except Exception:
            spearman_rho, spearman_p = np.nan, np.nan
        try:
            pearson_r, pearson_p = pearsonr(x_corr, y_corr)
        except Exception:
            pearson_r, pearson_p = np.nan, np.nan
        try:
            distance_corr = _distance_correlation_1d(x_corr, y_corr)
        except Exception:
            distance_corr = np.nan

    # 网格关联强度图：标准化后逐格乘积，正值代表共同高/共同低
    x2 = np.asarray(eddy_grid, dtype=float)
    y2 = np.asarray(high_do_grid, dtype=float)
    xlog = np.log1p(np.where(np.isfinite(x2), np.maximum(x2, 0.0), np.nan))
    if high_do_is_rate:
        ylog = np.where(np.isfinite(y2), np.maximum(y2, 0.0), np.nan)
    else:
        ylog = np.log1p(np.where(np.isfinite(y2), np.maximum(y2, 0.0), np.nan))
    valid_map = np.isfinite(xlog) & np.isfinite(ylog)
    if analysis_mask is not None:
        am2 = np.asarray(analysis_mask, dtype=bool)
        if am2.shape == valid_map.shape:
            valid_map = valid_map & am2
    assoc_map = np.full_like(xlog, np.nan, dtype=float)
    if np.any(valid_map):
        xmean = np.nanmean(xlog[valid_map])
        ymean = np.nanmean(ylog[valid_map])
        xstd = np.nanstd(xlog[valid_map])
        ystd = np.nanstd(ylog[valid_map])
        if xstd > 0 and ystd > 0:
            zx = (xlog - xmean) / xstd
            zy = (ylog - ymean) / ystd
            assoc_map[valid_map] = zx[valid_map] * zy[valid_map]

    return {
        'n_cells_total': int(x.size),
        'n_cells_active': int(np.count_nonzero(active)),
        'n_cells_inactive': int(np.count_nonzero(inactive)),
        'n_cells_eddy_active': int(np.count_nonzero(active)),
        'n_cells_eddy_inactive': int(np.count_nonzero(inactive)),
        'mean_high_do_active': float(mean_active) if np.isfinite(mean_active) else np.nan,
        'mean_high_do_inactive': float(mean_inactive) if np.isfinite(mean_inactive) else np.nan,
        'do_rate_mean_eddy_active': float(mean_active) if np.isfinite(mean_active) else np.nan,
        'do_rate_mean_eddy_inactive': float(mean_inactive) if np.isfinite(mean_inactive) else np.nan,
        'uplift_ratio_mean': float(uplift_ratio) if np.isfinite(uplift_ratio) else np.nan,
        'uplift_pct_mean': float(uplift_pct) if np.isfinite(uplift_pct) else np.nan,
        'do_rate_ratio_eddy_active_vs_inactive': float(uplift_ratio) if np.isfinite(uplift_ratio) else np.nan,
        'do_rate_uplift_pct_eddy_active_vs_inactive': float(uplift_pct) if np.isfinite(uplift_pct) else np.nan,
        'high_do_occurrence_active': float(p_active) if np.isfinite(p_active) else np.nan,
        'high_do_occurrence_inactive': float(p_inactive) if np.isfinite(p_inactive) else np.nan,
        'uplift_ratio_occurrence': float(occurrence_ratio) if np.isfinite(occurrence_ratio) else np.nan,
        'uplift_pct_occurrence': float(occurrence_uplift_pct) if np.isfinite(occurrence_uplift_pct) else np.nan,
        'do_rate_nonzero_ratio_eddy_active_vs_inactive': float(occurrence_ratio) if np.isfinite(occurrence_ratio) else np.nan,
        'do_rate_nonzero_uplift_pct_eddy_active_vs_inactive': float(occurrence_uplift_pct) if np.isfinite(occurrence_uplift_pct) else np.nan,
        'do_rate_threshold_for_high_group': float(do_rate_threshold) if np.isfinite(do_rate_threshold) else np.nan,
        'n_cells_do_rate_high': int(np.count_nonzero(high_do_rate_mask)),
        'n_cells_do_rate_low': int(np.count_nonzero(low_do_rate_mask)),
        'eddy_days_mean_do_rate_high': float(mean_eddy_high_do_rate) if np.isfinite(mean_eddy_high_do_rate) else np.nan,
        'eddy_days_mean_do_rate_low': float(mean_eddy_low_do_rate) if np.isfinite(mean_eddy_low_do_rate) else np.nan,
        'eddy_days_ratio_do_rate_high_vs_low': (
            float(eddy_days_ratio_high_vs_low_do_rate)
            if np.isfinite(eddy_days_ratio_high_vs_low_do_rate)
            else np.nan
        ),
        'eddy_days_uplift_pct_do_rate_high_vs_low': (
            float(eddy_days_uplift_pct_high_vs_low_do_rate)
            if np.isfinite(eddy_days_uplift_pct_high_vs_low_do_rate)
            else np.nan
        ),
        'spearman_rho': float(spearman_rho) if np.isfinite(spearman_rho) else np.nan,
        'spearman_p': float(spearman_p) if np.isfinite(spearman_p) else np.nan,
        'pearson_r': float(pearson_r) if np.isfinite(pearson_r) else np.nan,
        'pearson_p': float(pearson_p) if np.isfinite(pearson_p) else np.nan,
        'distance_corr': float(distance_corr) if np.isfinite(distance_corr) else np.nan,
        'association_map': assoc_map,
    }


def _scan_active_thresholds(
    ace_days: np.ndarray,
    ce_days: np.ndarray,
    high_do_grid: np.ndarray,
    analysis_mask: np.ndarray | None,
    *,
    min_group_cells: int = 20,
) -> tuple[float, pd.DataFrame]:
    """为 ACE/CE 共享扫描单一阈值并返回推荐值与诊断表。"""
    ace = np.asarray(ace_days, dtype=float)
    ce = np.asarray(ce_days, dtype=float)
    y = np.asarray(high_do_grid, dtype=float)
    mask = np.isfinite(ace) & np.isfinite(ce) & np.isfinite(y)
    if analysis_mask is not None:
        am = np.asarray(analysis_mask, dtype=bool)
        if am.shape == mask.shape:
            mask = mask & am

    ace_v = ace[mask]
    ce_v = ce[mask]
    y_v = y[mask]

    total = int(ace_v.size)
    if total == 0:
        diag = pd.DataFrame([
            {
                'threshold': 1.0,
                'ace_active': 0,
                'ace_inactive': 0,
                'ce_active': 0,
                'ce_inactive': 0,
                'ace_inactive_nonzero': 0,
                'ce_inactive_nonzero': 0,
                'balance': 0,
                'total_cells': 0,
            }
        ])
        return 1.0, diag

    pos = np.concatenate([ace_v[ace_v > 0], ce_v[ce_v > 0]])
    if pos.size == 0:
        diag = pd.DataFrame([
            {
                'threshold': 1.0,
                'ace_active': 0,
                'ace_inactive': total,
                'ce_active': 0,
                'ce_inactive': total,
                'ace_inactive_nonzero': int(np.count_nonzero(y_v > 0)),
                'ce_inactive_nonzero': int(np.count_nonzero(y_v > 0)),
                'balance': 0,
                'total_cells': total,
            }
        ])
        return 1.0, diag

    quantiles = [0.5, 0.6, 0.7, 0.8, 0.9, 0.95, 0.98]
    cand = sorted(set([1.0] + [float(np.quantile(pos, q)) for q in quantiles]))

    rows = []
    for thr in cand:
        ace_active_mask = ace_v >= thr
        ce_active_mask = ce_v >= thr

        ace_active = int(np.count_nonzero(ace_active_mask))
        ce_active = int(np.count_nonzero(ce_active_mask))
        ace_inactive = total - ace_active
        ce_inactive = total - ce_active

        ace_inactive_nonzero = int(np.count_nonzero(y_v[~ace_active_mask] > 0)) if ace_inactive > 0 else 0
        ce_inactive_nonzero = int(np.count_nonzero(y_v[~ce_active_mask] > 0)) if ce_inactive > 0 else 0

        balance = min(ace_active, ace_inactive, ce_active, ce_inactive)
        rows.append(
            {
                'threshold': float(thr),
                'ace_active': ace_active,
                'ace_inactive': ace_inactive,
                'ce_active': ce_active,
                'ce_inactive': ce_inactive,
                'ace_inactive_nonzero': ace_inactive_nonzero,
                'ce_inactive_nonzero': ce_inactive_nonzero,
                'balance': int(balance),
                'total_cells': total,
            }
        )

    diag = pd.DataFrame(rows)
    feasible = diag[
        (diag['ace_active'] >= min_group_cells)
        & (diag['ace_inactive'] >= min_group_cells)
        & (diag['ce_active'] >= min_group_cells)
        & (diag['ce_inactive'] >= min_group_cells)
        & (diag['ace_inactive_nonzero'] > 0)
        & (diag['ce_inactive_nonzero'] > 0)
    ]
    if not feasible.empty:
        rec = float(
            feasible.sort_values(
                ['balance', 'ace_inactive_nonzero', 'ce_inactive_nonzero', 'threshold'],
                ascending=[False, False, False, True],
            ).iloc[0]['threshold']
        )
    else:
        ok = diag[
            (diag['ace_active'] >= min_group_cells)
            & (diag['ace_inactive'] >= min_group_cells)
            & (diag['ce_active'] >= min_group_cells)
            & (diag['ce_inactive'] >= min_group_cells)
        ]
        if not ok.empty:
            rec = float(
                ok.sort_values(
                    ['balance', 'ace_inactive_nonzero', 'ce_inactive_nonzero', 'threshold'],
                    ascending=[False, False, False, True],
                ).iloc[0]['threshold']
            )
        else:
            rec = float(diag.sort_values(['balance', 'threshold'], ascending=[False, True]).iloc[0]['threshold'])

    return rec, diag


def analyze_euler_ace_ce_association(
    summary: dict,
    *,
    active_threshold: float | str | None = 'auto',
    do_rate_high_low_threshold: float | str | None = 'auto',
    min_group_cells: int = 20,
    show_fig: bool = True,
    save_fig: bool = False,
    output_prefix: str = 'EddyDays',
    output_dir: str | Path | None = None,
) -> dict:
    """在同一网格上分析并绘制 ACE/CE 与 Argo 异常出现率的关联性。

    参数:
        - summary (dict): build_euler_grid_summary 的输出。
        - active_threshold (float | str | None): 涡旋活跃区分组阈值，可传数值或 'auto'，默认 'auto'。
        - do_rate_high_low_threshold (float | str | None): 异常出现率高低分组阈值，可传数值或 'auto'，默认 'auto'。
        - min_group_cells (int): 自动阈值扫描时每组最小网格数，默认 20。
        - show_fig (bool): 是否显示图像，默认 True。
        - save_fig (bool): 是否保存图像，默认 False。
        - output_prefix (str): 输出文件名前缀，默认 'EddyDays'。
        - output_dir (str | Path | None): 输出目录；None 时使用默认路径。

    返回:
        - dict: 含关联指标、阈值扫描结果与图像保存路径等。

    说明:
        分析输出:

            - 关联地图（ACE-Anomaly、CE-Anomaly）。
            - 正向指标：涡旋活跃区 vs 非活跃区的异常出现率提升。
            - 反向指标：异常出现率高区 vs 低区的涡旋天数提升。
    """
    ace_days = np.asarray(summary['ace_days'], dtype=float)
    ce_days = np.asarray(summary['ce_days'], dtype=float)
    high_do = np.asarray(summary.get('high_do_occurrence_ratio', summary['high_do_profiles']), dtype=float)
    analysis_mask = np.asarray(summary.get('analysis_mask'), dtype=bool) if 'analysis_mask' in summary else None
    meta = summary.get('meta', {})
    grid = summary['grid']
    detection_label = meta.get('detection_label')
    if not detection_label:
        detection_label = f"ΔDO ≥ {meta.get('do_threshold', np.nan):g} μmol kg⁻¹"

    threshold_scan = None
    shared_threshold = float(active_threshold) if isinstance(active_threshold, (int, float)) else 1.0
    if active_threshold is None or str(active_threshold).lower() == 'auto':
        shared_threshold, threshold_scan = _scan_active_thresholds(
            ace_days,
            ce_days,
            high_do,
            analysis_mask,
            min_group_cells=min_group_cells,
        )

    shared_do_rate_threshold = None
    if not (do_rate_high_low_threshold is None or str(do_rate_high_low_threshold).lower() == 'auto'):
        shared_do_rate_threshold = float(do_rate_high_low_threshold)
    else:
        y_for_thr = np.asarray(high_do, dtype=float)
        if analysis_mask is not None and analysis_mask.shape == y_for_thr.shape:
            y_for_thr = y_for_thr[np.asarray(analysis_mask, dtype=bool)]
        y_for_thr = y_for_thr[np.isfinite(y_for_thr)]
        y_for_thr = y_for_thr[y_for_thr > 0]
        if y_for_thr.size > 0:
            shared_do_rate_threshold = float(np.quantile(y_for_thr, 0.5))

    ace_stats = _single_eddy_do_association(
        ace_days,
        high_do,
        active_threshold=shared_threshold,
        do_rate_high_low_threshold=shared_do_rate_threshold,
        analysis_mask=analysis_mask,
        high_do_is_rate=True,
    )
    ce_stats = _single_eddy_do_association(
        ce_days,
        high_do,
        active_threshold=shared_threshold,
        do_rate_high_low_threshold=shared_do_rate_threshold,
        analysis_mask=analysis_mask,
        high_do_is_rate=True,
    )

    lon_edges = np.asarray(grid['lon_edges'], dtype=float)
    lat_edges = np.asarray(grid['lat_edges'], dtype=float)
    lon_mesh, lat_mesh = np.meshgrid(lon_edges, lat_edges)

    ace_map = ace_stats['association_map']
    ce_map = ce_stats['association_map']

    vmax = np.nanmax(np.abs(np.concatenate([
        ace_map[np.isfinite(ace_map)],
        ce_map[np.isfinite(ce_map)],
    ]))) if (np.any(np.isfinite(ace_map)) or np.any(np.isfinite(ce_map))) else 1.0
    if not np.isfinite(vmax) or vmax <= 0:
        vmax = 1.0

    crosses_dateline = bool(grid.get('crosses_dateline', False))
    central_lon = 180 if crosses_dateline else 0
    data_crs = ccrs.PlateCarree()
    map_crs = ccrs.PlateCarree(central_longitude=central_lon)

    fig, axes = plt.subplots(
        1,
        2,
        figsize=(20, 8),
        subplot_kw={'projection': map_crs},
        constrained_layout=True,
    )

    base_ocean = _BASEMAP_COLORS['ocean']
    base_land = _BASEMAP_COLORS['land']
    coast_color = _BASEMAP_COLORS['coastline']
    grid_color = _BASEMAP_COLORS['grid']

    panel_defs = [
        ('ACE', ace_map, ace_stats),
        ('CE', ce_map, ce_stats),
    ]

    mappables = []
    for ax, (title, amap, st) in zip(axes, panel_defs):
        ax.set_facecolor(base_ocean)
        ax.add_feature(cfeature.OCEAN, facecolor=base_ocean, zorder=0)
        ax.add_feature(cfeature.LAND, facecolor=base_land, edgecolor=coast_color, linewidth=0.5, zorder=0)
        ax.add_feature(cfeature.COASTLINE, linewidth=0.7, edgecolor=coast_color, zorder=1)
        gl = ax.gridlines(draw_labels=True, linewidth=0.4, color=grid_color, alpha=0.45, linestyle='--')
        gl.top_labels = False
        gl.right_labels = False

        hm = ax.pcolormesh(
            lon_mesh,
            lat_mesh,
            amap,
            cmap='RdBu_r',
            norm=Normalize(vmin=-vmax, vmax=vmax),
            shading='auto',
            transform=data_crs,
            zorder=2,
        )
        mappables.append(hm)
        forward_uplift = st.get('do_rate_uplift_pct_eddy_active_vs_inactive', np.nan)
        reverse_uplift = st.get('eddy_days_uplift_pct_do_rate_high_vs_low', np.nan)
        rho = st.get('spearman_rho', np.nan)
        rho_p = st.get('spearman_p', np.nan)
        dcor = st.get('distance_corr', np.nan)
        rho_p_label = "p=nan"
        if np.isfinite(rho_p):
            rho_p_label = "p<0.0001" if rho_p < 1e-4 else f"p={rho_p:.4f}"
        ax.set_title(
            (
                f"{title}\n"
                f"Anomaly-rate uplift (eddy high vs low)={forward_uplift:.1f}%\n"
                f"Eddy-days uplift (anomaly-rate high vs low)={reverse_uplift:.1f}%\n"
                f"{rho_p_label} | rho={rho:.3f} | dCor={dcor:.3f}"
            ),
            fontsize=12,
        )
        ax.set_extent([float(lon_edges[0]), float(lon_edges[-1]), float(lat_edges[0]), float(lat_edges[-1])], crs=data_crs)

    cbar = fig.colorbar(mappables[0], ax=axes, orientation='horizontal', fraction=0.05, pad=0.08)
    cbar.set_label('Grid-wise Association Index (z(log1p(EddyDays))*z(AnomalyRate))', fontsize=11)

    fig.suptitle(
        (
            f"Eddy - Argo Anomaly Association ({meta.get('region_key', 'region')}, "
            f"{meta.get('start_year', '')}-{meta.get('end_year', '')})\n"
            f"{detection_label}"
            f" | depth ≥ {meta.get('anomaly_min_depth', np.nan):g} m"
            f" | eddy-days threshold ≥ {shared_threshold:g} day"
            f" | anomaly-rate threshold ≥ {shared_do_rate_threshold if shared_do_rate_threshold is not None else np.nan:.4f}"
        ),
        fontsize=14,
    )

    if output_dir is None:
        output_dir = _detection_output_dir_from_meta('plot_euler_grid_summary', meta)
    else:
        output_dir = Path(output_dir)

    prefix_safe = ''.join(ch if (ch.isalnum() or ch in {'_', '-'}) else '_' for ch in str(output_prefix).strip())
    if not prefix_safe:
        prefix_safe = 'EddyDays'

    def _compact(st: dict) -> dict:
        return {
            'n_cells_active': int(st.get('n_cells_eddy_active', st.get('n_cells_active', 0))),
            'n_cells_inactive': int(st.get('n_cells_eddy_inactive', st.get('n_cells_inactive', 0))),
            'uplift': {
                'do_rate_active_vs_inactive_ratio': float(st.get('do_rate_ratio_eddy_active_vs_inactive', np.nan)),
                'do_rate_active_vs_inactive_pct': float(st.get('do_rate_uplift_pct_eddy_active_vs_inactive', np.nan)),
                'eddy_days_do_rate_high_vs_low_ratio': float(st.get('eddy_days_ratio_do_rate_high_vs_low', np.nan)),
                'eddy_days_do_rate_high_vs_low_pct': float(st.get('eddy_days_uplift_pct_do_rate_high_vs_low', np.nan)),
            },
            'corr': {
                'spearman_rho': float(st.get('spearman_rho', np.nan)),
                'spearman_p': float(st.get('spearman_p', np.nan)),
                'pearson_r': float(st.get('pearson_r', np.nan)),
                'pearson_p': float(st.get('pearson_p', np.nan)),
                'distance_corr': float(st.get('distance_corr', np.nan)),
            },
        }

    ace_compact = _compact(ace_stats)
    ce_compact = _compact(ce_stats)
    fwd_mean = float(np.nanmean([
        ace_compact['uplift']['do_rate_active_vs_inactive_pct'],
        ce_compact['uplift']['do_rate_active_vs_inactive_pct'],
    ]))
    rev_mean = float(np.nanmean([
        ace_compact['uplift']['eddy_days_do_rate_high_vs_low_pct'],
        ce_compact['uplift']['eddy_days_do_rate_high_vs_low_pct'],
    ]))

    out = {
        'target': 'eddy_days',
        'thresholds': {
            'active': float(shared_threshold),
            'do_rate_high_low': float(shared_do_rate_threshold) if shared_do_rate_threshold is not None else np.nan,
        },
        'n_cells_analysis': int(np.count_nonzero(analysis_mask)) if analysis_mask is not None else int(np.size(high_do)),
        'uplift': {
            'do_rate_active_vs_inactive_pct_mean': fwd_mean,
            'eddy_days_do_rate_high_vs_low_pct_mean': rev_mean,
        },
        'corr': {
            'spearman_rho_mean': float(np.nanmean([
                ace_compact['corr']['spearman_rho'],
                ce_compact['corr']['spearman_rho'],
            ])),
            'pearson_r_mean': float(np.nanmean([
                ace_compact['corr']['pearson_r'],
                ce_compact['corr']['pearson_r'],
            ])),
            'distance_corr_mean': float(np.nanmean([
                ace_compact['corr']['distance_corr'],
                ce_compact['corr']['distance_corr'],
            ])),
        },
        'by_type': {
            'ace': ace_compact,
            'ce': ce_compact,
        },
    }

    if save_fig:
        output_dir.mkdir(parents=True, exist_ok=True)
        run_tag = str(meta.get('detection_file_stem') or f"do{_format_detection_value(meta.get('do_threshold', 'NA'))}")
        fname = output_dir / (
            f"{prefix_safe}_Association_{meta.get('start_year', 'NA')}_{meta.get('end_year', 'NA')}_"
            f"{meta.get('grid_step_deg', 1.0):g}deg_{run_tag}.png"
        )
        fig.savefig(fname, dpi=300, bbox_inches='tight')
        out['figure'] = str(fname)
        print(f"Association figure saved: {fname}")

    if show_fig:
        plt.show()
    plt.close(fig)
    return out


def analyze_euler_eke_do_association(
    summary: dict,
    eke_euler: dict | np.ndarray,
    *,
    active_threshold: float | str | None = 'auto',
    do_rate_high_low_threshold: float | str | None = 'auto',
    show_fig: bool = True,
    save_fig: bool = False,
    output_prefix: str = 'EKE',
    output_dir: str | Path | None = None,
) -> dict:
    """分析欧拉网格 EKE 与 Argo 异常出现率之间的关系并绘图。

    参数:
        - summary (dict): build_euler_grid_summary 的输出。
        - eke_euler (dict | np.ndarray): 欧拉网格 EKE（remap/build_euler_eke_grid 的输出或其网格数组）。
        - active_threshold (float | str | None): EKE active 分组阈值，可传数值或 'auto'，默认 'auto'。
        - do_rate_high_low_threshold (float | str | None): 异常出现率高低分组阈值，可传数值或 'auto'，默认 'auto'。
        - show_fig (bool): 是否显示图像，默认 True。
        - save_fig (bool): 是否保存图像，默认 False。
        - output_prefix (str): 输出文件名前缀，默认 'EKE'。
        - output_dir (str | Path | None): 输出目录；None 时使用默认路径。

    返回:
        - dict: 含双向 uplift 指标与图像保存路径等。

    说明:
        分析输出:

            - 两张图：EKE 与 Argo Anomaly Occurrence Rate 概览图、EKE-Anomaly Association 图。
            - 双向 uplift：正向 anomaly-rate uplift（EKE active vs inactive）、反向 EKE uplift（anomaly-rate high vs low）。
    """
    grid = summary['grid']
    meta = summary.get('meta', {})
    detection_label = meta.get('detection_label')
    if not detection_label:
        detection_label = f"ΔDO ≥ {meta.get('do_threshold', np.nan):g} μmol kg⁻¹"
    high_do = np.asarray(summary.get('high_do_occurrence_ratio', summary['high_do_profiles']), dtype=float)
    base_mask = np.asarray(summary.get('analysis_mask'), dtype=bool) if 'analysis_mask' in summary else np.ones_like(high_do, dtype=bool)

    if isinstance(eke_euler, dict):
        if 'eke_grid' in eke_euler:
            eke_grid = np.asarray(eke_euler['eke_grid'], dtype=float)
        elif 'euler' in eke_euler and isinstance(eke_euler['euler'], dict) and 'eke_grid' in eke_euler['euler']:
            eke_grid = np.asarray(eke_euler['euler']['eke_grid'], dtype=float)
        else:
            raise ValueError("eke_euler 字典中未找到 `eke_grid`")
    else:
        eke_grid = np.asarray(eke_euler, dtype=float)

    if eke_grid.shape != high_do.shape:
        raise ValueError("EKE 网格与 high_do 网格形状不一致")

    analysis_mask = base_mask & np.isfinite(eke_grid) & np.isfinite(high_do)
    core = _compute_grid_association_core(
        eke_grid,
        high_do,
        analysis_mask=analysis_mask,
        active_threshold=active_threshold,
        active_quantile=0.5,
        active_positive_only=True,
        x_log_for_map=False,
        y_log_for_map=False,
    )
    active_thr = core['active_threshold']
    mean_active = core['mean_y_active']
    mean_inactive = core['mean_y_inactive']
    uplift_pct = core['uplift_pct']
    uplift_ratio = core['uplift_ratio']
    spearman_rho = core['spearman_rho']
    spearman_p = core['spearman_p']
    pearson_r = core['pearson_r']
    pearson_p = core['pearson_p']
    distance_corr = core.get('distance_corr', np.nan)
    assoc_map = np.asarray(core['association_map'], dtype=float)

    # 反向指标：按异常出现率高低分组，比较两组 EKE 均值
    if do_rate_high_low_threshold is None or str(do_rate_high_low_threshold).lower() == 'auto':
        y_for_thr = np.asarray(high_do, dtype=float)
        y_for_thr = y_for_thr[analysis_mask]
        y_for_thr = y_for_thr[np.isfinite(y_for_thr)]
        y_for_thr = y_for_thr[y_for_thr > 0]
        do_rate_thr = float(np.quantile(y_for_thr, 0.5)) if y_for_thr.size > 0 else np.nan
    else:
        do_rate_thr = float(do_rate_high_low_threshold)

    if np.isfinite(do_rate_thr):
        do_high_mask = analysis_mask & (high_do >= do_rate_thr)
    else:
        do_high_mask = np.zeros_like(high_do, dtype=bool)
    do_low_mask = analysis_mask & (~do_high_mask)

    mean_eke_do_rate_high = np.nanmean(eke_grid[do_high_mask]) if np.any(do_high_mask) else np.nan
    mean_eke_do_rate_low = np.nanmean(eke_grid[do_low_mask]) if np.any(do_low_mask) else np.nan
    eke_uplift_ratio = np.nan
    eke_uplift_pct = np.nan
    if np.isfinite(mean_eke_do_rate_low) and mean_eke_do_rate_low > 0 and np.isfinite(mean_eke_do_rate_high):
        eke_uplift_ratio = mean_eke_do_rate_high / mean_eke_do_rate_low
        eke_uplift_pct = (eke_uplift_ratio - 1.0) * 100.0

    lon_edges = np.asarray(grid['lon_edges'], dtype=float)
    lat_edges = np.asarray(grid['lat_edges'], dtype=float)
    lon_mesh, lat_mesh = np.meshgrid(lon_edges, lat_edges)

    # 图1：EKE + Argo 异常出现率概览
    eke_vals = eke_grid[analysis_mask]
    eke_vals = eke_vals[np.isfinite(eke_vals)]
    eke_vmax = float(np.nanquantile(eke_vals, 0.99)) if eke_vals.size > 0 else np.nan
    if not np.isfinite(eke_vmax) or eke_vmax <= 0:
        eke_vmax = float(np.nanmax(eke_grid)) if np.any(np.isfinite(eke_grid)) else 1.0
    if not np.isfinite(eke_vmax) or eke_vmax <= 0:
        eke_vmax = 1.0

    do_vmax = float(np.nanmax(high_do)) if np.any(np.isfinite(high_do)) else np.nan
    if not np.isfinite(do_vmax) or do_vmax <= 0:
        do_vmax = 1.0

    suptitle_summary = (
        f"EKE & Argo Anomaly Summary ({meta.get('region_key', 'region')}, "
        f"{meta.get('start_year', '')}-{meta.get('end_year', '')})\n"
        f"{detection_label}"
        f" | depth ≥ {meta.get('anomaly_min_depth', np.nan):g} m"
        f" | EKE threshold ≥ {active_thr:.4e}"
        f" | anomaly-rate threshold ≥ {do_rate_thr if np.isfinite(do_rate_thr) else np.nan:.4f}"
    )
    fig_summary, _ = _plot_two_panel_association_figure(
        lon_edges=lon_edges,
        lat_edges=lat_edges,
        left_mat=eke_grid,
        right_mat=high_do,
        left_title='EKE (m² s⁻²)',
        right_title='Argo Anomaly Occurrence Rate',
        left_cmap='YlOrRd',
        right_cmap='Reds',
        left_norm=Normalize(vmin=0.0, vmax=eke_vmax),
        right_norm=Normalize(vmin=0.0, vmax=do_vmax),
        left_cbar_label='EKE (m² s⁻²)',
        right_cbar_label='Argo Anomaly Occurrence Rate per Grid Cell',
        suptitle=suptitle_summary,
    )

    # 图2：EKE-Anomaly 关联图
    vmax_assoc = np.nanmax(np.abs(assoc_map[np.isfinite(assoc_map)])) if np.any(np.isfinite(assoc_map)) else 1.0
    if not np.isfinite(vmax_assoc) or vmax_assoc <= 0:
        vmax_assoc = 1.0
    crosses_dateline = bool(grid.get('crosses_dateline', False))
    central_lon = 180 if crosses_dateline else 0
    data_crs = ccrs.PlateCarree()
    map_crs = ccrs.PlateCarree(central_longitude=central_lon)

    fig_assoc, ax = plt.subplots(
        1,
        1,
        figsize=(11, 8),
        subplot_kw={'projection': map_crs},
        constrained_layout=True,
    )
    base_ocean = _BASEMAP_COLORS['ocean']
    base_land = _BASEMAP_COLORS['land']
    coast_color = _BASEMAP_COLORS['coastline']
    grid_color = _BASEMAP_COLORS['grid']

    ax.set_facecolor(base_ocean)
    ax.add_feature(cfeature.OCEAN, facecolor=base_ocean, zorder=0)
    ax.add_feature(cfeature.LAND, facecolor=base_land, edgecolor=coast_color, linewidth=0.5, zorder=0)
    ax.add_feature(cfeature.COASTLINE, linewidth=0.7, edgecolor=coast_color, zorder=1)
    gl = ax.gridlines(draw_labels=True, linewidth=0.4, color=grid_color, alpha=0.45, linestyle='--')
    gl.top_labels = False
    gl.right_labels = False

    hm = ax.pcolormesh(
        lon_mesh,
        lat_mesh,
        assoc_map,
        cmap='RdBu_r',
        norm=Normalize(vmin=-vmax_assoc, vmax=vmax_assoc),
        shading='auto',
        transform=data_crs,
        zorder=2,
    )
    ax.set_title('EKE-Anomaly Association', fontsize=12)
    ax.set_extent([float(lon_edges[0]), float(lon_edges[-1]), float(lat_edges[0]), float(lat_edges[-1])], crs=data_crs)
    cbar_assoc = fig_assoc.colorbar(hm, ax=ax, orientation='horizontal', fraction=0.05, pad=0.08)
    cbar_assoc.set_label('z(EKE) * z(AnomalyRate)', fontsize=11)
    spearman_p_label = "p=nan"
    if np.isfinite(spearman_p):
        spearman_p_label = "p<0.0001" if spearman_p < 1e-4 else f"p={spearman_p:.4f}"
    fig_assoc.suptitle(
        (
            f"EKE - Argo Anomaly Association ({meta.get('region_key', 'region')}, "
            f"{meta.get('start_year', '')}-{meta.get('end_year', '')})\n"
            f"{detection_label}"
            f" | depth ≥ {meta.get('anomaly_min_depth', np.nan):g} m"
            f" | EKE threshold ≥ {active_thr:.4e}"
            f" | anomaly-rate threshold ≥ {do_rate_thr if np.isfinite(do_rate_thr) else np.nan:.4f}\n"
            f"Anomaly-rate uplift (EKE high vs low)={uplift_pct:.1f}%"
            f" | EKE uplift (anomaly-rate high vs low)={eke_uplift_pct:.1f}%\n"
            f"{spearman_p_label} | rho={spearman_rho:.3f} | dCor={distance_corr:.3f}"
        ),
        fontsize=14,
    )

    if output_dir is None:
        output_dir = _detection_output_dir_from_meta('plot_euler_grid_summary', meta)
    else:
        output_dir = Path(output_dir)

    prefix_safe = ''.join(ch if (ch.isalnum() or ch in {'_', '-'}) else '_' for ch in str(output_prefix).strip())
    if not prefix_safe:
        prefix_safe = 'EKE'

    out = {
        'target': 'eke',
        'thresholds': {
            'active': float(active_thr),
            'do_rate_high_low': float(do_rate_thr) if np.isfinite(do_rate_thr) else np.nan,
        },
        'n_cells_analysis': int(np.count_nonzero(analysis_mask)),
        'uplift': {
            'do_rate_active_vs_inactive_ratio': float(uplift_ratio) if np.isfinite(uplift_ratio) else np.nan,
            'do_rate_active_vs_inactive_pct': float(uplift_pct) if np.isfinite(uplift_pct) else np.nan,
            'eke_do_rate_high_vs_low_ratio': float(eke_uplift_ratio) if np.isfinite(eke_uplift_ratio) else np.nan,
            'eke_do_rate_high_vs_low_pct': float(eke_uplift_pct) if np.isfinite(eke_uplift_pct) else np.nan,
        },
        'means': {
            'do_rate_active': float(mean_active) if np.isfinite(mean_active) else np.nan,
            'do_rate_inactive': float(mean_inactive) if np.isfinite(mean_inactive) else np.nan,
            'eke_do_rate_high': float(mean_eke_do_rate_high) if np.isfinite(mean_eke_do_rate_high) else np.nan,
            'eke_do_rate_low': float(mean_eke_do_rate_low) if np.isfinite(mean_eke_do_rate_low) else np.nan,
        },
        'corr': {
            'spearman_rho': float(spearman_rho) if np.isfinite(spearman_rho) else np.nan,
            'spearman_p': float(spearman_p) if np.isfinite(spearman_p) else np.nan,
            'pearson_r': float(pearson_r) if np.isfinite(pearson_r) else np.nan,
            'pearson_p': float(pearson_p) if np.isfinite(pearson_p) else np.nan,
            'distance_corr': float(distance_corr) if np.isfinite(distance_corr) else np.nan,
        },
    }

    if save_fig:
        output_dir.mkdir(parents=True, exist_ok=True)
        run_tag = str(meta.get('detection_file_stem') or f"do{_format_detection_value(meta.get('do_threshold', 'NA'))}")
        fname_summary = output_dir / (
            f"{prefix_safe}_Summary_{meta.get('start_year', 'NA')}_{meta.get('end_year', 'NA')}_"
            f"{meta.get('grid_step_deg', 1.0):g}deg_{run_tag}.png"
        )
        fname_assoc = output_dir / (
            f"{prefix_safe}_Association_{meta.get('start_year', 'NA')}_{meta.get('end_year', 'NA')}_"
            f"{meta.get('grid_step_deg', 1.0):g}deg_{run_tag}.png"
        )
        fig_summary.savefig(fname_summary, dpi=300, bbox_inches='tight')
        fig_assoc.savefig(fname_assoc, dpi=300, bbox_inches='tight')
        out['figure_summary'] = str(fname_summary)
        out['figure_association'] = str(fname_assoc)
        out['figure'] = str(fname_assoc)
        print(f"EKE summary figure saved: {fname_summary}")
        print(f"EKE-Anomaly association figure saved: {fname_assoc}")

    if show_fig:
        plt.show()
    plt.close(fig_summary)
    plt.close(fig_assoc)
    return out


def run_euler_grid_analysis(
    start_year: int = 2002,
    end_year: int = 2022,
    *,
    analysis_targets: str | list[str] | tuple[str, ...] = ('eke',),
    run_association: bool = True,
    grid_step_deg: float = 1.0,
    detection_config: DetectionConfig | None = None,
    analysis_depth: float | int | None = None,
    active_threshold: float | str | None = 'auto',
    do_rate_high_low_threshold: float | str | None = 'auto',
    min_group_cells: int = 20,
    eke_file_path: str | Path | None = None,
    show_fig: bool = True,
    save_fig: bool = False,
    save_data: bool = False,
    output_prefix: str = 'EddyDays',
    eke_output_prefix: str = 'EKE',
    meta_output_root: str | Path | None = None,
    use_dask: bool = True,
    dask_scheduler: str = 'processes',
    dask_workers: int | None = None,
    dask_show_progress: bool = True,
) -> dict:
    """执行欧拉网格统计与关联分析流程。

    依次调用 build_euler_grid_summary 生成网格统计底图，当 targets 含 'eddy_days' 时绘制 ACE/CE 三联图，
    并按 analysis_targets 可选执行 eddy-days 与 EKE 关联分析。

    参数:
        - start_year (int): 统计起始年份（闭区间），默认 2002。
        - end_year (int): 统计结束年份（闭区间），默认 2022。
        - analysis_targets (str | list[str] | tuple[str, ...]): 分析内容开关，可选 'eddy_days'、'eke'，支持字符串或列表，默认 ('eke',)。
        - run_association (bool): 是否执行关联分析，默认 True。
        - grid_step_deg (float): 网格分辨率（°），默认 1.0。
        - detection_config (DetectionConfig | None): 异常识别配置；None 时使用 processing.yml 默认值。
        - analysis_depth (float | int | None): 统一分析深度（m），用于异常深度筛选与 EKE 采样深度；None 时回退配置默认。
        - active_threshold (float | str | None): 关联分析 active 分组阈值，可传数值或 'auto'，默认 'auto'。
        - do_rate_high_low_threshold (float | str | None): 关联分析异常出现率高低分组阈值，可传数值或 'auto'，默认 'auto'。
        - min_group_cells (int): 自动阈值扫描时每组最小网格数，默认 20。
        - eke_file_path (str | Path | None): 本地 EKE 原始分辨率 zarr 路径；None 时默认读取 GLORYS_processed/eke.zarr。
        - show_fig (bool): 是否显示图像，默认 True。
        - save_fig (bool): 是否保存图像，默认 False。
        - save_data (bool): 是否保存网格数据，默认 False。
        - output_prefix (str): eddy-days 输出文件名前缀，默认 'EddyDays'。
        - eke_output_prefix (str): EKE 关联图输出前缀，默认 'EKE'。
        - meta_output_root (str | Path | None): META 轨迹数据根目录（可选覆盖配置路径）。
        - use_dask (bool): 是否对按年 Argo 处理启用 Dask 并行，默认 True。
        - dask_scheduler (str): Dask 调度器（'processes'|'threads'|'synchronous'），默认 'processes'。
        - dask_workers (int | None): Dask worker 数；None 时采用 Dask 默认。
        - dask_show_progress (bool): 并行计算时是否显示 Dask 进度条，默认 True。

    返回:
        - dict: 含 'summary'（网格统计结果与元信息）、'outputs'（eddy-days 三联图/网格数据输出路径，未启用为空字典）、'association'（单目标时为该目标关联结果，多目标时为 {'eddy_days':..., 'eke':...} 平级字典）、'eke'（EKE 网格结果，未启用为 None）。
    """
    if isinstance(analysis_targets, str):
        targets = {analysis_targets.lower().strip()}
    else:
        targets = {str(t).lower().strip() for t in analysis_targets}
    if not targets:
        targets = {'eke'}

    unified_depth = float(analysis_depth) if analysis_depth is not None else None
    cfg = _resolve_detection_config(detection_config, anomaly_min_depth=unified_depth)

    summary = build_euler_grid_summary(
        start_year=start_year,
        end_year=end_year,
        grid_step_deg=grid_step_deg,
        detection_config=cfg,
        meta_output_root=meta_output_root,
        use_dask=use_dask,
        dask_scheduler=dask_scheduler,
        dask_workers=dask_workers,
        dask_show_progress=dask_show_progress,
    )
    outputs: dict = {}
    if 'eddy_days' in targets:
        outputs = plot_euler_grid_summary(
            summary,
            save_fig=save_fig,
            show_fig=show_fig,
            save_data=save_data,
            output_prefix=output_prefix,
        )

    if eke_file_path is None:
        eke_file_path = glorys_processed_root / 'eke.zarr'

    associations: dict[str, dict] = {}
    eke_bundle = None
    requested_assoc_targets = [k for k in ('eddy_days', 'eke') if k in targets]
    if run_association:
        if 'eddy_days' in targets:
            eddy_assoc = analyze_euler_ace_ce_association(
                summary,
                show_fig=show_fig,
                save_fig=save_fig,
                output_prefix=output_prefix,
                active_threshold=active_threshold,
                do_rate_high_low_threshold=do_rate_high_low_threshold,
                min_group_cells=min_group_cells,
            )
            associations['eddy_days'] = eddy_assoc

        if 'eke' in targets:
            eke_native = load_glorys_eke_native(file_path=eke_file_path)
            eke_euler = remap_glorys_eke_to_euler_grid(
                eke_native,
                grid_step_deg=grid_step_deg,
                method='linear',
            )
            eke_bundle = {
                'source_file': str(eke_file_path),
                'native': eke_native,
                'euler': eke_euler,
            }
            eke_assoc = analyze_euler_eke_do_association(
                summary,
                eke_euler,
                active_threshold=active_threshold,
                do_rate_high_low_threshold=do_rate_high_low_threshold,
                show_fig=show_fig,
                save_fig=save_fig,
                output_prefix=eke_output_prefix,
            )
            associations['eke'] = eke_assoc

    association_out = None
    if run_association:
        if len(requested_assoc_targets) == 1:
            association_out = associations.get(requested_assoc_targets[0])
        elif len(requested_assoc_targets) > 1:
            association_out = {
                k: associations[k]
                for k in requested_assoc_targets
                if k in associations
            }

    return {
        'summary': summary,
        'outputs': outputs,
        'association': association_out,
        'eke': eke_bundle,
    }

def calculate_delta_do(
    data: pd.DataFrame,
    detection_config: DetectionConfig | None = None,
    depth_col: str = 'Depth',
    do_col: str = 'DO',
    salinity_col: str = 'Salinity',
    temperature_col: str = 'Temperature',
    pi_col: str = 'PI',
    include_aou: bool = True,
    remove_outliers: bool = True,
    verbose: bool = False,
) -> pd.DataFrame:
    """
    计算 Argo 垂向剖面中的“潜在俯冲异常”并支持多种识别模式。

    参数:
        - data (pd.DataFrame): 含多个剖面数据的表；需含 Profile_number、深度、DO、盐度等列。
        - detection_config (DetectionConfig | None): 异常识别配置；None 时由 make_detection_config() 使用 processing.yml 默认值。
        - depth_col (str): 深度列名，默认 'Depth'。
        - do_col (str): 溶解氧列名，默认 'DO'。
        - salinity_col (str): 盐度列名，默认 'Salinity'。
        - temperature_col (str): 温度列名，默认 'Temperature'。
        - pi_col (str): aou 模式下优先使用的 π 列名；缺失时现场由 T/S/P 计算 surface-referenced potential spiciness，默认 'PI'。
        - include_aou (bool): 是否返回 AOU 相关结果（delta_aou），默认 True；do 模式中 delta_aou 是在 delta_do 对应深度上按同一参考线计算，并非“基于 AOU 自身阈值先定深度”得到。
        - remove_outliers (bool): 基础 QC 与规则过滤，默认 True。
        - verbose (bool): 是否打印进度信息，默认 False。

    返回:
        - pd.DataFrame: 每个满足条件的候选一行，始终含 Profile_number、depth、detection_method、primary_metric、primary_value、anomaly_score；do 模式额外含 delta_do/delta_salinity/delta_temperature/do_value 等，aou/trim 模式额外含 delta_aou/delta_pi/aou_value/pi_value/trim_* 等，并带 Year/Month/Day/Longitude/Latitude/Platform_number（若存在）；无满足记录时返回空表。

    说明:
        识别模式:

            - do（基线模式）：先找 DO 正峰（导数由正到负），以峰深度为中心在 [p-Δp, p+Δp] 内用两端点连线构造参考值，计算同深度 delta_do/delta_salinity/delta_temperature，以 delta_do ≥ do_threshold 为主判据、盐温阈值（若 >0）为附加过滤。
            - aou（论文式 AOU+π 模式）：找 AOU 负峰并在深度容差内配对最近 π 峰，在上下窗口取极值点连线计算 delta_aou 与 delta_pi，以 delta_aou ≤ aou_threshold 且 |delta_pi| ≥ pi_threshold 为判据。
            - trim（分箱 + trimmed mean 模式）：对 AOU 与绝对盐度分箱降采样，用 rolling trimmed mean 构造鲁棒残差并标准化筛选候选，再做局地梯度符号变化检查，最后回原剖面输出 delta_* 指标。
    """

    def _append_profile_metadata(result: dict, profile_df: pd.DataFrame) -> dict:
        for c in ('Year', 'Month', 'Day', 'Longitude', 'Latitude', 'Platform_number'):
            if c in profile_df.columns:
                result[c] = profile_df[c].iloc[0]
        return result

    def _compute_profile_pi_surface(profile_rows: pd.DataFrame) -> np.ndarray:
        if pi_col in profile_rows.columns:
            return pd.to_numeric(profile_rows[pi_col], errors='coerce').to_numpy(dtype=float)

        needed = [salinity_col, temperature_col, depth_col]
        if any(c not in profile_rows.columns for c in needed):
            return np.full(len(profile_rows), np.nan, dtype=float)

        sal_vals = pd.to_numeric(profile_rows[salinity_col], errors='coerce').to_numpy(dtype=float)
        temp_vals = pd.to_numeric(profile_rows[temperature_col], errors='coerce').to_numpy(dtype=float)
        pres_vals = pd.to_numeric(profile_rows[depth_col], errors='coerce').to_numpy(dtype=float)
        lon_vals = (
            pd.to_numeric(profile_rows['Longitude'], errors='coerce').to_numpy(dtype=float)
            if 'Longitude' in profile_rows.columns else np.zeros_like(sal_vals)
        )
        lat_vals = (
            pd.to_numeric(profile_rows['Latitude'], errors='coerce').to_numpy(dtype=float)
            if 'Latitude' in profile_rows.columns else np.zeros_like(sal_vals)
        )

        valid = np.isfinite(sal_vals) & np.isfinite(temp_vals) & np.isfinite(pres_vals)
        out = np.full(len(profile_rows), np.nan, dtype=float)
        if not valid.any():
            return out

        try:
            sa = gsw.SA_from_SP(sal_vals[valid], pres_vals[valid], lon_vals[valid], lat_vals[valid])
            ct = gsw.CT_from_t(sa, temp_vals[valid], pres_vals[valid])
            if not hasattr(gsw, 'spiciness0'):
                return out
            out[valid] = np.asarray(gsw.spiciness0(sa, ct), dtype=float)
        except Exception:
            return out
        return out

    def _compute_profile_abs_sal(profile_rows: pd.DataFrame) -> np.ndarray:
        needed = [salinity_col, depth_col]
        if any(c not in profile_rows.columns for c in needed):
            return np.full(len(profile_rows), np.nan, dtype=float)

        sal_vals = pd.to_numeric(profile_rows[salinity_col], errors='coerce').to_numpy(dtype=float)
        pres_vals = pd.to_numeric(profile_rows[depth_col], errors='coerce').to_numpy(dtype=float)
        lon_vals = (
            pd.to_numeric(profile_rows['Longitude'], errors='coerce').to_numpy(dtype=float)
            if 'Longitude' in profile_rows.columns else np.zeros_like(sal_vals)
        )
        lat_vals = (
            pd.to_numeric(profile_rows['Latitude'], errors='coerce').to_numpy(dtype=float)
            if 'Latitude' in profile_rows.columns else np.zeros_like(sal_vals)
        )

        valid = np.isfinite(sal_vals) & np.isfinite(pres_vals)
        out = np.full(len(profile_rows), np.nan, dtype=float)
        if not valid.any():
            return out

        try:
            out[valid] = np.asarray(
                gsw.SA_from_SP(sal_vals[valid], pres_vals[valid], lon_vals[valid], lat_vals[valid]),
                dtype=float,
            )
        except Exception:
            return out
        return out

    def _best_qc_pick(group: pd.DataFrame) -> pd.Series:
        qccol = f"{do_col}_Flag"
        priority = {1: 0, 2: 1, 5: 2, 8: 3}

        def rank(v):
            try:
                iv = int(v)
            except Exception:
                return 999
            return priority.get(iv, 999)

        if qccol in group.columns:
            ranks = group[qccol].apply(rank)
            min_rank = ranks.min()
            picked = group.loc[ranks[ranks == min_rank].index]
            if do_col in picked.columns:
                picked = picked[pd.notna(picked[do_col])]
                if picked.empty:
                    picked = group.loc[ranks[ranks == min_rank].index]
            return picked.iloc[0]
        return group.iloc[0]

    def _mean_pick(group: pd.DataFrame) -> pd.Series:
        first = group.iloc[0].copy()
        for c in [do_col, salinity_col, temperature_col]:
            if c in group.columns:
                first[c] = pd.to_numeric(group[c], errors='coerce').mean()
        return first

    def _dedupe_profile_depth(profile_df: pd.DataFrame) -> pd.DataFrame:
        df = profile_df.copy()

        if df[depth_col].duplicated().any():
            strategy = (duplicate_depth_strategy or 'best_qc').lower()
            if strategy not in {'best_qc', 'first', 'mean', 'max', 'min'}:
                strategy = 'best_qc'

            grouped = list(df.groupby(depth_col, sort=False))
            picked_rows = []
            for _, grp in grouped:
                if len(grp) == 1:
                    picked_rows.append(grp.iloc[0])
                    continue
                if strategy == 'best_qc':
                    picked_rows.append(_best_qc_pick(grp))
                elif strategy == 'first':
                    picked_rows.append(grp.iloc[0])
                elif strategy == 'mean':
                    picked_rows.append(_mean_pick(grp))
                elif strategy == 'max':
                    picked_rows.append(grp.loc[pd.to_numeric(grp[do_col], errors='coerce').idxmax()])
                elif strategy == 'min':
                    picked_rows.append(grp.loc[pd.to_numeric(grp[do_col], errors='coerce').idxmin()])

            df = pd.DataFrame(picked_rows)

        df = df.sort_values(by=depth_col).reset_index(drop=True)
        depth_arr = pd.to_numeric(df[depth_col], errors='coerce').to_numpy(dtype=float)
        if len(depth_arr) == 0:
            return df

        if np.any(np.diff(depth_arr) <= 0):
            keep_idx = [0]
            last_depth = depth_arr[0]
            for ridx in range(1, len(depth_arr)):
                dval = depth_arr[ridx]
                if np.isfinite(dval) and dval > last_depth:
                    keep_idx.append(ridx)
                    last_depth = dval
            df = df.iloc[keep_idx].reset_index(drop=True)
        return df

    def _find_peaks_by_slope(values: np.ndarray, depth_values: np.ndarray) -> list[tuple[int, str, float]]:
        if len(values) < 3:
            return []
        slopes = np.gradient(values, depth_values)
        peaks: list[tuple[int, str, float]] = []
        for i in range(1, len(slopes) - 1):
            prev_s = slopes[i - 1]
            next_s = slopes[i + 1]
            if not (np.isfinite(prev_s) and np.isfinite(next_s)):
                continue
            if prev_s > 0 and next_s < 0:
                peaks.append((i, 'positive', float(depth_values[i])))
            elif prev_s < 0 and next_s > 0:
                peaks.append((i, 'negative', float(depth_values[i])))
        return peaks

    def _endpoint_delta(
        depth_values: np.ndarray,
        values: np.ndarray,
        target_depth: float,
        half_window: float,
    ) -> tuple[float, float, float, float, float]:
        d_lower = max(0.0, float(target_depth) - float(half_window))
        d_upper = float(target_depth) + float(half_window)

        mask = (
            np.isfinite(depth_values)
            & np.isfinite(values)
            & (depth_values >= d_lower)
            & (depth_values <= d_upper)
        )
        if np.count_nonzero(mask) < 2:
            return np.nan, np.nan, np.nan, np.nan, np.nan

        d = depth_values[mask]
        v = values[mask]
        order = np.argsort(d)
        d = d[order]
        v = v[order]
        if len(d) < 2 or np.isclose(d[0], d[-1]):
            return np.nan, np.nan, np.nan, np.nan, np.nan

        obs = float(np.interp(target_depth, d, v))
        ref = float(np.interp(target_depth, [d[0], d[-1]], [v[0], v[-1]]))
        return obs - ref, obs, ref, float(d[0]), float(d[-1])

    def _window_extrema_delta(
        depth_values: np.ndarray,
        values: np.ndarray,
        target_depth: float,
        half_window: float,
        peak_type: str,
    ) -> tuple[float, float, float, float, float]:
        up_mask = (
            np.isfinite(depth_values)
            & np.isfinite(values)
            & (depth_values >= (target_depth - half_window))
            & (depth_values <= target_depth)
        )
        lo_mask = (
            np.isfinite(depth_values)
            & np.isfinite(values)
            & (depth_values >= target_depth)
            & (depth_values <= (target_depth + half_window))
        )

        if np.count_nonzero(up_mask) < 1 or np.count_nonzero(lo_mask) < 1:
            return np.nan, np.nan, np.nan, np.nan, np.nan

        up_d = depth_values[up_mask]
        up_v = values[up_mask]
        lo_d = depth_values[lo_mask]
        lo_v = values[lo_mask]

        try:
            if peak_type == 'negative':
                up_idx = int(np.nanargmax(up_v))
                lo_idx = int(np.nanargmax(lo_v))
            else:
                up_idx = int(np.nanargmin(up_v))
                lo_idx = int(np.nanargmin(lo_v))
        except ValueError:
            return np.nan, np.nan, np.nan, np.nan, np.nan

        d1 = float(up_d[up_idx])
        v1 = float(up_v[up_idx])
        d2 = float(lo_d[lo_idx])
        v2 = float(lo_v[lo_idx])
        if np.isclose(d1, d2):
            return np.nan, np.nan, np.nan, np.nan, np.nan

        if d1 > d2:
            d1, d2 = d2, d1
            v1, v2 = v2, v1

        valid_all = np.isfinite(depth_values) & np.isfinite(values)
        if np.count_nonzero(valid_all) < 2:
            return np.nan, np.nan, np.nan, np.nan, np.nan

        d_all = depth_values[valid_all]
        v_all = values[valid_all]
        order = np.argsort(d_all)
        d_all = d_all[order]
        v_all = v_all[order]

        obs = float(np.interp(target_depth, d_all, v_all))
        ref = float(np.interp(target_depth, [d1, d2], [v1, v2]))
        return obs - ref, obs, ref, d1, d2

    def _trimmed_mean_20_80(x: np.ndarray) -> float:
        arr = np.asarray(x, dtype=float)
        arr = arr[np.isfinite(arr)]
        if arr.size == 0:
            return np.nan
        q20 = np.nanquantile(arr, 0.2)
        q80 = np.nanquantile(arr, 0.8)
        subset = arr[(arr >= q20) & (arr <= q80)]
        if subset.size == 0:
            return np.nan
        return float(np.nanmean(subset))

    def _downscale_profile(
        profile_df: pd.DataFrame,
        var_cols: list[str],
        bin_width: float,
    ) -> pd.DataFrame:
        if bin_width <= 0:
            return pd.DataFrame(columns=[depth_col, *var_cols])

        work = profile_df[[depth_col, *var_cols]].copy()
        work[depth_col] = pd.to_numeric(work[depth_col], errors='coerce')
        for c in var_cols:
            work[c] = pd.to_numeric(work[c], errors='coerce')

        work = work[np.isfinite(work[depth_col])].copy()
        if work.empty:
            return pd.DataFrame(columns=[depth_col, *var_cols])

        dmin = float(work[depth_col].min())
        dmax = float(work[depth_col].max())
        if not np.isfinite(dmin) or not np.isfinite(dmax):
            return pd.DataFrame(columns=[depth_col, *var_cols])

        left = np.floor(dmin / bin_width) * bin_width
        right = np.ceil(dmax / bin_width) * bin_width
        if right <= left:
            right = left + bin_width
        bins = np.arange(left, right + bin_width, bin_width)
        if bins.size < 2:
            bins = np.array([left, left + bin_width], dtype=float)

        work['_bin'] = pd.cut(work[depth_col], bins=bins, include_lowest=True, labels=False)
        work = work.dropna(subset=['_bin'])
        if work.empty:
            return pd.DataFrame(columns=[depth_col, *var_cols])

        work['_bin'] = work['_bin'].astype(int)
        grouped = work.groupby('_bin', as_index=False)[var_cols].mean(numeric_only=True)
        centers = ((bins[grouped['_bin'].to_numpy(dtype=int)] + bins[grouped['_bin'].to_numpy(dtype=int) + 1]) / 2.0)
        grouped[depth_col] = centers
        grouped = grouped[[depth_col, *var_cols]].sort_values(depth_col).reset_index(drop=True)
        return grouped

    def _scaled_residual(values: np.ndarray) -> np.ndarray:
        s = pd.Series(values, dtype=float)
        tm9 = s.rolling(window=9, center=True, min_periods=9).apply(_trimmed_mean_20_80, raw=True)
        rob_res_raw = s - tm9

        arr = rob_res_raw.to_numpy(dtype=float)
        finite = np.isfinite(arr)
        out = np.full(arr.shape, np.nan, dtype=float)
        if np.count_nonzero(finite) < 3:
            return out

        p75 = np.nanpercentile(arr[finite], 75)
        p25 = np.nanpercentile(arr[finite], 25)
        iqrn = (p75 - p25) / 1.349
        if not np.isfinite(iqrn) or iqrn <= 0:
            return out

        nz = arr[finite & (arr != 0)]
        median_res = float(np.nanmedian(nz)) if nz.size else 0.0

        zero_mask = np.isfinite(arr) & (arr == 0)
        nz_mask = np.isfinite(arr) & (arr != 0)
        out[zero_mask] = 0.0
        out[nz_mask] = (arr[nz_mask] - median_res) / iqrn
        return out

    def _gradient_sign_change(
        depth_values: np.ndarray,
        values: np.ndarray,
        target_depth: float,
        window: float,
    ) -> bool:
        mask = (
            np.isfinite(depth_values)
            & np.isfinite(values)
            & (depth_values >= (target_depth - window))
            & (depth_values <= (target_depth + window))
        )
        if np.count_nonzero(mask) < 3:
            return False

        d = depth_values[mask]
        v = values[mask]
        order = np.argsort(d)
        d = d[order]
        v = v[order]
        if len(d) < 3:
            return False

        keep = np.r_[True, np.diff(d) > 0]
        d = d[keep]
        v = v[keep]
        if len(d) < 3:
            return False

        dv = np.diff(v) / np.diff(d)
        dv = dv[np.isfinite(dv)]
        if dv.size < 2:
            return False

        signs = np.sign(dv)
        return bool(np.any(np.diff(signs) != 0))
    
    # 检查必要的列是否存在
    required_cols = [depth_col, do_col, salinity_col, temperature_col, 'Profile_number']
    missing_cols = [col for col in required_cols if col not in data.columns]
    if missing_cols:
        print(f"警告：缺少必要的列: {missing_cols}")
        return pd.DataFrame()

    # 复制数据以避免修改原始数据
    input_data = data.copy()

    # 存储所有剖面的结果
    all_results = []

    # 按Profile_number分组处理
    profile_groups = input_data.groupby('Profile_number')
    total_profiles = len(profile_groups)
    if verbose:
        print(f"开始处理 {total_profiles} 个剖面...")

    processed_profiles = 0

    # 参数集中来自 DetectionConfig，避免下游函数签名按方法膨胀。
    cfg = _resolve_detection_config(detection_config)
    depth_interval = cfg.depth_interval
    do_threshold = cfg.do_threshold
    salinity_threshold = cfg.salinity_threshold
    temperature_threshold = cfg.temperature_threshold
    depth_merge_tolerance = cfg.depth_merge_tolerance
    duplicate_depth_strategy = cfg.duplicate_depth_strategy
    anomaly_min_depth = cfg.anomaly_min_depth
    anomaly_max_depth = cfg.anomaly_max_depth
    do_near_zero_threshold = cfg.do_near_zero_threshold
    do_near_zero_max_count = cfg.do_near_zero_max_count
    if do_near_zero_max_count is not None:
        do_near_zero_max_count = int(do_near_zero_max_count)

    method_norm = cfg.method
    aou_threshold = cfg.aou_threshold
    pi_threshold = cfg.pi_threshold
    aou_pi_depth_tolerance = cfg.aou_pi_depth_tolerance
    trim_cutoff = cfg.trim_cutoff
    trim_window = cfg.trim_window
    trim_bin_width_outlier = cfg.trim_bin_width_outlier
    trim_bin_width_check = cfg.trim_bin_width_check
    trim_depth_min = cfg.trim_depth_min
    trim_depth_max = cfg.trim_depth_max

    for profile_num, profile_data in profile_groups:
        profile_data = profile_data.copy()

        # 先统一为数值，避免字符串导致比较失败。
        for c in [depth_col, do_col, salinity_col, temperature_col]:
            if c in profile_data.columns:
                profile_data[c] = pd.to_numeric(profile_data[c], errors='coerce')

        # 质量控制：移除异常值和质量标记不良的数据
        if remove_outliers:
            # 应用Argo QC标准：仅保留等级为{1,2,5,8}的观测
            for var in [do_col, salinity_col, temperature_col]:
                qc_column_name = f"{var}_Flag"
                if qc_column_name in profile_data.columns:
                    good_qc_flags = ['1', '2', '5', '8', 1, 2, 5, 8]
                    bad_qc_mask = ~profile_data[qc_column_name].isin(good_qc_flags)
                    profile_data.loc[bad_qc_mask, var] = np.nan

            # 规则法：移除已知的错误值
            if do_col in profile_data.columns:
                do_numeric = pd.to_numeric(profile_data[do_col], errors='coerce')
                bad_do_mask = do_numeric <= do_near_zero_threshold
                if do_near_zero_max_count is not None and do_near_zero_max_count >= 0:
                    bad_do_count = int(np.count_nonzero(bad_do_mask.to_numpy()))
                    if bad_do_count > do_near_zero_max_count:
                        continue
                profile_data.loc[bad_do_mask, do_col] = np.nan

        # 移除包含NaN值的行
        drop_subset = [depth_col, do_col, salinity_col, temperature_col]
        profile_data_clean = profile_data.dropna(subset=drop_subset)
        if len(profile_data_clean) < 5:
            continue

        profile_data_clean = _dedupe_profile_depth(profile_data_clean)
        if len(profile_data_clean) < 5:
            continue

        depth_values = pd.to_numeric(profile_data_clean[depth_col], errors='coerce').to_numpy(dtype=float)
        do_values = pd.to_numeric(profile_data_clean[do_col], errors='coerce').to_numpy(dtype=float)
        salinity_values = pd.to_numeric(profile_data_clean[salinity_col], errors='coerce').to_numpy(dtype=float)
        temperature_values = pd.to_numeric(profile_data_clean[temperature_col], errors='coerce').to_numpy(dtype=float)

        if len(depth_values) < 5 or np.any(np.diff(depth_values) <= 0):
            continue

        need_aou = include_aou or (method_norm in {'aou', 'trim'})
        aou_values = None
        if need_aou:
            try:
                aou_series, _ = _compute_aou_for_plot(profile_data_clean.copy(), remove_outliers=remove_outliers)
                aou_values = pd.to_numeric(aou_series, errors='coerce').to_numpy(dtype=float)
            except Exception:
                aou_values = None

        if method_norm in {'aou', 'trim'}:
            if aou_values is None or np.count_nonzero(np.isfinite(aou_values)) < 5:
                continue

        profile_results = []

        if method_norm == 'do':
            # 重要说明：坐标系差异
            # 图像坐标系（海洋学习惯）：深度向下为正，图像上正斜率表示随深度增加变量增大。
            # 数据计算坐标系：np.gradient 计算 dVar/dDepth；图像上的“正斜率”在数值上对应负斜率。
            # 因此峰值条件看起来会和图像描述方向相反，但逻辑上是一致的。
            do_peaks = [pk for pk in _find_peaks_by_slope(do_values, depth_values) if pk[1] == 'positive']
            if not do_peaks:
                continue

            use_salinity_filter = (salinity_threshold is not None and salinity_threshold > 0)
            use_temperature_filter = (temperature_threshold is not None and temperature_threshold > 0)

            for do_idx, _, target_depth in do_peaks:
                if anomaly_max_depth is not None and anomaly_max_depth > 0 and target_depth > anomaly_max_depth:
                    continue

                delta_do, do_obs, _, _, _ = _endpoint_delta(depth_values, do_values, target_depth, depth_interval)
                if not np.isfinite(delta_do):
                    continue

                delta_salinity, sal_obs, _, _, _ = _endpoint_delta(
                    depth_values, salinity_values, target_depth, depth_interval
                )
                delta_temperature, temp_obs, _, _, _ = _endpoint_delta(
                    depth_values, temperature_values, target_depth, depth_interval
                )

                delta_aou = np.nan
                if include_aou and aou_values is not None:
                    delta_aou, _, _, _, _ = _endpoint_delta(depth_values, aou_values, target_depth, depth_interval)

                cond = (delta_do >= do_threshold)
                if use_salinity_filter:
                    cond = cond and np.isfinite(delta_salinity) and (abs(delta_salinity) >= salinity_threshold)
                if use_temperature_filter:
                    cond = cond and np.isfinite(delta_temperature) and (abs(delta_temperature) >= temperature_threshold)

                if not cond:
                    continue

                result = {
                    'Profile_number': profile_num,
                    'depth': float(target_depth),
                    'delta_do': float(delta_do),
                    'delta_salinity': float(delta_salinity),
                    'delta_temperature': float(delta_temperature),
                    'do_value': float(do_obs),
                    'salinity_value': float(sal_obs),
                    'temperature_value': float(temp_obs),
                    'detection_method': method_norm,
                    'primary_metric': 'delta_do',
                    'primary_value': float(delta_do),
                    'anomaly_score': float(delta_do),
                    '_rank_score': float(delta_do),
                }
                if include_aou:
                    result['delta_aou'] = float(delta_aou)
                profile_results.append(_append_profile_metadata(result, profile_data_clean))

        elif method_norm == 'aou':
            pi_values = _compute_profile_pi_surface(profile_data_clean)
            if np.count_nonzero(np.isfinite(pi_values)) < 5:
                continue

            aou_neg_peaks = [pk for pk in _find_peaks_by_slope(aou_values, depth_values) if pk[1] == 'negative']
            pi_peaks = _find_peaks_by_slope(pi_values, depth_values)
            if not aou_neg_peaks or not pi_peaks:
                continue

            for aou_idx, _, target_depth in aou_neg_peaks:
                if anomaly_max_depth is not None and anomaly_max_depth > 0 and target_depth > anomaly_max_depth:
                    continue

                near_pi = [pk for pk in pi_peaks if abs(pk[2] - target_depth) <= aou_pi_depth_tolerance]
                if not near_pi:
                    continue
                near_pi.sort(key=lambda x: abs(x[2] - target_depth))
                pi_idx, pi_peak_type, pi_peak_depth = near_pi[0]

                delta_aou, aou_obs, _, _, _ = _window_extrema_delta(
                    depth_values,
                    aou_values,
                    target_depth,
                    depth_interval,
                    peak_type='negative',
                )
                if not np.isfinite(delta_aou):
                    continue

                delta_pi, pi_obs, _, _, _ = _window_extrema_delta(
                    depth_values,
                    pi_values,
                    target_depth,
                    depth_interval,
                    peak_type=pi_peak_type,
                )
                if not np.isfinite(delta_pi):
                    continue

                if not (delta_aou <= aou_threshold and abs(delta_pi) >= pi_threshold):
                    continue

                delta_do, do_obs, _, _, _ = _endpoint_delta(depth_values, do_values, target_depth, depth_interval)
                delta_salinity, sal_obs, _, _, _ = _endpoint_delta(
                    depth_values, salinity_values, target_depth, depth_interval
                )
                delta_temperature, temp_obs, _, _, _ = _endpoint_delta(
                    depth_values, temperature_values, target_depth, depth_interval
                )

                result = {
                    'Profile_number': profile_num,
                    'depth': float(target_depth),
                    'delta_do': float(delta_do),
                    'delta_aou': float(delta_aou),
                    'delta_pi': float(delta_pi),
                    'delta_salinity': float(delta_salinity),
                    'delta_temperature': float(delta_temperature),
                    'do_value': float(do_obs),
                    'aou_value': float(aou_obs),
                    'pi_value': float(pi_obs),
                    'salinity_value': float(sal_obs),
                    'temperature_value': float(temp_obs),
                    'pi_peak_type': pi_peak_type,
                    'pi_peak_depth': float(pi_peak_depth),
                    'aou_peak_depth': float(target_depth),
                    'peak_depth_offset': float(pi_peak_depth - target_depth),
                    'detection_method': method_norm,
                    'primary_metric': 'delta_aou',
                    'primary_value': float(delta_aou),
                    'anomaly_score': float(-delta_aou),
                    '_rank_score': float(-delta_aou),
                }
                profile_results.append(_append_profile_metadata(result, profile_data_clean))

        elif method_norm == 'trim':
            abs_sal_values = _compute_profile_abs_sal(profile_data_clean)
            if np.count_nonzero(np.isfinite(abs_sal_values)) < 5:
                continue

            base_profile = pd.DataFrame(
                {
                    depth_col: depth_values,
                    'AOU_WORK': aou_values,
                    'ABS_SAL_WORK': abs_sal_values,
                }
            )

            down40 = _downscale_profile(base_profile, ['AOU_WORK', 'ABS_SAL_WORK'], trim_bin_width_outlier)
            down20 = _downscale_profile(base_profile, ['AOU_WORK', 'ABS_SAL_WORK'], trim_bin_width_check)
            if down40.empty or down20.empty:
                continue

            down40['SCALE_RES_ROB_AOU'] = _scaled_residual(
                pd.to_numeric(down40['AOU_WORK'], errors='coerce').to_numpy(dtype=float)
            )
            down40['SCALE_RES_ROB_ABS_SAL'] = _scaled_residual(
                pd.to_numeric(down40['ABS_SAL_WORK'], errors='coerce').to_numpy(dtype=float)
            )

            candidates = down40[
                np.isfinite(pd.to_numeric(down40['SCALE_RES_ROB_AOU'], errors='coerce'))
                & np.isfinite(pd.to_numeric(down40['SCALE_RES_ROB_ABS_SAL'], errors='coerce'))
                & (np.abs(pd.to_numeric(down40['SCALE_RES_ROB_AOU'], errors='coerce')) > trim_cutoff)
                & (np.abs(pd.to_numeric(down40['SCALE_RES_ROB_ABS_SAL'], errors='coerce')) > trim_cutoff)
                & (pd.to_numeric(down40['SCALE_RES_ROB_AOU'], errors='coerce') < 0)
            ].copy()

            if trim_depth_min is not None:
                candidates = candidates[pd.to_numeric(candidates[depth_col], errors='coerce') >= float(trim_depth_min)]
            if trim_depth_max is not None and trim_depth_max > 0:
                candidates = candidates[pd.to_numeric(candidates[depth_col], errors='coerce') <= float(trim_depth_max)]

            if candidates.empty:
                continue

            d20 = pd.to_numeric(down20[depth_col], errors='coerce').to_numpy(dtype=float)
            a20 = pd.to_numeric(down20['AOU_WORK'], errors='coerce').to_numpy(dtype=float)
            s20 = pd.to_numeric(down20['ABS_SAL_WORK'], errors='coerce').to_numpy(dtype=float)

            for _, row in candidates.iterrows():
                target_depth = float(pd.to_numeric(pd.Series([row.get(depth_col)]), errors='coerce').iloc[0])
                if not np.isfinite(target_depth):
                    continue
                if anomaly_max_depth is not None and anomaly_max_depth > 0 and target_depth > anomaly_max_depth:
                    continue

                if not _gradient_sign_change(d20, a20, target_depth, trim_window):
                    continue
                if not _gradient_sign_change(d20, s20, target_depth, trim_window):
                    continue

                delta_do, do_obs, _, _, _ = _endpoint_delta(depth_values, do_values, target_depth, depth_interval)
                delta_salinity, sal_obs, _, _, _ = _endpoint_delta(
                    depth_values, salinity_values, target_depth, depth_interval
                )
                delta_temperature, temp_obs, _, _, _ = _endpoint_delta(
                    depth_values, temperature_values, target_depth, depth_interval
                )
                delta_aou, aou_obs, _, _, _ = _endpoint_delta(
                    depth_values, aou_values, target_depth, depth_interval
                )

                score_aou = float(pd.to_numeric(pd.Series([row.get('SCALE_RES_ROB_AOU')]), errors='coerce').iloc[0])
                score_sal = float(pd.to_numeric(pd.Series([row.get('SCALE_RES_ROB_ABS_SAL')]), errors='coerce').iloc[0])
                trim_score = float(min(abs(score_aou), abs(score_sal)))

                result = {
                    'Profile_number': profile_num,
                    'depth': float(target_depth),
                    'delta_do': float(delta_do),
                    'delta_aou': float(delta_aou),
                    'delta_salinity': float(delta_salinity),
                    'delta_temperature': float(delta_temperature),
                    'do_value': float(do_obs),
                    'aou_value': float(aou_obs),
                    'salinity_value': float(sal_obs),
                    'temperature_value': float(temp_obs),
                    'trim_scale_res_rob_aou': score_aou,
                    'trim_scale_res_rob_abs_sal': score_sal,
                    'trim_score': trim_score,
                    'detection_method': method_norm,
                    'primary_metric': 'trim_score',
                    'primary_value': trim_score,
                    'anomaly_score': trim_score,
                    '_rank_score': trim_score,
                }
                profile_results.append(_append_profile_metadata(result, profile_data_clean))

        # 剖面内“深度近邻合并”：按模式得分降序贪心选取，避免近邻重复取点。
        if profile_results:
            if anomaly_min_depth is not None and anomaly_min_depth > 0:
                profile_results = [
                    rec for rec in profile_results
                    if np.isfinite(rec.get('depth', np.nan)) and rec.get('depth', np.nan) >= anomaly_min_depth
                ]
            if anomaly_max_depth is not None and anomaly_max_depth > 0:
                profile_results = [
                    rec for rec in profile_results
                    if np.isfinite(rec.get('depth', np.nan)) and rec.get('depth', np.nan) <= anomaly_max_depth
                ]
            if not profile_results:
                continue

            if depth_merge_tolerance is not None and depth_merge_tolerance > 0:
                profile_results.sort(
                    key=lambda r: np.nan_to_num(
                        r.get('_rank_score', r.get('delta_do', np.nan)), nan=-np.inf
                    ),
                    reverse=True,
                )
                kept = []
                kept_depths = []
                for rec in profile_results:
                    d = rec['depth']
                    if all(abs(d - kd) >= depth_merge_tolerance for kd in kept_depths):
                        kept.append(rec)
                        kept_depths.append(d)
                # 输出前按深度升序，便于查看
                kept.sort(key=lambda r: r['depth'])
                all_results.extend(kept)
            else:
                all_results.extend(profile_results)

        processed_profiles += 1
        if processed_profiles % 100 == 0 and verbose:
            print(f"已处理 {processed_profiles}/{total_profiles} 个剖面...")

    if not all_results:
        if verbose:
            print(f"未检测到满足条件的异常信号（method={method_norm}）。")
        return pd.DataFrame()

    results_df = pd.DataFrame(all_results)
    if '_rank_score' in results_df.columns:
        results_df = results_df.drop(columns=['_rank_score'])

    if anomaly_min_depth is not None and anomaly_min_depth > 0 and 'depth' in results_df.columns:
        results_df = results_df[results_df['depth'] >= anomaly_min_depth]
    if anomaly_max_depth is not None and anomaly_max_depth > 0 and 'depth' in results_df.columns:
        results_df = results_df[results_df['depth'] <= anomaly_max_depth]
    if verbose:
        print(
            f"总共检测到 {len(results_df)} 个潜在异常信号，"
            f"来自 {len(results_df['Profile_number'].unique())} 个剖面（method={method_norm}）"
        )

    return results_df

def _export_interacting_argo_worker(args):
    """
    Worker function for export_all_interacting_argo to support multiprocessing.
    Optimized for batch processing by Year with Low Memory Footprint.
    """
    y, kinds, cef, r_key, lmin, lmax, ltmin, ltmax = args

    try:
        # 1. 加载 Argo (按年加载)
        df = load_argo_data(y)
        if df.empty: return [], pd.DataFrame()
        
        # 地理过滤 (粗筛)
        pad = 2.0 
        lon_vals = df['Longitude'].to_numpy(dtype=float)
        lat_vals = df['Latitude'].to_numpy(dtype=float)
        
        if lmax < lmin: # 跨日界线
            lon_mask = (lon_vals >= lmin - pad) | (lon_vals <= lmax + pad)
        else:
            lon_mask = (lon_vals >= lmin - pad) & (lon_vals <= lmax + pad)
            
        lat_mask = (lat_vals >= ltmin - pad) & (lat_vals <= ltmax + pad)
        df_geo = df[lon_mask & lat_mask].copy()
        if df_geo.empty: return [], pd.DataFrame()
        
        # 提取 Baseline (去重)
        baseline = (
            df_geo.sort_values(['Profile_number', 'Depth'])
            .groupby('Profile_number', as_index=False)
            .first()[['Profile_number', 'Longitude', 'Latitude', 'Year', 'Month', 'Day']]
        )
        baseline['date'] = pd.to_datetime(baseline[['Year', 'Month', 'Day']])
        
        # 按天分组 Argo (Dict of DataFrames)
        argo_by_day = {d: grp for d, grp in baseline.groupby('date')}

        # 2. 加载当年涡旋数据 (仅元数据，不含轮廓，节省内存)
        start_d = pd.Timestamp(f"{y}-01-01")
        end_d = pd.Timestamp(f"{y}-12-31")
        
        # Pass 1: Load lightweight metadata
        eddy_data_meta = _load_eddy_datasets_for_range(
            kinds=kinds,
            start_date=start_d,
            end_date=end_d,
            region_key=r_key,
            include_contours=False # 关键：先不加载轮廓
        )
        
        # 3. 构建涡旋按天索引 (仅元数据)
        eddy_by_day = defaultdict(list)
        for ds_name, ds_tracks in eddy_data_meta.items():
            for item in ds_tracks:
                track = item[0] if isinstance(item, tuple) else item
                if not track: continue
                
                # 批量转换时间
                raw_times = [p[1] for p in track]
                converted_times = convert_date(raw_times)
                
                if isinstance(converted_times, pd.Series):
                    ts_index = pd.DatetimeIndex(converted_times)
                else:
                    ts_index = pd.DatetimeIndex([converted_times])
                
                mask = (ts_index >= start_d) & (ts_index <= end_d)
                valid_indices = np.where(mask)[0]
                
                for idx in valid_indices:
                    p = track[idx]
                    d = ts_index[idx].normalize()
                    
                    eddy_by_day[d].append({
                        'track_id': p[0],
                        'lon': p[2],
                        'lat': p[3],
                        'radius': p[8],
                        'ds_name': ds_name,
                        # 'contour_lon': p[6], # Pass 1 中这些是 None 或空
                        # 'contour_lat': p[7]
                    })

        interacting_records = []
        
        # 记录需要进一步检查轮廓的候选者
        # candidates[ds_name][track_id] = set(dates)
        candidates = defaultdict(lambda: defaultdict(set))
        
        # 4. Pass 1: 批量圆筛选
        common_days = sorted(set(argo_by_day.keys()) & set(eddy_by_day.keys()))
        
        # print(f"[Info] Year {y}: Checking {len(common_days)} days with Argo & Eddies...")

        for current_date in common_days:
            day_argo_df = argo_by_day[current_date]
            day_eddies = eddy_by_day[current_date]
            
            argo_lons = day_argo_df['Longitude'].to_numpy(dtype=float)
            argo_lats = day_argo_df['Latitude'].to_numpy(dtype=float)
            
            for eddy in day_eddies:
                e_lon, e_lat, e_rad = eddy['lon'], eddy['lat'], eddy['radius']
                eff_rad = e_rad * cef
                
                # 粗筛
                scale_rough = approximate_degree_length(e_lat)
                # 使用 1.5 倍 buffer，并分别计算经纬度方向的度数阈值
                # 注意：meters_per_degree_lon 随纬度增加而减小，因此同样的米数对应的度数会变大
                # 为安全起见，使用 local scale
                rad_deg_lat = (eff_rad / scale_rough['meters_per_degree_lat']) * 1.5
                # 防止极点附近除零或过大，设置上限（例如 180度）
                if scale_rough['meters_per_degree_lon'] < 1000: # 极靠近极点
                     rad_deg_lon = 180.0
                else:
                     rad_deg_lon = (eff_rad / scale_rough['meters_per_degree_lon']) * 1.5
                
                dlon = np.abs(_minimal_lon_diff_deg(argo_lons, e_lon))
                dlat = np.abs(argo_lats - e_lat)
                
                mask_bb = (dlon < rad_deg_lon) & (dlat < rad_deg_lat)
                if not np.any(mask_bb):
                    continue
                
                # 精细距离
                # 使用 adaptive_distance_m 替代手动平面计算，以解决高纬度畸变问题
                c_lons = argo_lons[mask_bb]
                c_lats = argo_lats[mask_bb]
                
                dists = adaptive_distance_m(c_lons, c_lats, e_lon, e_lat)
                mask_circle = dists <= eff_rad
                
                if np.any(mask_circle):
                    # 命中圆，记录为候选，稍后加载轮廓做精确检查
                    candidates[eddy['ds_name']][eddy['track_id']].add(current_date)

        # 5. Pass 2: 批量加载轮廓并精确匹配 (Batch Loading Optimization)
        # 收集所有需要加载轮廓的 track_ids
        for ds_name, track_map in candidates.items():
            target_ids = list(track_map.keys())
            if not target_ids:
                continue
            
            try:
                # 批量加载：一次性读取该数据集下所有候选涡旋的完整轨迹（含轮廓）
                # find_track 返回 {tid: list_of_rows} (当 len(target_ids) > 1)
                # 或 list_of_rows (当 len(target_ids) == 1)
                # 统一处理为 dict
                batch_res = find_track(
                    ds_name.lower(), 
                    target_ids, 
                    region=r_key, 
                    include_contours=True, 
                    return_list=True
                )
                
                tracks_dict = {}
                if len(target_ids) == 1:
                    # 单个 ID 返回的是 list
                    tracks_dict[target_ids[0]] = batch_res
                else:
                    tracks_dict = batch_res
                
                # 遍历每个涡旋进行多边形检测
                for track_id, full_track_list in tracks_dict.items():
                    dates_set = track_map.get(track_id, set())
                    if not dates_set: continue
                    
                    # 构建时间索引: YYYYMMDD int -> row
                    # full_track_list item: [tid, ymd, clon, clat, mlon, mlat, clon_poly, clat_poly, rad, ...]
                    # Index: 1=ymd, 2=clon, 3=clat, 6=clon_poly, 7=clat_poly, 8=rad
                    
                    # 快速筛选：只处理 dates_set 中的日期
                    # 将 dates_set 转为 YYYYMMDD 整数集合以便快速查找
                    target_ymds = {d.year * 10000 + d.month * 100 + d.day for d in dates_set}
                    
                    for row in full_track_list:
                        ymd = row[1]
                        if ymd not in target_ymds:
                            continue
                            
                        # 提取数据
                        e_lon = row[2]
                        e_lat = row[3]
                        c_lon_poly = row[6]
                        c_lat_poly = row[7]
                        e_rad = row[8]
                        
                        # 还原日期对象用于查找 Argo
                        # ymd is int YYYYMMDD
                        y_ = ymd // 10000
                        m_ = (ymd % 10000) // 100
                        d_ = ymd % 100
                        current_date = pd.Timestamp(year=y_, month=m_, day=d_)
                        
                        # 获取当天的 Argo
                        if current_date not in argo_by_day: continue
                        day_argo_df = argo_by_day[current_date]
                        
                        argo_lons = day_argo_df['Longitude'].to_numpy(dtype=float)
                        argo_lats = day_argo_df['Latitude'].to_numpy(dtype=float)
                        argo_ids = day_argo_df['Profile_number'].to_numpy()
                        
                        # 重复圆筛选 (为了拿到 mask_circle 对应的点)
                        eff_rad = e_rad * cef
                        
                        scale_rough = approximate_degree_length(e_lat)
                        rad_deg_lat = (eff_rad / scale_rough['meters_per_degree_lat']) * 1.5
                        if scale_rough['meters_per_degree_lon'] < 1000:
                             rad_deg_lon = 180.0
                        else:
                             rad_deg_lon = (eff_rad / scale_rough['meters_per_degree_lon']) * 1.5

                        dlon = np.abs(_minimal_lon_diff_deg(argo_lons, e_lon))
                        dlat = np.abs(argo_lats - e_lat)
                        mask_bb = (dlon < rad_deg_lon) & (dlat < rad_deg_lat)
                        
                        if not np.any(mask_bb): continue
                        
                        c_lons = argo_lons[mask_bb]
                        c_lats = argo_lats[mask_bb]
                        c_ids = argo_ids[mask_bb]
                        
                        # 使用 adaptive_distance_m 替代手动平面计算
                        dists = adaptive_distance_m(c_lons, c_lats, e_lon, e_lat)
                        mask_circle = dists <= eff_rad
                        
                        if np.any(mask_circle):
                            circle_hits_idx = np.where(mask_circle)[0]
                            
                            # 多边形检测
                            has_poly = False
                            is_inside = np.zeros(len(circle_hits_idx), dtype=bool)
                            
                            if c_lon_poly is not None and c_lat_poly is not None:
                                c_lon_arr = np.asarray(c_lon_poly)
                                c_lat_arr = np.asarray(c_lat_poly)
                                if c_lon_arr.size >= 3:
                                    has_poly = True
                                    c_lon_norm = e_lon + _minimal_lon_diff_deg(c_lon_arr, e_lon)
                                    verts = np.column_stack((c_lon_norm, c_lat_arr))
                                    path = MplPath(verts)
                                    
                                    p_lons = c_lons[mask_circle]
                                    p_lats = c_lats[mask_circle]
                                    p_lons_norm = e_lon + _minimal_lon_diff_deg(p_lons, e_lon)
                                    points = np.column_stack((p_lons_norm, p_lats))
                                    is_inside = path.contains_points(points)
                            
                            for idx_in_subset, inside in enumerate(is_inside):
                                real_idx = circle_hits_idx[idx_in_subset]
                                if has_poly:
                                    m_str = 'poly' if inside else 'circle'
                                else:
                                    m_str = 'circle'
                                    
                                interacting_records.append({
                                    'Profile_number': c_ids[real_idx],
                                    'Year': y,
                                    'Month': current_date.month,
                                    'Day': current_date.day,
                                    'track_id': track_id,
                                    'ds_name': ds_name,
                                    'method': m_str
                                })
            except Exception as e:
                print(f"[Warn] Batch processing failed for {ds_name} in Year {y}: {e}")
                traceback.print_exc()
                continue

        return interacting_records, baseline
    except Exception as e:
        print(f"[Error] Year {y}: {e}")
        traceback.print_exc()
        return [], pd.DataFrame()

def export_all_interacting_argo(
    start_year: int,
    end_year: int,
    eddy_datasets: dict | list[str] | tuple[str, ...] | None = None,
    circle_enlargement_factor: float | None = None,
    output_path: str | Path | None = None,
    num_workers: int = 1,
):
    """
    计算并导出指定年份范围内所有位于涡旋内部的 Argo 剖面数据。

    遍历指定年份的 Argo 数据，加载对应 META Tracks 涡旋轨迹，判断每个 Argo 剖面是否位于涡旋内部（支持
    多核并行），将命中的剖面（含匹配涡旋 ID、类型等）保存为 Parquet，并同时保存该区域全部 Argo 剖面的
    基础信息（用于后续计算交互率分母）。

    参数:
        - start_year (int): 起始年份。
        - end_year (int): 结束年份。
        - eddy_datasets (dict | list[str] | tuple[str, ...] | None): 使用的涡旋数据集（如 ['acl','acs','cs','cl']）；None 时使用所有可用数据集。
        - circle_enlargement_factor (float | None): 涡旋边界放大系数；None 时从 processing.yml 读取。
        - output_path (str | Path | None): 结果文件保存路径；None 时用 `plot_outputs/shared/<region>/statistics/all_interacting_argo_<years>.parquet`。
        - num_workers (int): 并行进程数，默认 1。

    输出:
        - 一个 Parquet 文件，含所有与涡旋发生交互的 Argo 剖面详细信息；另存区域基线 Argo 剖面信息。
    """
    if circle_enlargement_factor is None:
        circle_enlargement_factor = globals().get('circle_enlargement_factor', 1.2)
    
    region_slug = _current_region_key()
    if output_path is None:
        out_dir = _shared_output_dir("statistics", region_slug)
        out_dir.mkdir(parents=True, exist_ok=True)
        output_path = out_dir / f"all_interacting_argo_{start_year}_{end_year}.parquet"
    else:
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

    print(f"--- Exporting All Interacting Argo (Baseline) {start_year}-{end_year} ---")
    print(f"Output: {output_path}")
    print(f"Workers: {num_workers} (Task granularity: Yearly)")
    
    # 准备任务列表 (Year)
    years = list(range(start_year, end_year + 1))
    
    # 准备参数
    kinds = eddy_datasets if eddy_datasets else ['acs', 'acl', 'cs', 'cl']
    if isinstance(kinds, dict): kinds = list(kinds.keys())
    kinds = [str(k).lower() for k in kinds]
    
    worker_args = [
        (y, kinds, circle_enlargement_factor, region_slug, lonmin, lonmax, latmin, latmax)
        for y in years
    ]
    
    print(f"Total tasks: {len(worker_args)}")
    
    # 执行并行
    all_results = []
    all_profiles = []
    
    if num_workers > 1:
        with multiprocessing.Pool(processes=num_workers, initializer=init_worker, initargs=(None,), maxtasksperchild=1) as pool:
            for res_interactions, res_profiles in tqdm(pool.imap_unordered(_export_interacting_argo_worker, worker_args), total=len(worker_args), desc="Processing Years"):
                all_results.extend(res_interactions)
                if not res_profiles.empty:
                    all_profiles.append(res_profiles)
    else:
        for args in tqdm(worker_args, desc="Processing Years"):
            res_interactions, res_profiles = _export_interacting_argo_worker(args)
            all_results.extend(res_interactions)
            if not res_profiles.empty:
                all_profiles.append(res_profiles)
            
    # 保存交互记录
    if all_results:
        df_out = pd.DataFrame(all_results)
        df_out.to_parquet(output_path, index=False)
        print(f"Saved {len(df_out)} interacting records to: {output_path}")
    else:
        print("No interacting profiles found.")

    # 保存区域内所有 Argo 剖面 (Baseline)
    if all_profiles:
        df_profiles = pd.concat(all_profiles, ignore_index=True)
        # 构造 Baseline 文件名: all_region_argo_{start}_{end}.parquet
        baseline_path = output_path.parent / f"all_region_argo_{start_year}_{end_year}.parquet"
        df_profiles.to_parquet(baseline_path, index=False)
        print(f"Saved {len(df_profiles)} region profiles (Baseline) to: {baseline_path}")
    else:
        print("No region profiles found.")


def query_argo_inside_eddy(
    profile_number: int,
    year: int,
    month: int | None = None,
    day: int | None = None,
    precomputed_file: str | Path | None = None,
    region: str | None = None,
    return_matches: bool = True,
) -> dict:
    """查询指定 Argo 剖面是否位于任一涡旋内部（基于 export_all_interacting_argo 预计算结果）。

    仅查询已落盘的交互记录，不做实时几何匹配。

    参数:
        - profile_number (int): Argo 剖面编号（Profile_number）。
        - year (int): 年份过滤。
        - month (int | None): 月份过滤（可选）。
        - day (int | None): 日期过滤（可选）。
        - precomputed_file (str | Path | None): 显式指定交互记录 parquet 路径（可选）。
        - region (str | None): 区域 slug；None 时使用当前 region。
        - return_matches (bool): True 时返回命中的明细表（DataFrame），默认 True。

    返回:
        - dict: 含以下键：

            - inside_eddy (bool)：是否命中至少一条交互记录。
            - hit_count (int)：命中记录数。
            - source (str | None)：实际使用的数据源路径。
            - message (str)：查询状态说明。
            - matches (pd.DataFrame, 可选)：命中的明细（去重后）。

    说明:
        - 默认优先读取 `plot_outputs/shared/<region>/statistics/all_interacting_argo_*.parquet`。
        - 使用 parquet 过滤条件（Profile_number + Year + 可选 Month/Day），仅扫描少量行以快速判定。
    """
    pnum = int(profile_number)
    y = int(year)
    m = int(month) if month is not None else None
    d = int(day) if day is not None else None

    # 1) 解析候选数据源
    candidate_files: list[Path] = []
    if precomputed_file is not None:
        p = Path(precomputed_file)
        if p.exists() and p.is_file():
            candidate_files = [p]
    else:
        region_slug = region or _current_region_key()
        stats_dir = _shared_output_dir("statistics", region_slug)
        if stats_dir.exists():
            candidate_files = sorted(stats_dir.glob("all_interacting_argo_*.parquet"))

    if not candidate_files:
        msg = (
            "No precomputed interaction parquet found. "
            "Run export_all_interacting_argo(...) first or pass precomputed_file."
        )
        result = {
            'inside_eddy': False,
            'hit_count': 0,
            'source': None,
            'message': msg,
        }
        if return_matches:
            result['matches'] = pd.DataFrame(
                columns=['Profile_number', 'Year', 'Month', 'Day', 'track_id', 'ds_name', 'method']
            )
        return result

    # 2) 用 parquet 过滤条件快速查询（AND 条件）
    filters: list[tuple[str, str, int]] = [
        ('Profile_number', '==', pnum),
        ('Year', '==', y),
    ]
    if m is not None:
        filters.append(('Month', '==', m))
    if d is not None:
        filters.append(('Day', '==', d))

    keep_cols = ['Profile_number', 'Year', 'Month', 'Day', 'track_id', 'ds_name', 'method']
    hit_frames: list[pd.DataFrame] = []
    used_sources: list[str] = []

    for src in candidate_files:
        try:
            df_hit = pd.read_parquet(src, columns=keep_cols, filters=filters)
        except Exception:
            # 兼容极少数引擎不支持 filters 的情况
            try:
                df_hit = pd.read_parquet(src, columns=keep_cols)
                cond = pd.to_numeric(df_hit['Profile_number'], errors='coerce').eq(pnum) & pd.to_numeric(df_hit['Year'], errors='coerce').eq(y)
                if m is not None:
                    cond = cond & pd.to_numeric(df_hit['Month'], errors='coerce').eq(m)
                if d is not None:
                    cond = cond & pd.to_numeric(df_hit['Day'], errors='coerce').eq(d)
                df_hit = df_hit[cond].copy()
            except Exception:
                continue

        if not df_hit.empty:
            hit_frames.append(df_hit)
            used_sources.append(str(src))

    if hit_frames:
        hits = pd.concat(hit_frames, ignore_index=True)
        hits = hits.drop_duplicates(subset=keep_cols, keep='first')
        if {'Year', 'Month', 'Day'}.issubset(hits.columns):
            hits = hits.sort_values(['Year', 'Month', 'Day', 'ds_name', 'track_id'])
        hit_count = int(len(hits))
        result = {
            'inside_eddy': True,
            'hit_count': hit_count,
            'source': ';'.join(used_sources),
            'message': f'Found {hit_count} interacting record(s).',
        }
        if return_matches:
            result['matches'] = hits.reset_index(drop=True)
        return result

    result = {
        'inside_eddy': False,
        'hit_count': 0,
        'source': ';'.join(str(p) for p in candidate_files),
        'message': 'No interacting records found for the given profile/date filter.',
    }
    if return_matches:
        result['matches'] = pd.DataFrame(columns=keep_cols)
    return result

def calculate_interaction_statistics(
    start_year: int,
    end_year: int,
    eddy_datasets: dict | list[str] | tuple[str, ...] | None = None,
    detection_config: DetectionConfig | None = None,
    circle_enlargement_factor: float | None = None,
    save_report: bool = True,
    precomputed_file: str | Path | None = None,
    anomalies_file: str | Path | None = None,
    use_precomputed_anomalies: bool = True,
):
    """
    计算并对比 Argo 剖面与涡旋的交互概率（Baseline vs Anomalies）。

    加载指定年份的全部 Argo 数据与 export_all_interacting_argo 预计算的交互记录，计算 Baseline（全部剖面
    落在涡旋内的比例）与 Anomalies（按 detection_config 筛选的异常剖面落在涡旋内的比例），并输出对比报告。

    参数:
        - start_year (int): 起始年份。
        - end_year (int): 结束年份。
        - eddy_datasets (dict | list[str] | tuple[str, ...] | None): 使用的涡旋数据集；None 时使用所有可用数据集。
        - detection_config (DetectionConfig | None): 异常识别配置；None 时用默认。
        - circle_enlargement_factor (float | None): 涡旋边界放大系数；None 时从 processing.yml 读取。
        - save_report (bool): 是否保存对比报告，默认 True。
        - precomputed_file (str | Path | None): 显式指定交互记录 parquet；None 时用默认路径。
        - anomalies_file (str | Path | None): 显式指定异常剖面文件；None 时按默认查找。
        - use_precomputed_anomalies (bool): True 时优先读取 plot_argo_hotspots 生成的异常文件，找不到再实时调用 calculate_delta_do，默认 True。

    返回:
        - dict: Baseline 与 Anomalies 交互概率的对比统计。

    说明:
        - 必须先运行 export_all_interacting_argo 生成交互记录文件，否则报错。
    """
    cfg = _resolve_detection_config(detection_config)
    method_name = cfg.method
    run_tag = cfg.file_stem()
    if circle_enlargement_factor is None: circle_enlargement_factor = globals().get('circle_enlargement_factor', 1.2)

    print(f"--- Calculating Interaction Statistics {start_year}-{end_year} ---")
    
    # --- 0. 尝试加载预计算的交互记录 ---
    interacting_ids = set()
    loaded_precomputed = False
    
    region_slug = _current_region_key()
    if precomputed_file is None:
        # 尝试默认路径
        default_file = _shared_output_dir("statistics", region_slug) / f"all_interacting_argo_{start_year}_{end_year}.parquet"
        if default_file.exists():
            precomputed_file = default_file
            
    if precomputed_file and Path(precomputed_file).exists():
        print(f"[*] Loading precomputed interactions from: {precomputed_file}")
        try:
            df_int = pd.read_parquet(precomputed_file)
            if 'Profile_number' in df_int.columns:
                interacting_ids = set(df_int['Profile_number'].unique())
                loaded_precomputed = True
                print(f"[*] Loaded {len(interacting_ids)} unique interacting profiles.")
        except Exception as e:
            print(f"[WARN] Failed to read precomputed file: {e}")
            
    # 1. 加载 Argo 数据 (用于计算分母和检测异常)
    # 优化：尝试加载预计算的区域 Argo 列表 (Baseline)
    baseline_profiles = pd.DataFrame()
    loaded_baseline_file = False
    
    default_baseline_file = _shared_output_dir("statistics", region_slug) / f"all_region_argo_{start_year}_{end_year}.parquet"
    if default_baseline_file.exists():
        print(f"[*] Loading precomputed region profiles from: {default_baseline_file}")
        try:
            baseline_profiles = pd.read_parquet(default_baseline_file)
            if not baseline_profiles.empty:
                # 确保严格符合当前区域定义 (因为 export 可能包含 padding)
                lons = baseline_profiles['Longitude'].to_numpy()
                lats = baseline_profiles['Latitude'].to_numpy()
                mask_geo = _region_lon_mask(lons, lonmin, lonmax) & (lats >= latmin) & (lats <= latmax)
                baseline_profiles = baseline_profiles[mask_geo].copy()
                
                baseline_profiles['date'] = pd.to_datetime(baseline_profiles[['Year', 'Month', 'Day']])
                loaded_baseline_file = True
                print(f"[*] Loaded {len(baseline_profiles)} region profiles (Strictly inside region).")
        except Exception as e:
            print(f"[WARN] Failed to read region profiles file: {e}")

    argo_geo = pd.DataFrame() # 用于实时计算异常的原始数据
    
    if not loaded_baseline_file:
        print("[*] Loading Argo data (Real-time)...")
        argo_list = []
        for y in range(start_year, end_year + 1):
            try:
                df = load_argo_data(y)
                if not df.empty:
                    argo_list.append(df)
            except Exception:
                pass
        if not argo_list:
            print("No Argo data found.")
            return
        argo_all = pd.concat(argo_list, ignore_index=True)
        
        # 地理过滤
        lon_vals = argo_all['Longitude'].to_numpy(dtype=float)
        lat_vals = argo_all['Latitude'].to_numpy(dtype=float)
        lon_mask = _region_lon_mask(lon_vals, lonmin, lonmax)
        lat_mask = (lat_vals >= latmin) & (lat_vals <= latmax)
        argo_geo = argo_all[lon_mask & lat_mask].copy()
        
        if argo_geo.empty:
            print("No Argo data in region.")
            return

        # 提取 Baseline (所有剖面)
        # 按 Profile_number 去重，保留时间/位置
        baseline_profiles = (
            argo_geo.sort_values(['Profile_number', 'Depth'])
            .groupby('Profile_number', as_index=False)
            .first()[['Profile_number', 'Longitude', 'Latitude', 'Year', 'Month', 'Day']]
        )
        baseline_profiles['date'] = pd.to_datetime(baseline_profiles[['Year', 'Month', 'Day']])
    
    # 提取 Anomalies (异常剖面)
    anomalies = pd.DataFrame()
    
    # 自动查找默认异常文件
    if anomalies_file is None and use_precomputed_anomalies:
        region_slug = _current_region_key()
        out_dir = cfg.output_dir("plot_argo_hotspots", region_slug)
        default_anomalies_file = out_dir / f"anomalies_{start_year}_{end_year}_{run_tag}.parquet"
        if default_anomalies_file.exists():
            anomalies_file = default_anomalies_file
            print(f"[*] Found default precomputed anomalies file: {anomalies_file}")
        else:
            print(f"[*] Default precomputed anomalies file not found: {default_anomalies_file}")
            print("    Will proceed with real-time calculation.")

    if anomalies_file and Path(anomalies_file).exists():
        print(f"[*] Loading precomputed anomalies from: {anomalies_file}")
        try:
            anomalies = pd.read_parquet(anomalies_file)
            if 'detection_method' in anomalies.columns:
                method_mask = anomalies['detection_method'].astype(str).str.lower().eq(method_name)
                total_count = len(anomalies)
                mismatch_count = int((~method_mask).sum())
                if mismatch_count > 0:
                    print(f"[WARN] Mixed detection_method found: expected={method_name}, mismatched={mismatch_count}/{total_count}.")
                else:
                    print(f"[*] Method={method_name}, records={total_count}")
                anomalies = anomalies[method_mask].copy()
            # 确保只统计当前区域内的异常（如果文件包含更多区域）
            # 使用 Profile_number 与 baseline_profiles (已过滤区域) 取交集最稳妥
            valid_ids = set(baseline_profiles['Profile_number'])
            anomalies = anomalies[anomalies['Profile_number'].isin(valid_ids)].copy()
            print(f"[*] Loaded {len(anomalies)} anomalies (filtered by region).")
        except Exception as e:
            print(f"[WARN] Failed to read anomalies file: {e}. Falling back to calculation.")
            anomalies = pd.DataFrame() # 触发下方重新计算

    if anomalies.empty:
        print("[*] Detecting anomalies (Real-time calculation)...")
        # 如果没有加载 baseline file，说明 argo_geo 已经准备好了
        # 如果加载了 baseline file，但没有 anomalies file，我们需要重新加载 argo_geo 吗？
        # 是的，因为 calculate_delta_do 需要原始剖面数据 (argo_geo)
        
        if argo_geo.empty:
             print("[*] Reloading Argo data for anomaly calculation...")
             # 这里必须重新加载，因为之前可能跳过了
             argo_list = []
             for y in range(start_year, end_year + 1):
                try:
                    df = load_argo_data(y)
                    if not df.empty:
                        argo_list.append(df)
                except Exception:
                    pass
             if argo_list:
                 argo_all = pd.concat(argo_list, ignore_index=True)
                 lon_vals = argo_all['Longitude'].to_numpy(dtype=float)
                 lat_vals = argo_all['Latitude'].to_numpy(dtype=float)
                 lon_mask = _region_lon_mask(lon_vals, lonmin, lonmax)
                 lat_mask = (lat_vals >= latmin) & (lat_vals <= latmax)
                 argo_geo = argo_all[lon_mask & lat_mask].copy()

        if not argo_geo.empty:
            anomalies = calculate_delta_do(
                argo_geo,
                detection_config=cfg,
                remove_outliers=True,
                verbose=False
            )
        else:
            print("[Error] Cannot calculate anomalies: No Argo data available.")
    
    anomaly_ids = set()
    if not anomalies.empty:
        anomaly_ids = set(anomalies['Profile_number'].unique())
    
    # 2. 若未加载预计算文件，则报错返回
    if not loaded_precomputed:
        print(f"[Error] Precomputed interaction file not found or failed to load.")
        print(f"Please run 'export_all_interacting_argo(start_year={start_year}, end_year={end_year}, ...)' first.")
        return

    # 4. 统计
    total_baseline = len(baseline_profiles)
    interacted_baseline = len(interacting_ids)
    pct_baseline = (interacted_baseline / total_baseline * 100) if total_baseline > 0 else 0.0
    
    total_anom = len(anomaly_ids)
    # 异常且交互 = 异常ID集合 与 交互ID集合 的交集
    interacted_anom = len(anomaly_ids & interacting_ids)
    pct_anom = (interacted_anom / total_anom * 100) if total_anom > 0 else 0.0
    
    # 5. 输出报告
    source_str = f"Precomputed File ({Path(precomputed_file).name})" if loaded_precomputed else "Real-time Calculation"
    
    report = (
        f"========================================\n"
        f"Interaction Statistics Report ({start_year}-{end_year})\n"
        f"Region: {_current_region_key()}\n"
        f"Source: {source_str}\n"
        f"Method: {cfg.method}\n"
        f"Criteria: {cfg.threshold_label()}, Depth>={cfg.anomaly_min_depth}m\n"
        f"----------------------------------------\n"
        f"[Baseline] All Argo Profiles:\n"
        f"  Total Profiles:       {total_baseline}\n"
        f"  Inside Eddies:        {interacted_baseline}\n"
        f"  Interaction Rate:     {pct_baseline:.2f}%\n"
        f"----------------------------------------\n"
        f"[Subset] Anomalies:\n"
        f"  Total Anomalies:      {total_anom}\n"
        f"  Inside Eddies:        {interacted_anom}\n"
        f"  Interaction Rate:     {pct_anom:.2f}%\n"
        f"----------------------------------------\n"
        f"Ratio (Anom Rate / Base Rate): {(pct_anom/pct_baseline if pct_baseline > 0 else 0):.2f}x\n"
        f"========================================"
    )
    
    print(report)
    
    if save_report:
        region_slug = _current_region_key()
        out_dir = cfg.output_dir("statistics", region_slug)
        out_dir.mkdir(parents=True, exist_ok=True)
        fname = out_dir / f"interaction_stats_{start_year}_{end_year}_{run_tag}.txt"
        with open(fname, 'w') as f:
            f.write(report)
        print(f"Report saved to: {fname}")
